from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import TemplateView
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from datetime import date
import datetime
from django.db.models.functions import Coalesce
import calendar
from django.db.models import Sum, F, Q
import pandas as pd
import os
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from .models import Trabajador, Cargo, Puesto , ProgramacionMensual, Superintendencia, GerenciaGeneral, Gerencia, ProgramacionSemanal, ProgramacionDiaria, DetalleDiario
from .forms import TrabajadorForm, CargoForm , PuestoForm, CargaMasivaPlanForm, SuperintendenciaForm, GerenciaForm, GerenciaGeneralForm, InspeccionConjuntaForm, PatronFormSet, ReportadoFormSet
from .models import Agente
from .forms import AgenteForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import requests
from io import BytesIO
from datetime import timedelta, date
from openpyxl import load_workbook
import os
from django.conf import settings
from django.template.loader import render_to_string
from weasyprint import HTML
import logging
logger = logging.getLogger(__name__)
import math
import statistics
from decimal import Decimal, ROUND_HALF_UP
from openpyxl.drawing.image import Image
from django.db.models import OuterRef, Subquery
from django.db.models import Q
from dateutil.relativedelta import relativedelta
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils import timezone
import locale
from django.views.generic import FormView

from .models import Trabajador, Cargo, Dispositivo, Diagnostico, FotoDiagnostico, Calibracion, Mantenimiento, Operatividad, Verificacion, Evidencias, Resultados, Patrones, FotoPatron, HistorialCambioAccesorio, InventarioRepuestos, Accesorio, Inspeccion, DetalleInspeccion, PruebaTecnica, ProgramaMantenimiento, FilaCronograma
from .forms import TrabajadorForm, CargoForm, DiagnosticoForm, MantenimientoForm, OperatividadForm, CalibracionForm, HistorialCambioAccesorioForm, CambioDosimetroForm, CambioBombaForm, InspeccionEspecificaForm, DetalleMuestraFormSet, FilaCronogramaForm

try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    pass

# --- MIXIN DE SEGURIDAD ACTUALIZADO ---
class GroupRequiredMixin(UserPassesTestMixin):
    group_required = None # Se define en cada vista hija

    def test_func(self):
        user = self.request.user
        
        # 1. Permitir siempre si es superusuario
        if user.is_superuser:
            return True
            
        # 2. Permitir siempre si pertenece al grupo "Yeni_admin" (Llave maestra)
        if user.groups.filter(name='Yeni_admin').exists():
            return True
            
        # 3. Permitir si pertenece al grupo específico de la vista
        if self.group_required:
            return user.groups.filter(name=self.group_required).exists()
            
        return False

    def handle_no_permission(self):
        # Mensaje de error personalizado
        messages.error(self.request, f'Restringido: No tienes permisos para acceder a esta sección.')
        return redirect('proyecto_monitoreo_smcv:dashboard')


class DashboardYeniView(LoginRequiredMixin, TemplateView):
    template_name = 'proyecto_monitoreo_smcv/dashboard.html'

class SoporteTecnicoView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    template_name = 'proyecto_monitoreo_smcv/soporte/index.html'
    group_required = 'Yeni_Soporte'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['dispositivos'] = Dispositivo.objects.all().order_by('-id')[:10]
        context['diagnosticos'] = Diagnostico.objects.all().order_by('-fecha_revision')[:5]
        context['mantenimientos'] = Mantenimiento.objects.all().order_by('-created_at')[:5]
        context['operatividades'] = Operatividad.objects.all().order_by('-fecha_operatividad')[:5]
        context['calibraciones'] = Calibracion.objects.all().order_by('-fecha_calibracion')[:5]
        context['ultimas_inspecciones'] = Inspeccion.objects.filter(total_equipos=1).order_by('-fecha_inspeccion')[:5]
        context['ultimas_inspecciones_conjuntas'] = Inspeccion.objects.annotate(
            num_detalles=Count('detalles')
        ).filter(num_detalles__gt=1).order_by('-fecha_inspeccion')[:5]
        return context

class ProgramacionView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    template_name = 'proyecto_monitoreo_smcv/programacion/programacion.html'
    group_required = 'Yeni_Programacion'
    extra_context = {'show_sidebar': True}

class EvaluacionView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    template_name = 'proyecto_monitoreo_smcv/evaluacion/evaluacion.html'
    group_required = 'Yeni_Evaluacion'

class CapacitacionView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    template_name = 'proyecto_monitoreo_smcv/capacitacion/index.html'
    group_required = 'Yeni_Capacitacion'

class SupervisorView(LoginRequiredMixin, GroupRequiredMixin, TemplateView):
    template_name = 'proyecto_monitoreo_smcv/supervisor/supervisor.html'
    group_required = 'Yeni_Supervisor'

# ==========================================
# Vista para Trabajadores
# ==========================================

@login_required
def lista_trabajador(request):
    query = request.GET.get('q')
    if query:
        trabajadores = Trabajador.objects.filter(nombre__icontains=query) | Trabajador.objects.filter(dni__icontains=query)
    else:
        trabajadores = Trabajador.objects.all()

    context = {
        'trabajadores': trabajadores,
        'show_sidebar': True
    }
    return render(request, 'trabajador/trabajador_list.html', context)

@login_required
def crear_trabajador(request):
    if request.method == 'POST':
        form = TrabajadorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Trabajador creado con éxito")
            return redirect('proyecto_monitoreo_smcv:lista_trabajador')
    else:
        form = TrabajadorForm()
    
    context = {
        'form': form,
        'show_sidebar': True
    }
    return render(request, 'trabajador/trabajador_form.html', context)

# ==========================================
# Vistas para Ubicación
# ==========================================
from .models import Ubicacion

@login_required
def lista_ubicaciones(request):
    q = request.GET.get('q', '')
    ubicaciones = Ubicacion.objects.all().order_by('nombre')
    if q:
        ubicaciones = ubicaciones.filter(nombre__icontains=q)
    context = {
        'ubicaciones': ubicaciones,
        'q': q,
        'show_sidebar': True,
    }
    return render(request, 'proyecto_monitoreo_smcv/soporte/ubicaciones/lista_ubicaciones.html', context)

@login_required
def crear_ubicacion(request):
    from .forms import UbicacionForm
    if request.method == 'POST':
        form = UbicacionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Ubicación creada correctamente.")
            return redirect('proyecto_monitoreo_smcv:lista_ubicaciones')
    else:
        form = UbicacionForm()
    context = {
        'form': form,
        'titulo': 'Nueva Ubicación',
        'show_sidebar': True,
    }
    return render(request, 'proyecto_monitoreo_smcv/soporte/ubicaciones/form_ubicacion.html', context)

@login_required
def editar_ubicacion(request, pk):
    from .forms import UbicacionForm
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    if request.method == 'POST':
        form = UbicacionForm(request.POST, instance=ubicacion)
        if form.is_valid():
            form.save()
            messages.success(request, "Ubicación actualizada correctamente.")
            return redirect('proyecto_monitoreo_smcv:lista_ubicaciones')
    else:
        form = UbicacionForm(instance=ubicacion)
    context = {
        'form': form,
        'titulo': f'Editar: {ubicacion.nombre}',
        'ubicacion': ubicacion,
        'show_sidebar': True,
    }
    return render(request, 'proyecto_monitoreo_smcv/soporte/ubicaciones/form_ubicacion.html', context)

@login_required
def eliminar_ubicacion(request, pk):
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    if request.method == 'POST':
        nombre = ubicacion.nombre
        ubicacion.delete()
        messages.success(request, f"Ubicación '{nombre}' eliminada.")
        return redirect('proyecto_monitoreo_smcv:lista_ubicaciones')
    return redirect('proyecto_monitoreo_smcv:lista_ubicaciones')

# ==========================================
# vista para Agentes
# ==========================================

@login_required
def lista_agente(request):
    agentes = Agente.objects.all().order_by('nombre')
    context = {
        'agentes': agentes,
        'show_sidebar': True
    }
    return render(request, 'agente/lista_agente.html', context)

@login_required
def crear_agente(request):
    if request.method == 'POST':
        form = AgenteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Agente creado correctamente.")
            return redirect('proyecto_monitoreo_smcv:lista_agentes')
    else:
        form = AgenteForm()
    
    context = {
        'form': form,
        'show_sidebar': True
    }
    return render(request, 'agente/crear_agente.html', context)

# ==========================================
# Vistas para Puesto
# ==========================================

@login_required
def lista_puesto(request):
    puestos = Puesto.objects.all().order_by('nombre')
    context = {
        'puestos':puestos,
        'show_sidebar': True
    }
    return render(request, 'puesto/lista_puesto.html', context) 

@login_required
def crear_puesto(request):
    if request.method == 'POST':
        form = PuestoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Puesto creado correctamente.")
            return redirect('proyecto_monitoreo_smcv:lista_puestos')
    else:
        form = PuestoForm()

    context = {
        'form': form,
        'show_sidebar': True 
    }
    
    return render(request, 'puesto/crear_puesto.html', context)

# ==========================================
# Vista para Trabajadores
# ==========================================

@login_required
def gerencia_general_lista(request):
    gerencias = GerenciaGeneral.objects.all().order_by('nombre')
    context = {'gerencias_generales': gerencias, 'show_sidebar': True} # OJO: la variable en el HTML se llama 'gerencias_generales'
    return render(request, 'gerencia_general/lista_gg.html', context)

@login_required
def crear_gerencia_general(request):
    if request.method == 'POST':
        form = GerenciaGeneralForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Gerencia General creada.")
            return redirect('proyecto_monitoreo_smcv:lista_gerencias_generales')
    else:
        form = GerenciaGeneralForm()
    
    context = {'form': form, 'show_sidebar': True}
    return render(request, 'gerencia_general/crear_gg.html', context)

# ==========================================
# Vista para Gerencia
# ==========================================

@login_required
def gerencia_lista(request):
    # Nota: Usamos select_related para traer el nombre del padre optimizado
    gerencias = Gerencia.objects.all().select_related('gerencia_general').order_by('nombre')
    context = {'gerencias': gerencias, 'show_sidebar': True}
    return render(request, 'gerencia/lista_gerencia.html', context)

@login_required
def crear_gerencia(request):
    if request.method == 'POST':
        form = GerenciaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Gerencia creada correctamente.")
            return redirect('proyecto_monitoreo_smcv:lista_gerencias')
    else:
        form = GerenciaForm()
    
    context = {'form': form, 'show_sidebar': True}
    return render(request, 'gerencia/crear_gerencia.html', context)

# ==========================================
# vista para superintendencias
# ==========================================

# Vista para listar superintendencias
@login_required
def superintendencia_lista(request):
    # Traemos también el nombre de la gerencia para la tabla
    sups = Superintendencia.objects.all().select_related('gerencia').order_by('nombre')
    context = {'superintendencias': sups, 'show_sidebar': True}
    return render(request, 'superintendencia/lista_superintendencia.html', context)

# Vista para crear superintendencia
@login_required
def crear_superintendencia(request):
    if request.method == 'POST':
        form = SuperintendenciaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Superintendencia creada correctamente.")
            return redirect('proyecto_monitoreo_smcv:lista_superintendencias')
    else:
        form = SuperintendenciaForm()
    
    context = {'form': form, 'show_sidebar': True}
    return render(request, 'superintendencia/crear_superintendencia.html', context)

# ==========================================
# Vista para Importacion de Plan Mensual
# ==========================================

@login_required
def importar_plan_mensual(request):
    if request.method == 'POST':
        print("\n=== [DEBUG] INICIO PROCESO DE IMPORTACIÓN ===")
        
        form = CargaMasivaPlanForm(request.POST, request.FILES)
        
        if form.is_valid():
            print("--- [DEBUG] Formulario validado correctamente ---")
            
            archivo = request.FILES['archivo_excel']
            mes = form.cleaned_data['mes']
            anio = form.cleaned_data['anio']

            # 1. Guardar archivo temporal
            # Asegúrate de crear la carpeta 'media/tmp_import' manualmente si falla
            save_path = os.path.join(settings.MEDIA_ROOT, 'tmp_import')
            if not os.path.exists(save_path):
                os.makedirs(save_path)
                print(f"--- [DEBUG] Carpeta creada: {save_path} ---")

            fs = FileSystemStorage(location=save_path)
            filename = fs.save(archivo.name, archivo)
            file_path = fs.path(filename)
            
            print(f"--- [DEBUG] Archivo guardado temporalmente en: {file_path} ---")

            try:
                # 2. Leer Excel con Pandas
                print("--- [DEBUG] Intentando leer Excel con Pandas... ---")
                df = pd.read_excel(file_path)
                print(f"--- [DEBUG] Excel leído. Filas encontradas: {len(df)} ---")
                
                # 3. Limpieza de columnas
                print("--- [DEBUG] Limpiando nombres de columnas... ---")
                # Forzamos conversión a string para evitar errores si hay números en los encabezados
                df.columns = df.columns.astype(str).str.strip().str.lower().str.replace(' ', '_')
                
                print(f"--- [DEBUG] Columnas detectadas: {list(df.columns)} ---")

                # 4. Validación de columnas
                required = ['gerencia_general', 'gerencia', 'superintendencia', 'puesto_de_trabajo', 'agente', 'cantidad']
                missing = [c for c in required if c not in df.columns]
                
                if missing:
                    print(f"--- [ERROR] Faltan columnas: {missing} ---")
                    fs.delete(filename)
                    messages.error(request, f"Faltan columnas en el Excel: {', '.join(missing).upper()}")
                    return render(request, 'importar/importar_excel.html', {'form': form, 'show_sidebar': True})

                # 5. Generar Lógica de Resumen (Staging)
                print("--- [DEBUG] Iniciando agrupación de datos para resumen... ---")
                
                # Convertir cantidad a numérico, forzar 0 si hay error
                df['cantidad'] = pd.to_numeric(df['cantidad'], errors='coerce').fillna(0)
                
                # Agrupar por Gerencia General y Agente
                grupo = df.groupby(['gerencia_general', 'agente'])['cantidad'].sum().reset_index()
                
                resumen_data = {}
                total_global = 0

                for _, row in grupo.iterrows():
                    gg = str(row['gerencia_general']).strip()
                    ag = str(row['agente']).strip()
                    cant = int(row['cantidad'])

                    if gg not in resumen_data:
                        resumen_data[gg] = {
                            'agentes': [], 
                            'total_gerencia': 0,
                            'rowspan': 0
                        }
                    
                    resumen_data[gg]['agentes'].append({'nombre': ag, 'cantidad': cant})
                    resumen_data[gg]['total_gerencia'] += cant
                    resumen_data[gg]['rowspan'] += 1
                    total_global += cant

                print(f"--- [DEBUG] Resumen generado. Total global calculado: {total_global} ---")

                context = {
                    'resumen': resumen_data,
                    'total_global': total_global,
                    'filename': filename,
                    'mes': mes,
                    'anio': anio,
                    'show_sidebar': True
                }
                
                print("--- [DEBUG] Renderizando plantilla de previsualización (preview_importacion.html) ---")
                return render(request, 'importar/preview_importacion.html', context)

            except Exception as e:
                print(f"\n!!! [ERROR CRÍTICO EN TRY] !!!: {str(e)}")
                # Borrar archivo si falló
                if os.path.exists(file_path):
                    fs.delete(filename)
                    print("--- [DEBUG] Archivo temporal eliminado por error ---")
                
                messages.error(request, f"Error procesando el archivo: {str(e)}")
                return render(request, 'importar/importar_excel.html', {'form': form, 'show_sidebar': True})

        else:
            print("\n!!! [ERROR] EL FORMULARIO NO ES VÁLIDO !!!")
            print("Errores del formulario:", form.errors)
            # Devolvemos el form con errores para que se vean en pantalla
            return render(request, 'importar/importar_excel.html', {'form': form, 'show_sidebar': True})

    else:
        # GET Request
        form = CargaMasivaPlanForm()

    return render(request, 'importar/importar_excel.html', {'form': form, 'show_sidebar': True})

@login_required
def confirmar_importacion(request):
    if request.method == 'POST':
        filename = request.POST.get('filename')
        mes = request.POST.get('mes')
        anio = request.POST.get('anio')
        
        # Recuperar el archivo temporal
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'tmp_import'))
        file_path = fs.path(filename)

        if not os.path.exists(file_path):
            messages.error(request, "El archivo temporal ha expirado. Sube de nuevo.")
            return redirect('proyecto_monitoreo_smcv:importar_plan')

        conteo = 0
        try:
            # --- AQUÍ VA TU LÓGICA DE GUARDADO (La que ya tenías y funcionaba) ---
            df = pd.read_excel(file_path)
            df.columns = df.columns.astype(str).str.strip().str.lower().str.replace(' ', '_')
            
            with transaction.atomic():
                # (Opcional: Limpiar mes previo si quieres reemplazar)
                # ProgramacionMensual.objects.filter(mes=mes, anio=anio).delete()

                for index, row in df.iterrows():
                    # 1. Crear Estructura Orgánica
                    gg_nombre = str(row['gerencia_general']).strip()
                    gg, _ = GerenciaGeneral.objects.get_or_create(nombre__iexact=gg_nombre, defaults={'nombre': gg_nombre})

                    g_nombre = str(row['gerencia']).strip()
                    g, _ = Gerencia.objects.get_or_create(nombre__iexact=g_nombre, gerencia_general=gg, defaults={'nombre': g_nombre})

                    sup_nombre = str(row['superintendencia']).strip()
                    sup, _ = Superintendencia.objects.get_or_create(nombre__iexact=sup_nombre, gerencia=g, defaults={'nombre': sup_nombre})

                    puesto_nombre = str(row['puesto_de_trabajo']).strip()
                    puesto, _ = Puesto.objects.get_or_create(nombre__iexact=puesto_nombre, superintendencia=sup, defaults={'nombre': puesto_nombre})

                    # 2. Crear Agente
                    agente_nombre = str(row['agente']).strip()
                    agente, _ = Agente.objects.get_or_create(nombre__iexact=agente_nombre, defaults={'nombre': agente_nombre})

                    # 3. Guardar Meta
                    try:
                        cantidad = int(row['cantidad']) if pd.notna(row['cantidad']) else 0
                    except: cantidad = 0

                    ProgramacionMensual.objects.update_or_create(
                        puesto=puesto, agente=agente, mes=mes, anio=anio,
                        defaults={'cantidad_mes': cantidad, 'responsable_mes': None}
                    )
                    conteo += 1
            
            # Borrar archivo temporal tras éxito
            fs.delete(filename)
            messages.success(request, f"✅ Confirmado: Se importaron {conteo} registros al Plan {mes}/{anio}.")
            return redirect('proyecto_monitoreo_smcv:plan_mensual_index')

        except Exception as e:
            messages.error(request, f"Error al guardar en BD: {str(e)}")
            return redirect('proyecto_monitoreo_smcv:importar_plan')
            
    return redirect('proyecto_monitoreo_smcv:importar_plan')

@login_required
def eliminar_plan_mensual(request, anio, mes):
    # Borra todos los registros de ese mes y año
    if request.method == 'POST':
        registros = ProgramacionMensual.objects.filter(anio=anio, mes=mes)
        cantidad = registros.count()
        registros.delete()
        messages.warning(request, f"Se eliminó el plan del mes {mes}/{anio} ({cantidad} registros borrados).")
        return redirect('proyecto_monitoreo_smcv:plan_mensual_index')
    
    return redirect('proyecto_monitoreo_smcv:plan_mensual_index')

@login_required
def tablero_mensual_semanas(request, anio, mes):
    # 1. Obtener las Metas del Mes (Los padres)
    metas_mensuales = ProgramacionMensual.objects.filter(anio=anio, mes=mes)
    total_meta_mes = metas_mensuales.aggregate(Sum('cantidad_mes'))['cantidad_mes__sum'] or 0
    
    # 2. Calcular Total YA Asignado (Suma de todas las semanas de este mes)
    # Buscamos en la tabla Semanal todo lo relacionado a las metas de este mes
    total_asignado = ProgramacionSemanal.objects.filter(
        programacion_mensual__in=metas_mensuales
    ).aggregate(Sum('cantidad_semanal'))['cantidad_semanal__sum'] or 0

    # 3. Calcular Porcentaje de Avance Global
    porcentaje_avance = 0
    if total_meta_mes > 0:
        porcentaje_avance = int((total_asignado / total_meta_mes) * 100)
        # Tope visual 100% (por si se pasan asignando)
        if porcentaje_avance > 100: porcentaje_avance = 100

    # 4. Generar Semanas
    cal = calendar.Calendar(firstweekday=0)
    semanas_del_mes = cal.monthdatescalendar(anio, mes)
    
    tarjetas_semanas = []
    for i, week in enumerate(semanas_del_mes):
        inicio = week[0]
        fin = week[6]
        numero_semana = i + 1
        
        # --- CÁLCULO POR SEMANA ---
        # Sumar cuánto se asignó específicamente a la semana 1, 2, etc.
        carga_semana = ProgramacionSemanal.objects.filter(
            programacion_mensual__in=metas_mensuales,
            numero_semana=numero_semana
        ).aggregate(Sum('cantidad_semanal'))['cantidad_semanal__sum'] or 0

        # Determinar estado
        if carga_semana > 0:
            estado = "Planificado"
            color_estado = "bg-green-100 text-green-800"
        else:
            estado = "Pendiente"
            color_estado = "bg-gray-200 text-gray-600"

        tarjetas_semanas.append({
            'numero': numero_semana,
            'inicio': inicio,
            'fin': fin,
            'estado': estado,
            'color_estado': color_estado, # Pasamos el color para el HTML
            'carga_asignada': carga_semana # El número real
        })

    context = {
        'anio': anio,
        'mes': mes,
        'nombre_mes': date(anio, mes, 1).strftime('%B'),
        'total_meta_mes': total_meta_mes,
        'total_asignado': total_asignado,
        'porcentaje_avance': porcentaje_avance, # Nueva variable para la barra
        'semanas': tarjetas_semanas,
        'show_sidebar': True
    }
    return render(request, 'planificacion_mensual/tablero_semanal.html', context)

# ==========================================
# vista para la planificacion mensual
# ==========================================

@login_required
def plan_mensual_index(request):
    """
    VISTA UNIFICADA: Muestra los planes cargados y botón para importar.
    """
    # Agrupamos por Año y Mes
    planes = ProgramacionMensual.objects.values('anio', 'mes').annotate(
        total_meta=Sum('cantidad_mes'),
        total_puestos=Count('puesto', distinct=True)
    ).order_by('-anio', '-mes') # Los más recientes primero

    context = {
        'planes': planes,
        'show_sidebar': True
    }
    return render(request, 'planificacion_mensual/plan_mensual_index.html', context)

@login_required
def lista_planes_mensuales(request):
    planes = ProgramacionMensual.objects.values('anio', 'mes').annotate(
        total_muestreos=Sum('cantidad_mes'),
        total_puestos=Count('puesto', distinct=True),
        total_registros=Count('id')
    ).order_by('-anio', '-mes')

    context = {
        'planes': planes,
        'show_sidebar': True
    }
    return render(request, 'planificacion/plan_mensual_list.html', context)

@login_required
def asignar_carga_semanal(request, anio, mes, nro_semana):
    """
    Vista para gestionar la carga de trabajo de una semana específica.
    Permite edición masiva y filtrado por Superintendencia/Agente.
    """
    
    # 1. CALCULAR FECHAS DE LA SEMANA
    cal = calendar.Calendar(firstweekday=0) # 0 = Lunes
    try:
        semanas_del_mes = cal.monthdatescalendar(anio, mes)
        
        # Validar que el número de semana exista
        if nro_semana < 1 or nro_semana > len(semanas_del_mes):
            messages.error(request, "Número de semana inválido.")
            return redirect('proyecto_monitoreo_smcv:tablero_mensual', anio=anio, mes=mes)

        # Índices de lista empiezan en 0, por eso restamos 1
        fecha_inicio = semanas_del_mes[nro_semana - 1][0]
        fecha_fin = semanas_del_mes[nro_semana - 1][6]
    except Exception as e:
        messages.error(request, f"Error calculando fechas: {e}")
        return redirect('proyecto_monitoreo_smcv:plan_mensual_index')

    # 2. OBTENER METAS DEL MES (OPTIMIZADO)
    # Usamos select_related para traer los nombres de gerencias/puestos en una sola consulta
    metas_mensuales = ProgramacionMensual.objects.filter(anio=anio, mes=mes).select_related(
        'puesto__superintendencia__gerencia', 
        'puesto', 
        'agente'
    ).order_by('puesto__superintendencia__nombre', 'puesto__nombre')

    # --- PROCESO DE GUARDADO (POST) ---
    if request.method == 'POST':
        guardados = 0
        try:
            with transaction.atomic():
                for meta in metas_mensuales:
                    # Buscamos el input específico para esta meta (ej: name="meta_55")
                    input_name = f"meta_{meta.id}"
                    cantidad_input = request.POST.get(input_name)

                    # Si hay un valor válido, guardamos/actualizamos
                    if cantidad_input and cantidad_input.strip() != "":
                        cantidad = int(cantidad_input)
                        
                        if cantidad > 0:
                            ProgramacionSemanal.objects.update_or_create(
                                programacion_mensual=meta,
                                numero_semana=nro_semana,
                                defaults={
                                    'cantidad_semanal': cantidad,
                                    'fecha_inicio': fecha_inicio,
                                    'fecha_fin': fecha_fin,
                                    # Asignamos al usuario actual como responsable del cambio
                                    'responsable_semana': None
                                }
                            )
                            guardados += 1
                        else:
                            # Si el usuario puso 0, borramos el registro si existía
                            ProgramacionSemanal.objects.filter(programacion_mensual=meta, numero_semana=nro_semana).delete()
                    
                    # Si el input vino vacío pero antes había dato, también podríamos borrar,
                    # pero en este caso asumimos que vacío = no tocar o borrar (según tu lógica).
                    # Aquí asumo vacío = borrar para limpiar rápido.
                    elif cantidad_input == "":
                         ProgramacionSemanal.objects.filter(programacion_mensual=meta, numero_semana=nro_semana).delete()

            messages.success(request, f"✅ Se guardaron {guardados} asignaciones para la Semana {nro_semana}.")
            return redirect('proyecto_monitoreo_smcv:tablero_mensual', anio=anio, mes=mes)

        except Exception as e:
            messages.error(request, f"Error al guardar datos: {str(e)}")
            # No redirigimos para no perder lo que el usuario escribió, recargamos la página
    
    # --- PREPARACIÓN DE DATOS PARA LA TABLA (GET) ---
    lista_asignacion = []
    
    # Sets para los filtros desplegables (evitan duplicados)
    filtros_super = set()
    filtros_agente = set()

    for meta in metas_mensuales:
        # Datos para los filtros del HTML
        nombre_super = meta.puesto.superintendencia.nombre
        nombre_agente = meta.agente.nombre
        
        filtros_super.add(nombre_super)
        filtros_agente.add(nombre_agente)

        # 1. ¿Cuánto tiene asignado ESTA semana?
        # Optimizacion: Podríamos hacer esto fuera del bucle con diccionarios, 
        # pero para 1500 filas esto sigue siendo aceptable.
        try:
            semanal = ProgramacionSemanal.objects.get(programacion_mensual=meta, numero_semana=nro_semana)
            valor_actual = semanal.cantidad_semanal
        except ProgramacionSemanal.DoesNotExist:
            valor_actual = 0 # Campo vacío en el input

        # 2. ¿Cuánto se ha gastado en OTRAS semanas?
        otros_semanales = ProgramacionSemanal.objects.filter(programacion_mensual=meta).exclude(numero_semana=nro_semana)
        total_otros = otros_semanales.aggregate(Sum('cantidad_semanal'))['cantidad_semanal__sum'] or 0
        
        # 3. Saldo disponible
        saldo = meta.cantidad_mes - total_otros

        lista_asignacion.append({
            'meta_obj': meta,
            'valor_actual': valor_actual,
            'saldo': saldo,
            'total_otros': total_otros,
            # Variables para el Javascript de filtrado
            'js_super': nombre_super,
            'js_agente': nombre_agente,
        })

    context = {
        'anio': anio,
        'mes': mes,
        'nro_semana': nro_semana,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'lista_asignacion': lista_asignacion,
        # Enviamos las listas ordenadas alfabéticamente para los <select>
        'filtros_super': sorted(list(filtros_super)),
        'filtros_agente': sorted(list(filtros_agente)),
        'show_sidebar': True
    }
    
    return render(request, 'planificacion_mensual/asignar_carga.html', context)

# ==========================================
# vista para la planificacion semanal
# ==========================================

@login_required
def control_semanal_dashboard(request, anio=None, mes=None, semana=None):
    today = datetime.date.today()
    
    # 1. Si no hay parámetros, usar fecha actual
    if not anio: anio = today.year
    if not mes: mes = today.month
    
    # 2. Calcular Semanas del mes para las pestañas
    cal = calendar.Calendar(firstweekday=0)
    weeks = cal.monthdatescalendar(anio, mes)
    
    # Si no especifican semana, buscamos cuál es la semana actual (hoy)
    if not semana:
        for i, week_dates in enumerate(weeks):
            if today in week_dates:
                semana = i + 1
                break
        if not semana: semana = 1 # Fallback a semana 1

    # Validar rango de semana
    if semana < 1 or semana > len(weeks): semana = 1
    
    # Fechas de inicio y fin de la semana seleccionada
    fecha_inicio_sem = weeks[semana-1][0]
    fecha_fin_sem = weeks[semana-1][6]

    # 3. CONSULTA PRINCIPAL (El corazón del dashboard)
    # Buscamos la programación de esa semana y le "pegamos" la suma de lo ejecutado
    planes_semanales = ProgramacionSemanal.objects.filter(
        programacion_mensual__anio=anio,
        programacion_mensual__mes=mes,
        numero_semana=semana
    ).select_related(
        'programacion_mensual__puesto__superintendencia',
        'programacion_mensual__agente'
    ).annotate(
        # Sumamos la cantidad_real_ejecutada de los detalles diarios vinculados
        total_ejecutado=Coalesce(Sum('detalles_diarios__cantidad_real_ejecutada'), 0)
    ).order_by('programacion_mensual__puesto__superintendencia__nombre')

    # 4. Procesar datos para el template (Calcular % y Estados)
    tabla_control = []
    total_meta = 0
    total_real = 0

    for plan in planes_semanales:
        meta = plan.cantidad_semanal
        real = plan.total_ejecutado
        pendiente = meta - real
        
        # Porcentaje para la barra (tope 100%)
        porcentaje = int((real / meta) * 100) if meta > 0 else 0
        if porcentaje > 100: porcentaje = 100
        
        # Color del estado
        if real >= meta:
            estado_color = "bg-green-100 text-green-800"
            estado_texto = "Completado"
        elif real == 0:
            estado_color = "bg-red-100 text-red-800"
            estado_texto = "Sin Avance"
        else:
            estado_color = "bg-yellow-100 text-yellow-800"
            estado_texto = "En Progreso"

        tabla_control.append({
            'obj': plan,
            'puesto': plan.programacion_mensual.puesto,
            'agente': plan.programacion_mensual.agente,
            'meta': meta,
            'real': real,
            'pendiente': pendiente,
            'porcentaje': porcentaje,
            'estado_color': estado_color,
            'estado_texto': estado_texto
        })
        
        total_meta += meta
        total_real += real

    # Nombres para navegación
    nombre_mes = datetime.date(anio, mes, 1).strftime('%B').capitalize()
    
    # Navegación de meses
    prev_mes = mes - 1 if mes > 1 else 12
    prev_anio = anio if mes > 1 else anio - 1
    next_mes = mes + 1 if mes < 12 else 1
    next_anio = anio if mes < 12 else anio + 1

    context = {
        'anio': anio, 'mes': mes, 'semana_actual': semana,
        'nombre_mes': nombre_mes,
        'semanas_nav': range(1, len(weeks)+1), # Para las pestañas
        'tabla_control': tabla_control,
        'total_meta': total_meta,
        'total_real': total_real,
        'fecha_inicio': fecha_inicio_sem,
        'fecha_fin': fecha_fin_sem,
        'prev_mes': prev_mes, 'prev_anio': prev_anio,
        'next_mes': next_mes, 'next_anio': next_anio,
        'show_sidebar': True
    }

    return render(request, 'planificacion_semanal/semanal_dashboard.html', context)

@login_required
def gestion_dias_semana(request, anio, mes, nro_semana):
    """
    Muestra los 7 días de la semana seleccionada para entrar a programar el día.
    """
    # 1. Calcular fechas de esta semana
    cal = calendar.Calendar(firstweekday=0)
    semanas_del_mes = cal.monthdatescalendar(anio, mes)
    
    if nro_semana < 1 or nro_semana > len(semanas_del_mes):
        return redirect('proyecto_monitoreo_smcv:tablero_mensual', anio=anio, mes=mes)

    # Obtenemos la lista de fechas (Lunes a Domingo)
    fechas_semana = semanas_del_mes[nro_semana - 1]
    
    # 2. Verificar qué días ya tienen programación creada
    dias_data = []
    for fecha in fechas_semana:
        # Buscamos si existe una ProgramacionDiaria para esta fecha
        prog_diaria = ProgramacionDiaria.objects.filter(fecha_programacion=fecha).first()
        
        estado = "Sin Programar"
        monitoristas = 0
        id_diario = None
        
        if prog_diaria:
            estado = "Programado"
            id_diario = prog_diaria.id
            # Contamos cuántas tareas hay ese día
            monitoristas = prog_diaria.detalles.count()

        dias_data.append({
            'fecha': fecha,
            'dia_nombre': fecha.strftime("%A"), # Lunes, Martes...
            'estado': estado,
            'tareas': monitoristas,
            'id_diario': id_diario
        })

    context = {
        'anio': anio, 'mes': mes, 'nro_semana': nro_semana,
        'dias': dias_data,
        'show_sidebar': True
    }
    return render(request, 'planificacion/gestion_dias.html', context)

@login_required
def crear_programacion_diaria(request, fecha_str):
    """
    Vista principal de operación diaria.
    Permite asignar tareas de la 'Bolsa Semanal' a trabajadores específicos en un día.
    """
    # Convertir el string de la URL (YYYY-MM-DD) a objeto fecha real
    try:
        fecha = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return redirect('proyecto_monitoreo_smcv:programacion_diaria_index')
    
    # 1. BUSCAR O CREAR LA CABECERA DEL DÍA
    # Esto agrupa todos los trabajos de hoy. Si no existe, se crea sola.
    prog_diaria, created = ProgramacionDiaria.objects.get_or_create(
        fecha_programacion=fecha,
        defaults={
            # Intentamos asignar al usuario logueado si tiene trabajador asociado, si no, None
            'responsable_diario': getattr(request.user, 'trabajador_set', None) and request.user.trabajador_set.first() or None
        }
    )

    # --- LÓGICA DE GUARDADO (POST) ---
    if request.method == 'POST':
        # Obtener datos del formulario HTML
        plan_semanal_id = request.POST.get('plan_semanal_id')
        monitoristas_ids = request.POST.getlist('monitoristas') # Lista de IDs
        hora = request.POST.get('hora')
        lugar = request.POST.get('lugar')
        cantidad = int(request.POST.get('cantidad_programada', 1))
        
        # El checkbox o input hidden envía 'True' o 'False' como texto
        es_planeado = request.POST.get('es_planeado') == 'True'

        # Datos para NO PLANEADOS (Si implementaste los selects en el HTML)
        puesto_id_manual = request.POST.get('puesto_manual')
        agente_id_manual = request.POST.get('agente_manual')

        try:
            with transaction.atomic():
                # Instancia base del detalle
                detalle = DetalleDiario(
                    programacion=prog_diaria,
                    lugar=lugar,
                    hora=hora if hora else None, # Si viene vacío, guarda Null
                    cantidad_programada=cantidad,
                    es_planeado=es_planeado,
                    estado_tarea='PENDIENTE',
                    cantidad_real_ejecutada=0
                )

                if es_planeado:
                    # CASO A: Tarea Planeada (Viene de la Meta Semanal)
                    if not plan_semanal_id:
                        raise ValueError("Falta seleccionar una tarea de la lista.")
                    
                    plan_semanal = ProgramacionSemanal.objects.get(id=plan_semanal_id)
                    
                    # Vinculamos y heredamos datos automáticamente
                    detalle.plan_semanal = plan_semanal 
                    detalle.puesto = plan_semanal.programacion_mensual.puesto
                    detalle.agente = plan_semanal.programacion_mensual.agente
                
                else:
                    # CASO B: Tarea No Planeada (Emergencia)
                    # Aquí necesitamos que el usuario haya elegido puesto y agente manualmente
                    if puesto_id_manual and agente_id_manual:
                        detalle.puesto_id = puesto_id_manual
                        detalle.agente_id = agente_id_manual
                        detalle.plan_semanal = None # No descuenta meta
                    else:
                        # Si el formulario HTML aún no tiene los selects para no planeados, 
                        # lanzamos error o hardcodeamos para pruebas.
                        raise ValueError("Para tareas no planeadas debes seleccionar Puesto y Agente.")

                # Guardamos la tarea
                detalle.save()
                
                # Asignar los trabajadores (Many-to-Many siempre va DESPUÉS del save)
                if monitoristas_ids:
                    detalle.monitoristas_asignados.set(monitoristas_ids)

            messages.success(request, "Tarea asignada correctamente.")
            # Recargamos la misma página para ver los cambios
            return redirect('proyecto_monitoreo_smcv:crear_prog_diaria', fecha_str=fecha_str)

        except Exception as e:
            messages.error(request, f"Error al guardar: {e}")
            # No redirigimos para que el usuario no pierda datos, el render de abajo se encarga

    # --- PREPARACIÓN DE DATOS (GET) ---

    # 2. CALCULAR TAREAS DISPONIBLES (BOLSA SEMANAL)
    # Buscamos planes semanales que incluyan la fecha actual
    tareas_semanales = ProgramacionSemanal.objects.filter(
        fecha_inicio__lte=fecha, 
        fecha_fin__gte=fecha
    ).select_related(
        'programacion_mensual__puesto__superintendencia', # Optimización SQL
        'programacion_mensual__agente'
    )

    tareas_disponibles = []
    
    for tarea in tareas_semanales:
        # A. Cuánto se programó para esta semana (La meta)
        total_meta_semana = tarea.cantidad_semanal
        
        # B. Cuánto ya se ha asignado en días anteriores o hoy (Lo gastado)
        ya_asignado = DetalleDiario.objects.filter(
            plan_semanal=tarea
        ).aggregate(Sum('cantidad_programada'))['cantidad_programada__sum'] or 0
        
        # C. Saldo
        saldo = total_meta_semana - ya_asignado
        
        # Solo mostramos si queda saldo positivo
        if saldo > 0:
            tareas_disponibles.append({
                'tarea': tarea,
                'saldo': saldo
            })

    # 3. LISTAS PARA FORMULARIOS
    trabajadores = Trabajador.objects.all().order_by('nombre')

    cal = calendar.Calendar(firstweekday=0)
    semanas_del_mes = cal.monthdatescalendar(fecha.year, fecha.month)
    
    nro_semana_actual = 1
    for i, semana in enumerate(semanas_del_mes):
        if fecha in semana:
            nro_semana_actual = i + 1
            break
    
    # Opcional: Listas completas por si quieres implementar el "No Planeado" ahora
    puestos_all = Puesto.objects.all().order_by('nombre')
    agentes_all = Agente.objects.all().order_by('nombre')

    context = {
        'fecha': fecha,
        'prog_diaria': prog_diaria,
        'tareas_disponibles': tareas_disponibles, # Lista izquierda
        'trabajadores': trabajadores,             # Select de monitoristas
        'puestos_all': puestos_all,               # Para select No Planeado (futuro)
        'agentes_all': agentes_all,               # Para select No Planeado (futuro)
        'show_sidebar': True,
        'anio_back': fecha.year,
        'mes_back': fecha.month,
        'semana_back': nro_semana_actual
    }
    
    return render(request, 'planificacion_semanal/form_diario.html', context)

@login_required
def programacion_diaria_index(request):
    hoy = datetime.date.today()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    
    cal = calendar.Calendar(firstweekday=0)
    semanas_calendario = cal.monthdatescalendar(anio, mes)
    
    # --- CÁLCULO CORREGIDO ---
    planes_del_mes = ProgramacionSemanal.objects.filter(
        programacion_mensual__anio=anio,
        programacion_mensual__mes=mes
    ).values('numero_semana').annotate(
        meta=Sum('cantidad_semanal'),
        
        # 1. CORRECCIÓN AQUÍ: Usamos 'detalles_diarios'
        real=Coalesce(Sum('detalles_diarios__cantidad_real_ejecutada'), 0),
        
        # 2. CORRECCIÓN AQUÍ: Usamos 'detalles_diarios'
        asignado=Coalesce(Sum('detalles_diarios__cantidad_programada'), 0)
    )
    
    datos = {p['numero_semana']: p for p in planes_del_mes}

    lista_semanas = []
    for i, dias in enumerate(semanas_calendario):
        num = i + 1
        info = datos.get(num, {'meta': 0, 'real': 0, 'asignado': 0})
        
        meta = info['meta']
        real = info['real']
        asignado = info['asignado']
        
        # Calcular porcentajes
        pct_real = int((real / meta) * 100) if meta > 0 else 0
        if pct_real > 100: pct_real = 100
        
        pct_asignado = int((asignado / meta) * 100) if meta > 0 else 0
        if pct_asignado > 100: pct_asignado = 100
        
        lista_semanas.append({
            'numero': num,
            'inicio': dias[0],
            'fin': dias[6],
            'meta': meta,
            'real': real,
            'asignado': asignado,
            'pct_real': pct_real,
            'pct_asignado': pct_asignado
        })

    nombre_mes = datetime.date(anio, mes, 1).strftime('%B').capitalize()
    
    prev_mes = mes - 1 if mes > 1 else 12
    prev_anio = anio if mes > 1 else anio - 1
    next_mes = mes + 1 if mes < 12 else 1
    next_anio = anio if mes < 12 else anio + 1

    context = {
        'anio': anio, 'mes': mes, 
        'nombre_mes': nombre_mes,
        'semanas': lista_semanas,
        'show_sidebar': True,
        'prev_mes': prev_mes, 'prev_anio': prev_anio,
        'next_mes': next_mes, 'next_anio': next_anio
    }
    return render(request, 'planificacion_semanal/diaria_tarjetas.html', context)

@login_required
def vista_semana_operativa(request, anio, mes, nro_semana):
    """
    NIVEL 2: Vista de los 7 días de la semana seleccionada.
    """
    cal = calendar.Calendar(firstweekday=0)
    try:
        semanas = cal.monthdatescalendar(anio, mes)
        dias_semana = semanas[nro_semana - 1] # Lista de 7 objetos date
    except IndexError:
        return redirect('proyecto_monitoreo_smcv:programacion_diaria_index')

    detalle_dias = []
    hoy = datetime.date.today()

    for fecha in dias_semana:
        # Verificar estado del día
        prog = ProgramacionDiaria.objects.filter(fecha_programacion=fecha).first()
        
        tareas_count = 0
        estado = "Libre"
        
        if prog:
            tareas_count = prog.detalles.count()
            if tareas_count > 0:
                estado = "Programado"
        
        detalle_dias.append({
            'fecha': fecha,
            'es_hoy': (fecha == hoy),
            'dia_nombre': fecha.strftime("%A"), # Lunes, Martes...
            'estado': estado,
            'tareas': tareas_count
        })

    context = {
        'anio': anio, 'mes': mes, 'nro_semana': nro_semana,
        'dias': detalle_dias,
        'fecha_inicio': dias_semana[0],
        'fecha_fin': dias_semana[6],
        'show_sidebar': True
    }
    return render(request, 'planificacion_semanal/semana_operativa.html', context)

# ==========================================
# SOPORTE TÉCNICO
# ==========================================     

# ------ APIS --------

@require_POST
def actualizar_dispositivo_api(request):
    try:
        data = json.loads(request.body)
        dispositivo_id = data.get('id')
        campo = data.get('campo') # 'ubicacion', 'fecha_bajada' o 'fecha_subida'
        valor = data.get('valor')

        device = Dispositivo.objects.get(pk=dispositivo_id)

        # Validación simple y asignación
        if campo == 'ubicacion':
            # ubicacion es FK, buscamos o creamos por nombre
            from .models import Ubicacion
            try:
                ubicacion_obj = Ubicacion.objects.get(nombre=valor)
                device.ubicacion = ubicacion_obj
            except Ubicacion.DoesNotExist:
                return JsonResponse({'status': 'error', 'mensaje': f'Ubicación "{valor}" no encontrada'}, status=404)
        elif campo == 'observacion':
            device.observacion = valor
        elif campo == 'estado':
            device.estado = valor
        elif campo in ['fecha_bajada', 'fecha_subida']:
            # Si el valor viene vacío, ponemos None (null en BD)
            if not valor:
                setattr(device, campo, None)
            else:
                # Aseguramos formato fecha
                setattr(device, campo, valor)
        
        device.save()
        return JsonResponse({'status': 'success', 'mensaje': 'Actualizado correctamente'})

    except Dispositivo.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Dispositivo no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)

@require_POST
def actualizar_accesorio_api(request):
    try:
        data = json.loads(request.body)
        accesorio_id = data.get('id')
        campo = data.get('campo') # eje_x, eje_y, eje_z
        valor = data.get('valor')

        # Buscamos el accesorio
        acc = Accesorio.objects.get(pk=accesorio_id)

        # Solo permitimos editar estos campos específicos
        if campo in ['eje_x', 'eje_y', 'eje_z', 'serie']:
            setattr(acc, campo, valor)
            acc.save()
            return JsonResponse({'status': 'success', 'mensaje': 'Sensibilidad guardada'})
        
        return JsonResponse({'status': 'error', 'mensaje': 'Campo no válido'})

    except Accesorio.DoesNotExist:
        return JsonResponse({'status': 'error', 'mensaje': 'Accesorio no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)

@require_POST
def actualizar_cambio_api(request):
    try:
        data = json.loads(request.body)
        dispositivo_id = data.get('id')
        componente_key = data.get('campo') 
        valor_fecha = data.get('valor')

        if not valor_fecha:
            # Si borran la fecha, quizás quieras borrar el último registro o no hacer nada
            pass 
        else:
            try:
                accesorio = Accesorio.objects.get(dispositivo_id=dispositivo_id, nombre=componente_key)
                
                # Buscamos el registro MÁS RECIENTE de ese accesorio
                ultimo_cambio = HistorialCambioAccesorio.objects.filter(
                    accesorio=accesorio
                ).order_by('-fecha_cambio').first()

                if ultimo_cambio:
                    ultimo_cambio.fecha_cambio = valor_fecha
                    ultimo_cambio.save()
                else:
                    HistorialCambioAccesorio.objects.create(
                        accesorio=accesorio,
                        fecha_cambio=valor_fecha
                    )
            except Accesorio.DoesNotExist:
                return JsonResponse({'status': 'error', 'mensaje': f'Accesorio "{componente_key}" no encontrado'}, status=404)

        return JsonResponse({'status': 'success', 'mensaje': 'Fecha actualizada'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)

@require_POST
def actualizar_stock_api(request):
    try:
        data = json.loads(request.body)
        repuesto_id = data.get('id')
        accion = data.get('accion') # 'sumar' o 'restar'
        cantidad = int(data.get('cantidad', 1))

        repuesto = InventarioRepuestos.objects.get(pk=repuesto_id)

        if accion == 'sumar':
            repuesto.stock += cantidad
        elif accion == 'restar':
            if repuesto.stock >= cantidad:
                repuesto.stock -= cantidad
            else:
                return JsonResponse({'status': 'error', 'mensaje': 'Stock insuficiente'})
        
        repuesto.save()
        return JsonResponse({'status': 'success', 'nuevo_stock': repuesto.stock})

    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)

@require_POST
def toggle_ejecucion_api(request):
    try:
        data = json.loads(request.body)
        fila_id = data.get('id')
        estado = data.get('estado') # true o false

        fila = FilaCronograma.objects.get(pk=fila_id)
        fila.ejecutado = estado
        fila.save()

        return JsonResponse({'status': 'success', 'nuevo_estado': fila.ejecutado})
    except Exception as e:
        return JsonResponse({'status': 'error', 'mensaje': str(e)}, status=500)

    
# ------ FORMULARIOS --------

class DiagnosticoCreateView(LoginRequiredMixin, CreateView):
    model = Diagnostico
    form_class = DiagnosticoForm
    template_name = 'proyecto_monitoreo_smcv/soporte/diagnostico_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:soporte_index')

    def form_valid(self, form):
        # Usamos logger.warning para asegurar que salga en los logs de Render (a veces INFO se oculta)
        logger.warning("================ INICIO DEPURACIÓN (LOGS) ===================")
        
        try:
            self.object = form.save()
            logger.warning(f"[OK] Diagnóstico creado con ID: {self.object.id}")
        except Exception as e:
            logger.error(f"[ERROR CRITICO] Falló al crear el Diagnóstico: {e}")
            return self.form_invalid(form)
        
        desc_fotos = form.cleaned_data.get('descripcion_fotos')
        if not desc_fotos:
            desc_fotos = "Evidencia fotográfica de ingreso"

        # Verificar archivos
        files = self.request.FILES.getlist('imagenes')
        logger.warning(f"[INFO] Cantidad de imágenes recibidas: {len(files)}")

        if len(files) == 0:
            logger.error("[ALERTA] No se recibieron archivos. Verifica el enctype del form HTML.")

        # Subida
        for f in files:
            logger.warning(f"   -> Procesando archivo: {f.name} ({f.size} bytes)")
            try:
                foto = FotoDiagnostico.objects.create(
                    diagnostico=self.object,
                    img=f,
                    descripcion=desc_fotos,
                )
                # Si esto se imprime, la BD guardó la referencia. 
                # Si usas S3, verifica si la URL empieza con https://tu-bucket...
                logger.warning(f"   -> [EXITO] Foto guardada. URL: {foto.img.url}")
            except Exception as e:
                # Este es el error que buscamos
                logger.error(f"   -> [FALLO SUBIDA] Error: {e}")

        logger.warning("================ FIN DEPURACIÓN ===================")
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        logger.warning("=== [LOG] Entró a form_invalid (HAY ERRORES) ===")
        logger.warning("Django rechazó el formulario por esto:")
        
        # Esto imprimirá en los logs de Render exactamente qué campo falla
        for field, errors in form.errors.items():
            logger.error(f"   -> CAMPO: '{field}' | ERROR: {errors}")
            
        return super().form_invalid(form)

class MantenimientoCreateView(LoginRequiredMixin, CreateView):
    model = Mantenimiento
    form_class = MantenimientoForm
    template_name = 'proyecto_monitoreo_smcv/soporte/mantenimiento_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:soporte_index')

class OperatividadCreateView(LoginRequiredMixin, CreateView):
    model = Operatividad
    form_class = OperatividadForm
    template_name = 'proyecto_monitoreo_smcv/soporte/operatividad_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:soporte_index')

    # LISTA DE PREGUNTAS FIJAS (Según tu PDF)
    preguntas_fijas = [
        "Case en buen estado",
        "Botón de encendido/apagado y teclas en buen estado",
        "Puerto de carga en buen estado",
        "Pantalla en buen estado",
        "Puerto de batería en buen estado",
        "Se logro ingresar a opciones de programación satisfactoriamente",
        "Se realizo la descarga de data",
        "Accesorio en buen estado"
    ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pasamos las preguntas al HTML para dibujarlas
        context['preguntas_lista'] = self.preguntas_fijas
        return context

    def form_valid(self, form):
        # 1. Guardar la Operatividad (Padre)
        self.object = form.save()
        
        # 2. Guardar el Checklist (Verificaciones)
        # Recorremos la lista de preguntas y buscamos su respuesta en el POST
        # En el HTML los inputs se llamarán 'respuesta_0', 'respuesta_1', etc.
        for i, pregunta in enumerate(self.preguntas_fijas):
            respuesta = self.request.POST.get(f'respuesta_{i}')
            
            # Si no marcaron nada, guardamos un guion o N.A por defecto
            if not respuesta: 
                respuesta = '-'
                
            Verificacion.objects.create(
                operatividad=self.object,
                pregunta=pregunta,
                respuesta=respuesta
            )

        # 3. Guardar Evidencias (Fotos + Descripción individual)
        # En el HTML usaremos inputs arrays: name="fotos[]" y name="descripciones[]"
        fotos = self.request.FILES.getlist('fotos[]')
        descripciones = self.request.POST.getlist('descripciones[]')
        
        # Iteramos asumiendo que el orden coincide
        for i, foto in enumerate(fotos):
            # Intentamos obtener la descripción correspondiente, si no, cadena vacía
            desc = descripciones[i] if i < len(descripciones) else ""
            
            Evidencias.objects.create(
                operatividad=self.object,
                img=foto,
                descripcion=desc
            )

        return redirect(self.get_success_url())

class CalibracionCreateView(LoginRequiredMixin, CreateView):
    model = Calibracion
    form_class = CalibracionForm
    template_name = 'proyecto_monitoreo_smcv/soporte/calibracion_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:soporte_index')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Enviamos todos los dispositivos para que puedan ser elegidos como patrón
        # Opcional: Podrías filtrar por .filter(es_patron=True) si agregas ese campo al modelo
        context['posibles_patrones'] = Dispositivo.objects.filter(nombre__icontains='Calibrador').order_by('nombre')
        return context

    def form_valid(self, form):
        self.object = form.save(commit=False)
        
        # 2. CALCULAR PRÓXIMA FECHA (+1 AÑO)
        fecha_actual = self.object.fecha_calibracion
        try:
            # Intentamos sumar 1 al año
            proxima = fecha_actual.replace(year=fecha_actual.year + 1)
        except ValueError:
            # Si es 29 de febrero y el próximo año no es bisiesto, ponemos 28 feb
            proxima = fecha_actual.replace(year=fecha_actual.year + 1, month=2, day=28)
            
        self.object.fecha_proxima = proxima
        # 1. Guardar la Cabecera
        self.object.save()
        
        patron_id = self.request.POST.get('patron_id')
        trazabilidad = self.request.POST.get('patron_trazabilidad', 'Estándar')
        
        if patron_id:
            equipo_patron = Dispositivo.objects.get(pk=patron_id)
            
            # 2. GUARDAR RELACIÓN EN BD
            Patrones.objects.create(
                calibracion=self.object,
                equipo_patron=equipo_patron,
                trazabilidad=trazabilidad
            )

        # 3. GUARDAR FOTOS DE PATRONES (Manual)
        fotos = self.request.FILES.getlist('fotos_patron')
        for f in fotos:
            FotoPatron.objects.create(calibracion=self.object, img=f)
        

        print("\n" + "="*50)
        print(" INICIO DEPURACIÓN MATEMÁTICA CALIBRACIÓN")
        print("="*50)

        # 4. PROCESAR MEDICIONES Y CALCULOS MATEMÁTICOS
        # Flujos Nominales fijos: 1, 2, 3, 4, 5
        # Valores fijos de tu Excel (DMS, u_p, etc)
        # u_p extraídos de tu imagen fila 42 aprox (ajustar si varían)
        constantes_up = [
            Decimal('0.00704911341943089'), Decimal('0.0172078470472049'), Decimal('0.00877895210147534'), 
            Decimal('0.0172078470472046'), Decimal('0.0070491134194308')
        ]
        flujos_nominales = [1, 2, 3, 4, 5]
        
        # Constantes fijas
        u_res = Decimal('0.054')
        factor_student = Decimal('1.11')
        raiz_n = Decimal(math.sqrt(3)) # 1.732...

        for i, flujo_nom in enumerate(flujos_nominales):
            # 1. Obtener Inputs y convertir a Decimal
            try:
                m1 = Decimal(self.request.POST.get(f'medicion_{i}_0', 0))
                m2 = Decimal(self.request.POST.get(f'medicion_{i}_1', 0))
                m3 = Decimal(self.request.POST.get(f'medicion_{i}_2', 0))
                mediciones = [m1, m2, m3]
            except:
                mediciones = [Decimal(0), Decimal(0), Decimal(0)]

            # 2. PROMEDIO
            promedio = sum(mediciones) / Decimal(3)

            # 3. DESVIACIÓN ESTÁNDAR
            # La librería statistics de Python maneja Decimals correctamente
            try:
                desv_std = statistics.stdev(mediciones)
            except:
                desv_std = Decimal(0)

            # 4. CÁLCULOS
            # u rep = (s * 1.11) / 1.732...
            u_rep = (desv_std * factor_student) / raiz_n
            
            # u p
            u_p = constantes_up[i]

            # Suma Cuadrados (u^2)
            suma_cuadrados = (u_res**2) + (u_rep**2) + (u_p**2)

            # u combinada (Raiz)
            # Decimal.sqrt() es más preciso que math.sqrt()
            u_combinada = suma_cuadrados.sqrt()

            # Incertidumbre Expandida (k=2)
            incertidumbre_final = u_combinada * Decimal(2)

            # --- REDONDEO ESTILO EXCEL ---
            # Python 'round()' redondea 0.005 a 0.00 (par). Excel a 0.01 (arriba).
            # Usamos quantize para forzar estilo Excel.
            
            def redondear_excel(valor):
                return valor.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)

            promedio_guardar = redondear_excel(promedio)
            desviacion_guardar = redondear_excel(desv_std)
            incertidumbre_guardar = redondear_excel(incertidumbre_final)

            condicion = "Pasó" # Lógica placeholder

            # Guardar en Base de Datos
            Resultados.objects.create(
                calibracion=self.object,
                flujo_nominal=flujo_nom,
                flujo_establecido=flujo_nom, 
                lectura_promedio=promedio_guardar, # Fila 46
                
                # AQUÍ ESTABA LA CONFUSIÓN: Ahora guardamos la Desv Std (Fila 39)
                desviacion=desviacion_guardar, 
                
                icertidumbre=incertidumbre_guardar, # Fila 45
                condicion=condicion,
                m1=mediciones[0],
                m2=mediciones[1],
                m3=mediciones[2]
            )

        return redirect(self.get_success_url())
    
class HistorialCambioAccesorioCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = HistorialCambioAccesorio
    form_class = HistorialCambioAccesorioForm
    template_name = 'proyecto_monitoreo_smcv/soporte/cambios/cambio_componente_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:vibrometros_view')
    group_required = 'Yeni_Soporte'

    def form_valid(self, form):
        self.object = form.save()

        if self.object.accesorio and self.object.accesorio.codigo_inventario:
            try:
                repuesto = InventarioRepuestos.objects.get(codigo=self.object.accesorio.codigo_inventario)
                if repuesto.stock > 0:
                    repuesto.stock -= 1
                    repuesto.save()
                    messages.success(self.request, f"Cambio registrado y stock descontado para: {repuesto.articulo} (Quedan: {repuesto.stock})")
                else:
                    messages.warning(self.request, f"Cambio registrado, pero NO HABÍA STOCK de {repuesto.articulo} para descontar.")
            except InventarioRepuestos.DoesNotExist:
                messages.error(self.request, f"Error: No se encontró el repuesto con código '{self.object.accesorio.codigo_inventario}' en el inventario.")
        
        return redirect(self.success_url)

class CambioDosimetroCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = HistorialCambioAccesorio
    form_class = CambioDosimetroForm
    template_name = 'proyecto_monitoreo_smcv/soporte/cambios/cambio_dosimetro_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:dosimetros_view')
    group_required = 'Yeni_Soporte'

    def form_valid(self, form):
        self.object = form.save()

        if self.object.accesorio and self.object.accesorio.codigo_inventario:
            try:
                repuesto = InventarioRepuestos.objects.get(codigo=self.object.accesorio.codigo_inventario)
                if repuesto.stock > 0:
                    repuesto.stock -= 1
                    repuesto.save()
                    messages.success(self.request, f"Stock descontado: {repuesto.articulo}")
                else:
                    messages.warning(self.request, f"Sin stock de {repuesto.articulo}")
            except InventarioRepuestos.DoesNotExist:
                pass 
        
        return redirect(self.success_url)
    
class CambioBombaCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = HistorialCambioAccesorio
    form_class = CambioBombaForm
    template_name = 'proyecto_monitoreo_smcv/soporte/cambios/cambio_bomba_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:bombas_view')
    group_required = 'Yeni_Soporte'

    def form_valid(self, form):
        self.object = form.save()

        if self.object.accesorio and self.object.accesorio.codigo_inventario:
            try:
                repuesto = InventarioRepuestos.objects.get(codigo=self.object.accesorio.codigo_inventario)
                if repuesto.stock > 0:
                    repuesto.stock -= 1
                    repuesto.save()
                    messages.success(self.request, f"Stock descontado: {repuesto.articulo}")
                else:
                    messages.warning(self.request, f"Sin stock de {repuesto.articulo}")
            except InventarioRepuestos.DoesNotExist:
                pass
        
        return redirect(self.success_url)

class InspeccionEspecificaCreateView(LoginRequiredMixin, GroupRequiredMixin, FormView):
    template_name = 'proyecto_monitoreo_smcv/soporte/inspeccion_form.html'
    form_class = InspeccionEspecificaForm
    success_url = reverse_lazy('proyecto_monitoreo_smcv:soporte_index') # O historial
    group_required = 'Yeni_Soporte'

    def form_valid(self, form):
        # 1. Crear la Cabecera (Inspección)
        inspeccion = Inspeccion.objects.create(
            fecha_inspeccion=form.cleaned_data['fecha_inspeccion'],
            responsable=form.cleaned_data['responsable'],
            total_equipos=1, # Es específica
            observaciones_generales=form.cleaned_data['observaciones_generales'],
        )
        dispositivo = form.cleaned_data['dispositivo']
        # 2. Crear el Detalle
        detalle = DetalleInspeccion.objects.create(
            inspeccion=inspeccion,
            dispositivo=form.cleaned_data['dispositivo'],
            estado_case=form.cleaned_data['estado_case'],
            estado_botones=form.cleaned_data['estado_botones'],
            estado_pantalla=form.cleaned_data['estado_pantalla'],
            estado_bateria=form.cleaned_data['estado_bateria'],
            estado_accesorios=form.cleaned_data['estado_accesorios'],
            observaciones=form.cleaned_data['observaciones_estado'],
            resultado_final=form.cleaned_data['resultado_final']
        )

        # 3. Crear la Prueba Técnica (Datos numéricos)
        PruebaTecnica.objects.create(
            detalle_inspeccion=detalle,
            hora_inicio=form.cleaned_data['hora_inicio'],
            hora_fin=form.cleaned_data['hora_fin'],
            tiempo_total=form.cleaned_data['tiempo_total'],
            
            # Bomba
            flujo_pre_calibracion=form.cleaned_data['flujo_pre'],
            flujo_post_calibracion=form.cleaned_data['flujo_post'],
            flujo_promedio=form.cleaned_data['flujo_promedio'],
            respuesta=form.cleaned_data['respuesta'],
            
            # Dosímetro
            tasa_cambio=form.cleaned_data['tasa_cambio'],
            ponderacion=form.cleaned_data['ponderacion'],
            db_pre_calibracion=form.cleaned_data['db_pre'],
            db_post_calibracion=form.cleaned_data['db_post'],
            l_max=form.cleaned_data['lectura_max'],
            l_min=form.cleaned_data['lectura_min'],
            l_pico=form.cleaned_data['lectura_pico'],
            
            # Vibro (Usamos campos genéricos o creamos nuevos si faltan en modelo)
            serie_cuerpo=form.cleaned_data['serie_cuerpo'],
            sens_x_cuerpo=form.cleaned_data['sens_x_cuerpo'],
            sens_y_cuerpo=form.cleaned_data['sens_y_cuerpo'],
            sens_z_cuerpo=form.cleaned_data['sens_z_cuerpo'],
            
            serie_mano=form.cleaned_data['serie_mano'],
            sens_x_mano=form.cleaned_data['sens_x_mano'],
            sens_y_mano=form.cleaned_data['sens_y_mano'],
            sens_z_mano=form.cleaned_data['sens_z_mano'],
            fuentes_ruido=form.cleaned_data['fuentes_ruido'],
            
            equipo_patron=form.cleaned_data['equipo_patron'],

            bomba_flujo_constante=form.cleaned_data['bomba_flujo_constante'],
            bomba_ruido_excesivo=form.cleaned_data['bomba_ruido_excesivo'],
            
            vibro_aeq_x=form.cleaned_data['vibro_aeq_x'],
            vibro_aeq_y=form.cleaned_data['vibro_aeq_y'],
            vibro_aeq_z=form.cleaned_data['vibro_aeq_z'],
        )
        
        # Redirigir al PDF directamente para verlo
        messages.success(self.request, f"Inspección registrada correctamente para: {dispositivo.nombre}")
        return redirect(self.get_success_url())

class InspeccionConjuntaCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = Inspeccion
    form_class = InspeccionConjuntaForm
    template_name = 'proyecto_monitoreo_smcv/soporte/inspeccion_conjunta_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:soporte_index')
    group_required = 'Yeni_Soporte'
    def get_initial(self):
        initial = super().get_initial()
        
        # Contar dispositivos donde la ubicación contenga "mina" (insensible a mayúsculas)
        # Ejemplo: "Mina", "MINA", "Mina Yanacocha"
        conteo = Dispositivo.objects.filter(ubicacion__icontains='mina').count()
        
        initial['total_equipos'] = conteo
        return initial
    
    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        
        # Instancia vacía temporal para que el formset no sea None
        # Esto soluciona el error 'NoneType' object has no attribute '_state'
        instance = Inspeccion() 

        if self.request.POST:
            data['detalles'] = DetalleMuestraFormSet(self.request.POST, instance=instance)
            data['patrones'] = PatronFormSet(self.request.POST, instance=instance, prefix='patrones')
            data['reportados'] = ReportadoFormSet(self.request.POST, instance=instance, prefix='reportados')
        else:
            data['detalles'] = DetalleMuestraFormSet(instance=instance)
            data['patrones'] = PatronFormSet(instance=instance, prefix='patrones')
            data['reportados'] = ReportadoFormSet(instance=instance, prefix='reportados')
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalles = context['detalles']
        patrones = context['patrones']
        reportados = context['reportados']
        
        # Validamos todo
        if form.is_valid() and detalles.is_valid() and patrones.is_valid() and reportados.is_valid():
            
            # 1. Guardar la Cabecera REAL
            self.object = form.save()
            
            # 2. Asignar la instancia REAL a los formsets
            detalles.instance = self.object
            patrones.instance = self.object
            reportados.instance = self.object
            
            # 3. Guardar Detalles
            instances = detalles.save()
            # Guardar PruebaTecnica para cada detalle...
            for i, detalle_obj in enumerate(instances):
                form_row = detalles.forms[i]
                no_aplica = form_row.cleaned_data.get('no_aplica')
                pre = None if no_aplica else form_row.cleaned_data.get('pre_calibracion')
                post = None if no_aplica else form_row.cleaned_data.get('post_calibracion')
                equipo_patron = form_row.cleaned_data.get('equipo_patron')

                PruebaTecnica.objects.create(
                    detalle_inspeccion=detalle_obj,
                    db_pre_calibracion=pre,
                    db_post_calibracion=post,
                    flujo_pre_calibracion=pre,
                    no_aplica=no_aplica,
                    equipo_patron=equipo_patron
                )

            # 4. Guardar Patrones
            patrones.save()
            
            # 5. Guardar Reportados (Con filtro manual para no guardar vacíos)
            # Como usamos formset.save() arriba, este guarda todo. 
            # Si quieres filtrar vacíos como hicimos antes, usa el bucle manual:
            for form_rep in reportados:
                if form_rep.cleaned_data.get('dispositivo'):
                    obj = form_rep.save(commit=False)
                    obj.inspeccion = self.object
                    obj.save()

            return redirect(self.success_url)
        
        else:
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        print("\n" + "="*50)
        print("🔍 DEPURACIÓN: FORMULARIO INVÁLIDO DETECTADO")
        print("="*50)
        
        # 1. Errores del Formulario Principal
        if form.errors:
            print("❌ ERRORES CABECERA (MAIN FORM):")
            for field, errors in form.errors.items():
                print(f"   -> Campo '{field}': {errors}")
        else:
            print("✅ Cabecera OK")

        # Recuperar los formsets del contexto
        context = self.get_context_data()
        detalles = context.get('detalles')
        patrones = context.get('patrones')
        reportados = context.get('reportados')

        # 2. Errores en Detalles (Muestras)
        if detalles and not detalles.is_valid():
            print("\n❌ ERRORES EN TABLA MUESTRAS (Detalles):")
            if detalles.non_form_errors():
                print(f"   -> General: {detalles.non_form_errors()}")
            for i, error in enumerate(detalles.errors):
                if error: print(f"   -> Fila {i+1}: {error}")
        else:
            print("\n✅ Tabla Muestras OK")

        # 3. Errores en Patrones
        if patrones and not patrones.is_valid():
            print("\n❌ ERRORES EN TABLA PATRONES:")
            if patrones.non_form_errors():
                print(f"   -> General: {patrones.non_form_errors()}")
            for i, error in enumerate(patrones.errors):
                if error: print(f"   -> Fila {i+1}: {error}")
        else:
            print("\n✅ Tabla Patrones OK")

        # 4. Errores en Reportados
        if reportados and not reportados.is_valid():
            print("\n❌ ERRORES EN TABLA REPORTADOS:")
            if reportados.non_form_errors():
                print(f"   -> General: {reportados.non_form_errors()}")
            for i, error in enumerate(reportados.errors):
                if error: print(f"   -> Fila {i+1}: {error}")
        else:
            print("\n✅ Tabla Reportados OK")

        print("="*50 + "\n")
        return super().form_invalid(form)   

class ProgramaCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = ProgramaMantenimiento
    fields = ['anio', 'nombre', 'semestre']
    template_name = 'proyecto_monitoreo_smcv/soporte/programa_form.html'
    success_url = reverse_lazy('proyecto_monitoreo_smcv:lista_programas') # Define esta URL
    group_required = 'Yeni_Soporte'

class FilaCronogramaCreateView(LoginRequiredMixin, GroupRequiredMixin, CreateView):
    model = FilaCronograma
    form_class = FilaCronogramaForm
    template_name = 'proyecto_monitoreo_smcv/soporte/fila_form.html'
    group_required = 'Yeni_Soporte'

    def form_valid(self, form):
        programa = get_object_or_404(ProgramaMantenimiento, pk=self.kwargs['programa_id'])
        form.instance.programa = programa
        
        # Lógica: Si no seleccionan ejecutados, asumimos que son los mismos que programados
        response = super().form_valid(form)
        if not form.cleaned_data['equipos_ejecutados']:
            form.instance.equipos_ejecutados.set(form.cleaned_data['equipos_programados'])
        
        return response

    def get_success_url(self):
        # Redirigir al PDF o a una vista de detalle del programa
        return reverse_lazy('proyecto_monitoreo_smcv:detalle_programa', kwargs={'pk': self.object.programa.id})

# ------ VISTAS --------

class InventarioGeneralView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Dispositivo
    template_name = 'proyecto_monitoreo_smcv/soporte/tablas/general.html'
    context_object_name = 'dispositivos'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        # Optimizamos la consulta para traer accesorios y calibraciones de golpe
        qs = Dispositivo.objects.prefetch_related('accesorios', 'calibracion_set').all().order_by('id')
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Aquí procesaremos los datos para facilitar el template
        tabla_data = []
        
        for disp in context['dispositivos']:
            # Obtener última calibración para sacar fechas
            ultima_cal = disp.calibracion_set.order_by('-fecha_calibracion').first()
            
            # Obtener accesorios como texto
            accesorios_list = disp.accesorios.filter(serie__isnull=False)
            acc_nombres = ", ".join([a.nombre for a in accesorios_list]) if accesorios_list else "-"
            # Asumimos que la serie del accesorio va junto al nombre o es un campo pendiente
            acc_series = ", ".join([a.serie for a in accesorios_list]) if accesorios_list else "-"

            tiempo_retorno = "-"
            
            # Solo calculamos si ambas fechas existen
            if disp.fecha_subida and disp.fecha_bajada:
                # Calculamos la diferencia relativa (calendario exacto)
                # Asumimos que Fecha Subida es posterior a Fecha Bajada
                diferencia = relativedelta(disp.fecha_subida, disp.fecha_bajada)
                
                partes = []
                
                # Años (por si acaso tarda mucho)
                if diferencia.years > 0:
                    partes.append(f"{diferencia.years} año{'s' if diferencia.years != 1 else ''}")
                
                # Meses
                if diferencia.months > 0:
                    partes.append(f"{diferencia.months} mes{'es' if diferencia.months != 1 else ''}")
                
                # Días
                if diferencia.days > 0:
                    partes.append(f"{diferencia.days} día{'s' if diferencia.days != 1 else ''}")
                
                # Unimos las partes con coma (Ej: "1 mes, 5 días")
                if partes:
                    tiempo_retorno = ", ".join(partes)
                else:
                    tiempo_retorno = "0 días" # Si son la misma fecha

            item = {
                'obj': disp,
                'codigo_interno': disp.codigo_interno,
                'equipo': disp.nombre,
                'marca': disp.marca,
                'modelo': disp.modelo,
                'serie': disp.serie,
                'accesorios': acc_nombres,
                'serie_accesorio': acc_series,
                'ubicacion': disp.ubicacion,
                'estado': disp.estado,
                'fecha_bajada': disp.fecha_bajada, 
                'fecha_subida': disp.fecha_subida,
                'tiempo_retorno': tiempo_retorno, 
                'fecha_calibracion': ultima_cal.fecha_calibracion if ultima_cal else "-",
                'fecha_prox_calibracion': ultima_cal.fecha_proxima if (ultima_cal and ultima_cal.fecha_proxima) else "-"
            }
            tabla_data.append(item)
        
        context['tabla_data'] = tabla_data
        return context

class MatrizMantenimientoView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Dispositivo
    template_name = 'proyecto_monitoreo_smcv/soporte/tablas/matriz_mantenimiento.html'
    context_object_name = 'dispositivos'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        return Dispositivo.objects.all().order_by('id')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        matriz_data = []
        
        for index, disp in enumerate(context['dispositivos'], start=1):
            # 1. Obtener los últimos registros de cada módulo
            diag = disp.diagnostico_set.order_by('-fecha_recepcion').first()
            mant = disp.mantenimiento_set.order_by('-created_at').first()
            cal = disp.calibracion_set.order_by('-fecha_calibracion').first()
            op = disp.operatividad_set.order_by('-fecha_operatividad').first()

            diag_data = {'id': diag.id, 'fecha': diag.fecha_recepcion} if diag else None
            cal_data = {'id': cal.id, 'fecha': cal.fecha_calibracion} if cal else None
            mant_data = {'id': mant.id, 'fecha': mant.created_at} if mant else None # O la fecha que uses
            op_data = {'id': op.id, 'fecha': op.fecha_operatividad} if op else None
            
            # 2. Accesorio
            accesorios_list = disp.accesorios.filter(serie__isnull=False)
            acc_nombres = ", ".join([a.nombre for a in accesorios_list]) if accesorios_list else "-"
            # Asumimos que la serie del accesorio va junto al nombre o es un campo pendiente
            acc_series = ", ".join([a.serie for a in accesorios_list]) if accesorios_list else "-"

            # 3. Cálculo de Tiempo de Atención (Días desde recepción hasta hoy)
            # Si hay fecha de recepción en el diagnóstico, calculamos.
            fecha_recepcion = "-"
            observacion = "-"
            tiempo_retorno = "-"
            
            # Solo calculamos si ambas fechas existen
            if disp.fecha_subida and disp.fecha_bajada:
                # Calculamos la diferencia relativa (calendario exacto)
                # Asumimos que Fecha Subida es posterior a Fecha Bajada
                diferencia = relativedelta(disp.fecha_subida, disp.fecha_bajada)
                
                partes = []
                
                # Años (por si acaso tarda mucho)
                if diferencia.years > 0:
                    partes.append(f"{diferencia.years} año{'s' if diferencia.years != 1 else ''}")
                
                # Meses
                if diferencia.months > 0:
                    partes.append(f"{diferencia.months} mes{'es' if diferencia.months != 1 else ''}")
                
                # Días
                if diferencia.days > 0:
                    partes.append(f"{diferencia.days} día{'s' if diferencia.days != 1 else ''}")
                
                # Unimos las partes con coma (Ej: "1 mes, 5 días")
                if partes:
                    tiempo_retorno = ", ".join(partes)
                else:
                    tiempo_retorno = "0 días" # Si son la misma fecha
            

            row = {
                'no': index,
                'obj': disp, # Objeto dispositivo completo
                'equipo': disp.nombre,
                'marca': disp.marca,
                'modelo': disp.modelo,
                'serie': disp.serie,
                'observacion': disp.observacion, 
                'accesorios': acc_nombres,
                'serie_accesorio': acc_series,
                'fecha_recepcion': disp.fecha_bajada,
                
                # Objetos relacionados para sacar IDs y links
                'diag_data': diag_data, 
                'cal_data': cal_data,
                'mant_data': mant_data,
                'op_data': op_data,
                
                'estado': disp.estado,
                'fecha_subida': disp.fecha_subida,
                'tiempo_atencion': tiempo_retorno
            }
            matriz_data.append(row)
        
        context['matriz_data'] = matriz_data
        return context

class VibrometrosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Dispositivo
    template_name = 'proyecto_monitoreo_smcv/soporte/cambios/vibrometros.html'
    context_object_name = 'dispositivos'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        # Traemos vibrometros con sus accesorios
        return Dispositivo.objects.filter(nombre__icontains='Vibrometro').prefetch_related('accesorios').order_by('id')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tabla_data = []
        
        for disp in context['dispositivos']:
            # Lógica 1: Accesorio
            accesorios_con_serie = disp.accesorios.filter(serie__isnull=False).exclude(serie__exact='')
            row_count = accesorios_con_serie.count() or 1
            lista_accesorios = list(accesorios_con_serie) if accesorios_con_serie.exists() else [None]

            # Lógica 2: Mapear los cambios a un diccionario
            todos_los_cambios = HistorialCambioAccesorio.objects.filter(accesorio__dispositivo=disp).select_related('accesorio').order_by('fecha_cambio')
            
            # Como ordenamos por fecha ascendente, el último valor será la FECHA MÁS RECIENTE.
            cambios_dict = {c.accesorio.nombre: c.fecha_cambio for c in todos_los_cambios}

            tabla_data.append({
                'dispositivo': disp,
                'row_span': row_count,
                'accesorios': lista_accesorios,
                'cambios': cambios_dict # Pasamos el diccionario limpio
            })
            
        context['vibrometros_data'] = tabla_data
        return context

class InventarioRepuestosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = InventarioRepuestos
    template_name = 'proyecto_monitoreo_smcv/soporte/inventario/lista_repuestos.html'
    context_object_name = 'repuestos'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        return InventarioRepuestos.objects.all().order_by('articulo')
    
class HistorialCambiosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = HistorialCambioAccesorio
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_cambios_componentes.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'
    paginate_by = 50

    def get_queryset(self):
        # Traemos todos los cambios, ordenados por fecha descendente (lo último primero)
        return HistorialCambioAccesorio.objects.filter(accesorio__dispositivo__nombre__icontains='Vibrometro').order_by('-fecha_cambio')

class DosimetrosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Dispositivo
    template_name = 'proyecto_monitoreo_smcv/soporte/cambios/dosimetros.html'
    context_object_name = 'dispositivos'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        return Dispositivo.objects.filter(nombre__icontains='Dosimetro').prefetch_related('accesorios').order_by('id')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tabla_data = []
        
        for disp in context['dispositivos']:
            # Accesorio con serie (para generar filas)
            accesorios_con_serie = disp.accesorios.filter(serie__isnull=False).exclude(serie__exact='')
            row_count = accesorios_con_serie.count() or 1
            lista_accesorios = list(accesorios_con_serie) if accesorios_con_serie.exists() else [None]

            # Mapeo de cambios
            todos_los_cambios = HistorialCambioAccesorio.objects.filter(accesorio__dispositivo=disp).select_related('accesorio').order_by('fecha_cambio')
            cambios_dict = {c.accesorio.nombre: c.fecha_cambio for c in todos_los_cambios}
            
            # CÁLCULO DE PRÓXIMO CAMBIO (PANTALLA + 6 MESES)
            prox_pantalla = "-"
            # Buscamos alguna key que parezca pantalla
            pantalla_key = next((k for k in cambios_dict.keys() if 'pantalla' in k.lower()), None)
            if pantalla_key and cambios_dict[pantalla_key]:
                fecha_cambio = cambios_dict[pantalla_key]
                fecha_prox = fecha_cambio + relativedelta(months=6)
                prox_pantalla = fecha_prox

            tabla_data.append({
                'dispositivo': disp,
                'row_span': row_count,
                'accesorios': lista_accesorios,
                'cambios': cambios_dict,
                'prox_pantalla': prox_pantalla # Variable calculada
            })
            
        context['dosimetros_data'] = tabla_data
        return context

class HistorialCambiosDosimetrosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = HistorialCambioAccesorio
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_cambios_dosimetros.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        # Filtramos solo dispositivos que sean Dosimetros
        return HistorialCambioAccesorio.objects.filter(accesorio__dispositivo__nombre__icontains='Dosimetro').order_by('-fecha_cambio')

class BombasView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Dispositivo
    template_name = 'proyecto_monitoreo_smcv/soporte/cambios/bombas.html'
    context_object_name = 'dispositivos'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        return Dispositivo.objects.filter(nombre__icontains='Bomba').prefetch_related('accesorios').order_by('id')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tabla_data = []
        
        for disp in context['dispositivos']:
            # Diccionario de cambios
            todos_los_cambios = HistorialCambioAccesorio.objects.filter(accesorio__dispositivo=disp).select_related('accesorio').order_by('fecha_cambio')
            c = {obj.accesorio.nombre: obj.fecha_cambio for obj in todos_los_cambios}
            
            # --- CÁLCULOS DE PRÓXIMO CAMBIO ---
            def get_date(key_part):
                key = next((k for k in c.keys() if key_part.lower() in k.lower()), None)
                return c[key] if key else None

            # Motor: 4 años
            val_motor = get_date('motor')
            prox_motor = val_motor + relativedelta(years=4) if val_motor else "-"
            # Batería: 2 años
            val_bateria = get_date('bater')
            prox_bateria = val_bateria + relativedelta(years=2) if val_bateria else "-"
            # Filtro: 1 año
            val_filtro = get_date('filtro')
            prox_filtro = val_filtro + relativedelta(years=1) if val_filtro else "-"
            # Diafragma: 2 años
            val_diaf = get_date('diafragma')
            prox_diafragma = val_diaf + relativedelta(years=2) if val_diaf else "-"
            # Teclado: Sin cálculo
            prox_teclado = "-"

            tabla_data.append({
                'dispositivo': disp,
                'cambios': c,
                'prox': {
                    'MOTOR': prox_motor,
                    'BATERIA': prox_bateria,
                    'FILTRO': prox_filtro,
                    'DIAFRAGMA': prox_diafragma
                }
            })
            
        context['bombas_data'] = tabla_data
        return context

class HistorialCambiosBombasView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = HistorialCambioAccesorio
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_cambios_bombas.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        return HistorialCambioAccesorio.objects.filter(accesorio__dispositivo__nombre__icontains='Bomba').order_by('-fecha_cambio')

class HistorialInspeccionesView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Inspeccion
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_inspecciones.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'
    paginate_by = 20

    def get_queryset(self):
        # Filtramos inspecciones y ordenamos por fecha descendente
        # prefetch_related('detalles__dispositivo') es vital para no hacer 100 consultas SQL
        qs = Inspeccion.objects.filter(total_equipos=1).prefetch_related(
            'detalles', 
            'detalles__dispositivo',
            'responsable'
        ).order_by('-fecha_inspeccion')
        
        q = self.request.GET.get('q')
        if q:
            # Buscamos por nombre de equipo, serie o nombre del responsable
            qs = qs.filter(
                Q(detalles__dispositivo__nombre__icontains=q) |
                Q(detalles__dispositivo__serie__icontains=q) |
                Q(responsable__nombre__icontains=q)
            ).distinct() # distinct() evita duplicados si coincide en varios campos
            
        return qs

class HistorialConjuntasView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Inspeccion
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_conjuntas.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'
    paginate_by = 20

    def get_queryset(self):
        # Filtramos solo las que son masivas (más de 1 equipo o creadas con el form conjunto)
        # Usamos prefetch_related para optimizar la carga de detalles
        qs = Inspeccion.objects.filter(total_equipos__gt=1).prefetch_related(
            'detalles', 'detalles__dispositivo', 'responsable'
        ).order_by('-fecha_inspeccion')
        
        q = self.request.GET.get('q')
        if q:
            # Búsqueda por código o responsable
            qs = qs.filter(
                Q(codigo_documento__icontains=q) |
                Q(responsable__nombre__icontains=q)
            )
        return qs

class ListaProgramasView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = ProgramaMantenimiento
    template_name = 'proyecto_monitoreo_smcv/soporte/tablas/lista_programas.html'
    context_object_name = 'programas'
    group_required = 'Yeni_Soporte'
    ordering = ['-anio', '-semestre'] # Más recientes primero

# --- VER DETALLE Y AGREGAR FILAS (La "Matriz" de un año) ---
class DetalleProgramaView(LoginRequiredMixin, GroupRequiredMixin, DetailView):
    model = ProgramaMantenimiento
    template_name = 'proyecto_monitoreo_smcv/soporte/tablas/detalle_programa.html'
    context_object_name = 'programa'
    group_required = 'Yeni_Soporte'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Traemos las filas ordenadas por número (1, 2, 3...)
        context['filas'] = self.object.filas.prefetch_related(
            'equipos_programados', 'equipos_ejecutados'
        ).order_by('numero')
        return context

# --- 1. HISTORIAL DIAGNÓSTICOS ---
class HistorialDiagnosticosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Diagnostico
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_diagnosticos.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'
    paginate_by = 20 # Paginación opcional

    def get_queryset(self):
        qs = Diagnostico.objects.select_related('dispositivo').order_by('-fecha_recepcion')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(numero__icontains=q) | 
                Q(dispositivo__nombre__icontains=q) |
                Q(dispositivo__serie__icontains=q) |
                Q(estado__icontains=q)
            )
        return qs

# --- 2. HISTORIAL MANTENIMIENTOS ---
class HistorialMantenimientosView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Mantenimiento
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_mantenimientos.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        qs = Mantenimiento.objects.select_related('dispositivo', 'cliente', 'proveedor').order_by('-created_at')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(dispositivo__nombre__icontains=q) |
                Q(dispositivo__serie__icontains=q) |
                Q(cliente__razon_social__icontains=q) |
                Q(proveedor__trabajador__nombre__icontains=q)
            )
        return qs

# --- 3. HISTORIAL CALIBRACIONES ---
class HistorialCalibracionesView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Calibracion
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_calibraciones.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        qs = Calibracion.objects.select_related('dispositivo').order_by('-fecha_calibracion')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(dispositivo__nombre__icontains=q) |
                Q(dispositivo__serie__icontains=q) |
                Q(estado_final__icontains=q)
            )
        return qs

# --- 4. HISTORIAL OPERATIVIDAD ---
class HistorialOperatividadView(LoginRequiredMixin, GroupRequiredMixin, ListView):
    model = Operatividad
    template_name = 'proyecto_monitoreo_smcv/soporte/historial/lista_operatividad.html'
    context_object_name = 'registros'
    group_required = 'Yeni_Soporte'

    def get_queryset(self):
        qs = Operatividad.objects.select_related('dispositivo').order_by('-fecha_operatividad')
        q = self.request.GET.get('q')
        if q:
            qs = qs.filter(
                Q(dispositivo__nombre__icontains=q) |
                Q(dispositivo__serie__icontains=q) |
                Q(estado_final__icontains=q)
            )
        return qs

# --- FUNCIONES PARA GENERAR PDF ---

def generar_pdf_diagnostico(request, pk):
    diagnostico = get_object_or_404(Diagnostico, pk=pk)
    dispositivo = diagnostico.dispositivo
    fotos = diagnostico.fotos.all() # Relacion inversa de FotoDiagnostico
    ultima_calibracion = Calibracion.objects.filter(dispositivo=dispositivo).order_by('-fecha_calibracion').first()

    # Contexto para el HTML
    context = {
        'diagnostico': diagnostico,
        'dispositivo': dispositivo,
        'fotos': fotos,
        'ultima_calibracion': ultima_calibracion,
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/global_group.jpg')
    }

    # Renderizar HTML
    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/diagnostico_pdf.html', context)

    # Generar PDF
    # base_url es importante para que encuentre imagenes y estilos
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    # 'inline' abre el PDF en el navegador. 'attachment' lo descarga directo.
    response['Content-Disposition'] = f'inline; filename="Diagnostico_{diagnostico.numero}.pdf"'
    return response

def generar_pdf_mantenimiento(request, pk):
    mantenimiento = get_object_or_404(Mantenimiento, pk=pk)
    
    context = {
        'mantenimiento': mantenimiento,
        'dispositivo': mantenimiento.dispositivo,
        'cliente': mantenimiento.cliente,
        'proveedor': mantenimiento.proveedor,
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/global_group.jpg'),
        'firmaT_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_tecnico.png'),
        'firmaG_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_gerente.png'),
    }

    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/mantenimiento_pdf.html', context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Mantenimiento_{mantenimiento.id}.pdf"'
    return response

def generar_pdf_operatividad(request, pk):
    operatividad = get_object_or_404(Operatividad, pk=pk)
    dispositivo = operatividad.dispositivo
    cliente = dispositivo.cliente
    verificaciones = operatividad.verificaciones.all()
    evidencias = operatividad.evidencias.all()

    context = {
        'operatividad': operatividad,
        'dispositivo': dispositivo,
        'cliente': cliente,
        'verificaciones': verificaciones,
        'evidencias': evidencias,
        # URL absoluta para logo y marca de agua
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/global_group_Ope.jpg'),
        # Usaremos el mismo logo como marca de agua, o si tienes otra imagen cámbialo
        'watermark_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/fondo_Ope.png'), 
        'firmaT_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_tecnico.png'),
        'firmaG_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_gerente.png'),
    }

    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/operatividad_pdf.html', context)
    
    # Base_url es vital para cargar imágenes estáticas
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Certificado_Operatividad_{operatividad.id}.pdf"'
    return response

def generar_pdf_calibracion(request, pk):
    calibracion = get_object_or_404(Calibracion, pk=pk)
    # Traemos resultados y patrones
    resultados = calibracion.resultados.all().order_by('flujo_nominal')
    patrones = calibracion.patrones.first() # Asumimos 1 bloque de patrones
    fotos = FotoPatron.objects.filter(calibracion=calibracion)
    
    context = {
        'calibracion': calibracion,
        'dispositivo': calibracion.dispositivo,
        'cliente': calibracion.dispositivo.cliente,
        'resultados': resultados,
        'patrones': patrones,
        'fotos': fotos,
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/global_group_Ope.jpg'),
        # Usaremos el mismo logo como marca de agua, o si tienes otra imagen cámbialo
        'watermark_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/fondo_Ope.png'), 
        'firmaT_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_tecnico.png'),
        'firmaG_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_gerente.png'),
    }

    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/calibracion_pdf.html', context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Certificado_Calibracion_{calibracion.id}.pdf"'
    return response

def exportar_inventario_pdf(request):
    # 1. Obtener los mismos datos que la vista web
    dispositivos = Dispositivo.objects.prefetch_related('accesorios', 'calibracion_set').all().order_by('id')
    
    tabla_data = []
    
    # Reutilizamos la lógica de tiempo de retorno
    for disp in dispositivos:
        ultima_cal = disp.calibracion_set.order_by('-fecha_calibracion').first()
        
        # Calcular tiempo de retorno
        tiempo_retorno = "-"
        if disp.fecha_subida and disp.fecha_bajada:
            diferencia = relativedelta(disp.fecha_subida, disp.fecha_bajada)
            partes = []
            if diferencia.years > 0: partes.append(f"{diferencia.years} año{'s' if diferencia.years != 1 else ''}")
            if diferencia.months > 0: partes.append(f"{diferencia.months} mes{'es' if diferencia.months != 1 else ''}")
            if diferencia.days > 0: partes.append(f"{diferencia.days} día{'s' if diferencia.days != 1 else ''}")
            tiempo_retorno = ", ".join(partes) if partes else "0 días"

        item = {
            'obj': disp,
            'accesorios_str': ", ".join([a.nombre for a in disp.accesorios.all()]),
            'fecha_cal': ultima_cal.fecha_calibracion if ultima_cal else "-",
            'prox_cal': ultima_cal.fecha_proxima if (ultima_cal and ultima_cal.fecha_proxima) else "-",
            'tiempo_retorno': tiempo_retorno
        }
        tabla_data.append(item)

    # 2. Contexto para el PDF
    context = {
        'tabla_data': tabla_data,
        'fecha_actual': timezone.now(), # Para el título amarillo
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/logo.jpg'), # Asegúrate que la imagen exista
    }

    # 3. Renderizar HTML
    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/inventario_general_pdf.html', context)
    
    # 4. Generar PDF (Horizontal para que quepa todo)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Inventario_General.pdf"'
    return response

def generar_pdf_inspeccion_especifica(request, pk):
    inspeccion = get_object_or_404(Inspeccion, pk=pk)
    detalle = inspeccion.detalles.first() # Solo hay uno en este caso
    prueba = detalle.prueba_tecnica
    dispositivo = detalle.dispositivo
    
    # Determinar tipo para la plantilla
    tipo_equipo = "OTRO"
    if "Bomba" in dispositivo.nombre: tipo_equipo = "BOMBA"
    elif "Dosimetro" in dispositivo.nombre or "Dosímetro" in dispositivo.nombre: tipo_equipo = "DOSIMETRO"
    elif "Vibrometro" in dispositivo.nombre or "Vibrómetro" in dispositivo.nombre: tipo_equipo = "VIBROMETRO"

    context = {
        'inspeccion': inspeccion,
        'detalle': detalle,
        'prueba': prueba,
        'dispositivo': dispositivo,
        'tipo_equipo': tipo_equipo,
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/logo.jpg'),
        'firmaT_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_tecnico.png'),
    }

    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/inspeccion_especifica_pdf.html', context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Inspeccion_{dispositivo.serie}.pdf"'
    return response

def generar_pdf_inspeccion_conjunta(request, pk):
    inspeccion = get_object_or_404(Inspeccion, pk=pk)
    # Traemos los detalles y sus pruebas técnicas asociadas
    detalles = inspeccion.detalles.select_related('dispositivo', 'prueba_tecnica').all()
    
    context = {
        'inspeccion': inspeccion,
        'detalles': detalles,
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/logo.jpg'),
        'firmaT_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/firma_tecnico.png'),
    }

    html_string = render_to_string('proyecto_monitoreo_smcv/soporte/pdf/inspeccion_conjunta_pdf.html', context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Inspeccion_Conjunta_{inspeccion.fecha_inspeccion}.pdf"'
    return response

def generar_pdf_programa(request, pk):
    programa = get_object_or_404(ProgramaMantenimiento, pk=pk)
    filas = programa.filas.prefetch_related('equipos_programados').all().order_by('numero')

    # 1. FECHAS (Igual que antes)
    primera_fila = filas.order_by('fecha_programada').first()
    fecha_inicio = primera_fila.fecha_programada if primera_fila else date(programa.anio, 1, 1)
    
    # Aseguramos empezar un poco antes si es necesario o desde la fecha exacta
    # Para que cuadre con tu imagen que empieza el 03/02, usaremos la fecha exacta de la primera fila
    
    fecha_fin = fecha_inicio + relativedelta(months=6)
    
    todas_semanas = []
    current = fecha_inicio
    while current <= fecha_fin:
        todas_semanas.append(current)
        current += timedelta(days=7)

    # CORTE DE MESES (Ajusta este número si quieres más o menos semanas por hoja)
    corte = 13 
    semanas_p1 = todas_semanas[:corte]
    semanas_p2 = todas_semanas[corte:]

    def obtener_meses_header(lista_semanas):
        if not lista_semanas: return []
        headers = []
        mes_actual = lista_semanas[0].month
        contador = 0
        # Diccionario de meses en español
        nombres = {1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio', 
                   7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'}
        
        for sem in lista_semanas:
            if sem.month == mes_actual:
                contador += 1
            else:
                headers.append({'nombre': nombres[mes_actual], 'colspan': contador})
                mes_actual = sem.month
                contador = 1
        headers.append({'nombre': nombres[mes_actual], 'colspan': contador})
        return headers

    meses_p1 = obtener_meses_header(semanas_p1)
    meses_p2 = obtener_meses_header(semanas_p2)

    # 2. PROCESAR FILAS (AQUÍ ESTÁ LA CORRECCIÓN DEL MONITOR)
    rows_data = []
    for fila in filas:
        all_equipos = fila.equipos_programados.all()
        
        # Filtros más flexibles
        vibs_prog = []
        dosis_prog = []
        bombas_prog = []
        otros_prog = [] # <--- AQUÍ IRÁ EL MONITOR

        for eq in all_equipos:
            nombre = eq.nombre.lower()
            if 'vibrometro' in nombre or 'vibrómetro' in nombre:
                vibs_prog.append(eq)
            elif 'dosimetro' in nombre or 'dosímetro' in nombre:
                dosis_prog.append(eq)
            elif 'bomba' in nombre:
                bombas_prog.append(eq)
            else:
                otros_prog.append(eq) # Monitores, Sonómetros, etc.

        # Lógica de Color Rojo:
        # Si fila.ejecutado es TRUE -> Todo Negro (Bien)
        # Si fila.ejecutado es FALSE -> Todo Rojo (Pendiente)
        def procesar_equipos(lista_prog):
            resultado = []
            for eq in lista_prog:
                # Si NO se ejecutó la fila entera, el equipo sale rojo
                es_rojo = not fila.ejecutado 
                resultado.append({'serie': eq.serie, 'nombre': eq.nombre, 'rojo': es_rojo})
            return resultado

        data = {
            'numero': fila.numero,
            'vibros': procesar_equipos(vibs_prog),
            'dosis': procesar_equipos(dosis_prog),
            'bombas': procesar_equipos(bombas_prog),
            'otros': procesar_equipos(otros_prog), # Enviamos los otros
            
            'fecha_p': fila.fecha_programada,
            'fecha_e': fila.fecha_programada if fila.ejecutado else None,
            'observaciones': fila.observaciones,
            
            # Conteo total corregido
            'total_eq': len(vibs_prog) + len(dosis_prog) + len(bombas_prog) + len(otros_prog)
        }
        rows_data.append(data)

    context = {
        'programa': programa,
        'logo_url': request.build_absolute_uri(settings.STATIC_URL + 'proyecto_monitoreo_smcv/img/ceneris_logo.png'),
        'rows_data': rows_data,
        'parte1': {'semanas': semanas_p1, 'meses': meses_p1},
        'parte2': {'semanas': semanas_p2, 'meses': meses_p2},
    }

    html_string = render_to_string('proyecto_monitoreo_smcv/programa/pdf/programa_pdf.html', context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Programa_{programa.anio}.pdf"'
    return response

# --- FUNCIONES PARA GENERAR EXCEL ---

def generar_excel_calibracion(request, pk):
    calibracion = get_object_or_404(Calibracion, pk=pk)
    # Ordenamos por flujo nominal para asegurar que el Flujo 1 vaya a la Columna C, el 2 a la D...
    resultados = calibracion.resultados.all().order_by('flujo_nominal')
    dispositivo = calibracion.dispositivo
    cliente = dispositivo.cliente
    
    # 1. RUTA DE LA PLANTILLA
    # Debes guardar tu excel "vacío" (con fórmulas pero sin datos) aquí:
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'plantillas', 'plantilla_calibracion.xlsx')
    
    if not os.path.exists(ruta_plantilla):
        return HttpResponse("Error: No se encuentra la plantilla Excel en static.", status=404)

    # 2. CARGAR WORKBOOK
    wb = load_workbook(ruta_plantilla)
    ws = wb.active # O usa ws = wb['NombreDeTuHoja']

    ruta_u_exp = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u_exp.png')
    
    if os.path.exists(ruta_u_exp):
        img = Image(ruta_u_exp)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 180
        img.height = 60
        ws.add_image(img, 'B10')
    
    ruta_u_raiz = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u_raiz.png')
    
    if os.path.exists(ruta_u_raiz):
        img = Image(ruta_u_raiz)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 230
        img.height = 60
        ws.add_image(img, 'B14')
    
    ruta_u_rep = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u_rep.png')
    
    if os.path.exists(ruta_u_rep):
        img = Image(ruta_u_rep)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 150
        img.height = 60
        ws.add_image(img, 'B18')
    
    ruta_u_res = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u_res.png')
    
    if os.path.exists(ruta_u_res):
        img = Image(ruta_u_res)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 150
        img.height = 60
        ws.add_image(img, 'B23')
    
    k = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'k.png')
    
    if os.path.exists(k):
        img = Image(k)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'F10')

    u = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u.png')
    
    if os.path.exists(u):
        img = Image(u)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'F11')

    u_exp2 = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u_exp2.png')
    
    if os.path.exists(u_exp2):
        img = Image(u_exp2)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 35
        img.height = 20
        ws.add_image(img, 'F12')

    u = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u.png')
    
    if os.path.exists(u):
        img = Image(u)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'G14')
    
    u = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'u.png')
    
    if os.path.exists(u):
        img = Image(u)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'G15')

    o = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'o.png')
    
    if os.path.exists(o):
        img = Image(o)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'E24')
    
    t = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 't.png')
    
    if os.path.exists(t):
        img = Image(t)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'E25')

    n = os.path.join(settings.BASE_DIR, 'proyecto_monitoreo_smcv', 'static', 'proyecto_monitoreo_smcv', 'img', 'n.png')
    
    if os.path.exists(n):
        img = Image(n)
        # Opcional: Ajustar tamaño si sale muy grande
        img.width = 20
        img.height = 20
        ws.add_image(img, 'E26')

    # 4. LLENAR LA MATRIZ DE MEDICIONES (Filas 36, 37, 38)
    # Según tu imagen:
    # Columna C = Flujo 1
    # Columna D = Flujo 2
    # Columna E = Flujo 3
    # Columna F = Flujo 4
    # Columna G = Flujo 5
    
    columnas = ['C', 'D', 'E', 'F', 'G']

    for i, res in enumerate(resultados):
        if i < 5: # Proteccion para no salirnos de rango
            col = columnas[i]
            
            # Escribimos los valores crudos. 
            # Excel recalculará filas 39, 40, 41... automáticamente al abrirse.
            ws[f'{col}36'] = res.m1
            ws[f'{col}37'] = res.m2
            ws[f'{col}38'] = res.m3

    # 5. DESCARGAR
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Calculos_{calibracion.id}.xlsx"'
    wb.save(response)
    return response
