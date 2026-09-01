from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError, transaction
from django.db import connection
from django.db.models import Avg, Sum, Count, Q, Min
from django.core.management.color import no_style
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils.crypto import get_random_string
from django.utils import timezone
from django.utils.timezone import make_aware
from django.core.paginator import Paginator

import datetime
import calendar
import json
from dateutil.easter import easter
import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint

from openpyxl.chart.label import DataLabelList
from .models import EvaluacionMensual, Puntaje, NotaConocimiento
from .forms import EvaluacionForm, ESTRUCTURA_EVALUACION
from recursoshumanos.models import Trabajador, Asistencia, TareoDiario, Justificacion, Sede, Area, Sancion


TARDANZA_MINIMA_MINUTOS = 1
TARDANZA_MINIMA_HORAS = TARDANZA_MINIMA_MINUTOS / 60
TARDANZA_LEVE_MAX_MINUTOS = 15
HORA_REFERENCIA_ENTRADA = datetime.time(8, 30)
HORA_REFERENCIA_ENTRADA_SABADO = datetime.time(9, 0)
HORA_TARDANZA_LEVE_MAX = datetime.time(8, 45)


def _hora_referencia_entrada_por_fecha(fecha):
    if fecha and fecha.weekday() == 5:  # Sabado
        return HORA_REFERENCIA_ENTRADA_SABADO
    return HORA_REFERENCIA_ENTRADA


def _hora_tardanza_leve_max_por_fecha(fecha):
    hora_referencia = _hora_referencia_entrada_por_fecha(fecha)
    dt_ref = datetime.datetime.combine(datetime.date.today(), hora_referencia)
    return (dt_ref + datetime.timedelta(minutes=TARDANZA_LEVE_MAX_MINUTOS)).time()


def _cargo_jerarquico_actual(trabajador):
    if not trabajador:
        return Trabajador.CargoJerarquico.TRABAJADOR

    cargo = getattr(trabajador, 'cargo_jerarquico', None)
    if cargo in [
        Trabajador.CargoJerarquico.TRABAJADOR,
        Trabajador.CargoJerarquico.SUPERVISOR,
        Trabajador.CargoJerarquico.RESPONSABLE,
        Trabajador.CargoJerarquico.GERENTE,
    ]:
        return cargo

    return Trabajador.CargoJerarquico.TRABAJADOR


def _es_responsable(trabajador):
    return _cargo_jerarquico_actual(trabajador) == Trabajador.CargoJerarquico.RESPONSABLE


def _es_supervisor(trabajador):
    return _cargo_jerarquico_actual(trabajador) == Trabajador.CargoJerarquico.SUPERVISOR


def _es_gerente_puro(trabajador):
    return _cargo_jerarquico_actual(trabajador) == Trabajador.CargoJerarquico.GERENTE


def _es_lider(trabajador):
    return _cargo_jerarquico_actual(trabajador) in [
        Trabajador.CargoJerarquico.SUPERVISOR,
        Trabajador.CargoJerarquico.RESPONSABLE,
        Trabajador.CargoJerarquico.GERENTE,
    ]


def _es_trabajador_base(trabajador):
    return _cargo_jerarquico_actual(trabajador) == Trabajador.CargoJerarquico.TRABAJADOR


def _q_supervisor(prefix=''):
    return Q(**{f'{prefix}es_jefe': True, f'{prefix}es_gerente': False})


def _q_responsable(prefix=''):
    return Q(**{f'{prefix}es_jefe': True, f'{prefix}es_gerente': True})


def _q_gerente_puro(prefix=''):
    return Q(**{f'{prefix}es_jefe': False, f'{prefix}es_gerente': True})


def _q_trabajador_base(prefix=''):
    return Q(**{f'{prefix}es_jefe': False, f'{prefix}es_gerente': False})


def _areas_operativas_activas_qs():
    """
    Devuelve solo areas reales en uso (con trabajadores o asignaciones activas),
    excluyendo Gerencia General para evitar mostrar ambitos ficticios.
    """
    ids = set(
        Trabajador.objects.filter(
            activo=True,
            area__isnull=False,
        ).exclude(
            area__nombre__istartswith='Gerencia General'
        ).values_list('area_id', flat=True)
    )
    ids.update(
        Trabajador.objects.filter(
            activo=True,
            areas_supervisadas__isnull=False,
        ).exclude(
            areas_supervisadas__nombre__istartswith='Gerencia General'
        ).values_list('areas_supervisadas__id', flat=True)
    )

    if not ids:
        return Area.objects.none()

    return Area.objects.filter(id__in=ids).exclude(nombre__istartswith='Gerencia General').order_by('nombre')


def _areas_liderazgo_activas_qs(exclude_trabajador_id=None):
    """
    Areas que actualmente tienen liderazgo activo (supervisor/responsable).
    """
    lideres = Trabajador.objects.filter(activo=True).filter(_q_supervisor() | _q_responsable())
    if exclude_trabajador_id:
        lideres = lideres.exclude(id=exclude_trabajador_id)

    ids = set(
        lideres.filter(area__isnull=False).exclude(area__nombre__istartswith='Gerencia General').values_list('area_id', flat=True)
    )
    ids.update(
        lideres.filter(areas_supervisadas__isnull=False).exclude(areas_supervisadas__nombre__istartswith='Gerencia General').values_list('areas_supervisadas__id', flat=True)
    )

    if not ids:
        return Area.objects.none()

    return Area.objects.filter(id__in=ids).exclude(nombre__istartswith='Gerencia General').order_by('nombre')


def _areas_bajo_responsabilidad(trabajador):
    if not trabajador:
        return []

    # Supervisor: solo areas explicitamente asignadas + su area base.
    if _es_supervisor(trabajador):
        ids = set(trabajador.areas_supervisadas.values_list('id', flat=True))
        if trabajador.area_id:
            ids.add(trabajador.area_id)

        return list(
            Area.objects.filter(id__in=ids)
            .exclude(nombre__istartswith='Gerencia General')
            .order_by('nombre')
        )

    # Responsable: prioriza areas explicitamente asignadas; si no tiene,
    # usa areas operativas con liderazgo activo como fallback seguro.
    if _es_responsable(trabajador):
        ids_asignadas = set(trabajador.areas_supervisadas.values_list('id', flat=True))
        if trabajador.area_id:
            ids_asignadas.add(trabajador.area_id)

        if ids_asignadas:
            return list(
                Area.objects.filter(id__in=ids_asignadas)
                .exclude(nombre__istartswith='Gerencia General')
                .order_by('nombre')
            )

        areas_liderazgo = list(_areas_liderazgo_activas_qs(exclude_trabajador_id=trabajador.id))
        if areas_liderazgo:
            return areas_liderazgo

        return list(_areas_operativas_activas_qs())

    # Gerente puro: alcance en areas operativas reales.
    if _es_gerente_puro(trabajador):
        return list(
            _areas_operativas_activas_qs()
        )

    # Trabajador sin jerarquia de evaluacion.
    return []



def _supervisores_activos_por_area(area):
    return Trabajador.objects.filter(
        activo=True,
    ).filter(
        _q_supervisor()
    ).filter(
        Q(area=area) | Q(areas_supervisadas=area)
    ).distinct()


def _responsables_activos_por_area(area):
    return Trabajador.objects.filter(
        activo=True,
    ).filter(
        _q_responsable()
    ).filter(
        Q(area=area) | Q(areas_supervisadas=area)
    ).distinct()


def _objetivos_evaluacion_por_area(evaluador, area):
    if not area:
        return Trabajador.objects.none()

    if _es_supervisor(evaluador):
        return Trabajador.objects.filter(
            area=area,
            activo=True,
        ).filter(
            _q_trabajador_base()
        ).exclude(id=evaluador.id).order_by('apellido_paterno', 'apellido_materno', 'nombres')

    if _es_responsable(evaluador):
        supervisores = _supervisores_activos_por_area(area).exclude(id=evaluador.id)
        if supervisores.exists():
            return supervisores.order_by('apellido_paterno', 'apellido_materno', 'nombres')

        return Trabajador.objects.filter(
            area=area,
            activo=True,
        ).filter(
            _q_trabajador_base()
        ).exclude(id=evaluador.id).order_by('apellido_paterno', 'apellido_materno', 'nombres')

    return Trabajador.objects.none()


def _clasificar_areas_responsable_por_supervision(responsable, areas):
    areas_con_supervisor = []
    areas_sin_supervisor = []

    for area in areas:
        tiene_supervisor = _supervisores_activos_por_area(area).exclude(id=responsable.id).exists()
        if tiene_supervisor:
            areas_con_supervisor.append(area)
        else:
            areas_sin_supervisor.append(area)

    return areas_con_supervisor, areas_sin_supervisor


def _objetivos_supervisores_consolidados(responsable, areas):
    if not areas:
        return Trabajador.objects.none()

    return Trabajador.objects.filter(
        activo=True,
    ).filter(
        _q_supervisor()
    ).filter(
        Q(area__in=areas) | Q(areas_supervisadas__in=areas)
    ).exclude(
        id=responsable.id,
    ).distinct().order_by('apellido_paterno', 'apellido_materno', 'nombres')


def _resumen_operativo_area(area, hoy):
    colaboradores_qs = Trabajador.objects.filter(
        area=area,
        activo=True,
    ).filter(_q_trabajador_base())

    total_colaboradores = colaboradores_qs.count()
    evaluados_mes = EvaluacionMensual.objects.filter(
        trabajador__in=colaboradores_qs,
        fecha_evaluacion__year=hoy.year,
        fecha_evaluacion__month=hoy.month,
    ).values('trabajador_id').distinct().count()

    return {
        'total_colaboradores': total_colaboradores,
        'evaluados_mes': evaluados_mes,
        'pendientes_mes': max(total_colaboradores - evaluados_mes, 0),
    }


def _resolver_cascada_objetivos_gerencia(gerente, hoy):
    areas_objetivo = list(_areas_operativas_activas_qs())
    objetivos_ids = set()
    areas_con_liderazgo = []
    areas_directas = []

    for area_obj in areas_objetivo:
        responsables_area = _responsables_activos_por_area(area_obj).exclude(id=gerente.id)
        supervisores_area = _supervisores_activos_por_area(area_obj).exclude(id=gerente.id)

        resumen_area = _resumen_operativo_area(area_obj, hoy)

        if responsables_area.exists():
            objetivos_ids.update(responsables_area.values_list('id', flat=True))
            areas_con_liderazgo.append({
                'area': area_obj,
                'nivel_cascada': 'RESPONSABLE',
                'total_lideres_objetivo': responsables_area.count(),
                **resumen_area,
            })
            continue

        if supervisores_area.exists():
            objetivos_ids.update(supervisores_area.values_list('id', flat=True))
            areas_con_liderazgo.append({
                'area': area_obj,
                'nivel_cascada': 'SUPERVISOR',
                'total_lideres_objetivo': supervisores_area.count(),
                **resumen_area,
            })
            continue

        areas_directas.append({
            'area': area_obj,
            **resumen_area,
        })

    lideres_objetivo = Trabajador.objects.filter(
        id__in=objetivos_ids,
        activo=True,
    ).select_related('area').prefetch_related('areas_supervisadas').order_by(
        'apellido_paterno',
        'apellido_materno',
        'nombres',
    )

    return {
        'lideres_objetivo': lideres_objetivo,
        'areas_con_liderazgo': areas_con_liderazgo,
        'areas_directas': areas_directas,
    }


def _puede_evaluar_objetivo(evaluador, objetivo):
    if not evaluador or not objetivo or evaluador.id == objetivo.id:
        return False

    cargo_objetivo = _cargo_jerarquico_actual(objetivo)
    areas_ids = {a.id for a in _areas_bajo_responsabilidad(evaluador)}

    # Un objetivo puede pertenecer por area principal o por areas_supervisadas.
    areas_objetivo_ids = set()
    if getattr(objetivo, 'area_id', None):
        areas_objetivo_ids.add(objetivo.area_id)
    areas_objetivo_ids.update(objetivo.areas_supervisadas.values_list('id', flat=True))

    areas_objetivo = list(Area.objects.filter(id__in=areas_objetivo_ids)) if areas_objetivo_ids else []

    def _hay_area_compartida():
        return bool(areas_objetivo_ids & areas_ids)

    if _es_supervisor(evaluador):
        return (
            cargo_objetivo == Trabajador.CargoJerarquico.TRABAJADOR
            and _hay_area_compartida()
        )

    if _es_responsable(evaluador):
        if not _hay_area_compartida():
            return False

        if cargo_objetivo == Trabajador.CargoJerarquico.SUPERVISOR:
            return True

        if cargo_objetivo != Trabajador.CargoJerarquico.TRABAJADOR:
            return False

        area_control = objetivo.area
        if area_control is None:
            area_control = next((a for a in areas_objetivo if a.id in areas_ids), None)
        if area_control is None:
            return False

        # Solo puede evaluar trabajadores cuando el area no tiene supervisor activo.
        return not _supervisores_activos_por_area(area_control).exclude(id=evaluador.id).exists()

    if _es_gerente_puro(evaluador):
        if cargo_objetivo == Trabajador.CargoJerarquico.RESPONSABLE:
            return True

        if not areas_objetivo:
            return False

        if cargo_objetivo == Trabajador.CargoJerarquico.SUPERVISOR:
            for area in areas_objetivo:
                hay_responsable = _responsables_activos_por_area(area).exclude(id=evaluador.id).exists()
                if not hay_responsable:
                    return True
            return False

        if cargo_objetivo != Trabajador.CargoJerarquico.TRABAJADOR:
            return False

        for area in areas_objetivo:
            hay_responsable = _responsables_activos_por_area(area).exclude(id=evaluador.id).exists()
            hay_supervisor = _supervisores_activos_por_area(area).exclude(id=evaluador.id).exists()
            if (not hay_responsable) and (not hay_supervisor):
                return True
        return False

    return False


def _obtener_feriados(anio):
    # Calcular Pascua
    pascua = easter(anio)
    jueves_santo = pascua - datetime.timedelta(days=3)
    viernes_santo = pascua - datetime.timedelta(days=2)
    
    feriados = [
        datetime.date(anio, 1, 1),   # Año Nuevo
        jueves_santo,                # Jueves Santo
        viernes_santo,               # Viernes Santo
        datetime.date(anio, 5, 1),   # Día del Trabajo
        datetime.date(anio, 6, 7),   # Batalla de Arica
        datetime.date(anio, 6, 29),  # San Pedro y San Pablo
        datetime.date(anio, 7, 23),  # Día de la Fuerza Aérea
        datetime.date(anio, 7, 28),  # Fiestas Patrias
        datetime.date(anio, 7, 29),  # Fiestas Patrias
        datetime.date(anio, 8, 6),   # Batalla de Junín
        datetime.date(anio, 8, 30),  # Santa Rosa de Lima
        datetime.date(anio, 10, 8),  # Combate de Angamos
        datetime.date(anio, 11, 1),  # Todos los Santos
        datetime.date(anio, 12, 8),  # Inmaculada Concepción
        datetime.date(anio, 12, 9),  # Batalla de Ayacucho
        datetime.date(anio, 12, 25), # Navidad
    ]
    return feriados


def _inicio_periodo(hoy, periodo):
    return hoy.replace(day=1) if periodo == 'mes' else hoy.replace(month=1, day=1)


def _filtrar_tardanzas_validas(tareos):
    return tareos.filter(horas_tardanza__gte=TARDANZA_MINIMA_HORAS)


def _normalizar_semestre(valor, default=1):
    try:
        semestre = int(valor)
    except (TypeError, ValueError):
        semestre = default
    return semestre if semestre in [1, 2] else default


def _normalizar_mes(valor, default):
    try:
        mes = int(valor)
    except (TypeError, ValueError):
        mes = default
    return mes if 1 <= mes <= 12 else default


def _anio_semestre_por_fecha(fecha_ref):
    # Semestre 1: Febrero -> Julio (anio actual)
    if 2 <= fecha_ref.month <= 7:
        return fecha_ref.year, 1

    # Semestre 2: Agosto -> Enero
    # Para enero, el anio de referencia del semestre es el anio previo.
    if fecha_ref.month == 1:
        return fecha_ref.year - 1, 2
    return fecha_ref.year, 2


def _resolver_rango_semestre_por_filtro(anio, semestre):
    anio_objetivo = int(anio)
    semestre_objetivo = _normalizar_semestre(semestre, default=1)

    if semestre_objetivo == 1:
        return datetime.date(anio_objetivo, 2, 1), datetime.date(anio_objetivo, 7, 31)
    return datetime.date(anio_objetivo, 8, 1), datetime.date(anio_objetivo + 1, 1, 31)


def _resolver_anio_semestre_request(request, hoy):
    anio_param = request.GET.get('anio')
    try:
        anio = int(anio_param) if anio_param is not None else None
    except (TypeError, ValueError):
        anio = None

    semestre_param = request.GET.get('semestre')
    if semestre_param in ['1', '2']:
        if anio is None:
            anio = _anio_semestre_por_fecha(hoy)[0]
        return anio, int(semestre_param)

    # Compatibilidad con filtros antiguos (periodo/mes).
    mes_param = request.GET.get('mes')
    if mes_param:
        try:
            mes = int(mes_param)
        except (TypeError, ValueError):
            mes = hoy.month

        if mes < 1 or mes > 12:
            mes = hoy.month

        anio_mes = anio if anio is not None else hoy.year
        return _anio_semestre_por_fecha(datetime.date(anio_mes, mes, 1))

    anio_hoy, semestre_hoy = _anio_semestre_por_fecha(hoy)
    return (anio if anio is not None else anio_hoy), semestre_hoy


def _resolver_filtro_periodo_request(request, hoy, periodo_default='semestre'):
    try:
        anio = int(request.GET.get('anio', hoy.year))
    except (TypeError, ValueError):
        anio = hoy.year

    periodo = request.GET.get('periodo')
    if periodo not in ['mes', 'semestre']:
        if request.GET.get('semestre') in ['1', '2']:
            periodo = 'semestre'
        elif request.GET.get('mes'):
            periodo = 'mes'
        else:
            periodo = periodo_default

    if periodo == 'mes':
        mes = _normalizar_mes(request.GET.get('mes'), default=hoy.month)
        return {
            'periodo': 'mes',
            'anio': anio,
            'mes': mes,
            'semestre': None,
        }

    anio_semestre, semestre = _resolver_anio_semestre_request(request, hoy)
    return {
        'periodo': 'semestre',
        'anio': anio_semestre,
        'mes': None,
        'semestre': semestre,
    }


def _resolver_rango_periodo(hoy, periodo, anio=None, mes=None, semestre=None):
    anio_objetivo = int(anio) if anio else hoy.year

    if periodo == 'semestre':
        semestre_default = _anio_semestre_por_fecha(hoy)[1]
        return _resolver_rango_semestre_por_filtro(
            anio_objetivo,
            _normalizar_semestre(semestre, default=semestre_default)
        )
    if periodo == 'mes':
        mes_objetivo = _normalizar_mes(mes, default=hoy.month)
        inicio = datetime.date(anio_objetivo, mes_objetivo, 1)
        ultimo_dia = calendar.monthrange(anio_objetivo, mes_objetivo)[1]
        fin = datetime.date(anio_objetivo, mes_objetivo, ultimo_dia)
    else:
        inicio = datetime.date(anio_objetivo, 1, 1)
        fin = datetime.date(anio_objetivo, 12, 31)

    return inicio, fin


def _resolver_rango_semestre(fecha_ref):
    """
    Semestres calendario:
    - Semestre 1: Febrero -> Julio
    - Semestre 2: Agosto -> Enero
    """
    if 2 <= fecha_ref.month <= 7:
        inicio = datetime.date(fecha_ref.year, 2, 1)
        fin = datetime.date(fecha_ref.year, 7, 31)
    elif fecha_ref.month == 1:
        inicio = datetime.date(fecha_ref.year - 1, 8, 1)
        fin = datetime.date(fecha_ref.year, 1, 31)
    else:
        inicio = datetime.date(fecha_ref.year, 8, 1)
        fin = datetime.date(fecha_ref.year + 1, 1, 31)
    return inicio, fin


def _meses_antiguedad_por_fecha(fecha_ingreso, fecha_ref):
    if not fecha_ingreso or not fecha_ref or fecha_ref < fecha_ingreso:
        return 0
    return ((fecha_ref.year - fecha_ingreso.year) * 12) + (fecha_ref.month - fecha_ingreso.month) + 1


def _esta_en_periodo_prueba(trabajador, fecha_ref):
    meses_antiguedad = _meses_antiguedad_por_fecha(getattr(trabajador, 'fecha_ingreso', None), fecha_ref)
    return 1 <= meses_antiguedad <= 3


def _estado_evaluacion_para_fecha(trabajador, fecha_ref):
    evals = trabajador.evaluaciones.all()
    ya_evaluado_este_mes = evals.filter(
        fecha_evaluacion__year=fecha_ref.year,
        fecha_evaluacion__month=fecha_ref.month,
    ).exists()

    en_prueba = _esta_en_periodo_prueba(trabajador, fecha_ref)
    es_mes_semestral = fecha_ref.month in [1, 7]

    if en_prueba:
        return {
            'es_nuevo': True,
            'hito_tipo': 'NUEVO',
            'btn_hito_activo': not ya_evaluado_este_mes,
            'btn_mensual_activo': False,
            'en_periodo_prueba': True,
        }

    if es_mes_semestral:
        return {
            'es_nuevo': False,
            'hito_tipo': 'SEMESTRAL',
            'btn_hito_activo': not ya_evaluado_este_mes,
            'btn_mensual_activo': False,
            'en_periodo_prueba': False,
        }

    return {
        'es_nuevo': False,
        'hito_tipo': 'MENSUAL',
        'btn_hito_activo': False,
        'btn_mensual_activo': not ya_evaluado_este_mes,
        'en_periodo_prueba': False,
    }


def _fecha_referencia_periodo(hoy, periodo, anio=None, mes=None, semestre=None):
    _, fin_periodo = _resolver_rango_periodo(hoy, periodo, anio=anio, mes=mes, semestre=semestre)
    return fin_periodo


def _calcular_desempeno_por_periodo(trabajador, inicio_periodo, fin_periodo):
    inicio_calculo = inicio_periodo
    fin_calculo = fin_periodo

    if fin_calculo < inicio_calculo:
        return {
            'promedio_mensual': 0.0,
            'promedio_semestral': 0.0,
            'desempeno_compuesto': 0.0,
            'promedio_general_semestre': 0.0,
            'cantidad_evaluaciones_semestre': 0,
            'evaluaciones_requeridas_semestre': 6,
            'semestre_inicio': inicio_calculo,
            'semestre_fin': fin_calculo,
        }

    evaluaciones = trabajador.evaluaciones.filter(
        fecha_evaluacion__gte=inicio_calculo,
        fecha_evaluacion__lte=fin_calculo,
    )

    cantidad_evaluaciones = evaluaciones.count()
    promedio_general = evaluaciones.aggregate(
        Avg('promedio_final')
    )['promedio_final__avg'] or 0.0

    promedio_mensual = evaluaciones.filter(tipo__in=['MENSUAL', 'NUEVO']).aggregate(
        Avg('promedio_final')
    )['promedio_final__avg'] or 0.0

    promedio_semestral = evaluaciones.filter(tipo='SEMESTRAL').aggregate(
        Avg('promedio_final')
    )['promedio_final__avg'] or 0.0

    tiene_semestral = evaluaciones.filter(tipo='SEMESTRAL').exists()
    evaluaciones_requeridas = 6

    # Regla acordada: si no llega a 6 evaluaciones en el semestre, desempeño = promedio simple.
    if 0 < cantidad_evaluaciones < evaluaciones_requeridas:
        desempeno_compuesto = float(promedio_general)
    else:
        if tiene_semestral:
            desempeno_compuesto = (float(promedio_mensual) * 0.6) + (float(promedio_semestral) * 0.4)
        elif cantidad_evaluaciones > 0:
            desempeno_compuesto = float(promedio_general)
        else:
            desempeno_compuesto = 0.0

    return {
        'promedio_mensual': float(promedio_mensual),
        'promedio_semestral': float(promedio_semestral),
        'desempeno_compuesto': float(desempeno_compuesto),
        'promedio_general_semestre': float(promedio_general),
        'cantidad_evaluaciones_semestre': int(cantidad_evaluaciones),
        'evaluaciones_requeridas_semestre': evaluaciones_requeridas,
        'semestre_inicio': inicio_calculo,
        'semestre_fin': fin_calculo,
    }


def _calcular_asistencia_por_periodo(trabajador, inicio_periodo, fin_periodo):
    hoy = timezone.now().date()
    
    # 1. EL ESCUDO: Definimos un límite para nunca mirar al futuro
    limite_fecha = min(fin_periodo, hoy)

    anios_involucrados = set([inicio_periodo.year, fin_periodo.year])
    feriados_lista = []
    for anio in anios_involucrados:
        feriados_lista.extend(_obtener_feriados(anio))

    # 2. Consultar Tareos: Filtramos HASTA EL LÍMITE (Hoy)
    tareos = TareoDiario.objects.filter(
        trabajador=trabajador,
        fecha__gte=inicio_periodo,
        fecha__lte=limite_fecha, # <--- IGNORA TODO LO QUE PASE DE HOY
    ).exclude(
        fecha__in=feriados_lista
    ).exclude(
        fecha__week_day=1 
    )

    # 3. Faltas Reales HASTA HOY
    faltas_bd = tareos.filter(resultado='F').exclude(
        justificacion__estado_solicitud='APROBADO'
    ).distinct().count()

    # 4. Días laborables matemáticos (Hasta hoy)
    faltas_sinteticas = 0
    total_dias_reales_hasta_hoy = 0
    fechas_con_tareo = set(tareos.values_list('fecha', flat=True))

    delta = limite_fecha - inicio_periodo
    for i in range(delta.days + 1):
        dia_evaluado = inicio_periodo + datetime.timedelta(days=i)
        if dia_evaluado.weekday() != 6 and dia_evaluado not in feriados_lista:
            total_dias_reales_hasta_hoy += 1
            if dia_evaluado not in fechas_con_tareo:
                faltas_sinteticas += 1

    total_faltas = faltas_bd + faltas_sinteticas

    # 5. Cálculo de Tardanzas
    descuento_tardanza_dias = 0.0
    tardanzas_penalizadas = 0
    
    tareos_con_marcacion = tareos.exclude(hora_entrada_real__isnull=True)
    
    for tareo in tareos_con_marcacion:
        hora_entrada = tareo.hora_entrada_real
        hora_referencia = _hora_referencia_entrada_por_fecha(tareo.fecha)
        hora_tardanza_leve_max = _hora_tardanza_leve_max_por_fecha(tareo.fecha)

        if hora_referencia < hora_entrada <= hora_tardanza_leve_max:
            descuento_tardanza_dias += 0.25
            tardanzas_penalizadas += 1
        elif hora_entrada > hora_tardanza_leve_max:
            descuento_tardanza_dias += 0.5
            tardanzas_penalizadas += 1

    # 6. Score Final
    faltas_ajustadas = total_faltas + descuento_tardanza_dias

    porc_asistencia = 0
    if total_dias_reales_hasta_hoy > 0:
        porc_asistencia = ((total_dias_reales_hasta_hoy - faltas_ajustadas) / total_dias_reales_hasta_hoy) * 100
        porc_asistencia = max(0, porc_asistencia)

    # Nota 1-10 según la metodología (PPT): 10 − (Tardanzas×0.2) − (Inasistencias×0.3),
    # con piso 1. Es la que alimenta el score. `porcentaje` queda como % real de
    # asistencia solo para mostrar (kpi_asistencia).
    nota = max(1.0, 10.0 - (tardanzas_penalizadas * 0.2) - (total_faltas * 0.3))

    return {
        'porcentaje': float(porc_asistencia),
        'nota': float(nota),
        'faltas': total_faltas,
        'tardanzas': tardanzas_penalizadas,
        'total_dias': total_dias_reales_hasta_hoy
    }

# Pesos de la metodología de evaluación (metodología de RRHH). Los 4 aspectos
# aplican por igual a todos los trabajadores (oficina y mina); la modalidad no
# cambia el cálculo, es solo informativa. Suma 1.0.
PESOS_EVALUACION = {
    'DESEMPENO': 0.10,
    'MEDIDAS_DISCIPLINARIAS': 0.40,
    'CONOCIMIENTO': 0.20,
    'ASISTENCIA': 0.30,
}

# Factor que resta cada tipo de sanción en la nota de Medidas disciplinarias
# (fórmula 10 − Σ, con piso 1). Solo Verbal/Escrita/Suspensión descuentan;
# Oral y Otra quedan en 0 por decisión de negocio (no contempladas en la metodología).
FACTOR_SANCION = {
    Sancion.Tipo.VERBAL: 2,
    Sancion.Tipo.ESCRITA: 3,
    Sancion.Tipo.SUSPENSION: 5,
}


def _pesos_para_modalidad(trabajador):
    """Pesos de los 4 aspectos (iguales para todos). Devuelve además la
    modalidad real del trabajador solo como etiqueta informativa."""
    modalidad = getattr(trabajador, 'modalidad_evaluacion', None)
    return modalidad, PESOS_EVALUACION


def _calcular_medidas_disciplinarias(trabajador, inicio, fin):
    """Nota 1-10 de Medidas disciplinarias en el periodo.

    10 − (Verbal×2) − (Escrita×3) − (Suspensión×5), con piso 1. Sin sanciones = 10.
    """
    sanciones = trabajador.sanciones.filter(fecha_sancion__range=(inicio, fin))
    descuento = 0
    for tipo, factor in FACTOR_SANCION.items():
        descuento += sanciones.filter(tipo=tipo).count() * factor
    return max(1.0, 10.0 - descuento)


# La nota de Conocimiento se carga como examen sobre 20 (metodología PPT) y se
# convierte a escala 1-10 dividiendo entre 2. Sin examen cargado inicia en 10.
CONOCIMIENTO_EXAMEN_INICIAL = 20  # equivalente a 10/10
CONOCIMIENTO_INICIAL = 10.0       # nota 1-10 por defecto (20/20 ÷ 2)


def _calcular_conocimiento(trabajador, inicio, fin):
    """Nota 1-10 de Conocimiento. Se guarda el examen sobre 20 y se convierte
    ÷2 a escala 1-10. Sin examen cargado en el periodo, vale 10."""
    nota_examen = (trabajador.notas_conocimiento
                   .filter(fecha__range=(inicio, fin))
                   .order_by('-fecha', '-id')
                   .values_list('nota', flat=True)
                   .first())
    if nota_examen is None:
        return CONOCIMIENTO_INICIAL
    return min(10.0, float(nota_examen) / 2.0)


def _ponderar_score(trabajador, nota_desempeno, asistencia_escala_10, inicio, fin):
    """Combina los aspectos con los pesos de la modalidad. Fuente única de verdad
    para el score ponderado (la usan el cálculo del periodo y los exports)."""
    nota_disciplinaria = _calcular_medidas_disciplinarias(trabajador, inicio, fin)
    nota_conocimiento = _calcular_conocimiento(trabajador, inicio, fin)

    # Conocimiento inicia en 10 para todos, así que siempre es un número: no hay
    # aspecto faltante ni re-normalización, se aplican los 4 pesos completos.
    componentes = {
        'DESEMPENO': float(nota_desempeno),
        'MEDIDAS_DISCIPLINARIAS': float(nota_disciplinaria),
        'ASISTENCIA': float(asistencia_escala_10),
        'CONOCIMIENTO': float(nota_conocimiento),
    }

    modalidad, pesos = _pesos_para_modalidad(trabajador)
    score = sum(componentes[k] * p for k, p in pesos.items())

    return {
        'score_total': float(score),
        'nota_disciplinaria': float(nota_disciplinaria),
        'nota_conocimiento': float(nota_conocimiento),
        'modalidad': modalidad,
        'pesos_aplicados': pesos,
        'conocimiento_pendiente': False,
    }


def _calcular_score_total(trabajador, hoy, periodo, anio=None, mes=None, semestre=None):
    inicio_periodo, fin_periodo = _resolver_rango_periodo(hoy, periodo, anio=anio, mes=mes, semestre=semestre)

    desempeno = _calcular_desempeno_por_periodo(trabajador, inicio_periodo, fin_periodo)
    asistencia_data = _calcular_asistencia_por_periodo(trabajador, inicio_periodo, fin_periodo)

    nota_desempeno_periodo = desempeno['desempeno_compuesto']
    # Nota de asistencia (PPT) para el score; `porcentaje` es el % real (display).
    asistencia_escala_10 = asistencia_data['nota']

    ponderacion = _ponderar_score(
        trabajador, nota_desempeno_periodo, asistencia_escala_10, inicio_periodo, fin_periodo)

    return {
        'promedio_mensual': desempeno['promedio_mensual'],
        'promedio_semestral': desempeno['promedio_semestral'],
        'desempeno_compuesto': desempeno['desempeno_compuesto'],
        'nota_desempeno_periodo': float(nota_desempeno_periodo),
        'asistencia_pct': asistencia_data['porcentaje'],
        'nota_asistencia': asistencia_data['nota'],
        'nota_disciplinaria': ponderacion['nota_disciplinaria'],
        'nota_conocimiento': ponderacion['nota_conocimiento'],
        'modalidad': ponderacion['modalidad'],
        'pesos_aplicados': ponderacion['pesos_aplicados'],
        'conocimiento_pendiente': ponderacion['conocimiento_pendiente'],
        'score_total': ponderacion['score_total'],
        'faltas': asistencia_data['faltas'],
        'tardanzas': asistencia_data['tardanzas'],
        'total_dias': asistencia_data['total_dias'],
        'en_periodo_prueba': _esta_en_periodo_prueba(trabajador, fin_periodo),
        'semestre_inicio': inicio_periodo,
        'semestre_fin': fin_periodo,
    }


def _puede_evaluar_semestral(trabajador, hoy):
    if _esta_en_periodo_prueba(trabajador, hoy):
        return False

    if hoy.month not in [1, 7]:
        return False

    ya_tiene_semestral_en_mes = trabajador.evaluaciones.filter(
        tipo='SEMESTRAL',
        fecha_evaluacion__year=hoy.year,
        fecha_evaluacion__month=hoy.month,
    ).exists()
    return not ya_tiene_semestral_en_mes


def _serializar_item_ranking_tabla(item, tipo_persona):
    trabajador = item['trabajador']
    return {
        'id': trabajador.id,
        'nombres': trabajador.nombres or '',
        'apellido_paterno': trabajador.apellido_paterno or '',
        'apellido_materno': trabajador.apellido_materno or '',
        'cargo': trabajador.cargo or 'Sin cargo',
        'nombre_completo': f"{trabajador.nombres or ''} {trabajador.apellido_paterno or ''} {trabajador.apellido_materno or ''}".strip(),
        'tipo_persona': tipo_persona,
        'area_nombre': item.get('area_nombre') or 'Sin Área',
        'sede_nombre': trabajador.sede.nombre if getattr(trabajador, 'sede', None) else 'Sin sede',
        'promedio_mensual': item.get('promedio_mensual', 0),
        'promedio_mensual_porc': item.get('promedio_mensual_porc', 0),
        'promedio_semestral': item.get('promedio_semestral', 0),
        'promedio_semestral_porc': item.get('promedio_semestral_porc', 0),
        'asistencia_avg': item.get('asistencia_avg', 0),
        'nota_disciplinaria': item.get('nota_disciplinaria', 0),
        'nota_conocimiento': item.get('nota_conocimiento', None),
        'modalidad': item.get('modalidad', None),
        'score': item.get('score', 0),
    }


def _guardar_evaluacion_con_reintento_secuencia(evaluacion):
    """
    En PostgreSQL puede quedar desfasada la secuencia del PK tras migraciones/importaciones.
    Si ocurre duplicate key en pkey, resincroniza la secuencia y reintenta una vez.
    """
    try:
        # Savepoint local: si falla este save, no rompe la transacción externa.
        with transaction.atomic():
            evaluacion.save()
            return
    except IntegrityError as exc:
        error_texto = str(exc)
        if 'evaluacionmensual_pkey' not in error_texto:
            raise

    sql_reset = connection.ops.sequence_reset_sql(no_style(), [EvaluacionMensual])
    if not sql_reset:
        raise IntegrityError('No fue posible resincronizar la secuencia de EvaluacionMensual.')

    with connection.cursor() as cursor:
        for sentencia in sql_reset:
            cursor.execute(sentencia)

    with transaction.atomic():
        evaluacion.save()

@login_required
def panel_jefe(request):
    try:
        perfil = request.user.trabajador
    except Exception:
        return render(request, 'metricas_ceneris/error.html', {'mensaje': 'No tienes perfil asignado.'})

    # Gerente puro usa su panel dedicado (sin cambios de vista).
    if _es_gerente_puro(perfil):
        return redirect('metricas_ceneris:dashboard_gerente')

    ruta_actual = getattr(getattr(request, 'resolver_match', None), 'url_name', '')

    if _es_responsable(perfil) and ruta_actual != 'panel_responsable':
        return redirect('metricas_ceneris:panel_responsable')

    if not (_es_supervisor(perfil) or _es_responsable(perfil)):
        return redirect('metricas_ceneris:dashboard_trabajador')

    jefe = perfil
    hoy = timezone.now().date()

    areas_disponibles = _areas_bajo_responsabilidad(jefe)
    if not areas_disponibles:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes áreas asignadas para evaluación. Contacta a RRHH/Calidad.'
        })

    area_id_param = request.GET.get('area_id')
    area = None

    if area_id_param:
        try:
            area_id = int(area_id_param)
            area = next((a for a in areas_disponibles if a.id == area_id), None)
        except (TypeError, ValueError):
            area = None

    if area is None and areas_disponibles:
        area = areas_disponibles[0]

    if not area:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No hay áreas disponibles para evaluación. Contacta a RRHH/Calidad.'
        })

    es_responsable = _es_responsable(jefe)

    tarjetas_areas = []
    tarjetas_sin_jefe = []
    supervisores_asignados_map = {}
    for area_item in areas_disponibles:
        objetivos_area = _objetivos_evaluacion_por_area(jefe, area_item)
        evaluados = EvaluacionMensual.objects.filter(
            trabajador__in=objetivos_area,
            fecha_evaluacion__year=hoy.year,
            fecha_evaluacion__month=hoy.month,
        ).values('trabajador_id').distinct().count()

        hay_supervisores = _supervisores_activos_por_area(area_item).exclude(id=jefe.id).exists()
        tipo_objetivo = 'supervisores' if es_responsable and hay_supervisores else 'trabajadores'
        
        tarjeta = {
            'area': area_item,
            'total_colaboradores': objetivos_area.count(),
            'evaluados_mes': evaluados,
            'pendientes_mes': max(objetivos_area.count() - evaluados, 0),
            'activa': area_item.id == area.id,
            'tipo_objetivo': tipo_objetivo,
        }

        if es_responsable and not hay_supervisores:
            tarjetas_sin_jefe.append(tarjeta)
        else:
            tarjetas_areas.append(tarjeta)

        if es_responsable and hay_supervisores:
            for supervisor in _supervisores_activos_por_area(area_item).exclude(id=jefe.id):
                supervisor_data = supervisores_asignados_map.setdefault(
                    supervisor.id,
                    {
                        'supervisor': supervisor,
                        'areas': set(),
                        'area_ids': set(),
                    }
                )
                supervisor_data['areas'].add(area_item.nombre)
                supervisor_data['area_ids'].add(area_item.id)

    supervisores_asignados = []
    for supervisor_data in supervisores_asignados_map.values():
        supervisor = supervisor_data['supervisor']
        evaluado_mes = EvaluacionMensual.objects.filter(
            trabajador=supervisor,
            fecha_evaluacion__year=hoy.year,
            fecha_evaluacion__month=hoy.month,
        ).exists()

        estado_supervisor = _estado_evaluacion_para_fecha(supervisor, hoy)
        tipo_eval = None
        if estado_supervisor['btn_hito_activo']:
            if estado_supervisor['hito_tipo'] == 'NUEVO':
                tipo_eval = 'nuevo'
            else:
                tipo_eval = 'semestral'
        elif estado_supervisor['btn_mensual_activo']:
            tipo_eval = 'mensual'

        areas_ordenadas = sorted(supervisor_data['areas'])
        area_ids_ordenados = sorted(supervisor_data['area_ids'])

        supervisores_asignados.append({
            'supervisor': supervisor,
            'areas_label': ', '.join(areas_ordenadas) if areas_ordenadas else 'Sin área operativa',
            'total_areas': len(areas_ordenadas),
            'area_foco_id': area_ids_ordenados[0] if area_ids_ordenados else None,
            'evaluado_mes': evaluado_mes,
            'tipo_eval': tipo_eval,
        })

    supervisores_asignados = sorted(
        supervisores_asignados,
        key=lambda item: (
            (item['supervisor'].apellido_paterno or '').lower(),
            (item['supervisor'].apellido_materno or '').lower(),
            (item['supervisor'].nombres or '').lower(),
        ),
    )

    trabajadores = _objetivos_evaluacion_por_area(jefe, area)

    hay_supervisores_area_activa = _supervisores_activos_por_area(area).exclude(id=jefe.id).exists()
    if es_responsable and hay_supervisores_area_activa:
        objetivo_label_plural = 'Supervisores'
    else:
        objetivo_label_plural = 'Trabajadores'
    
    # KPIs y Gráficos
    total_evaluaciones_mes = 0
    promedios_area = []
    mejor_trabajador = None
    mejor_nota = -1
    promedios_por_categoria = {'OPERACIONAL': [], 'ADMINISTRATIVO': [], 'HABILIDADES': []}

    for t in trabajadores:
        evals = t.evaluaciones.all()
        ultima_ev = evals.first()

        # Usamos tu función auxiliar de estado
        estado_eval = _estado_evaluacion_para_fecha(t, hoy)
        t.hito_tipo = estado_eval['hito_tipo']
        t.btn_hito_activo = estado_eval['btn_hito_activo']
        t.btn_mensual_activo = estado_eval['btn_mensual_activo']

        if ultima_ev:
            t.ultima_nota = ultima_ev.promedio_final
            promedios_area.append(t.ultima_nota)
            total_evaluaciones_mes += 1
            if t.ultima_nota > mejor_nota:
                mejor_nota = t.ultima_nota
                mejor_trabajador = t
                
            for cat in ['OPERACIONAL', 'ADMINISTRATIVO', 'HABILIDADES']:
                avg_cat = ultima_ev.puntajes.filter(categoria=cat).aggregate(Avg('nota'))['nota__avg']
                if avg_cat: promedios_por_categoria[cat].append(avg_cat)
        else:
            t.ultima_nota = 0

    promedio_area_total = round(sum(promedios_area) / len(promedios_area), 1) if promedios_area else 0
    radar_data = [
        round(sum(promedios_por_categoria[cat]) / len(promedios_por_categoria[cat]), 1) if promedios_por_categoria[cat] else 0 
        for cat in ['OPERACIONAL', 'ADMINISTRATIVO', 'HABILIDADES']
    ]

    if es_responsable:
        areas_section_title = 'Supervisores Asignados'
        areas_section_desc = 'Visualiza todos los supervisores asignados dentro de tus áreas y evalúalos de forma directa.'
    else:
        areas_section_title = 'Tus Áreas Asignadas'
        areas_section_desc = 'Selecciona un área bajo tu responsabilidad para evaluar su equipo.'

    return render(request, 'metricas_ceneris/evaluaciones/panel_jefe.html', {
        'jefe': jefe,
        'area': area,
        'es_responsable': es_responsable,
        'tarjetas_areas': tarjetas_areas,
        'tarjetas_sin_jefe': tarjetas_sin_jefe,
        'supervisores_asignados': supervisores_asignados,
        'modo_gerencia_directa': False,
        'trabajadores': trabajadores,
        'objetivo_label_plural': objetivo_label_plural,
        'areas_section_title': areas_section_title,
        'areas_section_desc': areas_section_desc,
        'hero_badge_label': 'RESPONSABILIDAD ACTIVA' if es_responsable else 'SUPERVISION ACTIVA',
        'kpi_promedio_label': 'Promedio de Objetivos' if es_responsable else 'Promedio del Área',
        'kpi_mejor_label': 'Objetivo Destacado Mensual' if es_responsable else 'Mejor Desempeño Mensual',
        'estado_objetivos_title': 'Estado de Objetivos' if es_responsable else 'Estado del Equipo',
        'empty_objetivos_message': 'No hay objetivos activos en esta área.' if es_responsable else 'No hay trabajadores asignados a esta área.',
        'promedio_area': promedio_area_total,
        'total_evaluaciones': total_evaluaciones_mes,
        'mejor_trabajador': mejor_trabajador,
        'radar_data': radar_data, 
    })


@login_required
def panel_responsable(request):
    try:
        perfil = request.user.trabajador
    except Exception:
        return render(request, 'metricas_ceneris/error.html', {'mensaje': 'No tienes perfil asignado.'})

    if not _es_responsable(perfil):
        return redirect('metricas_ceneris:inicio_metricas')

    return panel_jefe(request)


def _meses_disponibles_evaluacion(hoy):
    meses_nombres = [
        '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    return [
        {'numero': mes, 'nombre': meses_nombres[mes]}
        for mes in range(1, hoy.month + 1)
    ]


def _resolver_mes_objetivo_request(request, hoy):
    anio_objetivo = hoy.year
    mes_origen = request.POST.get('mes_objetivo') if request.method == 'POST' else request.GET.get('mes_objetivo')

    try:
        mes_objetivo = int(mes_origen) if mes_origen else hoy.month
    except (TypeError, ValueError):
        mes_objetivo = hoy.month

    if mes_objetivo < 1:
        mes_objetivo = 1
    if mes_objetivo > hoy.month:
        mes_objetivo = hoy.month

    return anio_objetivo, mes_objetivo


def _crear_puntajes_evaluacion_desde_form(evaluacion, form):
    total_puntos = 0
    cantidad_items = 0

    for categoria, lista_inds in ESTRUCTURA_EVALUACION.items():
        for clave, nombre, _ in lista_inds:
            campo = f"ind_{clave}"
            if campo in form.cleaned_data:
                valor = int(form.cleaned_data[campo])
                Puntaje.objects.create(
                    evaluacion=evaluacion,
                    categoria=categoria,
                    indicador_nombre=nombre,
                    nota=valor,
                )
                total_puntos += valor
                cantidad_items += 1

    evaluacion.promedio_final = round(total_puntos / cantidad_items, 1) if cantidad_items > 0 else 0
    evaluacion.save()


def _puntajes_iniciales_para_evaluacion(evaluacion):
    mapa_nombre_a_campo = {}
    for _categoria, indicadores in ESTRUCTURA_EVALUACION.items():
        for clave, nombre, _desc in indicadores:
            mapa_nombre_a_campo[nombre] = f"ind_{clave}"

    iniciales = {}
    for puntaje in evaluacion.puntajes.all():
        campo = mapa_nombre_a_campo.get(puntaje.indicador_nombre)
        if campo:
            iniciales[campo] = str(puntaje.nota)

    return iniciales

@login_required
def evaluar_trabajador(request, trabajador_id, tipo):
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    evaluador_actual = getattr(request.user, 'trabajador', None)

    if not _puede_evaluar_objetivo(evaluador_actual, trabajador):
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes permisos para evaluar a este trabajador con la jerarquía actual.'
        })
    
    # Validamos que el tipo sea correcto, si escriben otra cosa por error, será MENSUAL
    tipo_upper = tipo.upper()
    if tipo_upper not in ['MENSUAL', 'SEMESTRAL', 'NUEVO']:
        tipo_upper = 'MENSUAL'

    hoy = timezone.now().date()
    anio_objetivo, mes_objetivo = _resolver_mes_objetivo_request(request, hoy)
    meses_disponibles_evaluacion = _meses_disponibles_evaluacion(hoy)

    evaluaciones_mensuales_anio = set(
        trabajador.evaluaciones.filter(
            tipo='MENSUAL',
            fecha_evaluacion__year=anio_objetivo,
        ).values_list('fecha_evaluacion__month', flat=True)
    )

    for item in meses_disponibles_evaluacion:
        item['ya_evaluado'] = item['numero'] in evaluaciones_mensuales_anio

    mes_objetivo_nombre = next(
        (item['nombre'] for item in meses_disponibles_evaluacion if item['numero'] == mes_objetivo),
        hoy.strftime('%B'),
    )

    if request.method == 'POST':
        form = EvaluacionForm(request.POST)
        if form.is_valid():
            existe_mismo_tipo_mes = EvaluacionMensual.objects.filter(
                trabajador=trabajador,
                tipo=tipo_upper,
                fecha_evaluacion__year=anio_objetivo,
                fecha_evaluacion__month=mes_objetivo,
            ).exists()

            if existe_mismo_tipo_mes:
                messages.warning(request, 'Ya existe una evaluación de este tipo para el mes seleccionado. Si hubo un error, corrígela desde el historial.')
                return render(request, 'metricas_ceneris/evaluaciones/evaluar.html', {
                    'trabajador': trabajador,
                    'form': form,
                    'estructura': ESTRUCTURA_EVALUACION,
                    'tipo_evaluacion': tipo_upper,
                    'modo_correccion': False,
                    'anio_objetivo': anio_objetivo,
                    'mes_objetivo': mes_objetivo,
                    'mes_objetivo_nombre': mes_objetivo_nombre,
                    'meses_disponibles_evaluacion': meses_disponibles_evaluacion,
                })

            with transaction.atomic():
                evaluacion = form.save(commit=False)
                evaluacion.trabajador = trabajador
                evaluacion.tipo = tipo_upper
                evaluacion.fecha_evaluacion = datetime.date(anio_objetivo, mes_objetivo, 1)
                if evaluador_actual:
                    evaluacion.evaluador = evaluador_actual
                    evaluacion.cargo_evaluador = evaluador_actual.cargo or ''
                _guardar_evaluacion_con_reintento_secuencia(evaluacion)

                _crear_puntajes_evaluacion_desde_form(evaluacion, form)

            return redirect('metricas_ceneris:historial_trabajador', trabajador_id=trabajador.id)
    else:
        form = EvaluacionForm()

    return render(request, 'metricas_ceneris/evaluaciones/evaluar.html', {
        'trabajador': trabajador,
        'form': form,
        'estructura': ESTRUCTURA_EVALUACION,
        'tipo_evaluacion': tipo_upper,
        'modo_correccion': False,
        'anio_objetivo': anio_objetivo,
        'mes_objetivo': mes_objetivo,
        'mes_objetivo_nombre': mes_objetivo_nombre,
        'meses_disponibles_evaluacion': meses_disponibles_evaluacion,
    })


@login_required
def corregir_evaluacion(request, evaluacion_id):
    evaluacion = get_object_or_404(
        EvaluacionMensual.objects.select_related('trabajador', 'evaluador', 'trabajador__area').prefetch_related('puntajes'),
        id=evaluacion_id,
    )
    evaluador_actual = getattr(request.user, 'trabajador', None)

    if not evaluador_actual:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes perfil asignado para corregir evaluaciones.'
        })

    puede_corregir = (
        evaluacion.evaluador_id == evaluador_actual.id
        or _puede_evaluar_objetivo(evaluador_actual, evaluacion.trabajador)
    )
    if not puede_corregir:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes permisos para corregir esta evaluación.'
        })

    puntajes_iniciales = _puntajes_iniciales_para_evaluacion(evaluacion)

    if request.method == 'POST':
        form = EvaluacionForm(request.POST, instance=evaluacion, puntajes_iniciales=puntajes_iniciales)
        if form.is_valid():
            with transaction.atomic():
                evaluacion_actualizada = form.save(commit=False)
                evaluacion_actualizada.evaluador = evaluador_actual
                evaluacion_actualizada.cargo_evaluador = evaluador_actual.cargo or evaluacion_actualizada.cargo_evaluador
                evaluacion_actualizada.save()

                evaluacion_actualizada.puntajes.all().delete()
                _crear_puntajes_evaluacion_desde_form(evaluacion_actualizada, form)

            messages.success(request, 'Evaluación corregida correctamente.')
            return redirect('metricas_ceneris:historial_trabajador', trabajador_id=evaluacion.trabajador_id)
    else:
        form = EvaluacionForm(instance=evaluacion, puntajes_iniciales=puntajes_iniciales)

    return render(request, 'metricas_ceneris/evaluaciones/evaluar.html', {
        'trabajador': evaluacion.trabajador,
        'form': form,
        'estructura': ESTRUCTURA_EVALUACION,
        'tipo_evaluacion': evaluacion.tipo,
        'modo_correccion': True,
        'evaluacion_objetivo': evaluacion,
        'anio_objetivo': evaluacion.fecha_evaluacion.year,
        'mes_objetivo': evaluacion.fecha_evaluacion.month,
        'mes_objetivo_nombre': evaluacion.fecha_evaluacion.strftime('%B'),
        'meses_disponibles_evaluacion': [],
    })

@login_required
def historial_trabajador(request, trabajador_id):
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    evaluaciones = list(trabajador.evaluaciones.select_related('evaluador').all())

    # Para cada mes evaluado, calculamos el SCORE ponderado de los 4 aspectos
    # (no solo el Desempeño de la evaluación) y lo adjuntamos a cada registro.
    hoy = timezone.now().date()
    for e in evaluaciones:
        d = _calcular_score_total(
            trabajador, hoy, 'mes', anio=e.fecha_evaluacion.year, mes=e.fecha_evaluacion.month)
        e.score_periodo = round(d['score_total'], 2)
        e.nota_desempeno_periodo = round(d['nota_desempeno_periodo'], 2)
        e.nota_disciplinaria = round(d['nota_disciplinaria'], 2)
        e.nota_conocimiento = round(d['nota_conocimiento'], 2)
        e.nota_asistencia = round(d['nota_asistencia'], 2)

    # Datos para gráfica (histórico del score de 4 aspectos)
    fechas = [e.fecha_evaluacion.strftime("%b %Y") for e in evaluaciones]
    notas = [e.score_periodo for e in evaluaciones]

    return render(request, 'metricas_ceneris/historial.html', {
        'trabajador': trabajador,
        'evaluaciones': evaluaciones,
        'fechas_grafica': fechas,
        'notas_grafica': notas
    })

#========================
# GESTION DE USUARIOS
#========================

def tiene_acceso_al_portal(user):
    grupos = user.groups.values_list('name', flat=True)
    es_admin = 'Administrador' in grupos
    # Puede entrar si es Admin Y (tiene metricas O tiene rrhh)
    return es_admin and ('Metricas' in grupos or 'Recursos Humanos' in grupos)

@login_required
def mis_evaluaciones(request):
    try:
        # Intentamos obtener el perfil del trabajador logueado
        trabajador = request.user.trabajador
    except ObjectDoesNotExist:
        # Si entra un usuario que no es trabajador (ej. admin puro)
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes un perfil de trabajador asignado para ver métricas.'
        })

    # Obtenemos las evaluaciones del trabajador
    evaluaciones = trabajador.evaluaciones.all().order_by('-fecha_evaluacion')
    
    # Calculamos el promedio de las evaluaciones
    promedio_total = 0
    if evaluaciones.exists():
        promedio_total = sum([e.promedio_final for e in evaluaciones]) / evaluaciones.count()
    
    context = {
        'trabajador': trabajador,
        'evaluaciones': evaluaciones,
        'promedio_total': round(promedio_total, 1)
    }
    return render(request, 'metricas_ceneris/evaluaciones/mis_evaluaciones.html', context)

def detalle_evaluacion(request, evaluacion_id):
    evaluacion = get_object_or_404(
        EvaluacionMensual.objects.select_related('evaluador', 'trabajador', 'trabajador__area'),
        id=evaluacion_id,
    )
    puntajes = evaluacion.puntajes.all()

    if evaluacion.evaluador:
        nombre_evaluador = evaluacion.evaluador.nombre_completo
        cargo_evaluador = evaluacion.cargo_evaluador or evaluacion.evaluador.cargo or 'Sin cargo registrado'
    else:
        nombre_evaluador = 'Sin evaluador registrado'
        cargo_evaluador = evaluacion.cargo_evaluador or 'Sin cargo registrado'
    
    # Agrupamos los puntajes por categoría para mostrarlos ordenados
    grupos = {
        'OPERACIONAL': puntajes.filter(categoria='OPERACIONAL'),
        'ADMINISTRATIVO': puntajes.filter(categoria='ADMINISTRATIVO'),
        'HABILIDADES': puntajes.filter(categoria='HABILIDADES'),
    }

    return render(request, 'metricas_ceneris/detalle.html', {
        'evaluacion': evaluacion,
        'grupos': grupos,
        'nombre_evaluador': nombre_evaluador,
        'cargo_evaluador': cargo_evaluador,
    })

@login_required
def dashboard_trabajador(request):
    try:
        trabajador = request.user.trabajador
    except:
        return render(request, 'metricas_ceneris/error.html', {'mensaje': 'No tienes perfil asignado.'})

    # Obtener todas sus evaluaciones. Adjuntamos a cada una el SCORE ponderado
    # (4 aspectos) del mes evaluado, para mostrarlo en el Historial.
    evaluaciones = list(trabajador.evaluaciones.select_related('evaluador').all().order_by('-fecha_evaluacion'))
    _hoy_hist = timezone.now().date()
    for _ev in evaluaciones:
        _d = _calcular_score_total(
            trabajador, _hoy_hist, 'mes', anio=_ev.fecha_evaluacion.year, mes=_ev.fecha_evaluacion.month)
        _ev.score_periodo = round(_d['score_total'], 2)
        _ev.nota_desempeno_periodo = round(_d['nota_desempeno_periodo'], 2)
        _ev.nota_disciplinaria = round(_d['nota_disciplinaria'], 2)
        _ev.nota_conocimiento = round(_d['nota_conocimiento'], 2)
        _ev.nota_asistencia = round(_d['nota_asistencia'], 2)

    # 1. KPI: Última Nota y Estado
    ultima_ev = evaluaciones[0] if evaluaciones else None
    ultima_nota = ultima_ev.promedio_final if ultima_ev else 0
    
    # Clasificación (AD, A, B, C)
    clasificacion = "Pendiente"
    color_clase = "gray"
    if ultima_nota > 0:
        if ultima_nota >= 9:
            clasificacion = "AD - Muy Alto"
            color_clase = "indigo"
        elif ultima_nota >= 7:
            clasificacion = "A - Alto"
            color_clase = "green"
        elif ultima_nota >= 5:
            clasificacion = "B - Medio"
            color_clase = "yellow"
        else:
            clasificacion = "C - Bajo"
            color_clase = "red"

    # 2. Datos para Gráfico de Línea (Histórico del SCORE ponderado).
    # Se calcula EN VIVO para los últimos 6 meses (sin tabla ni cron).
    MESES_ABREV = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                   'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    hoy_ref = timezone.now().date()
    historial_score = []
    cursor_mes = hoy_ref.replace(day=1)
    for _ in range(6):
        d_mes = _calcular_score_total(
            trabajador, hoy_ref, 'mes', anio=cursor_mes.year, mes=cursor_mes.month)
        historial_score.append(
            (f"{MESES_ABREV[cursor_mes.month]} {cursor_mes.year}", round(d_mes['score_total'], 2)))
        # Retroceder al primer día del mes anterior.
        cursor_mes = (cursor_mes - datetime.timedelta(days=1)).replace(day=1)
    historial_score.reverse()  # cronológico
    fechas_grafica = [h[0] for h in historial_score]
    notas_grafica = [h[1] for h in historial_score]

    # 3. Datos para Gráfico de Radar (Sus fortalezas actuales)
    radar_data = []
    if ultima_ev:
        for cat in ['OPERACIONAL', 'ADMINISTRATIVO', 'HABILIDADES']:
            avg = ultima_ev.puntajes.filter(categoria=cat).aggregate(Avg('nota'))['nota__avg']
            radar_data.append(round(avg, 1) if avg else 0)
    else:
        radar_data = [0, 0, 0]

    # -------------------------------------------------------------
    # NUEVO: LÓGICA DE FECHAS DINÁMICA (Permite ver cualquier mes)
    # -------------------------------------------------------------
    hoy = timezone.now().date()
    
    mes_param = request.GET.get('mes')
    anio_param = request.GET.get('anio')
    
    if mes_param and anio_param:
        try:
            mes = int(mes_param)
            anio = int(anio_param)
            # Obtenemos el último día de ese mes específico
            ultimo_dia = calendar.monthrange(anio, mes)[1]
            
            inicio_mes = datetime.date(anio, mes, 1)
            fin_mes = datetime.date(anio, mes, ultimo_dia)
        except ValueError:
            # Si mandan texto inválido, volvemos al mes actual
            mes = hoy.month
            anio = hoy.year
            ultimo_dia = calendar.monthrange(anio, mes)[1]
            inicio_mes = hoy.replace(day=1)
            fin_mes = datetime.date(anio, mes, ultimo_dia)
    else:
        # Comportamiento por defecto: Mes actual completo
        mes = hoy.month
        anio = hoy.year
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        inicio_mes = hoy.replace(day=1)
        fin_mes = datetime.date(anio, mes, ultimo_dia)
    # -------------------------------------------------------------
    _, num_dias = calendar.monthrange(anio, mes)

    # Obtenemos los registros de este trabajador con el NUEVO filtro de fechas (fin_mes)
    tareos_qs = TareoDiario.objects.filter(
        trabajador=trabajador,
        fecha__gte=inicio_mes,
        fecha__lte=fin_mes
    ).order_by('-fecha')

    tareos = list(tareos_qs)
    tareos_por_fecha = {tareo.fecha: tareo for tareo in tareos}

    # Cálculos de ASISTENCIA del mes seleccionado
    dias_laborables = 0
    faltas = 0
    asistencias = 0
    # Obtener feriados del año seleccionado
    feriados = _obtener_feriados(anio)

    registros_diarios = []
    for dia in range(1, num_dias + 1):
        fecha_actual = datetime.date(anio, mes, dia)
        es_futuro = fecha_actual > hoy
        es_descanso = fecha_actual.weekday() == 6
        es_feriado = fecha_actual in feriados 
        tareo = tareos_por_fecha.get(fecha_actual)

        if tareo:
            # ¡CORRECCIÓN CLAVE 1!
            # Solo sumamos a los KPIs de asistencia si el día YA PASÓ o es HOY
            if not es_futuro and not es_descanso and not es_feriado:
                dias_laborables += 1
                if tareo.resultado == 'F':
                    faltas += 1
                else:
                    asistencias += 1

            # Cálculo de horas trabajadas
            horas_trabajadas = None
            if tareo.hora_entrada_real and tareo.hora_salida_real:
                entrada_dt = datetime.datetime.combine(fecha_actual, tareo.hora_entrada_real)
                salida_dt = datetime.datetime.combine(fecha_actual, tareo.hora_salida_real)
                if salida_dt < entrada_dt: 
                    salida_dt += datetime.timedelta(days=1)
                diferencia_segundos = (salida_dt - entrada_dt).total_seconds()
                horas_trabajadas = round(diferencia_segundos / 3600, 2)

            registros_diarios.append({
                'fecha': fecha_actual,
                'tareo': tareo,
                'es_descanso': es_descanso,
                # ¡CORRECCIÓN CLAVE 2! Le pasamos la variable real, no "False" por defecto
                'es_futuro': es_futuro, 
                'es_falta_sintetica': False,
                'horas_trabajadas': horas_trabajadas,
            })
            
        elif es_futuro:
            registros_diarios.append({
                'fecha': fecha_actual,
                'tareo': None,
                'es_descanso': False,
                'es_futuro': True,
                'es_falta_sintetica': False,
            })
        elif es_feriado:
            # ¡LA SOLUCIÓN AQUÍ! Es feriado y no vino. No pasa nada, no es falta.
            registros_diarios.append({
                'fecha': fecha_actual,
                'tareo': None,
                'es_descanso': False,
                'es_feriado': True,
                'es_futuro': False,
                'es_falta_sintetica': False,
            })
        elif not es_descanso:
            dias_laborables += 1
            faltas += 1
            registros_diarios.append({
                'fecha': fecha_actual,
                'tareo': None,
                'es_descanso': False,
                'es_futuro': False,
                'es_falta_sintetica': True,
            })
        else:
            registros_diarios.append({
                'fecha': fecha_actual,
                'tareo': None,
                'es_descanso': True,
                'es_futuro': False,
                'es_falta_sintetica': False,
            })

    dias_tarde = 0
    horas_tardanza_total = 0.0

    for item in registros_diarios:
        tareo = item['tareo']
        if not tareo:
            continue

        minutos_tardanza_ajustada = 0

        if tareo.hora_entrada_real:
            hora_referencia = _hora_referencia_entrada_por_fecha(tareo.fecha)
            segundos_referencia = (hora_referencia.hour * 3600) + (hora_referencia.minute * 60) + hora_referencia.second
            segundos_entrada = (tareo.hora_entrada_real.hour * 3600) + (tareo.hora_entrada_real.minute * 60) + tareo.hora_entrada_real.second
            segundos_tarde = max(0, segundos_entrada - segundos_referencia)
            minutos_tardanza_ajustada = int(segundos_tarde // 60)

        tareo.minutos_tardanza_ajustada = minutos_tardanza_ajustada
        tareo.horas_tardanza_ajustada = minutos_tardanza_ajustada / 60

        if minutos_tardanza_ajustada > 0:
            dias_tarde += 1
            horas_tardanza_total += tareo.horas_tardanza_ajustada

    # Porcentajes
    # KPI de asistencia visual: solo descuenta faltas
    porc_asistencia = 0
    if dias_laborables > 0:
        porc_asistencia = round((asistencias / dias_laborables) * 100, 1)

    # Nota de asistencia para score: aplica sanción con la lógica unificada del semestre.
    score_data = _calcular_score_total(trabajador, hoy, 'mes', anio=anio, mes=mes)
    nota_asistencia_pct = round(score_data['asistencia_pct'], 1)
    porc_puntualidad = 100 # Asumimos 100% si no hay datos aún
    
    if asistencias > 0:
        # Puntualidad = 100% - % de veces que llegaste tarde
        tasa_tardanza = (dias_tarde / asistencias) * 100
        porc_puntualidad = round(100 - tasa_tardanza, 1)

    # --------- CÁLCULO DE NOTA DE DESEMPEÑO CON REGLA DE SEMESTRES ---------
    nota_desempeno_mes = round(score_data['nota_desempeno_periodo'], 1)
    
    # Nota de asistencia (fórmula PPT) que alimenta el score; el % real se
    # muestra aparte como kpi_asistencia.
    nota_asistencia = round(score_data['nota_asistencia'], 1)
    
    # --------- CÁLCULO DEL SCORE FINAL ---------
    # Score ponderado por modalidad (oficina 3 aspectos / mina 4). Ver
    # _ponderar_score / PESOS_EVALUACION.
    score_final = round(score_data['score_total'], 2)

    # Aspectos a mostrar en el desglose, según la modalidad del trabajador.
    _labels_aspecto = {
        'DESEMPENO': 'Desempeño',
        'MEDIDAS_DISCIPLINARIAS': 'Medidas disciplinarias',
        'CONOCIMIENTO': 'Conocimiento',
        'ASISTENCIA': 'Asistencia',
    }
    _notas_aspecto = {
        'DESEMPENO': nota_desempeno_mes,
        'MEDIDAS_DISCIPLINARIAS': round(score_data['nota_disciplinaria'], 2),
        'CONOCIMIENTO': round(score_data['nota_conocimiento'], 2) if score_data['nota_conocimiento'] is not None else None,
        'ASISTENCIA': nota_asistencia,
    }
    aspectos_score = [
        {
            'clave': clave,
            'label': _labels_aspecto[clave],
            'peso': round(peso * 100),
            'nota': _notas_aspecto[clave],
        }
        for clave, peso in score_data['pesos_aplicados'].items()
    ]

    return render(request, 'metricas_ceneris/trabajador/dashboard_trabajador.html', {
        'aspectos_score': aspectos_score,
        'modalidad': score_data['modalidad'],
        'conocimiento_pendiente': score_data['conocimiento_pendiente'],
        'trabajador': trabajador,
        'evaluaciones': evaluaciones,
        'ultima_nota': ultima_nota,
        'clasificacion': clasificacion,
        'color_clase': color_clase,
        'fechas_grafica': fechas_grafica,
        'notas_grafica': notas_grafica,
        'radar_data': radar_data,
        'tareos': tareos,
        'registros_diarios': registros_diarios,
        'kpi_asistencia': porc_asistencia,
        'kpi_puntualidad': porc_puntualidad,
        'kpi_faltas': faltas,
        'kpi_horas_tardanza': round(horas_tardanza_total, 2),
        'kpi_nota_asistencia': nota_asistencia,
        # Nuevas variables para el mes seleccionado
        'nota_desempeno_mes': nota_desempeno_mes,
        'nota_asistencia': nota_asistencia,
        'score_final': score_final,
        # Mandamos el mes y año para que el HTML sepa qué estamos viendo
        'mes_seleccionado': mes,
        'anio_seleccionado': anio,
    })

@login_required
def inicio_inteligente(request):
    try:
        perfil = request.user.trabajador
        
        # --- DIAGNÓSTICO EN TERMINAL ---
        print(f"--- LOGIN: {perfil.nombres} ---")
        print(f"1. Cargo jerarquico: {_cargo_jerarquico_actual(perfil)}")
        # -------------------------------

        # PRIORIDAD 1: GERENTE (solo gerente puro)
        if _es_gerente_puro(perfil):
            print(">>> Redirigiendo a Panel GERENTE")
            return redirect('metricas_ceneris:dashboard_gerente')

        # PRIORIDAD 2: RESPONSABLE
        if _es_responsable(perfil):
            print(">>> Redirigiendo a Dashboard de RESPONSABLE")
            return redirect('metricas_ceneris:dashboard_responsable')

        # PRIORIDAD 3: SUPERVISOR
        if _es_supervisor(perfil):
            print(">>> Redirigiendo a Dashboard de EQUIPO")
            return redirect('metricas_ceneris:dashboard_jefe')

        # PRIORIDAD 4: TRABAJADOR (Si no es nada de lo anterior)
        print(">>> Redirigiendo a Panel TRABAJADOR")
        return redirect('metricas_ceneris:dashboard_trabajador')
    except Exception as e:
        print(f"ERROR CRÍTICO EN INICIO: {e}")
        if request.user.is_superuser:
            return redirect('/admin/')
        return render(request, 'metricas_ceneris/error.html', {'mensaje': 'Usuario sin perfil de trabajador.'})


def _resolver_filtro_mes_anio(request, hoy):
    mes_param = request.GET.get('mes')
    anio_param = request.GET.get('anio')

    try:
        mes = int(mes_param) if mes_param else hoy.month
        anio = int(anio_param) if anio_param else hoy.year
    except (TypeError, ValueError):
        mes = hoy.month
        anio = hoy.year

    if mes < 1 or mes > 12:
        mes = hoy.month

    ultimo_dia = calendar.monthrange(anio, mes)[1]
    inicio_mes = datetime.date(anio, mes, 1)
    fin_mes = datetime.date(anio, mes, ultimo_dia)

    return mes, anio, inicio_mes, fin_mes


@login_required
def panel_conocimiento(request):
    """Carga de notas de Conocimiento (aspecto 03).

    Accesible por jefaturas (supervisor/responsable/gerente). Cada jefe solo ve a
    quienes le corresponde evaluar según la jerarquía (`_puede_evaluar_objetivo`):
    el supervisor a los trabajadores de su área; el gerente a los supervisores.
    """
    perfil = getattr(request.user, 'trabajador', None)
    if not perfil or not _es_lider(perfil):
        return redirect('metricas_ceneris:inicio_metricas')

    hoy = timezone.now().date()
    try:
        mes = int(request.POST.get('mes') or request.GET.get('mes') or hoy.month)
        anio = int(request.POST.get('anio') or request.GET.get('anio') or hoy.year)
    except (TypeError, ValueError):
        mes, anio = hoy.month, hoy.year
    if not (1 <= mes <= 12):
        mes = hoy.month
    fecha_periodo = datetime.date(anio, mes, 1)

    # Alcance = a quiénes puede evaluar este jefe (misma jerarquía que Desempeño).
    candidatos = (Trabajador.objects.filter(activo=True)
                  .exclude(id=perfil.id)
                  .select_related('area'))
    trabajadores = sorted(
        (t for t in candidatos if _puede_evaluar_objetivo(perfil, t)),
        key=lambda t: ((t.apellido_paterno or ''), (t.nombres or '')),
    )

    if request.method == 'POST':
        guardadas = 0
        for t in trabajadores:
            raw = (request.POST.get(f'nota_{t.id}') or '').strip()
            if not raw:
                continue
            try:
                val = float(raw)
            except ValueError:
                continue
            # La nota del examen va de 1 a 20 (se convierte ÷2 para el score).
            if 1 <= val <= 20:
                NotaConocimiento.objects.update_or_create(
                    trabajador=t, fecha=fecha_periodo,
                    defaults={'nota': val, 'comentario': 'Carga desde panel de Conocimiento'})
                guardadas += 1
        messages.success(request, f'Se guardaron {guardadas} nota(s) de Conocimiento para {mes:02d}/{anio}.')
        return redirect(f"/metricas_ceneris/conocimiento/?mes={mes}&anio={anio}")

    notas = {
        nc.trabajador_id: nc.nota
        for nc in NotaConocimiento.objects.filter(fecha=fecha_periodo, trabajador__in=trabajadores)
    }
    # El examen inicia en 20/20 cuando aún no se cargó nota en el periodo.
    filas = [{'trabajador': t, 'nota': notas.get(t.id, CONOCIMIENTO_EXAMEN_INICIAL)} for t in trabajadores]

    return render(request, 'metricas_ceneris/conocimiento/panel_conocimiento.html', {
        'filas': filas,
        'mes_seleccionado': mes,
        'anio_seleccionado': anio,
        'meses_disponibles': _meses_disponibles_evaluacion(hoy),
        'anios_disponibles': [hoy.year - 1, hoy.year],
        'total_mina': len(filas),
    })


def _construir_resumen_asistencia(trabajadores, inicio_mes, fin_mes, incluir_en_detalle=None):
    if incluir_en_detalle is None:
        incluir_en_detalle = lambda *_: True

    datos = []
    acumulado_asistencia = 0
    acumulado_nota_asistencia = 0
    total_faltas = 0
    total_horas_tardanza = 0
    total_personas = 0

    for trabajador in trabajadores:
        # Obtenemos los tareos para el cálculo detallado de horas abajo
        tareos = TareoDiario.objects.filter(
            trabajador=trabajador,
            fecha__gte=inicio_mes,
            fecha__lte=fin_mes
        )
        tareos_por_fecha = {t.fecha: t for t in tareos}

        # ==============================================================
        # 1. LLAMAMOS A LA FUNCIÓN BLINDADA PRIMERO
        # ==============================================================
        asistencia_data = _calcular_asistencia_por_periodo(trabajador, inicio_mes, fin_mes)
        
        # 2. EXTRAEMOS LOS DATOS MATEMÁTICAMENTE EXACTOS
        dias_totales = asistencia_data['total_dias']
        faltas = asistencia_data['faltas'] # ¡Ahora sí incluye las faltas por días vacíos!
        nota_asistencia = round(asistencia_data['nota'], 1)  # nota PPT (consistente con el score)
        
        # 3. Calculamos las asistencias reales restando las faltas exactas
        asistencias = dias_totales - faltas
        # ==============================================================

        tardanzas_detalle = []
        for tareo in tareos.exclude(hora_entrada_real__isnull=True).order_by('fecha'):
            hora_referencia = _hora_referencia_entrada_por_fecha(tareo.fecha)

            segundos_referencia = (
                (hora_referencia.hour * 3600)
                + (hora_referencia.minute * 60)
                + hora_referencia.second
            )
            segundos_entrada = (
                (tareo.hora_entrada_real.hour * 3600)
                + (tareo.hora_entrada_real.minute * 60)
                + tareo.hora_entrada_real.second
            )
            segundos_tarde = max(0, segundos_entrada - segundos_referencia)

            if segundos_tarde < (TARDANZA_MINIMA_MINUTOS * 60):
                continue

            tardanzas_detalle.append({
                'fecha': tareo.fecha,
                'hora_entrada_real': tareo.hora_entrada_real,
                'horas_tardanza': round(segundos_tarde / 3600, 2),
            })

        dias_tarde = len(tardanzas_detalle)

        asistencias_qs = tareos.exclude(resultado='F').order_by('fecha')
        asistencias_detalle = [
            {
                'fecha': tareo_asistencia.fecha,
                'hora_entrada_real': tareo_asistencia.hora_entrada_real,
                'resultado': tareo_asistencia.resultado,
            }
            for tareo_asistencia in asistencias_qs
        ]

        # Detalle de faltas (registradas y sinteticas) para modal de tabla.
        faltas_detalle = []
        fecha_cursor = inicio_mes
        feriados_por_anio = {}
        hoy = timezone.now().date()

        while fecha_cursor <= fin_mes:
            if fecha_cursor.year not in feriados_por_anio:
                feriados_por_anio[fecha_cursor.year] = set(_obtener_feriados(fecha_cursor.year))

            es_futuro = fecha_cursor > hoy
            es_descanso = fecha_cursor.weekday() == 6
            es_feriado = fecha_cursor in feriados_por_anio[fecha_cursor.year]

            if not es_futuro and not es_descanso and not es_feriado:
                tareo_dia = tareos_por_fecha.get(fecha_cursor)
                if tareo_dia and tareo_dia.resultado == 'F':
                    faltas_detalle.append({
                        'fecha': fecha_cursor,
                        'tipo': 'REGISTRADA',
                    })
                elif not tareo_dia:
                    faltas_detalle.append({
                        'fecha': fecha_cursor,
                        'tipo': 'SINTETICA',
                    })

            fecha_cursor += datetime.timedelta(days=1)

        horas_tardanza = sum(t['horas_tardanza'] for t in tardanzas_detalle)

        # Calculamos los porcentajes visuales para el frontend
        porc_asistencia = 0
        porc_tardanza = 0

        if dias_totales > 0:
            porc_asistencia = round((asistencias / dias_totales) * 100, 1)
        if asistencias > 0:
            porc_tardanza = round((dias_tarde / asistencias) * 100, 1)

        total_personas += 1
        acumulado_asistencia += porc_asistencia
        acumulado_nota_asistencia += nota_asistencia
        
        # Sumamos las faltas reales (BD + Sintéticas) al KPI Global del Área
        total_faltas += faltas 
        total_horas_tardanza += horas_tardanza

        if not incluir_en_detalle(trabajador, nota_asistencia):
            continue

        justificaciones = Justificacion.objects.filter(
            tareo__trabajador=trabajador,
            tareo__fecha__gte=inicio_mes,
            tareo__fecha__lte=fin_mes
        ).order_by('-tareo__fecha')

        datos.append({
            'trabajador': trabajador,
            'dias_totales': dias_totales,
            'asistencias': asistencias,
            'faltas': faltas,
            'faltas_detalle': faltas_detalle,
            'dias_tarde': dias_tarde,
            'tardanzas_detalle': tardanzas_detalle,
            'asistencias_detalle': asistencias_detalle,
            'horas_tardanza': horas_tardanza,
            'porc_asistencia': porc_asistencia,
            'porc_tardanza': porc_tardanza,
            'nota_asistencia': nota_asistencia,
            'cant_justificadas': justificaciones.count(),
            'lista_justificaciones': justificaciones,
        })

    promedio_asistencia = 0
    promedio_nota_asistencia = 0
    if total_personas > 0:
        promedio_asistencia = round(acumulado_asistencia / total_personas, 1)
        promedio_nota_asistencia = round(acumulado_nota_asistencia / total_personas, 1)

    return {
        'datos': datos,
        'promedio_asistencia_area': promedio_asistencia,
        'promedio_nota_asistencia_area': promedio_nota_asistencia,
        'total_faltas_area': total_faltas,
        'total_horas_tardanza': round(total_horas_tardanza, 2),
    }

@login_required
def panel_asistencias(request):
    from .asistencia_views import panel_asistencias as _panel_asistencias
    return _panel_asistencias(request)

@login_required
def dashboard_gerente(request):
    """ Vista principal ultra-rápida. Solo calcula Podio y Gráficos. """
    try:
        usuario = request.user.trabajador
        if not usuario.es_gerente:
            return redirect('metricas_ceneris:inicio_metricas')
    except:
        return redirect('metricas_ceneris:inicio_metricas')

    hoy = timezone.now().date()
    _, semestre_hoy = _anio_semestre_por_fecha(hoy)
    filtro_periodo = _resolver_filtro_periodo_request(request, hoy, periodo_default='semestre')
    periodo_tabla_defecto = filtro_periodo['periodo']
    anio_tabla_defecto = filtro_periodo['anio']
    mes_tabla_defecto = filtro_periodo['mes'] or hoy.month
    semestre_tabla_defecto = filtro_periodo['semestre'] or semestre_hoy

    parametros_periodo = {'anio': anio_tabla_defecto}
    if periodo_tabla_defecto == 'mes':
        parametros_periodo['mes'] = mes_tabla_defecto
    else:
        parametros_periodo['semestre'] = semestre_tabla_defecto

    sede_id = request.GET.get('sede', 'todas')
    todos_trabajadores = Trabajador.objects.filter(activo=True).select_related('area', 'sede')
    
    if sede_id != 'todas' and sede_id:
        try:
            todos_trabajadores = todos_trabajadores.filter(sede_id=int(sede_id))
        except ValueError:
            pass

    ranking_trabajadores_periodo = []
    ranking_jefes_periodo = []
    scores_por_area_periodo = {}
    fecha_ref_periodo = _fecha_referencia_periodo(hoy, periodo_tabla_defecto, **parametros_periodo)

    # Solo calculamos lo necesario para el Podio y el Gráfico
    for t in todos_trabajadores:
        datos_periodo = _calcular_score_total(t, hoy, periodo_tabla_defecto, **parametros_periodo)
        puntaje = datos_periodo['score_total']

        if t.area and puntaje > 0:
            if t.area.nombre not in scores_por_area_periodo:
                scores_por_area_periodo[t.area.nombre] = []
            scores_por_area_periodo[t.area.nombre].append(puntaje)
        
        # Guardamos datos mínimos para el podio
        if puntaje > 0 and not _esta_en_periodo_prueba(t, fecha_ref_periodo):
            info_podio = {
                'trabajador': t,
                'score': round(puntaje, 2),
                'eval_avg': round(datos_periodo['desempeno_compuesto'], 2),
                'asistencia_avg': round(datos_periodo['nota_asistencia'], 1),
                'area_nombre': t.area.nombre if t.area else "Sin Área"
            }
            if _es_trabajador_base(t):
                ranking_trabajadores_periodo.append(info_podio)
            elif _es_supervisor(t) or _es_responsable(t):
                ranking_jefes_periodo.append(info_podio)

    ranking_trabajadores_periodo.sort(key=lambda x: x['score'], reverse=True)
    ranking_jefes_periodo.sort(key=lambda x: x['score'], reverse=True)

    lista_areas_promedio = []
    for nombre_area, puntajes in scores_por_area_periodo.items():
        promedio = sum(puntajes) / len(puntajes)
        lista_areas_promedio.append({'nombre': nombre_area, 'promedio': round(promedio, 2)})

    lista_areas_promedio.sort(key=lambda x: x['promedio'], reverse=True)
    labels_areas = [item['nombre'] for item in lista_areas_promedio]
    data_areas = [item['promedio'] for item in lista_areas_promedio]

    sedes = Sede.objects.filter(activo=True).order_by('nombre')
    
    # Rango de años dinámico
    min_eval = EvaluacionMensual.objects.aggregate(valor=Min('fecha_evaluacion'))['valor']
    anio_inicio = min_eval.year if min_eval else max(hoy.year - 2, 2020)
    anios_disponibles = list(range(hoy.year, anio_inicio - 1, -1))
    meses_disponibles = [(i, nombre) for i, nombre in enumerate(['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'], 1)]

    return render(request, 'metricas_ceneris/gerente/dashboard_gerente.html', {
        'usuario': usuario,
        'anio': anio_tabla_defecto,
        'podio_trabajadores_mes': ranking_trabajadores_periodo[:3],
        'podio_jefes_mes': ranking_jefes_periodo[:3],
        'mejor_area_mes': lista_areas_promedio[0] if lista_areas_promedio else None,
        'labels_areas_json': json.dumps(labels_areas, ensure_ascii=False),
        'data_areas_json': json.dumps(data_areas),
        'sedes': sedes,
        'sede_seleccionada': sede_id,
        'anios_disponibles': anios_disponibles,
        'meses_disponibles': meses_disponibles,
        'periodo_tabla_defecto': periodo_tabla_defecto,
        'anio_tabla_defecto': anio_tabla_defecto,
        'mes_tabla_defecto': mes_tabla_defecto,
        'semestre_tabla_defecto': semestre_tabla_defecto,
    })


@login_required
def ranking_general(request):
    """ Nueva vista dedicada exclusivamente al Ranking General como herramienta de consulta """
    try:
        usuario = request.user.trabajador
        if not usuario.es_gerente:
            return redirect('metricas_ceneris:inicio_metricas')
    except:
        return redirect('metricas_ceneris:inicio_metricas')

    hoy = timezone.now().date()
    sedes = Sede.objects.filter(activo=True).order_by('nombre')
    
    min_eval = EvaluacionMensual.objects.aggregate(valor=Min('fecha_evaluacion'))['valor']
    anio_inicio = min_eval.year if min_eval else max(hoy.year - 2, 2020)
    anios_disponibles = list(range(hoy.year, anio_inicio - 1, -1))
    meses_disponibles = [(i, nombre) for i, nombre in enumerate(['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'], 1)]

    return render(request, 'metricas_ceneris/ranking_general/ranking_general.html', {
        'sedes': sedes,
        'anios_disponibles': anios_disponibles,
        'meses_disponibles': meses_disponibles,
        'anio_actual': hoy.year,
        'mes_actual': hoy.month
    })

@login_required
def panel_evaluacion_gerente(request):
    from .gerencia_views import panel_evaluacion_gerente as _panel_evaluacion_gerente
    return _panel_evaluacion_gerente(request)


@login_required
def panel_area_directa_gerencia(request, area_id):
    from .gerencia_views import panel_area_directa_gerencia as _panel_area_directa_gerencia
    return _panel_area_directa_gerencia(request, area_id)

@login_required
def ajax_datos_ranking(request):
    from .gerencia_views import ajax_datos_ranking as _ajax_datos_ranking
    return _ajax_datos_ranking(request)


@login_required
def ajax_datos_grafico_areas(request):
    from .gerencia_views import ajax_datos_grafico_areas as _ajax_datos_grafico_areas
    return _ajax_datos_grafico_areas(request)


@login_required
def ajax_datos_podio(request):
    from .gerencia_views import ajax_datos_podio as _ajax_datos_podio
    return _ajax_datos_podio(request)


@login_required
def ajax_datos_area_lider(request):
    from .gerencia_views import ajax_datos_area_lider as _ajax_datos_area_lider
    return _ajax_datos_area_lider(request)


# ==================== VISTAS DE EXPORTACIÓN ====================

def _usuario_puede_exportar_metricas(user):
    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    try:
        perfil = user.trabajador
        if _es_gerente_puro(perfil):
            return True
    except Exception:
        pass

    return user.groups.filter(name='Recursos Humanos').exists()


def _contexto_exportacion_metricas(request):
    hoy = timezone.now().date()
    _, semestre_hoy = _anio_semestre_por_fecha(hoy)
    filtro_periodo = _resolver_filtro_periodo_request(request, hoy, periodo_default='semestre')
    periodo_tabla_defecto = filtro_periodo['periodo']
    anio_tabla_defecto = filtro_periodo['anio']
    mes_tabla_defecto = filtro_periodo['mes'] or hoy.month
    semestre_tabla_defecto = filtro_periodo['semestre'] or semestre_hoy

    sedes = Sede.objects.all().order_by('nombre')
    areas = Trabajador.objects.filter(
        activo=True,
    ).exclude(
        area__isnull=True,
    ).exclude(
        area__nombre__istartswith='Gerencia General'
    ).values_list('area__id', 'area__nombre').distinct().order_by('area__nombre')

    anio_actual = hoy.year
    anios_disponibles = list(range(anio_actual - 2, anio_actual + 1))
    meses_disponibles = [
        (1, 'Enero'),
        (2, 'Febrero'),
        (3, 'Marzo'),
        (4, 'Abril'),
        (5, 'Mayo'),
        (6, 'Junio'),
        (7, 'Julio'),
        (8, 'Agosto'),
        (9, 'Septiembre'),
        (10, 'Octubre'),
        (11, 'Noviembre'),
        (12, 'Diciembre'),
    ]

    return {
        'sedes': sedes,
        'areas': areas,
        'anios_disponibles': anios_disponibles,
        'meses_disponibles': meses_disponibles,
        'periodo_tabla_defecto': periodo_tabla_defecto,
        'anio_tabla_defecto': anio_tabla_defecto,
        'mes_tabla_defecto': mes_tabla_defecto,
        'semestre_tabla_defecto': semestre_tabla_defecto,
    }

@login_required
def exportar_ranking_excel_area(request):
    """Exporta ranking de trabajadores del área del supervisor/responsable (solo hoja 1)"""
    try:
        perfil = request.user.trabajador
    except Exception:
        return HttpResponse('No autorizado', status=403)

    if not (_es_supervisor(perfil) or _es_responsable(perfil)):
        return HttpResponse('No autorizado', status=403)

    areas_asignadas = _areas_bajo_responsabilidad(perfil)
    if not areas_asignadas:
        return HttpResponse('No tienes áreas asignadas', status=403)

    areas_ids = [a.id for a in areas_asignadas]

    # Filtro de área específica (seguro: solo permite áreas del usuario)
    area_export = request.GET.get('area_export', 'todas')
    if area_export != 'todas':
        try:
            area_id_filtro = int(area_export)
            if area_id_filtro in areas_ids:
                areas_ids = [area_id_filtro]
        except (TypeError, ValueError):
            pass

    hoy = timezone.now().date()
    nombres_meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                     'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    mes_inicio_param = request.GET.get('mes_inicio') or request.GET.get('mes')
    mes_fin_param = request.GET.get('mes_fin') or request.GET.get('mes')
    anio_param = request.GET.get('anio')

    try:
        mes_inicio = int(mes_inicio_param) if mes_inicio_param else hoy.month
        mes_fin = int(mes_fin_param) if mes_fin_param else mes_inicio
        anio = int(anio_param) if anio_param else hoy.year
    except (TypeError, ValueError):
        mes_inicio = hoy.month
        mes_fin = hoy.month
        anio = hoy.year

    if mes_inicio > mes_fin:
        mes_inicio, mes_fin = mes_fin, mes_inicio

    inicio_periodo = datetime.date(anio, mes_inicio, 1)
    fin_periodo = datetime.date(anio, mes_fin, calendar.monthrange(anio, mes_fin)[1])

    if mes_inicio == mes_fin:
        periodo_texto = f"{nombres_meses[mes_inicio]}-{anio}"
    else:
        periodo_texto = f"{nombres_meses[mes_inicio]}-{nombres_meses[mes_fin]}-{anio}"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    trabajadores = Trabajador.objects.filter(
        activo=True, area_id__in=areas_ids
    ).select_related('area', 'sede')

    trabajador_ids = list(trabajadores.values_list('id', flat=True))
    evaluaciones_periodo = EvaluacionMensual.objects.filter(
        trabajador_id__in=trabajador_ids,
        fecha_evaluacion__gte=inicio_periodo,
        fecha_evaluacion__lte=fin_periodo,
    ).select_related('evaluador').order_by('-fecha_evaluacion', '-id')

    evaluacion_reciente = {}
    for ev in evaluaciones_periodo:
        evaluacion_reciente.setdefault(ev.trabajador_id, ev)

    def _nombre_evaluador(t):
        ev = evaluacion_reciente.get(t.id)
        if not ev or not ev.evaluador:
            return 'Sin registro historico'
        e = ev.evaluador
        return f"{e.nombres} {e.apellido_paterno} {e.apellido_materno}".strip() or 'Sin registro historico'

    def _clasificacion(score):
        if score >= 9:
            return 'AD'
        if score >= 7:
            return 'A'
        if score >= 5:
            return 'B'
        return 'C'

    ranking = []
    for t in trabajadores:
        if _esta_en_periodo_prueba(t, inicio_periodo):
            continue
        desempeno_data = _calcular_desempeno_por_periodo(t, inicio_periodo, fin_periodo)
        asistencia_data = _calcular_asistencia_por_periodo(t, inicio_periodo, fin_periodo)

        nota_desempeno = round(desempeno_data['desempeno_compuesto'], 2)
        nota_asistencia = round(asistencia_data['nota'], 2)
        score = round(_ponderar_score(
            t, nota_desempeno, nota_asistencia, inicio_periodo, fin_periodo)['score_total'], 2)

        if score > 0:
            ranking.append({
                'trabajador': t,
                'score': score,
                'asistencia': nota_asistencia,
                'desempeno': nota_desempeno,
                'tardanzas': asistencia_data['tardanzas'],
                'faltas': asistencia_data['faltas'],
            })

    ranking.sort(key=lambda x: x['score'], reverse=True)

    headers = ['Puesto', 'Apellidos', 'Nombres', 'Área', 'Cargo evaluado', 'Sede', 'Evaluador',
               'Score Total', 'Calificación', 'Asistencia (0-10)', 'Desempeño', 'Tardanzas', 'Faltas']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for idx, item in enumerate(ranking, start=1):
        t = item['trabajador']
        row = [
            idx,
            f"{t.apellido_paterno} {t.apellido_materno}",
            t.nombres,
            t.area.nombre if t.area else 'Sin área',
            t.cargo or 'Sin cargo',
            t.sede.nombre if t.sede else 'Sin sede',
            _nombre_evaluador(t),
            round(item['score'], 2),
            _clasificacion(item['score']),
            round(item['asistencia'], 2),
            round(item['desempeno'], 2),
            item['tardanzas'],
            item['faltas'],
        ]
        ws.append(row)

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=idx + 1, column=col_idx).alignment = center_alignment

        ws.cell(row=idx + 1, column=10).number_format = '0.00'

        score_cell = ws.cell(row=idx + 1, column=8)
        if item['score'] >= 8:
            score_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif item['score'] >= 5:
            score_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for col, width in {'A': 8, 'B': 22, 'C': 18, 'D': 18, 'E': 20, 'F': 15, 'G': 28,
                       'H': 12, 'I': 12, 'J': 12, 'K': 10, 'L': 10, 'M': 10}.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Ranking_MiArea_{periodo_texto}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportacion_metricas_view(request):
    """Vista principal para exportación de métricas"""
    if not _usuario_puede_exportar_metricas(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta página')
        return redirect('metricas_ceneris:dashboard_gerente')

    context = _contexto_exportacion_metricas(request)
    return render(request, 'metricas_ceneris/exportacion_metricas.html', context)


@login_required
def exportacion_metricas_rrhh_view(request):
    """Vista aislada para RRHH: solo filtros y exportación."""
    if not _usuario_puede_exportar_metricas(request.user):
        messages.error(request, 'No tienes permisos para acceder a esta página')
        return redirect('recursoshumanos:dashboard')

    context = _contexto_exportacion_metricas(request)
    context['current_view'] = 'exportar_metricas_rrhh'
    return render(request, 'metricas_ceneris/exportacion_metricas_rrhh.html', context)


@login_required
def exportar_ranking_excel(request):
    """Exporta TODAS las métricas en un solo Excel con múltiples hojas"""
    if not _usuario_puede_exportar_metricas(request.user):
        return HttpResponse('No autorizado', status=403)

    hoy = timezone.now().date()

    # Obtener filtros de periodo
    filtro_periodo = _resolver_filtro_periodo_request(request, hoy, periodo_default='semestre')
    periodo = filtro_periodo['periodo']
    anio = filtro_periodo['anio']
    mes = filtro_periodo['mes']
    semestre = filtro_periodo['semestre']

    parametros_periodo = {'anio': anio}
    if periodo == 'mes':
        parametros_periodo['mes'] = mes
    else:
        parametros_periodo['semestre'] = semestre

    tipo = request.GET.get('tipo', 'todos')
    sede_id = request.GET.get('sede', 'todas')
    area_id = request.GET.get('area', 'todas')
    orden = request.GET.get('orden', 'score_desc')
    limite = request.GET.get('limite', 'todos')
    inicio_periodo, fin_periodo = _resolver_rango_periodo(hoy, periodo, **parametros_periodo)
    fecha_ref_periodo = _fecha_referencia_periodo(hoy, periodo, **parametros_periodo)

    # Estilos comunes
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    info_font = Font(bold=True, size=11)
    
    # Crear workbook con múltiples hojas
    wb = Workbook()
    
    # ============ HOJA 1: RANKING COMPLETO ============
    if wb.worksheets:
        ws_ranking = wb.active
        ws_ranking.title = "1_Ranking"
    else:
        ws_ranking = wb.create_sheet("1_Ranking")
    
    header_fill_ranking = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    # Filtrar trabajadores
    trabajadores = Trabajador.objects.filter(activo=True).select_related('area', 'sede')
    
    if tipo == 'operativo':
        trabajadores = trabajadores.filter(_q_trabajador_base())
    elif tipo in ['jefatura', 'supervisor', 'lider']:
        trabajadores = trabajadores.filter(_q_supervisor() | _q_responsable())
    
    if sede_id != 'todas' and sede_id:
        try:
            trabajadores = trabajadores.filter(sede_id=int(sede_id))
        except ValueError:
            pass
    
    if area_id != 'todas' and area_id:
        try:
            trabajadores = trabajadores.filter(area_id=int(area_id))
        except ValueError:
            pass

    trabajador_ids = list(trabajadores.values_list('id', flat=True))
    evaluaciones_periodo = EvaluacionMensual.objects.filter(
        trabajador_id__in=trabajador_ids,
        fecha_evaluacion__gte=inicio_periodo,
        fecha_evaluacion__lte=fin_periodo,
    ).select_related('evaluador').order_by('-fecha_evaluacion', '-id')

    evaluacion_reciente_por_trabajador = {}
    for evaluacion in evaluaciones_periodo:
        evaluacion_reciente_por_trabajador.setdefault(evaluacion.trabajador_id, evaluacion)

    def _datos_evaluador(trabajador):
        evaluacion = evaluacion_reciente_por_trabajador.get(trabajador.id)
        if not evaluacion or not evaluacion.evaluador:
            return 'Sin registro historico', 'Sin registro historico'

        evaluador = evaluacion.evaluador
        nombre_evaluador = f"{evaluador.nombres} {evaluador.apellido_paterno} {evaluador.apellido_materno}".strip()
        cargo_evaluador = evaluacion.cargo_evaluador or evaluador.cargo or 'Sin registro historico'
        return nombre_evaluador or 'Sin registro historico', cargo_evaluador

    def _clasificacion_score(score):
        if score >= 9:
            return 'AD'
        if score >= 7:
            return 'A'
        if score >= 5:
            return 'B'
        return 'C'
    
    # Calcular scores
    ranking = []
    for t in trabajadores:
        if _esta_en_periodo_prueba(t, fecha_ref_periodo):
            continue
        datos = _calcular_score_total(t, hoy, periodo, **parametros_periodo)
        
        if datos['score_total'] > 0:
            ranking.append({
                'trabajador': t,
                'score': datos['score_total'],
                'asistencia_pct': datos['asistencia_pct'],
                'asistencia_nota': round(datos['nota_asistencia'], 2),
                'desempeno': datos['desempeno_compuesto'],
                'tardanzas': datos['tardanzas'],
                'faltas': datos['faltas'],
                'area': t.area.nombre if t.area else 'Sin área',
                'sede': t.sede.nombre if t.sede else 'Sin sede',
            })
    
    # Ordenar
    if orden == 'score_desc':
        ranking.sort(key=lambda x: x['score'], reverse=True)
    elif orden == 'score_asc':
        ranking.sort(key=lambda x: x['score'])
    elif orden == 'nombre':
        ranking.sort(key=lambda x: (x['trabajador'].apellido_paterno, x['trabajador'].nombres))
    elif orden == 'area':
        ranking.sort(key=lambda x: x['area'])
    
    # Aplicar límite
    if limite != 'todos':
        try:
            limite_int = int(limite)
            ranking = ranking[:limite_int]
        except ValueError:
            pass
    
    # Headers
    headers_ranking = [
        'Puesto',
        'Apellidos',
        'Nombres',
        'Área',
        'Cargo evaluado',
        'Sede',
        'Evaluador',
        'Score Total',
        'Calificación',
        'Asistencia (0-10)',
        'Desempeño',
        'Tardanzas',
        'Faltas',
    ]
    ws_ranking.append(headers_ranking)
    
    for cell in ws_ranking[1]:
        cell.font = header_font
        cell.fill = header_fill_ranking
        cell.alignment = header_alignment
    
    # Datos
    for idx, item in enumerate(ranking, start=1):
        t = item['trabajador']
        nombre_evaluador, _ = _datos_evaluador(t)
        cargo_evaluado = t.cargo or 'Sin cargo'
        row = [
            idx,
            f"{t.apellido_paterno} {t.apellido_materno}",
            t.nombres,
            item['area'],
            cargo_evaluado,
            item['sede'],
            nombre_evaluador,
            round(item['score'], 2),
            _clasificacion_score(item['score']),
            round(item['asistencia_nota'], 2),
            round(item['desempeno'], 2),
            item['tardanzas'],
            item['faltas']
        ]
        ws_ranking.append(row)

        for col_idx in range(1, len(headers_ranking) + 1):
            ws_ranking.cell(row=idx + 1, column=col_idx).alignment = center_alignment

        # Asistencia en formato numerico (0.00), sin estilo de porcentaje.
        ws_ranking.cell(row=idx + 1, column=10).number_format = '0.00'
        
        # Estilo condicional para score
        score_cell = ws_ranking.cell(row=idx+1, column=8)
        score_val = item['score']
        if score_val >= 8:
            score_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        elif score_val >= 5:
            score_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Ajustar anchos
    ranking_column_widths = {
        'A': 8,
        'B': 22,
        'C': 18,
        'D': 18,
        'E': 20,
        'F': 15,
        'G': 28,
        'H': 12,
        'I': 12,
        'J': 12,
        'K': 10,
        'L': 10,
        'M': 10,
    }
    for col, width in ranking_column_widths.items():
        ws_ranking.column_dimensions[col].width = width
    
    # Datos auxiliares para hojas 2 y 3
    header_fill_podio = PatternFill(start_color="9333EA", end_color="9333EA", fill_type="solid")
    category_fill = PatternFill(start_color="A855F7", end_color="A855F7", fill_type="solid")

    # Agrupar por área
    areas_dict = {}
    for t in trabajadores:
        if not t.area:
            continue
        if _esta_en_periodo_prueba(t, fecha_ref_periodo):
            continue
        area_nombre = t.area.nombre
        datos = _calcular_score_total(t, hoy, periodo, **parametros_periodo)
        if datos['score_total'] > 0:
            if area_nombre not in areas_dict:
                areas_dict[area_nombre] = []
            areas_dict[area_nombre].append({
                'trabajador': t,
                'score': datos['score_total'],
                'asistencia_pct': datos['asistencia_pct'],
                'asistencia_nota': round(datos['nota_asistencia'], 2),
                'desempeno': datos['desempeno_compuesto'],
            })
    
    # Segmentaciones de ranking
    medallas = ['🥇', '🥈', '🥉']
    headers_podio = ['Pos', 'Nombre Completo', 'Área', 'Sede', 'Score', 'Calificación', 'Asistencia (0-10)', 'Desemp']

    ranking_lideres = [
        item for item in ranking
        if _es_supervisor(item['trabajador']) or _es_responsable(item['trabajador'])
    ]
    ranking_operativos = [item for item in ranking if _es_trabajador_base(item['trabajador'])]
    # ============ HOJA 2: COMPARATIVA DE ÁREAS ============
    ws_areas = wb.create_sheet("2_ComparativaAreas")
    header_fill_areas = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    
    # Calcular promedios por área
    comparativa = []
    for area_nombre, lista in areas_dict.items():
        if not lista:
            continue
        scores = [item['score'] for item in lista]
        asistencias = [item['asistencia_nota'] for item in lista]
        desempenos = [item['desempeno'] for item in lista]
        
        comparativa.append({
            'area': area_nombre,
            'trabajadores': len(lista),
            'score_promedio': sum(scores) / len(scores),
            'asistencia_promedio': sum(asistencias) / len(asistencias),
            'desempeno_promedio': sum(desempenos) / len(desempenos),
        })
    
    comparativa.sort(key=lambda x: x['score_promedio'], reverse=True)
    
    headers_areas = ['Pos', 'Área', 'Trabajadores', 'Score Prom', 'Asist Prom (0-10)', 'Desemp Prom']
    ws_areas.append(headers_areas)
    
    for cell in ws_areas[1]:
        cell.font = header_font
        cell.fill = header_fill_areas
        cell.alignment = header_alignment
    
    for idx, item in enumerate(comparativa, start=1):
        row = [
            idx,
            item['area'],
            item['trabajadores'],
            round(item['score_promedio'], 2),
            round(item['asistencia_promedio'], 2),
            round(item['desempeno_promedio'], 2),
        ]
        ws_areas.append(row)

        for col_idx in range(1, len(headers_areas) + 1):
            ws_areas.cell(row=idx + 1, column=col_idx).alignment = center_alignment

        ws_areas.cell(row=idx + 1, column=5).number_format = '0.00'
        
        score_cell = ws_areas.cell(row=idx+1, column=4)
        if idx == 1:
            score_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        elif idx == 2:
            score_cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        elif idx == 3:
            score_cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
    
    ws_areas.column_dimensions['A'].width = 6
    ws_areas.column_dimensions['B'].width = 36
    ws_areas.column_dimensions['C'].width = 14
    ws_areas.column_dimensions['D'].width = 12
    ws_areas.column_dimensions['E'].width = 14
    ws_areas.column_dimensions['F'].width = 14
    ws_areas.column_dimensions['G'].width = 12
    ws_areas.column_dimensions['H'].width = 12
    
    # GRÁFICO: Comparativa de Áreas
    chart_bottom_row = 0
    if len(comparativa) > 0:
        comparativa_chart = comparativa[:5]
        total_areas_chart = len(comparativa_chart)
        max_area_len = max(len(item['area']) for item in comparativa_chart) if comparativa_chart else 12

        chart_areas = BarChart()
        chart_areas.type = "bar"  # Horizontal
        chart_areas.title = "Score Promedio por Area"
        chart_areas.style = 10
        chart_areas.width = min(28, max(20, 15 + (max_area_len * 0.22)))
        chart_areas.height = max(10, 4 + (total_areas_chart * 1.5))
        chart_areas.gapWidth = 45
        
        # Referencias a los datos (Score Promedio columna 4) - SIN título de columna
        data_ref = Reference(ws_areas, min_col=4, min_row=2, max_row=len(comparativa_chart) + 1)
        cats_ref = Reference(ws_areas, min_col=2, min_row=2, max_row=len(comparativa_chart) + 1)
        
        chart_areas.add_data(data_ref, titles_from_data=False)
        chart_areas.set_categories(cats_ref)
        chart_areas.x_axis.title = "Puntaje"
        chart_areas.y_axis.delete = False
        chart_areas.legend = None
        
        # Etiquetas de datos bien espaciadas hacia afuera
        chart_areas.dataLabels = DataLabelList()
        chart_areas.dataLabels.showVal = True
        chart_areas.dataLabels.showCatName = True
        chart_areas.dataLabels.showSerName = False
        chart_areas.dataLabels.showLegendKey = False
        chart_areas.dataLabels.separator = "; "
        chart_areas.dataLabels.position = "outEnd"
        chart_areas.y_axis.tickLblPos = "nextTo"

        # Dar margen al eje X para que el texto no se corte
        max_score = max([item['score_promedio'] for item in comparativa_chart]) if comparativa_chart else 5
        chart_areas.x_axis.scaling.max = max_score * 1.3

        # Colorear primeras 3 barras como Oro, Plata y Bronce
        colores_podio = ["FFD700", "C0C0C0", "CD7F32"]
        serie_areas = chart_areas.series[0]
        for i, color in enumerate(colores_podio[:len(comparativa_chart)]):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            serie_areas.dPt.append(pt)
        
        # Posicionar gráfico a la derecha de los datos (columna H)
        ws_areas.add_chart(chart_areas, "H2")
        chart_bottom_row = 2 + int(chart_areas.height * 5.2)

    # Detalle completo por área
    fila_area_podio = len(comparativa) + 4
    if chart_bottom_row:
        fila_area_podio = max(fila_area_podio, chart_bottom_row + 2)
    ws_areas.merge_cells(f'A{fila_area_podio}:H{fila_area_podio}')
    cell = ws_areas.cell(row=fila_area_podio, column=1)
    cell.value = "DETALLE COMPLETO POR AREA"
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    cell.alignment = header_alignment
    fila_area_podio += 1

    for area_nombre, lista in sorted(areas_dict.items()):
        lista.sort(key=lambda x: x['score'], reverse=True)

        ws_areas.merge_cells(f'A{fila_area_podio}:H{fila_area_podio}')
        cell = ws_areas.cell(row=fila_area_podio, column=1)
        cell.value = f"Area: {area_nombre}"
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = category_fill
        cell.alignment = header_alignment
        fila_area_podio += 1

        for col_idx, header in enumerate(headers_podio, start=1):
            cell = ws_areas.cell(row=fila_area_podio, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = header_fill_podio
            cell.alignment = header_alignment
        fila_area_podio += 1

        for idx, item in enumerate(lista, start=1):
            t = item['trabajador']
            posicion = medallas[idx - 1] if idx <= 3 else idx
            row = [
                posicion,
                f"{t.nombres} {t.apellido_paterno} {t.apellido_materno}",
                t.area.nombre if t.area else '-',
                t.sede.nombre if t.sede else '-',
                round(item['score'], 2),
                _clasificacion_score(item['score']),
                round(item['asistencia_nota'], 2),
                round(item['desempeno'], 2),
            ]

            for col_idx, val in enumerate(row, start=1):
                cell = ws_areas.cell(row=fila_area_podio, column=col_idx, value=val)
                cell.alignment = center_alignment
                if col_idx == 7:
                    cell.number_format = '0.00'
                if idx == 1:
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                elif idx == 2:
                    cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                elif idx == 3:
                    cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

            fila_area_podio += 1

        fila_area_podio += 1
    # ============ HOJA 3: RESUMEN EJECUTIVO ============
    ws_resumen = wb.create_sheet("3_ResumenEjecutivo")
    header_fill_resumen = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    section_fill_trab = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    section_fill_jef = PatternFill(start_color="334155", end_color="334155", fill_type="solid")

    if periodo == 'mes':
        meses_nombres = {
            1: 'Enero',
            2: 'Febrero',
            3: 'Marzo',
            4: 'Abril',
            5: 'Mayo',
            6: 'Junio',
            7: 'Julio',
            8: 'Agosto',
            9: 'Septiembre',
            10: 'Octubre',
            11: 'Noviembre',
            12: 'Diciembre',
        }
        periodo_texto = f"{meses_nombres.get(mes, 'Mes')} {anio}"
    else:
        periodo_texto = f"Semestre {semestre} de {anio}"
    ws_resumen.merge_cells('A1:R1')
    ws_resumen['A1'] = 'RESUMEN EJECUTIVO'
    ws_resumen['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_resumen['A1'].fill = header_fill_resumen
    ws_resumen['A1'].alignment = header_alignment
    ws_resumen['B3'] = 'Período:'
    ws_resumen['C3'] = periodo_texto
    ws_resumen['E3'] = 'Total evaluados:'
    ws_resumen['F3'] = len(ranking)
    ws_resumen['K3'] = 'Score promedio general:'
    ws_resumen['L3'] = round(sum(item['score'] for item in ranking) / len(ranking), 2) if ranking else 0
    for celda in ['B3', 'E3', 'K3']:
        ws_resumen[celda].font = Font(bold=True, size=13)
    for celda in ['C3', 'F3', 'L3']:
        ws_resumen[celda].font = Font(bold=True, size=14)

    top_operativos = sorted(ranking_operativos, key=lambda x: x['score'], reverse=True)[:5]
    top_lideres = sorted(ranking_lideres, key=lambda x: x['score'], reverse=True)[:5]
    bajos_operativos = sorted(ranking_operativos, key=lambda x: x['score'])[:5]
    bajos_lideres = sorted(ranking_lideres, key=lambda x: x['score'])[:5]

    def _escribir_tabla_top(row_start, col_start, titulo, items, fill_titulo, chart_anchor, chart_title):
        col_end = col_start + 7
        helper_col = col_start + 8
        ws_resumen.merge_cells(f'{get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_start}')
        cell_titulo = ws_resumen.cell(row=row_start, column=col_start, value=titulo)
        cell_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        cell_titulo.fill = fill_titulo
        cell_titulo.alignment = header_alignment

        headers = ['Pos', 'Nombre Completo', 'Área', 'Sede', 'Evaluador', 'Cargo evaluado', 'Score', 'Calificación']
        header_row = row_start + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_resumen.cell(row=header_row, column=col_start + col_idx - 1, value=header)
            cell.font = header_font
            cell.fill = header_fill_resumen
            cell.alignment = header_alignment

        data_start = header_row + 1
        for idx, item in enumerate(items, start=1):
            t = item['trabajador']
            nombre_evaluador, _ = _datos_evaluador(t)
            cargo_evaluado = t.cargo or 'Sin cargo'
            nombre_tabla = f"{t.nombres} {t.apellido_paterno} {t.apellido_materno}".strip()
            fila = [
                medallas[idx - 1] if idx <= 3 else idx,
                nombre_tabla,
                item['area'],
                item['sede'],
                nombre_evaluador,
                cargo_evaluado,
                round(item['score'], 2),
                _clasificacion_score(item['score']),
            ]
            for col_idx, val in enumerate(fila, start=1):
                col_dest = col_start + col_idx - 1
                cell = ws_resumen.cell(row=data_start + idx - 1, column=col_dest, value=val)
                cell.alignment = center_alignment
                if idx == 1:
                    if col_idx in (1, 7):
                        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                elif idx == 2:
                    if col_idx in (1, 7):
                        cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
                elif idx == 3:
                    if col_idx in (1, 7):
                        cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

            ws_resumen.cell(row=data_start + idx - 1, column=helper_col, value=nombre_tabla)

        table_bottom_row = data_start + max(len(items), 1) - 1
        chart_bottom_row_local = table_bottom_row

        if items:
            chart_items = items[:3]
            data_end = data_start + len(chart_items) - 1
            max_name_len = max(len((ws_resumen.cell(row=data_start + i, column=helper_col).value or '')) for i in range(len(chart_items)))
            chart = BarChart()
            chart.type = "bar"
            chart.title = chart_title
            chart.style = 10
            chart.width = min(18, max(14, 11 + (max_name_len * 0.10)))
            chart.height = max(7.5, 3.5 + (len(chart_items) * 1.5))
            chart.gapWidth = 55
            data_ref = Reference(ws_resumen, min_col=col_start + 6, min_row=data_start, max_row=data_end)
            cats_ref = Reference(ws_resumen, min_col=helper_col, min_row=data_start, max_row=data_end)
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)
            chart.x_axis.title = None
            chart.y_axis.title = None
            chart.legend = None
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            chart.dataLabels.showCatName = True
            chart.dataLabels.showSerName = False
            chart.dataLabels.showLegendKey = False
            chart.dataLabels.separator = "; "
            chart.dataLabels.position = "outEnd"
            chart.y_axis.delete = False
            chart.y_axis.tickLblPos = "nextTo"

            max_score = max([item['score'] for item in chart_items]) if chart_items else 5
            chart.x_axis.scaling.max = max_score * 1.4

            serie = chart.series[0]
            colores = ["FFD700", "C0C0C0", "CD7F32"]
            for i, color in enumerate(colores[:len(chart_items)]):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = color
                serie.dPt.append(pt)

            ws_resumen.add_chart(chart, chart_anchor)
            anchor_row = int(''.join(ch for ch in chart_anchor if ch.isdigit()) or '1')
            chart_bottom_row_local = max(chart_bottom_row_local, anchor_row + int(chart.height * 5.2))

        return chart_bottom_row_local

    def _escribir_tabla_bajos(row_start, col_start, titulo, items, fill_titulo):
        col_end = col_start + 7
        ws_resumen.merge_cells(f'{get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_start}')
        cell_titulo = ws_resumen.cell(row=row_start, column=col_start, value=titulo)
        cell_titulo.font = Font(bold=True, size=12, color="FFFFFF")
        cell_titulo.fill = fill_titulo
        cell_titulo.alignment = header_alignment

        headers = ['Pos', 'Nombre Completo', 'Área', 'Sede', 'Evaluador', 'Cargo evaluado', 'Score', 'Calificación']
        header_row = row_start + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_resumen.cell(row=header_row, column=col_start + col_idx - 1, value=header)
            cell.font = header_font
            cell.fill = header_fill_resumen
            cell.alignment = header_alignment

        for idx, item in enumerate(items, start=1):
            t = item['trabajador']
            nombre_evaluador, _ = _datos_evaluador(t)
            cargo_evaluado = t.cargo or 'Sin cargo'
            fila = [
                medallas[idx - 1] if idx <= 3 else idx,
                f"{t.nombres} {t.apellido_paterno} {t.apellido_materno}",
                item['area'],
                item['sede'],
                nombre_evaluador,
                cargo_evaluado,
                round(item['score'], 2),
                _clasificacion_score(item['score']),
            ]
            for col_idx, val in enumerate(fila, start=1):
                cell = ws_resumen.cell(row=header_row + idx, column=col_start + col_idx - 1, value=val)
                cell.alignment = center_alignment
                if col_idx in (1, 7):
                    if idx == 1:
                        # Destacar solo la nota mas baja con rojo suave.
                        cell.fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    fila_top = 7
    top_trab_bottom = _escribir_tabla_top(
        fila_top,
        1,
        'TOP 5 TRABAJADORES',
        top_operativos,
        section_fill_trab,
        'B14',
        'Top 3 Trabajadores',
    )
    top_lider_bottom = _escribir_tabla_top(
        fila_top,
        11,
        'TOP 5 LIDERES',
        top_lideres,
        section_fill_jef,
        'K14',
        'Top 3 Lideres',
    )

    fila_bajos = max(top_trab_bottom, top_lider_bottom) + 2
    _escribir_tabla_bajos(
        fila_bajos,
        1,
        '5 NOTAS MÁS BAJAS - TRABAJADORES',
        bajos_operativos,
        section_fill_trab,
    )
    _escribir_tabla_bajos(
        fila_bajos,
        11,
        '5 NOTAS MÁS BAJAS - LIDERES',
        bajos_lideres,
        section_fill_jef,
    )

    for col in ('A', 'K'):
        ws_resumen.column_dimensions[col].width = 8
    for col in ('B', 'L'):
        ws_resumen.column_dimensions[col].width = 32
    for col in ('C', 'M'):
        ws_resumen.column_dimensions[col].width = 20
    for col in ('D', 'N'):
        ws_resumen.column_dimensions[col].width = 14
    for col in ('E', 'O'):
        ws_resumen.column_dimensions[col].width = 18
    for col in ('F', 'P'):
        ws_resumen.column_dimensions[col].width = 18
    for col in ('G', 'Q'):
        ws_resumen.column_dimensions[col].width = 8
    for col in ('H', 'R'):
        ws_resumen.column_dimensions[col].width = 12

    ws_resumen.column_dimensions['I'].width = 3
    ws_resumen.column_dimensions['I'].hidden = True
    ws_resumen.column_dimensions['J'].width = 4
    ws_resumen.column_dimensions['S'].width = 3
    ws_resumen.column_dimensions['S'].hidden = True
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    periodo_texto_archivo = f"S{semestre}-{anio}"
    response['Content-Disposition'] = f'attachment; filename="Metricas_Completas_{periodo_texto_archivo}.xlsx"'
    wb.save(response)
    return response

def importar_asistencias(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        anio = int(request.POST.get('anio'))
        mes = int(request.POST.get('mes'))

        if not archivo:
            messages.error(request, "Por favor sube un archivo Excel.")
            return render(request, 'metricas_ceneris/importar_asistencia.html')

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            sheet = wb.active

            empleados_procesados = 0
            dnis_procesados = []       
            dnis_no_encontrados = []   
            
            # 1. Obtener las fechas de las columnas (Fila 6)
            fechas_por_columna = {}
            for col in range(4, sheet.max_column + 1):
                dia_valor = sheet.cell(row=6, column=col).value
                if dia_valor is not None:
                    try:
                        dia = int(dia_valor)
                        mes_calculado = mes
                        anio_calculado = anio
                        # Si el día es mayor a 25, pertenece al mes anterior
                        if dia > 25:
                            mes_calculado = mes - 1
                            if mes_calculado == 0:  
                                mes_calculado = 12
                                anio_calculado = anio - 1
                                
                        fechas_por_columna[col] = datetime.date(anio_calculado, mes_calculado, dia)
                    except ValueError:
                        pass 

            # 2. Iterar sobre cada empleado (Desde la fila 7)
            for fila in sheet.iter_rows(min_row=7, values_only=False):
                dni_excel = fila[2].value 
                
                if not dni_excel:
                    continue

                dni_excel = str(dni_excel).strip()
                
                # Buscamos al TRABAJADOR por DNI
                try:
                    trabajador_db = Trabajador.objects.get(dni=dni_excel) 
                except Trabajador.DoesNotExist:
                    if dni_excel not in dnis_no_encontrados:
                        dnis_no_encontrados.append(dni_excel)
                    continue 

                # 3. Procesar las celdas de este empleado
                for col_idx in range(4, sheet.max_column + 1):
                    valor_celda = fila[col_idx - 1].value
                    
                    if col_idx not in fechas_por_columna:
                        continue
                        
                    fecha_solo_dia = fechas_por_columna[col_idx]
                    
                    resultado_tareo = 'A'  
                    horas_tardanza_tareo = 0.0
                    es_falta = False
                    minutos_tardanza = 0

                    # Evaluamos qué dice la celda del Excel
                    if str(valor_celda).strip().upper() == 'F':
                        resultado_tareo = 'F'
                        es_falta = True
                    else:
                        if valor_celda is not None and str(valor_celda).strip() != '':
                            if isinstance(valor_celda, datetime.time):
                                minutos_tardanza = (valor_celda.hour * 60) + valor_celda.minute
                            elif isinstance(valor_celda, str) and ':' in valor_celda:
                                try:
                                    h, m = map(int, valor_celda.split(':')[:2])
                                    minutos_tardanza = (h * 60) + m
                                except ValueError:
                                    pass
                        
                        if minutos_tardanza > 0:
                            horas_tardanza_tareo = round(minutos_tardanza / 60.0, 2)

                    # ---------------------------------------------------------
                    # ACCIÓN 1: GUARDAR EN TAREO DIARIO (Panel Gerencial)
                    # ---------------------------------------------------------
                    tareo, created = TareoDiario.objects.get_or_create(
                        trabajador=trabajador_db,
                        fecha=fecha_solo_dia,
                        defaults={'estado': 'O'}
                    )
                    tareo.resultado = resultado_tareo
                    tareo.horas_tardanza = horas_tardanza_tareo
                    tareo.save()

                    # ---------------------------------------------------------
                    # ACCIÓN 2: GUARDAR EN ASISTENCIA (App Móvil / Flutter)
                    # ---------------------------------------------------------
                    # Solo creamos la marcación si NO es falta y si el trabajador tiene un User vinculado
                    if not es_falta and hasattr(trabajador_db, 'user') and trabajador_db.user:
                        
                        # Definimos la hora oficial a las 08:30 AM
                        hora_base = datetime.time(8, 30, 0)
                        fecha_hora_oficial = datetime.datetime.combine(fecha_solo_dia, hora_base)
                        
                        # Sumamos la tardanza
                        timestamp_final = fecha_hora_oficial + datetime.timedelta(minutes=minutos_tardanza)
                        timestamp_aware = make_aware(timestamp_final)

                        # Verificamos si ya existe para no duplicar
                        registro_existente = Asistencia.objects.filter(
                            usuario=trabajador_db.user,
                            tipo_marcacion='Entrada',
                            timestamp__date=fecha_solo_dia
                        ).first()

                        if registro_existente:
                            registro_existente.timestamp = timestamp_aware
                            registro_existente.nombre_ubicacion = 'Importado desde RRHH (Excel)'
                            registro_existente.save()
                        else:
                            Asistencia.objects.create(
                                usuario=trabajador_db.user,
                                timestamp=timestamp_aware,
                                tipo_marcacion='Entrada',
                                nombre_ubicacion='Importado desde RRHH (Excel)'
                            )
                
                empleados_procesados += 1
                if dni_excel not in dnis_procesados:
                    dnis_procesados.append(dni_excel)

            if dnis_procesados:
                lista_exito = ", ".join(dnis_procesados)
                messages.success(request, f"¡Éxito! Se procesaron {empleados_procesados} trabajadores. Tareos y Asistencias actualizados. DNIs: {lista_exito}")
            
            if dnis_no_encontrados:
                lista_errores = ", ".join(dnis_no_encontrados)
                messages.warning(request, f"Atención: Los siguientes DNIs no existen en la base de datos y fueron omitidos: {lista_errores}")

        except Exception as e:
            messages.error(request, f"Ocurrió un error general leyendo el Excel: {str(e)}")

    return render(request, 'metricas_ceneris/importar_asistencia.html')

@login_required
def dashboard_jefe(request):
    try:
        jefe = request.user.trabajador
    except:
        return render(request, 'metricas_ceneris/error.html', {'mensaje': 'No tienes perfil asignado.'})

    if _es_responsable(jefe):
        return redirect('metricas_ceneris:dashboard_responsable')

    if not _es_supervisor(jefe):
        return redirect('metricas_ceneris:dashboard_trabajador')

    areas_asignadas = []
    area_activa_id = None

    areas_asignadas = _areas_bajo_responsabilidad(jefe)

    if not areas_asignadas:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes areas asignadas para visualizar este panel.'
        })

    area_id_param = request.GET.get('area_id')
    area = None

    if area_id_param:
        try:
            area_id = int(area_id_param)
            area = next((a for a in areas_asignadas if a.id == area_id), None)
        except (TypeError, ValueError):
            area = None

    if area is None:
        area = areas_asignadas[0]

    area_activa_id = area.id

    trabajadores = _objetivos_evaluacion_por_area(jefe, area)
    es_responsable = _es_responsable(jefe)
    hay_supervisores_area = _supervisores_activos_por_area(area).exclude(id=jefe.id).exists()

    if es_responsable and hay_supervisores_area:
        objetivo_label_singular = 'Supervisor'
        objetivo_label_plural = 'Supervisores'
    else:
        objetivo_label_singular = 'Trabajador'
        objetivo_label_plural = 'Trabajadores'

    # 1. LÓGICA DE FECHAS
    hoy = timezone.now().date()
    mes_param = request.GET.get('mes')
    anio_param = request.GET.get('anio')
    
    if mes_param and anio_param:
        try:
            mes = int(mes_param)
            anio = int(anio_param)
            ultimo_dia = calendar.monthrange(anio, mes)[1]
            inicio_mes = datetime.date(anio, mes, 1)
            fin_mes = datetime.date(anio, mes, ultimo_dia)
        except ValueError:
            mes = hoy.month
            anio = hoy.year
            ultimo_dia = calendar.monthrange(anio, mes)[1]
            inicio_mes = hoy.replace(day=1)
            fin_mes = datetime.date(anio, mes, ultimo_dia)
    else:
        mes = hoy.month
        anio = hoy.year
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        inicio_mes = hoy.replace(day=1)
        fin_mes = datetime.date(anio, mes, ultimo_dia)

    # 2. CÁLCULOS POR TRABAJADOR Y DEL ÁREA
    datos_equipo = []
    suma_score_final = 0
    suma_desempeno = 0
    suma_asistencia = 0
    suma_disciplinaria = 0
    suma_conocimiento = 0
    total_evaluados = 0
    total_faltas_area = 0

    for t in trabajadores:
        score_data = _calcular_score_total(t, hoy, 'mes', anio=anio, mes=mes)
        nota_desempeno = score_data['nota_desempeno_periodo']
        porc_asistencia = score_data['asistencia_pct']
        faltas = score_data['faltas']
        
        total_faltas_area += faltas
        # Nota de asistencia (fórmula PPT), la misma que usa el score.
        nota_asistencia = round(score_data['nota_asistencia'], 2)

        # Score Final
        score_trabajador = round(score_data['score_total'], 2)

        # Letra individual
        clasif_ind = "C"
        if score_trabajador >= 9: clasif_ind = "AD"
        elif score_trabajador >= 7: clasif_ind = "A"
        elif score_trabajador >= 5: clasif_ind = "B"

        # Guardamos en la lista
        datos_equipo.append({
            'trabajador': t,
            'desempeno': nota_desempeno,
            'asistencia': nota_asistencia,
            'nota_disciplinaria': round(score_data['nota_disciplinaria'], 2),
            'nota_conocimiento': round(score_data['nota_conocimiento'], 2) if score_data['nota_conocimiento'] is not None else None,
            'score': score_trabajador,
            'faltas': faltas,
            'clasificacion': clasif_ind
        })

        suma_score_final += score_trabajador
        suma_desempeno += nota_desempeno
        suma_asistencia += nota_asistencia
        suma_disciplinaria += score_data['nota_disciplinaria']
        suma_conocimiento += score_data['nota_conocimiento']
        total_evaluados += 1

    # Ordenar Ranking (De mayor a menor score)
    datos_equipo = sorted(datos_equipo, key=lambda x: x['score'], reverse=True)
    mejor_trabajador = datos_equipo[0] if datos_equipo else None

    # Promedios del Área
    promedio_area_final = round(suma_score_final / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_desempeno = round(suma_desempeno / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_asistencia = round(suma_asistencia / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_disciplinaria = round(suma_disciplinaria / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_conocimiento = round(suma_conocimiento / total_evaluados, 2) if total_evaluados > 0 else 0

    clasificacion = "Pendiente"
    if promedio_area_final > 0:
        if promedio_area_final >= 9: clasificacion = "AD"
        elif promedio_area_final >= 7: clasificacion = "A"
        elif promedio_area_final >= 5: clasificacion = "B"
        else: clasificacion = "C"

    # Preparar datos para Chart.js
    nombres_chart = [d['trabajador'].nombres.split()[0] for d in datos_equipo] # Solo primer nombre
    desempeno_chart = [d['desempeno'] for d in datos_equipo]
    asistencia_chart = [d['asistencia'] for d in datos_equipo]
    score_chart = [d['score'] for d in datos_equipo]

    return render(request, 'metricas_ceneris/jefe/dashboard_jefe.html', {
        'jefe': jefe,
        'area': area,
        'areas_asignadas': areas_asignadas,
        'area_activa_id': area_activa_id,
        'datos_equipo': datos_equipo,
        'mejor_trabajador': mejor_trabajador,
        'total_faltas_area': total_faltas_area,
        'promedio_area': promedio_area_final,        
        'promedio_desempeno': promedio_area_desempeno,
        'promedio_asistencia': promedio_area_asistencia,
        'promedio_disciplinaria': promedio_area_disciplinaria,
        'promedio_conocimiento': promedio_area_conocimiento,
        'clasificacion': clasificacion,
        'mes_seleccionado': mes,
        'anio_seleccionado': anio,
        'scope_label': 'Ámbito responsable' if es_responsable else 'Área operativa',
        'objetivo_label_singular': objetivo_label_singular,
        'objetivo_label_plural': objetivo_label_plural,
        'composicion_title': f'Composición de {objetivo_label_plural}',
        'clasificacion_title': f'Clasificación Detallada de {objetivo_label_plural}',
        
        # Datos para los gráficos
        'nombres_chart': nombres_chart,
        'desempeno_chart': desempeno_chart,
        'asistencia_chart': asistencia_chart,
        'score_chart': score_chart,
    })


@login_required
def dashboard_responsable(request):
    try:
        perfil = request.user.trabajador
    except Exception:
        return render(request, 'metricas_ceneris/error.html', {'mensaje': 'No tienes perfil asignado.'})

    if not _es_responsable(perfil):
        return redirect('metricas_ceneris:inicio_metricas')

    areas_asignadas = _areas_bajo_responsabilidad(perfil)
    if not areas_asignadas:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No tienes areas asignadas para visualizar este panel.'
        })

    areas_con_supervisor, areas_sin_supervisor = _clasificar_areas_responsable_por_supervision(
        perfil,
        areas_asignadas,
    )

    modo_param = (request.GET.get('modo') or '').strip().lower()
    if modo_param in ['supervisor', 'supervisores']:
        modo_actual = 'supervisores'
    elif modo_param in ['equipo', 'equipos', 'trabajador', 'trabajadores']:
        modo_actual = 'equipos'
    else:
        modo_actual = ''

    if areas_con_supervisor and areas_sin_supervisor:
        if modo_actual not in ['supervisores', 'equipos']:
            modo_actual = 'supervisores'
    elif areas_con_supervisor:
        modo_actual = 'supervisores'
    elif areas_sin_supervisor:
        modo_actual = 'equipos'
    else:
        return render(request, 'metricas_ceneris/error.html', {
            'mensaje': 'No hay objetivos activos para este responsable. Contacta a RRHH/Calidad.'
        })

    hoy = timezone.now().date()
    nombres_meses = [
        '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
    ]

    mes_inicio_param = request.GET.get('mes_inicio') or request.GET.get('mes')
    mes_fin_param = request.GET.get('mes_fin') or request.GET.get('mes')
    anio_param = request.GET.get('anio')

    try:
        mes_inicio = _normalizar_mes(mes_inicio_param, default=hoy.month)
        mes_fin = _normalizar_mes(mes_fin_param, default=mes_inicio)
        anio = int(anio_param) if anio_param else hoy.year
    except (TypeError, ValueError):
        mes_inicio = hoy.month
        mes_fin = hoy.month
        anio = hoy.year

    if mes_inicio > mes_fin:
        mes_inicio, mes_fin = mes_fin, mes_inicio

    inicio_periodo = datetime.date(anio, mes_inicio, 1)
    fin_periodo = datetime.date(anio, mes_fin, calendar.monthrange(anio, mes_fin)[1])
    rango_meses_label = f'{nombres_meses[mes_inicio]} a {nombres_meses[mes_fin]} {anio}'

    area = None
    areas_equipos = areas_sin_supervisor

    if modo_actual == 'supervisores':
        trabajadores = _objetivos_supervisores_consolidados(perfil, areas_con_supervisor)
        scope_label = 'Ámbito responsable consolidado'
        objetivo_label_singular = 'Supervisor'
        objetivo_label_plural = 'Supervisores'
        vista_detalle = f'{len(areas_con_supervisor)} áreas con supervisión activa'
    else:
        area_id_param = request.GET.get('area_id')
        if area_id_param:
            try:
                area_id = int(area_id_param)
                area = next((a for a in areas_equipos if a.id == area_id), None)
            except (TypeError, ValueError):
                area = None

        if area is None and areas_equipos:
            area = areas_equipos[0]

        trabajadores = _objetivos_evaluacion_por_area(perfil, area)
        scope_label = 'Ámbito responsable por equipo'
        objetivo_label_singular = 'Trabajador'
        objetivo_label_plural = 'Trabajadores'
        vista_detalle = area.nombre if area else 'Sin área operativa'

    datos_equipo = []
    suma_score_final = 0
    suma_desempeno = 0
    suma_asistencia = 0
    suma_disciplinaria = 0
    suma_conocimiento = 0
    total_evaluados = 0
    total_faltas_area = 0

    for t in trabajadores:
        desempeno_data = _calcular_desempeno_por_periodo(t, inicio_periodo, fin_periodo)
        asistencia_data = _calcular_asistencia_por_periodo(t, inicio_periodo, fin_periodo)

        nota_desempeno = round(desempeno_data['desempeno_compuesto'], 2)
        porc_asistencia = asistencia_data['porcentaje']
        faltas = asistencia_data['faltas']

        total_faltas_area += faltas
        # Nota de asistencia (fórmula PPT), la misma que alimenta el score.
        nota_asistencia = round(asistencia_data['nota'], 2)
        ponderacion = _ponderar_score(t, nota_desempeno, nota_asistencia, inicio_periodo, fin_periodo)
        score_trabajador = round(ponderacion['score_total'], 2)

        clasif_ind = 'C'
        if score_trabajador >= 9:
            clasif_ind = 'AD'
        elif score_trabajador >= 7:
            clasif_ind = 'A'
        elif score_trabajador >= 5:
            clasif_ind = 'B'

        datos_equipo.append({
            'trabajador': t,
            'desempeno': nota_desempeno,
            'asistencia': nota_asistencia,
            'nota_disciplinaria': round(ponderacion['nota_disciplinaria'], 2),
            'nota_conocimiento': round(ponderacion['nota_conocimiento'], 2) if ponderacion['nota_conocimiento'] is not None else None,
            'score': score_trabajador,
            'faltas': faltas,
            'clasificacion': clasif_ind,
        })

        suma_score_final += score_trabajador
        suma_desempeno += nota_desempeno
        suma_asistencia += nota_asistencia
        suma_disciplinaria += ponderacion['nota_disciplinaria']
        suma_conocimiento += ponderacion['nota_conocimiento']
        total_evaluados += 1

    datos_equipo = sorted(datos_equipo, key=lambda x: x['score'], reverse=True)
    mejor_trabajador = datos_equipo[0] if datos_equipo else None

    promedio_area_final = round(suma_score_final / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_desempeno = round(suma_desempeno / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_asistencia = round(suma_asistencia / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_disciplinaria = round(suma_disciplinaria / total_evaluados, 2) if total_evaluados > 0 else 0
    promedio_area_conocimiento = round(suma_conocimiento / total_evaluados, 2) if total_evaluados > 0 else 0

    clasificacion = 'Pendiente'
    if promedio_area_final > 0:
        if promedio_area_final >= 9:
            clasificacion = 'AD'
        elif promedio_area_final >= 7:
            clasificacion = 'A'
        elif promedio_area_final >= 5:
            clasificacion = 'B'
        else:
            clasificacion = 'C'

    nombres_chart = [d['trabajador'].nombres.split()[0] for d in datos_equipo]
    desempeno_chart = [d['desempeno'] for d in datos_equipo]
    asistencia_chart = [d['asistencia'] for d in datos_equipo]
    score_chart = [d['score'] for d in datos_equipo]

    mostrar_selector_modo = bool(areas_con_supervisor and areas_sin_supervisor)
    primera_area_equipos_id = areas_equipos[0].id if areas_equipos else None

    return render(request, 'metricas_ceneris/responsable/dashboard_responsable.html', {
        'jefe': perfil,
        'area': area,
        'areas_asignadas': areas_asignadas,
        'areas_con_supervisor': areas_con_supervisor,
        'areas_sin_supervisor': areas_sin_supervisor,
        'areas_equipos': areas_equipos,
        'primera_area_equipos_id': primera_area_equipos_id,
        'modo_actual': modo_actual,
        'mostrar_selector_modo': mostrar_selector_modo,
        'vista_detalle': vista_detalle,
        'scope_label': scope_label,
        'datos_equipo': datos_equipo,
        'mejor_trabajador': mejor_trabajador,
        'total_faltas_area': total_faltas_area,
        'promedio_area': promedio_area_final,
        'promedio_desempeno': promedio_area_desempeno,
        'promedio_asistencia': promedio_area_asistencia,
        'promedio_disciplinaria': promedio_area_disciplinaria,
        'promedio_conocimiento': promedio_area_conocimiento,
        'clasificacion': clasificacion,
        'mes_inicio_seleccionado': mes_inicio,
        'mes_fin_seleccionado': mes_fin,
        'anio_seleccionado': anio,
        'rango_meses_label': rango_meses_label,
        'objetivo_label_singular': objetivo_label_singular,
        'objetivo_label_plural': objetivo_label_plural,
        'composicion_title': f'Composición de {objetivo_label_plural}',
        'clasificacion_title': f'Clasificación de {objetivo_label_plural}',
        'nombres_chart': nombres_chart,
        'desempeno_chart': desempeno_chart,
        'asistencia_chart': asistencia_chart,
        'score_chart': score_chart,
    })