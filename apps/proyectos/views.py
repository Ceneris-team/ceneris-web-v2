from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from collections import defaultdict
from django.contrib import messages
import json
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.http import require_POST 
from django.db.models import F, Max, Min, Count, Sum, Q, Case, When, FloatField, Value
from datetime import timedelta, datetime, date
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from .models import RegistroActividad
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

import calendar
from PIL import Image as PilImage # Importamos la biblioteca de imágenes de Python
from io import BytesIO 


from .models import Proyecto, TareaP, SubTarea
from .forms import ProyectoForm, TareaPForm, SubTareaForm, ProyectoEditForm, SubTareaEditForm
from inventario.models import Insumo, AsignacionInsumo, ItemInsumo
from personal.models import Personal
from inventario.models import RegistroReparacion 
from django.db.models.functions import Cast
from inventario.forms import RegistroReparacionForm



def index_view(request):
    """
    Vista raíz que redirige al usuario al lugar correcto.
    """
    
    # --- BLOQUE DE DEPURACIÓN ---
    print("--- DEBUG: Ejecutando index_view ---")
    
    # Comprobamos el estado de autenticación del usuario
    is_authenticated = request.user.is_authenticated
    print(f"Usuario autenticado: {is_authenticated}")
    
    # Si está autenticado, vemos quién es
    if is_authenticated:
        print(f"Nombre de usuario: {request.user.username}")
        # Redirigimos al dashboard
        print("➡️ Redirigiendo a 'proyectos:dashboard'")
        return redirect('proyectos:dashboard')
    else:
        # Si es anónimo, redirigimos al login
        print("➡️ Redirigiendo a 'login'")
        return redirect('login')
        
#PROYECTOS
@login_required 
def lista_proyectos(request):
    """
    Muestra la lista de proyectos con capacidades avanzadas de filtrado y ordenamiento.
    """
    # Empezamos con un queryset que tiene los cálculos que necesitaremos
    proyectos_qs = Proyecto.objects.annotate(
        num_subtareas_completadas=Count('tareas__subtareas', filter=Q(tareas__subtareas__completada=True)),
        num_personal_asignado=Count('tareas__subtareas__personal_asignado', distinct=True)
    )

    # --- 1. FILTRADO ---
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    date_range_filter = request.GET.get('date_range', '')

    # Filtro por nombre (buscador de texto)
    if query:
        proyectos_qs = proyectos_qs.filter(nombre__icontains=query)

    # Filtro por estado
    if status_filter == 'activo':
        proyectos_qs = proyectos_qs.filter(completada=False)
    elif status_filter == 'completado':
        proyectos_qs = proyectos_qs.filter(completada=True)

    # Filtro por Rango de Fechas
    today = timezone.now().date()
    if date_range_filter == 'end_this_month':
        end_of_month = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        proyectos_qs = proyectos_qs.filter(subtareas__fecha_fin__lte=end_of_month, subtareas__fecha_fin__gte=today)
    elif date_range_filter == 'start_next_month':
        next_month_start = (today.replace(day=1) + timedelta(days=32)).replace(day=1)
        next_month_end = (next_month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        proyectos_qs = proyectos_qs.filter(subtareas__fecha_inicio__range=[next_month_start, next_month_end])
    
    # --- 2. ORDENAMIENTO ---
    sort_by = request.GET.get('sort_by', '-id') # Por defecto, los más nuevos primero

    if sort_by == 'mas_completadas':
        proyectos_qs = proyectos_qs.order_by('-num_subtareas_completadas')
    elif sort_by == 'menos_completadas':
        proyectos_qs = proyectos_qs.order_by('num_subtareas_completadas')
    elif sort_by == 'mas_personal':
        proyectos_qs = proyectos_qs.order_by('-num_personal_asignado')
    else:
        proyectos_qs = proyectos_qs.order_by(sort_by)
        
    context = {
        'proyectos': proyectos_qs.distinct(), # .distinct() para evitar duplicados por los joins de los filtros
        'query': query,
        'status_filter': status_filter,
        'date_range_filter': date_range_filter,
        'sort_by': sort_by,
    }
    return render(request, 'proyectos/lista_proyectos.html', context)

@login_required 
@transaction.atomic 
def crear_proyecto(request):
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            nuevo_proyecto =form.save()
            return redirect('proyectos:asignar_tareas', proyecto_id=nuevo_proyecto.id)
    else:
        form = ProyectoForm()

    return render(request, 'proyectos/crear_proyecto.html', {'form': form})

@login_required 
@transaction.atomic # Asegura que si algo falla, todos los cambios se deshagan
def editar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    if request.method == 'POST':
        form = ProyectoEditForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, "Proyecto actualizado correctamente.")
            return redirect('proyectos:lista_proyectos')
    else:
        form = ProyectoEditForm(instance=proyecto)

    context = {
        'form': form,
        'proyecto': proyecto
    }
    return render(request, 'proyectos/editar_proyecto.html', context)

def detalle_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    # --- 1. OBTENER OPCIONES PARA LOS DESPLEGABLES DE FILTROS ---
    tareas_principales_opciones = TareaP.objects.filter(proyecto=proyecto)
    personal_opciones = Personal.objects.filter(
        subtareas_asignadas__tarea__proyecto=proyecto
    ).distinct()

    # --- 2. EMPEZAR CON TODAS LAS SUBTAREAS DEL PROYECTO ---
    # Este es nuestro queryset base. Le iremos añadiendo filtros.
    subtareas_qs = SubTarea.objects.filter(tarea__proyecto=proyecto)

    # --- 3. RECOGER Y APLICAR FILTROS DE LA URL ---
    # Recogemos todos los posibles filtros de los parámetros GET
    tarea_filter_id = request.GET.get('tarea_p')
    personal_filter_id = request.GET.get('personal')
    status_filter = request.GET.get('status')
    query_subtarea = request.GET.get('q_subtarea', '') # Buscador por nombre
    start_date_filter = request.GET.get('start_date') # Filtro por fecha inicio
    end_date_filter = request.GET.get('end_date') # Filtro por fecha fin

    # Aplicamos los filtros de forma secuencial al queryset
    if tarea_filter_id:
        subtareas_qs = subtareas_qs.filter(tarea__id=tarea_filter_id)
    if personal_filter_id:
        subtareas_qs = subtareas_qs.filter(personal_asignado__id=personal_filter_id)
    if status_filter == 'completada':
        subtareas_qs = subtareas_qs.filter(completada=True)
    elif status_filter == 'pendiente':
        subtareas_qs = subtareas_qs.filter(completada=False)
    if query_subtarea:
        subtareas_qs = subtareas_qs.filter(titulo__icontains=query_subtarea)
    if start_date_filter:
        subtareas_qs = subtareas_qs.filter(fecha_inicio__gte=start_date_filter)
    if end_date_filter:
        subtareas_qs = subtareas_qs.filter(fecha_fin__lte=end_date_filter)

    # --- 4. PREPARAR DATOS PARA LA PLANTILLA (AHORA USANDO EL QUERYSET FILTRADO) ---
    # ¡LA CORRECCIÓN CLAVE! Ahora trabajamos con 'subtareas_qs', que ya contiene solo los resultados filtrados.
    subtareas_ordenadas = subtareas_qs.order_by('fecha_inicio', 'fecha_fin').distinct()

    # Agrupamos las tareas filtradas por fecha de inicio
    tareas_agrupadas = defaultdict(list)
    for subtarea in subtareas_ordenadas:
        tareas_agrupadas[subtarea.fecha_inicio].append(subtarea)

    # Creamos la lista final para la línea de tiempo (con 'chunks' y 'delay')
    timeline_items = []
    delay_counter = 0
    for fecha, subtareas_del_dia in sorted(tareas_agrupadas.items()): # Usamos sorted() para ordenar los días
        for i in range(0, len(subtareas_del_dia), 2):
            chunk = subtareas_del_dia[i:i + 2]
            timeline_items.append({
                'fecha': fecha,
                'subtareas': chunk,
                'delay': delay_counter * 0.15
            })
            delay_counter += 1
            
    # --- 5. CONSTRUIR EL CONTEXTO ---
    context = {
        'proyecto': proyecto,
        'timeline_items': timeline_items,
        'tareas_principales_opciones': tareas_principales_opciones,
        'personal_opciones': personal_opciones,
        # Pasamos los valores de los filtros para mantener la selección en la plantilla
        'filtros_aplicados': {
            'tarea_p': tarea_filter_id,
            'personal': personal_filter_id,
            'status': status_filter,
            'q_subtarea': query_subtarea,
            'start_date': start_date_filter,
            'end_date': end_date_filter,
        }
    }
    return render(request, 'proyectos/detalle_proyecto.html', context)

def search_proyectos(request):
    term = request.GET.get('term', '').strip()
    term_id = request.GET.get('term_id')

    if term_id:
        proyectos_qs = Proyecto.objects.filter(pk=term_id)
    elif term:
        proyectos_qs = Proyecto.objects.filter(nombre__icontains=term)[:10]
    else:
        return JsonResponse({'results': []})

    proyectos_qs = Proyecto.objects.filter(nombre__icontains=term)[:10]

    results = [{"id": proyecto.id, "text": proyecto.nombre} for proyecto in proyectos_qs]
    
    return JsonResponse({'results': results})

@login_required 
@transaction.atomic
def asignar_tareas(request, proyecto_id):
    """
    Vista para la carga INICIAL de tareas. Se encarga de CREAR nuevos elementos
    y validar los datos antes de guardarlos en la base de datos.
    """
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    
    if request.method == 'POST':
        data = request.POST
        print("--- DEBUG: Datos brutos de request.POST ---")
        print(data)
        # --- CAPA 1: ESTRUCTURACIÓN DE DATOS Y VALIDACIÓN DE DUPLICADOS ---
        structured_data = defaultdict(lambda: {
            'id': None, 'titulo': '', 'subtareas': defaultdict(lambda: {
                'id': None, 'titulo': '', 'inicio': '', 'fin': '', 
                'insumos': defaultdict(dict), 'personal': []
            })
        })
        
        insumos_por_subtarea = defaultdict(set)
        has_errors = False

        for key, value in data.items():
            parts = key.split('_')
            if len(parts) < 3: continue

            item_type = parts[0]
            field = parts[1]

            try:
                if item_type == 'main' and field == 'task':
                    main_idx = '_'.join(parts[3:]) # Reconstruye el índice ej: "new_0"
                    structured_data[main_idx][parts[2]] = value
                
                elif item_type == 'subtask':
                    main_idx = parts[2]
                    # Reconstruimos el sub_idx completo (ej: "new_12345678")
                    sub_idx = '_'.join(parts[3:]) 
                    structured_data[main_idx]['subtareas'][sub_idx][field] = value

                elif item_type == 'insumo':
                    main_idx = parts[2]
                    sub_idx = parts[3]
                    insumo_idx = '_'.join(parts[4:])
                    structured_data[main_idx]['subtareas'][sub_idx]['insumos'][insumo_idx][field] = value

                elif item_type == 'personal' and field == 'asignado':
                    main_idx = parts[2]
                    sub_idx = parts[3]
                    structured_data[main_idx]['subtareas'][sub_idx]['personal'] = data.getlist(key)

            except (IndexError, ValueError):
                continue
        print("--- DEBUG: Datos estructurados del formulario (antes de guardar) ---")


    # Usamos pprint para una visualización bonita de diccionarios anidados
        import pprint
        pprint.pprint(dict(structured_data))
        # Si se encontraron errores, se devuelve el formulario con los datos y mensajes
        if has_errors:
            form_data_for_template = json.dumps(dict(structured_data))
            print(f"Valor de form_data_for_template en caso de error: {form_data_for_template}")
            return render(request, 'proyectos/asignar_tareas.html', {
                'proyecto': proyecto,
                'form_data': form_data_for_template
            })

        # --- CAPA 2: GUARDADO EN LA BASE DE DATOS (Solo si no hay errores) ---
        try:
            for main_idx, main_data in sorted(structured_data.items()):
                main_task_titulo = main_data.get('titulo')
                if not main_task_titulo: 
                    continue

                main_task_obj = TareaP.objects.create(proyecto=proyecto, titulo=main_task_titulo)
                print(f"--- DEBUG: Creando Tarea Principal: {main_data['titulo']} ---")
                for sub_idx, sub_data in sorted(main_data.get('subtareas', {}).items()):
                    sub_task_titulo = sub_data.get('titulo')
                    if not sub_task_titulo: 
                        continue
                    fecha_inicio_str = sub_data.get('inicio')
                    fecha_fin_str = sub_data.get('fin')
                    
                    # --- ¡VALIDACIÓN DE FECHAS MANUAL Y DIRECTA! ---
                    # 1. Nos aseguramos de que ambas fechas estén presentes
                    if fecha_inicio_str and fecha_fin_str:
                        # 2. Convertimos los strings de fecha (ej: '2025-09-30') a objetos de fecha de Python
                        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                        
                        # 3. Comparamos los objetos de fecha
                        if fecha_fin < fecha_inicio:
                            # Si la validación falla, lanzamos un error claro
                            raise ValueError(
                                f"Error en la subtarea '{sub_task_titulo}': La fecha de fin ({fecha_fin_str}) "
                                f"no puede ser anterior a la fecha de inicio ({fecha_inicio_str})."
                            )

                    sub_task_obj = SubTarea.objects.create(
                        tarea=main_task_obj,
                        titulo=sub_task_titulo,
                        fecha_inicio=sub_data.get('inicio'),
                        fecha_fin=sub_data.get('fin')
                    )
                    print(f"--- DEBUG: Creando Subtarea: {sub_data['titulo']} para {main_data['titulo']} ---")
                    # Procesa y asigna los Insumos para ESTA subtarea
                    for ins_idx, insumo_data in sorted(sub_data.get('insumos', {}).items()):
                        insumo_id = insumo_data.get('id')
                        cantidad_str = insumo_data.get('cantidad')
                        
                        if not insumo_id or not cantidad_str: 
                            continue
                        
                        stock_calculado = int(cantidad_str)
                        insumo_tipo_obj = get_object_or_404(Insumo, id=insumo_id)

                        items_disponibles = ItemInsumo.objects.filter(
                            insumo_padre=insumo_tipo_obj,
                            estado='EN STOCK'
                        )
                        stock_real = items_disponibles.count()

                        if stock_real < stock_calculado:
                            raise ValueError(f"No hay suficiente stock para '{insumo_tipo_obj.nombre}'. Se necesitan {stock_calculado}, pero solo hay {stock_real} disponibles.")
                        
                        items_a_asignar = items_disponibles[:stock_calculado]

                        for item in items_a_asignar:
                            AsignacionInsumo.objects.create(
                                subtarea=sub_task_obj,
                                item_insumo=item, 
                                stock_calculado=1,
                                costo_unitario_registrado=insumo_tipo_obj.costo_unitario_actual
                            )
                            item.estado = 'INSTALADO'
                            item.save()
                    print(f"--- DEBUG: Insumos para la subtarea {sub_data['titulo']}: {sub_data.get('insumos', {})}")

                    # Procesa y asigna el Personal para ESTA subtarea
                    personal_ids = sub_data.get('personal', [])
                    if personal_ids:
                        sub_task_obj.personal_asignado.set(personal_ids)
                    print(f"--- DEBUG: Personal para la subtarea {sub_data['titulo']}: {sub_data.get('personal', [])}")
        
        except (ValueError, Exception) as e:
            messages.error(request, str(e))
            form_data_for_template = json.dumps(dict(structured_data))
            print("--- DEBUG: Excepción capturada en la base de datos ---")
            print(e)
            print(f"Valor de form_data_for_template en caso de error: {form_data_for_template}")
            return render(request, 'proyectos/asignar_tareas.html', {
                'proyecto': proyecto,
                'form_data': form_data_for_template
            })

        messages.success(request, "Plan de trabajo inicial guardado correctamente.")
        return redirect('proyectos:detalle_proyecto', pk=proyecto_id)
    
    # Lógica para la petición GET (mostrar el formulario)
    else:
        # ¡LA CONSULTA CORREGIDA Y EFICIENTE!
        insumos_disponibles = Insumo.objects.annotate(
            # 1. Calcula el stock contando los items con estado 'EN STOCK'
            stock_actual=Count('items', filter=Q(items__estado='EN STOCK'))
        ).filter(
            # 2. Filtra por el resultado de ese cálculo
            stock_actual__gt=0
        )
        
        context = {
            'proyecto': proyecto,
            'insumos_disponibles': insumos_disponibles # Pasamos la nueva lista filtrada
        }
        return render(request, 'proyectos/asignar_tareas.html', context)

@login_required 
@transaction.atomic # Si algo falla dentro de esta vista, se deshacen todos los cambios en la BBDD.
def anadir_tareas(request, proyecto_id):
    """
    Vista dedicada exclusivamente a AÑADIR nuevas tareas, subtareas e insumos
    a un proyecto existente. No maneja lógica de edición ni eliminación.
    """
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    
    if request.method == 'POST':
        data = request.POST
        
        # --- CAPA 1: VALIDACIÓN PREVIA (ANTES DE TOCAR LA BASE DE DATOS) ---
        
        # 1.1: Validar duplicados de insumos en una misma subtarea
        insumos_por_subtarea = defaultdict(set)
        has_errors = False
        
        for key, value in data.items():
            if key.startswith('insumo_id_') and value:
                try:
                    parts = key.split('_')
                    main_index, sub_index = parts[2], parts[3]
                    subtarea_key = f"{main_index}_{sub_index}"
                    
                    if value in insumos_por_subtarea[subtarea_key]:
                        has_errors = True
                        main_task_titulo = data.get(f'main_task_titulo_{main_index}', 'esta tarea')
                        insumo = get_object_or_404(Insumo, pk=value)
                        messages.error(request, f'Error: El insumo "{insumo.nombre}" está duplicado en una subtarea de "{main_task_titulo}".')
                    else:
                        insumos_por_subtarea[subtarea_key].add(value)
                except (IndexError, ValueError):
                    continue

        for main_idx, main_data in structured_data.items():
            for sub_idx, sub_data in main_data.get('subtareas', {}).items():
                form = SubTareaForm(sub_data) # Creamos un form por cada subtarea
                if not form.is_valid():
                    has_errors = True
                    # Añadimos los errores del form a los mensajes
                    for field, error_list in form.errors.items():
                        for error in error_list:
                            messages.error(request, f"Error en '{main_data.get('titulo')}' / '{sub_data.get('titulo')}': {error}")

        if has_errors:
            structured_data = defaultdict(lambda: {'titulo': '', 'subtareas': defaultdict(lambda: {'titulo': '', 'inicio': '', 'fin': '', 'insumos': defaultdict(dict)})})
            
            for key, value in request.POST.items():
                parts = key.split('_')
                if key.startswith('main_task_titulo_'):
                    structured_data[parts[3]]['titulo'] = value
                elif key.startswith('subtask_'):
                    field, main_idx, sub_idx = parts[1], parts[2], parts[3]
                    structured_data[main_idx]['subtareas'][sub_idx][field] = value
                elif key.startswith('insumo_id_') and value:
                    main_idx, sub_idx, insumo_idx = parts[2], parts[3], parts[4]
                    insumo = get_object_or_404(Insumo, pk=value)
                    structured_data[main_idx]['subtareas'][sub_idx]['insumos'][insumo_idx]['id'] = value
                    # Añadimos el texto del insumo para que Select2 lo pueda mostrar
                    structured_data[main_idx]['subtareas'][sub_idx]['insumos'][insumo_idx]['text'] = insumo.nombre
                elif key.startswith('insumo_cantidad_'):
                    main_idx, sub_idx, insumo_idx = parts[2], parts[3], parts[4]
                    structured_data[main_idx]['subtareas'][sub_idx]['insumos'][insumo_idx]['cantidad'] = value

            form_data_for_template = json.dumps(dict(structured_data))
            return render(request, 'proyectos/asignar_tareas.html', {'proyecto': proyecto, 'form_data': form_data_for_template})

        for main_idx, main_data in structured_data.items():
            for sub_idx, sub_data in main_data.get('subtareas', {}).items():
                form = SubTareaForm(sub_data) # Creamos un form por cada subtarea
                if not form.is_valid():
                    has_errors = True
                    # Añadimos los errores del form a los mensajes
                    for field, error_list in form.errors.items():
                        for error in error_list:
                            messages.error(request, f"Error en '{main_data.get('titulo')}' / '{sub_data.get('titulo')}': {error}")

        # --- CAPA 2: ESTRUCTURACIÓN DE DATOS ---
        structured_data = defaultdict(lambda: {'titulo': '', 'subtareas': defaultdict(lambda: {'titulo': '', 'inicio': '', 'fin': '', 'insumos': defaultdict(dict)})})
        for key, value in data.items():
            parts = key.split('_')
            if len(parts) < 4: continue

            item_type, field_type = parts[0], parts[1]
            
            if item_type == 'main' and field_type == 'task':
                field, main_idx = parts[2], parts[3]
                structured_data[main_idx][field] = value
            elif item_type == 'subtask':
                field, main_idx, sub_idx = parts[1], parts[2], parts[3]
                structured_data[main_idx]['subtareas'][sub_idx][field] = value
            elif item_type == 'insumo':
                field, main_idx, sub_idx, ins_idx = parts[1], parts[2], parts[3], parts[4]
                if field == 'id': structured_data[main_idx]['subtareas'][sub_idx]['insumos'][ins_idx]['id'] = value
                if field == 'cantidad': structured_data[main_idx]['subtareas'][sub_idx]['insumos'][ins_idx]['cantidad'] = value


        # --- CAPA 3: GUARDADO EN LA BASE DE DATOS ---
        try:
            for main_idx, main_data in structured_data.items():
                main_task_titulo = main_data.get('titulo')
                if not main_task_titulo: continue

                # Siempre CREA una nueva Tarea Principal
                main_task_obj = TareaP.objects.create(proyecto=proyecto, titulo=main_task_titulo)

                for sub_idx, sub_data in main_data.get('subtareas', {}).items():
                    sub_task_titulo = sub_data.get('titulo')
                    if not sub_task_titulo: continue

                    # Siempre CREA una nueva Subtarea
                    sub_task_obj = SubTarea.objects.create(
                        tarea=main_task_obj,
                        titulo=sub_task_titulo,
                        fecha_inicio=sub_data.get('inicio'),
                        fecha_fin=sub_data.get('fin')
                    )
                    
                    # Procesar y CREAR asignaciones de insumos
                    for ins_idx, insumo_data in sub_data.get('insumos', {}).items():
                        insumo_id = insumo_data.get('id')
                        cantidad_str = insumo_data.get('cantidad')

                        if not insumo_id or not cantidad_str: continue
                        
                        stock_calculado = int(cantidad_str)
                        insumo_obj = get_object_or_404(Insumo, id=insumo_id)

                        # 1.2: Validar y restar stock
                        if insumo_obj.stock_disponible < stock_calculado:
                             # Lanzamos una excepción para que la transacción se deshaga
                             raise ValueError(f"No hay suficiente stock para '{insumo_obj.nombre}'. Disponible: {insumo_obj.stock_disponible}, Requerido: {stock_calculado}")
                        
                        insumo_obj.stock_disponible -= stock_calculado
                        insumo_obj.save()

                        # Crear el registro de asignación "congelando" el precio
                        AsignacionInsumo.objects.create(
                            subtarea=sub_task_obj,
                            insumo=insumo_obj,
                            stock_calculado=stock_calculado,
                            costo_unitario_registrado=insumo_obj.costo_unitario_actual
                        )
                    personal_ids = data.getlist(f'personal_asignado_{main_idx}_{sub_idx}')
                    if personal_ids:
                        sub_task_obj.personal_asignado.set(personal_ids)
                
        except ValueError as e:
            # Capturamos el error de stock (o cualquier otro ValueError)
            messages.error(request, str(e))
            # De nuevo, idealmente reconstruiríamos el formulario aquí
            return render(request, 'anadir_tareas.html', {'proyecto': proyecto})

        messages.success(request, "Nuevas tareas han sido añadidas al proyecto.")
        return redirect('proyectos:detalle_proyecto', pk=proyecto_id)
    
    # --- LÓGICA PARA CARGAR LA PÁGINA (MÉTODO GET) ---
    else:
        # Para esta vista, el formulario siempre empieza vacío.
        return render(request, 'proyectos/anadir_tareas.html', {'proyecto': proyecto})

def api_get_subtarea_data(request, pk):
    subtarea = get_object_or_404(SubTarea, pk=pk)

    personal_asignado = []
    for persona in subtarea.personal_asignado.all():
        personal_asignado.append({
            'id': persona.id,
            'text': f"{persona.nombre} {persona.apellido}"
        })

    data = {
        'id': subtarea.id,
        'titulo': subtarea.titulo,
        'fecha_inicio': subtarea.fecha_inicio,
        'fecha_fin': subtarea.fecha_fin,
        'personal_asignado': personal_asignado,
    }
    return JsonResponse(data)
@require_POST
@transaction.atomic # Usamos transacción para que si algo falla, se deshaga todo
def api_update_subtarea(request, pk):
    subtarea = get_object_or_404(SubTarea, pk=pk)
    
    original_fecha_inicio = subtarea.fecha_inicio
    original_fecha_fin = subtarea.fecha_fin

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)

    form_data = {k: v for k, v in data.items() if k in ['titulo', 'fecha_inicio', 'fecha_fin']}
    form = SubTareaEditForm(form_data, instance=subtarea)

    form = SubTareaEditForm(data, instance=subtarea)
    if not form.is_valid():
        return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)

    
    updated_subtarea = form.save()

    nueva_fecha_fin = updated_subtarea.fecha_fin

    if original_fecha_fin != nueva_fecha_fin:
        
        # 4. Calculamos la diferencia
        delta = nueva_fecha_fin - original_fecha_fin
        
        proyecto_actual = updated_subtarea.tarea.proyecto

        subtareas_a_mover = SubTarea.objects.filter(
            tarea__proyecto=proyecto_actual,
            fecha_inicio__gt=original_fecha_inicio 
        ).exclude(pk=updated_subtarea.pk)

        if subtareas_a_mover.exists():
            subtareas_a_mover.update(
                fecha_inicio = F('fecha_inicio') + delta,
                fecha_fin = F('fecha_fin') + delta
            )


    # 2. AÑADIR NUEVO PERSONAL ASIGNADO (Lógica Aditiva)
    personal_ids_a_anadir = data.get('personal_asignado', [])
    if personal_ids_a_anadir:
        # El método .add() simplemente añade las nuevas relaciones sin borrar las existentes
        updated_subtarea.personal_asignado.add(*personal_ids_a_anadir)
    
    # 3. AÑADIR NUEVOS INSUMOS Y RESTAR STOCK
    nuevos_insumos = data.get('nuevos_insumos', [])
    for item in nuevos_insumos:
        try:
            insumo_id = int(item['id'])
            cantidad = int(item['cantidad'])
            
            insumo = Insumo.objects.get(pk=insumo_id)
            
            # Verificar y restar stock
            if insumo.stock_disponible < cantidad:
                raise ValueError(f"No hay suficiente stock para '{insumo.nombre}'.")
            
            insumo.stock_disponible -= cantidad
            insumo.save()

            # Crear el nuevo registro de asignación
            AsignacionInsumo.objects.create(
                subtarea=updated_subtarea,
                insumo=insumo,
                stock_calculado=cantidad,
                costo_unitario_registrado=insumo.costo_unitario_actual
            )
        except (Insumo.DoesNotExist, ValueError, KeyError, TypeError):
            # Ignora insumos mal formados o con errores para no romper la operación
            continue

    # 4. PREPARAR Y DEVOLVER LA RESPUESTA DE ÉXITO
    # (Tu código para preparar la respuesta JSON era perfecto, lo reutilizamos)
    personal_data = [{'id': p.id, 'text': f"{p.nombre} {p.apellido}"} for p in updated_subtarea.personal_asignado.all()]
    
    return JsonResponse({
        'status': 'success',
        'message': 'Subtarea actualizada y planificación ajustada.',
        'subtarea': {
            'id': updated_subtarea.id,
            'titulo': updated_subtarea.titulo,
            'fecha_inicio': updated_subtarea.fecha_inicio.strftime('%d %b %Y'),
            'fecha_fin': updated_subtarea.fecha_fin.strftime('%d %b %Y'),
            'dia_semana': updated_subtarea.dia_semana(),
            'personal_asignado': personal_data,
        }
    })

@require_POST
@transaction.atomic
def api_delete_subtarea(request, pk):
    
    subtarea = get_object_or_404(SubTarea, pk=pk)
    
    # --- Guardamos una referencia a la Tarea Principal ANTES de borrar la subtarea ---
    tarea_principal = subtarea.tarea
    
    # --- Lógica de Reposición de Stock (sin cambios) ---
    for asignacion in subtarea.insumos_asignados.all():
        insumo = asignacion.insumo
        insumo.stock_disponible += asignacion.stock_calculado
        insumo.save()
        
    titulo_subtarea = subtarea.titulo
    
    subtarea.delete()
    
    if tarea_principal.subtareas.count() == 0:

        titulo_tarea_principal = tarea_principal.titulo
        tarea_principal.delete()

        return JsonResponse({
            'status': 'success',
            'task_deleted': True, # Le decimos al frontend que la tarea principal también se fue
            'message': f'Subtarea "{titulo_subtarea}" eliminada. La Tarea Principal "{titulo_tarea_principal}" también fue eliminada por estar vacía.'
        })
    return JsonResponse({
        'status': 'success',
        'task_deleted': False,
        'message': f'Subtarea "{titulo_subtarea}" eliminada y el stock ha sido repuesto.'
    })


@require_POST
@transaction.atomic
def api_toggle_complete_subtarea(request, pk):
    """
    Endpoint para marcar/desmarcar una subtarea como completada.
    Actualiza el estado de la Tarea Principal y del Proyecto en cascada.
    """
    subtarea = get_object_or_404(SubTarea, pk=pk)
    
    # 1. Cambia el estado de la subtarea
    subtarea.completada = not subtarea.completada
    subtarea.save()

    # 2. Comprueba y actualiza la Tarea Principal
    tarea_principal = subtarea.tarea
    # Si TODAS las subtareas de esta Tarea Principal están completadas
    if not tarea_principal.subtareas.filter(completada=False).exists():
        tarea_principal.completada = True
    else:
        tarea_principal.completada = False
    tarea_principal.save()
    
    # 3. Comprueba y actualiza el Proyecto
    proyecto = tarea_principal.proyecto
    # Si TODAS las Tareas Principales de este Proyecto están completadas
    if not proyecto.tareas.filter(completada=False).exists():
        proyecto.completada = True
    else:
        proyecto.completada = False
    proyecto.save()
    
    return JsonResponse({
        'status': 'success',
        'subtarea_completada': subtarea.completada,
        'message': 'Estado de la tarea actualizado.'
    })

@login_required 
def dashboard_view(request):
    """
    Calcula todos los KPIs y datos para los gráficos del dashboard.
    """
    # --- 1. KPIs Numéricos Principales ---
    proyectos_activos_qs = Proyecto.objects.filter(completada=False)
    proyectos_activos = proyectos_activos_qs.count()
    total_proyectos = Proyecto.objects.count()
    
    tareas_pendientes = SubTarea.objects.filter(completada=False, tarea__proyecto__in=proyectos_activos_qs).count()
    
    costo_total_estimado = AsignacionInsumo.objects.filter(
        subtarea__tarea__proyecto__completada=False
    ).aggregate(
        total=Sum('costo_unitario_registrado') 
    )['total'] or 0.00

    # KPI de Progreso General
    agregado_total = SubTarea.objects.filter(tarea__proyecto__completada=False).aggregate(
        total=Count('id'),
        completadas=Count('id', filter=Q(completada=True))
    )
    total_subtareas_activas = agregado_total.get('total', 0)
    subtareas_completadas_activas = agregado_total.get('completadas', 0)
    
    progreso_general = 0
    if total_subtareas_activas > 0:
        progreso_general = (subtareas_completadas_activas / total_subtareas_activas) * 100

    # --- 2. Datos para Gráfico de Dona (Estado de Proyectos) ---
    # ¡LA LÓGICA QUE FALTABA!
    proyectos_completados = total_proyectos - proyectos_activos
    project_status_data = {
        'labels': ['Activos', 'Completados'],
        'data': [proyectos_activos, proyectos_completados],
    }

    # --- 3. Datos para Gráfico de Barras (Carga de Trabajo) ---
    personal_workload = Personal.objects.filter(subtareas_asignadas__completada=False).annotate(
        num_tareas_pendientes=Count('subtareas_asignadas')
    ).order_by('-num_tareas_pendientes')[:10]
    
    workload_data = {
        'labels': [f"{p.nombre} {p.apellido}" for p in personal_workload],
        'data': [p.num_tareas_pendientes for p in personal_workload],
    }

    # --- 4. Datos para Gráfico de Líneas (Tendencia de Tareas) ---
    today = timezone.now().date()
    meses = []
    tareas_completadas_por_mes = []
    for i in range(5, -1, -1):
        target_month_start = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        next_month_start = (target_month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        meses.append(target_month_start.strftime("%b %Y"))
        
        count = SubTarea.objects.filter(
            completada=True,
            fecha_fin__gte=target_month_start,
            fecha_fin__lt=next_month_start
        ).count()
        tareas_completadas_por_mes.append(count)
        
    completion_trend_data = {'labels': meses, 'data': tareas_completadas_por_mes}

    # --- 5. Datos para Gráfico de Barras (Progreso por Proyecto) ---
    proyectos_con_progreso = proyectos_activos_qs.annotate(
        total_subtareas=Count('tareas__subtareas'),
        subtareas_completadas=Count('tareas__subtareas', filter=Q(tareas__subtareas__completada=True))
    ).annotate(
        progreso=Case(
            When(total_subtareas=0, then=0.0),
            default=(Cast('subtareas_completadas', FloatField()) * 100.0 / Cast('total_subtareas', FloatField())),
            output_field=FloatField()
        )
    ).order_by('-progreso')

    progress_data = {
        'labels': [p.nombre for p in proyectos_con_progreso],
        'data': [p.progreso for p in proyectos_con_progreso],
    }

    # --- Construcción del Contexto Final ---
    context = {
        'proyectos_activos': proyectos_activos,
        'tareas_pendientes': tareas_pendientes,
        'costo_total_estimado': costo_total_estimado,
        'progreso_general': progreso_general,
        'project_status_data': json.dumps(project_status_data),
        'workload_data': json.dumps(workload_data),
        'completion_trend_data': json.dumps(completion_trend_data),
        'progress_data': json.dumps(progress_data),
    }
    
    return render(request, 'proyectos/dashboard.html', context)

@login_required 
def detalle_subtarea(request, pk):
    subtarea = get_object_or_404(
        SubTarea.objects.select_related('tarea__proyecto')
        .prefetch_related('personal_asignado', 'insumos_asignados__insumo'),
        pk=pk
    )
    
    # Lógica para manejar el formulario de registro de reparaciones
    if request.method == 'POST':
        form = RegistroReparacionForm(request.POST)
        if form.is_valid():
            # Obtenemos el ID del ITEM específico del formulario
            item_insumo_id = request.POST.get('item_insumo_id')
            if item_insumo_id:
                item_insumo = get_object_or_404(ItemInsumo, pk=item_insumo_id)
                
                # Guardamos el registro de reparación
                nueva_reparacion = form.save(commit=False)
                nueva_reparacion.item_insumo = item_insumo
                nueva_reparacion.save()
                
                # (Bonus) Actualizamos el estado del item a "En Reparación"
                item_insumo.estado = 'EN REPARACION'
                item_insumo.save()

                messages.success(request, f"Reparación registrada para el insumo S/N: {item_insumo.numero_serie}.")
            else:
                messages.error(request, "Error: No se especificó un número de serie para la reparación.")
            
            return redirect('proyectos:detalle_subtarea', pk=subtarea.pk)
        else:
            # Si el form no es válido, mostramos los errores
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
            return redirect('proyectos:detalle_subtarea', pk=subtarea.pk)
    
    # La vista GET no necesita pasar el formulario, lo manejaremos con un modal
    context = {
        'subtarea': subtarea,
    }
    return render(request, 'proyectos/detalle_subtarea.html', context)
# en proyectos/views.py
from django.db import transaction
from inventario.models import AsignacionInsumo, ItemInsumo
# ... otros imports

@require_POST
@transaction.atomic # ¡Crucial para una operación tan destructiva!
def api_delete_proyecto(request, pk):
    """
    Endpoint de API para eliminar un proyecto.
    - Repone el stock de todos los insumos que estaban asignados al proyecto.
    - Elimina el proyecto y todos sus datos asociados.
    """
    proyecto = get_object_or_404(Proyecto, pk=pk)
    nombre_proyecto = proyecto.nombre

    try:
        # --- 1. LÓGICA DE REPOSICIÓN DE STOCK Y ESTADO ---
        # Buscamos todos los items específicos asignados a CUALQUIER subtarea de este proyecto.
        items_asignados = ItemInsumo.objects.filter(
            asignacioninsumo__subtarea__tarea__proyecto=proyecto,
            estado='INSTALADO'
        )

        # Actualizamos su estado de vuelta a "EN STOCK" en una sola consulta eficiente.
        items_asignados.update(estado='EN STOCK')
        
        # (Nota: La lógica de 'cantidad' se vuelve más simple. Al cambiar el estado,
        # la propiedad 'stock_calculado' del Insumo se actualizará automáticamente
        # la próxima vez que se calcule.)

        # --- 2. ELIMINACIÓN DEL PROYECTO ---
        # Ahora que hemos manejado la lógica de negocio, podemos eliminar de forma segura.
        # El CASCADE se encargará del resto (TareaP, SubTarea, AsignacionInsumo).
        proyecto.delete()
        
        return JsonResponse({
            'status': 'success',
            'message': f'El proyecto "{nombre_proyecto}", sus tareas y asignaciones han sido eliminados. El stock de los insumos ha sido repuesto.'
        })

    except Exception as e:
        # Captura cualquier error inesperado durante el proceso
        return JsonResponse({'status': 'error', 'message': f'Ocurrió un error inesperado: {str(e)}'}, status=500)
def vista_seguimiento(request):
    # Obtenemos todos los registros de actividad
    registros_list = RegistroActividad.objects.all().select_related('usuario', 'content_type')
    
    # Paginación para manejar grandes volúmenes de registros
    paginator = Paginator(registros_list, 25) # 25 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # --- Datos para Gráficos ---
    # Ejemplo: Actividad en los últimos 7 días
    today = timezone.now()
    last_7_days = today - timedelta(days=7)
    actividad_semanal = RegistroActividad.objects.filter(
        timestamp__gte=last_7_days
    ).extra(
        {'dia': "date(timestamp)"}
    ).values('dia').annotate(
        cantidad=Count('id')
    ).order_by('dia')

    chart_data = {
        'labels': [a['dia'].strftime('%a, %d') for a in actividad_semanal],
        'data': [a['cantidad'] for a in actividad_semanal],
    }

    context = {
        'page_obj': page_obj,
        'chart_data': json.dumps(chart_data),
    }
    return render(request, 'proyectos/vista_seguimiento.html', context)


TARGET_WIDTH = 2000 # Ancho deseado
TARGET_HEIGHT = 100
def exportar_proyecto_excel(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    # 1. PREPARAR DATOS
    tareas_principales = TareaP.objects.filter(proyecto=proyecto).prefetch_related('subtareas__personal_asignado')
    all_subtareas = SubTarea.objects.filter(tarea__in=tareas_principales)

    

    # Calcular el rango de fechas del proyecto completo
    min_start_date = min(st.fecha_inicio for st in all_subtareas)
    max_end_date = max(st.fecha_fin for st in all_subtareas)
    date_range = [min_start_date + timedelta(days=i) for i in range((max_end_date - min_start_date).days + 1)]

    # 2. CREAR WORKBOOK Y HOJA DE CÁLCULO
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "CalendarioProyecto"

    import locale
    try:
        # Esta es la configuración para sistemas Linux y macOS.
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except locale.Error:
        try:
            # Esta es la configuración alternativa.
            locale.setlocale(locale.LC_TIME, 'es_ES')
        except locale.Error:
            # Esta es la configuración para sistemas Windows.
            locale.setlocale(locale.LC_TIME, 'Spanish_Spain')

    import os
    from openpyxl.drawing.image import Image
    # La ruta base de tu proyecto
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        img_path = os.path.join(BASE_DIR, 'proyectos', 'static', 'logoCC.png')
        
        # 1. Cargamos la imagen con PIL
        pil_img = PilImage.open(img_path)
        
        # 2. Redimensionamos la imagen
        pil_img.thumbnail((TARGET_WIDTH, TARGET_HEIGHT), PilImage.Resampling.LANCZOS)
        
        # 3. Guardamos la imagen redimensionada en un objeto BytesIO en memoria
        resized_img_bytes = BytesIO()
        pil_img.save(resized_img_bytes, format='PNG')
        resized_img_bytes.seek(0)
        
        # 4. Creamos el objeto Image de openpyxl desde los datos en memoria
        logo = Image(resized_img_bytes)
        
        # 5. Añadimos la imagen redimensionada a la hoja
        sheet.add_image(logo, 'A1')
        
    except FileNotFoundError:
        print(f"Advertencia: No se encontró la imagen en {img_path}. La exportación continuará sin logo.")
    except Exception as e:
        print(f"Ocurrió un error al agregar la imagen: {e}")

    if not all_subtareas.exists():
        # Manejar caso sin subtareas
        messages.error(request, "El proyecto no tiene subtareas para generar un cronograma.")
        return redirect('proyectos:detalle_proyecto', pk=proyecto.pk)

    # 3. DEFINIR ESTILOS (para un look profesional)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
    tp_font = Font(bold=True)
    tp_fill = PatternFill(start_color="D2E5D2", end_color="D2E5D2", fill_type="solid") # Verde claro
    st_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Azul claro
    weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    gantt_bar_fills = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in ["8DB4E2", "C4D79B", "F79646", "B1A0C7"]] # Colores para las barras
    domingo_font = Font(bold=True, color="FF0000")
    # 4. CONSTRUIR ENCABEZADO DEL PROYECTO
    sheet.merge_cells('A1:A2')
    sheet['A1'] = ""
    sheet['B1'] = "Elaborado por:"
    sheet['C1'] = "Equipo de Telemetría"
    sheet['B2'] = "Inicio del proyecto:"
    sheet['C2'] = proyecto.fecha_inicio_calculada.strftime('%d %b %Y')
    sheet['A3'] = "PROYECTO:"
    sheet['B3'] = proyecto.nombre
    sheet.merge_cells('B3:D3')
    
    # 5. CONSTRUIR CABECERA DE LA TABLA Y GANTT
    main_headers = ["TAREA", "ASIGNADO A", "INICIO", "FIN"]
    gantt_start_col = len(main_headers) + 1
    
    for i, header in enumerate(main_headers, 1):
        cell = sheet.cell(row=5, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    current_month = None
    month_start_col = None

    for i, day in enumerate(date_range, gantt_start_col):
        # Fila 5: Letra inicial del día en español
        cell_day_letter = sheet.cell(row=5, column=i, value=day.strftime('%A')[0].upper())
        
        # Apply standard header styles first
        cell_day_letter.alignment = Alignment(horizontal='center')
        cell_day_letter.fill = header_fill
        cell_day_letter.font = header_font
        
        # Now, check for Sunday and override the font if necessary
        if cell_day_letter.value == 'D':
            cell_day_letter.font = domingo_font
            
        # Check for weekends to change fill color
        
        
        # Fila 4: Día numérico
        sheet.cell(row=4, column=i, value=day.day).alignment = Alignment(horizontal='center')
        
        # Lógica para la fila 3: Mes
        if day.month != current_month:
            # Si es un nuevo mes, fusiona las celdas del mes anterior
            if current_month is not None:
                sheet.merge_cells(start_row=3, start_column=month_start_col, end_row=3, end_column=i-1)
                
            # Configura el nuevo mes
            current_month = day.month
            month_start_col = i
            
            # Obtén el nombre completo del mes en español
            # %B devuelve el nombre completo del mes en el locale actual (ej. "septiembre")
            month_name = day.strftime('%B').capitalize()
            
            # Fila 3: Establece el nombre del mes
            cell_month = sheet.cell(row=3, column=i, value=month_name)
            cell_month.alignment = Alignment(horizontal='center')
            cell_month.fill = header_fill
            cell_month.font = header_font
            
    # Finalmente, fusiona las celdas para el último mes
    if current_month is not None:
        sheet.merge_cells(start_row=3, start_column=month_start_col, end_row=3, end_column=gantt_start_col + len(date_range) - 1)
    
    # 6. LLENAR LOS DATOS Y DIBUJAR EL GANTT
    current_row = 6
    gantt_fill_cycle = 0
    numero= 1
    for tarea_p in tareas_principales:
        # Fila de la Tarea Principal
        
        cell = sheet.cell(row=current_row, column=1, value=f"{numero}. {tarea_p.titulo.upper()}")
        cell.font = tp_font
        for i in range(1, 5):
            sheet.cell(row=current_row, column=i).fill = tp_fill
        current_row += 1

        # Filas de las Subtareas
        for subtarea in tarea_p.subtareas.all():
            personal_nombres = ", ".join([p.nombre for p in subtarea.personal_asignado.all()])
            sheet.cell(row=current_row, column=1, value=subtarea.titulo)
            sheet.cell(row=current_row, column=2, value=personal_nombres)
            sheet.cell(row=current_row, column=3, value=subtarea.fecha_inicio.strftime('%d/%m/%Y'))
            sheet.cell(row=current_row, column=4, value=subtarea.fecha_fin.strftime('%d/%m/%Y'))
            
            # Dibujar la barra del Gantt
            if subtarea.fecha_inicio and subtarea.fecha_fin:
                for i, date in enumerate(date_range, gantt_start_col):
                    if subtarea.fecha_inicio <= date <= subtarea.fecha_fin:
                        sheet.cell(row=current_row, column=i).fill = gantt_bar_fills[gantt_fill_cycle % len(gantt_bar_fills)]

            for i in range(1, gantt_start_col):
                sheet.cell(row=current_row, column=i).fill = st_fill

            current_row += 1
        numero += 1
        gantt_fill_cycle += 1

    # Ajustar ancho de columnas
    sheet.row_dimensions[1].height = 50
    sheet.row_dimensions[2].height = 33
    sheet.column_dimensions['A'].width = 38
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12 

    # 7. DEVOLVER EL ARCHIVO EN LA RESPUESTA HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Cronograma - {proyecto.nombre}.xlsx"'
    workbook.save(response)
    
    return response