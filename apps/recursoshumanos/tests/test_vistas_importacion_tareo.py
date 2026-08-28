"""Pruebas de las vistas de importación de tareo.

Cubren dos fallos vistos en uso real:

* recargar la previsualización (F5) devolvía un 405 en blanco, porque la URL
  solo aceptaba POST y el navegador la pedía por GET;
* la sospecha de que importar cerraba la sesión del usuario. No es así: en
  esta instalación `SESION_UNICA_EXIMIR_SUPERUSUARIO=0`, y lo que cerraba
  sesiones era iniciar sesión con la misma cuenta desde otro sitio. Esta
  prueba deja fijado que la importación no la toca.
"""

import io
import json
from datetime import date

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse

from accesos.models import SesionCerradaRemotamente
from recursoshumanos.models import (Area, AsignacionProyecto, Proyecto,
                                    TareoDiario, Trabajador)


def libro_de_una_semana():
    """Un bloque semanal de agosto de 2026 con una sola persona marcada."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026'
    ws.cell(2, 5, 'AGOSTO')
    ws.cell(4, 3, 'POSICIÓN')
    ws.cell(4, 4, 'NOMBRE')
    for i, numero in enumerate([3, 4, 5, 6, 7, 8, 9]):
        ws.cell(3, 5 + i, numero)
    ws.cell(5, 3, 'M1')
    ws.cell(5, 4, 'DIEGO HERNANI')
    for i in range(3):
        ws.cell(5, 5 + i, 'M1')

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class ImportacionTareoVistaTests(TestCase):
    def setUp(self):
        # El grupo ya lo crea una migración del proyecto.
        grupo, _ = Group.objects.get_or_create(name='Recursos Humanos')
        self.usuario = get_user_model().objects.create_user(
            username='rrhh', password='clave-de-prueba')
        self.usuario.groups.add(grupo)

        self.area = Area.objects.create(nombre='Operaciones QA')
        self.trabajador = Trabajador.objects.create(
            dni='10000001', nombres='DIEGO ARMANDO', apellido_paterno='HERNANI',
            apellido_materno='ROJAS', area=self.area, activo=True)

        self.url_importar = reverse('recursoshumanos:importar_tareo')
        self.url_tareo = reverse('recursoshumanos:gestion_tareo')
        self.client.login(username='rrhh', password='clave-de-prueba')

    def _subir(self):
        archivo = SimpleUploadedFile(
            'tareo.xlsx', libro_de_una_semana(),
            content_type=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'))
        return self.client.post(self.url_importar, {
            'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
            'area': str(self.area.id), 'archivo': archivo,
        })

    # ---- recarga de la previsualización ----

    def test_get_a_importar_redirige_al_tareo_en_vez_de_dar_405(self):
        respuesta = self.client.get(
            self.url_importar, {'mes': '2026-08', 'area': str(self.area.id)})

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(self.url_tareo, respuesta['Location'])
        self.assertIn('mes=2026-08', respuesta['Location'])
        self.assertIn(f'area={self.area.id}', respuesta['Location'])

    def test_get_a_confirmar_redirige_al_tareo_en_vez_de_dar_405(self):
        respuesta = self.client.get(
            reverse('recursoshumanos:importar_tareo_confirmar'), {'mes': '2026-08'})

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(self.url_tareo, respuesta['Location'])

    def test_get_sin_mes_no_revienta_y_vuelve_al_tareo(self):
        respuesta = self.client.get(self.url_importar)

        self.assertEqual(respuesta.status_code, 302)
        self.assertIn(self.url_tareo, respuesta['Location'])

    # ---- la previsualización mantiene la estructura de la matriz ----

    def test_la_previsualizacion_conserva_la_cabecera_de_la_matriz(self):
        """El título y los cinco filtros con sus dos botones son los mismos
        que en la matriz: la previsualización no cambia de página, solo cambia
        lo que se ve dentro de ella."""
        respuesta = self._subir()
        contenido = respuesta.content.decode()

        self.assertIn('Planificación Matricial', contenido)
        for control in ('id="buscadorTrabajador"', 'id="mes"', 'id="proyecto"',
                        'id="subproyecto"', 'id="area"',
                        'id="filtrarBtn"', 'id="limpiarFiltrosBtn"'):
            self.assertIn(control, contenido)

    def test_los_filtros_reciben_area_y_proyectos_de_cada_trabajador(self):
        """Los filtros de la previsualización corren en el navegador, así que
        necesitan por DNI lo mismo que filtra la matriz. Quien está asignado a
        un subproyecto también cuenta como del proyecto padre."""
        padre = Proyecto.objects.create(nombre='Monitoreo QA')
        hijo = Proyecto.objects.create(nombre='Monitoreo QA - Fase 2', parent=padre)
        AsignacionProyecto.objects.create(
            trabajador=self.trabajador, proyecto=hijo, activo=True)

        respuesta = self._subir()
        meta = respuesta.context['meta_trabajadores'][self.trabajador.dni]

        self.assertEqual(meta['a'], str(self.area.id))
        self.assertEqual(set(meta['p']), {str(hijo.id), str(padre.id)})

    # ---- la importación no toca la sesión ----

    def test_previsualizar_no_cierra_la_sesion(self):
        clave_antes = self.client.session.session_key

        respuesta = self._subir()

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(self.client.session.session_key, clave_antes)
        self.assertEqual(SesionCerradaRemotamente.objects.count(), 0)
        self.assertEqual(
            int(self.client.session['_auth_user_id']), self.usuario.pk)

    def test_confirmar_escribe_el_tareo_y_mantiene_la_sesion(self):
        previa = self._subir()
        fila = previa.context['filas'][0]
        clave_antes = self.client.session.session_key

        respuesta = self.client.post(
            reverse('recursoshumanos:importar_tareo_confirmar'), {
                'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
                'area': str(self.area.id),
                'incluir_0': 'on', 'dni_0': self.trabajador.dni,
                'datos_0': fila['datos_json'],
            })

        self.assertEqual(respuesta.status_code, 302)
        self.assertEqual(self.client.session.session_key, clave_antes)
        self.assertEqual(SesionCerradaRemotamente.objects.count(), 0)

        dias = TareoDiario.objects.filter(trabajador=self.trabajador)
        self.assertEqual(dias.count(), 3)
        self.assertEqual(sorted(d.fecha for d in dias),
                         [date(2026, 8, 3), date(2026, 8, 4), date(2026, 8, 5)])
        # Personal en campo entra como H (12 h, sin horario fijo).
        self.assertEqual({d.estado for d in dias}, {'H'})
        for dia in dias:
            self.assertEqual(float(dia.jornada_horas), 12.0)
            self.assertIsNone(dia.hora_entrada)


def libro_con_nombres(nombres):
    """Un bloque de agosto de 2026 con las personas indicadas, 3 días marcados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026'
    ws.cell(2, 5, 'AGOSTO')
    ws.cell(4, 3, 'POSICIÓN')
    ws.cell(4, 4, 'NOMBRE')
    for i, numero in enumerate([3, 4, 5, 6, 7, 8, 9]):
        ws.cell(3, 5 + i, numero)
    for orden, nombre in enumerate(nombres):
        fila = 5 + orden
        ws.cell(fila, 3, f'M{orden + 1}')
        ws.cell(fila, 4, nombre)
        for i in range(3):
            ws.cell(fila, 5 + i, f'M{orden + 1}')

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class AlertasDeFilaTests(TestCase):
    """La previsualización tiene que explicarle a RRHH qué decidir y por qué.

    El Excel no trae DNI: cuando el nombre no resuelve solo, la fila lleva una
    alerta con lo que dice literalmente el archivo, en vez de desaparecer o
    importarse a ciegas.
    """

    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name='Recursos Humanos')
        self.usuario = get_user_model().objects.create_user(
            username='rrhh2', password='clave-de-prueba')
        self.usuario.groups.add(grupo)
        self.area = Area.objects.create(nombre='Operaciones QA')
        Trabajador.objects.create(
            dni='10000001', nombres='DIEGO ARMANDO', apellido_paterno='HERNANI',
            apellido_materno='ROJAS', area=self.area, activo=True)
        self.client.login(username='rrhh2', password='clave-de-prueba')

    def _subir(self, nombres):
        archivo = SimpleUploadedFile(
            'tareo.xlsx', libro_con_nombres(nombres),
            content_type=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'))
        return self.client.post(reverse('recursoshumanos:importar_tareo'), {
            'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
            'area': str(self.area.id), 'archivo': archivo,
        })

    def _fila(self, respuesta, nombre_excel):
        return next(f for f in respuesta.context['filas']
                    if f['nombre_excel'] == nombre_excel)

    def test_nombre_desconocido_avisa_y_no_desaparece(self):
        respuesta = self._subir(['PERSONA INEXISTENTE'])

        fila = self._fila(respuesta, 'PERSONA INEXISTENTE')
        self.assertEqual(fila['alerta']['tipo'], 'sin_match')
        self.assertIn('PERSONA INEXISTENTE', fila['alerta']['detalle'])
        self.assertEqual(respuesta.context['total_por_revisar'], 1)
        self.assertContains(respuesta, 'data-alerta-tipo="sin_match"')

    def test_nota_entre_parentesis_se_menciona_en_la_alerta(self):
        # La nota se ignora al emparejar, así que el nombre coincide igual; pero
        # una nota que NO es vacaciones no cambia el estado sola: la fila entra
        # como campo y el aviso cita la nota literal para que RRHH decida.
        respuesta = self._subir(['DIEGO HERNANI (soporte)'])

        fila = self._fila(respuesta, 'DIEGO HERNANI (soporte)')
        self.assertEqual(fila['alerta']['tipo'], 'nota')
        self.assertIn('«soporte»', fila['alerta']['detalle'])
        self.assertIn('desmarca la fila', fila['alerta']['detalle'])

    def test_fila_normal_no_lleva_alerta(self):
        respuesta = self._subir(['DIEGO HERNANI'])

        self.assertIsNone(self._fila(respuesta, 'DIEGO HERNANI')['alerta'])
        self.assertEqual(respuesta.context['total_por_revisar'], 0)
        self.assertNotContains(respuesta, 'data-alerta-tipo')


def libro_con_secciones():
    """Agosto de 2026: una fila de campo, una del turno Vallecito y una de una
    sección desconocida (Soporte/Optalert)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026'
    ws.cell(2, 5, 'AGOSTO')
    ws.cell(4, 2, 'PERSONAL EN CAMPO')
    ws.merge_cells(start_row=4, end_row=5, start_column=2, end_column=2)
    ws.cell(6, 2, 'PERSONAL VALLECITO\nFINALIZACIÓN DE CADENAS DE CUSTODIA')
    ws.merge_cells(start_row=6, end_row=7, start_column=2, end_column=2)
    ws.cell(8, 2, 'SOPORTE MONITOREO - OPTALERT')
    ws.cell(4, 3, 'POSICIÓN')
    ws.cell(4, 4, 'NOMBRE')
    for i, numero in enumerate([3, 4, 5, 6, 7, 8, 9]):
        ws.cell(3, 5 + i, numero)

    ws.cell(5, 3, 'M1')
    ws.cell(5, 4, 'DIEGO HERNANI')
    for i in range(3):
        ws.cell(5, 5 + i, 'M1')

    ws.cell(6, 3, 'M17')
    ws.cell(6, 4, 'JULIO HUAMAN')
    for i in (0, 1, 4):
        ws.cell(6, 5 + i, 'M17')

    ws.cell(8, 3, 'SOP')
    ws.cell(8, 4, 'YERSON MOLLO')
    for i in (0, 1):
        ws.cell(8, 5 + i, 'YERSON')

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def libro_con_vacaciones():
    """Agosto 2026, una semana. Una fila normal (H) y la fila aparte de
    vacaciones ("WASHINGTON (vacaciones)"), ambas bajo PERSONAL EN CAMPO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026'
    ws.cell(2, 5, 'AGOSTO')
    ws.cell(4, 2, 'PERSONAL EN CAMPO')
    ws.merge_cells(start_row=4, end_row=6, start_column=2, end_column=2)
    ws.cell(4, 3, 'POSICIÓN')
    ws.cell(4, 4, 'NOMBRE')
    for i, numero in enumerate([3, 4, 5, 6, 7, 8, 9]):
        ws.cell(3, 5 + i, numero)

    ws.cell(5, 3, 'M1')
    ws.cell(5, 4, 'DIEGO HERNANI')
    for i in range(3):
        ws.cell(5, 5 + i, 'M1')

    ws.cell(6, 3, 'AS SUP')
    ws.cell(6, 4, 'WASHINGTON (vacaciones)')
    for i in range(2):
        ws.cell(6, 5 + i, 'X')

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def libro_persona_mixta():
    """Agosto de 2026, dos semanas. La MISMA persona (EDWIN PUMA, M12) cae bajo
    PERSONAL EN CAMPO la primera semana (varios días → H) y bajo PERSONAL
    VALLECITO la segunda (un día → P), porque el borde de la banda Vallecito se
    corre de una semana a otra. La fila debe salir con tipo dominante H —nunca
    'mixto'— y cada día conserva su estado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026'
    ws.cell(2, 5, 'AGOSTO')

    # --- Bloque 1 (semana Ago 3-9): rótulo col 2, POSICIÓN col 3, días 5-11 ---
    ws.cell(4, 2, 'PERSONAL EN CAMPO')
    ws.merge_cells(start_row=4, end_row=6, start_column=2, end_column=2)
    ws.cell(4, 3, 'POSICIÓN')
    ws.cell(4, 4, 'NOMBRE')
    for i, numero in enumerate([3, 4, 5, 6, 7, 8, 9]):
        ws.cell(3, 5 + i, numero)
    ws.cell(5, 3, 'M12')
    ws.cell(5, 4, 'EDWIN PUMA')
    for i in range(5):           # Ago 3,4,5,6,7 → 5 días de campo (H)
        ws.cell(5, 5 + i, 'M12')

    # --- Bloque 2 (semana Ago 10-16): rótulo col 12, POSICIÓN col 13, días 15-21 ---
    ws.cell(4, 12, 'PERSONAL EN CAMPO')
    ws.merge_cells(start_row=4, end_row=5, start_column=12, end_column=12)
    ws.cell(6, 12, 'PERSONAL VALLECITO\nFINALIZACIÓN DE CADENAS DE CUSTODIA')
    ws.merge_cells(start_row=6, end_row=7, start_column=12, end_column=12)
    ws.cell(4, 13, 'POSICIÓN')
    ws.cell(4, 14, 'NOMBRE')
    for i, numero in enumerate([10, 11, 12, 13, 14, 15, 16]):
        ws.cell(3, 15 + i, numero)
    ws.cell(6, 13, 'M12')
    ws.cell(6, 14, 'EDWIN PUMA')
    ws.cell(6, 15, 'M12')        # Ago 10 → 1 día Vallecito (P 13:00-21:00)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class PersonalDeOtraSeccionTests(TestCase):
    """El personal que el Excel pone fuera de PERSONAL EN CAMPO se muestra.

    Antes se descartaba en silencio, y como las hojas de junio en adelante ya no
    traen la columna de rótulo, la misma persona entraba o no según el mes.
    Ahora llega marcada y sin asignar: la decisión es de RRHH.
    """

    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name='Recursos Humanos')
        u = get_user_model().objects.create_user(username='rrhh3', password='clave-de-prueba')
        u.groups.add(grupo)
        self.area = Area.objects.create(nombre='Operaciones QA')
        Trabajador.objects.create(
            dni='10000001', nombres='DIEGO ARMANDO', apellido_paterno='HERNANI',
            apellido_materno='ROJAS', area=self.area, activo=True)
        Trabajador.objects.create(
            dni='10000005', nombres='JULIO', apellido_paterno='HUAMAN',
            apellido_materno='CONDORI', area=self.area, activo=True)
        Trabajador.objects.create(
            dni='10000006', nombres='YERSON', apellido_paterno='MOLLO',
            apellido_materno='FLORES', area=self.area, activo=True)
        self.client.login(username='rrhh3', password='clave-de-prueba')

        archivo = SimpleUploadedFile(
            'tareo.xlsx', libro_con_secciones(),
            content_type=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'))
        self.respuesta = self.client.post(
            reverse('recursoshumanos:importar_tareo'), {
                'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
                'area': str(self.area.id), 'archivo': archivo,
            })
        self.filas = {f['nombre_excel']: f for f in self.respuesta.context['filas']}

    def test_la_fila_de_campo_entra_normal(self):
        fila = self.filas['DIEGO HERNANI']
        self.assertTrue(fila['asignada'])
        self.assertFalse(fila['otra_seccion'])
        self.assertIsNone(fila['alerta'])

    def test_la_fila_de_campo_entra_como_h(self):
        fila = self.filas['DIEGO HERNANI']
        self.assertEqual(fila['tipo_inicial'], 'H')
        self.assertEqual({d['estado'] for d in fila['dias'].values()}, {'H'})

    def test_la_fila_de_vallecito_entra_como_personalizado(self):
        fila = self.filas['JULIO HUAMAN']
        # Vallecito es un turno conocido: entra ya resuelto como P 13:00-21:00,
        # sin marca de revisión ni alerta.
        self.assertFalse(fila['otra_seccion'])
        self.assertTrue(fila['asignada'])
        self.assertIsNone(fila['alerta'])
        self.assertEqual(fila['tipo_inicial'], 'P')
        self.assertEqual({d['estado'] for d in fila['dias'].values()}, {'P'})
        self.assertEqual(fila['horario_inicial'], {'entrada': '13:00', 'salida': '21:00'})

    def test_la_fila_de_seccion_desconocida_entra_como_dia_libre_marcada(self):
        fila = self.filas['YERSON MOLLO']
        self.assertTrue(fila['otra_seccion'])
        self.assertTrue(fila['asignada'])
        # El Excel no dice qué turno es esta sección, así que NO se da por bueno
        # trabajo de campo: entra como día libre y RRHH cambia el tipo.
        self.assertEqual(fila['tipo_inicial'], 'D')
        self.assertEqual({d['estado'] for d in fila['dias'].values()}, {'D'})

    def test_la_alerta_dice_la_seccion_el_nombre_y_los_dias(self):
        alerta = self.filas['YERSON MOLLO']['alerta']
        self.assertEqual(alerta['tipo'], 'otra_seccion')
        self.assertIn('OPTALERT', alerta['titulo'])
        self.assertIn('YERSON MOLLO', alerta['detalle'])
        self.assertIn('MOLLO FLORES, YERSON', alerta['detalle'])
        self.assertIn('3 y 4 de Agosto de 2026', alerta['detalle'])

    def test_el_resumen_cuenta_las_filas_de_otra_seccion(self):
        # Solo la sección desconocida cuenta; Vallecito ya no.
        self.assertEqual(self.respuesta.context['total_otra_seccion'], 1)
        self.assertContains(self.respuesta, 'data-alerta-tipo="otra_seccion"')
        self.assertContains(self.respuesta, 'trabajador-col sticky-col otra-seccion')


class FilaSinMixtoTests(TestCase):
    """Una persona con días de dos secciones (campo H + Vallecito P) ya no sale
    'mixto': la fila toma el tipo dominante y cada día conserva el suyo."""

    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name='Recursos Humanos')
        u = get_user_model().objects.create_user(username='rrhh_mix', password='clave-de-prueba')
        u.groups.add(grupo)
        self.area = Area.objects.create(nombre='Operaciones QA')
        Trabajador.objects.create(
            dni='10000012', nombres='EDWIN', apellido_paterno='PUMA',
            apellido_materno='QUISPE', area=self.area, activo=True)
        self.client.login(username='rrhh_mix', password='clave-de-prueba')

        archivo = SimpleUploadedFile(
            'tareo.xlsx', libro_persona_mixta(),
            content_type=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'))
        self.respuesta = self.client.post(
            reverse('recursoshumanos:importar_tareo'), {
                'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
                'area': str(self.area.id), 'archivo': archivo,
            })
        self.filas = {f['nombre_excel']: f for f in self.respuesta.context['filas']}

    def test_el_tipo_de_la_fila_es_el_dominante_no_mixto(self):
        fila = self.filas['EDWIN PUMA']
        # 5 días de campo contra 1 de Vallecito: gana H, nunca queda ''.
        self.assertEqual(fila['tipo_inicial'], 'H')

    def test_cada_dia_conserva_su_estado(self):
        fila = self.filas['EDWIN PUMA']
        estados = {d['estado'] for d in fila['dias'].values()}
        self.assertEqual(estados, {'H', 'P'})

    def test_el_dia_vallecito_trae_su_horario(self):
        fila = self.filas['EDWIN PUMA']
        self.assertEqual(fila['horario_inicial'],
                         {'entrada': '13:00', 'salida': '21:00'})

    def test_la_previsualizacion_ya_no_ofrece_la_opcion_mixto(self):
        self.assertNotContains(self.respuesta, 'Mixto')


class FilaDeVacacionesTests(TestCase):
    """"NOMBRE (vacaciones)" entra como Día Libre (D) ya resuelto: sin alerta,
    con el tipo de fila en D."""

    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name='Recursos Humanos')
        u = get_user_model().objects.create_user(username='rrhh_vac', password='clave-de-prueba')
        u.groups.add(grupo)
        self.area = Area.objects.create(nombre='Operaciones QA')
        Trabajador.objects.create(
            dni='10000001', nombres='DIEGO ARMANDO', apellido_paterno='HERNANI',
            apellido_materno='ROJAS', area=self.area, activo=True)
        Trabajador.objects.create(
            dni='10000009', nombres='WASHINGTON', apellido_paterno='PALOMINO',
            apellido_materno='RAMOS', area=self.area, activo=True)
        self.client.login(username='rrhh_vac', password='clave-de-prueba')

        archivo = SimpleUploadedFile(
            'tareo.xlsx', libro_con_vacaciones(),
            content_type=('application/vnd.openxmlformats-officedocument'
                          '.spreadsheetml.sheet'))
        self.respuesta = self.client.post(
            reverse('recursoshumanos:importar_tareo'), {
                'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
                'area': str(self.area.id), 'archivo': archivo,
            })
        self.filas = {f['nombre_excel']: f for f in self.respuesta.context['filas']}

    def test_los_dias_de_vacaciones_entran_como_dia_libre(self):
        fila = self.filas['WASHINGTON (vacaciones)']
        self.assertEqual(fila['tipo_inicial'], 'D')
        self.assertEqual({d['estado'] for d in fila['dias'].values()}, {'D'})

    def test_la_fila_de_vacaciones_no_lleva_alerta_ni_marca_de_seccion(self):
        fila = self.filas['WASHINGTON (vacaciones)']
        self.assertIsNone(fila['alerta'])
        self.assertFalse(fila['otra_seccion'])


class TiposDeJornadaAlConfirmarTests(TestCase):
    """RRHH puede reclasificar la fila antes de importar.

    El Excel solo sabe de campo; el tipo real (oficina, personalizado, jornada
    por horas, libre) lo decide RRHH en la previsualización, y P y J pueden
    llevar sus horas desde ahí.
    """

    def setUp(self):
        grupo, _ = Group.objects.get_or_create(name='Recursos Humanos')
        u = get_user_model().objects.create_user(username='rrhh4', password='clave-de-prueba')
        u.groups.add(grupo)
        self.area = Area.objects.create(nombre='Operaciones QA')
        self.trabajador = Trabajador.objects.create(
            dni='10000001', nombres='DIEGO ARMANDO', apellido_paterno='HERNANI',
            apellido_materno='ROJAS', area=self.area, activo=True)
        self.client.login(username='rrhh4', password='clave-de-prueba')
        self.url = reverse('recursoshumanos:importar_tareo_confirmar')

    def _confirmar(self, estado, extra=None):
        datos = json.dumps({'3': {'e': estado, 'u': None},
                            '4': {'e': estado, 'u': None}})
        campos = {
            'mes': '2026-08', 'q': '', 'proyecto': '', 'subproyecto': '',
            'area': str(self.area.id),
            'incluir_0': 'on', 'dni_0': self.trabajador.dni, 'datos_0': datos,
        }
        campos.update(extra or {})
        return self.client.post(self.url, campos)

    def test_dia_libre_se_guarda_como_libre(self):
        self._confirmar('D')
        dias = TareoDiario.objects.filter(trabajador=self.trabajador)
        self.assertEqual(dias.count(), 2)
        self.assertEqual({d.estado for d in dias}, {'D'})

    def test_oficina_toma_el_horario_de_oficina(self):
        self._confirmar('O')
        # 3 y 4 de agosto de 2026 son lunes y martes: 08:30 a 18:00.
        for dia in TareoDiario.objects.filter(trabajador=self.trabajador):
            self.assertEqual(dia.estado, 'O')
            self.assertEqual(dia.hora_entrada.strftime('%H:%M'), '08:30')
            self.assertEqual(dia.hora_salida.strftime('%H:%M'), '18:00')

    def test_personalizado_guarda_las_horas_que_escribio_rrhh(self):
        self._confirmar('P', {'hora_entrada_0': '07:15', 'hora_salida_0': '15:45'})
        for dia in TareoDiario.objects.filter(trabajador=self.trabajador):
            self.assertEqual(dia.estado, 'P')
            self.assertEqual(dia.hora_entrada.strftime('%H:%M'), '07:15')
            self.assertEqual(dia.hora_salida.strftime('%H:%M'), '15:45')

    def test_jornada_por_horas_guarda_las_horas(self):
        self._confirmar('J', {'jornada_horas_0': '6.5'})
        for dia in TareoDiario.objects.filter(trabajador=self.trabajador):
            self.assertEqual(dia.estado, 'J')
            self.assertEqual(float(dia.jornada_horas), 6.5)

    def test_campo_h_guarda_jornada_de_12_horas_sin_horario_fijo(self):
        self._confirmar('H')
        dias = TareoDiario.objects.filter(trabajador=self.trabajador)
        self.assertEqual(dias.count(), 2)
        for dia in dias:
            self.assertEqual(dia.estado, 'H')
            self.assertEqual(float(dia.jornada_horas), 12.0)
            self.assertIsNone(dia.hora_entrada)
            self.assertIsNone(dia.hora_salida)

    def test_personalizado_sin_horas_entra_igual_para_completarlo_luego(self):
        self._confirmar('P')
        for dia in TareoDiario.objects.filter(trabajador=self.trabajador):
            self.assertEqual(dia.estado, 'P')
            self.assertIsNone(dia.hora_entrada)

    def test_un_estado_inventado_no_entra(self):
        respuesta = self._confirmar('F')
        self.assertEqual(TareoDiario.objects.count(), 0)
        self.assertEqual(respuesta.status_code, 302)
