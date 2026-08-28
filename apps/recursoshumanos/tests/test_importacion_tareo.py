"""Pruebas del parseo del Excel de tareo del proyecto.

Se construye un libro con la misma forma que el archivo real (bloques
semanales en horizontal) en vez de depender de un .xlsx de ejemplo.
"""

import io
from datetime import date

import openpyxl
from django.test import TestCase

from recursoshumanos import servicios_importacion_tareo as srv
from recursoshumanos.models import Trabajador, Ubicacion


def construir_libro(semanas, filas, rotulos=None):
    """semanas: lista de (mes_texto, [7 numeros de dia]).
    filas: lista de (posicion, nombre, {indice_semana: [7 celdas]}).
    rotulos: [(fila_ini, fila_fin, texto)] para la columna de seccion de la
    izquierda; None deja el bloque sin rotulo (formato de 9 columnas)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026'

    for indice, (mes_texto, numeros) in enumerate(semanas):
        col = 2 + indice * 10          # columna "PERSONAL EN CAMPO"
        ws.cell(2, col + 3, mes_texto)
        if rotulos:
            for ini, fin, texto in rotulos:
                ws.cell(ini, col, texto)
                if fin > ini:
                    ws.merge_cells(start_row=ini, end_row=fin,
                                   start_column=col, end_column=col)
        ws.cell(4, col + 1, 'POSICIÓN')
        ws.cell(4, col + 2, 'NOMBRE')
        for i, numero in enumerate(numeros):
            ws.cell(3, col + 3 + i, numero)

    for offset, (posicion, nombre, celdas_por_semana) in enumerate(filas):
        fila = 5 + offset
        for indice in range(len(semanas)):
            col = 2 + indice * 10
            ws.cell(fila, col + 1, posicion)
            ws.cell(fila, col + 2, nombre)
            for i, valor in enumerate(celdas_por_semana.get(indice, [None] * 7)):
                if valor:
                    ws.cell(fila, col + 3 + i, valor)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class ParseoExcelTests(TestCase):
    """Julio 2026 arranca en miércoles, así que la primera y la última semana
    del archivo cruzan de mes: buen banco de pruebas para el fechado."""

    def _libro_julio(self):
        semanas = [
            ('JUNIO', [29, 30, 1, 2, 3, 4, 5]),
            ('JULIO', [6, 7, 8, 9, 10, 11, 12]),
            ('JULIO', [13, 14, 15, 16, 17, 18, 19]),
            ('JULIO', [20, 21, 22, 23, 24, 25, 26]),
            ('JULIO', [27, 28, 29, 30, 31, 1, 2]),
        ]
        filas = [
            ('M1', 'DIEGO HERNANI', {
                0: [None, None, 'M1', 'M1', 'M1', None, None],
                1: [None, 'M1', 'M1', 'M1', 'M1', None, None],
            }),
            ('GABINETE 1', 'JEYSSON SOLIS', {
                1: ['JEYSSON', 'JEYSSON', None, None, None, None, None],
            }),
            ('M4', 'SHAMIR ACHO', {
                1: ['M4', 'M4 (GAB)', None, None, None, None, None],
            }),
        ]
        return construir_libro(semanas, filas)

    def test_fecha_los_dias_respetando_los_cruces_de_mes(self):
        resultado = srv.leer_excel(self._libro_julio(), 2026, 7)

        diego = next(p for p in resultado.personas if p.nombre_excel == 'DIEGO HERNANI')
        # De la semana que cruza junio solo entran 1, 2 y 3 de julio.
        self.assertEqual(sorted(diego.dias), [1, 2, 3, 7, 8, 9, 10])
        self.assertEqual(diego.dias[1].fecha, date(2026, 7, 1))

    def test_ignora_las_semanas_de_otro_mes(self):
        resultado = srv.leer_excel(self._libro_julio(), 2026, 7)
        self.assertEqual(resultado.semanas_leidas, 5)

    def test_todo_lo_marcado_entra_como_campo(self):
        """Sin rótulo a la izquierda (formato de junio en adelante) todo es C,
        incluidas las posiciones de gabinete."""
        resultado = srv.leer_excel(self._libro_julio(), 2026, 7)

        for persona in resultado.personas:
            for dia in persona.dias.values():
                self.assertEqual(dia.estado, 'C', persona.nombre_excel)

        jeysson = next(p for p in resultado.personas if p.nombre_excel == 'JEYSSON SOLIS')
        self.assertEqual(jeysson.posicion_resumen, 'GABINETE 1')

    def test_la_anotacion_gab_ya_no_cambia_el_estado(self):
        resultado = srv.leer_excel(self._libro_julio(), 2026, 7)

        shamir = next(p for p in resultado.personas if p.nombre_excel == 'SHAMIR ACHO')
        self.assertEqual(shamir.dias[6].estado, 'C')   # "M4"
        self.assertEqual(shamir.dias[7].estado, 'C')   # "M4 (GAB)"
        self.assertEqual(shamir.dias[7].anotacion, 'GAB')

    def test_solo_importa_las_filas_bajo_personal_en_campo(self):
        """Con rótulo, las secciones que no son campo quedan fuera."""
        libro = construir_libro(
            [('JULIO', [6, 7, 8, 9, 10, 11, 12])],
            [('M1', 'DIEGO HERNANI', {0: ['M1', 'M1', None, None, None, None, None]}),
             ('M2', 'SONNY ALVIRI', {0: ['M2', 'M2', None, None, None, None, None]})],
            rotulos=[(4, 5, 'PERSONAL EN CAMPO'),
                     (6, 8, 'PERSONAL VALLECITO\nFINALIZACIÓN DE CADENAS DE CUSTODIA')],
        )
        resultado = srv.leer_excel(libro, 2026, 7)

        nombres = [p.nombre_excel for p in resultado.personas]
        self.assertEqual(nombres, ['DIEGO HERNANI'])          # fila 5, campo
        self.assertTrue(any('VALLECITO' in a for a in resultado.avisos))

    def test_avisa_cuando_la_semana_no_trae_rotulo(self):
        resultado = srv.leer_excel(self._libro_julio(), 2026, 7)
        self.assertTrue(any('no traen el rotulo' in a for a in resultado.avisos))

    def test_celda_vacia_no_genera_dia(self):
        resultado = srv.leer_excel(self._libro_julio(), 2026, 7)
        jeysson = next(p for p in resultado.personas if p.nombre_excel == 'JEYSSON SOLIS')
        self.assertNotIn(8, jeysson.dias)

    def test_no_importa_celdas_compartidas_por_dos_personas(self):
        libro = construir_libro(
            [('JULIO', [6, 7, 8, 9, 10, 11, 12])],
            [('APOYO GABINETE', 'EDWIN PUMA / HENRRY CERPA',
              {0: ['EDWIN', None, None, None, None, None, None]})],
        )
        resultado = srv.leer_excel(libro, 2026, 7)

        self.assertEqual(resultado.personas, [])
        self.assertTrue(any('mas de una persona' in a for a in resultado.avisos))

    def test_descarta_filas_sin_ningun_dia_marcado(self):
        libro = construir_libro(
            [('JULIO', [6, 7, 8, 9, 10, 11, 12])],
            [('M1', 'CARLOS ROJAS', {})],
        )
        resultado = srv.leer_excel(libro, 2026, 7)
        self.assertEqual(resultado.personas, [])

    def test_archivo_sin_el_formato_esperado_falla_con_mensaje_claro(self):
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'cualquier cosa'
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        with self.assertRaises(srv.ErrorImportacion):
            srv.leer_excel(buffer, 2026, 7)


class DiasDeLaSemanaTests(TestCase):
    """Agosto de 2026 empieza sábado, así que el sábado 1 y el domingo 2 solo
    pueden entrar por la semana que arranca el lunes 27 de julio. Sirve para
    fijar que se capturan los siete días y que cada uno cae donde debe."""

    def _libro_agosto(self):
        semanas = [
            ('JULIO', [27, 28, 29, 30, 31, 1, 2]),
            ('AGOSTO', [3, 4, 5, 6, 7, 8, 9]),
        ]
        filas = [('M1', 'DIEGO HERNANI', {0: ['M1'] * 7, 1: ['M1'] * 7})]
        return construir_libro(semanas, filas)

    def test_captura_sabado_y_domingo_igual_que_el_resto_de_dias(self):
        resultado = srv.leer_excel(self._libro_agosto(), 2026, 8)

        diego = resultado.personas[0]
        self.assertEqual(sorted(diego.dias), [1, 2, 3, 4, 5, 6, 7, 8, 9])

    def test_cada_celda_queda_en_su_dia_de_la_semana(self):
        resultado = srv.leer_excel(self._libro_agosto(), 2026, 8)

        dias = resultado.personas[0].dias
        # 5 = sábado, 6 = domingo, 0 = lunes ... 4 = viernes.
        self.assertEqual([dias[n].fecha.weekday() for n in range(1, 10)],
                         [5, 6, 0, 1, 2, 3, 4, 5, 6])


class EmparejamientoTests(TestCase):
    def setUp(self):
        self.diego = Trabajador.objects.create(
            dni='10000001', nombres='DIEGO ARMANDO',
            apellido_paterno='HERNANI', apellido_materno='ROJAS')
        self.kleyder = Trabajador.objects.create(
            dni='10000002', nombres='KLEIDER',
            apellido_paterno='JAITA', apellido_materno='MAMANI')

    def _resultado_con(self, *nombres):
        resultado = srv.ResultadoImportacion(anio=2026, mes=7)
        resultado.personas = [srv.PersonaImportada(nombre_excel=n) for n in nombres]
        return resultado

    def test_empareja_por_nombre_y_apellido_paterno(self):
        resultado = self._resultado_con('DIEGO HERNANI')
        srv.emparejar_personas(resultado, Trabajador.objects.all())

        persona = resultado.personas[0]
        self.assertEqual(persona.trabajador, self.diego)
        self.assertEqual(persona.confianza, 'exacta')

    def test_tolera_las_variantes_de_escritura_del_excel(self):
        # El archivo escribe "KLEYDER" y el maestro dice "KLEIDER".
        resultado = self._resultado_con('KLEYDER JAITA')
        srv.emparejar_personas(resultado, Trabajador.objects.all())

        persona = resultado.personas[0]
        self.assertEqual(persona.trabajador, self.kleyder)
        self.assertEqual(persona.confianza, 'aproximada')

    def test_un_desconocido_queda_sin_match(self):
        resultado = self._resultado_con('PERSONA INEXISTENTE')
        srv.emparejar_personas(resultado, Trabajador.objects.all())

        persona = resultado.personas[0]
        self.assertIsNone(persona.trabajador)
        self.assertEqual(persona.confianza, 'sin_match')


class UbicacionesTests(TestCase):
    def setUp(self):
        self.vallecito = Ubicacion.objects.create(nombre='Vallecito')

    def _resultado_con_anotacion(self, anotacion):
        resultado = srv.ResultadoImportacion(anio=2026, mes=7)
        persona = srv.PersonaImportada(nombre_excel='VICTOR QUIÑONES')
        persona.dias[6] = srv.DiaImportado(
            fecha=date(2026, 7, 6), estado='C', posicion='SUP',
            texto_celda=f'VICTOR ({anotacion})', anotacion=anotacion)
        resultado.personas = [persona]
        return resultado

    def test_resuelve_la_anotacion_contra_el_catalogo(self):
        resultado = self._resultado_con_anotacion('Vallecito')
        srv.resolver_ubicaciones(resultado, Ubicacion.objects.all())

        dia = resultado.personas[0].dias[6]
        self.assertEqual(dia.ubicacion_id, self.vallecito.id)

    def test_los_codigos_de_funcion_no_son_ubicaciones(self):
        resultado = self._resultado_con_anotacion('GAB')
        srv.resolver_ubicaciones(resultado, Ubicacion.objects.all())

        self.assertIsNone(resultado.personas[0].dias[6].ubicacion_id)
        self.assertEqual(resultado.avisos, [])

    def test_anotacion_desconocida_se_avisa_pero_no_bloquea(self):
        resultado = self._resultado_con_anotacion('C2')
        srv.resolver_ubicaciones(resultado, Ubicacion.objects.all())

        self.assertIsNone(resultado.personas[0].dias[6].ubicacion_id)
        self.assertTrue(any('sin ubicacion equivalente' in a for a in resultado.avisos))
