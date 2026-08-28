"""
Importacion del tareo mensual desde el Excel de planificacion del proyecto
(formato "TAREO PROY - MONITOREO <anio>").

El archivo NO es una tabla: son bloques semanales puestos uno al lado del otro
en horizontal. Cada bloque ocupa ~10 columnas:

    [PERSONAL EN CAMPO | POSICION | NOMBRE | LUN | MAR | MIE | JUE | VIE | SAB | DOM]
     fila 2 = mes,  fila 3 = numero de dia,  fila 4 = dia de la semana

La celda de un dia trae texto cuando esa persona trabaja ese dia y esta vacia
cuando no. Ese binario (con la POSICION de la fila) es todo lo que el archivo
sabe: no hay DNI, ni horas, ni forma de distinguir descanso de falta. El estado
que se le asigna a cada dia sale de la seccion del propio archivo:

  * "PERSONAL EN CAMPO"  -> H (campo): sin horario fijo, pero cumple 12 horas.
  * "PERSONAL VALLECITO" -> P (horario personalizado) 13:00-21:00, que es el
    turno fijo de esa cuadrilla; entra ya resuelto, sin pedir revision.
  * cualquier otra seccion -> D (dia libre): el estado inocuo, senalado para
    que RRHH decida el tipo fila por fila en la previsualizacion.

Aparte de la seccion, la nota del nombre manda: una fila "NOMBRE (vacaciones)"
entra como D (dia libre) ya resuelto, sea cual sea su seccion.

Los dias sin marca no se tocan.

El parseo es puro: no toca la base de datos. El emparejamiento con Trabajador y
la resolucion de ubicaciones reciben los catalogos ya consultados.
"""

from __future__ import annotations

import re
import unicodedata
import warnings
from dataclasses import dataclass, field
from datetime import date, timedelta
from difflib import SequenceMatcher

import openpyxl

# --- Vocabulario del archivo -------------------------------------------------

# Personal de campo: no tiene horario fijo, pero cumple 12 horas. Se guarda como
# jornada por horas de 12 (la vista pone las 12 horas al escribir el TareoDiario).
ESTADO_CAMPO = 'H'

# Secciones que el Excel pone fuera de PERSONAL EN CAMPO y que el sistema NO sabe
# clasificar. No son trabajo de campo, y el archivo tampoco dice que turno son:
# entran como dia libre, el estado inocuo, senalado para que RRHH decida a mano.
ESTADO_LIBRE = 'D'

# Turno Vallecito: la cuadrilla de finalizacion de cadenas de custodia trabaja un
# horario personalizado fijo de 13:00 a 21:00. Como el turno es conocido, se
# importa ya resuelto (P con esas horas), sin marcarlo para revision.
ESTADO_VALLECITO = 'P'
HORARIO_VALLECITO = ('13:00', '21:00')

# La columna que va INMEDIATAMENTE A LA IZQUIERDA de POSICION rotula la seccion
# a la que pertenece cada fila, en una celda combinada vertical. El rotulo decide
# el estado: "PERSONAL EN CAMPO" -> H; "PERSONAL VALLECITO ..." -> P 13:00-21:00;
# cualquier otra seccion -> D (dia libre) y senalada para que RRHH la revise.
SECCION_CAMPO = 'PERSONAL EN CAMPO'
SECCION_VALLECITO = 'VALLECITO'

# Ojo: la columna de rotulo solo existe hasta la semana del 1 de junio de 2026.
# A partir de ahi el archivo pasa de bloques de 10 columnas a bloques de 9 y la
# columna desaparece. Sin rotulo se asume campo (ver `_secciones_del_bloque`).

# Anotaciones que describen la funcion del dia, no un lugar. No se intenta
# resolverlas contra el catalogo de Ubicacion.
ANOTACIONES_SIN_UBICACION = {'GAB', 'GABINETE', 'LAB', 'SOP', 'SOPORTE'}

# Umbral del emparejamiento aproximado de nombres. Por debajo de esto la fila
# queda "sin match" y se resuelve a mano en la previsualizacion.
UMBRAL_SIMILITUD = 0.82

MAX_FILA_BLOQUE = 74  # mas abajo solo hay notas y listas borrador


# --- Utilidades --------------------------------------------------------------

def normalizar(texto) -> str:
    """Mayusculas, sin tildes, sin puntuacion y con espacios colapsados.

    El archivo mezcla espacios normales con espacios tipograficos y escribe el
    mismo nombre de varias formas, asi que todo emparejamiento pasa por aca.
    """
    if texto is None:
        return ''
    s = unicodedata.normalize('NFKD', str(texto))
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper()
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def _entero(valor):
    """Numero de dia de una celda de la fila 3, o None."""
    if isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    if isinstance(valor, str) and valor.strip().isdigit():
        return int(valor.strip())
    return None


def partir_anotacion(texto):
    """"M4 (GAB)" -> ("M4", "GAB"). Sin parentesis la anotacion queda vacia."""
    bruto = str(texto).strip()
    encontrado = re.search(r'\(([^)]*)\)', bruto)
    anotacion = encontrado.group(1).strip() if encontrado else ''
    base = re.sub(r'\([^)]*\)', '', bruto).strip()
    return base, anotacion


def es_seccion_campo(rotulo) -> bool:
    """True si el rotulo de la izquierda dice PERSONAL EN CAMPO.

    Sin rotulo (bloques de 9 columnas, de junio en adelante) tambien se toma
    como campo: en ese formato la hoja ya no separa secciones y todo lo que
    aparece es personal de campo.
    """
    if rotulo is None:
        return True
    return SECCION_CAMPO in normalizar(rotulo)


def es_seccion_vallecito(rotulo) -> bool:
    """True si el rotulo de la izquierda corresponde al turno Vallecito.

    Sin rotulo devuelve False: `es_seccion_campo` ya se queda con ese caso.
    """
    if rotulo is None:
        return False
    return SECCION_VALLECITO in normalizar(rotulo)


# El Excel abre una fila aparte para los dias de vacaciones, con la nota escrita
# en el propio nombre: "WASHINGTON (vacaciones)" convive con "WASHINGTON PALOMINO"
# y solo trae los dias libres. Esos dias NO son trabajo de campo: entran como D
# (dia libre), ya resueltos, sin marcar la fila para revision.
MARCADORES_DIA_LIBRE = ('VACACION',)  # cubre VACACION / VACACIONES


def es_nombre_dia_libre(nombre) -> bool:
    """True si el nombre trae una nota de vacaciones (dias que van como D)."""
    normalizado = normalizar(nombre)
    return any(marca in normalizado for marca in MARCADORES_DIA_LIBRE)


# --- Estructuras de resultado ------------------------------------------------

@dataclass
class DiaImportado:
    fecha: date
    estado: str
    posicion: str
    texto_celda: str
    anotacion: str = ''
    ubicacion_id: object = None
    ubicacion_nombre: str = ''
    # Horario por defecto que trae la seccion, cuando lo tiene (turno Vallecito:
    # '13:00'/'21:00'). Vacio para campo y para el resto. Lo usa la vista para
    # precargar las horas de la fila cuando el estado es P.
    hora_entrada: str = ''
    hora_salida: str = ''
    # Rotulo de la seccion del Excel cuando es una seccion AJENA que RRHH tiene
    # que revisar (ni campo ni Vallecito). Vacio = seccion ya resuelta.
    seccion: str = ''

    @property
    def dia(self) -> int:
        return self.fecha.day


@dataclass
class PersonaImportada:
    """Una persona del Excel con todos sus dias del mes ya agregados."""
    nombre_excel: str
    posiciones: list = field(default_factory=list)
    dias: dict = field(default_factory=dict)  # {numero_dia: DiaImportado}

    # Resultado del emparejamiento contra Trabajador
    trabajador: object = None
    confianza: str = 'sin_match'   # exacta | aproximada | ambigua | sin_match
    candidatos: list = field(default_factory=list)
    similitud: float = 0.0

    @property
    def posicion_resumen(self) -> str:
        return ' / '.join(dict.fromkeys(self.posiciones))

    @property
    def dias_ordenados(self) -> list:
        return [self.dias[k] for k in sorted(self.dias)]

    @property
    def total_dias(self) -> int:
        return len(self.dias)

    @property
    def dias_otra_seccion(self) -> list:
        """Dias que el Excel puso bajo una seccion distinta de campo."""
        return [d for d in self.dias_ordenados if d.seccion]

    @property
    def secciones_ajenas(self) -> list:
        """Rotulos no-campo bajo los que aparece esta persona, sin repetir."""
        return list(dict.fromkeys(d.seccion for d in self.dias_otra_seccion))

    @property
    def solo_otra_seccion(self) -> bool:
        """True si NINGUN dia de la persona viene de PERSONAL EN CAMPO."""
        return bool(self.dias) and all(d.seccion for d in self.dias.values())


@dataclass
class ResultadoImportacion:
    anio: int
    mes: int
    personas: list = field(default_factory=list)
    avisos: list = field(default_factory=list)
    semanas_leidas: int = 0
    hoja: str = ''
    # Semanas que se dejaron FUERA por no traer los numeros de dia en el archivo:
    # su fecha solo se puede deducir por la posicion del bloque, y una cabecera
    # mal copiada haria escribir toda la semana en los dias equivocados. Cada
    # entrada: {'inicio': date, 'fin': date, 'dias': [numeros del mes]}.
    semanas_omitidas: list = field(default_factory=list)

    @property
    def total_dias(self) -> int:
        return sum(p.total_dias for p in self.personas)

    @property
    def emparejadas(self) -> list:
        return [p for p in self.personas if p.trabajador is not None]

    @property
    def sin_emparejar(self) -> list:
        return [p for p in self.personas if p.trabajador is None]


class ErrorImportacion(Exception):
    """El archivo no se puede leer o no tiene el formato esperado."""


# --- Lectura del Excel -------------------------------------------------------

def _localizar_bloques(ws) -> list:
    """Columnas donde arranca cada bloque semanal (la de POSICION).

    Se detecta por la fila 4 y no por la de NOMBRE porque en varias semanas el
    encabezado "NOMBRE" quedo pisado con el nombre de una persona.
    """
    bloques = []
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(4, col).value
        if valor and 'POSICI' in normalizar(valor):
            bloques.append(col)
    return bloques


def _lunes_candidatos(numeros, anio_objetivo) -> list:
    """Lunes del calendario cuyos 7 dias coinciden con los numeros del bloque.

    Casi siempre hay mas de uno: "6,7,8,...,12" encaja tanto en abril como en
    julio de 2026. La eleccion final la hace `_resolver_calendario`.
    """
    ancla = next(((i, n) for i, n in enumerate(numeros) if n is not None), None)
    if ancla is None:
        return []
    indice_ancla, numero_ancla = ancla

    candidatos = []
    for anio in (anio_objetivo - 1, anio_objetivo, anio_objetivo + 1):
        for mes in range(1, 13):
            try:
                fecha_ancla = date(anio, mes, numero_ancla)
            except ValueError:
                continue
            lunes = fecha_ancla - timedelta(days=indice_ancla)
            if lunes.weekday() != 0:
                continue
            if all(n is None or (lunes + timedelta(days=i)).day == n
                   for i, n in enumerate(numeros)):
                candidatos.append(lunes)
    return candidatos


def _resolver_calendario(numeros_por_bloque, anio_objetivo, mes_objetivo) -> dict:
    """Asigna las 7 fechas reales a cada bloque semanal.

    Un bloque suelto es ambiguo, pero la hoja completa no lo es: los bloques
    van en orden y cada uno arranca exactamente 7 dias despues del anterior.
    Se vota el lunes de origen que mas bloques respaldan y desde ahi se deduce
    el resto de la hoja; asi tambien quedan fechadas las semanas a las que les
    falta la fila de numeros.
    """
    candidatos = {i: _lunes_candidatos(nums, anio_objetivo)
                  for i, nums in enumerate(numeros_por_bloque)}

    votos = {}
    for indice, lunes_posibles in candidatos.items():
        for lunes in lunes_posibles:
            origen = lunes - timedelta(days=7 * indice)
            votos[origen] = votos.get(origen, 0) + 1

    total = len(numeros_por_bloque)
    referencia = date(anio_objetivo, mes_objetivo, 1)

    def _puntaje(par):
        """Mas bloques respaldando gana; a igualdad, la cadena que cubre el mes.

        Un calendario se repite: "6..12 empezando lunes" encaja igual en abril
        de 2026 que en julio, y la hoja entera desplazada tambien encaja. El
        desempate mira que la cadena contenga el mes que se esta importando.
        """
        origen, apoyo = par
        fin = origen + timedelta(days=7 * total - 1)
        if origen <= referencia <= fin:
            distancia = 0
        else:
            distancia = min(abs((origen - referencia).days),
                            abs((fin - referencia).days))
        return (apoyo, -distancia)

    if votos:
        origen, apoyo = max(votos.items(), key=_puntaje)
        if apoyo >= max(3, total // 2):
            calendario = {}
            for indice in range(total):
                lunes = origen + timedelta(days=7 * indice)
                # Si el bloque trae numeros y no cuadran con la cadena, se
                # respeta lo que dice el propio bloque antes que la deduccion.
                posibles = candidatos.get(indice) or []
                if posibles and lunes not in posibles:
                    lunes = min(posibles, key=lambda f: abs((f - lunes).days))
                calendario[indice] = [lunes + timedelta(days=d) for d in range(7)]
            return calendario

    # Respaldo: hoja sin cadena reconocible. Se resuelve bloque por bloque
    # eligiendo el candidato mas cercano al mes que se esta importando.
    calendario = {}
    for indice, posibles in candidatos.items():
        if not posibles:
            continue
        lunes = min(posibles, key=lambda f: abs((f - referencia).days))
        calendario[indice] = [lunes + timedelta(days=d) for d in range(7)]
    return calendario


def _secciones_del_bloque(hoja, col_rotulo) -> dict:
    """Rotulo de seccion que aplica a cada fila del bloque.

    El rotulo vive en la columna anterior a POSICION, escrito una sola vez en
    una celda combinada vertical que abarca todas las filas de esa seccion.
    Devuelve {fila: texto_del_rotulo}. Si esa columna no rotula nada (bloques
    de 9 columnas, donde a la izquierda esta el domingo de la semana previa)
    devuelve un diccionario vacio y el bloque se trata entero como campo.
    """
    if col_rotulo < 1:
        return {}
    encabezado = hoja.cell(4, col_rotulo).value
    if not encabezado or 'PERSONAL' not in normalizar(encabezado):
        return {}

    secciones = {}
    for rango in hoja.merged_cells.ranges:
        if rango.min_col != col_rotulo or rango.max_col != col_rotulo:
            continue
        texto = hoja.cell(rango.min_row, col_rotulo).value
        if not texto:
            continue
        for fila in range(rango.min_row, rango.max_row + 1):
            secciones[fila] = texto

    # Una seccion de una sola fila no queda combinada: se toma tal cual.
    for fila in range(4, MAX_FILA_BLOQUE):
        valor = hoja.cell(fila, col_rotulo).value
        if valor and fila not in secciones:
            secciones[fila] = valor
    return secciones


def _es_fila_valida(posicion, nombre) -> bool:
    """Descarta filas de notas, encabezados repetidos y horas sueltas."""
    if not posicion or not nombre:
        return False
    if '\n' in str(posicion) or len(str(posicion)) > 30:
        return False   # "GRUPO MINA - TURNO A\n4:30 AM..." y similares
    nombre_norm = normalizar(nombre)
    if len(nombre_norm) < 4:
        return False
    if not re.search(r'[A-Z]{3}', nombre_norm):
        return False   # "01:40:00"
    if nombre_norm in {'NOMBRE', 'NOMBRES', 'POSICION'}:
        return False
    return True


def leer_excel(archivo, anio, mes) -> ResultadoImportacion:
    """Parsea el archivo y devuelve solo los dias del mes pedido."""
    with warnings.catch_warnings():
        # El archivo trae imagenes wmf que openpyxl no sabe leer; son irrelevantes.
        warnings.simplefilter('ignore')
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
        except Exception as exc:
            raise ErrorImportacion(
                "No se pudo abrir el archivo. Debe ser un Excel .xlsx valido."
            ) from exc

    resultado = ResultadoImportacion(anio=anio, mes=mes)

    hoja, bloques = None, []
    for candidata in wb.worksheets:
        encontrados = _localizar_bloques(candidata)
        if encontrados:
            hoja, bloques = candidata, encontrados
            if normalizar(candidata.title) == str(anio):
                break
    if hoja is None:
        raise ErrorImportacion(
            "El archivo no tiene el formato del tareo de proyecto: no se encontro "
            "ninguna columna 'POSICION' en la fila 4."
        )
    resultado.hoja = hoja.title

    agregadas = {}
    compartidas = set()
    conflictos = set()
    otras_secciones = {}
    vallecito_filas = 0
    vacaciones_filas = 0
    bloques_sin_rotulo = 0

    numeros_por_bloque = [
        [_entero(hoja.cell(3, c).value) for c in range(col + 2, col + 9)]
        for col in bloques
    ]
    calendario = _resolver_calendario(numeros_por_bloque, anio, mes)

    # Bloques que no traen NINGUN numero de dia en la fila 3. Su fecha solo sale
    # de la cadena (bloque anterior + 7 dias), y una cabecera mal armada por
    # quien edita el Excel (rotulo del mes corrido a la fila de numeros, numeros
    # pegados de otra semana) haria escribir la semana entera en dias que no son.
    # No se adivina: esa semana se deja fuera y se le avisa a RRHH.
    semanas_sin_numeros = {
        i for i, nums in enumerate(numeros_por_bloque)
        if not any(n is not None for n in nums)
    }

    for indice_bloque, col_pos in enumerate(bloques):
        col_nombre = col_pos + 1
        col_primer_dia = col_pos + 2

        fechas = calendario.get(indice_bloque)
        if fechas is None:
            continue
        if not any(f.year == anio and f.month == mes for f in fechas):
            continue  # semana de otro mes: se ignora entera
        if indice_bloque in semanas_sin_numeros:
            resultado.semanas_omitidas.append({
                'inicio': fechas[0],
                'fin': fechas[6],
                'dias': [f.day for f in fechas
                         if f.year == anio and f.month == mes],
            })
            continue
        resultado.semanas_leidas += 1

        secciones = _secciones_del_bloque(hoja, col_pos - 1)
        if not secciones:
            bloques_sin_rotulo += 1

        for fila in range(5, MAX_FILA_BLOQUE):
            posicion = hoja.cell(fila, col_pos).value
            nombre = hoja.cell(fila, col_nombre).value
            celdas = [hoja.cell(fila, c).value
                      for c in range(col_primer_dia, col_primer_dia + 7)]

            if not any(c is not None and str(c).strip() for c in celdas):
                continue  # fila sin ningun dia marcado (listas borrador)
            if not _es_fila_valida(posicion, nombre):
                continue

            # La seccion de la izquierda dice de que turno es la fila:
            #   * campo     -> H (12 horas, sin horario fijo)
            #   * Vallecito -> P 13:00-21:00 (turno conocido, entra ya resuelto)
            #   * otra      -> D (dia libre) y marcada para que RRHH la revise;
            #                  no se descarta en silencio, se le muestra a RRHH
            #                  (antes se perdia sin dejar rastro).
            # Ojo: los limites de cada seccion varian de una semana a otra (una
            # cuadrilla crece o encoge), asi que se leen por bloque de las celdas
            # combinadas de la izquierda, nunca de un rango de filas fijo.
            nombre_texto = str(nombre).strip()
            if '/' in nombre_texto:
                # "EDWIN PUMA / HENRRY CERPA": no se adivina a quien le toca.
                compartidas.add(nombre_texto)
                continue

            rotulo = secciones.get(fila) if secciones else None
            seccion_fila = ''
            horario_seccion = ('', '')
            if es_nombre_dia_libre(nombre_texto):
                # "WASHINGTON (vacaciones)": manda la nota del nombre por encima de
                # la seccion. Son dias libres, entran como D ya resuelto (verde),
                # no como campo ni marcados para revision.
                estado_seccion = ESTADO_LIBRE
                vacaciones_filas += 1
            elif es_seccion_campo(rotulo):
                estado_seccion = ESTADO_CAMPO
            elif es_seccion_vallecito(rotulo):
                estado_seccion = ESTADO_VALLECITO
                horario_seccion = HORARIO_VALLECITO
                vallecito_filas += 1
            else:
                estado_seccion = ESTADO_LIBRE
                seccion_fila = ' '.join(str(rotulo).split())
                otras_secciones[seccion_fila] = otras_secciones.get(seccion_fila, 0) + 1

            clave = normalizar(nombre_texto)
            persona = agregadas.get(clave)
            if persona is None:
                persona = PersonaImportada(nombre_excel=nombre_texto)
                agregadas[clave] = persona
            persona.posiciones.append(str(posicion).strip())

            for indice_dia, celda in enumerate(celdas):
                if celda is None or not str(celda).strip():
                    continue
                fecha = fechas[indice_dia]
                if fecha.year != anio or fecha.month != mes:
                    continue  # dia de la semana que cae en el mes vecino
                _, anotacion = partir_anotacion(celda)
                nuevo = DiaImportado(
                    fecha=fecha,
                    estado=estado_seccion,
                    posicion=str(posicion).strip(),
                    texto_celda=str(celda).strip(),
                    anotacion=anotacion,
                    hora_entrada=horario_seccion[0],
                    hora_salida=horario_seccion[1],
                    seccion=seccion_fila,
                )
                previo = persona.dias.get(fecha.day)
                # Si el mismo dia aparece en una seccion ajena (marcada) y en una
                # ya resuelta (campo o Vallecito), manda la resuelta: es la que el
                # importador sabe escribir sin pedir intervencion.
                if previo is None or (previo.seccion and not seccion_fila):
                    persona.dias[fecha.day] = nuevo

    # Una fila puede tener marcas solo en los dias que caen en el mes vecino
    # (semanas a caballo entre dos meses, tipo "MARZO - ABRIL"). Esa persona no
    # aporta ningun dia al mes que se importa: se descarta para que no salga
    # como fila vacia en la previsualizacion.
    resultado.personas = sorted((p for p in agregadas.values() if p.dias),
                                key=lambda p: normalizar(p.nombre_excel))

    if not resultado.semanas_leidas:
        if resultado.semanas_omitidas:
            raise ErrorImportacion(
                "Ninguna semana de ese mes trae los numeros de dia en el archivo, "
                "asi que no se puede saber a que fechas corresponden. Corrige la "
                "fila de numeros de las semanas en el Excel y vuelve a subirlo."
            )
        raise ErrorImportacion("El archivo no contiene ninguna semana de ese mes.")
    for semana in resultado.semanas_omitidas:
        resultado.avisos.append(
            f"Semana del {semana['inicio'].strftime('%d/%m')} al "
            f"{semana['fin'].strftime('%d/%m')}: NO se importo. Su bloque en el "
            "Excel no trae la fila con los numeros de dia (la cabecera esta mal "
            "armada), asi que la fecha no es confiable. Esos dias quedan vacios "
            "para que RRHH los complete a mano."
        )
    if compartidas:
        resultado.avisos.append(
            "Celdas con mas de una persona, no importadas: "
            + ', '.join(sorted(compartidas))
        )
    if vallecito_filas:
        resultado.avisos.append(
            "Turno Vallecito: sus dias se importaron como \"P\" (horario "
            "personalizado) de 13:00 a 21:00. Si alguna fila lleva otro horario, "
            "ajustalo antes de confirmar."
        )
    if vacaciones_filas:
        resultado.avisos.append(
            f"Vacaciones: {vacaciones_filas} fila(s) con la nota \"vacaciones\" en "
            "el nombre se importaron como \".\" (dia libre), no como campo."
        )
    for seccion, cuantas in sorted(otras_secciones.items()):
        resultado.avisos.append(
            f"Seccion \"{seccion}\": {cuantas} fila(s) marcadas para revision "
            "(no estan bajo PERSONAL EN CAMPO). Sus dias entran como \".\" "
            "(dia libre), no como campo: revisa el tipo fila por fila."
        )
    if bloques_sin_rotulo:
        resultado.avisos.append(
            f"{bloques_sin_rotulo} de {resultado.semanas_leidas} semana(s) no traen "
            "el rotulo PERSONAL EN CAMPO a la izquierda (el archivo cambio de "
            "formato en junio y esa columna desaparecio). Se importaron como campo."
        )
    return resultado


# --- Emparejamiento con el maestro de trabajadores ---------------------------

def _indice_trabajadores(trabajadores):
    indice = []
    for t in trabajadores:
        completo = normalizar(f"{t.nombres} {t.apellido_paterno} {t.apellido_materno}")
        indice.append((t, completo, set(completo.split())))
    return indice


def emparejar_personas(resultado, trabajadores) -> None:
    """Asocia cada nombre del Excel con un Trabajador activo.

    El Excel escribe "CHRISTOPHER BEGAZO" (nombre + apellido paterno) y con
    varias grafias del mismo apellido, asi que se prueba primero por tokens
    contenidos y despues por similitud. Nunca decide solo: la confianza viaja a
    la pantalla de previsualizacion para que RRHH confirme.
    """
    indice = _indice_trabajadores(trabajadores)

    for persona in resultado.personas:
        # La nota entre parentesis del nombre ("(vacaciones)", "(VALLECITO)") no
        # es parte del nombre: se quita antes de emparejar para que la fila
        # empareje con el mismo trabajador que su fila normal.
        sin_nota = re.sub(r'\([^)]*\)', '', persona.nombre_excel)
        objetivo = normalizar(sin_nota)
        tokens = set(objetivo.split())

        contenidos = [t for t, _, tks in indice if tokens and tokens <= tks]
        if len(contenidos) == 1:
            persona.trabajador = contenidos[0]
            persona.confianza = 'exacta'
            persona.similitud = 1.0
            continue
        if len(contenidos) > 1:
            persona.candidatos = contenidos
            persona.confianza = 'ambigua'
            continue

        mejor, mejor_ratio = None, 0.0
        for trabajador, completo, _ in indice:
            ratio = SequenceMatcher(None, objetivo, completo).ratio()
            # Tambien contra "nombre + apellido paterno", que es como suele
            # venir escrito en el Excel.
            partes = completo.split()
            if len(partes) > 2:
                corto = f"{partes[0]} {partes[-2]}"
                ratio = max(ratio, SequenceMatcher(None, objetivo, corto).ratio())
            if ratio > mejor_ratio:
                mejor, mejor_ratio = trabajador, ratio

        if mejor is not None and mejor_ratio >= UMBRAL_SIMILITUD:
            persona.trabajador = mejor
            persona.confianza = 'aproximada'
            persona.similitud = round(mejor_ratio, 3)
        else:
            persona.confianza = 'sin_match'
            persona.similitud = round(mejor_ratio, 3)


# --- Ubicaciones (anotaciones entre parentesis) ------------------------------

def resolver_ubicaciones(resultado, ubicaciones) -> None:
    """Traduce la anotacion del dia a una Ubicacion del catalogo, si existe.

    Solo se resuelve lo que coincide por nombre con una ubicacion ya registrada.
    Codigos de funcion como (GAB) o (LAB), y las anotaciones de una sola letra
    tipo (M)/(G), no son lugares: quedan como metadato informativo.
    """
    catalogo = {normalizar(u.nombre): u for u in ubicaciones}
    sin_equivalente = set()

    for persona in resultado.personas:
        for dia in persona.dias.values():
            anotacion = normalizar(dia.anotacion)
            if not anotacion:
                continue
            if anotacion in ANOTACIONES_SIN_UBICACION or len(anotacion) < 2:
                continue
            ubicacion = catalogo.get(anotacion)
            if ubicacion is not None:
                dia.ubicacion_id = ubicacion.id
                dia.ubicacion_nombre = ubicacion.nombre
            else:
                sin_equivalente.add(dia.anotacion)

    if sin_equivalente:
        resultado.avisos.append(
            "Anotaciones sin ubicacion equivalente en el catalogo (se importan "
            "igual, pero sin ubicacion): " + ', '.join(sorted(sin_equivalente))
        )


def analizar(archivo, anio, mes, trabajadores, ubicaciones) -> ResultadoImportacion:
    """Parseo + emparejamiento + ubicaciones, en un solo paso."""
    resultado = leer_excel(archivo, anio, mes)
    emparejar_personas(resultado, trabajadores)
    resolver_ubicaciones(resultado, ubicaciones)
    return resultado
