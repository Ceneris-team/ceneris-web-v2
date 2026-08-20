from django.shortcuts import get_object_or_404, render, redirect
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.utils import timezone as django_timezone
from django.utils import timezone
from django.contrib.auth import logout
from metricas_ceneris.views import _calcular_asistencia_por_periodo, _hora_referencia_entrada_por_fecha
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from datetime import timedelta
from decimal import Decimal
from .models import Area, IntentoFraude, Trabajador, Empresa # ... etc
from django.shortcuts import render, get_object_or_404, redirect
from rest_framework import permissions
from admin_panel.settings import db
from .forms import TrabajadorForm, UbicacionForm, JustificacionForm, EmpresaForm, ProyectoForm, CargoForm, PermisosAsistenciaForm, CentroCostoForm, AreaForm, DispositivoForm
from .models import Cargo, Empresa, Proyecto, Trabajador, CentroCosto, Ubicacion, TareoDiario
import pandas as pd
from recursoshumanos.services import recalcular_asistencia_diaria
from .models import Sede, ConfiguracionTolerancia, ToleranciaAuditoria
from .motor_reglas import EstadoMarca
from .services import listar_tolerancias, crear_o_actualizar_tolerancia, actualizar_tolerancia
from firebase_admin import firestore
from google.cloud import firestore
from google.cloud.firestore_v1.field_path import FieldPath
from google.cloud.firestore_v1.base_query import FieldFilter
from datetime import datetime, time, date
import pytz
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import MyTokenObtainPairSerializer, AsistenciaSerializer
import csv
from .models import Trabajador, AsignacionProyecto
from .models import Asistencia, Dispositivo
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from .models import Justificacion, TareoDiario
from .serializers import FaltaPendienteSerializer, CrearJustificacionSerializer
from .decorators import group_required
from rest_framework import serializers
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy, reverse
from django.views.decorators.http import require_POST
import json
from django.db.models import Q, Prefetch
from django.utils.http import url_has_allowed_host_and_scheme
from collections import Counter, defaultdict
from .templatetags.custom_filters_rrhh import has_group
import calendar
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import IntegrityError
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import SolicitudHorasExtra
from urllib.parse import urlparse
from openpyxl import Workbook
from .models import Pregunta, ResultadoCuestionario, Respuesta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated
from .serializers import SolicitudHorasExtraSerializer
import logging
from rest_framework.decorators import api_view
import traceback


# --- CONFIGURACIÓN GLOBAL PARA ESTE ARCHIVO ---
LOCAL_TIMEZONE = pytz.timezone('America/Lima') # ¡Asegúrate de que esta sea tu zona horaria!
TOLERANCIA_TARDANZA_MINUTOS = 15
TARDANZA_MINIMA_HORAS = Decimal('0.25')
HORA_LIMITE_TARDANZA_OFICINA = time(8, 45, 0)


# --- CONFIGURACIÓN DE HORARIOS ---
HORARIOS = {
    'C': {'entrada': time(6, 0), 'salida': time(19, 0), 'tolerancia_minutos': TOLERANCIA_TARDANZA_MINUTOS, 'duracion_jornada': timedelta(hours=8)},
    'O': {'entrada': time(8, 30), 'salida': time(18, 0), 'tolerancia_minutos': TOLERANCIA_TARDANZA_MINUTOS, 'duracion_jornada': timedelta(hours=8)},
}

# --- FILTRO POR MEDIO DE MARCACIÓN (campo Asistencia.origen) ---
# Los valores deben coincidir con los choices del modelo Asistencia.
ORIGENES_MARCACION = [
    ('APP', 'Aplicación Móvil'),
    ('BIOMETRICO', 'Reloj Biométrico'),
    ('MANUAL', 'Manual'),
]
ORIGENES_VALIDOS = {clave for clave, _ in ORIGENES_MARCACION}


def _normalizar_filtro_origen(valor):
    """
    Valida el filtro de medio recibido por GET. Devuelve el codigo en
    mayusculas si es valido, o None si viene vacio/basura (en cuyo caso el
    reporte sale sin filtrar, que es el comportamiento historico).
    """
    if not valor:
        return None
    valor = str(valor).strip().upper()
    return valor if valor in ORIGENES_VALIDOS else None


def _etiqueta_origen(codigo):
    """Nombre legible del medio, para mostrarlo en el encabezado del Excel."""
    return dict(ORIGENES_MARCACION).get(codigo, codigo or '')

# --- FUNCIONES AUXILIARES ---

def formatear_timedelta(td):
    if not td or td.total_seconds() <= 0: return ""
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours}:{minutes:02d}"

def calcular_resumen_diario(marcaciones_del_dia, tareo_del_dia):
    """
    Calcula el resumen de un día con lógica avanzada de negocio.
    1. Suma intervalos de Entrada/Salida para las horas trabajadas.
    2. Calcula tardanza solo si hay hora de entrada fija.
    3. Calcula horas extra solo si el tareo del día lo permite.
    4. Distingue entre Descanso (sin tareo) y Falta (con tareo pero sin marcación).
    """
    resumen = {
        'hrs_trabajadas': timedelta(), 
        'hrs_tardanza': timedelta(), 
        'hrs_extra': timedelta(), 
        'hrs_faltantes': timedelta(),
        'estado_celda': '' # Empezamos con estado vacío
    }
    
    tipo_jornada_planeada = tareo_del_dia.get('estado')

    # --- REGLA 5: Si no se elige horario, es descanso ---
    if not tipo_jornada_planeada or tipo_jornada_planeada in ['.', 'D']:
        resumen['estado_celda'] = '.'
        return resumen

    # --- Obtenemos las reglas del horario planeado ---
    reglas_horario = None
    calcular_tardanza_flag = True
    # Asumimos que las horas extra se permiten por defecto, a menos que se indique lo contrario.
    # Para más control, podrías añadir un campo 'permite_horas_extra' en tu tareo.
    permite_horas_extra = tareo_del_dia.get('permite_horas_extra', True) 

    if tipo_jornada_planeada == 'P' and 'horario' in tareo_del_dia:
        horario_p = tareo_del_dia['horario']
        entrada_dt = datetime.strptime(horario_p.get('entrada', '00:00'), '%H:%M').time()
        salida_dt = datetime.strptime(horario_p.get('salida', '00:00'), '%H:%M').time()
        duracion = datetime.combine(date.today(), salida_dt) - datetime.combine(date.today(), entrada_dt)
        reglas_horario = {'entrada': entrada_dt, 'salida': salida_dt, 'tolerancia_minutos': TOLERANCIA_TARDANZA_MINUTOS, 'duracion_jornada': duracion}
    elif tipo_jornada_planeada == 'J' and 'jornada_horas' in tareo_del_dia:
        horas_a_cumplir = float(tareo_del_dia.get('jornada_horas', 8))
        reglas_horario = {'duracion_jornada': timedelta(hours=horas_a_cumplir)}
        calcular_tardanza_flag = False
    else:
        reglas_horario = HORARIOS.get(tipo_jornada_planeada)

    if not reglas_horario:
        resumen['estado_celda'] = 'ERR' # Error, horario no definido
        return resumen

    # --- REGLA 6: Si hay horario pero no marcaciones, es FALTA ---
    if not marcaciones_del_dia:
        resumen['estado_celda'] = 'F'
        resumen['hrs_faltantes'] = reglas_horario['duracion_jornada']
        return resumen

    # --- LÓGICA DE CÁLCULO DE HORAS TRABAJADAS (Hrs Trabs) ---
    marcaciones = sorted(marcaciones_del_dia, key=lambda x: x['timestamp'])
    jornada_efectiva_bruta = timedelta()
    ultima_entrada = None
    for m in marcaciones:
        tipo = m.get('tipoMarcacion')
        timestamp = m.get('timestamp').astimezone(LOCAL_TIMEZONE)
        
        if tipo == 'Entrada':
            ultima_entrada = timestamp
        elif tipo == 'Salida' and ultima_entrada:
            jornada_efectiva_bruta += timestamp - ultima_entrada
            ultima_entrada = None # Reseteamos para el próximo par

    jornada_neta = max(timedelta(0), jornada_efectiva_bruta) # Aseguramos que no sea negativo
    resumen['hrs_trabajadas'] = jornada_neta
    resumen['estado_celda'] = tipo_jornada_planeada
    
    # --- CÁLCULO DE TARDANZA (Hrs Tard) ---
    if calcular_tardanza_flag and 'entrada' in reglas_horario:
        primera_entrada = next((m['timestamp'].astimezone(LOCAL_TIMEZONE) for m in marcaciones if m.get('tipoMarcacion') == 'Entrada'), None)
        if primera_entrada:
            hora_limite = primera_entrada.replace(hour=reglas_horario['entrada'].hour, minute=reglas_horario['entrada'].minute, second=0)
            tolerancia = timedelta(minutes=reglas_horario.get('tolerancia_minutos', TOLERANCIA_TARDANZA_MINUTOS))
            if primera_entrada > hora_limite + tolerancia:
                resumen['hrs_tardanza'] = primera_entrada - (hora_limite + tolerancia)

    # --- CÁLCULO DE HORAS FALTANTES Y EXTRAS (Hrs Faltas / H Ext) ---
    jornada_regular = reglas_horario['duracion_jornada']
    if jornada_neta < jornada_regular:
        resumen['hrs_faltantes'] = jornada_regular - jornada_neta
    elif jornada_neta > jornada_regular and permite_horas_extra:
        resumen['hrs_extra'] = jornada_neta - jornada_regular
        
    return resumen

# --- LÓGICA DE REPORTE GENERAL Y EXPORTACIÓN ---

def _get_report_data(filtro_fecha_inicio_str, filtro_fecha_fin_str, filtro_ubicacion):
    """
    Función auxiliar MEJORADA para calcular los datos del reporte. Es usada por la vista y la exportación.
    """
    fecha_inicio = LOCAL_TIMEZONE.localize(datetime.strptime(filtro_fecha_inicio_str, '%Y-%m-%d').replace(hour=0, minute=0))
    fecha_fin = LOCAL_TIMEZONE.localize(datetime.strptime(filtro_fecha_fin_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59))

    trabajadores = {}
    trabajadores_query = db.collection('trabajadores').where(filter=FieldFilter('activo', '==', True))
    if filtro_ubicacion:
        trabajadores_query = trabajadores_query.where(filter=FieldFilter('ubicacionesPermitidas', 'array_contains', filtro_ubicacion))
    for doc in trabajadores_query.stream():
        trabajadores[doc.id] = doc.to_dict()

    asistencias_query = db.collection('asistencias').where(filter=FieldFilter('timestamp', '>=', fecha_inicio)).where(filter=FieldFilter('timestamp', '<=', fecha_fin))
    if filtro_ubicacion:
        asistencias_query = asistencias_query.where(filter=FieldFilter('locationName', '==', filtro_ubicacion))
    
    justificaciones_query = db.collection('justificaciones').where(filter=FieldFilter('fechaInicio', '<=', fecha_fin)).where(filter=FieldFilter('fechaFin', '>=', fecha_inicio))
    
    asistencias_por_dni = {dni: [] for dni in trabajadores.keys()}
    for doc in asistencias_query.stream():
        data = doc.to_dict()
        dni = data.get('userDni')
        if dni in asistencias_por_dni:
            asistencias_por_dni[dni].append(data)

    justificaciones_por_dni = {dni: [] for dni in trabajadores.keys()}
    for doc in justificaciones_query.stream():
        data = doc.to_dict()
        dni = data.get('dniTrabajador')
        if dni in justificaciones_por_dni:
            justificaciones_por_dni[dni].append(data)

    reporte_final = []
    HORA_LIMITE_TARDANZA = HORA_LIMITE_TARDANZA_OFICINA
    for dni, data_trabajador in trabajadores.items():
        stats = {'asistencias_a_tiempo': 0, 'tardanzas': 0, 'faltas': 0, 'justificaciones': 0}
        
        dias_asistidos = set()
        for asistencia in asistencias_por_dni.get(dni, []):
            timestamp_local = asistencia['timestamp'].astimezone(LOCAL_TIMEZONE)
            dias_asistidos.add(timestamp_local.date())
            if timestamp_local.time() > HORA_LIMITE_TARDANZA:
                stats['tardanzas'] += 1
            else:
                stats['asistencias_a_tiempo'] += 1
        
        # --- CÁLCULO PRECISO DE DÍAS JUSTIFICADOS DENTRO DEL RANGO ---
        dias_justificados_en_rango = set()
        for justificacion in justificaciones_por_dni.get(dni, []):
            start_date = justificacion['fechaInicio'].astimezone(LOCAL_TIMEZONE).date()
            end_date = justificacion['fechaFin'].astimezone(LOCAL_TIMEZONE).date()
            
            # Calculamos la intersección del rango de la justificación con el rango del filtro
            rango_justificacion = pd.date_range(start=start_date, end=end_date)
            rango_filtro = pd.date_range(start=fecha_inicio.date(), end=fecha_fin.date())
            dias_comunes = rango_justificacion.intersection(rango_filtro)
            
            for dia_comun in dias_comunes:
                dias_justificados_en_rango.add(dia_comun.date())
        stats['justificaciones'] = len(dias_justificados_en_rango)
        
        # --- CÁLCULO PRECISO DE FALTAS ---
        current_date = fecha_inicio.date()
        while current_date <= fecha_fin.date():
            # Asumimos que no se trabaja los domingos (weekday == 6)
            if current_date.weekday() != 6:
                if current_date not in dias_asistidos and current_date not in dias_justificados_en_rango:
                    stats['faltas'] += 1
            current_date += timedelta(days=1)
        
        reporte_final.append({'dni': dni, 'nombre': data_trabajador.get('nombre'), 'stats': stats})
        
    return reporte_final


@login_required
def asistencias_dashboard(request):
    # --- OBTENER LISTA DE UBICACIONES PARA FILTRO ---
    ubicaciones_disponibles = []
    ubicaciones_horas = {}
    try:
        ubicaciones_ref = db.collection('ubicaciones').order_by('nombre').stream()
        for doc in ubicaciones_ref:
            data = doc.to_dict()
            if 'nombre' in data:
                ubicaciones_disponibles.append(data['nombre'])
                if 'horaEntrada' in data:
                    ubicaciones_horas[data['nombre']] = data['horaEntrada']
    except Exception as e:
        print(f"Error al cargar ubicaciones: {e}")

    # --- CÁLCULO DE MÉTRICAS DEL DASHBOARD ---
    hoy_inicio = datetime.combine(date.today(), time.min).astimezone(LOCAL_TIMEZONE)
    hoy_fin = datetime.combine(date.today(), time.max).astimezone(LOCAL_TIMEZONE)
    total_asistencias_hoy, total_fraudes_hoy = 0, 0
    nombres_asistieron_hoy = []
    try:
        # Métrica de asistencias de hoy
        asistencias_hoy_docs = db.collection('asistencias').where(filter=FieldFilter('timestamp', '>=', hoy_inicio)).where(filter=FieldFilter('timestamp', '<=', hoy_fin)).stream()
        asistencias_hoy_lista = [doc.to_dict() for doc in asistencias_hoy_docs]
        total_asistencias_hoy = len(asistencias_hoy_lista)
        if asistencias_hoy_lista:
            nombres_asistieron_hoy = sorted(list(set([a.get('userName', '') for a in asistencias_hoy_lista])))
        # Métrica de fraudes de hoy
        fraudes_hoy_query = db.collection('asistencias_fraudulentas').where(filter=FieldFilter('timestamp', '>=', hoy_inicio)).where(filter=FieldFilter('timestamp', '<=', hoy_fin)).count().get()
        total_fraudes_hoy = fraudes_hoy_query[0][0].value
    except Exception as e:
        print(f"Error al calcular métricas: {e}")

    # --- LÓGICA DE FILTRADO ---
    filtro_dni = request.GET.get('dni', '').strip()
    filtro_ubicacion = request.GET.get('ubicacion', '')
    filtro_fecha_inicio_str = request.GET.get('fecha_inicio', '')
    filtro_fecha_fin_str = request.GET.get('fecha_fin', '')
    
    # CHIVATO 1: Imprimimos el valor que recibimos del filtro
    # ==========================================================
    print(f"--- Filtro de ubicación recibido del navegador: '{filtro_ubicacion}' ---")

    query = db.collection('asistencias')
    
    if filtro_dni: query = query.where(filter=FieldFilter('userDni', '==', filtro_dni))
    if filtro_ubicacion:
        query = query.where(filter=FieldFilter('locationName', '==', filtro_ubicacion))
    if filtro_fecha_inicio_str:
        fecha_inicio = datetime.strptime(filtro_fecha_inicio_str, '%Y-%m-%d')
        query = query.where(filter=FieldFilter('timestamp', '>=', fecha_inicio))
    if filtro_fecha_fin_str:
        fecha_fin = datetime.strptime(filtro_fecha_fin_str, '%Y-%m-%d')
        fecha_fin_con_hora = datetime.combine(fecha_fin, time.max)
        query = query.where(filter=FieldFilter('timestamp', '<=', fecha_fin_con_hora))

    asistencias_list = []
    try:
        docs = query.order_by('timestamp', direction='DESCENDING').stream()
        print("--- Comparando con los siguientes valores de Firestore: ---")
        
        for i, doc in enumerate(docs): # Usamos enumerate para no imprimir todo
            asistencia = doc.to_dict()

            # CHIVATO 2: Imprimimos los primeros 5 valores de la base de datos
            # ==========================================================
            if i < 5:
                location_name_from_db = asistencia.get('locationName', 'CAMPO NO ENCONTRADO')
                print(f"  Registro #{i+1}: '{location_name_from_db}'")
            
            # --- LÓGICA DE PROCESAMIENTO POR CADA ASISTENCIA ---
            if 'timestamp' in asistencia and asistencia['timestamp']:
                timestamp_utc = asistencia['timestamp']
                timestamp_local = timestamp_utc.astimezone(LOCAL_TIMEZONE)
                asistencia['timestamp_formateado'] = timestamp_local.strftime('%d/%m/%Y %H:%M:%S')
                
                nombre_ubicacion = asistencia.get('locationName')
                hora_entrada_str = ubicaciones_horas.get(nombre_ubicacion, "08:30")
                partes_hora_entrada = hora_entrada_str.split(':')
                hora_entrada = time(hour=int(partes_hora_entrada[0]), minute=int(partes_hora_entrada[1]))
                
                minutos_entrada_con_tolerancia = (hora_entrada.hour * 60) + hora_entrada.minute + TOLERANCIA_TARDANZA_MINUTOS
                minutos_marcacion = (timestamp_local.hour * 60) + timestamp_local.minute
                
                asistencia['estadoPuntualidad'] = "A Tiempo"
                if minutos_marcacion > minutos_entrada_con_tolerancia:
                    asistencia['estadoPuntualidad'] = "Tardanza"

    
            
            asistencias_list.append(asistencia)
    except Exception as e:
        print(f"Error en la consulta de asistencias: {e}")
        messages.error(request, "Error al filtrar los datos. Puede que necesites crear un índice en Firestore (revisa la terminal).")

    # INICIO DE LA NUEVA LÓGICA DE PAGINACIÓN
    # ==========================================================
    paginator = Paginator(asistencias_list, 25) # Muestra 25 asistencias por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # ==========================================================

    context = {
        'total_asistencias_hoy': total_asistencias_hoy,
        'nombres_asistieron_hoy': nombres_asistieron_hoy,
        'total_fraudes_hoy': total_fraudes_hoy,
        'asistencias': page_obj, # <-- CAMBIO: Ya no pasamos la lista completa
        'filtro_dni': filtro_dni,
        'filtro_ubicacion': filtro_ubicacion,
        'ubicaciones_disponibles': ubicaciones_disponibles,
        'filtro_fecha_inicio': filtro_fecha_inicio_str,
        'filtro_fecha_fin': filtro_fecha_fin_str,
    }
    return render(request, 'recursoshumanos/asistencias.html', context)
# --- VISTAS DE GESTIÓN DE TRABAJADORES ---

@login_required
@group_required("Recursos Humanos", "Supervisores", "Calidad")
def lista_trabajadores(request):
    # --- 1. Obtener parámetros de filtrado ---
    search_query = request.GET.get('search', '')
    empresa_id = request.GET.get('empresa')
    proyecto_padre_id = request.GET.get('proyecto_padre')
    subproyecto_id = request.GET.get('subproyecto')
    cargo_id = request.GET.get('cargo')
    estado = request.GET.get('estado')

    # --- 2. Consulta Base ---
    # Usamos prefetch_related para traer las asignaciones (Muchos-a-Muchos)
    trabajadores_queryset = Trabajador.objects.select_related(
        'empresa', 
        'centro_costo'
    ).prefetch_related(
        'asignaciones',             
        'asignaciones__proyecto',   
        'asignaciones__cargo'       
    ).all().order_by('apellido_paterno')

    # --- 3. Aplicar Filtros ---
    
    # Búsqueda por texto
    if search_query:
        trabajadores_queryset = trabajadores_queryset.filter(
            Q(nombres__icontains=search_query) |
            Q(apellido_paterno__icontains=search_query) |
            Q(apellido_materno__icontains=search_query) |
            Q(dni__icontains=search_query)
        )

    if empresa_id:
        trabajadores_queryset = trabajadores_queryset.filter(empresa_id=empresa_id)

    if cargo_id:
        trabajadores_queryset = trabajadores_queryset.filter(
            asignaciones__cargo_id=cargo_id
        ).distinct()

    if proyecto_padre_id:
        trabajadores_queryset = trabajadores_queryset.filter(
            Q(asignaciones__proyecto__parent_id=proyecto_padre_id) | 
            Q(asignaciones__proyecto_id=proyecto_padre_id)
        ).distinct()

    if subproyecto_id:
        trabajadores_queryset = trabajadores_queryset.filter(
            asignaciones__proyecto_id=subproyecto_id
        ).distinct()

    if estado == 'activo':
        trabajadores_queryset = trabajadores_queryset.filter(activo=True)
    elif estado == 'inactivo':
        trabajadores_queryset = trabajadores_queryset.filter(activo=False)

    # --- 4. Datos para los selectores (CORREGIDO AQUÍ) ---
    
    # ERROR ANTERIOR: Empresa.objects.filter(activo=True) -> FALLABA
    # SOLUCIÓN: Usamos .all() porque tu modelo Empresa no tiene campo 'activo'
    opciones_empresas = Empresa.objects.all()
    
    opciones_cargos = Cargo.objects.all().order_by('nombre')
    opciones_proyectos_padre = Proyecto.objects.filter(parent__isnull=True, activo=True)
    
    # Mapa de subproyectos para JS
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent')
    for sub in all_subproyectos:
        if sub.parent_id not in subproyectos_map:
            subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({
            'id': sub.id, 
            'nombre': sub.nombre, 
            'codigo': sub.codigo or ''
        })
    
    import json
    subproyectos_map_json = json.dumps(subproyectos_map)

    # Firestore data placeholder
    firestore_data = {}

    context = {
        'trabajadores': trabajadores_queryset,
        'opciones_empresas': opciones_empresas,
        'opciones_cargos': opciones_cargos,
        'opciones_proyectos_padre': opciones_proyectos_padre,
        'subproyectos_map_json': subproyectos_map_json,
        'firestore_data': firestore_data,
        
        # Filtros activos
        'filtro_busqueda': search_query,
        'filtro_activo_empresa': int(empresa_id) if empresa_id else None,
        'filtro_activo_cargo': int(cargo_id) if cargo_id else None,
        'filtro_activo_proyecto_padre': int(proyecto_padre_id) if proyecto_padre_id else None,
        'filtro_activo_subproyecto': int(subproyecto_id) if subproyecto_id else None,
        'filtro_activo_estado': estado,
    }

    return render(request, 'recursoshumanos/empleados/lista_trabajadores.html', context)
# --- VISTAS DE CREACIÓN Y EDICIÓN DE TRABAJADORES ---

@login_required
@group_required('Recursos Humanos', 'Calidad', 'Supervisores', 'Gerencia')
def gestion_empleados(request):
    """Muestra el dashboard de tarjetas para la gestión de empleados."""
    context = {'current_view': 'gestion_empleados'}
    return render(request, 'recursoshumanos/dashboards/gestion_empleados.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def crear_trabajador(request):
    
    # 1. Capturamos el 'next'
    next_url = request.POST.get('next') or request.GET.get('next') or ''

    if request.method == 'POST':
        form = TrabajadorForm(request.POST)
        if form.is_valid():
            try:
                # --- CAMBIO IMPORTANTE: Eliminamos la creación automática de usuario --   
                nuevo_trabajador = form.save(commit=False)
                
                # El campo 'user' se queda en None (Vacío) por defecto
                nuevo_trabajador.save()
                
                # Guardamos las relaciones ManyToMany (proyectos, etc.)
                form.save_m2m() 

                messages.success(request, f"Trabajador registrado exitosamente (Sin usuario de sistema).")
                
                # 2. Redirección
                if next_url:
                    return redirect(next_url)
                
                return redirect('recursoshumanos:lista_trabajadores')

            except Exception as e:
                messages.error(request, f"Error inesperado al guardar: {e}")
    else:
        form = TrabajadorForm()

    context = {
        'form': form, 
        'action': 'Crear', 
        'current_view': 'gestion_empleados',
        'next_url': next_url
    }
    return render(request, 'recursoshumanos/empleados/crear_trabajador_multistep.html', context)
    
@login_required
@group_required('Recursos Humanos', 'Calidad')
def editar_trabajador(request, pk):
    trabajador = get_object_or_404(Trabajador, pk=pk)
    next_url = request.POST.get('next') or request.GET.get('next') or request.session.get('editar_trabajador_next') or ''

    if next_url and not url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = ''

    if not next_url:
        referer = request.META.get('HTTP_REFERER', '')
        if referer:
            parsed_referer = urlparse(referer)
            mismo_host = (not parsed_referer.netloc) or (parsed_referer.netloc == request.get_host())
            if mismo_host:
                if parsed_referer.path.startswith('/calidad/'):
                    next_url = reverse('calidad:lista_trabajadores')
                elif parsed_referer.path.startswith('/recursoshumanos/'):
                    next_url = reverse('recursoshumanos:lista_trabajadores')

    if not next_url:
        es_rrhh = request.user.groups.filter(name='Recursos Humanos').exists()
        es_calidad = request.user.groups.filter(name='Calidad').exists()

        if es_calidad and not es_rrhh:
            next_url = reverse('calidad:lista_trabajadores')
        else:
            next_url = reverse('recursoshumanos:lista_trabajadores')

    request.session['editar_trabajador_next'] = next_url
    
    if request.method == 'POST':
        form = TrabajadorForm(request.POST, instance=trabajador)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Trabajador actualizado con éxito.")
                request.session.pop('editar_trabajador_next', None)
                return redirect(next_url)
            except Exception as e:
                messages.error(request, f"Ocurrió un error al actualizar: {e}")
    else:
        form = TrabajadorForm(instance=trabajador)
    
    context = {
        'form': form,
        'action': 'Editar',
        'current_view': 'gestion_empleados',
        'trabajador': trabajador,
        'next_url': next_url,
    }
    return render(request, 'recursoshumanos/empleados/crear_trabajador_multistep.html', context)


@login_required
@group_required("Recursos Humanos", "Supervisores")
def info_trabajador(request):
    trabajador = None
    ultimo_emo_realizado = None
    dni_buscado = request.GET.get('dni', '')
    
    if dni_buscado:
        try:
            trabajador = Trabajador.objects.get(dni=dni_buscado)
            
            # --- ¡LÓGICA CORREGIDA! ---
            # Ordenamos primero por fecha, y luego por ID para desempatar.
            ultimo_emo_realizado = trabajador.historial_emo.filter(
                estado='Realizado'
            ).order_by('-fecha_realizacion', '-pk').first() # <-- Se añade '-pk'
            
        except Trabajador.DoesNotExist:
            messages.error(request, f"No se encontró ningún trabajador con el DNI {dni_buscado}.")
    
    context = {
        'trabajador': trabajador,
        'ultimo_emo_realizado': ultimo_emo_realizado,
        'dni_buscado': dni_buscado,
        'current_view': 'info_trabajador'
    }
    return render(request, 'recursoshumanos/empleados/info_trabajador.html', context)

# --- VISTAS DE REPORTES Y ANÁLISIS ---

@login_required
@group_required("Recursos Humanos")
def trabajador_detail_view(request, dni):
    """
    Muestra la ficha detallada de un trabajador con su historial de asistencias
    y fraudes, obteniendo los datos desde PostgreSQL.
    """
    trabajador = get_object_or_404(Trabajador, dni=dni)

    # Obtenemos las asistencias relacionadas con el TRABAJADOR, no con el USER
    # (limitadas a las 100 más recientes para no cargar años de marcaciones).
    if trabajador.user_id:
        asistencias = trabajador.user.asistencias.all().order_by('-timestamp')[:100]
    else:
        asistencias = []

    fraudes = trabajador.alertas_fraude.all().order_by('-fecha_hora')[:50]

    context = {
        'trabajador': trabajador,
        'asistencias': asistencias,
        'fraudes': fraudes,
    }
    return render(request, 'recursoshumanos/empleados/trabajador_detail.html', context)


@login_required
def ranking_tardanzas(request):
    """ Calcula y muestra el ranking de tardanzas para un mes seleccionado. """
    HORA_LIMITE = HORA_LIMITE_TARDANZA_OFICINA
    mes_seleccionado_str = request.GET.get('mes', datetime.now().strftime('%Y-%m'))
    
    tardanzas_por_trabajador = Counter()
    mes_nombre = ""
    try:
        mes_dt = datetime.strptime(mes_seleccionado_str, '%Y-%m')
        mes_nombre = mes_dt.strftime('%B de %Y').capitalize()
        
        primer_dia = mes_dt.replace(day=1).astimezone(LOCAL_TIMEZONE)
        num_dias = calendar.monthrange(mes_dt.year, mes_dt.month)[1]
        ultimo_dia = mes_dt.replace(day=num_dias, hour=23, minute=59, second=59).astimezone(LOCAL_TIMEZONE)

        query = db.collection('asistencias').where(filter=FieldFilter('timestamp', '>=', primer_dia)).where(filter=FieldFilter('timestamp', '<=', ultimo_dia))
        
        docs = query.stream()
        for doc in docs:
            data = doc.to_dict()
            timestamp_utc = data.get('timestamp')
            if timestamp_utc:
                hora_marcacion_local = timestamp_utc.astimezone(LOCAL_TIMEZONE)
                if hora_marcacion_local.time() > HORA_LIMITE:
                    nombre = data.get('userName', 'Desconocido')
                    tardanzas_por_trabajador[nombre] += 1
    except Exception as e:
        messages.error(request, f"Error al procesar las tardanzas: {e}")

    ranking = tardanzas_por_trabajador.most_common()
    labels = [item[0] for item in ranking]
    data_values = [item[1] for item in ranking]

    context = {
        'ranking': ranking,
        'chart_labels': json.dumps(labels),
        'chart_data': json.dumps(data_values),
        'mes_seleccionado': mes_seleccionado_str,
        'mes_nombre': mes_nombre,
    }
    return render(request, 'recursoshumanos/ranking_tardanzas.html', context)


@login_required
def exportar_asistencias_csv(request):
    """ Exporta las asistencias filtradas a un archivo CSV. """
    filtro_dni = request.GET.get('dni', '').strip()
    filtro_fecha_inicio_str = request.GET.get('fecha_inicio', '')
    filtro_fecha_fin_str = request.GET.get('fecha_fin', '')

    
    
    response = HttpResponse(content_type='text/csv', headers={'Content-Disposition': 'attachment; filename="reporte_asistencias.csv"'})
    
    writer = csv.writer(response)
    writer.writerow(['Nombre', 'DNI', 'Fecha y Hora', 'Ubicación', 'Estado', 'Device ID'])
    
    query = db.collection('asistencias')
    if filtro_dni:
        query = query.where(filter=FieldFilter('userDni', '==', filtro_dni))
    if filtro_fecha_inicio_str:
        fecha_inicio = datetime.strptime(filtro_fecha_inicio_str, '%Y-%m-%d')
        query = query.where(filter=FieldFilter('timestamp', '>=', fecha_inicio))
    if filtro_fecha_fin_str:
        fecha_fin = datetime.strptime(filtro_fecha_fin_str, '%Y-%m-%d')
        fecha_fin_con_hora = datetime.combine(fecha_fin, time.max)
        query = query.where(filter=FieldFilter('timestamp', '<=', fecha_fin_con_hora))

    try:
        docs = query.order_by('timestamp', direction='DESCENDING').stream()
        for doc in docs:
            asistencia = doc.to_dict()
            fecha_formateada = ''
            if 'timestamp' in asistencia and asistencia['timestamp']:
                timestamp_utc = asistencia['timestamp']
                timestamp_local = timestamp_utc.astimezone(LOCAL_TIMEZONE)
                fecha_formateada = timestamp_local.strftime('%d/%m/%Y %H:%M:%S')
            
            writer.writerow([
                asistencia.get('userName', ''),
                asistencia.get('userDni', ''),
                fecha_formateada,
                asistencia.get('locationName', ''),
                asistencia.get('status', ''),
                asistencia.get('deviceId', '')
            ])
    except Exception as e:
        print(f"Error al exportar a CSV: {e}")

    return response

@login_required
def reportes_fraude_view(request):
    print("\n" + "="*50)
    print(" 🕵️‍♂️ [DEBUG FRAUDES] INICIANDO CARGA DE VISTA ")
    print("="*50)
    
    try:
        print("[DEBUG] 1. Ejecutando consulta a la base de datos (IntentoFraude)...")
        fraudes_list = IntentoFraude.objects.select_related('trabajador').all()
        
        # Evaluamos la consulta contando los registros
        total_fraudes = fraudes_list.count()
        print(f"[DEBUG] 2. Consulta exitosa. Registros encontrados: {total_fraudes}")
        
        # Imprimimos una pequeña muestra (máximo 3) para ver qué hay adentro
        if total_fraudes > 0:
            print("[DEBUG] 3. Mostrando muestra de datos:")
            for f in fraudes_list[:3]:
                print(f"    -> DNI: {f.trabajador.dni} | Motivo: {f.motivo_detectado} | Fecha: {f.fecha_hora}")
        else:
            print("[DEBUG] 3. La tabla existe, pero está VACÍA. (Aún no se ha guardado ningún fraude).")

        fuentes_detectadas = [f"Base de Datos Local ({total_fraudes})"]

    except Exception as e:
        print(f"\n[DEBUG] ❌ ERROR CRÍTICO AL CARGAR LA TABLA:")
        print(f"    -> Tipo de error: {type(e).__name__}")
        print(f"    -> Mensaje: {e}\n")
        print("[DEBUG] --- TRAZA EXACTA DEL ERROR ---")
        traceback.print_exc()  # Esto imprimirá la línea exacta del error
        print("--------------------------------------\n")
        
        messages.error(request, "No se pudieron cargar los reportes de fraude.")
        fraudes_list = []
        fuentes_detectadas = []

    print("="*50 + "\n")
    
    context = {
        'fraudes': fraudes_list,
        'fuentes_fraude': fuentes_detectadas,
    }
    return render(request, 'recursoshumanos/reportes_fraude.html', context)
#========================================================================d=========
# --- VISTAS DE GESTIÓN DE UBICACIONES ---
#=================================================================================

@login_required
def gestion_ubicaciones(request):
    """ Muestra, crea, edita y elimina Ubicaciones usando modelos de Django. """
    
    if request.method == 'POST':
        # Lógica para eliminar una ubicación
        if 'delete_id' in request.POST:
            ubicacion_a_eliminar = get_object_or_404(Ubicacion, pk=request.POST.get('delete_id'))
            ubicacion_a_eliminar.delete()
            messages.success(request, f"Ubicación '{ubicacion_a_eliminar.nombre}' eliminada con éxito.")
            return redirect('recursoshumanos:gestion_ubicaciones') # Asegúrate de que el nombre de la ruta sea correcto
        
        # Lógica para crear o editar (esto requeriría un formulario,
        # pero por ahora lo dejamos simple. Lo puedes hacer desde el admin).

    # Obtenemos todas las ubicaciones de la base de datos PostgreSQL
    ubicaciones_list = Ubicacion.objects.all().order_by('nombre')

    context = {
        'ubicaciones': ubicaciones_list,
        'current_view': 'gestion_ubicaciones'
    }
    return render(request, 'recursoshumanos/ubicaciones/gestion_ubicaciones.html', context)

@login_required
def lista_ubicaciones(request):
    """ Muestra la lista de todas las ubicaciones. """
    ubicaciones_list = Ubicacion.objects.all().order_by('nombre')
    context = {
        'ubicaciones': ubicaciones_list,
        'current_view': 'lista_ubicaciones'
    }
    return render(request, 'recursoshumanos/ubicaciones/lista_ubicaciones.html', context)

@login_required
@require_POST # Es más seguro usar este decorador para acciones de borrado
def ubicacion_eliminar(request, pk):
    """ Elimina una ubicación específica de PostgreSQL. """
    ubicacion = get_object_or_404(Ubicacion, pk=pk)
    nombre_ubicacion = ubicacion.nombre
    ubicacion.delete()
    messages.success(request, f"Ubicación '{nombre_ubicacion}' eliminada con éxito.")
    return redirect('recursoshumanos:gestion_ubicaciones')

@login_required
def ubicacion_crear(request):
        """ Maneja el formulario para añadir una nueva ubicación. """
        if request.method == 'POST':
            form = UbicacionForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Ubicación creada con éxito.")
                return redirect('recursoshumanos:gestion_ubicaciones')
        else:
            form = UbicacionForm()
        
        return render(request, 'recursoshumanos/ubicaciones/ubicacion_form_multistep.html', {'form': form, 'action': 'Añadir'})


@login_required
def ubicacion_editar(request, pk):
    """ Maneja el formulario para editar una ubicación existente (CON DEPURACIÓN). """
    
    print(f"\n--- [EDITAR UBICACIÓN] Iniciando vista para PK={pk} ---")
    ubicacion = get_object_or_404(Ubicacion, pk=pk)

    if request.method == 'POST':
        print("[EDITAR UBICACIÓN] Se recibió una petición POST.")
        
        # 1. Imprimimos los datos crudos que llegan del formulario
        print(f"[EDITAR UBICACIÓN] request.POST data: {request.POST}")
        
        form = UbicacionForm(request.POST, instance=ubicacion)

        # 2. Verificamos si el formulario es válido
        if form.is_valid():
            print("[EDITAR UBICACIÓN] El formulario ES VÁLIDO. Guardando cambios...")
            form.save()
            messages.success(request, f"Ubicación '{ubicacion.nombre}' actualizada con éxito.")
            print("[EDITAR UBICACIÓN] Redirigiendo a la lista de ubicaciones.")
            return redirect('recursoshumanos:gestion_ubicaciones')
        else:
            # 3. Si el formulario NO es válido, imprimimos los errores
            print("[EDITAR UBICACIÓN] ¡¡¡FALLO DE VALIDACIÓN!!! El formulario NO ES VÁLIDO.")
            print(f"[EDITAR UBICACIÓN] Errores del formulario: {form.errors.as_json()}")
            
            # También puedes mostrar los errores como mensajes en la plantilla
            # para que el usuario sepa qué corregir.
            for field, errors in form.errors.items():
                for error in errors:
                    # El prefijo 'Error en el campo...' ayuda a distinguirlos de los mensajes de éxito
                    messages.error(request, f"Error en el campo '{field}': {error}")
    else:
        print("[EDITAR UBICACIÓN] Se recibió una petición GET. Mostrando formulario para editar.")
        form = UbicacionForm(instance=ubicacion)

    context = {
        'form': form,
        'action': 'Editar',
        # Asegúrate de pasar la pk a la plantilla si la necesitas para la URL del POST
        'ubicacion': ubicacion, 
    }
    return render(request, 'recursoshumanos/ubicaciones/ubicacion_form_multistep.html', context)

@login_required
@group_required('Recursos Humanos', 'Supervisores')
def asignar_ubicaciones_trabajador(request, pk): # Ahora usamos la Primary Key (pk) del trabajador
    """ Página para asignar ubicaciones a un trabajador específico usando datos de PostgreSQL. """
    trabajador = get_object_or_404(Trabajador, pk=pk)
    
    # Obtenemos las ubicaciones ya asignadas a través de la relación ManyToManyField
    ubicaciones_asignadas = trabajador.ubicaciones_permitidas.all().order_by('nombre')
    
    # Obtenemos las ubicaciones que NO están asignadas
    ubicaciones_disponibles = Ubicacion.objects.exclude(pk__in=ubicaciones_asignadas.values_list('pk', flat=True)).order_by('nombre')

    context = {
        'trabajador': trabajador,
        'ubicaciones_asignadas': ubicaciones_asignadas,
        'ubicaciones_disponibles': ubicaciones_disponibles,
        'current_view': 'gestion_empleados'
    }
    return render(request, 'recursoshumanos/empleados/asignar_ubicaciones.html', context)



@login_required
def actualizar_ubicaciones_trabajador(request, pk): # Usamos la PK del trabajador
    if request.method == 'POST':
        try:
            trabajador = get_object_or_404(Trabajador, pk=pk)
            data = json.loads(request.body)
            nuevas_ubicaciones_ids = data.get('ubicaciones_ids', [])
            
            # set() es la forma más eficiente de actualizar una relación ManyToManyField
            trabajador.ubicaciones_permitidas.set(nuevas_ubicaciones_ids)
            
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

@login_required
@group_required("Recursos Humanos", "Supervisores")
def asignacion_masiva_ubicaciones(request):
    """ Página para asignar trabajadores a una ubicación específica usando datos de PostgreSQL. """
    
    # Obtenemos todas las ubicaciones para el selector
    todas_las_ubicaciones = Ubicacion.objects.all().order_by('nombre')
    selected_ubicacion_id = request.GET.get('ubicacion_id')
    
    trabajadores_asignados = Trabajador.objects.none() # QuerySet vacío por defecto
    trabajadores_disponibles = Trabajador.objects.filter(activo=True).order_by('nombres') # Todos los activos por defecto
    selected_ubicacion = None

    if selected_ubicacion_id:
        try:
            # Usamos get_object_or_404 para más seguridad
            selected_ubicacion = get_object_or_404(Ubicacion, pk=selected_ubicacion_id)
            
            # Obtenemos los trabajadores ya asignados a esta ubicación.
            # prefetch_related es más eficiente para relaciones ManyToMany.
            trabajadores_asignados = selected_ubicacion.trabajadores.filter(activo=True).order_by('nombres').prefetch_related('user')
            
            # Obtenemos los trabajadores activos que NO están en la lista de asignados.
            trabajadores_disponibles = Trabajador.objects.filter(activo=True).exclude(pk__in=trabajadores_asignados.values_list('pk', flat=True)).order_by('nombres')

        except (ValueError, TypeError):
             # Esto ocurre si selected_ubicacion_id no es un número válido
             messages.error(request, "El ID de la ubicación no es válido.")

    context = {
        'todas_las_ubicaciones': todas_las_ubicaciones,
        'selected_ubicacion': selected_ubicacion,
        'trabajadores_asignados': trabajadores_asignados,
        'trabajadores_disponibles': trabajadores_disponibles,
        'current_view': 'asignar_ubicaciones'
    }
    return render(request, 'recursoshumanos/ubicaciones/asignacion_masiva_ubicaciones.html', context)

@login_required
@require_POST # Aseguramos que solo sea POST
def actualizar_asignacion_ubicacion(request):
    """ Endpoint AJAX para añadir o quitar un trabajador de una ubicación. """
    try:
        data = json.loads(request.body)
        worker_pk = data.get('worker_pk')
        ubicacion_id = data.get('ubicacion_id')
        accion = data.get('action')

        if not all([worker_pk, ubicacion_id, accion]):
            return JsonResponse({'status': 'error', 'message': 'Faltan datos.'}, status=400)

        ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)
        trabajador = get_object_or_404(Trabajador, pk=worker_pk)

        if accion == 'add':
            ubicacion.trabajadores.add(trabajador)
        elif accion == 'remove':
            ubicacion.trabajadores.remove(trabajador)
        else:
            return JsonResponse({'status': 'error', 'message': 'Acción no válida.'}, status=400)
        
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

#=================================================================================
# --- VISTAS DE GESTIÓN DE TOLERANCIA DE HORARIO (HU-06 / CAV-15) ---
#=================================================================================

@login_required
@group_required("Recursos Humanos", "Gerencia", "Administracion")
def gestion_tolerancia(request):
    """
    Pantalla administrativa (CAV-72) para configurar los minutos de tolerancia
    de tardanza por Sede y horario/turno. La edición se hace de forma visual
    mediante AJAX contra los endpoints API de CAV-71 (sin recargar la página),
    para que el cambio surta efecto de inmediato en el cálculo de asistencia
    (CAV-154), sin reiniciar el servidor.
    """
    configuraciones = listar_tolerancias().order_by('sede__nombre', 'tipo_horario')
    sedes = Sede.objects.filter(activo=True).order_by('nombre')
    auditorias_recientes = ToleranciaAuditoria.objects.select_related('usuario').all()[:20]

    context = {
        'configuraciones': configuraciones,
        'sedes': sedes,
        'tipos_horario': ConfiguracionTolerancia.TipoHorario.choices,
        'auditorias_recientes': auditorias_recientes,
        'current_view': 'gestion_tolerancia',
    }
    return render(request, 'recursoshumanos/tolerancia/gestion_tolerancia.html', context)

#vista para reporte mensual

@login_required
def reporte_mensual(request):
    trabajadores = []
    try:
        trabajadores_ref = db.collection('trabajadores').order_by('nombre').stream()
        trabajadores = [{'dni': doc.id, 'nombre': doc.to_dict().get('nombre')} for doc in trabajadores_ref]
    except Exception as e:
        messages.error(request, f"No se pudieron cargar los trabajadores: {e}")

    dni_seleccionado = request.GET.get('dni', '')
    mes_seleccionado_str = request.GET.get('mes', datetime.now().strftime('%Y-%m'))
    
    # Inicializamos todas las variables
    reporte_dias = {}
    dias_del_mes_range = range(0)
    mes_nombre = ""
    total_faltas = 0
    total_tardanzas = 0
    total_justificados = 0
    asistencias_por_dia = {}
    dias_laborables = 0
    
    if dni_seleccionado and mes_seleccionado_str:
        try:
            mes_dt = datetime.strptime(mes_seleccionado_str, '%Y-%m')
            mes_nombre = mes_dt.strftime('%B de %Y').capitalize()
            
            primer_dia_naive = mes_dt.replace(day=1)
            num_dias = calendar.monthrange(mes_dt.year, mes_dt.month)[1]
            ultimo_dia_naive = primer_dia_naive.replace(day=num_dias, hour=23, minute=59, second=59)
            
            primer_dia_aware = LOCAL_TIMEZONE.localize(primer_dia_naive)
            ultimo_dia_aware = LOCAL_TIMEZONE.localize(ultimo_dia_naive)

            dias_del_mes_range = range(1, num_dias + 1)

            # ==========================================================
            # CORRECCIÓN: CONSULTA DE JUSTIFICACIONES SIMPLIFICADA
            # ==========================================================
            justificaciones_docs = db.collection('justificaciones').where(
                filter=FieldFilter('dniTrabajador', '==', dni_seleccionado)
            ).where(
                filter=FieldFilter('fechaInicio', '<=', ultimo_dia_aware) # Un solo filtro de rango
            ).stream()
            
            dias_justificados = {}
            for doc in justificaciones_docs:
                data = doc.to_dict()
                # El segundo filtro de rango lo hacemos en Python
                if data['fechaFin'] >= primer_dia_aware:
                    start_date = data['fechaInicio'].astimezone(LOCAL_TIMEZONE).date()
                    end_date = data['fechaFin'].astimezone(LOCAL_TIMEZONE).date()
                    
                    current_date = start_date
                    while current_date <= end_date:
                        if current_date.year == mes_dt.year and current_date.month == mes_dt.month:
                            dias_justificados[current_date.day] = data['tipo']
                        current_date += timedelta(days=1)
            # ==========================================================

            asistencias_docs = db.collection('asistencias').where(filter=FieldFilter('userDni', '==', dni_seleccionado)).where(filter=FieldFilter('timestamp', '>=', primer_dia_aware)).where(filter=FieldFilter('timestamp', '<=', ultimo_dia_aware)).stream()
            for doc in asistencias_docs:
                data = doc.to_dict()
                timestamp = data.get('timestamp')
                if timestamp:
                    dia = timestamp.astimezone(LOCAL_TIMEZONE).day
                    if dia not in asistencias_por_dia: asistencias_por_dia[dia] = timestamp

            HORA_LIMITE_TARDANZA = HORA_LIMITE_TARDANZA_OFICINA
            hoy_local = date.today()

            for dia in dias_del_mes_range:
                fecha_actual = date(mes_dt.year, mes_dt.month, dia)
                
                # --- NUEVO ORDEN DE PRIORIDADES ---
                
                # Prioridad 1: ¿Es Domingo?
                if fecha_actual.weekday() == 6:
                    reporte_dias[dia] = {'estado': 'Domingo', 'texto': 'No Laborable'}
                
                # Prioridad 2: ¿Está Justificado? (Esto se aplica incluso a días futuros)
                elif dia in dias_justificados:
                    reporte_dias[dia] = {'estado': 'Justificado', 'texto': dias_justificados[dia]}
                    total_justificados += 1
                    # Contamos como día laborable si está justificado (excepto domingos)
                    if fecha_actual.weekday() != 6:
                        dias_laborables += 1
                
                # Prioridad 3: ¿Es un día futuro (y no está justificado)?
                elif fecha_actual > hoy_local:
                    reporte_dias[dia] = {'estado': 'Futuro', 'texto': ''}
                
                # Si hemos llegado hasta aquí, es un día laborable del pasado o presente
                else:
                    dias_laborables += 1 # Es un día que ya debería haberse trabajado
                    
                    # Prioridad 4: ¿Asistió (Tarde o a Tiempo)?
                    if dia in asistencias_por_dia:
                        timestamp_asistencia = asistencias_por_dia[dia].astimezone(LOCAL_TIMEZONE)
                        if timestamp_asistencia.time() > HORA_LIMITE_TARDANZA:
                            reporte_dias[dia] = {'estado': 'Tardanza', 'texto': f"Tarde ({timestamp_asistencia.strftime('%H:%M')})"}
                            total_tardanzas += 1
                        else:
                            reporte_dias[dia] = {'estado': 'Asistio', 'texto': f"Asistió ({timestamp_asistencia.strftime('%H:%M')})"}
                    
                    # Prioridad 5: Si no, es Falta
                    else:
                        reporte_dias[dia] = {'estado': 'Falta', 'texto': 'Faltó'}
                        total_faltas += 1

        except Exception as e:
            messages.error(request, f"Error al generar el reporte: {e}")


    context = {
        'trabajadores': trabajadores, 'dni_seleccionado': dni_seleccionado,
        'mes_seleccionado': mes_seleccionado_str, 'mes_nombre': mes_nombre,
        'reporte_dias': reporte_dias, 'dias_del_mes': dias_del_mes_range,
        'total_faltas': total_faltas, 'total_tardanzas': total_tardanzas,
        'total_justificados': total_justificados, 'total_asistencias': len(asistencias_por_dia),
        'dias_laborables': dias_laborables,
    }
    return render(request, 'recursoshumanos/reportes/reporte_mensual.html', context)

@login_required
def gestion_justificaciones(request):
    """ Muestra la lista de justificaciones y maneja la creación y eliminación. """
    
    # Manejo de eliminación
    if request.method == 'POST' and 'delete_id' in request.POST:
        doc_id_to_delete = request.POST.get('delete_id')
        try:
            db.collection('justificaciones').document(doc_id_to_delete).delete()
            messages.success(request, "Justificación eliminada con éxito.")
        except Exception as e:
            messages.error(request, f"Error al eliminar: {e}")
        return redirect('gestion_justificaciones')

    # Obtenemos la lista de trabajadores para el formulario
    trabajadores_choices = []
    try:
        trabajadores_ref = db.collection('trabajadores').order_by('nombre').stream()
        trabajadores_choices = [(doc.id, doc.to_dict().get('nombre')) for doc in trabajadores_ref]
    except Exception as e:
        messages.error(request, f"No se pudieron cargar los trabajadores: {e}")

    # Manejo de creación
    if request.method == 'POST':
        form = JustificacionForm(request.POST)
        form.fields['trabajador_dni'].choices = trabajadores_choices
        if form.is_valid():
            data = form.cleaned_data
            try:
                # Obtenemos el nombre del trabajador a partir del DNI seleccionado
                nombre_trabajador = dict(trabajadores_choices).get(data['trabajador_dni'])
                
                # --- CORRECCIÓN DE ZONA HORARIA ---
                # --- CORRECCIÓN DE ZONA HORARIA ---
                # Hacemos que las fechas sean "conscientes" de la zona horaria local ANTES de guardarlas
                fecha_inicio_naive = datetime.combine(data['fechaInicio'], time.min)
                fecha_fin_naive = datetime.combine(data['fechaFin'], time.max)
                
                fecha_inicio_aware = LOCAL_TIMEZONE.localize(fecha_inicio_naive)
                fecha_fin_aware = LOCAL_TIMEZONE.localize(fecha_fin_naive)


                db.collection('justificaciones').add({
                    'dniTrabajador': data['trabajador_dni'],
                    'nombreTrabajador': nombre_trabajador,
                    'fechaInicio': fecha_inicio_aware, # Guardamos la fecha con zona horaria
                    'fechaFin': fecha_fin_aware,
                    'tipo': data['tipo'],
                    'descripcion': data['descripcion'],
                    'creadoEn': firestore.SERVER_TIMESTAMP
                })
                messages.success(request, "Justificación añadida con éxito.")
                return redirect('gestion_justificaciones')
            except Exception as e:
                messages.error(request, f"Error al guardar: {e}")
    else:
        form = JustificacionForm()
        form.fields['trabajador_dni'].choices = [(None, '-- Seleccione un trabajador --')] + trabajadores_choices

    # Obtenemos la lista de justificaciones existentes para mostrar en la tabla
    justificaciones_list = []
    try:
        docs = db.collection('justificaciones').order_by('fechaInicio', direction='DESCENDING').stream()
        for doc in docs:
            justificacion = doc.to_dict()
            justificacion['id'] = doc.id
            if 'fechaInicio' in justificacion:
                justificacion['fechaInicio'] = justificacion['fechaInicio'].strftime('%d/%m/%Y')
            if 'fechaFin' in justificacion:
                justificacion['fechaFin'] = justificacion['fechaFin'].strftime('%d/%m/%Y')
            justificaciones_list.append(justificacion)
    except Exception as e:
        messages.error(request, f"Error al cargar justificaciones: {e}")
        
    context = {
        'form': form,
        'justificaciones': justificaciones_list,
    }
    return render(request, 'recursoshumanos/gestion_justificaciones.html', context)

@login_required
def main_dashboard(request):
    """
    Vista para el DASHBOARD DE TARJETAS. Filtra las tarjetas según el
    parámetro 'view' de la URL.
    """
    # El valor por defecto no importa mucho, pero 'gestion_marcas' es un buen comienzo.
    current_view = request.GET.get('view', 'gestion_marcas')
    
    context = {
        'current_view': current_view
    }
    return render(request, 'recursoshumanos/dashboards/main_dashboard.html', context)

# PEGA ESTA NUEVA FUNCIÓN COMPLETA
@login_required
def estadisticas_general(request):
    current_view = request.GET.get('view', 'dashboard')
    context = {'current_view': current_view}

    if current_view == 'dashboard':
        mes_str = request.GET.get('mes', datetime.now().strftime('%Y-%m'))
        try:
            mes_dt = datetime.strptime(mes_str, '%Y-%m')
        except ValueError:
            mes_dt = datetime.now()

        primer_dia = LOCAL_TIMEZONE.localize(mes_dt.replace(day=1, hour=0, minute=0))
        num_dias_mes = calendar.monthrange(mes_dt.year, mes_dt.month)[1]
        ultimo_dia = LOCAL_TIMEZONE.localize(mes_dt.replace(day=num_dias_mes, hour=23, minute=59, second=59))

        trabajadores_ref = db.collection('trabajadores').where(filter=FieldFilter('activo', '==', True)).stream()
        trabajadores_fs = {doc.id: doc.to_dict() for doc in trabajadores_ref}

        # --- AGREGA A LOS TRABAJADORES LOCALES ACTIVOS PARA QUE NO SE EXCLUYAN ---
        for t in Trabajador.objects.filter(activo=True):
            if t.dni and t.dni not in trabajadores_fs:
                trabajadores_fs[t.dni] = {'nombre': t.nombre_completo}

        asistencias_mes_docs = db.collection('asistencias').where(filter=FieldFilter('timestamp', '>=', primer_dia)).where(filter=FieldFilter('timestamp', '<=', ultimo_dia)).stream()
        fraudes_mes = list(db.collection('asistencias_fraudulentas').where(filter=FieldFilter('timestamp', '>=', primer_dia)).where(filter=FieldFilter('timestamp', '<=', ultimo_dia)).stream())

        tareo_doc = db.collection('tareos').document(mes_dt.strftime('%Y-%m')).get()
        tareo_del_mes = tareo_doc.to_dict().get('trabajadores', {}) if tareo_doc.exists else {}

        # --- NUEVO: Traer tareos locales (SQL) e integrarlos ---
        tareos_locales = TareoDiario.objects.filter(
            fecha__gte=primer_dia.date(),
            fecha__lte=ultimo_dia.date()
        ).select_related('trabajador')

        for t in tareos_locales:
            if not t.trabajador or not t.trabajador.dni:
                continue
            dni = t.trabajador.dni
            dia_str = str(t.fecha.day)

            if dni not in tareo_del_mes:
                tareo_del_mes[dni] = {}
            
            # Cargar los datos del tareo local en el formato que espera el dashboard
            tareo_del_mes[dni][dia_str] = {
                'estado': t.estado,
                'jornada_horas': float(t.jornada_horas) if t.jornada_horas else 8,
                'horario': {
                    'entrada': t.hora_entrada.strftime('%H:%M') if t.hora_entrada else '00:00',
                    'salida': t.hora_salida.strftime('%H:%M') if t.hora_salida else '00:00'
                }
            }

        asistencias_locales = Asistencia.objects.filter(
            timestamp__gte=primer_dia,
            timestamp__lte=ultimo_dia,
        ).select_related('usuario')

        dni_por_user_id = {
            t.user_id: t.dni
            for t in Trabajador.objects.filter(user__isnull=False, activo=True)
        }

        trabajadores_nombre = {
            dni: data.get('nombre', dni)
            for dni, data in trabajadores_fs.items()
        }

        def clasificar_origen_firestore(registro):
            status = str(registro.get('status', '')).lower()
            location = str(registro.get('locationName', '')).lower()
            device = str(registro.get('deviceId', '')).lower()

            if 'import' in status or 'historico' in location or 'histórico' in location:
                return 'Excel importado'
            if 'biometr' in location or 'huell' in location:
                return 'Biometrico/Huellero'
            if device:
                return 'App movil'
            return 'Biometrico/Huellero'

        def clasificar_origen_local(asistencia):
            location = str(asistencia.nombre_ubicacion or '').lower()
            if 'biometr' in location or 'huell' in location:
                return 'Biometrico/Huellero'
            if asistencia.device_id:
                return 'App movil'
            return 'Biometrico/Huellero'

        marcas_unificadas = {}

        for doc in asistencias_mes_docs:
            data = doc.to_dict()
            dni = data.get('userDni')
            timestamp = data.get('timestamp')
            if not dni or not timestamp:
                continue

            ts_local = timestamp.astimezone(LOCAL_TIMEZONE)
            tipo = data.get('tipoMarcacion') or data.get('tipo_marcacion') or 'Entrada'
            origen = clasificar_origen_firestore(data)
            nombre = data.get('userName') or trabajadores_nombre.get(dni, dni)

            key = (dni, ts_local.strftime('%Y-%m-%d %H:%M'), tipo)
            marcas_unificadas[key] = {
                'dni': dni,
                'nombre': nombre,
                'timestamp': ts_local,
                'tipo': tipo,
                'origen': origen,
            }

        for asistencia in asistencias_locales:
            dni = dni_por_user_id.get(asistencia.usuario_id) or getattr(asistencia.usuario, 'username', None)
            if not dni:
                continue

            ts_local = asistencia.timestamp.astimezone(LOCAL_TIMEZONE)
            tipo = asistencia.tipo_marcacion or 'Entrada'
            origen = clasificar_origen_local(asistencia)

            nombre = trabajadores_nombre.get(dni)
            if not nombre:
                try:
                    trabajador = Trabajador.objects.only('dni', 'nombres', 'apellido_paterno', 'apellido_materno').get(dni=dni)
                    nombre = trabajador.nombre_completo
                except Trabajador.DoesNotExist:
                    nombre = dni

            key = (dni, ts_local.strftime('%Y-%m-%d %H:%M'), tipo)
            if key not in marcas_unificadas:
                marcas_unificadas[key] = {
                    'dni': dni,
                    'nombre': nombre,
                    'timestamp': ts_local,
                    'tipo': tipo,
                    'origen': origen,
                }

        asistencias_agrupadas = defaultdict(list)
        for evento in marcas_unificadas.values():
            dni = evento['dni']
            if dni in trabajadores_fs:
                dia = evento['timestamp'].day
                asistencias_agrupadas[(dni, dia)].append({
                    'timestamp': evento['timestamp'],
                    'tipoMarcacion': evento['tipo'],
                })

        faltas_por_trabajador = defaultdict(int)
        tardanzas_por_trabajador = defaultdict(timedelta)
        tardanzas_count_por_trabajador = defaultdict(int)
        asistencias_puntuales = defaultdict(int)
        balance_horas_extra = defaultdict(timedelta)
        balance_horas_faltantes = defaultdict(timedelta)
        cumplimiento_estado = {'Puntual': 0, 'Con tardanza': 0, 'Falta': 0}
        total_dias_laborables = 0
        total_dias_puntuales = 0

        for dni, data_trabajador in trabajadores_fs.items():
            nombre_trabajador = data_trabajador.get('nombre', dni)
            tareo_trabajador = tareo_del_mes.get(dni, {})
            for dia in range(1, num_dias_mes + 1):
                tareo_dia = tareo_trabajador.get(str(dia), {})
                if tareo_dia.get('estado') in ['C', 'O', 'P', 'J']:
                    total_dias_laborables += 1
                    marcaciones_dia = asistencias_agrupadas.get((dni, dia), [])
                    resumen_diario = calcular_resumen_diario(marcaciones_dia, tareo_dia)

                    if resumen_diario['estado_celda'] == 'F':
                        faltas_por_trabajador[nombre_trabajador] += 1
                        cumplimiento_estado['Falta'] += 1
                    elif resumen_diario['hrs_tardanza'].total_seconds() > 0:
                        tardanzas_count_por_trabajador[nombre_trabajador] += 1
                        cumplimiento_estado['Con tardanza'] += 1
                    else:
                        cumplimiento_estado['Puntual'] += 1

                    tardanzas_por_trabajador[nombre_trabajador] += resumen_diario['hrs_tardanza']
                    balance_horas_extra[nombre_trabajador] += resumen_diario['hrs_extra']
                    balance_horas_faltantes[nombre_trabajador] += resumen_diario['hrs_faltantes']

                    if resumen_diario['estado_celda'] != 'F' and resumen_diario['hrs_tardanza'].total_seconds() == 0:
                        asistencias_puntuales[nombre_trabajador] += 1
                        total_dias_puntuales += 1

        puntualidad_mes = (total_dias_puntuales / total_dias_laborables * 100) if total_dias_laborables > 0 else 0

        nombres_balance = set(balance_horas_extra.keys()) | set(balance_horas_faltantes.keys())
        balance_rows = []

        for nombre_trabajador in nombres_balance:
            horas_extra = round(balance_horas_extra[nombre_trabajador].total_seconds() / 3600, 2)
            horas_faltantes = round(balance_horas_faltantes[nombre_trabajador].total_seconds() / 3600, 2)

            # Excluimos valores nulos para evitar ruido visual.
            if horas_extra == 0 and horas_faltantes == 0:
                continue

            balance_rows.append({
                'nombre': nombre_trabajador,
                'extra': horas_extra,
                'faltantes': horas_faltantes,
                'impacto': horas_extra + horas_faltantes,
            })

        balance_rows.sort(key=lambda item: item['impacto'], reverse=True)
        balance_rows = balance_rows[:15]

        trabajadores_balance = [item['nombre'] for item in balance_rows]
        balance_extra_data = [item['extra'] for item in balance_rows]
        balance_faltantes_data = [item['faltantes'] for item in balance_rows]

        fuentes = ['App movil', 'Biometrico/Huellero', 'Excel importado']
        fuente_counter = Counter(ev['origen'] for ev in marcas_unificadas.values())
        fuente_data = [fuente_counter.get(f, 0) for f in fuentes]

        dias_labels = [str(d) for d in range(1, num_dias_mes + 1)]
        flujo_por_dia = {f: [0] * num_dias_mes for f in fuentes}
        reloj_por_hora = {f: [0] * 24 for f in fuentes}

        for evento in marcas_unificadas.values():
            indice_dia = evento['timestamp'].day - 1
            hora = evento['timestamp'].hour
            origen = evento['origen'] if evento['origen'] in flujo_por_dia else 'Biometrico/Huellero'
            flujo_por_dia[origen][indice_dia] += 1
            reloj_por_hora[origen][hora] += 1

        top_10_puntuales = sorted(asistencias_puntuales.items(), key=lambda x: x[1], reverse=True)[:10]
        top_10_tardanzas = sorted(tardanzas_count_por_trabajador.items(), key=lambda x: x[1], reverse=True)[:10]
        top_10_faltas = sorted(faltas_por_trabajador.items(), key=lambda x: x[1], reverse=True)[:10]

        fraudes_por_ubicacion = defaultdict(int)
        for doc in fraudes_mes:
            fraudes_por_ubicacion[doc.to_dict().get('locationName', 'Intento Remoto')] += 1

        fraudes_ubicacion_ordenado = sorted(fraudes_por_ubicacion.items(), key=lambda x: x[1], reverse=True)
        fraude_top_ubicacion = fraudes_ubicacion_ordenado[0][0] if fraudes_ubicacion_ordenado else 'Sin registros'
        fraude_top_cantidad = fraudes_ubicacion_ordenado[0][1] if fraudes_ubicacion_ordenado else 0
        fraude_total_eventos = len(fraudes_mes)
        fraude_top_participacion = round((fraude_top_cantidad / fraude_total_eventos) * 100, 1) if fraude_total_eventos else 0
        fraude_ranking = [
            {'ubicacion': ubicacion, 'cantidad': cantidad}
            for ubicacion, cantidad in fraudes_ubicacion_ordenado[:3]
        ]

        context.update({
            'mes_seleccionado': mes_dt.strftime('%Y-%m'),
            'total_marcas_kpi': len(marcas_unificadas),
            'fuentes_activas_kpi': sum(1 for v in fuente_data if v > 0),
            'total_faltas_kpi': sum(faltas_por_trabajador.values()),
            'total_fraudes_kpi': len(fraudes_mes),
            'puntualidad_mes_kpi': f"{puntualidad_mes:.2f}%",

            'cumplimiento_labels': json.dumps(list(cumplimiento_estado.keys())),
            'cumplimiento_data': json.dumps(list(cumplimiento_estado.values())),

            'fuente_labels': json.dumps(fuentes),
            'fuente_data': json.dumps(fuente_data),

            'flujo_dias_labels': json.dumps(dias_labels),
            'flujo_app_data': json.dumps(flujo_por_dia['App movil']),
            'flujo_bio_data': json.dumps(flujo_por_dia['Biometrico/Huellero']),
            'flujo_excel_data': json.dumps(flujo_por_dia['Excel importado']),

            'reloj_horas_labels': json.dumps([f"{h:02d}:00" for h in range(24)]),
            'reloj_app_data': json.dumps(reloj_por_hora['App movil']),
            'reloj_bio_data': json.dumps(reloj_por_hora['Biometrico/Huellero']),
            'reloj_excel_data': json.dumps(reloj_por_hora['Excel importado']),

            'top_puntuales_labels': json.dumps([item[0] for item in top_10_puntuales]),
            'top_puntuales_data': json.dumps([item[1] for item in top_10_puntuales]),
            'top_tardanzas_labels': json.dumps([item[0] for item in top_10_tardanzas]),
            'top_tardanzas_data': json.dumps([item[1] for item in top_10_tardanzas]),
            'top_faltas_labels': json.dumps([item[0] for item in top_10_faltas]),
            'top_faltas_data': json.dumps([item[1] for item in top_10_faltas]),

            'fraude_top_ubicacion': fraude_top_ubicacion,
            'fraude_top_cantidad': fraude_top_cantidad,
            'fraude_total_eventos': fraude_total_eventos,
            'fraude_top_participacion': fraude_top_participacion,
            'fraude_ranking': fraude_ranking,

            'fraudes_ubicacion_labels': json.dumps(list(fraudes_por_ubicacion.keys())),
            'fraudes_ubicacion_data': json.dumps(list(fraudes_por_ubicacion.values())),
            'balance_trabajadores_labels': json.dumps(trabajadores_balance),
            'balance_extra_data': json.dumps(balance_extra_data),
            'balance_faltantes_data': json.dumps(balance_faltantes_data),
        })

    return render(request, 'recursoshumanos/estadisticas_general.html', context)


@login_required
def exportar_reporte_general_excel(request):
    """ Exporta el reporte general con formato profesional a un archivo .xlsx """
    filtro_fecha_inicio_str = request.GET.get('fecha_inicio')
    filtro_fecha_fin_str = request.GET.get('fecha_fin')
    filtro_ubicacion = request.GET.get('ubicacion')

    try:
        reporte_data = _get_report_data(filtro_fecha_inicio_str, filtro_fecha_fin_str, filtro_ubicacion)
    except Exception as e:
        return HttpResponse(f"Error al generar los datos para exportar: {e}", status=500)

    # CREACIÓN Y CONFIGURACIÓN DEL LIBRO DE EXCEL
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte General Asistencias"

    # DEFINICIÓN DE ESTILOS
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='D45B04', end_color='D45B04', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # TÍTULO Y SUBTÍTULO
    sheet.merge_cells('A1:F1')
    titulo_cell = sheet['A1']
    titulo_cell.value = 'Reporte General de Personal'
    titulo_cell.font = Font(name='Calibri', size=16, bold=True)
    titulo_cell.alignment = Alignment(horizontal='center')
    
    subtitulo = f"Período: {filtro_fecha_inicio_str} al {filtro_fecha_fin_str}"
    if filtro_ubicacion:
        subtitulo += f" | Ubicación: {filtro_ubicacion}"
    
    sheet.merge_cells('A2:F2')
    subtitulo_cell = sheet['A2']
    subtitulo_cell.value = subtitulo
    subtitulo_cell.font = Font(name='Calibri', size=12, italic=True)
    subtitulo_cell.alignment = Alignment(horizontal='center')
    
    # CREACIÓN DE ENCABEZADOS EN LA FILA 4
    headers = [
        'DNI', 'Nombre del Trabajador', 'Asistencias a Tiempo', 
        'Tardanzas', 'Faltas', 'Días Justificados'
    ]
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # POBLAR EL EXCEL CON LOS DATOS (empezando en la fila 5)
    for row_num, data in enumerate(reporte_data, 5):
        stats = data['stats']
        row_data = [
            data['dni'],
            data['nombre'],
            stats['asistencias_a_tiempo'],
            stats['tardanzas'],
            stats['faltas'],
            stats['justificaciones']
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = thin_border
            # Centramos los datos numéricos (columnas 3 a 6)
            if cell.column >= 3:
                cell.alignment = center_alignment

    # AJUSTE FINAL DE ANCHO DE COLUMNAS
    column_widths = {'A': 15, 'B': 40, 'C': 20, 'D': 15, 'E': 15, 'F': 20}
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # GENERAR LA RESPUESTA HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f'Reporte_General_{date.today().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    
    return response

@login_required
def gestion_solicitudes(request):
    """ Muestra y gestiona las solicitudes pendientes de los trabajadores. """

    # --- Lógica para APROBAR o RECHAZAR una solicitud ---
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        solicitud_ref = db.collection('solicitudes').document(solicitud_id)
        
        try:
            if 'aprobar' in request.POST:
                solicitud_data = solicitud_ref.get().to_dict()
                
                # Si es una justificación, la creamos en la colección de justificaciones
                if solicitud_data.get('tipoSolicitud') == 'Justificación':
                    db.collection('justificaciones').add({
                        'dniTrabajador': solicitud_data.get('dniTrabajador'),
                        'nombreTrabajador': solicitud_data.get('nombreTrabajador'),
                        'fechaInicio': solicitud_data.get('fechaInicio'),
                        'fechaFin': solicitud_data.get('fechaFin'),
                        'tipo': solicitud_data.get('tipoSolicitud'),
                        'descripcion': solicitud_data.get('motivo'),
                        'creadoEn': firestore.SERVER_TIMESTAMP
                    })
                
                solicitud_ref.update({'estado': 'Aprobado'})
                messages.success(request, "Solicitud aprobada con éxito.")

            elif 'rechazar' in request.POST:
                solicitud_ref.update({'estado': 'Rechazado'})
                messages.warning(request, "Solicitud rechazada.")
        
        except Exception as e:
            messages.error(request, f"Error al procesar la solicitud: {e}")
            
        return redirect('gestion_solicitudes')

    # --- Lógica para mostrar las solicitudes pendientes ---
    solicitudes_pendientes = []
    try:
        docs = db.collection('solicitudes').where(
            filter=FieldFilter('estado', '==', 'Pendiente')
        ).order_by('fechaSolicitud').stream()

        for doc in docs:
            solicitud = doc.to_dict()
            solicitud['id'] = doc.id
            if 'fechaSolicitud' in solicitud and solicitud['fechaSolicitud']:
                solicitud['fechaSolicitud'] = solicitud['fechaSolicitud'].astimezone(LOCAL_TIMEZONE).strftime('%d/%m/%Y %H:%M')
            if 'fechaInicio' in solicitud and solicitud['fechaInicio']:
                solicitud['fechaInicio'] = solicitud['fechaInicio'].astimezone(LOCAL_TIMEZONE).strftime('%d/%m/%Y')
            if 'fechaFin' in solicitud and solicitud['fechaFin']:
                solicitud['fechaFin'] = solicitud['fechaFin'].astimezone(LOCAL_TIMEZONE).strftime('%d/%m/%Y')
            solicitudes_pendientes.append(solicitud)
    except Exception as e:
        messages.error(request, f"Error al cargar las solicitudes pendientes: {e}")

    context = {'solicitudes': solicitudes_pendientes}
    return render(request, 'recursoshumanos/gestion_solicitudes.html', context)


@login_required
def gestion_tareo(request):
    """
    Muestra y gestiona el tareo mensual (Planificación de Horarios).
    Aplica lógica automática de horarios para Campo (C) y Oficina (O).
    """
    # --- 1. PROCESAR FECHAS Y PARÁMETROS GET ---
    mes_seleccionado_str = request.GET.get('mes', datetime.now().strftime('%Y-%m'))
    try:
        anio, mes = map(int, mes_seleccionado_str.split('-'))
        mes_dt = date(anio, mes, 1)
    except ValueError:
        mes_dt = datetime.now().date()
        anio, mes = mes_dt.year, mes_dt.month

    meses_es = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    mes_nombre = f"{meses_es[mes-1]} de {anio}"
    
    _, num_dias = calendar.monthrange(anio, mes)
    inicio_mes = date(anio, mes, 1)
    fin_mes = date(anio, mes, num_dias)

    # --- 2. OBTENER FILTROS Y TRABAJADORES ---
    busqueda = request.GET.get('q', '').strip()
    proyecto_seleccionado_id = request.GET.get('proyecto')
    subproyecto_seleccionado_id = request.GET.get('subproyecto') # NUEVO FILTRO
    area_seleccionada_id = request.GET.get('area')
    
    proyectos = Proyecto.objects.filter(activo=True).order_by('nombre')
    subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).order_by('nombre')
    areas = Area.objects.all().order_by('nombre') 
    
    trabajadores_qs = Trabajador.objects.filter(activo=True)

    if proyecto_seleccionado_id:
        trabajadores_qs = trabajadores_qs.filter(
            asignaciones__proyecto_id=proyecto_seleccionado_id,
            asignaciones__activo=True
        )
        
    if subproyecto_seleccionado_id:
        trabajadores_qs = trabajadores_qs.filter(
            asignaciones__proyecto_id=subproyecto_seleccionado_id,
            asignaciones__activo=True
        )

    if area_seleccionada_id:
        trabajadores_qs = trabajadores_qs.filter(area_id=area_seleccionada_id)

    if busqueda:
        trabajadores_qs = trabajadores_qs.filter(
            Q(dni__icontains=busqueda) |
            Q(nombres__icontains=busqueda) |
            Q(apellido_paterno__icontains=busqueda) |
            Q(apellido_materno__icontains=busqueda)
        )

    trabajadores = trabajadores_qs.distinct().order_by('apellido_paterno')

    # --- 3. LÓGICA DE GUARDADO (POST) ---
    # (Toda tu lógica POST se mantiene EXACTAMENTE igual, no le he tocado nada 
    # para que tu guardado siga funcionando perfecto)
    if request.method == 'POST':
        try:
            cambios_json = request.POST.get('cambios', '{}')
            cambios = json.loads(cambios_json)
            
            with transaction.atomic():
                for dni, dias_data in cambios.items():
                    try:
                        trabajador = Trabajador.objects.get(dni=dni)
                    except Trabajador.DoesNotExist:
                        continue

                    for dia_str, data in dias_data.items():
                        fecha_actual = date(anio, mes, int(dia_str))
                        nuevo_estado = data.get('estado')

                        if not nuevo_estado:
                            TareoDiario.objects.filter(trabajador=trabajador, fecha=fecha_actual).delete()
                            continue

                        entrada_prog, salida_prog, jornada_prog = None, None, None

                        if nuevo_estado == 'C':
                            entrada_prog, salida_prog = time(9, 0), time(17, 0)
                        elif nuevo_estado == 'O':
                            # weekday() devuelve 5 para el Sábado
                            if fecha_actual.weekday() == 5: 
                                entrada_prog = time(9, 0)
                                salida_prog = time(13, 0) # Asumo que salen a la 1pm, cámbialo si es otra hora
                            else:
                                entrada_prog = time(8, 30)
                                salida_prog = time(18, 0)
                        elif nuevo_estado == 'P':
                            horario_dict = data.get('horario') or {}
                            str_ent = horario_dict.get('entrada')
                            str_sal = horario_dict.get('salida')
                            if str_ent: entrada_prog = datetime.strptime(str_ent, '%H:%M').time()
                            if str_sal: salida_prog = datetime.strptime(str_sal, '%H:%M').time()
                        elif nuevo_estado == 'J':
                            jornada_str = data.get('jornada_horas')
                            if jornada_str: jornada_prog = float(jornada_str)

                        TareoDiario.objects.update_or_create(
                            trabajador=trabajador,
                            fecha=fecha_actual,
                            defaults={
                                'estado': nuevo_estado,
                                'hora_entrada': entrada_prog,
                                'hora_salida': salida_prog,
                                'jornada_horas': jornada_prog,
                            }
                        )
            messages.success(request, f"Tareo de {mes_nombre} actualizado correctamente.")
        except Exception as e:
            messages.error(request, f"Error al guardar: {str(e)}")
        
        query_params = f"?mes={mes_seleccionado_str}"
        if busqueda: query_params += f"&q={busqueda}"
        if proyecto_seleccionado_id: query_params += f"&proyecto={proyecto_seleccionado_id}"
        if subproyecto_seleccionado_id: query_params += f"&subproyecto={subproyecto_seleccionado_id}"
        if area_seleccionada_id: query_params += f"&area={area_seleccionada_id}"
        return redirect(f"{request.path}{query_params}")

    # --- 4. PREPARAR DATOS PARA LA MATRIZ (GET) ---
    paginator = Paginator(trabajadores, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    trabajadores_pagina = page_obj.object_list

    dias_del_mes_info = []
    nombres_dias = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do"]
    for dia in range(1, num_dias + 1):
        fecha = date(anio, mes, dia)
        dias_del_mes_info.append({'numero': dia, 'nombre': nombres_dias[fecha.weekday()]})

    registros_mes = TareoDiario.objects.filter(
        fecha__range=(inicio_mes, fin_mes),
        trabajador__in=trabajadores_pagina
    ).select_related('trabajador')

    tareo_del_mes = {}
    for registro in registros_mes:
        dni = registro.trabajador.dni
        dia_str = str(registro.fecha.day)
        if dni not in tareo_del_mes: tareo_del_mes[dni] = {}
        
        tareo_del_mes[dni][dia_str] = {
            'estado': registro.estado,
            'jornada_horas': registro.jornada_horas,
            'horario': {
                'entrada': registro.hora_entrada.strftime('%H:%M') if registro.hora_entrada else None,
                'salida': registro.hora_salida.strftime('%H:%M') if registro.hora_salida else None
            }
        }

    context = {
        'trabajadores': trabajadores_pagina,
        'page_obj': page_obj,
        'mes_seleccionado': mes_seleccionado_str,
        'mes_nombre': mes_nombre,
        'dias_del_mes_info': dias_del_mes_info,
        'tareo_del_mes': tareo_del_mes,
        'proyectos': proyectos,
        'busqueda': busqueda,
        'proyecto_seleccionado_id': proyecto_seleccionado_id,
        'subproyectos': subproyectos,
        'subproyecto_seleccionado_id': subproyecto_seleccionado_id,
        'areas': areas, # Pasamos las áreas a la plantilla
        'area_seleccionada_id': area_seleccionada_id,
        'current_view': 'gestion_tareo',
    }
    
    return render(request, 'recursoshumanos/horarios/gestion_tareo_matricial.html', context)

class CustomLoginView(LoginView):
    template_name = 'login/login.html'
    # 1. APAGAMOS ESTO para evitar que Django se salte nuestras validaciones
    redirect_authenticated_user = False 

    def tiene_permisos_necesarios(self, user):
        """
        Función centralizada. Devuelve True si puede entrar, False si no tiene nada.
        """
        if user.is_superuser:
            return True
        if not user.groups.exists():
            return False

        grupos_administrativos = [
            'Administrador', 'Metricas', 'Recursos Humanos', 'Calidad',
            'Administracion', 'Gases', 'Proyectos', 'Cotizaciones',
            'proyecto_monitoreo_smcv', 'Yeni_admin', 'Supervisores', 'Gerencia'
        ]
        es_admin_o_grupo = user.is_superuser or user.groups.filter(name__in=grupos_administrativos).exists()
        tiene_perfil = hasattr(user, 'trabajador')
        
        return es_admin_o_grupo or tiene_perfil

    def dispatch(self, request, *args, **kwargs):
        """
        2. EL BARRERA DE ENTRADA: Interceptamos antes de que cargue la página.
        """
        if request.user.is_authenticated:
            # Si ya tiene una sesión iniciada en el navegador, revisamos si AÚN tiene roles.
            if not self.tiene_permisos_necesarios(request.user):
                # Si le quitaron los roles, lo deslogueamos a la fuerza.
                logout(request)
                messages.error(request, 'Tu cuenta no tiene grupos asignados o no tiene perfil válido. Contacta al Administrador.')
                return redirect('login') # Asegúrate de que el nombre de tu url sea 'login'
            
            # Si sí tiene roles, lo mandamos directo a su panel
            return redirect(self.get_success_url())
            
        return super().dispatch(request, *args, **kwargs)

    def form_invalid(self, form):
        """Mensaje genérico si se equivocan en la contraseña."""
        if not form.non_field_errors() and not self.request.POST.get('username') == '':
            messages.error(self.request, 'Nombre de usuario o contraseña incorrectos.')
        return super().form_invalid(form)

    def form_valid(self, form):
        """
        3. LA VALIDACIÓN AL DAR CLIC EN 'INGRESAR'
        """
        user = form.get_user()

        if not self.tiene_permisos_necesarios(user):
            # Rechazamos el formulario ANTES de que Django loguee al usuario
            messages.error(self.request, 'Tu usuario no tiene grupos asignados o no tiene perfil válido. Contacta al Administrador.')
            return self.render_to_response(self.get_context_data(form=form))

        # Solo si pasa la prueba, Django hace el login real
        return super().form_valid(form)

    def get_success_url(self):
        """
        A este punto es IMPOSIBLE que llegue un usuario sin roles.
        """
        user = self.request.user
        grupos = list(user.groups.values_list('name', flat=True))

        es_admin = 'Administrador' in grupos
        tiene_metricas = 'Metricas' in grupos
        tiene_rrhh = 'Recursos Humanos' in grupos
        tiene_calidad = 'Calidad' in grupos

        if es_admin:
            if tiene_metricas or tiene_rrhh:
                return reverse_lazy('accesos:dashboard_seleccion')
            return reverse_lazy('accesos:dashboard_seleccion')

        if tiene_metricas:
            return reverse_lazy('metricas_ceneris:inicio_metricas') 

        if tiene_calidad:
            return reverse_lazy('calidad:estadisticas_calidad')

        if tiene_rrhh:
            return reverse_lazy('recursoshumanos:dashboard')

        elif 'Supervisores' in grupos or 'Gerencia' in grupos:
            return reverse_lazy('recursoshumanos:dashboard')

        elif 'Administracion' in grupos:
            return reverse_lazy('administracion:dashboard_estadistico')

        elif 'Gases' in grupos:
            return reverse_lazy('cenerisapp:inicio')

        elif 'Proyectos' in grupos:
            return reverse_lazy('proyectos:dashboard')
        
        elif 'proyecto_monitoreo_smcv' in grupos or 'Yeni_admin' in grupos:
            # El módulo de monitoreo vive en el sistema externo monitoreo-web
            # (Fase 4). La app local no está en INSTALLED_APPS ni en urls.py,
            # así que sin esta redirección externa el usuario caería en un 404.
            if settings.MONITOREO_WEB_URL:
                return settings.MONITOREO_WEB_URL
            return '/proyecto_monitoreo_smcv/'
            
        elif 'Cotizaciones' in grupos:
            return reverse_lazy('cotizaciones:inicio')

        # LÓGICA DE TRABAJADOR
        try:
            perfil = user.trabajador 
            if perfil.es_jefe:
                return reverse_lazy('metricas_ceneris:panel_jefe')
            else:
                return reverse_lazy('metricas_ceneris:mis_evaluaciones') 
                
        except ObjectDoesNotExist:
            # Seguro final por si algo muy raro pasa
            logout(self.request)
            return reverse_lazy('login')
        
@login_required
def proyectos_dashboard(request):
    # Asegúrate que solo usuarios del grupo "Proyectos" entren aquí
    if not has_group(request.user, 'Calidad'):
        return redirect('recursoshumanos:main_dashboard') # O a una página de "acceso denegado"
    
    # El contexto para tu nueva app
    context = {} 
    return render(request, 'calidad/index.html', context)


# --- ¡NUEVA VISTA PARA EL DASHBOARD DE HORARIOS! ---
@login_required
@group_required('Recursos Humanos')
def gestion_horarios(request):
    """
    Muestra el dashboard de tarjetas para la gestión de horarios.
    """
    context = {
        'current_view': 'gestion_horarios',
    }
    return render(request, 'recursoshumanos/horarios/gestion_horarios.html', context)

# ==============================================================================
# VISTAS CRUD PARA EMPRESAS
# ==============================================================================

@login_required
@group_required('Recursos Humanos', 'Calidad')
def gestion_empresas(request):
    """Muestra el dashboard de tarjetas para la gestión de empresas."""
    context = {
        'current_view': 'gestion_empresas',
    }
    return render(request, 'recursoshumanos/empresas/gestion_empresas.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def lista_empresas(request):
    """Muestra una lista de todas las empresas."""
    empresas = Empresa.objects.all().order_by('nombre')
    context = {
        'empresas': empresas,
        'current_view': 'gestion_empresas',
    }
    return render(request, 'recursoshumanos/empresas/lista_empresas.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def crear_empresa(request):
    """Formulario para crear una nueva empresa."""
    if request.method == 'POST':
        form = EmpresaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa creada con éxito.')
            return redirect('recursoshumanos:lista_empresas')
    else:
        form = EmpresaForm()
    context = {
        'form': form,
        'form_title': 'Añadir Nueva Empresa',
        'current_view': 'gestion_empresas',
    }
    return render(request, 'recursoshumanos/empresas/empresa_form_multistep.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def editar_empresa(request, pk):
    """Formulario para editar una empresa."""
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa actualizada con éxito.')
            return redirect('recursoshumanos:lista_empresas')
    else:
        form = EmpresaForm(instance=empresa)
    context = {
        'form': form,
        'form_title': 'Editar Empresa',
        'current_view': 'gestion_empresas',
    }
    return render(request, 'recursoshumanos/empresas/empresa_form_multistep.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def eliminar_empresa(request, pk):
    """Elimina una empresa."""
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        # Comprobamos si tiene trabajadores de CUALQUIER tipo asociados
        # El acceso inverso por defecto para Trabajador.empresa es 'trabajador_set'
        if empresa.proyectos.exists() or empresa.trabajador_set.exists():
            messages.error(request, f'No se puede eliminar "{empresa.nombre}" porque tiene proyectos o trabajadores asociados.')
        else:
            empresa.delete()
            messages.success(request, f'Empresa "{empresa.nombre}" eliminada con éxito.')
        return redirect('recursoshumanos:lista_empresas')
    return redirect('recursoshumanos:lista_empresas')

@login_required
@group_required("Recursos Humanos", "Calidad")
def asignacion_masiva_empresas(request):
    todos_los_empresas = Empresa.objects.all().order_by('nombre')
    selected_empresa_id = request.GET.get('empresa_id')
    
    trabajadores_asignados = []
    selected_empresa = None

    if selected_empresa_id:
        try:
            selected_empresa = Empresa.objects.get(pk=selected_empresa_id)
            trabajadores_asignados = selected_empresa.trabajador_set.all().prefetch_related('asignaciones').order_by('apellido_paterno')
        except Empresa.DoesNotExist:
            messages.error(request, "La empresa seleccionada no existe.")
            selected_empresa_id = None

    # Filtramos trabajadores sin empresa que tengan asignaciones de proyecto activas
    trabajadores_disponibles = Trabajador.objects.filter(
        empresa__isnull=True,
        asignaciones__activo=True
    ).prefetch_related('asignaciones__proyecto', 'asignaciones__cargo').distinct().order_by('apellido_paterno')

    context = {
        'todos_los_empresas': todos_los_empresas,
        'selected_empresa': selected_empresa,
        'trabajadores_asignados': trabajadores_asignados,
        'trabajadores_disponibles': trabajadores_disponibles,
        'current_view': 'gestion_empresas'
    }
    return render(request, 'recursoshumanos/empresas/asignacion_masiva_empresas.html', context)

@login_required
def actualizar_asignacion_empresa(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            worker_dni = data.get('worker_dni')
            new_empresa_id = data.get('new_empresa_id')

            trabajador = Trabajador.objects.get(dni=worker_dni)
            if new_empresa_id:
                empresa = Empresa.objects.get(pk=new_empresa_id)
                trabajador.empresa = empresa
            else:
                trabajador.empresa = None
            
            trabajador.save(update_fields=['empresa'])
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error'}, status=400)

# ==============================================================================
# VISTAS CRUD PARA PROYECTOS
# ==============================================================================

@login_required
@group_required('Recursos Humanos', 'Calidad')
def gestion_proyectos(request):
    """Muestra el dashboard de tarjetas para la gestión de proyectos."""
    context = {
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'recursoshumanos/proyectos/gestion_proyectos.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def lista_proyectos(request):
    """
    Muestra una lista jerárquica de proyectos y sus subproyectos.
    """
    # 1. Obtenemos solo los proyectos de nivel superior (los que no son subproyectos).
    # 2. Usamos prefetch_related para cargar todos los subproyectos asociados en una 
    #    sola consulta adicional, lo que es muy eficiente.
    proyectos_padre = Proyecto.objects.filter(parent__isnull=True).prefetch_related(
        'subproyectos'
    ).order_by('nombre')
    
    context = {
        'proyectos_padre': proyectos_padre,
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'recursoshumanos/proyectos/lista_proyectos.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def crear_proyecto(request):
    """
    Formulario para crear un nuevo Proyecto Principal o un Subproyecto.
    """
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            proyecto = form.save(commit=False)
            tipo_proyecto = form.cleaned_data.get('tipo_proyecto')

            if tipo_proyecto == 'subproyecto':
                # BORRAMOS LA LÍNEA QUE SOBRESCRIBÍA EL NOMBRE CON EL CÓDIGO
                
                # Mantenemos solo la lógica de heredar la empresa
                if proyecto.parent:
                    proyecto.empresa = proyecto.parent.empresa
            
            proyecto.save()
            messages.success(request, '¡Proyecto guardado exitosamente!')
            return redirect('recursoshumanos:lista_proyectos')
        else:
            messages.error(request, "Por favor, corrige los errores mostrados en el formulario.")
            
    else: # Método GET
        form = ProyectoForm()

    context = {
        'form': form,
        'form_title': 'Crear Nuevo Proyecto',
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'recursoshumanos/proyectos/proyecto_form_tarjetas.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def editar_proyecto(request, pk):
    """
    Formulario para editar un proyecto o subproyecto existente.
    """
    # 1. RECUPERAR EL PROYECTO (Faltaba esta línea crítica)
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    padre_original = proyecto.parent 
    is_subproyecto = padre_original is not None
    
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        
        if form.is_valid():
            proyecto_editado = form.save(commit=False)

            if is_subproyecto:
                proyecto_editado.parent = padre_original
                
                if padre_original.empresa:
                    proyecto_editado.empresa = padre_original.empresa
            try:
                proyecto_editado.save()
                messages.success(request, 'Proyecto actualizado con éxito.')
                return redirect('recursoshumanos:lista_proyectos')
            
            except IntegrityError:
                form.add_error('nombre', 'Ya existe un proyecto con este nombre (o código). Por favor elige otro.')
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
            
    else: # GET
        form = ProyectoForm(instance=proyecto, is_editing_subproject=is_subproyecto)

    context = {
        'form': form,
        'proyecto': proyecto,
        'is_subproyecto': is_subproyecto,
        'form_title': 'Editar Proyecto',
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'recursoshumanos/proyectos/proyecto_form_editar.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def eliminar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    
    if proyecto.asignaciones.exists():
        messages.error(request, f"No se puede eliminar '{proyecto.nombre}' porque tiene trabajadores asignados. Primero desasigna al personal.")
        return redirect('recursoshumanos:lista_proyectos')

    if proyecto.subproyectos.exists():
        messages.error(request, f"No se puede eliminar '{proyecto.nombre}' porque tiene subproyectos dependientes.")
        return redirect('recursoshumanos:lista_proyectos')

    try:
        proyecto.delete()
        messages.success(request, 'Proyecto eliminado correctamente.')
    except Exception as e:
        messages.error(request, f"Error al eliminar el proyecto: {e}")
        
    return redirect('recursoshumanos:lista_proyectos')

@login_required
@group_required("Recursos Humanos", "Calidad")
def asignacion_masiva_proyectos(request):
    proyectos_padre = Proyecto.objects.filter(parent__isnull=True, activo=True)
    
    selected_proyecto_padre_id = request.GET.get('proyecto_padre_id')
    selected_subproyecto_id = request.GET.get('subproyecto_id')
    
    selected_subproyecto = None
    trabajadores_asignados = []
    
    # --- LÓGICA IMPORTANTE ---
    # Definimos primero la consulta base optimizada que trae todas las asignaciones
    base_queryset = Trabajador.objects.filter(activo=True).prefetch_related(
        'asignaciones', 
        'asignaciones__proyecto', 
        'asignaciones__cargo'
    ).order_by('apellido_paterno')

    if selected_subproyecto_id:
        try:
            selected_subproyecto = Proyecto.objects.get(pk=selected_subproyecto_id)
            
            # 1. Asignados: Los que SÍ tienen una asignación a este proyecto específico
            trabajadores_asignados = base_queryset.filter(
                asignaciones__proyecto=selected_subproyecto
            ).distinct()
            
            # 2. Disponibles: Todos los trabajadores activos MENOS los que ya están en la columna de asignados
            trabajadores_disponibles = base_queryset.exclude(
                pk__in=trabajadores_asignados.values_list('pk', flat=True)
            )

        except Proyecto.DoesNotExist:
            trabajadores_disponibles = base_queryset.all() # Si falla, mostramos todos
    else:
        # Si no hay subproyecto seleccionado, mostramos todos como disponibles
        trabajadores_disponibles = base_queryset.all()

    # ... (El resto del código del mapa JSON se queda igual) ...
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True)
    for sub in all_subproyectos:
        if sub.parent_id not in subproyectos_map:
            subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({'id': sub.id, 'nombre': sub.nombre, 'codigo': sub.codigo or ''})
    
    import json
    subproyectos_map_json = json.dumps(subproyectos_map)

    return render(request, 'recursoshumanos/proyectos/asignacion_masiva_proyectos.html', {
        'proyectos_padre': proyectos_padre,
        'trabajadores_disponibles': trabajadores_disponibles,
        'trabajadores_asignados': trabajadores_asignados,
        'selected_proyecto_padre_id': int(selected_proyecto_padre_id) if selected_proyecto_padre_id else None,
        'selected_subproyecto_id': int(selected_subproyecto_id) if selected_subproyecto_id else None,
        'selected_subproyecto': selected_subproyecto,
        'subproyectos_map_json': subproyectos_map_json,
    })

@login_required
@require_POST
def actualizar_asignacion_proyecto(request):
    try:
        data = json.loads(request.body)
        dni = data.get('worker_dni')
        new_proyecto_id = data.get('new_proyecto_id') # Si es NULL, significa desasignar
        
        trabajador = Trabajador.objects.get(dni=dni)

        if new_proyecto_id:
            # --- CASO ASIGNAR ---
            proyecto = Proyecto.objects.get(pk=new_proyecto_id)
            
            # Creamos la nueva asignación (o la obtenemos si ya existe)
            AsignacionProyecto.objects.get_or_create(
                trabajador=trabajador,
                proyecto=proyecto,
                defaults={'activo': True}
            )
            # Opcional: Si quieres que el trabajador SOLO pueda estar en un subproyecto a la vez dentro de este módulo,
            # podrías borrar sus otras asignaciones aquí antes de crear la nueva.
            
        else:
            # --- CASO DESASIGNAR (Mover a disponibles) ---
            # Aquí hay un dilema: "Disponible" significa sin asignaciones?
            # O significa "sacarlo de ESTE proyecto que estoy viendo"?
            # Asumiremos que queremos borrar TODAS sus asignaciones para dejarlo libre.
            
            # Borramos todas las asignaciones de este trabajador
            AsignacionProyecto.objects.filter(trabajador=trabajador).delete()

        return JsonResponse({'status': 'ok'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
# ==============================================================================
# VISTAS CRUD PARA CARGOS
# ==============================================================================

@login_required
@group_required('Calidad', 'Recursos Humanos')
def gestion_cargos(request):
    """Muestra el dashboard de tarjetas para la gestión de cargos."""
    context = {'current_view': 'gestion_cargos'}
    return render(request, 'recursoshumanos/cargos/gestion_cargos.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def lista_cargos(request):
    """Muestra una lista de todos los cargos."""
    cargos = Cargo.objects.all().order_by('nombre')
    context = {'cargos': cargos, 'current_view': 'gestion_cargos'}
    return render(request, 'recursoshumanos/cargos/lista_cargos.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def crear_cargo(request):
    """Formulario para crear un nuevo cargo."""
    if request.method == 'POST':
        form = CargoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo creado con éxito.')
            return redirect('recursoshumanos:lista_cargos')
    else:
        form = CargoForm()
    context = {'form': form, 'form_title': 'Añadir Nuevo Cargo'}
    return render(request, 'recursoshumanos/cargos/cargo_form_multistep.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def editar_cargo(request, pk):
    """Formulario para editar un cargo."""
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo actualizado con éxito.')
            return redirect('recursoshumanos:lista_cargos')
    else:
        form = CargoForm(instance=cargo)
    context = {'form': form, 'form_title': 'Editar Cargo'}
    return render(request, 'recursoshumanos/cargos/cargo_form_multistep.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def eliminar_cargo(request, pk):
    """
    Elimina un cargo, pero solo si no está siendo utilizado por ningún trabajador.
    """
    cargo = get_object_or_404(Cargo, pk=pk)
    
    # Esta vista solo debería procesar solicitudes POST para seguridad
    if request.method == 'POST':
        
        # --- ¡LA CORRECCIÓN ESTÁ AQUÍ! ---
        # Comprobamos si algún trabajador está usando este cargo.
        # El acceso inverso correcto es 'trabajador_set'.
        if cargo.trabajador_set.exists():
            # Si el cargo está en uso, mostramos un error y no lo eliminamos.
            messages.error(request, f'No se puede eliminar el cargo "{cargo.nombre}" porque está asignado a uno o más trabajadores.')
        else:
            # Si no está en uso, lo eliminamos y mostramos un mensaje de éxito.
            nombre_cargo = cargo.nombre
            cargo.delete()
            messages.success(request, f'El cargo "{nombre_cargo}" ha sido eliminado exitosamente.')
            
        # Redirigimos siempre a la lista de cargos.
        return redirect('recursoshumanos:lista_cargos')
    
    # Si se accede por GET, simplemente redirigir a la lista (o mostrar una página de confirmación).
    # Como usas un confirm de JS, redirigir es suficiente.
    return redirect('recursoshumanos:lista_cargos')


@login_required
@group_required("Recursos Humanos", "Calidad")
def asignacion_masiva_cargos(request):
    cargos = Cargo.objects.all().order_by('nombre')
    selected_cargo_id = request.GET.get('cargo_id')
    
    selected_cargo = None
    # Cambiamos la lógica: Ahora trabajamos con 'Asignaciones', no con 'Trabajadores' puros
    asignaciones_asignadas = []
    asignaciones_disponibles = []
    
    # Traemos todas las asignaciones activas con sus datos relacionados
    base_queryset = AsignacionProyecto.objects.filter(activo=True).select_related(
        'trabajador', 'proyecto', 'cargo'
    ).order_by('trabajador__apellido_paterno')

    if selected_cargo_id:
        try:
            selected_cargo = Cargo.objects.get(pk=selected_cargo_id)
            
            # 1. Asignados: Asignaciones que YA tienen este cargo específico
            asignaciones_asignadas = base_queryset.filter(
                cargo=selected_cargo
            )
            
            # 2. Disponibles: Asignaciones que NO tienen este cargo
            # (Mostramos todas las demás para que puedas cambiar a alguien de cargo o asignar a alguien nuevo)
            asignaciones_disponibles = base_queryset.exclude(
                cargo=selected_cargo
            )
            
        except Cargo.DoesNotExist:
            asignaciones_disponibles = base_queryset
    else:
        # Si no hay cargo seleccionado, mostramos las asignaciones que están "HUÉRFANAS" de cargo
        asignaciones_disponibles = base_queryset.filter(cargo__isnull=True)

    return render(request, 'recursoshumanos/cargos/asignacion_masiva_cargos.html', {
        'cargos': cargos,
        'asignaciones_disponibles': asignaciones_disponibles, # OJO: Cambió el nombre de la variable
        'asignaciones_asignadas': asignaciones_asignadas,     # OJO: Cambió el nombre de la variable
        'selected_cargo_id': int(selected_cargo_id) if selected_cargo_id else None,
        'selected_cargo': selected_cargo,
    })

@login_required
# En recursoshumanos/views.py (Reemplaza la función existente)

@require_POST
def actualizar_asignacion_cargo(request):
    try:
        data = json.loads(request.body)
        
        # AHORA RECIBIMOS EL ID DE LA ASIGNACIÓN (La relación específica)
        asignacion_id = data.get('asignacion_id') 
        new_cargo_id = data.get('new_cargo_id')
        
        # Buscamos la asignación exacta (Ej: Chipana en Proyecto X)
        asignacion = AsignacionProyecto.objects.get(pk=asignacion_id)

        if new_cargo_id:
            # Asignar Cargo
            cargo = Cargo.objects.get(pk=new_cargo_id)
            asignacion.cargo = cargo
            asignacion.save()
            mensaje = f"Asignado: {asignacion.trabajador} ahora es {cargo.nombre} en {asignacion.proyecto.nombre}"
        else:
            # Quitar Cargo
            asignacion.cargo = None
            asignacion.save()
            mensaje = f"Cargo removido de {asignacion.proyecto.nombre}"

        return JsonResponse({'status': 'ok', 'message': mensaje})

    except AsignacionProyecto.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Asignación no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

#falta crear vista para permisos de asistencia

@login_required
@group_required("Recursos Humanos", "Supervisores")
def gestionar_permisos(request, dni):
    """
    Página dedicada para editar las ubicaciones permitidas y ver el estado
    del dispositivo de un trabajador en Firestore.
    """
    try:
        doc_ref = db.collection('trabajadores').document(dni)
        trabajador_fs = doc_ref.get().to_dict()
        if not trabajador_fs:
            raise FileNotFoundError
        trabajador_fs['dni'] = dni
    except Exception as e:
        messages.error(request, f"No se pudo encontrar al trabajador de asistencia con DNI {dni}.")
        return redirect('recursoshumanos:lista_trabajadores') # Vuelve a la lista principal

    location_choices = _get_all_locations_choices()

    if request.method == 'POST':
        form = PermisosAsistenciaForm(request.POST)
        form.fields['ubicacionesPermitidas'].choices = location_choices
        if form.is_valid():
            try:
                update_data = {
                    'ubicacionesPermitidas': form.cleaned_data.get('ubicacionesPermitidas', [])
                }
                doc_ref.update(update_data)
                messages.success(request, f"Permisos de marcación para {trabajador_fs['nombre']} actualizados con éxito.")
                return redirect('recursoshumanos:lista_trabajadores')
            except Exception as e:
                messages.error(request, f"Error al actualizar permisos: {e}")
    else:
        # Preparamos los datos iniciales para el formulario
        initial_data = {
            'ubicacionesPermitidas': trabajador_fs.get('ubicacionesPermitidas', [])
        }
        form = PermisosAsistenciaForm(initial=initial_data)
        form.fields['ubicacionesPermitidas'].choices = location_choices

    context = {
        'form': form,
        'trabajador': trabajador_fs,
        'current_view': 'gestion_empleados'
    }
    return render(request, 'recursoshumanos/empleados/gestionar_permisos.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def gestion_centro_costo(request):
    """
    Muestra el dashboard de tarjetas para la gestión de Centros de Costo.
    """
    context = {
        'current_view': 'gestion_centro_costo', # Para resaltar la opción en el sidebar
    }
    return render(request, 'recursoshumanos/centro_costo/gestion_centro_costo.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def lista_centro_costo(request):
    """
    Muestra una tabla con todos los Centros de Costo registrados.
    """
    # Usamos .prefetch_related para optimizar el conteo de trabajadores en la plantilla
    centros_de_costo = CentroCosto.objects.prefetch_related('trabajadores').all().order_by('codigo')
    
    context = {
        'centros_de_costo': centros_de_costo,
        'current_view': 'gestion_centro_costo', # Para resaltar la opción en el sidebar
    }
    return render(request, 'recursoshumanos/centro_costo/lista_centro_costo.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def crear_centro_costo(request):
    """Muestra y procesa el formulario para crear un nuevo Centro de Costo."""
    if request.method == 'POST':
        form = CentroCostoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Centro de Costo creado con éxito.')
            return redirect('recursoshumanos:lista_centro_costo')
    else:
        form = CentroCostoForm()
    context = {
        'form': form,
        'form_title': 'Añadir Nuevo Centro de Costo',
        'current_view': 'gestion_centro_costo',
    }
    return render(request, 'recursoshumanos/centro_costo/centro_costo_form_multistep.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def editar_centro_costo(request, pk):
    """Muestra y procesa el formulario para editar un Centro de Costo existente."""
    centro_costo = get_object_or_404(CentroCosto, pk=pk)
    if request.method == 'POST':
        form = CentroCostoForm(request.POST, instance=centro_costo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Centro de Costo actualizado con éxito.')
            return redirect('recursoshumanos:lista_centro_costo')
    else:
        form = CentroCostoForm(instance=centro_costo)
    context = {
        'form': form,
        'form_title': 'Editar Centro de Costo',
        'current_view': 'gestion_centro_costo',
    }
    return render(request, 'recursoshumanos/centro_costo/centro_costo_form_multistep.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def eliminar_centro_costo(request, pk):
    """Elimina un Centro de Costo, con protección si está en uso."""
    centro_costo = get_object_or_404(CentroCosto, pk=pk)
    if request.method == 'POST':
        # Comprobamos si está siendo usado por algún trabajador
        if centro_costo.trabajadores.exists():
            messages.error(request, f'No se puede eliminar "{centro_costo.nombre}" porque está asignado a uno o más trabajadores.')
        else:
            centro_costo.delete()
            messages.success(request, f'Centro de Costo "{centro_costo.nombre}" eliminado con éxito.')
        return redirect('recursoshumanos:lista_centro_costo')
    return redirect('recursoshumanos:lista_centro_costo')

@login_required
@group_required('Calidad', 'Recursos Humanos')
def asignar_trabajadores_centro_costo(request):
    """
    Página interactiva para asignar trabajadores a centros de costo.
    """
    # Obtenemos todos los centros de costo para el selector
    todos_los_centros = CentroCosto.objects.all().order_by('nombre')
    
    # Obtenemos el centro de costo seleccionado desde la URL (ej. ?cc_id=1)
    selected_cc_id = request.GET.get('cc_id')
    
    trabajadores_asignados = []
    selected_cc = None

    if selected_cc_id:
        try:
            selected_cc = CentroCosto.objects.get(pk=selected_cc_id)
            trabajadores_asignados = selected_cc.trabajadores.all().order_by('apellido_paterno')
        except CentroCosto.DoesNotExist:
            messages.error(request, "El centro de costo seleccionado no existe.")
            selected_cc_id = None # Limpiamos para que no falle la plantilla

    # Obtenemos todos los trabajadores que NO tienen un centro de costo asignado
    trabajadores_disponibles = Trabajador.objects.filter(centro_costo__isnull=True).order_by('apellido_paterno')

    context = {
        'todos_los_centros': todos_los_centros,
        'selected_cc': selected_cc,
        'trabajadores_asignados': trabajadores_asignados,
        'trabajadores_disponibles': trabajadores_disponibles,
        'current_view': 'asignar_personal',
    }
    return render(request, 'recursoshumanos/centro_costo/asignacion_masiva.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def trabajadores_por_centro_costo(request, pk):
    centro_costo = get_object_or_404(CentroCosto, pk=pk)
    trabajadores = centro_costo.trabajadores.all().order_by('apellido_paterno')
    
    context = {
        'centro_costo': centro_costo,
        'trabajadores': trabajadores,
        'current_view': 'gestion_centro_costo' 
    }
    return render(request, 'recursoshumanos/centro_costo/trabajadores_por_centro_costo.html', context)

@login_required
def actualizar_asignacion_trabajador(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        worker_dni = data.get('worker_dni')
        new_cc_id = data.get('new_cc_id')

        try:
            trabajador = Trabajador.objects.get(dni=worker_dni)
            if new_cc_id:
                centro_costo = CentroCosto.objects.get(pk=new_cc_id)
                trabajador.centro_costo = centro_costo
            else:
                trabajador.centro_costo = None
            
            trabajador.save(update_fields=['centro_costo'])
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

@login_required
def gestion_dispositivos(request):
    """
    Página interactiva para asignar trabajadores a un dispositivo específico,
    usando modelos de Django (PostgreSQL). VERSIÓN CORREGIDA.
    """
    todos_los_dispositivos = Dispositivo.objects.all().order_by('nombre')
    selected_device_id = request.GET.get('device_id')
    
    trabajadores_asignados = User.objects.none() # QuerySet de User vacío
    trabajadores_disponibles = Trabajador.objects.filter(activo=True).order_by('apellido_paterno', 'nombres')
    selected_dispositivo = None

    if selected_device_id:
        try:
            selected_dispositivo = get_object_or_404(Dispositivo, id=selected_device_id)
            
            # Obtenemos los USUARIOS asignados
            usuarios_asignados = selected_dispositivo.trabajadores_permitidos.all()
            
            # A partir de los usuarios, obtenemos los perfiles de Trabajador
            trabajadores_asignados = Trabajador.objects.filter(user__in=usuarios_asignados, activo=True).order_by('apellido_paterno', 'nombres')
            
            # Obtenemos los trabajadores disponibles que NO están asignados
            trabajadores_disponibles = Trabajador.objects.filter(activo=True).exclude(
                user__in=usuarios_asignados
            ).order_by('apellido_paterno', 'nombres')

        except Dispositivo.DoesNotExist:
            messages.error(request, "El dispositivo seleccionado no existe.")
            selected_dispositivo = None
        except Exception as e:
            messages.error(request, f"Ocurrió un error: {e}")
            selected_dispositivo = None

    context = {
        'todos_los_dispositivos': todos_los_dispositivos,
        'selected_dispositivo': selected_dispositivo,
        'trabajadores_asignados': trabajadores_asignados,
        'trabajadores_disponibles': trabajadores_disponibles,
        'current_view': 'gestion_dispositivos',
    }
    return render(request, 'recursoshumanos/dispositivos/gestion_dispositivos.html', context)

@login_required
@require_POST
def actualizar_asignacion_dispositivo(request):
    try:
        data = json.loads(request.body)
        worker_pk = data.get('worker_pk')
        device_id = data.get('device_id')
        action = data.get('action')

        if not all([worker_pk, device_id, action]):
            return JsonResponse({'status': 'error', 'message': 'Faltan datos.'}, status=400)

        dispositivo = get_object_or_404(Dispositivo, id=device_id)
        trabajador = get_object_or_404(Trabajador, pk=worker_pk)
        
        usuario = trabajador.user

        if not usuario:
            return JsonResponse({'status': 'error', 'message': 'El trabajador no tiene un usuario vinculado.'}, status=400)

        if action == 'add':
            dispositivo.trabajadores_permitidos.add(usuario)
        elif action == 'remove':
            dispositivo.trabajadores_permitidos.remove(usuario)
        else:
            return JsonResponse({'status': 'error', 'message': 'Acción no válida.'}, status=400)
            
        return JsonResponse({'status': 'ok'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
@require_POST
def renombrar_dispositivo(request):
    """
    Endpoint para actualizar el nombre de un dispositivo.
    """
    try:
        data = json.loads(request.body)
        device_id = data.get('device_id')
        new_name = data.get('new_name')

        if not all([device_id, new_name]):
            return JsonResponse({'status': 'error', 'message': 'Faltan datos.'}, status=400)

        device_ref = db.collection('dispositivos').document(device_id)
        device_ref.update({'nombreDispositivo': new_name})
        
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# ======================= APARTADDO PARA REPORTES EXCEL ==========================

@login_required
def exportar_reporte_diario(request):
    """
    Exporta un reporte de todas las marcaciones (exclusivamente de PostgreSQL)
    de un día específico en un formato de tabla simple.
    """
    # 1. OBTENER FILTROS
    fecha_seleccionada_str = request.GET.get('fecha')
    filtro_ubicacion_id  = request.GET.get('ubicacion_id')
    filtro_proyecto_id   = request.GET.get('proyecto_id')
    filtro_trabajador_id = request.GET.get('trabajador_id')
    filtro_origen        = _normalizar_filtro_origen(request.GET.get('origen'))

    if not fecha_seleccionada_str:
        return render(request, 'recursoshumanos/reportes/pagina_exportar_diario.html', {
            'today': date.today().strftime('%Y-%m-%d'),
            'ubicaciones':       Ubicacion.objects.all().order_by('nombre'),
            'proyectos':         Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
            'trabajadores_lista': Trabajador.objects.filter(activo=True).order_by('apellido_paterno', 'nombres'),
            'origenes':          ORIGENES_MARCACION,
        })
    
    try:
        fecha_dt = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d')
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return redirect('exportar_reporte_diario')

    # 2. PREPARAR LIBRO Y ESTILOS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Marcas del {fecha_seleccionada_str}"
    
    bold_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')

    # 3. CONSTRUIR ESTRUCTURA Y ENCABEZADOS DEL EXCEL
    # Títulos principales
    sheet.merge_cells('B1:C1'); sheet['B1'].value = 'Ceneris E.I.R.L'
    sheet.merge_cells('B2:D2'); sheet['B2'].value = 'Registro de Marcas por día'
    
    # Fecha y hora de generación
    sheet['H2'] = fecha_dt.strftime('%d/%m/%Y')
    sheet['H3'] = f"Exportado: {datetime.now(LOCAL_TIMEZONE).strftime('%d/%m/%Y %H:%M:%S')}"

    # Aviso visible cuando el reporte esta filtrado por medio de marcacion.
    if filtro_origen:
        sheet.merge_cells('A4:E4')
        aviso = sheet['A4']
        aviso.value = f"FILTRADO POR MEDIO: {_etiqueta_origen(filtro_origen)} — solo se muestran las marcas de este medio"
        aviso.font = Font(bold=True, color='9C5700')
        aviso.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    # Encabezados de la tabla
    sheet['A5'] = 'Nombre'
    sheet['B5'] = 'Hora Entrada'
    sheet['C5'] = 'Hora Salida'
    sheet['D5'] = 'Punto'
    sheet['E5'] = 'Origen'
    for cell in sheet[5]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4. OBTENER DATOS SOLO DE POSTGRESQL
    inicio_dia = LOCAL_TIMEZONE.localize(fecha_dt.replace(hour=0, minute=0, second=0))
    fin_dia = LOCAL_TIMEZONE.localize(fecha_dt.replace(hour=23, minute=59, second=59))

    try:
        # Construir queryset de trabajadores con filtros opcionales
        trab_qs = Trabajador.objects.filter(user__isnull=False, activo=True)
        if filtro_ubicacion_id:
            trab_qs = trab_qs.filter(ubicaciones_permitidas__id=filtro_ubicacion_id)
        if filtro_proyecto_id:
            trab_qs = trab_qs.filter(proyectos__id=filtro_proyecto_id)
        if filtro_trabajador_id:
            trab_qs = trab_qs.filter(id=filtro_trabajador_id)

        dni_por_user_id = {t.user_id: t for t in trab_qs}

        # Filtrar marcaciones al rango del día
        asis_filter = dict(timestamp__gte=inicio_dia, timestamp__lte=fin_dia)
        if any([filtro_ubicacion_id, filtro_proyecto_id, filtro_trabajador_id]):
            asis_filter['usuario_id__in'] = list(dni_por_user_id.keys())
        # Filtro por medio de marcacion: al filtrar, el reporte solo muestra
        # las marcas hechas por ese medio (quien no marco por ahi no aparece).
        if filtro_origen:
            asis_filter['origen__iexact'] = filtro_origen
        asistencias_pg = Asistencia.objects.filter(**asis_filter).select_related('usuario').order_by('usuario_id', 'timestamp')

        # Agrupar por trabajador: guardar primera Entrada y última Salida del día
        from collections import defaultdict
        registros_por_usuario = defaultdict(lambda: {
            'nombre_completo': '', 'nombre_orden': '',
            'hora_entrada': None, 'hora_salida': None,
            'punto': '', 'origen': ''
        })

        for asis in asistencias_pg:
            trabajador = dni_por_user_id.get(asis.usuario_id)
            if trabajador:
                codigo = trabajador.dni[:5] if trabajador.dni else ''
                nombre = trabajador.nombre_completo
            else:
                codigo = ''
                nombre = asis.usuario.username

            timestamp_local = asis.timestamp.astimezone(LOCAL_TIMEZONE)
            origen_display = asis.get_origen_display() if hasattr(asis, 'get_origen_display') else asis.origen
            uid = asis.usuario_id
            reg = registros_por_usuario[uid]

            reg['nombre_completo'] = f"{codigo}: {nombre}" if codigo else nombre
            reg['nombre_orden'] = nombre.lower()
            reg['punto'] = asis.nombre_ubicacion or 'Sede Principal'
            # Origen de la PRIMERA marca del dia (igual que Consulta de
            # Asistencias, Semanal y Mensual individuales). Antes se
            # sobrescribia en cada vuelta sin condicion, asi que terminaba
            # mostrando el origen de la ULTIMA marca (normalmente la salida)
            # en vez de la primera (la entrada).
            if not reg['origen']:
                reg['origen'] = origen_display

            hora_str = timestamp_local.strftime('%H:%M')
            if asis.tipo_marcacion == 'Entrada':
                # Guardar la primera entrada del día
                if reg['hora_entrada'] is None:
                    reg['hora_entrada'] = hora_str
            elif asis.tipo_marcacion == 'Salida':
                # Guardar la última salida del día
                reg['hora_salida'] = hora_str

        # 5. ORDENAR Y LLENAR EXCEL
        filas = sorted(registros_por_usuario.values(), key=lambda x: x['nombre_orden'])

        current_row = 6
        for reg in filas:
            sheet.cell(row=current_row, column=1, value=reg['nombre_completo'])
            sheet.cell(row=current_row, column=2, value=reg['hora_entrada'] or '-')
            sheet.cell(row=current_row, column=3, value=reg['hora_salida'] or '-')
            sheet.cell(row=current_row, column=4, value=reg['punto'])
            sheet.cell(row=current_row, column=5, value=reg['origen'])
            current_row += 1

    except Exception as e:
        messages.error(request, f"Error al generar el reporte: {e}")
        return redirect('exportar_reporte_diario')

    # 6. AJUSTES FINALES DE ANCHO DE COLUMNAS
    sheet.column_dimensions['A'].width = 366 / 7
    sheet.column_dimensions['B'].width = 90 / 7
    sheet.column_dimensions['C'].width = 90 / 7
    sheet.column_dimensions['D'].width = 150 / 7
    sheet.column_dimensions['E'].width = 110 / 7

    # 7. GENERAR Y DEVOLVER EL ARCHIVO
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'Registro_Marcas_{fecha_seleccionada_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    
    return response

@login_required
def exportar_reporte_semanal(request):
    """
    Exporta el registro de marcaciones de la semana (lunes a domingo) que contiene
    la fecha seleccionada. Los días posteriores a la fecha seleccionada muestran 'X'.
    """
    fecha_seleccionada_str = request.GET.get('fecha')
    filtro_ubicacion_id  = request.GET.get('ubicacion_id')
    filtro_proyecto_id   = request.GET.get('proyecto_id')
    filtro_trabajador_id = request.GET.get('trabajador_id')
    filtro_origen        = _normalizar_filtro_origen(request.GET.get('origen'))

    if not fecha_seleccionada_str:
        # La plantilla 'pagina_exportar_semanal.html' nunca existio: entrar a
        # esta URL sin fecha reventaba con TemplateDoesNotExist (error 500).
        # Mandamos al panel de reportes, que ya tiene la seccion semanal con
        # todos sus filtros.
        return redirect('recursoshumanos:gestion_reportes')

    try:
        fecha_dt = datetime.strptime(fecha_seleccionada_str, '%Y-%m-%d')
        fecha_seleccionada = fecha_dt.date()
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return redirect('recursoshumanos:exportar_reporte_semanal')

    # Lunes y domingo de la semana seleccionada
    lunes   = fecha_seleccionada - timedelta(days=fecha_seleccionada.weekday())
    domingo = lunes + timedelta(days=6)
    dias_semana = [lunes + timedelta(days=i) for i in range(7)]

    # --- PREPARAR LIBRO Y ESTILOS ---
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Semana {lunes.strftime('%d-%m')}"

    bold_font    = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    x_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    x_font = Font(bold=True, color='808080')

    # --- OBTENER DATOS DE POSTGRESQL (comun a ambos formatos) ---
    inicio_periodo = LOCAL_TIMEZONE.localize(datetime.combine(lunes, time(0, 0, 0)))
    fin_periodo    = LOCAL_TIMEZONE.localize(datetime.combine(fecha_seleccionada, time(23, 59, 59)))

    trab_qs = Trabajador.objects.filter(user__isnull=False, activo=True)
    if filtro_ubicacion_id:
        trab_qs = trab_qs.filter(ubicaciones_permitidas__id=filtro_ubicacion_id)
    if filtro_proyecto_id:
        trab_qs = trab_qs.filter(proyectos__id=filtro_proyecto_id)
    if filtro_trabajador_id:
        trab_qs = trab_qs.filter(id=filtro_trabajador_id)

    trab_map = {t.user_id: t for t in trab_qs}

    asis_filter = dict(timestamp__gte=inicio_periodo, timestamp__lte=fin_periodo)
    if any([filtro_ubicacion_id, filtro_proyecto_id, filtro_trabajador_id]):
        asis_filter['usuario_id__in'] = list(trab_map.keys())
    # Filtro por medio de marcacion: solo las marcas de ese medio.
    if filtro_origen:
        asis_filter['origen__iexact'] = filtro_origen

    asistencias = (Asistencia.objects
                   .filter(**asis_filter)
                   .select_related('usuario')
                   .order_by('usuario_id', 'timestamp'))

    # Agrupar por trabajador y día
    datos_por_uid = defaultdict(lambda: {
        'nombre': '', 'nombre_orden': '',
        'dias': defaultdict(lambda: {'entrada': None, 'salida': None, 'origen': None}),
    })

    for asis in asistencias:
        trab = trab_map.get(asis.usuario_id)
        if not trab:
            continue
        uid = asis.usuario_id
        codigo = trab.dni[:5] if trab.dni else ''
        nombre = trab.nombre_completo
        datos_por_uid[uid]['nombre'] = f"{codigo}: {nombre}" if codigo else nombre
        datos_por_uid[uid]['nombre_orden'] = nombre.lower()

        ts_local  = asis.timestamp.astimezone(LOCAL_TIMEZONE)
        fecha_asis = ts_local.date()
        hora_str  = ts_local.strftime('%H:%M')

        if datos_por_uid[uid]['dias'][fecha_asis]['origen'] is None:
            datos_por_uid[uid]['dias'][fecha_asis]['origen'] = asis.origen

        if asis.tipo_marcacion == 'Entrada':
            if datos_por_uid[uid]['dias'][fecha_asis]['entrada'] is None:
                datos_por_uid[uid]['dias'][fecha_asis]['entrada'] = hora_str
        else:
            datos_por_uid[uid]['dias'][fecha_asis]['salida'] = hora_str

    # =========================================================================
    # FORMATO VERTICAL: se activa solo cuando se filtra a UN trabajador
    # especifico (una fila por dia, en vez de un trabajador por fila con los
    # 7 dias como columnas). El formato horizontal de abajo (para todos los
    # trabajadores) no se modifica.
    # =========================================================================
    if filtro_trabajador_id:
        trabajador_obj = trab_qs.first()
        if not trabajador_obj:
            messages.error(request, 'No se encontró el trabajador seleccionado.')
            return redirect('recursoshumanos:exportar_reporte_semanal')

        ESTADO_LABELS = {'A': 'ASISTIÓ', 'F': 'FALTA', 'J': 'JUSTIFICADO'}
        ESTADO_FILLS = {
            'ASISTIÓ':     PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
            'FALTA':       PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
            'JUSTIFICADO': PatternFill(start_color='E9D5FF', end_color='E9D5FF', fill_type='solid'),
            'PROGRAMADO':  x_fill,
            'SIN REGISTRO': x_fill,
        }

        # resultado_efectivo (no el campo crudo): si hay marcaciones reales el
        # dia cuenta como Asistio aunque el campo guardado siga en 'F'.
        tareo_map = {
            t.fecha: t.resultado_efectivo
            for t in TareoDiario.objects.filter(trabajador=trabajador_obj, fecha__range=(lunes, domingo))
        }
        reg = datos_por_uid.get(trabajador_obj.user_id, {'dias': {}})

        # --- TÍTULOS ---
        sheet.merge_cells('A1:E1')
        sheet['A1'] = 'Ceneris E.I.R.L'
        sheet['A1'].font = bold_font

        sheet.merge_cells('A2:E2')
        sheet['A2'] = 'Registro de Marcas Semanal — Vista Individual'
        sheet['A2'].font = bold_font

        codigo_trab = trabajador_obj.dni[:5] if trabajador_obj.dni else ''
        sheet.merge_cells('A3:E3')
        sheet['A3'] = f"{trabajador_obj.nombre_completo} — DNI {trabajador_obj.dni or '-'}"
        sheet['A3'].font = bold_font

        sheet.merge_cells('A4:E4')
        area_nombre = trabajador_obj.area.nombre if trabajador_obj.area_id else 'Sin área'
        sheet['A4'] = f"Área: {area_nombre}   |   Semana: {lunes.strftime('%d/%m/%Y')} - {domingo.strftime('%d/%m/%Y')}"
        sheet['G1'] = f"Exportado: {datetime.now(LOCAL_TIMEZONE).strftime('%d/%m/%Y %H:%M:%S')}"

        fila_encabezado = 6
        if filtro_origen:
            sheet.merge_cells('A5:E5')
            aviso = sheet['A5']
            aviso.value = f"FILTRADO POR MEDIO: {_etiqueta_origen(filtro_origen)} — solo se listan los días marcados por este medio"
            aviso.font = Font(bold=True, color='9C5700')
            aviso.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        # --- ENCABEZADOS ---
        headers = ['Fecha', 'Estado', 'Entrada', 'Salida', 'Medio']
        for col, header in enumerate(headers, 1):
            c = sheet.cell(row=fila_encabezado, column=col, value=header)
            c.font = bold_font
            c.alignment = header_align
        sheet.row_dimensions[fila_encabezado].height = 20

        # --- FILAS: una por día de la semana ---
        dias_abrev_es = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do']
        current_row = fila_encabezado + 1
        for dia in dias_semana:
            dia_data = reg['dias'].get(dia, {})

            # Con filtro por medio: solo se listan los dias que tienen marca de
            # ese medio. No se muestran faltas ni justificaciones.
            if filtro_origen and not dia_data.get('entrada') and not dia_data.get('salida'):
                continue

            fecha_label = f"{dia.strftime('%d/%m/%Y')} ({dias_abrev_es[dia.weekday()]})"
            sheet.cell(row=current_row, column=1, value=fecha_label).alignment = center_align

            if filtro_origen:
                estado_label = 'ASISTIÓ'
                entrada = dia_data.get('entrada') or '-'
                salida = dia_data.get('salida') or '-'
                medio = dia_data.get('origen') or '-'
            elif dia > fecha_seleccionada:
                estado_label = 'PROGRAMADO'
                entrada, salida, medio = '-', '-', '-'
            else:
                resultado = tareo_map.get(dia)
                estado_label = ESTADO_LABELS.get(resultado, 'SIN REGISTRO')
                entrada = dia_data.get('entrada') or '-'
                salida = dia_data.get('salida') or '-'
                medio = dia_data.get('origen') or '-'

            estado_cell = sheet.cell(row=current_row, column=2, value=estado_label)
            estado_cell.alignment = center_align
            estado_cell.font = bold_font
            fill = ESTADO_FILLS.get(estado_label)
            if fill:
                estado_cell.fill = fill

            sheet.cell(row=current_row, column=3, value=entrada).alignment = center_align
            sheet.cell(row=current_row, column=4, value=salida).alignment = center_align
            sheet.cell(row=current_row, column=5, value=medio).alignment = center_align

            sheet.row_dimensions[current_row].height = 18
            current_row += 1

        # --- ANCHOS DE COLUMNA ---
        sheet.column_dimensions['A'].width = 22
        sheet.column_dimensions['B'].width = 16
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 14

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'Registro_Semanal_{trabajador_obj.dni or trabajador_obj.id}_{lunes.strftime("%Y-%m-%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    # =========================================================================
    # FORMATO HORIZONTAL (sin cambios): todos los trabajadores, un trabajador
    # por fila y los 7 dias de la semana como columnas.
    # =========================================================================

    # --- TÍTULOS ---
    sheet.merge_cells('B1:C1')
    sheet['B1'] = 'Ceneris E.I.R.L'
    sheet['B1'].font = bold_font

    sheet.merge_cells('B2:E2')
    sheet['B2'] = 'Registro de Marcas Semanal'
    sheet['B2'].font = bold_font

    sheet['N2'] = f"{lunes.strftime('%d/%m/%Y')} - {domingo.strftime('%d/%m/%Y')}"
    sheet['N3'] = f"Exportado: {datetime.now(LOCAL_TIMEZONE).strftime('%d/%m/%Y %H:%M:%S')}"

    if filtro_origen:
        sheet.merge_cells('B3:I3')
        aviso = sheet['B3']
        aviso.value = f"FILTRADO POR MEDIO: {_etiqueta_origen(filtro_origen)} — solo se muestran las marcas de este medio"
        aviso.font = Font(bold=True, color='9C5700')
        aviso.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

    # --- ENCABEZADOS FILA 5: Nombre (merge con fila 6) + días (2 cols cada uno) ---
    nombres_dias = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do']
    sheet.merge_cells(start_row=5, end_row=6, start_column=1, end_column=1)
    cell_nombre = sheet.cell(row=5, column=1, value='Nombre')
    cell_nombre.font = bold_font
    cell_nombre.alignment = header_align

    COL_INICIO = 2  # columna B
    for i, dia in enumerate(dias_semana):
        col_e = COL_INICIO + i * 2
        col_s = COL_INICIO + i * 2 + 1
        dia_str = f"{nombres_dias[i]}\n{dia.strftime('%d/%m')}"

        sheet.merge_cells(start_row=5, end_row=5, start_column=col_e, end_column=col_s)
        cell_dia = sheet.cell(row=5, column=col_e, value=dia_str)
        cell_dia.font = bold_font
        cell_dia.alignment = header_align

        for col, label in [(col_e, 'Entrada'), (col_s, 'Salida')]:
            c = sheet.cell(row=6, column=col, value=label)
            c.font = bold_font
            c.alignment = header_align

    sheet.row_dimensions[5].height = 30
    sheet.row_dimensions[6].height = 18

    filas_ordenadas = sorted(datos_por_uid.values(), key=lambda r: r['nombre_orden'])

    # --- ESCRIBIR FILAS DE DATOS ---
    current_row = 7
    for reg in filas_ordenadas:
        sheet.cell(row=current_row, column=1, value=reg['nombre'])

        for i, dia in enumerate(dias_semana):
            col_e = COL_INICIO + i * 2
            col_s = COL_INICIO + i * 2 + 1

            if dia > fecha_seleccionada:
                sheet.merge_cells(start_row=current_row, end_row=current_row,
                                  start_column=col_e, end_column=col_s)
                cell_x = sheet.cell(row=current_row, column=col_e, value='X')
                cell_x.alignment = center_align
                cell_x.fill = x_fill
                cell_x.font = x_font
            else:
                dia_data = reg['dias'].get(dia, {})
                entrada = dia_data.get('entrada') or '-'
                salida  = dia_data.get('salida')  or '-'
                c_e = sheet.cell(row=current_row, column=col_e, value=entrada)
                c_s = sheet.cell(row=current_row, column=col_s, value=salida)
                c_e.alignment = center_align
                c_s.alignment = center_align

        sheet.row_dimensions[current_row].height = 18
        current_row += 1

    # --- ANCHOS DE COLUMNA ---
    sheet.column_dimensions['A'].width = 40
    for i in range(7):
        sheet.column_dimensions[get_column_letter(COL_INICIO + i * 2)].width = 10
        sheet.column_dimensions[get_column_letter(COL_INICIO + i * 2 + 1)].width = 10

    # --- GENERAR RESPUESTA ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'Registro_Semanal_{lunes.strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


#exportar formato para planilla
@login_required
def exportar_formato_planilla(request):
    """
    Genera un archivo Excel con el formato específico de carga de datos de planilla,
    resumiendo las tardanzas (TARD) y faltas (DIFA) del mes por trabajador
    usando datos de PostgreSQL.
    """
    # --- 1. OBTENER FILTROS ---
    mes_seleccionado_str = request.GET.get('mes')
    filtro_ubicacion_id  = request.GET.get('ubicacion_id')
    filtro_proyecto_id   = request.GET.get('proyecto_id')
    filtro_trabajador_id = request.GET.get('trabajador_id')

    if not mes_seleccionado_str:
        return render(request, 'recursoshumanos/reportes/pagina_exportar_planilla.html', {
            'ubicaciones':       Ubicacion.objects.all().order_by('nombre'),
            'proyectos':         Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
            'trabajadores_lista': Trabajador.objects.filter(activo=True).order_by('apellido_paterno', 'nombres'),
        })
    
    try:
        mes_dt = datetime.strptime(mes_seleccionado_str, '%Y-%m')
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return redirect('exportar_formato_planilla')

    # --- 2. PREPARAR LIBRO DE EXCEL ---
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Carga Planilla {mes_dt.strftime('%B %Y')}"

    center_align = Alignment(horizontal='center', vertical='center')
    
    # --- 3. DEFINIR RANGOS DE FECHA ---
    primer_dia = mes_dt.replace(day=1).date()
    num_dias = calendar.monthrange(mes_dt.year, mes_dt.month)[1]
    ultimo_dia = mes_dt.replace(day=num_dias).date()

    trabajadores = Trabajador.objects.filter(activo=True)
    if filtro_ubicacion_id:
        trabajadores = trabajadores.filter(ubicaciones_permitidas__id=filtro_ubicacion_id)
    if filtro_proyecto_id:
        trabajadores = trabajadores.filter(proyectos__id=filtro_proyecto_id)
    if filtro_trabajador_id:
        trabajadores = trabajadores.filter(id=filtro_trabajador_id)
    trabajadores = trabajadores.order_by('apellido_paterno', 'nombres').distinct()

    # --- 4. CONSTRUIR ESTRUCTURA Y ENCABEZADOS DEL EXCEL ---
    sheet.merge_cells('B1:C1'); sheet['B1'].value = 'METRALAB S.A.C.'
    sheet.merge_cells('B2:C2'); sheet['B2'].value = 'Formato de carga de datos de planilla'
    
    header_font = Font(bold=True)
    sheet['A4'].value = 'CODIGO'; sheet['A4'].font = header_font
    sheet['B4'].value = 'IDTRAB'; sheet['B4'].font = header_font
    sheet['C4'].value = 'NOMBRE'; sheet['C4'].font = header_font
    sheet['D4'].value = 'TARD';   sheet['D4'].font = header_font
    sheet['E4'].value = 'DIFA';   sheet['E4'].font = header_font

    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 38

    # --- 5. LLENAR DATOS DE TRABAJADORES ---
    current_row = 5 
    for trabajador in trabajadores:
        
        # A) CÁLCULO DE FALTAS (DIFA) - Usamos la función blindada
        asistencia_data = _calcular_asistencia_por_periodo(trabajador, primer_dia, ultimo_dia)
        total_faltas_mes = asistencia_data['faltas']

        # B) CÁLCULO DE TARDANZAS (TARD) en Horas:Minutos
        tareos_mes = TareoDiario.objects.filter(
            trabajador=trabajador,
            fecha__gte=primer_dia,
            fecha__lte=ultimo_dia
        ).exclude(hora_entrada_real__isnull=True)
        
        total_segundos_tardanza = 0
        
        for t in tareos_mes:
            hora_ref = _hora_referencia_entrada_por_fecha(t.fecha)
            
            segundos_ref = (hora_ref.hour * 3600) + (hora_ref.minute * 60) + hora_ref.second
            segundos_ent = (t.hora_entrada_real.hour * 3600) + (t.hora_entrada_real.minute * 60) + t.hora_entrada_real.second
            
            segundos_tarde = max(0, segundos_ent - segundos_ref)
            
            # Usamos TOLERANCIA_TARDANZA_MINUTOS que está definida al inicio de tu archivo
            if segundos_tarde >= (TOLERANCIA_TARDANZA_MINUTOS * 60):
                total_segundos_tardanza += segundos_tarde

        td_tardanza = timedelta(seconds=total_segundos_tardanza)
        str_tardanza = formatear_timedelta(td_tardanza) if total_segundos_tardanza > 0 else "0:00"

        # C) ESCRIBIR EN EXCEL
        codigo = trabajador.dni[:5] if trabajador.dni else ''
        
        sheet.cell(row=current_row, column=1, value=codigo) 
        sheet.cell(row=current_row, column=2, value=trabajador.dni) 
        sheet.cell(row=current_row, column=3, value=trabajador.nombre_completo) 
        
        tard_cell = sheet.cell(row=current_row, column=4, value=str_tardanza)
        tard_cell.alignment = center_align
        
        difa_cell = sheet.cell(row=current_row, column=5, value=total_faltas_mes)
        difa_cell.alignment = center_align
        
        current_row += 1

    # --- 6. GENERAR Y DEVOLVER EL ARCHIVO ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'Formato_carga_planilla_{mes_seleccionado_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    
    return response



@login_required
def exportar_reporte_maestro(request):

    # --- 1. OBTENER FILTROS ---
    mes_seleccionado_str = request.GET.get('mes')
    filtro_ubicacion_id  = request.GET.get('ubicacion_id')
    filtro_proyecto_id   = request.GET.get('proyecto_id')
    filtro_trabajador_id = request.GET.get('trabajador_id')
    filtro_origen        = _normalizar_filtro_origen(request.GET.get('origen'))

    if not mes_seleccionado_str:
        return render(request, 'recursoshumanos/reportes/pagina_exportar_maestro.html', {
            'ubicaciones':       Ubicacion.objects.all().order_by('nombre'),
            'proyectos':         Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
            'trabajadores_lista': Trabajador.objects.filter(activo=True).order_by('apellido_paterno', 'nombres'),
            'origenes':          ORIGENES_MARCACION,
        })
    try:
        mes_dt = datetime.strptime(mes_seleccionado_str, '%Y-%m')
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return redirect('exportar_reporte_maestro')

     # --- 2. PREPARAR LIBRO Y ESTILOS DE EXCEL ---
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Tareo {mes_dt.strftime('%Y-%m')}"
    
    # Fuentes y Alineaciones
    header_font = Font(name='Arial', size=9, bold=True)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True) # wrap_text=True es clave para los saltos de línea
    center_align = Alignment(horizontal='center', vertical='center') # centrar el contenido de las celdas

    # Colores de Relleno
    rojo_fill = PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid')
    naranja_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    morado_fill = PatternFill(start_color='E9D5FF', end_color='E9D5FF', fill_type='solid')
    rojo_font = Font(name='Arial', size=9, bold=True, color='9C0006')
    naranja_font = Font(name='Arial', size=9, bold=True, color='9C5700')
    morado_font = Font(name='Arial', size=9, bold=True, color='6D28D9')
    normal_font = Font(name='Arial', size=9)


    # --- 3. OBTENER TODOS LOS DATOS DE FIRESTORE ---
    primer_dia = LOCAL_TIMEZONE.localize(mes_dt.replace(day=1))
    num_dias = calendar.monthrange(mes_dt.year, mes_dt.month)[1]
    ultimo_dia = LOCAL_TIMEZONE.localize(mes_dt.replace(day=num_dias, hour=23, minute=59))

    trabajadores_docs = list(
        db.collection('trabajadores').where(filter=FieldFilter('activo', '==', True)).stream()
    )

    trabajadores = {}
    for doc in trabajadores_docs:
        data = doc.to_dict() or {}
        dni = str(
            data.get('dni')
            or data.get('userDni')
            or data.get('documento')
            or doc.id
        ).strip()
        if not dni:
            continue
        trabajadores[dni] = data

    # Enriquecemos con BD local (RRHH): nombres/cargos oficiales y cobertura completa.
    for trabajador_db in Trabajador.objects.filter(activo=True).order_by('apellido_paterno', 'apellido_materno', 'nombres'):
        dni = str(trabajador_db.dni or '').strip()
        if not dni:
            continue

        data_existente = trabajadores.get(dni, {})
        nombre_db = f"{trabajador_db.nombres or ''} {trabajador_db.apellido_paterno or ''} {trabajador_db.apellido_materno or ''}".strip()
        data_existente.setdefault('dni', dni)
        data_existente.setdefault('nombre', nombre_db)
        data_existente.setdefault('nombres', trabajador_db.nombres or '')
        data_existente.setdefault('cargo', str(trabajador_db.cargo or ''))
        data_existente.setdefault('codigo', dni[:6])
        trabajadores[dni] = data_existente

    # Aplicar filtros de ubicación, proyecto y personal al diccionario de trabajadores
    if any([filtro_ubicacion_id, filtro_proyecto_id, filtro_trabajador_id]):
        trab_filtrados_qs = Trabajador.objects.filter(activo=True)
        if filtro_ubicacion_id:
            trab_filtrados_qs = trab_filtrados_qs.filter(ubicaciones_permitidas__id=filtro_ubicacion_id)
        if filtro_proyecto_id:
            trab_filtrados_qs = trab_filtrados_qs.filter(proyectos__id=filtro_proyecto_id)
        if filtro_trabajador_id:
            trab_filtrados_qs = trab_filtrados_qs.filter(id=filtro_trabajador_id)
        dnis_permitidos = set(str(t.dni).strip() for t in trab_filtrados_qs if t.dni)
        trabajadores = {dni: data for dni, data in trabajadores.items() if dni in dnis_permitidos}

    # Fallback local: si Firestore no trae trabajadores, usamos BD Django.
    if not trabajadores:
        trab_fallback_qs = Trabajador.objects.filter(activo=True)
        if filtro_ubicacion_id:
            trab_fallback_qs = trab_fallback_qs.filter(ubicaciones_permitidas__id=filtro_ubicacion_id)
        if filtro_proyecto_id:
            trab_fallback_qs = trab_fallback_qs.filter(proyectos__id=filtro_proyecto_id)
        if filtro_trabajador_id:
            trab_fallback_qs = trab_fallback_qs.filter(id=filtro_trabajador_id)
        for trabajador_db in trab_fallback_qs.order_by('apellido_paterno', 'apellido_materno', 'nombres'):
            dni = str(trabajador_db.dni or '').strip()
            if not dni:
                continue
            nombre_db = f"{trabajador_db.nombres or ''} {trabajador_db.apellido_paterno or ''} {trabajador_db.apellido_materno or ''}".strip()
            trabajadores[dni] = {
                'dni': dni,
                'nombre': nombre_db,
                'cargo': str(trabajador_db.cargo or ''),
                'codigo': dni[:6],
            }

    # Orden estable por nombre para mantener el formato esperado.
    trabajadores_ordenados = sorted(
        trabajadores.items(),
        key=lambda item: str(
            item[1].get('nombre')
            or item[1].get('nombres')
            or item[1].get('fullName')
            or ''
        ).lower()
    )
    asistencias_ref = db.collection('asistencias').where(filter=FieldFilter('timestamp', '>=', primer_dia)).where(filter=FieldFilter('timestamp', '<=', ultimo_dia)).stream()
    tareo_doc = db.collection('tareos').document(mes_dt.strftime('%Y-%m')).get()
    tareo_del_mes = tareo_doc.to_dict().get('trabajadores', {}) if tareo_doc.exists else {}

    # Justificaciones del mes (Firestore + BD local) para poblar Falt Just.
    justificaciones_por_dni_dia = defaultdict(set)
    justificaciones_ref = db.collection('justificaciones').where(
        filter=FieldFilter('fechaInicio', '<=', ultimo_dia)
    ).where(
        filter=FieldFilter('fechaFin', '>=', primer_dia)
    ).stream()

    primer_dia_date = primer_dia.date()
    ultimo_dia_date = ultimo_dia.date()
    for doc in justificaciones_ref:
        data = doc.to_dict() or {}
        dni = str(data.get('dniTrabajador') or data.get('dni') or '').strip()
        fecha_inicio = data.get('fechaInicio')
        fecha_fin = data.get('fechaFin')
        if not dni or not fecha_inicio or not fecha_fin:
            continue

        fecha_inicio = fecha_inicio.astimezone(LOCAL_TIMEZONE).date()
        fecha_fin = fecha_fin.astimezone(LOCAL_TIMEZONE).date()
        inicio_rango = max(fecha_inicio, primer_dia_date)
        fin_rango = min(fecha_fin, ultimo_dia_date)
        if inicio_rango > fin_rango:
            continue

        fecha_cursor = inicio_rango
        while fecha_cursor <= fin_rango:
            justificaciones_por_dni_dia[dni].add(fecha_cursor.day)
            fecha_cursor += timedelta(days=1)

    justificaciones_locales = Justificacion.objects.filter(
        tareo__fecha__gte=primer_dia_date,
        tareo__fecha__lte=ultimo_dia_date,
        tareo__trabajador__activo=True,
    ).select_related('tareo__trabajador')
    for justificacion in justificaciones_locales:
        estado_solicitud = (justificacion.estado_solicitud or '').upper()
        if estado_solicitud == 'RECHAZADO':
            continue

        dni_local = str(justificacion.tareo.trabajador.dni or '').strip()
        if not dni_local:
            continue
        justificaciones_por_dni_dia[dni_local].add(justificacion.tareo.fecha.day)
    
    # Prioridad 1: asistencias sincronizadas del biométrico en BD local.
    trabajador_por_usuario = {
        t.user_id: str(t.dni or '').strip()
        for t in Trabajador.objects.filter(activo=True).exclude(user__isnull=True)
        if t.user_id
    }

    asistencias_biometrico_por_dni_dia = {}
    asistencias_app_por_dni_dia = {}
    if trabajador_por_usuario:
        asistencias_biometricas = Asistencia.objects.filter(
            usuario_id__in=trabajador_por_usuario.keys(),
            timestamp__date__gte=primer_dia_date,
            timestamp__date__lte=ultimo_dia_date,
            origen__iexact='BIOMETRICO',
        ).values('usuario_id', 'timestamp', 'tipo_marcacion')

        asistencias_app = Asistencia.objects.filter(
            usuario_id__in=trabajador_por_usuario.keys(),
            timestamp__date__gte=primer_dia_date,
            timestamp__date__lte=ultimo_dia_date,
        ).exclude(
            origen__iexact='BIOMETRICO',
        ).values('usuario_id', 'timestamp', 'tipo_marcacion', 'origen')

        for asistencia in asistencias_biometricas:
            dni = trabajador_por_usuario.get(asistencia['usuario_id'])
            timestamp = asistencia.get('timestamp')
            if not dni or not timestamp or dni not in trabajadores:
                continue

            dia = timestamp.astimezone(LOCAL_TIMEZONE).day
            tipo = str(asistencia.get('tipo_marcacion') or '').strip().lower()
            tipo_normalizado = 'Entrada' if tipo in {'entrada', 'e', 'in'} else 'Salida'

            if dni not in asistencias_biometrico_por_dni_dia:
                asistencias_biometrico_por_dni_dia[dni] = {}
            if dia not in asistencias_biometrico_por_dni_dia[dni]:
                asistencias_biometrico_por_dni_dia[dni][dia] = []
            asistencias_biometrico_por_dni_dia[dni][dia].append({
                'timestamp': timestamp,
                'tipoMarcacion': tipo_normalizado,
                'origen': 'BIOMETRICO',
            })

        for asistencia in asistencias_app:
            dni = trabajador_por_usuario.get(asistencia['usuario_id'])
            timestamp = asistencia.get('timestamp')
            if not dni or not timestamp or dni not in trabajadores:
                continue

            dia = timestamp.astimezone(LOCAL_TIMEZONE).day
            tipo = str(asistencia.get('tipo_marcacion') or '').strip().lower()
            tipo_normalizado = 'Entrada' if tipo in {'entrada', 'e', 'in'} else 'Salida'

            if dni not in asistencias_app_por_dni_dia:
                asistencias_app_por_dni_dia[dni] = {}
            if dia not in asistencias_app_por_dni_dia[dni]:
                asistencias_app_por_dni_dia[dni][dia] = []
            asistencias_app_por_dni_dia[dni][dia].append({
                'timestamp': timestamp,
                'tipoMarcacion': tipo_normalizado,
                'origen': str(asistencia.get('origen') or 'APP').upper(),
            })

    # Marcaciones del medio seleccionado (filtro por Asistencia.origen). Se
    # arma aparte porque las propiedades del modelo (hora_entrada_real_calculada)
    # miran TODAS las marcas del dia y no respetarian el filtro.
    asistencias_filtradas_por_dni_dia = {}
    if filtro_origen and trabajador_por_usuario:
        asistencias_filtro_qs = Asistencia.objects.filter(
            usuario_id__in=trabajador_por_usuario.keys(),
            timestamp__date__gte=primer_dia_date,
            timestamp__date__lte=ultimo_dia_date,
            origen__iexact=filtro_origen,
        ).values('usuario_id', 'timestamp', 'tipo_marcacion', 'origen')

        for asistencia in asistencias_filtro_qs:
            dni = trabajador_por_usuario.get(asistencia['usuario_id'])
            timestamp = asistencia.get('timestamp')
            if not dni or not timestamp or dni not in trabajadores:
                continue

            dia = timestamp.astimezone(LOCAL_TIMEZONE).day
            tipo = str(asistencia.get('tipo_marcacion') or '').strip().lower()
            tipo_normalizado = 'Entrada' if tipo in {'entrada', 'e', 'in'} else 'Salida'

            asistencias_filtradas_por_dni_dia.setdefault(dni, {}).setdefault(dia, []).append({
                'timestamp': timestamp,
                'tipoMarcacion': tipo_normalizado,
                'origen': str(asistencia.get('origen') or filtro_origen).upper(),
            })

    # Prioridad 2 (fallback): asistencias históricas en Firestore.
    asistencias_firestore_por_dni_dia = {}
    for doc in asistencias_ref:
        data = doc.to_dict()
        dni, timestamp = data.get('userDni'), data.get('timestamp')
        if dni and timestamp and dni in trabajadores:
            dia = timestamp.astimezone(LOCAL_TIMEZONE).day
            if dni not in asistencias_firestore_por_dni_dia:
                asistencias_firestore_por_dni_dia[dni] = {}
            if dia not in asistencias_firestore_por_dni_dia[dni]:
                asistencias_firestore_por_dni_dia[dni][dia] = []
            asistencias_firestore_por_dni_dia[dni][dia].append(data)

    # Fuente principal para reporte maestro: TareoDiario local (misma base que desempeño).
    tareos_locales = TareoDiario.objects.filter(
        fecha__gte=primer_dia.date(),
        fecha__lte=ultimo_dia.date(),
        trabajador__activo=True,
    ).select_related('trabajador')

    tareos_locales_por_dni_dia = {}
    for tareo_local in tareos_locales:
        dni_local = str(tareo_local.trabajador.dni or '').strip()
        if not dni_local:
            continue
        dia_local = tareo_local.fecha.day
        if dni_local not in tareos_locales_por_dni_dia:
            tareos_locales_por_dni_dia[dni_local] = {}
        tareos_locales_por_dni_dia[dni_local][dia_local] = tareo_local


    # =========================================================================
    # FORMATO VERTICAL: se activa solo cuando se filtra a UN trabajador
    # especifico (una fila por dia del mes, en vez de un trabajador por fila
    # con los dias del mes como columnas). Reutiliza los mismos datos ya
    # obtenidos arriba (Firestore + Postgres); el formato horizontal de mas
    # abajo (para todos los trabajadores) no se modifica en absoluto.
    # =========================================================================
    if filtro_trabajador_id:
        trabajador_obj = Trabajador.objects.filter(id=filtro_trabajador_id, activo=True).first()
        if not trabajador_obj:
            messages.error(request, 'No se encontró el trabajador seleccionado.')
            return redirect('exportar_reporte_maestro')

        dni_v = str(trabajador_obj.dni or '').strip()

        def _v_primera_entrada(marcaciones):
            horas = []
            for m in (marcaciones or []):
                if str(m.get('tipoMarcacion') or '').strip().lower() != 'entrada':
                    continue
                ts = m.get('timestamp')
                if not ts:
                    continue
                try:
                    horas.append(ts.astimezone(LOCAL_TIMEZONE).time())
                except Exception:
                    continue
            return min(horas) if horas else None

        def _v_ultima_salida(marcaciones):
            horas = []
            for m in (marcaciones or []):
                if str(m.get('tipoMarcacion') or '').strip().lower() != 'salida':
                    continue
                ts = m.get('timestamp')
                if not ts:
                    continue
                try:
                    horas.append(ts.astimezone(LOCAL_TIMEZONE).time())
                except Exception:
                    continue
            return max(horas) if horas else None

        def _v_medio(marcaciones):
            for m in (marcaciones or []):
                origen = m.get('origen')
                if origen:
                    return str(origen).upper()
            return '-'

        ESTADO_FILLS_V = {
            'ASISTIÓ':        PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
            'FALTA':          rojo_fill,
            'JUSTIFICADO':    morado_fill,
            'SIN PROGRAMAR':  PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
            'SIN REGISTRO':   PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        }

        # --- TÍTULOS ---
        sheet.merge_cells('A1:E1')
        sheet['A1'] = 'Ceneris E.I.R.L'
        sheet['A1'].font = header_font

        sheet.merge_cells('A2:E2')
        sheet['A2'] = 'Tareo del Periodo — Vista Individual'
        sheet['A2'].font = header_font

        sheet.merge_cells('A3:E3')
        sheet['A3'] = f"{trabajador_obj.nombre_completo} — DNI {dni_v or '-'}"
        sheet['A3'].font = header_font

        meses_es_v = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        sheet.merge_cells('A4:E4')
        area_nombre_v = trabajador_obj.area.nombre if trabajador_obj.area_id else 'Sin área'
        sheet['A4'] = f"Área: {area_nombre_v}   |   Periodo: {meses_es_v[mes_dt.month - 1]} de {mes_dt.year}"
        sheet['G1'] = f"Exportado: {datetime.now(LOCAL_TIMEZONE).strftime('%d/%m/%Y %H:%M:%S')}"

        if filtro_origen:
            sheet.merge_cells('A5:E5')
            aviso_v = sheet['A5']
            aviso_v.value = f"FILTRADO POR MEDIO: {_etiqueta_origen(filtro_origen)} — solo se listan los días marcados por este medio"
            aviso_v.font = Font(name='Arial', size=9, bold=True, color='9C5700')
            aviso_v.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        headers_v = ['Fecha', 'Estado', 'Entrada', 'Salida', 'Medio']
        for col, header in enumerate(headers_v, 1):
            c = sheet.cell(row=6, column=col, value=header)
            c.font = header_font
            c.alignment = header_align
        sheet.row_dimensions[6].height = 20

        dias_abrev_es = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do']
        current_row_v = 7
        for dia in range(1, num_dias + 1):
            fecha_actual = date(mes_dt.year, mes_dt.month, dia)

            tareo_local = tareos_locales_por_dni_dia.get(dni_v, {}).get(dia)
            dia_justificado = dia in justificaciones_por_dni_dia.get(dni_v, set())
            marcaciones_del_dia_priorizadas = (
                asistencias_biometrico_por_dni_dia.get(dni_v, {}).get(dia)
                or asistencias_app_por_dni_dia.get(dni_v, {}).get(dia)
                or asistencias_firestore_por_dni_dia.get(dni_v, {}).get(dia, [])
            )

            estado_label, entrada_str, salida_str, medio_str = 'SIN REGISTRO', '-', '-', '-'

            # --- CON FILTRO POR MEDIO ---
            # Solo se listan los dias con marca de ese medio; no van faltas ni
            # justificaciones. Se calcula desde las marcas filtradas, no desde
            # las propiedades del tareo (que miran todas las marcas del dia).
            if filtro_origen:
                marcas_filtradas = asistencias_filtradas_por_dni_dia.get(dni_v, {}).get(dia, [])
                if not marcas_filtradas:
                    continue

                hora_e = _v_primera_entrada(marcas_filtradas)
                hora_s = _v_ultima_salida(marcas_filtradas)
                if not hora_e and not hora_s:
                    continue

                fecha_label_f = f"{fecha_actual.strftime('%d/%m/%Y')} ({dias_abrev_es[fecha_actual.weekday()]})"
                sheet.cell(row=current_row_v, column=1, value=fecha_label_f).alignment = center_align

                celda_estado_f = sheet.cell(row=current_row_v, column=2, value='ASISTIÓ')
                celda_estado_f.alignment = center_align
                celda_estado_f.font = normal_font
                celda_estado_f.fill = ESTADO_FILLS_V['ASISTIÓ']

                sheet.cell(row=current_row_v, column=3, value=hora_e.strftime('%H:%M') if hora_e else '-').alignment = center_align
                sheet.cell(row=current_row_v, column=4, value=hora_s.strftime('%H:%M') if hora_s else '-').alignment = center_align
                sheet.cell(row=current_row_v, column=5, value=_v_medio(marcas_filtradas)).alignment = center_align

                sheet.row_dimensions[current_row_v].height = 18
                current_row_v += 1
                continue

            if tareo_local is not None:
                estado_local = (tareo_local.resultado or '').upper()
                # hora_entrada_real_calculada consulta Asistencia en vivo (la
                # marca mas temprana real), en vez de leer el campo guardado
                # que puede haber quedado congelado con un dato viejo.
                hora_entrada_local = tareo_local.hora_entrada_real_calculada or _v_primera_entrada(marcaciones_del_dia_priorizadas)

                if estado_local == 'J' or dia_justificado:
                    estado_label = 'JUSTIFICADO'
                elif hora_entrada_local:
                    # Las marcaciones reales mandan sobre un 'F' desactualizado:
                    # si hay marcas, el dia fue trabajado aunque el campo
                    # guardado siga diciendo Falta (ej. carga manual en la BD).
                    hora_salida_local = tareo_local.hora_salida_real_calculada or _v_ultima_salida(marcaciones_del_dia_priorizadas)
                    estado_label = 'ASISTIÓ'
                    entrada_str = hora_entrada_local.strftime('%H:%M')
                    salida_str = hora_salida_local.strftime('%H:%M') if hora_salida_local else '-'
                    # Usamos tareo_local.origen (la misma propiedad que usa
                    # Consulta de Asistencias) en vez de la lista priorizada:
                    # esa lista ignora por completo la App si hay CUALQUIER
                    # marca de biometrico ese dia, aunque la hora mostrada
                    # (arriba) haya venido realmente de la App.
                    medio_str = tareo_local.origen or _v_medio(marcaciones_del_dia_priorizadas) or '-'
                elif estado_local == 'F':
                    estado_label = 'FALTA'
            else:
                tareo_del_dia = tareo_del_mes.get(dni_v, {}).get(str(dia), {}) or {}
                marcaciones_del_dia = marcaciones_del_dia_priorizadas

                if not tareo_del_dia and not marcaciones_del_dia:
                    estado_label = 'SIN PROGRAMAR'
                else:
                    tareo_para_calculo = tareo_del_dia or {'estado': 'O'}
                    resumen_diario = calcular_resumen_diario(marcaciones_del_dia, tareo_para_calculo)
                    hora_entrada_fallback = _v_primera_entrada(marcaciones_del_dia)

                    if dia_justificado:
                        estado_label = 'JUSTIFICADO'
                    elif resumen_diario['estado_celda'] == 'F':
                        estado_label = 'FALTA'
                    elif hora_entrada_fallback:
                        hora_salida_fallback = _v_ultima_salida(marcaciones_del_dia)
                        estado_label = 'ASISTIÓ'
                        entrada_str = hora_entrada_fallback.strftime('%H:%M')
                        salida_str = hora_salida_fallback.strftime('%H:%M') if hora_salida_fallback else '-'
                        medio_str = _v_medio(marcaciones_del_dia)

            fecha_label_v = f"{fecha_actual.strftime('%d/%m/%Y')} ({dias_abrev_es[fecha_actual.weekday()]})"
            sheet.cell(row=current_row_v, column=1, value=fecha_label_v).alignment = center_align

            estado_cell = sheet.cell(row=current_row_v, column=2, value=estado_label)
            estado_cell.alignment = center_align
            estado_cell.font = normal_font
            fill_v = ESTADO_FILLS_V.get(estado_label)
            if fill_v:
                estado_cell.fill = fill_v

            sheet.cell(row=current_row_v, column=3, value=entrada_str).alignment = center_align
            sheet.cell(row=current_row_v, column=4, value=salida_str).alignment = center_align
            sheet.cell(row=current_row_v, column=5, value=medio_str).alignment = center_align

            sheet.row_dimensions[current_row_v].height = 18
            current_row_v += 1

        sheet.column_dimensions['A'].width = 22
        sheet.column_dimensions['B'].width = 16
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 14

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'Reporte_Maestro_{dni_v or trabajador_obj.id}_{mes_seleccionado_str}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    # =========================================================================
    # FORMATO HORIZONTAL (sin cambios): todos los trabajadores, un trabajador
    # por fila y los dias del mes como columnas.
    # =========================================================================

    # --- 4. CONSTRUIR ENCABEZADOS DEL EXCEL ---
    # Título principal
    sheet['B1'] = 'Ceneris E.I.R.L'
    sheet['B2'] = 'Tareo del Periodo'
    sheet['B3'] = mes_dt.strftime('%Y-%m')
    sheet['D1'] = f"Exportado: {datetime.now(LOCAL_TIMEZONE).strftime('%d/%m/%Y %H:%M:%S')}"

    if filtro_origen:
        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=12)
        aviso_h = sheet.cell(row=4, column=1)
        aviso_h.value = (
            f"REPORTE FILTRADO POR MEDIO: {_etiqueta_origen(filtro_origen)} — "
            f"solo marcas de este medio. Los totales del mes NO se calculan en vista filtrada: NO usar para planilla."
        )
        aviso_h.font = Font(name='Arial', size=9, bold=True, color='9C0006')
        aviso_h.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        sheet.row_dimensions[4].height = 18

    # Encabezados de trabajador
    sheet.cell(row=5, column=1, value='Cod.Trab')
    sheet.cell(row=5, column=2, value='Nombre')
    sheet.cell(row=5, column=3, value='Cargo')
    sheet.column_dimensions[get_column_letter(3)].width = 233 / 7 # Convertimos píxeles a unidades de Excel (aprox)

    
    # NUEVA COLUMNA "N. Err"
    sheet.cell(row=5, column=4, value='N.\nErr') # \n para el salto de línea
    sheet.cell(row=5, column=4).alignment = header_align
    sheet.column_dimensions[get_column_letter(4)].width = 27 / 7 # Convertimos píxeles a unidades de Excel (aprox)
    
    # Encabezados de días
    nombres_dias_es = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sá", "Do"] # Lunes=0
    col_inicio_dias = 5
    for dia in range(1, num_dias + 1):
        col = col_inicio_dias + dia - 1
        fecha = date(mes_dt.year, mes_dt.month, dia)
            # Obtén la celda en la fila 5
        cell = sheet.cell(row=5, column=col)
        cell.value = f"{nombres_dias_es[fecha.weekday()]}\n{dia:02d}"
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        
        # --- AÑADIR ESTA LÍNEA PARA AJUSTAR EL ANCHO ---
        # Ancho ampliado (antes 26/7) para que quepa "HH:MM-HH:MM" en vez de solo "15m"/"F"/"J"
        sheet.column_dimensions[get_column_letter(col)].width = 85 / 7 # Convertimos píxeles a unidades de Excel (aprox)
        # --- AÑADIR ESTA LÍNEA PARA AJUSTAR LA ALTURA DE LA FILA 5 ---
        sheet.row_dimensions[5].height = 27.60

    # Encabezados de Resumen (ajústalos a tu necesidad)
    col_resumen_inicio = col_inicio_dias + num_dias
    encabezados_resumen = ['Hrs\nTrab', 'H.Ext\nTipo 1', 'H.Ext Tipo 2', 'H.Ext Tipo 3', 'H.Ext Tipo 4', 'H.Ext Tipo 5', 
    'Hrs\nTard', 'Hrs\nFaltas', 'Dias\nTard', 'Dias\nTrab', 'Dias\nDesc', 'Dias\nFalt', 'Dias\nVaca', 'Dias\nMedi', 'Dias\nPerm', 
    'Falt\nJust', 'Dias\nLicen', 'Lic\n/Goce', 'Dias\nSubsi', 'Lic Mater', 'Lic Pater', 'Dias\nHuelg', 'Dias\nSusp', 'Dias\nCapa', 'Dias\nComi', 'Sali\nTard', 'C', 'O']
    for i, header in enumerate(encabezados_resumen):
        col = col_resumen_inicio + i
        sheet.cell(row=5, column=col, value=header).alignment = header_align
        # Ajustamos los anchos de columna según tus especificaciones
        if i < 8: # Las primeras 8 columnas de resumen
            sheet.column_dimensions[get_column_letter(col)].width = 55 / 7
        else: # El resto
            sheet.column_dimensions[get_column_letter(col)].width = 41 / 7

    # --- 5. LLENAR MATRIZ DE DATOS ---
    current_row = 6
    def _formatear_tardanza_celda(td):
        if not td or td.total_seconds() <= 0:
            return ''
        total_min = int(td.total_seconds() // 60)
        if total_min < 60:
            return f"{total_min}m"
        horas, minutos = divmod(total_min, 60)
        return f"{horas}:{minutos:02d}"

    def _formatear_tiempo_con_unidad(td, mostrar_cero=False):
        if not td or td.total_seconds() <= 0:
            return '0 min' if mostrar_cero else ''

        total_min = int(td.total_seconds() // 60)
        horas, minutos = divmod(total_min, 60)
        if horas and minutos:
            return f"{horas} h {minutos} min"
        if horas:
            return f"{horas} h"
        return f"{minutos} min"

    def _formatear_entrada_salida_celda(hora_entrada, hora_salida):
        entrada_str = hora_entrada.strftime('%H:%M') if hora_entrada else '--:--'
        salida_str = hora_salida.strftime('%H:%M') if hora_salida else '--:--'
        return f"{entrada_str}-{salida_str}"

    def _decimal_horas_a_timedelta(valor):
        try:
            horas = float(valor or 0)
            return timedelta(minutes=int(round(horas * 60)))
        except (TypeError, ValueError):
            return timedelta(0)

    def _calcular_tardanza_desde_entrada(hora_entrada, fecha_actual):
        if not hora_entrada:
            return timedelta(0)

        if fecha_actual and fecha_actual.weekday() == 5:  # Sabado
            referencia_segundos = (9 * 3600)  # 09:00
        else:
            referencia_segundos = (8 * 3600) + (30 * 60)  # 08:30
        entrada_segundos = (hora_entrada.hour * 3600) + (hora_entrada.minute * 60) + hora_entrada.second
        segundos_tarde = max(0, entrada_segundos - referencia_segundos)
        return timedelta(seconds=segundos_tarde)

    def _primera_entrada_marcaciones(marcaciones_del_dia):
        if not marcaciones_del_dia:
            return None

        entradas = []
        for marca in marcaciones_del_dia:
            tipo = str(marca.get('tipoMarcacion') or '').strip().lower()
            if tipo != 'entrada':
                continue

            timestamp = marca.get('timestamp')
            if not timestamp:
                continue

            try:
                entradas.append(timestamp.astimezone(LOCAL_TIMEZONE).time())
            except Exception:
                continue

        if not entradas:
            return None

        return min(entradas)

    def _ultima_salida_marcaciones(marcaciones_del_dia):
        if not marcaciones_del_dia:
            return None

        salidas = []
        for marca in marcaciones_del_dia:
            tipo = str(marca.get('tipoMarcacion') or '').strip().lower()
            if tipo != 'salida':
                continue

            timestamp = marca.get('timestamp')
            if not timestamp:
                continue

            try:
                salidas.append(timestamp.astimezone(LOCAL_TIMEZONE).time())
            except Exception:
                continue

        if not salidas:
            return None

        return max(salidas)

    def _jornada_programada_desde_tareo_local(tareo_local):
        estado_prog = (tareo_local.estado or '').upper()
        if estado_prog in ['', 'D', '.']:
            return timedelta(0)

        if estado_prog == 'J':
            return _decimal_horas_a_timedelta(tareo_local.jornada_horas)

        if estado_prog == 'P' and tareo_local.hora_entrada and tareo_local.hora_salida:
            dt_inicio = datetime.combine(date.today(), tareo_local.hora_entrada)
            dt_fin = datetime.combine(date.today(), tareo_local.hora_salida)
            if dt_fin < dt_inicio:
                dt_fin += timedelta(days=1)
            return max(timedelta(0), dt_fin - dt_inicio)

        reglas = HORARIOS.get(estado_prog)
        if reglas and reglas.get('duracion_jornada'):
            return reglas['duracion_jornada']

        return _decimal_horas_a_timedelta(tareo_local.horas_trabajadas_validas)

    def _horas_trabajadas_netas_local(tareo_local, tardanza_td):
        estado_resultado = (tareo_local.resultado or '').upper()
        if estado_resultado in ['F', 'D', '.']:
            return timedelta(0)

        jornada_programada = _jornada_programada_desde_tareo_local(tareo_local)
        if jornada_programada > timedelta(0):
            return max(timedelta(0), jornada_programada - tardanza_td)

        horas_validas = _decimal_horas_a_timedelta(tareo_local.horas_trabajadas_validas)
        return max(timedelta(0), horas_validas - tardanza_td)

    for dni, data_trabajador in trabajadores_ordenados:
        codigo_trab = (
            data_trabajador.get('codigo')
            or data_trabajador.get('codTrab')
            or data_trabajador.get('codigoTrabajador')
            or str(dni)[:6]
        )
        sheet.cell(row=current_row, column=1, value=codigo_trab)

        nombre_trab = (
            data_trabajador.get('nombre')
            or data_trabajador.get('nombres')
            or data_trabajador.get('fullName')
            or (
                f"{data_trabajador.get('apellido_paterno', '')} {data_trabajador.get('apellido_materno', '')} {data_trabajador.get('dni', dni)}"
            ).strip()
        )
        sheet.cell(
            row=current_row,
            column=2,
            value=nombre_trab
        )
        sheet.cell(
            row=current_row,
            column=3,
            value=data_trabajador.get('cargo') or data_trabajador.get('puesto') or ''
        )
        
        # Inicializamos los acumuladores del mes para este trabajador
        total_hrs_trab_mes, total_hrs_tard_mes, total_hrs_ext_mes, total_hrs_faltantes_mes = timedelta(), timedelta(), timedelta(), timedelta()
        total_faltas_mes, total_dias_tard_mes, total_dias_trab_mes, total_dias_desc_mes = 0, 0, 0, 0
        total_faltas_just_mes = len(justificaciones_por_dni_dia.get(dni, set()))
        total_min_tardanza_mes = 0
        

        for dia in range(1, num_dias + 1):
            col = 5 + dia - 1
            cell = sheet.cell(row=current_row, column=col)
            fecha_actual = date(mes_dt.year, mes_dt.month, dia)

            # --- CON FILTRO POR MEDIO ---
            # Solo se pinta el horario marcado por ese medio. No van faltas (F)
            # ni justificaciones (J), y los totales del mes quedan en blanco
            # (mas abajo) para no publicar cifras parciales como si fueran las
            # reales.
            if filtro_origen:
                marcas_filtradas_dia = asistencias_filtradas_por_dni_dia.get(dni, {}).get(dia, [])
                hora_e_f = _primera_entrada_marcaciones(marcas_filtradas_dia)
                hora_s_f = _ultima_salida_marcaciones(marcas_filtradas_dia)

                cell.value = ''
                cell.alignment = center_align
                cell.fill = PatternFill(fill_type=None)
                cell.font = normal_font

                if hora_e_f or hora_s_f:
                    cell.value = _formatear_entrada_salida_celda(hora_e_f, hora_s_f)
                    if _calcular_tardanza_desde_entrada(hora_e_f, fecha_actual) > timedelta(0):
                        cell.fill = naranja_fill
                        cell.font = naranja_font
                continue

            tareo_local = tareos_locales_por_dni_dia.get(dni, {}).get(dia)
            dia_justificado = dia in justificaciones_por_dni_dia.get(dni, set())
            sin_tareo_programado = False
            marcaciones_del_dia_priorizadas = (
                asistencias_biometrico_por_dni_dia.get(dni, {}).get(dia)
                or asistencias_app_por_dni_dia.get(dni, {}).get(dia)
                or asistencias_firestore_por_dni_dia.get(dni, {}).get(dia, [])
            )

            # Fallback al cálculo cuando no existe tareo local.
            if tareo_local is None:
                tareo_del_dia = tareo_del_mes.get(dni, {}).get(str(dia), {}) or {}
                marcaciones_del_dia = marcaciones_del_dia_priorizadas

                if not tareo_del_dia and not marcaciones_del_dia:
                    sin_tareo_programado = True
                    resumen_diario = {
                        'hrs_trabajadas': timedelta(),
                        'hrs_tardanza': timedelta(),
                        'hrs_extra': timedelta(),
                        'hrs_faltantes': timedelta(),
                        'estado_celda': '',
                    }
                else:
                    # Si hay marcaciones pero no tareo programado en app, asumimos jornada oficina para evitar "X" injustificada.
                    tareo_para_calculo = tareo_del_dia or {'estado': 'O'}
                    resumen_diario = calcular_resumen_diario(marcaciones_del_dia, tareo_para_calculo)

                hora_entrada_fallback = _primera_entrada_marcaciones(marcaciones_del_dia)
                tardanza_fallback_td = _calcular_tardanza_desde_entrada(hora_entrada_fallback, fecha_actual)
            else:
                resumen_diario = None
                tardanza_fallback_td = timedelta(0)

            # Reglas de visualización diaria:
            # - Puntual: celda en blanco
            # - Tardanza: mostrar minutos u horas:minutos y fondo naranja suave
            # - Falta: mostrar F y fondo rojo con texto visible
            cell.value = ''
            cell.alignment = center_align # Centramos el contenido
            cell.fill = PatternFill(fill_type=None)
            cell.font = normal_font

            if tareo_local is not None:
                estado_local = (tareo_local.resultado or '').upper()
                # hora_entrada_real_calculada consulta Asistencia en vivo (la
                # marca mas temprana real), en vez de leer el campo guardado
                # que puede haber quedado congelado con un dato viejo.
                hora_entrada_local = tareo_local.hora_entrada_real_calculada or _primera_entrada_marcaciones(marcaciones_del_dia_priorizadas)
                tardanza_local_td = _calcular_tardanza_desde_entrada(hora_entrada_local, fecha_actual)
                horas_trabajadas_netas_dia = _horas_trabajadas_netas_local(tareo_local, tardanza_local_td)

                if estado_local == 'J' or dia_justificado:
                    cell.value = 'J'
                    cell.fill = morado_fill
                    cell.font = morado_font
                elif hora_entrada_local:
                    # Las marcaciones reales mandan sobre un 'F' desactualizado:
                    # si hay marcas, el dia fue trabajado aunque el campo
                    # guardado siga diciendo Falta (ej. carga manual en la BD).
                    hora_salida_local = tareo_local.hora_salida_real_calculada or _ultima_salida_marcaciones(marcaciones_del_dia_priorizadas)
                    cell.value = _formatear_entrada_salida_celda(hora_entrada_local, hora_salida_local)
                    if tardanza_local_td > timedelta(0):
                        cell.fill = naranja_fill
                        cell.font = naranja_font
                elif estado_local == 'F':
                    cell.value = 'F'
                    cell.fill = rojo_fill
                    cell.font = rojo_font

                total_hrs_trab_mes += horas_trabajadas_netas_dia
                total_hrs_tard_mes += tardanza_local_td
                total_min_tardanza_mes += int(tardanza_local_td.total_seconds() // 60)
                if estado_local == 'F':
                    total_faltas_mes += 1
                if (tareo_local.estado or '').upper() in ['D', '.']:
                    total_dias_desc_mes += 1
                if tardanza_local_td > timedelta(0):
                    total_dias_tard_mes += 1
                if estado_local not in ['F', 'D', '.']:
                    total_dias_trab_mes += 1
            else:
                if sin_tareo_programado:
                    cell.value = 'X'
                elif dia_justificado:
                    cell.value = 'J'
                    cell.fill = morado_fill
                    cell.font = morado_font
                elif resumen_diario['estado_celda'] == 'F':
                    cell.value = 'F'
                    cell.fill = rojo_fill
                    cell.font = rojo_font
                elif hora_entrada_fallback:
                    hora_salida_fallback = _ultima_salida_marcaciones(marcaciones_del_dia)
                    cell.value = _formatear_entrada_salida_celda(hora_entrada_fallback, hora_salida_fallback)
                    if tardanza_fallback_td > timedelta(0):
                        cell.fill = naranja_fill
                        cell.font = naranja_font

                total_hrs_trab_mes += max(timedelta(0), resumen_diario['hrs_trabajadas'] - resumen_diario['hrs_tardanza'])
                total_hrs_tard_mes += tardanza_fallback_td
                total_min_tardanza_mes += int(tardanza_fallback_td.total_seconds() // 60)
                total_hrs_ext_mes += resumen_diario['hrs_extra']
                total_hrs_faltantes_mes += resumen_diario['hrs_faltantes']
                if resumen_diario['estado_celda'] == 'F':
                    total_faltas_mes += 1
                if resumen_diario['estado_celda'] == '.':
                    total_dias_desc_mes += 1
                if tardanza_fallback_td > timedelta(0):
                    total_dias_tard_mes += 1
                if resumen_diario['estado_celda'] not in ['F', 'D', '.']:
                    total_dias_trab_mes += 1

        # 5. ESCRIBIR COLUMNAS DE RESUMEN
        for i, header in enumerate(encabezados_resumen):
            col = col_resumen_inicio + i
            valor_resumen = "" # Por defecto, celda vacía

            # Con filtro por medio los totales quedan en blanco: contarian solo
            # los dias de ese medio y alguien podria tomarlos como los totales
            # reales del mes (ej. "Dias Trab: 12" cuando trabajo 22).
            if filtro_origen:
                resumen_cell = sheet.cell(row=current_row, column=col, value="")
                resumen_cell.alignment = header_align
                resumen_cell.font = normal_font
                continue

            # Aquí mapeas cada encabezado a su valor calculado
            if header == "Hrs\nTrab": valor_resumen = ""
            if header == "Hrs\nTard": valor_resumen = _formatear_tiempo_con_unidad(timedelta(minutes=total_min_tardanza_mes), mostrar_cero=True)
            if header == "H.Ext\nTipo 1": valor_resumen = _formatear_tiempo_con_unidad(total_hrs_ext_mes)
            if header == "Hrs\nFaltas": valor_resumen = _formatear_tiempo_con_unidad(total_hrs_faltantes_mes)
            if header == "Dias\nTard": valor_resumen = total_dias_tard_mes
            if header == "Dias\nDesc": valor_resumen = total_dias_desc_mes 
            if header == "Dias\nTrab": valor_resumen = total_dias_trab_mes 
            if header == "Dias\nFalt": valor_resumen = total_faltas_mes
            if header == "Falt\nJust": valor_resumen = total_faltas_just_mes
            # ... y así sucesivamente para todas las otras columnas de resumen
            
            resumen_cell = sheet.cell(row=current_row, column=col, value=valor_resumen)
            resumen_cell.alignment = header_align
            resumen_cell.font = normal_font

        current_row += 1
        

    # --- 7. AJUSTES FINALES Y GENERACIÓN DEL ARCHIVO ---
    # Ajustar anchos
    sheet.column_dimensions['A'].width = 62/7
    sheet.column_dimensions['B'].width = 255/7
    sheet.column_dimensions['C'].width = 182/7
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'Reporte_Maestro_{mes_seleccionado_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    
    return response


@login_required
@group_required('Recursos Humanos', 'Calidad')
def gestion_reporte_maestro(request):
    """Muestra el dashboard de gestión de reportes con filtros de exportación."""
    context = {
        'current_view': 'gestion_reportes',
        'ubicaciones':        Ubicacion.objects.all().order_by('nombre'),
        'proyectos':          Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
        'trabajadores_lista': Trabajador.objects.filter(activo=True).order_by('apellido_paterno', 'nombres'),
        'origenes':           ORIGENES_MARCACION,
    }
    return render(request, 'recursoshumanos/reportes/gestion_reporte_maestro.html', context)


@login_required
@group_required('Recursos Humanos')
def reporte_justificaciones_rrhh(request):
    """Vista RRHH para revisar y gestionar justificaciones enviadas desde la app móvil."""
    hoy = timezone.localdate()
    mes_param = (request.GET.get('mes') or hoy.strftime('%Y-%m')).strip()
    estado_param = (request.GET.get('estado') or 'todos').strip().upper()
    busqueda = (request.GET.get('q') or '').strip()

    try:
        mes_dt = datetime.strptime(mes_param, '%Y-%m')
    except ValueError:
        mes_dt = datetime(hoy.year, hoy.month, 1)
        mes_param = mes_dt.strftime('%Y-%m')
        messages.warning(request, 'Mes inválido. Se mostró el mes actual.')

    primer_dia = mes_dt.replace(day=1).date()
    ultimo_dia = mes_dt.replace(day=calendar.monthrange(mes_dt.year, mes_dt.month)[1]).date()

    base_mes_qs = Justificacion.objects.filter(
        tareo__fecha__gte=primer_dia,
        tareo__fecha__lte=ultimo_dia,
        tareo__trabajador__activo=True,
    ).exclude(
        descripcion__istartswith='Importado desde ERP:'
    ).select_related('tareo__trabajador').order_by('-creado_en', '-id')

    if request.method == 'POST':
        justificacion_id = request.POST.get('justificacion_id')
        accion = (request.POST.get('accion') or '').strip().lower()
        justificacion = base_mes_qs.filter(id=justificacion_id).first()

        if not justificacion:
            messages.error(request, 'No se encontró la justificación seleccionada para este periodo.')
            return redirect(request.get_full_path())

        if accion == 'aprobar':
            justificacion.estado_solicitud = 'APROBADO'
            justificacion.motivo_rechazo = ''
            justificacion.save(update_fields=['estado_solicitud', 'motivo_rechazo'])

            # Igual que hace la sincronizacion del ERP/biometrico: al aprobar,
            # el dia debe reflejarse como Justificado. Sin esto, la solicitud
            # quedaba "Aprobada" pero el tareo seguia contando como Falta en
            # Consulta de Asistencias y en el Reporte de Faltas.
            tareo = justificacion.tareo
            if tareo.resultado != 'J':
                tareo.resultado = 'J'
                tareo.save(update_fields=['resultado'])

            messages.success(request, 'Justificación aprobada correctamente.')
        elif accion == 'rechazar':
            motivo_rechazo = (request.POST.get('motivo_rechazo') or '').strip()
            if not motivo_rechazo:
                messages.error(request, 'Debes ingresar un motivo de rechazo.')
                return redirect(request.get_full_path())

            justificacion.estado_solicitud = 'RECHAZADO'
            justificacion.motivo_rechazo = motivo_rechazo
            justificacion.save(update_fields=['estado_solicitud', 'motivo_rechazo'])

            # Si se rechaza una justificacion que ya estaba aprobada (y por
            # tanto ya habia marcado el dia como 'J'), el dia vuelve a Falta.
            tareo = justificacion.tareo
            if tareo.resultado == 'J':
                tareo.resultado = 'F'
                tareo.save(update_fields=['resultado'])

            messages.success(request, 'Justificación rechazada correctamente.')
        elif accion == 'editar':
            motivo_nuevo = (request.POST.get('motivo') or '').strip().upper()
            descripcion_nueva = (request.POST.get('descripcion') or '').strip()
            motivo_rechazo_nuevo = (request.POST.get('motivo_rechazo') or '').strip()

            motivos_validos = {clave for clave, _ in Justificacion.MOTIVOS}
            if motivo_nuevo not in motivos_validos:
                messages.error(request, 'Motivo inválido para la justificación.')
                return redirect(request.get_full_path())
            if not descripcion_nueva:
                messages.error(request, 'La descripción de la justificación no puede estar vacía.')
                return redirect(request.get_full_path())

            justificacion.motivo = motivo_nuevo
            justificacion.descripcion = descripcion_nueva
            update_fields = ['motivo', 'descripcion']

            if justificacion.estado_solicitud == 'RECHAZADO':
                justificacion.motivo_rechazo = motivo_rechazo_nuevo
                update_fields.append('motivo_rechazo')

            justificacion.save(update_fields=update_fields)
            messages.success(request, 'Justificación actualizada correctamente.')
        else:
            messages.error(request, 'Acción inválida.')

        return redirect(request.get_full_path())

    justificaciones_qs = base_mes_qs
    if estado_param in {'PENDIENTE', 'APROBADO', 'RECHAZADO'}:
        justificaciones_qs = justificaciones_qs.filter(estado_solicitud=estado_param)
    else:
        estado_param = 'TODOS'

    if busqueda:
        justificaciones_qs = justificaciones_qs.filter(
            Q(tareo__trabajador__dni__icontains=busqueda)
            | Q(tareo__trabajador__nombres__icontains=busqueda)
            | Q(tareo__trabajador__apellido_paterno__icontains=busqueda)
            | Q(tareo__trabajador__apellido_materno__icontains=busqueda)
        )

    resumen = {
        'TOTAL': base_mes_qs.count(),
        'PENDIENTE': base_mes_qs.filter(estado_solicitud='PENDIENTE').count(),
        'APROBADO': base_mes_qs.filter(estado_solicitud='APROBADO').count(),
        'RECHAZADO': base_mes_qs.filter(estado_solicitud='RECHAZADO').count(),
    }

    context = {
        'current_view': 'reporte_justificaciones_rrhh',
        'justificaciones': justificaciones_qs,
        'mes_param': mes_param,
        'estado_param': estado_param,
        'busqueda': busqueda,
        'resumen': resumen,
        'motivos_disponibles': Justificacion.MOTIVOS,
    }
    return render(request, 'recursoshumanos/reportes/justificaciones_rrhh.html', context)


@csrf_exempt
@require_POST
def recibir_respuesta_google_form(request):
    secret_key = request.headers.get('X-Secret-Key')
    if secret_key != settings.GOOGLE_FORM_SECRET_KEY:
        return HttpResponseForbidden('Acceso denegado.')

    try:
        data = json.loads(request.body)
        email = data.get('email')
        timestamp_str = data.get('timestamp')
        respuestas_dict = data.get('respuestas', {})
        
        if not email or not timestamp_str or not respuestas_dict:
            return HttpResponse('Datos incompletos.', status=400)

        # Convierte el timestamp de Google a un formato que Django entiende
        from dateutil.parser import parse
        timestamp_dt = parse(timestamp_str)

        with transaction.atomic():
            # Intenta encontrar un trabajador con ese email
            trabajador = Trabajador.objects.filter(email__iexact=email).first()

            # Crea el registro principal del resultado
            resultado = ResultadoCuestionario.objects.create(
                trabajador=trabajador,
                email=email,
                timestamp=timestamp_dt
            )

            # Busca todas las preguntas de una vez para eficiencia
            textos_preguntas = list(respuestas_dict.keys())
            preguntas_qs = Pregunta.objects.filter(texto__in=textos_preguntas).in_bulk(field_name='texto')

            # Guarda cada respuesta individualmente
            respuestas_a_crear = []
            for texto_pregunta, valor_respuesta in respuestas_dict.items():
                pregunta_obj = preguntas_qs.get(texto_pregunta)
                if pregunta_obj:
                    respuestas_a_crear.append(
                        Respuesta(resultado=resultado, pregunta=pregunta_obj, valor=valor_respuesta)
                    )
            Respuesta.objects.bulk_create(respuestas_a_crear)

        # Aquí podrías lanzar una tarea en segundo plano para calcular los puntajes
        # ej: calcular_puntajes.delay(resultado.id)

        return HttpResponse(f'OK: Respuesta de {email} recibida y guardada.')

    except Exception as e:
        print(f"ERROR al procesar la respuesta de Google Form: {e}")
        return HttpResponse(f'Error: {e}', status=500)

@login_required
@require_POST # Asegura que solo se pueda eliminar con una solicitud POST
def eliminar_trabajador(request, pk):
    trabajador = get_object_or_404(Trabajador, pk=pk)
    try:
        # 1. Eliminar de Firestore primero, usando el PK
        db.collection('trabajadores').document(str(pk)).delete()
        
        # 2. Luego, eliminar de la base de datos de Django
        nombre_completo = str(trabajador)
        trabajador.delete()
        
        messages.success(request, f'El trabajador "{nombre_completo}" ha sido eliminado exitosamente.')
    except Exception as e:
        messages.error(request, f'Error al eliminar al trabajador: {e}.')
    
    return redirect('calidad:lista_trabajadores')

@login_required
def gestion_dispositivos_panel(request):
    """
    Muestra el panel de gestión con tarjetas para las diferentes
    funcionalidades de administración de dispositivos.
    """
    context = {
        'current_view': 'gestion_dispositivos', # Mantenemos esto para que el sidebar se ilumine
    }
    return render(request, 'recursoshumanos/dispositivos/gestion_dispositivos_panel.html', context)

#==================================
# GESTIÓN DE DISPOSITIVOS
#==================================

@login_required
def lista_dispositivos(request):
    """
    Muestra una lista de todos los dispositivos registrados en la base de datos
    de Django (PostgreSQL).
    """
    try:
        dispositivos = Dispositivo.objects.prefetch_related('trabajadores_permitidos').all().order_by('nombre')
        
    except Exception as e:
        messages.error(request, f"No se pudieron cargar los dispositivos: {e}")
        dispositivos = [] # Si hay un error, pasamos una lista vacía a la plantilla

    context = {
        'dispositivos': dispositivos,
        'current_view': 'gestion_dispositivos',
    }
    return render(request, 'recursoshumanos/dispositivos/lista_dispositivos.html', context)

@login_required
def crear_dispositivo(request):
    if request.method == 'POST':
        form = DispositivoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Dispositivo creado exitosamente.")
            return redirect('recursoshumanos:lista_dispositivos')
    else:
        form = DispositivoForm()

    return render(request, 'recursoshumanos/dispositivos/dispositivo_form.html', {
        'form': form,
        'titulo': 'Crear Nuevo Dispositivo'
    })

@login_required
def editar_dispositivo(request, pk):
    dispositivo = get_object_or_404(Dispositivo, pk=pk)
    
    if request.method == 'POST':
        # Pasamos 'instance=dispositivo' para que sepa que va a editar y no a crear
        form = DispositivoForm(request.POST, instance=dispositivo)
        if form.is_valid():
            form.save()
            messages.success(request, "Dispositivo actualizado correctamente.")
            return redirect('recursoshumanos:lista_dispositivos')
    else:
        form = DispositivoForm(instance=dispositivo)

    return render(request, 'recursoshumanos/dispositivos/dispositivo_form.html', {
        'form': form,
        'titulo': f'Editar {dispositivo.nombre}'
    })

@login_required
def eliminar_dispositivo(request, pk):
    dispositivo = get_object_or_404(Dispositivo, pk=pk)
    
    if request.method == 'POST':
        dispositivo.delete()
        messages.success(request, "Dispositivo eliminado correctamente.")
        return redirect('recursoshumanos:lista_dispositivos')

    # Renderiza una plantilla de confirmación
    return render(request, 'recursoshumanos/dispositivos/dispositivo_confirm_delete.html', {
        'dispositivo': dispositivo
    })

def group_check(user):
    return user.groups.filter(name__in=["Recursos Humanos", "Supervisores", "Gerencia"]).exists()

@login_required
@user_passes_test(group_check)
def panel_aprobacion_horas_extra(request):
    user = request.user
    
    # Optimizamos la consulta trayendo también los datos de los aprobadores anteriores
    # para evitar consultas extra en el template cuando el Gerente vea la lista.
    base_query = SolicitudHorasExtra.objects.select_related(
        'trabajador', 
        'aprobado_por_operador', 
        'aprobado_por_admin',
        'aprobado_por_gerente'
    )

    solicitudes_pendientes = base_query.none()

    # Banderas para usar en el template
    is_gerente = user.groups.filter(name='Gerencia').exists()
    is_rrhh = user.groups.filter(name='Recursos Humanos').exists()
    is_supervisor = user.groups.filter(name='Supervisores').exists()

    # Etiquetas del flujo (Supervisión -> RRHH -> Gerencia) para informar al usuario
    # en qué paso está y a quién le llega la solicitud después de su aprobación.
    nivel_actual = ''
    paso_actual = 0
    siguiente_nivel = ''

    if is_supervisor:
        # Supervisor ve PENDIENTE_OPERADOR
        solicitudes_pendientes = base_query.filter(estado=SolicitudHorasExtra.Estado.PENDIENTE_OPERADOR)
        nivel_actual = 'Supervisión'
        paso_actual = 1
        siguiente_nivel = 'Recursos Humanos'

    elif is_rrhh:
        # RRHH ve PENDIENTE_ADMIN
        solicitudes_pendientes = base_query.filter(estado=SolicitudHorasExtra.Estado.PENDIENTE_ADMIN)
        nivel_actual = 'Recursos Humanos'
        paso_actual = 2
        siguiente_nivel = 'Gerencia'

    elif is_gerente:
        # Gerencia ve PENDIENTE_GERENTE (y necesita ver quién aprobó antes)
        solicitudes_pendientes = base_query.filter(estado=SolicitudHorasExtra.Estado.PENDIENTE_GERENTE)
        nivel_actual = 'Gerencia'
        paso_actual = 3
        siguiente_nivel = ''  # Último nivel: su aprobación cierra el flujo

    historial_qs = base_query.filter(
        estado__in=[
            SolicitudHorasExtra.Estado.APROBADO,
            SolicitudHorasExtra.Estado.RECHAZADO,
        ]
    ).order_by('-creado_en', '-id')

    historial_paginator = Paginator(historial_qs, 15)
    historial_page_number = request.GET.get('historial_page')
    historial_page_obj = historial_paginator.get_page(historial_page_number)

    context = {
        'solicitudes_pendientes': solicitudes_pendientes,
        'historial_page_obj': historial_page_obj,
        'historial_total': historial_qs.count(),
        'current_view': 'panel_aprobacion_he',
        'is_gerente': is_gerente, # Pasamos esto para activar la columna extra en HTML
        'is_rrhh': is_rrhh,       # Opcional: por si RRHH también quiere ver quién fue el operador
        'is_supervisor': is_supervisor,
        'nivel_actual': nivel_actual,
        'paso_actual': paso_actual,
        'siguiente_nivel': siguiente_nivel,
    }
    return render(request, 'recursoshumanos/horas_extra/panel_aprobacion.html', context)

@login_required
@require_POST
def procesar_solicitud_horas_extra(request, solicitud_id, accion):
    """
    Vista de acción para aprobar o rechazar una solicitud siguiendo el flujo:
    Operador -> Admin (RRHH) -> Gerente -> APROBADO FINAL.
    
    IMPORTANTE: Al aprobarse finalmente, se dispara el recálculo del TareoDiario
    para que las horas extra se sumen a las horas válidas del trabajador.
    """
    solicitud = get_object_or_404(SolicitudHorasExtra, pk=solicitud_id)
    user = request.user

    # Pre-cargamos permisos
    es_supervisor = user.groups.filter(name='Supervisores').exists()
    es_rrhh = user.groups.filter(name='Recursos Humanos').exists()
    es_gerente = user.groups.filter(name='Gerencia').exists()

    # Nombre del trabajador para que el mensaje indique sobre qué solicitud se actuó
    nombre_trabajador = str(solicitud.trabajador)

    # --- LÓGICA DE APROBACIÓN ---
    if accion == 'aprobar':
        
        # 1. Nivel OPERADOR (Pasa a Admin)
        if solicitud.estado == SolicitudHorasExtra.Estado.PENDIENTE_OPERADOR:
            if es_supervisor:
                solicitud.aprobado_por_operador = user
                solicitud.fecha_aprobacion_operador = timezone.now()
                solicitud.estado = SolicitudHorasExtra.Estado.PENDIENTE_ADMIN
                solicitud.save()
                messages.success(
                    request,
                    f'Paso 1 de 3 completado: aprobaste la solicitud de {nombre_trabajador}. '
                    'Se envió a Recursos Humanos para su aprobación.'
                )
            else:
                messages.error(request, 'No tienes permisos de Supervisor.')

        # 2. Nivel ADMIN/RRHH (Pasa a Gerencia)
        elif solicitud.estado == SolicitudHorasExtra.Estado.PENDIENTE_ADMIN:
            if es_rrhh:
                solicitud.aprobado_por_admin = user
                solicitud.fecha_aprobacion_admin = timezone.now()
                solicitud.estado = SolicitudHorasExtra.Estado.PENDIENTE_GERENTE
                solicitud.save()
                messages.success(
                    request,
                    f'Paso 2 de 3 completado: Recursos Humanos aprobó la solicitud de {nombre_trabajador}. '
                    'Se envió a Gerencia para la aprobación final.'
                )
            else:
                messages.error(request, 'No tienes permisos de RRHH.')

        # 3. Nivel GERENCIA (Aprobación Final)
        elif solicitud.estado == SolicitudHorasExtra.Estado.PENDIENTE_GERENTE:
            if es_gerente:
                solicitud.aprobado_por_gerente = user
                solicitud.fecha_aprobacion_gerente = timezone.now()
                solicitud.estado = SolicitudHorasExtra.Estado.APROBADO
                
                # Guardamos PRIMERO para que el estado 'APROBADO' esté en la BD
                solicitud.save() 
                
                # --- AQUÍ ESTÁ EL CAMBIO CLAVE ---
                # Buscamos el tareo de ese día y forzamos el recálculo
                try:
                    tareo = TareoDiario.objects.get(
                        trabajador=solicitud.trabajador,
                        fecha=solicitud.fecha_horas_extra
                    )
                    recalcular_asistencia_diaria(tareo)
                    messages.success(
                        request,
                        f'Paso 3 de 3 completado: Gerencia dio la aprobación final a la solicitud de '
                        f'{nombre_trabajador}. El proceso terminó y las horas ya se sumaron a su tareo.'
                    )
                except TareoDiario.DoesNotExist:
                    # Si no existe tareo (ej: pidieron hora extra para un día futuro), solo aprobamos la solicitud.
                    messages.success(
                        request,
                        f'Paso 3 de 3 completado: Gerencia dio la aprobación final a la solicitud de '
                        f'{nombre_trabajador}. Las horas se sumarán cuando exista el tareo de ese día.'
                    )
                except Exception as e:
                    messages.warning(
                        request,
                        f'Solicitud aprobada por Gerencia, pero hubo un error recalculando el tareo: {e}'
                    )
            else:
                messages.error(request, 'No tienes permisos de Gerencia.')
        
        else:
            messages.warning(request, 'El estado de la solicitud no permite aprobación o ya fue procesada.')

    # --- LÓGICA DE RECHAZO ---
    elif accion == 'rechazar':
        estado_actual = solicitud.estado
        motivo = request.POST.get('motivo_rechazo', 'Sin motivo especificado.')

        # Trazabilidad de quién rechazó
        if estado_actual == SolicitudHorasExtra.Estado.PENDIENTE_OPERADOR and es_supervisor:
            solicitud.aprobado_por_operador = user
            solicitud.fecha_aprobacion_operador = timezone.now()
            nivel_rechazo = 'Supervisión'
        elif estado_actual == SolicitudHorasExtra.Estado.PENDIENTE_ADMIN and es_rrhh:
            solicitud.aprobado_por_admin = user
            solicitud.fecha_aprobacion_admin = timezone.now()
            nivel_rechazo = 'Recursos Humanos'
        elif estado_actual == SolicitudHorasExtra.Estado.PENDIENTE_GERENTE and es_gerente:
            solicitud.aprobado_por_gerente = user
            solicitud.fecha_aprobacion_gerente = timezone.now()
            nivel_rechazo = 'Gerencia'
        else:
            messages.error(request, 'No tienes permisos para rechazar esta solicitud en su estado actual.')
            return redirect('recursoshumanos:panel_aprobacion_he')

        solicitud.estado = SolicitudHorasExtra.Estado.RECHAZADO
        solicitud.motivo_rechazo = f"[{user.get_full_name() or user.username}] {motivo}"

        solicitud.save()
        
        # Opcional: Si se rechaza algo que ya estaba aprobado (raro, pero posible),
        # también deberíamos recalcular para quitarle las horas. 
        # Pero en este flujo lineal (Pendiente->Aprobado) no es estrictamente necesario.
        
        messages.warning(
            request,
            f'Solicitud de {nombre_trabajador} rechazada en el nivel de {nivel_rechazo}. '
            'El flujo se detiene aquí y no continúa a los siguientes niveles.'
        )

    return redirect('recursoshumanos:panel_aprobacion_he')


@login_required
def reporte_faltas_web(request):
    # 1. Filtro de Mes (Por defecto el actual)
    mes_str = request.GET.get('mes', datetime.now().strftime('%Y-%m'))
    try:
        anio, mes = map(int, mes_str.split('-'))
    except ValueError:
        # Fallback si el formato es incorrecto
        hoy_temp = datetime.now()
        anio, mes = hoy_temp.year, hoy_temp.month
        mes_str = hoy_temp.strftime('%Y-%m')

    # Fecha de referencia
    hoy = timezone.now().date()

    # 2. QuerySet Base
    faltas = TareoDiario.objects.filter(
        resultado='F',
        fecha__year=anio,
        fecha__month=mes,
        fecha__lt=hoy  # Excluye hoy y futuro
    ).select_related('trabajador', 'justificacion').order_by('-fecha')

    # 3. Filtro por Búsqueda (Nombre o DNI)
    search_query = request.GET.get('q', '')
    if search_query:
        # Aquí buscamos en nombres, apellido paterno, materno o DNI
        faltas = faltas.filter(
            Q(trabajador__nombres__icontains=search_query) |
            Q(trabajador__apellido_paterno__icontains=search_query) |
            Q(trabajador__apellido_materno__icontains=search_query) |
            Q(trabajador__dni__icontains=search_query)
        )

    # 4. Filtro por Estado de Justificación
    estado_filter = request.GET.get('estado_justificacion', '')
    if estado_filter == 'SIN_JUSTIFICAR':
        faltas = faltas.filter(justificacion__isnull=True)
    elif estado_filter:
        # Asumiendo que tu modelo Justificacion tiene un campo 'estado_solicitud'
        faltas = faltas.filter(justificacion__estado_solicitud=estado_filter)

    context = {
        'faltas': faltas,
        'mes_actual': mes_str,
        'search_query': search_query,
        'estado_filter': estado_filter,
        'current_view': 'reporte_faltas',
    }
    return render(request, 'recursoshumanos/reportes/lista_faltas.html', context)


@login_required
@group_required("Recursos Humanos", "Gerencia", "Administracion","Calidad") # Ajusta los grupos según necesites
def gestion_areas(request):
    """Panel principal (Dashboard) de Áreas"""
    return render(request, 'recursoshumanos/areas/gestion_areas.html', {
        'current_view': 'gestion_areas'
    })

@login_required
@group_required("Recursos Humanos", "Gerencia", "Administracion","Calidad")
def lista_areas(request):
    """Lista tipo tabla de todas las áreas con sus respectivos jefes"""
    
    # 1. OPTIMIZACIÓN: Pre-cargamos solo los jefes activos usando el nombre por defecto de Django
    jefes_prefetch = Prefetch(
        'trabajador_set', # <--- EL NOMBRE CORRECTO SEGÚN TU MODELO
        queryset=Trabajador.objects.filter(es_jefe=True, activo=True),
        to_attr='jefes_activos'
    )
    
    # 2. CONSULTA: Traemos las áreas e inyectamos los jefes pre-cargados
    areas = Area.objects.all().order_by('nombre').prefetch_related(jefes_prefetch)
    
    # 3. RENDER: Enviamos a la plantilla
    return render(request, 'recursoshumanos/areas/lista_areas.html', {
        'areas': areas,
        'current_view': 'gestion_areas'
    })

@login_required
@group_required("Recursos Humanos", "Gerencia", "Calidad")
def crear_editar_area(request, pk=None):
    """Vista combinada para Crear (pk=None) o Editar (pk!=None)"""
    if pk:
        area = get_object_or_404(Area, pk=pk)
        titulo = "Editar Área"
    else:
        area = None
        titulo = "Registrar Nueva Área"

    if request.method == 'POST':
        form = AreaForm(request.POST, instance=area)
        if form.is_valid():
            form.save()
            messages.success(request, f'Área "{form.instance.nombre}" guardada correctamente.')
            return redirect('recursoshumanos:lista_areas')
    else:
        form = AreaForm(instance=area)

    return render(request, 'recursoshumanos/areas/crear_area.html', {
        'form': form,
        'titulo': titulo,
        'current_view': 'gestion_areas'
    })

@login_required
@group_required("Recursos Humanos", "Gerencia")
def eliminar_area(request, pk):
    area = get_object_or_404(Area, pk=pk)
    nombre = area.nombre
    area.delete()
    messages.success(request, f'Área "{nombre}" eliminada.')
    return redirect('recursoshumanos:lista_areas')

@login_required
def consulta_asistencias_view(request):
    empresas     = Empresa.objects.all().order_by('nombre')
    proyectos    = Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre')
    subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent').order_by('nombre')
    areas        = Area.objects.all().order_by('nombre')
    trabajadores = Trabajador.objects.filter(activo=True).prefetch_related('proyectos').order_by('apellido_paterno')

    empresa_id     = request.GET.get('empresa', '').strip()
    proyecto_id    = request.GET.get('proyecto', '').strip()
    subproyecto_id = request.GET.get('subproyecto', '').strip()
    area_id        = request.GET.get('area', '').strip()
    trabajador_id  = request.GET.get('trabajador', '').strip()
    fecha_inicio   = request.GET.get('inicio', '').strip()
    fecha_fin      = request.GET.get('fin', '').strip()
    etiqueta       = request.GET.get('etiqueta', '').strip()
    page_number    = request.GET.get('page', 1)

    busqueda_realizada = any([empresa_id, proyecto_id, subproyecto_id, area_id, trabajador_id, fecha_inicio, fecha_fin, etiqueta])
    hoy = timezone.localdate()

    page_obj = None
    if busqueda_realizada:
        qs = (TareoDiario.objects
              .select_related('trabajador', 'trabajador__area', 'trabajador__empresa')
              .filter(fecha__lte=hoy)  # Nunca mostrar dias futuros (programados): apenas
              # llega la fecha, el dia aparece solo con su estado real.
              .order_by('-fecha', 'trabajador__apellido_paterno'))

        if empresa_id:
            qs = qs.filter(trabajador__empresa_id=empresa_id)
        if proyecto_id:
            qs = qs.filter(trabajador__proyectos__id=proyecto_id).distinct()
        if subproyecto_id:
            qs = qs.filter(trabajador__proyectos__id=subproyecto_id).distinct()
        if area_id:
            qs = qs.filter(trabajador__area_id=area_id)
        if trabajador_id:
            qs = qs.filter(trabajador_id=trabajador_id)
        if fecha_inicio:
            qs = qs.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha__lte=fecha_fin)
        if etiqueta:
            qs = qs.filter(etiqueta_estado=etiqueta)

        paginator = Paginator(qs, 20)
        page_obj  = paginator.get_page(page_number)

    context = {
        'empresas': empresas,
        'proyectos': proyectos,
        'subproyectos': subproyectos,
        'areas': areas,
        'trabajadores': trabajadores,
        'page_obj': page_obj,
        'busqueda_realizada': busqueda_realizada,
        'hoy': hoy,
        'current_empresa': empresa_id,
        'current_proyecto': int(proyecto_id) if proyecto_id else '',
        'current_subproyecto': int(subproyecto_id) if subproyecto_id else '',
        'current_area': area_id,
        'current_trabajador': int(trabajador_id) if trabajador_id else '',
        'current_inicio': fecha_inicio,
        'current_fin': fecha_fin,
        'current_etiqueta': etiqueta,
        'opciones_etiqueta': EstadoMarca.choices,
    }
    return render(request, 'recursoshumanos/consulta_asistencias/lista_consulta.html', context)