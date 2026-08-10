"""Exportaciones a Excel.

Extraido de cenerisapp/views.py durante la modularizacion.
Los cuerpos de las funciones son identicos al original.
"""

import calendar
import os

from collections import defaultdict
from datetime import date, datetime, time, timedelta
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from cenerisapp.models import (
    Alarma,
    Dispositivo,
    Empleado,
    FotoDispositivo,
    Modificacion,
    Programa,
    Registro,
    Reporte,
    SeguimientoDiario,
    Sensor,
)


@login_required
def exportar_indice(request):
    
    años_disponibles = Programa.objects.values_list('ano', flat=True).distinct().order_by('-ano')
    tipos_disponibles = Programa.objects.values_list('tipo_dispositivo', flat=True).distinct().order_by('tipo_dispositivo')
    
    fechas_de_reportes = Reporte.objects.values_list('fecReport', flat=True).distinct()
    
    años_unicos = set()
    for fecha in fechas_de_reportes:
        if fecha: # Nos aseguramos de que la fecha no sea Nula
            años_unicos.add(fecha.year)
    
    años_reportes = sorted(list(años_unicos), reverse=True)

    if not años_disponibles:
        años_disponibles = [date.today().year]
    if not tipos_disponibles:
        tipos_disponibles = ['Fijo', 'Portatil']

    tipos_disponibles = Programa.objects.values_list('tipo_dispositivo', flat=True).distinct().order_by('tipo_dispositivo')
    if not tipos_disponibles.exists():
        tipos_disponibles = ['Fijo', 'Portatil']

    areas_generales_disponibles = Dispositivo.objects.values_list('area_general', flat=True)\
                                                                                .distinct()\
                                                                                    .order_by('area_general')

    areas_generales_portatiles = Dispositivo.objects.filter(tipoDisp='Portatil')\
                                                    .values_list('area_general', flat=True)\
                                                    .distinct().order_by('area_general')
    
    turnos_disponibles = Registro.objects.values_list('turno', flat=True).distinct().order_by('turno')

    operadores = Empleado.objects.filter(
        Q(puesto__iexact="Operador de Servicio Tecnico") | 
        Q(puesto__iexact="Supervisor Operativo")
    ).order_by('nomEmpleado')

    areas_generales_fijos = Dispositivo.objects.filter(tipoDisp='Fijo')\
                                               .values_list('area_general', flat=True)\
                                               .distinct()\
                                               .order_by('area_general')
    meses_disponibles = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    context = {
        'titulo': "Centro de Exportaciones",
        'años_disponibles': años_disponibles,
        'tipos_disponibles': tipos_disponibles,
        'años_reportes': años_reportes,
        'areas_generales_disponibles': [area for area in areas_generales_disponibles if area],
        'turnos_disponibles': [t for t in turnos_disponibles if t], # Filtramos valores vacíos
        'operadores': operadores,
        'areas_generales_fijos': [area for area in areas_generales_fijos if area],
        'areas_generales_portatiles': [area for area in areas_generales_portatiles if area],
        'meses_disponibles': meses_disponibles,
    }
    return render(request, 'exportar/indice.html', context)


@login_required
def exportar_portatiles_excel(request):
    
    
    dispositivos_portatiles = Dispositivo.objects.filter(tipoDisp='Portatil').prefetch_related(
        'sensor_set', 'modificacion_set'
    ).order_by('id_dispositivo')

    
    sensores_unicos_qs = Sensor.objects.filter(dispositivo_instalado__in=dispositivos_portatiles)\
                                      .values('tipGas', 'nomComp')\
                                      .distinct()
    tipos_de_sensor_unicos = sorted([s['tipGas'] for s in sensores_unicos_qs if s['tipGas']])
    
    SENSOR_HEADER_CONFIG = {
        sensor_data.get('tipGas'): f"SENSOR {sensor_data.get('tipGas')}\n{sensor_data.get('nomComp', '')}"
        for sensor_data in sensores_unicos_qs
    }

    CAMPOS_POR_SENSOR_DETALLES = [
        'Nº DE SERIE ANTERIOR', 'Nº DE SERIE ACTUAL', 'FECHA DE FABRICACIÓN', 'FECHA DE INSTALACIÓN',
        'RESPONSABLE DEL CAMBIO', 'NÚMERO DE GUÍA DE INGRESO', 'ITEM DE GÚIA',
        'VENCIMIENTO DE GARANTIA', 'ESTATUS DE GARANTIA', 'ESTATUS DEL SENSOR'
    ]

    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Base de Datos Portátiles"
    hoy = date.today()

    
    super_header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=8, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    
    
    headers_fijos_inicio = [
        'N°', 'MODELO', 'NÚMERO SERIE', 'Fecha de Fabricación', 'Fecha de ingreso',
        'Fecha de vencimiento de garantía', 'NS', 'CÓDIGO DEL EQUIPO', 'Última fecha de Mantto',
        'Responsable de Mantto', 'SENSOR', 'UBICACIÓN DEL EQUIPO', 'OBSERVACION',
        'ESTATUS GARANTÍA DEL EQUIPO'
    ]
    headers_dinamicos_estatus = tipos_de_sensor_unicos # Ej: ['DUAL', 'LEL', 'O2']
    
    headers_fijos_finales = [
        'Número de Guía', 'OBSERVACION POR FALTA DE MANTENIMIENTO INDICADO POR MSA',
        'FECHA QUE PASA A IRREPARABLE', 'FECHA QUE PASA A INOPEATIVO', 'ESTADO DEL EQUIPO',
        'PROPIEDAD DEL EQUIPO', 'FECHA DE ULTIMA MODIFICACIÓN'
    ]
    
    headers_dinamicos_detalles = []
    for tipo in tipos_de_sensor_unicos:
        headers_dinamicos_detalles.extend([f'{campo} {tipo}' for campo in CAMPOS_POR_SENSOR_DETALLES])

    headers_fijos_finales_finales = [ 'PLACA ELECTRÓNICA - FECHA FABRICACIÓN', 'SENSOR CANIBALIZADO',
        'PCBA', 'CARCASA', 'CLIP', 'CARDEX']
        
    headers_completos = headers_fijos_inicio + headers_dinamicos_estatus + headers_fijos_finales + headers_dinamicos_detalles + headers_fijos_finales_finales
    
    
    
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers_fijos_inicio))
    sheet['B1'].value = 'INFORMACIÓN DEL EQUIPO'
    
    start_col = len(headers_fijos_inicio) + 1

    if tipos_de_sensor_unicos:    

        end_col = start_col + len(headers_dinamicos_estatus) -1
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        sheet['O1'].value = 'ESTATUS GARANTÍA DE SENSORES' # Título para el grupo de estatus
        
        start_col_detalles = len(headers_fijos_inicio) + len(headers_dinamicos_estatus) + len(headers_fijos_finales) + 1
        for tipo in tipos_de_sensor_unicos:
            end_col_detalles = start_col_detalles + len(CAMPOS_POR_SENSOR_DETALLES) - 1
            sheet.merge_cells(start_row=1, start_column=start_col_detalles, end_row=1, end_column=end_col_detalles)
            sheet.cell(row=1, column=start_col_detalles).value = SENSOR_HEADER_CONFIG.get(tipo)
            start_col_detalles = end_col_detalles + 1   
        
    
    sheet.append(headers_completos)
    
    for cell in sheet[2]:  # Fila 2 es la de los encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, dispositivo in enumerate(dispositivos_portatiles, 1):
        
        ultima_modificacion = dispositivo.modificacion_set.order_by('-fecInstalacionMod').first()
        sensores_del_dispositivo = {s.tipGas: s for s in dispositivo.sensor_set.all()}
        ultimo_mantenimiento = dispositivo.mantenimientos.order_by('-fecha_intervencion').first()
        
        ultima_fecha_mant_str = ''
        if ultimo_mantenimiento and ultimo_mantenimiento.fecha_intervencion:
            # 1. Convertir de UTC a tu zona horaria local
            fecha_local = timezone.localtime(ultimo_mantenimiento.fecha_intervencion)
            # 2. Formatear solo la parte de la fecha
            ultima_fecha_mant_str = fecha_local.strftime('%d/%m/%Y')


        estatus_garantia_equipo = 'VIGENTE' if dispositivo.fecVencimientoGarantia and dispositivo.fecVencimientoGarantia >= hoy else 'CADUCADO'
        ns_codigo = ''.join(filter(str.isdigit, dispositivo.tag or ''))

        observaciones = dispositivo.observaciones.all().values_list('comentario', flat=True)
    
        # 2. Unimos todos los comentarios en una sola cadena de texto.
        #    Cada comentario estará en una nueva línea, precedido por un guion.
        comentarios_concatenados = "\n".join([f"- {obs}" for obs in observaciones])
        

        placa_electronica = dispositivo.partes.filter(nomPart__icontains='Placa Electrónica').first()
        fecha_fab_placa = ''
        if placa_electronica:
            fecha_fab_placa = placa_electronica.fecFab if hasattr(placa_electronica, 'fecFab') else ''

        
        sensor_canibalizado_info = ''
        sensor_c = Sensor.objects.filter(dispositivo_instalado=dispositivo, info_canibalizado__isnull=False).exclude(info_canibalizado__exact='').first()
        if sensor_c:
            sensor_canibalizado_info = sensor_c.info_canibalizado

        cambios_partes_clave = { 'PCBA': None, 'CARCASA': None, 'CLIP': None }
    
        todas_mods_partes = dispositivo.modificacion_set.filter(
            MotivoCambio__in=cambios_partes_clave.keys()
        ).order_by('MotivoCambio', '-fecInstalacionMod')

        
        
        cambios_partes_clave = { 'PCBA': None, 'CARCASA': None, 'CLIP': None }
        for mod in todas_mods_partes:
            motivo = mod.MotivoCambio.upper()  # Convertimos el motivo a mayúsculas para la búsqueda en el diccionario
            if motivo in cambios_partes_clave and cambios_partes_clave[motivo] is None:
                cambios_partes_clave[motivo] = mod
            
        
        cardex_status = "Revisado" if getattr(dispositivo, 'cardex_revisado', False) else ""

        
        row_data = [
            idx, dispositivo.nomDisp, dispositivo.num_serie, dispositivo.fecFabricacion,
            dispositivo.fecIngreso, dispositivo.fecVencimientoGarantia, ns_codigo, dispositivo.tag,
            ultima_fecha_mant_str if ultima_fecha_mant_str else '',
            ultimo_mantenimiento.tecnico_a_cargo.nomEmpleado if ultimo_mantenimiento and ultimo_mantenimiento.tecnico_a_cargo else '',
            ", ".join(sensores_del_dispositivo.keys()),
            dispositivo.area_general if dispositivo.area_general else '',
            comentarios_concatenados,
            estatus_garantia_equipo,
        ]
        
        
        for tipo in tipos_de_sensor_unicos:
            sensor = sensores_del_dispositivo.get(tipo)
            if sensor and sensor.fecVencGarantia:
                estatus = 'VIGENTE' if sensor.fecVencGarantia >= hoy else 'CADUCADO'
                row_data.append(estatus)
            else:
                row_data.append('N/A')

        
        row_data.extend([
            next((s.nro_guia_ingreso for s in sensores_del_dispositivo.values() if s.nro_guia_ingreso), ''), # Número de Guía
            ultimo_mantenimiento.observacion_msa if ultimo_mantenimiento else '',
            dispositivo.fec_irreparable,
            dispositivo.fec_inoperativo,
            dispositivo.estadoD,
            getattr(dispositivo, 'propiedad', 'CENERIS'),
            ultima_modificacion.fecInstalacionMod if ultima_modificacion else ''
        ])
        
        
        for tipo in tipos_de_sensor_unicos:
            sensor = sensores_del_dispositivo.get(tipo)
            if sensor:
                mod_sensor = Modificacion.objects.filter(Q(sensor_saliente=sensor) | Q(componente_entrante=sensor.componente_ptr)).order_by('-fecInstalacionMod').first()
                ns_anterior = mod_sensor.sensor_saliente.nSerieActual if mod_sensor and mod_sensor.sensor_saliente else ''
                responsable_cambio = mod_sensor.id_trabajador.nomEmpleado if mod_sensor and mod_sensor.id_trabajador else ''
                estatus_garantia = 'VIGENTE' if sensor.fecVencGarantia and sensor.fecVencGarantia >= hoy else 'CADUCADO'

                row_data.extend([
                    ns_anterior, sensor.nSerieActual, sensor.fecFabComp, sensor.fecInst,
                    responsable_cambio, sensor.nro_guia_ingreso, sensor.item_guia,
                    sensor.fecVencGarantia, estatus_garantia, sensor.estComp
                ])
            else:
                row_data.extend([''] * len(CAMPOS_POR_SENSOR_DETALLES))

        

        pcba_mod = cambios_partes_clave['PCBA']
        carcasa_mod = cambios_partes_clave['CARCASA']
        clip_mod = cambios_partes_clave['CLIP']

        row_data.extend([
            fecha_fab_placa,
            sensor_canibalizado_info,
            f"Se cambió por {pcba_mod.id_trabajador.nomEmpleado}" if pcba_mod and pcba_mod.id_trabajador else '',
            f"Se cambió por {carcasa_mod.id_trabajador.nomEmpleado}" if carcasa_mod and carcasa_mod.id_trabajador else '',
            f"Se cambió por {clip_mod.id_trabajador.nomEmpleado}" if clip_mod and clip_mod.id_trabajador else '',
            cardex_status
        ])
        sheet.append(row_data)

        current_row = sheet.max_row
        
        row_fill = gray_fill if current_row % 2 == 0 else white_fill
        
        
        for col_idx, cell in enumerate(sheet[current_row]):
            
            if col_idx >= 0: # Comienza en la segunda columna (índice 1)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = header_alignment # Puedes usar una alineación diferente si lo necesitas


    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Portatiles_Final_{date.today().strftime("%Y-%m-%d")}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_registros_diarios_excel(request):
    
    fecha_str = request.GET.get('fecha')
    area_general_seleccionada = request.GET.get('area_general')
    if area_general_seleccionada:
        
        area_general_seleccionada = area_general_seleccionada.strip()
    turno_seleccionado = request.GET.get('turno')
    operador_id = request.GET.get('operador_id')

    print(f"[PASO 1] Filtros recibidos desde la URL:")
    print(f"  - fecha: '{fecha_str}'")
    print(f"  - area_general: '{area_general_seleccionada}'")
    print(f"  - turno: '{turno_seleccionado}'")
    print(f"  - operador_id: '{operador_id}'")
    if not all([fecha_str, area_general_seleccionada, turno_seleccionado, operador_id]):
        print(">>> ERROR: Faltan parámetros de filtro. Abortando.")
        return HttpResponse("Faltan parámetros de filtro.", status=400)
    operador_seleccionado = get_object_or_404(Empleado, pk=operador_id)
    try:
        
        fecha_seleccionada = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido.", status=400)

    
    tz = timezone.get_current_timezone()
    start_of_day = timezone.make_aware(datetime.combine(fecha_seleccionada, time.min), tz)
    end_of_day = timezone.make_aware(datetime.combine(fecha_seleccionada, time.max), tz)

    
    
    registros_del_dia = Registro.objects.filter(
        fecRegistro__range=(start_of_day, end_of_day),
        id_dispositivo__area_general__iexact=area_general_seleccionada.strip(),
        id_dispositivo__tipoDisp='Portatil',
        turno=turno_seleccionado
    ).select_related(
        'trabajador_receptor__empresa',
        'operador_responsable',
        'id_dispositivo',
        'area_trabajo_operacion',
        'punto_exacto_operacion'
    ).prefetch_related('trabajador_receptor__telefono_set').order_by('fecRegistro')

    print(f"-> Consulta SQL (aproximada): \n{registros_del_dia.query}\n")
    print(f"\n-> Número de registros encontrados con TODOS los filtros combinados: {registros_del_dia.count()}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Registro {fecha_seleccionada.strftime('%Y-%m-%d')}"
    
    
    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses_ano = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    dia_semana_nombre = dias_semana[fecha_seleccionada.weekday()]
    mes_nombre = meses_ano[fecha_seleccionada.month]
    
    fecha_formateada = f"{dia_semana_nombre}, {fecha_seleccionada.day} de {mes_nombre} de {fecha_seleccionada.year}"
    
    sheet.merge_cells('A1:C1')
    sheet['A1'].value = area_general_seleccionada.upper()
    sheet['A1'].font = Font(bold=True, size=12)

    sheet['A3'].value = "Turno :"
    sheet['A4'].value = "Operador"
    sheet['B3'].value = turno_seleccionado
    operador_encargado = registros_del_dia.first().operador_responsable if registros_del_dia.exists() else None
    if operador_encargado:
        sheet['B4'].value = operador_encargado.nomEmpleado
    
    sheet.merge_cells('F1:I1')
    sheet['F1'].value = fecha_formateada
    sheet['F1'].font = Font(bold=True, size=12)
    sheet['F1'].alignment = Alignment(horizontal='right')

    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid') # Verde
    blue_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')  # Azul
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')# Amarillo
    orange_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')# Naranja
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    
    sheet.merge_cells('A6:B6'); sheet['A6'].value = 'ORDEN'
    sheet.merge_cells('C6:F6'); sheet['C6'].value = 'DATOS PERSONALES'
    sheet.merge_cells('G6:K6'); sheet['G6'].value = 'INFORMACIÓN DEL EQUIPO DETECTOR DE GASES'
    sheet.merge_cells('L6:N6'); sheet['L6'].value = 'INFORMACION DE LA ACTIVIDAD'
    sheet.merge_cells('O6:S6'); sheet['O6'].value = 'CONTROL DE EQUIPO'

    super_headers = ['A6', 'C6', 'G6', 'L6', 'O6']
    fills = [blue_fill, green_fill, blue_fill, yellow_fill, orange_fill]
    for i, cell_coord in enumerate(super_headers):
        cell = sheet[cell_coord]
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = fills[i]
        cell.border = thin_border

    
    headers = [
        'ITEM', 'FECHA', 'IDENTIFICACIÓN', 'NOMBRE',
        'EMPRESA', 'TELÉFONO', 'EQUIPO ENTREGADO',
        'ÁREA ASIGNADA AL EQUIPO', 'MODELO', 'UBICACIÓN DEL EQUIPO', 'EQUIPO DEVUELTO',
        'ÁREA DE TRABAJO', 'PUNTO EXACTO', 'DURACIÓN (HRS)', 'TURNO',
        'ADMIN', 'HORA ENTREGA', 'HORA DEVOLUCIÓN', 'ESTADO'
    ]
    sheet.append(headers)
    
    
    header_row_num = sheet.max_row
    for cell in sheet[header_row_num]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    
    for idx, registro in enumerate(registros_del_dia, 1):
        fecha_registro_peru = registro.fecRegistro.strftime('%d/%m/%Y')
        
        trabajador_receptor = registro.trabajador_receptor
        operador_responsable = registro.operador_responsable
        dispositivo = registro.id_dispositivo
        telefonos = ", ".join([t.numero for t in trabajador_receptor.telefono_set.all()])
        duracion_horas = (registro.durPrestamo.total_seconds() / 3600) if registro.durPrestamo else ''
        
        
        row_data = [
            idx, 
            fecha_registro_peru,    
            trabajador_receptor.dni,
            
            trabajador_receptor.nomEmpleado,
            trabajador_receptor.empresa.nombreE if trabajador_receptor.empresa else '',
            
            telefonos,
            dispositivo.tag, 
            registro.id_dispositivo.area_general,
            dispositivo.nomDisp, 
            registro.id_dispositivo.area_general,
            dispositivo.tag if registro.fecDevol else '',
            
            registro.area_trabajo_operacion.nombreA if registro.area_trabajo_operacion else '', 
            registro.punto_exacto_operacion.nombre_punto if registro.punto_exacto_operacion else '', 
            f'{duracion_horas:.2f}' if duracion_horas else '',
            registro.turno,
            operador_responsable.nomEmpleado if operador_responsable else '',
            registro.fecRegistro.time().strftime('%H:%M:%S'),
            registro.fecDevol.time().strftime('%H:%M:%S') if registro.fecDevol else '',
            "DEVUELTO" if registro.fecDevol else "PRESTADO"
        ]
        
        sheet.append(row_data)

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Registro_Diario_{fecha_seleccionada.strftime("%Y-%m-%d")}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_fijos_excel_certificado(request):
    
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dispositivos Fijos"
    
    
    super_header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    super_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    super_header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid') # Naranja oscuro
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
    sheet.merge_cells('H2:J2'); sheet['H2'].value = 'CONDICIONES AMBIENTALES'
    sheet.merge_cells('L2:Q2'); sheet['L2'].value = 'GASES DE CALIBRACIÓN UTILIZADO'
    sheet.merge_cells('R2:T2'); sheet['R2'].value = 'RESULTADOS'

    for col in ['H', 'L', 'R']:
        cell = sheet[f'{col}2']
        cell.font = super_header_font
        cell.alignment = super_header_alignment
        cell.border = thin_border
        cell.fill = super_header_fill

    headers = [
        'N°', 'MODELO DEL DETECTOR', 'ÁREA', 'UBICACIÓN EN ÁREA', 'Código SMCV:',
        'TIPO DE SENSOR', 'GASES Y RANGO DE MEDICIÓN', 'TEMPERATURA', 'PRESIÓN', 'HUMEDAD RELATIVA',
        'ESTADO', 'CILINDROS', 'GASES', 'N/P', 'Lote N°', 'N° CERTIFICADO', 'EXPIRACIÓN',
        'LECTURA PATRÓN', 'LECTURA DEL EQUIPO', '% ERROR TRAS CALIBRACIÓN',
        'FECHA DE CALIBRACIÓN POR CENERIS', 'PRÓXIMA CALIBRACIÓN'
    ]
    
    sheet.append(headers)
    
    
    for cell in sheet[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    dispositivos_fijos = Dispositivo.objects.filter(tipoDisp='Fijo').select_related(
        'id_areaTrabajo_fijo'
    ).prefetch_related(
        'certificados__componente__sensor__informes_calibracion__empresa_realizadora' ,'certificados__patronescalibracion_set', 'certificados__resultados_set'
    ).order_by('nomDisp')

    for idx, dispositivo in enumerate(dispositivos_fijos, 1):
        
        ultimo_certificado = dispositivo.certificados.order_by('-fechCertificado').first()

        if ultimo_certificado:
            # Obtenemos el sensor VINCULADO a ESE certificado
            if ultimo_certificado.componente and hasattr(ultimo_certificado.componente, 'sensor'):
                sensor = ultimo_certificado.componente.sensor
        
        patrones = ultimo_certificado.patronescalibracion_set.all() if ultimo_certificado else []
        resultados = ultimo_certificado.resultados_set.all() if ultimo_certificado else []

        estado_calibracion = 'No Calibrado' 
        if ultimo_certificado and ultimo_certificado.estadoFinal:
            estado_calibracion = 'Calibrado'

        row = [
            idx,
            dispositivo.nomDisp,
            dispositivo.area_general,
            dispositivo.id_areaTrabajo_fijo.nombreA if dispositivo.id_areaTrabajo_fijo else '',
            dispositivo.tag,
            sensor.nomComp if sensor else '',
            ultimo_certificado.rango_medicion if ultimo_certificado else '', 
            
            
            ultimo_certificado.temp if ultimo_certificado else '',
            ultimo_certificado.presion if ultimo_certificado else '',
            ultimo_certificado.humedadRelativa if ultimo_certificado else '',
            
            
            estado_calibracion,
            ", ".join([p.numPatron for p in patrones if p.numPatron]),
            ", ".join([p.patronUtil for p in patrones]),
            ", ".join([p.n_p for p in patrones]),
            ", ".join([p.n_lote for p in patrones]),
            ", ".join([p.n_certificado for p in patrones]),
            ", ".join([p.fechaExpiracion.strftime('%d/%m/%Y') for p in patrones if p.fechaExpiracion]),
            
            
            ", ".join([r.lecturaPatron for r in resultados]),
            ", ".join([r.lecturaEquipo for r in resultados]),
            ", ".join([r.prob_error for r in resultados]),
            
            
            ultimo_certificado.fechCertificado.strftime('%d/%m/%Y') if ultimo_certificado and ultimo_certificado.fechCertificado else '',
            ultimo_certificado.proxFecha.strftime('%d/%m/%Y') if ultimo_certificado and ultimo_certificado.proxFecha else '',
        ]
        
        sheet.append(row)


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Complejo_Fijos_{date.today().strftime("%Y-%m-%d")}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_fijos_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dispositivos Fijos"
    
    # --- ESTILOS (sin cambios) ---
    super_header_font = Font(name='Arial', size=10, bold=True)
    super_header_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
    super_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid') 
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- CABECERAS (ligeramente ajustadas para que coincidan con la vista HTML) ---
    # Fila de Super-Headers
    # NOTA: Los headers se escriben manualmente para controlar las celdas combinadas.
    sheet.merge_cells('A1:H1'); sheet['A1'].value = 'INFORMACIÓN DEL EQUIPO'
    sheet.merge_cells('I1:K1'); sheet['I1'].value = 'CALIBRACIÓN ENCONTRADA'
    sheet.merge_cells('L1:M1'); sheet['L1'].value = 'FECHA VENCIMIENTO SENSOR'
    sheet.merge_cells('N1:O1'); sheet['N1'].value = 'INFO INFORME'
    sheet.merge_cells('P1:R1'); sheet['P1'].value = 'ALARMAS'
    sheet.merge_cells('S1:U1'); sheet['S1'].value = 'VALOR SPAM'
    sheet.merge_cells('V1:Y1'); sheet['V1'].value = 'ESTADO FINAL'

    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = super_header_font
            cell.fill = super_header_fill
            cell.alignment = super_header_alignment
            cell.border = thin_border
    sheet.row_dimensions[1].height = 30

    # Fila de Headers Principales
    headers = [
        'N°', 'NOMBRE DEL DISPOSITIVO','MODELO DE SENSOR', 'ÁREA', 'UBICACIÓN EN ÁREA', 'TAG', 'UBICACIÓN SENSOR',
        'TIPO DE GAS', 'INFORME', 'ENCONTRADO', 'SENSOR CAMBIADO',
        'MES', 'AÑO', 'FECHA', 'REALIZADA POR', '1RA', '2DA', '3RA', 'EQUIPO',
        'CILINDRO', 'UND', 'OBSERVACION', 'ESTADO CALIBRACIÓN', 'NRO CERTIFICADO',
        'FECHA CALIBRACIÓN CENERIS',
    ]
    sheet.append(headers)
    
    header_row_num = 2
    for cell in sheet[header_row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- LÓGICA DE DATOS (NUEVA ESTRUCTURA) ---
    dispositivos_fijos = Dispositivo.objects.filter(tipoDisp='Fijo').select_related(
        'id_areaTrabajo_fijo'
    ).prefetch_related(
        'sensor_set__certificados_de_componente',
        'sensor_set__informes_calibracion__empresa_realizadora',
        'sensor_set__alarmas',
    ).order_by('nomDisp')
    
    current_row = header_row_num + 1
    
    for idx, dispositivo in enumerate(dispositivos_fijos, 1):
        sensores = list(dispositivo.sensor_set.all())
        num_sensores = len(sensores) if sensores else 1
        start_row_for_merge = current_row
        
        # Iteramos sobre los sensores para escribir las filas de datos
        if not sensores:
            sensores = [None] # Creamos una lista con un 'None' para que el bucle se ejecute una vez
        
        for sensor in sensores:

            alarma = None
            # Datos específicos del sensor
            if sensor:
                ultimo_certificado = sensor.certificados_de_componente.order_by('-fechCertificado').first()
                ultimo_informe = sensor.informes_calibracion.order_by('-fecha_informe').first()
                try:
                    alarma = Alarma.objects.get(sensor=sensor)
                except Alarma.DoesNotExist:
                    pass # Es normal, alarma se queda como None
                except Alarma.MultipleObjectsReturned:
                    # Si, a pesar de la limpieza, hay duplicados,
                    # tomamos el más reciente y continuamos.
                    print(f"ADVERTENCIA: Se encontraron múltiples alarmas para el sensor {sensor}. Se usará la más reciente.")
                    alarma = Alarma.objects.filter(sensor=sensor).order_by('-id_alarma').first()

                estado_calibracion = 'No Calibrado'
                if ultimo_certificado and timezone.now() - ultimo_certificado.fechCertificado <= timedelta(days=183):
                    estado_calibracion = 'Calibrado'
            else:
                ultimo_certificado, ultimo_informe, estado_calibracion = None, None, 'N/A'
            
            # Escribimos los datos de esta fila
            row_data = [
                idx, dispositivo.nomDisp,sensor.nomComp if sensor else 'Sin sensor', dispositivo.area_general,
                dispositivo.id_areaTrabajo_fijo.nombreA if dispositivo.id_areaTrabajo_fijo else '',
                dispositivo.tag, '',
                # Datos específicos del sensor
                
                sensor.tipGas if sensor else '',
                ultimo_informe.informe if ultimo_informe else '',
                ultimo_informe.encontrado_calibracion if ultimo_informe else '',
                "Sí" if ultimo_informe and ultimo_informe.sensor_cambiado else "No",
                sensor.fecVencGarantia.strftime('%B').capitalize() if sensor and sensor.fecVencGarantia else '',
                sensor.fecVencGarantia.year if sensor and sensor.fecVencGarantia else '',
                ultimo_informe.fecha_informe.strftime('%d/%m/%Y') if ultimo_informe and ultimo_informe.fecha_informe else '',
                ultimo_informe.empresa_realizadora.nombreE if ultimo_informe and ultimo_informe.empresa_realizadora else '',
                alarma.primera if alarma else '', alarma.segunda if alarma else '', alarma.tercera if alarma else '',
                alarma.equipo if alarma else '', alarma.cilindro if alarma else '', alarma.und if alarma else '',
                ultimo_informe.observacion if ultimo_informe else '',
                estado_calibracion,
                ultimo_certificado.nro_certificado if ultimo_certificado else '',
                ultimo_certificado.fechCertificado.date().strftime('%d/%m/%Y') if ultimo_certificado else '',
            ]
            sheet.append(row_data)
            current_row += 1

        # --- LA MAGIA: COMBINAR CELDAS VERTICALMENTE ---
        # Si hay más de una fila para este dispositivo, combinamos las celdas comunes.
        if num_sensores > 1:
            end_row_for_merge = start_row_for_merge + num_sensores - 1
            # Columnas comunes (índices de columna, 1-based)
            columnas_a_combinar_inicio = [1, 2]
            for col_idx in columnas_a_combinar_inicio:
                sheet.merge_cells(start_row=start_row_for_merge, start_column=col_idx,
                                  end_row=end_row_for_merge, end_column=col_idx)

            # Columnas comunes DESPUÉS de la última columna de sensor
            # Área (4) hasta UND (21), y luego Observación (22)
            # Adaptamos los rangos a la nueva estructura.
            # Los datos específicos de sensor ahora son: Modelo(3), Tipo Gas(8), Informe(9), ..., Fecha Cal Ceneris(25)
            # El resto es común.
            columnas_a_combinar_final = list(range(4, 8)) + list(range(16, 22))
            for col_idx in columnas_a_combinar_final:
                 sheet.merge_cells(start_row=start_row_for_merge, start_column=col_idx,
                                  end_row=end_row_for_merge, end_column=col_idx)

            # Centramos verticalmente todas las celdas combinadas
            for col_idx in (columnas_a_combinar_inicio + columnas_a_combinar_final):
                cell = sheet.cell(row=start_row_for_merge, column=col_idx)
                cell.alignment = Alignment(vertical='center')
    # --- RESPUESTA HTTP (sin cambios) ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Fijos_Completo_{date.today().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_programas_excel(request):
    ano_seleccionado = request.GET.get('ano', date.today().year) 
    tipo_dispositivo_seleccionado = request.GET.get('tipo', 'Fijo')
    
    programas = Programa.objects.filter(
        ano=ano_seleccionado,
        tipo_dispositivo=tipo_dispositivo_seleccionado
    ).order_by('mes')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Programa {ano_seleccionado}"

    
    bold_font_11 = Font(name='Arial', size=11, bold=True)
    bold_font_10 = Font(name='Arial', size=10, bold=True)
    normal_font_10 = Font(name='Arial', size=10)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid') # Naranja claro
    green_fill = PatternFill(start_color='E2F0D5', end_color='E2F0D5', fill_type='solid') # Verde claro

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))



    column_widths = {'A': 20, 'B': 25, 'C': 25, 'D': 10, 'E': 8, 'F': 8, 'G': 18, 'H': 18, 'I': 40}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    
    sheet.merge_cells('A2:H2')
    cell_a2 = sheet['A2']
    cell_a2.value = f'PROGRAMA DE CALIBRACIÓN DE DETECTORES {tipo_dispositivo_seleccionado.upper()} {ano_seleccionado}'
    cell_a2.font = bold_font_11
    cell_a2.alignment = center_align

    
    
    meta_data = {
        'OBJETIVO': 'Calibrar los detectores de fijos del  Proyecto de Gases de Cerro Verde',
        'Presupuesto': 'Proyecto de Gases de Cerro Verde',
        'Mecanismos Legales': 'Ley N° 29783 Ley de SST, D.S. N°005-2012-TR Reglamento de la Ley de SST, 050-2013-TR Registros obligatorios del SGSST',
        'Recursos': 'HHT, Laptop, Internet, Luz, Papel, Impresora , Controlador, Gases, Detectores',
    }
    
    for i, (key, value) in enumerate(meta_data.items(), 5):
        sheet[f'A{i}'].value = key
        sheet[f'A{i}'].font = bold_font_10
        sheet[f'A{i}'].border = thin_border
        sheet.merge_cells(f'B{i}:I{i}')
        sheet[f'B{i}'].value = value
        sheet[f'B{i}'].border = thin_border

    
    sheet.merge_cells('A10:A11'); sheet['A10'].value = 'CERTIFICACIÓN'
    sheet.merge_cells('B10:C11'); sheet['B10'].value = 'RESPONSABLE'
    sheet.merge_cells('D10:D11'); sheet['D10'].value = 'META'
    sheet.merge_cells('E10:F10'); sheet['E10'].value = 'AVANCE'
    sheet.merge_cells('G10:G11'); sheet['G10'].value = 'TOTAL PROGRAMADO'
    sheet.merge_cells('H10:H11'); sheet['H10'].value = 'TOTAL EJECUTADO'
    sheet.merge_cells('I10:I11'); sheet['I10'].value = 'COMENTARIOS'
    
    sheet['E11'].value = 'P'; sheet['F11'].value = 'E'

    
    for row in sheet.iter_rows(min_row=10, max_row=11, min_col=1, max_col=9):
        for cell in row:
            cell.font = bold_font_10
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
    
    
    current_row = 12 # Empezamos a escribir los datos de los meses en la fila 12
    total_programado_anual = 0
    total_ejecutado_anual = 0

    for programa in programas:
        
        sheet.cell(row=current_row, column=1).value = calendar.month_name[programa.mes].upper()
        sheet.cell(row=current_row, column=2).value = "SUPERVISOR" # Placeholder
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)

        sheet.cell(row=current_row, column=4).value = "100%" # META
        sheet.cell(row=current_row, column=5).value = "100%" # AVANCE P
        
        porcentaje_ejecutado = programa.porcentaje_progreso
        cell_avance_e = sheet.cell(row=current_row, column=6)
        cell_avance_e.value = f'{porcentaje_ejecutado}%' # AVANCE E
        cell_avance_e.fill = green_fill # Pintamos la celda de verde
        
        sheet.cell(row=current_row, column=7).value = programa.totalPrograma
        sheet.cell(row=current_row, column=8).value = programa.totalEjecutado
        sheet.cell(row=current_row, column=9).value = programa.comentarios

        
        total_programado_anual += programa.totalPrograma
        total_ejecutado_anual += programa.totalEjecutado

        
        for col_num in range(1, 10):
            sheet.cell(row=current_row, column=col_num).border = thin_border
            sheet.cell(row=current_row, column=col_num).alignment = center_align if col_num != 9 else left_align
        
        current_row += 1

    
    current_row += 1
    sheet.merge_cells(f'A{current_row}:F{current_row}')
    sheet[f'A{current_row}'].value = 'CUMPLIMIENTO ACTUAL DEL PROGRAMA:'
    sheet[f'A{current_row}'].font = bold_font_10
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right')

    sheet[f'G{current_row}'].value = total_programado_anual
    sheet[f'H{current_row}'].value = total_ejecutado_anual
    
    promedio_cumplimiento = 0
    if total_programado_anual > 0:
        promedio_cumplimiento = (total_ejecutado_anual / total_programado_anual) * 100
        
    sheet[f'I{current_row}'].value = f'{promedio_cumplimiento:.0f}%'
    sheet[f'I{current_row}'].font = bold_font_10

    
    for col_num in range(1, 10):
        sheet.cell(row=current_row, column=col_num).border = thin_border
    

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Programa_Calibracion_{ano_seleccionado}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_mantenimiento_indice(request):
    
    modelos_disponibles = Dispositivo.objects.values_list('nomDisp', flat=True).distinct().order_by('nomDisp')
    
    context = {
        'modelos_disponibles': modelos_disponibles,
        'titulo': "Exportar Reporte de Mantenimiento"
    }
    return render(request, 'exportar/mantenimiento_indice.html', context)


@login_required
def exportar_mantenimiento_excel(request):
    modelo_seleccionado = request.GET.get('modelo', None)
    
    if not modelo_seleccionado:
        return HttpResponse("Error: Debe seleccionar un modelo de dispositivo.", status=400)
    
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Mantenimiento {modelo_seleccionado}"
    
    
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    super_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
    dispositivo_ejemplo = Dispositivo.objects.filter(nomDisp=modelo_seleccionado).prefetch_related('partes', 'sensor_set').first()
    if not dispositivo_ejemplo:
        return HttpResponse(f"No se encontraron dispositivos del modelo '{modelo_seleccionado}'.", status=404)

    columnas_fijas = [
        'N°', 'MODELO', 'NÚMERO SERIE', 'FECHA DE FABRICACIÓN', 'FECHA DE INGRESO',
        'VENCIMIENTO DE GARANTÍA', 'CÓDIGO', 'SENSOR', 'UBICACIÓN',
        'ESTADO INICIAL DEL EQUIPO', 'FECHA DE INTERVENCIÓN', 'TÉCNICO A CARGO'
    ]
    
    partes_del_modelo = sorted([p.nomPart for p in dispositivo_ejemplo.partes.all()])
    sensores_del_modelo = sorted([f"Sensor {s.tipGas}" for s in dispositivo_ejemplo.sensor_set.all()])
    columnas_checklist = partes_del_modelo + sensores_del_modelo

    
    columnas_fotos_evidencia = ['Fotos de evidencia1', 'Fotos de evidencia2', 'Fotos de evidencia3']

    columnas_finales = [
        'Componentes en mal estado', 'Componentes en estado regular',
        'Cambios Realizados', 'ESTADO DEL EQUIPO'
    ]
    
    headers_completos = columnas_fijas + columnas_checklist + columnas_fotos_evidencia + columnas_finales

    
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_completos))
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"MANTENIMIENTO DE DETECTORES PORTÁTILES ({modelo_seleccionado.upper()})"
    title_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    title_cell.fill = header_fill
    title_cell.alignment = super_header_alignment
    title_cell.border = thin_border
    
    sheet.append(headers_completos)
    header_row_num = sheet.max_row
    for col_idx, cell in enumerate(sheet[header_row_num], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        
        
        if get_column_letter(col_idx) in ['S', 'T', 'U']: # Ajusta estas letras según la posición de las fotos.
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15
        else:
            sheet.column_dimensions[get_column_letter(col_idx)].width = 12

    
    dispositivos_del_modelo = Dispositivo.objects.filter(nomDisp=modelo_seleccionado).prefetch_related(
        'mantenimientos__tecnico_a_cargo',
        'partes',
        'sensor_set'
    ).order_by('num_serie')

    for idx, dispositivo in enumerate(dispositivos_del_modelo, 1):
        ultimo_mantenimiento = dispositivo.mantenimientos.first()
        sensores_str = ", ".join([s.tipGas for s in dispositivo.sensor_set.all() if s.tipGas])
        
        fotos_evidencia = []
        if ultimo_mantenimiento:
            fotos_evidencia = dispositivo.fotos.filter(
                contexto='MANTENIMIENTO'
            ).order_by('fecha_carga')[:3]

        row_data = [
            idx,
            dispositivo.nomDisp,
            dispositivo.num_serie,
            dispositivo.fecFabricacion,
            dispositivo.fecIngreso,
            dispositivo.fecVencimientoGarantia,
            dispositivo.tag,
            sensores_str,
            dispositivo.area_general if dispositivo.area_general else '',
            ultimo_mantenimiento.estado_inicial_equipo if ultimo_mantenimiento else 'Operativo',
            ultimo_mantenimiento.fecha_intervencion.strftime('%d/%m/%Y') if ultimo_mantenimiento else '',
            ultimo_mantenimiento.tecnico_a_cargo.nomEmpleado if ultimo_mantenimiento and ultimo_mantenimiento.tecnico_a_cargo else ''
        ]
        
        checklist_data = ultimo_mantenimiento.checklist_partes if ultimo_mantenimiento else {}
        for item_columna in columnas_checklist:
            datos_parte = checklist_data.get(item_columna, {})
            estado_parte = datos_parte.get('estado', 'N/A')
            comentario_parte = datos_parte.get('comentario', '')
            celda_texto = estado_parte
            if comentario_parte:
                celda_texto += f":\n{comentario_parte}"
            row_data.append(celda_texto)

        
        row_data.extend([''] * len(columnas_fotos_evidencia))
        
        
        if ultimo_mantenimiento:
            row_data.extend([
                ultimo_mantenimiento.componentes_mal_estado,
                ultimo_mantenimiento.componentes_estado_regular,
                ultimo_mantenimiento.cambios_realizados,
                ultimo_mantenimiento.estado_final_equipo
            ])
        else:
            row_data.extend([''] * len(columnas_finales))

        sheet.append(row_data) 
        current_row = sheet.max_row
        
        
        
        start_col_photos_idx = len(columnas_fijas) + len(columnas_checklist) + 1
        
        
        sheet.row_dimensions[current_row].height = 85 

        
        for i, foto in enumerate(fotos_evidencia):
            
            col_idx_photos = start_col_photos_idx + i
            col_letter_photos = get_column_letter(col_idx_photos)

            if foto.imagen_original and os.path.exists(foto.imagen_original.path):
                try:
                    img = OpenpyxlImage(foto.imagen_original.path)
                    
                    
                    img.height = 60 # Altura de la imagen en píxeles
                    img.width = 90 # Ancho de la imagen en píxeles
                    
                    sheet.add_image(img, f"{col_letter_photos}{current_row}")
                except Exception as e:
                    print(f"Error al insertar imagen: {e}")
                    sheet.cell(row=current_row, column=col_idx_photos).value = "Error img"
            else:
                 sheet.cell(row=current_row, column=col_idx_photos).value = "No img"

    
    for row in sheet.iter_rows(min_row=3, min_col=len(columnas_fijas) + 1, max_col=len(columnas_fijas) + len(columnas_checklist)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
    
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Mantenimiento_{modelo_seleccionado}.xlsx"'
    workbook.save(response)
    return response


@login_required
def exportar_fijos_por_area(request, area_general):
    
    
    dispositivos = Dispositivo.objects.filter(
        tipoDisp='Fijo', 
        #aqui va area general
    ).order_by('tag').prefetch_related('sensor_set', 'fotos')

    if not dispositivos.exists():
        return HttpResponse(f"No se encontraron dispositivos fijos en el área '{area_general}'.")

    tipos_sensor_unicos = set()
    for dispositivo in dispositivos:
        for sensor in dispositivo.sensor_set.all():
            if sensor.tipGas:
                tipos_sensor_unicos.add(sensor.tipGas)
    
    filas_reporte = sorted(list(tipos_sensor_unicos)) + ['EVIDENCIA']
    
    print(f"Filas a generar en el reporte: {filas_reporte}")

    fotos_dict = {}
    fotos_relevantes = FotoDispositivo.objects.filter(
        dispositivo__in=dispositivos,
        tipo_foto__in=filas_reporte
    )
    for foto in fotos_relevantes:
        fotos_dict[(foto.dispositivo_id, foto.tipo_foto)] = foto.imagen_original.path

    workbook = Workbook()
    sheet = workbook.active
    
    
    title_font = Font(name='Calibri', size=16, bold=True)
    title_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')

    header_font = Font(name='Calibri', size=11, bold=True)
    detector_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    sensor_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    ubicacion_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    header_col_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    content_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dispositivos) + 2)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"PLANTA CONCENTRADORA {area_general.upper()}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = thin_border
    
    
    sheet.cell(row=2, column=1).value = "MODELO DETECTOR"
    sheet.cell(row=3, column=1).value = "MODELO SENSOR"
    sheet.cell(row=4, column=1).value = "UBICACIÓN"

    for row_idx in range(2, 5):
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = header_font
        cell.fill = detector_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    
    for col_idx, dispositivo in enumerate(dispositivos, 2):
        
        cell_r2 = sheet.cell(row=2, column=col_idx)
        cell_r2.value = dispositivo.nomDisp
        cell_r2.font = header_font
        cell_r2.fill = header_col_fill
        cell_r2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_r2.border = thin_border

        
        modelos_sensor = ", ".join([s.nomComp for s in dispositivo.sensor_set.all()])
        cell_r3 = sheet.cell(row=3, column=col_idx)
        cell_r3.value = modelos_sensor
        cell_r3.font = header_font
        cell_r3.fill = header_col_fill
        cell_r3.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_r3.border = thin_border

        
        cell_r4 = sheet.cell(row=4, column=col_idx)
        cell_r4.value = dispositivo.tag
        cell_r4.font = header_font
        cell_r4.fill = header_col_fill
        cell_r4.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_r4.border = thin_border
        
        sheet.column_dimensions[get_column_letter(col_idx)].width = 25

    
    
    
    
    start_row = 5
    for row_idx, tipo_fila in enumerate(filas_reporte, start_row):
        
        cell_header_dinamica = sheet.cell(row=row_idx, column=1)
        cell_header_dinamica.value = tipo_fila
        cell_header_dinamica.font = header_font
        cell_header_dinamica.fill = content_header_fill
        cell_header_dinamica.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_header_dinamica.border = thin_border

        sheet.row_dimensions[row_idx].height = 120
        
        for col_idx, dispositivo in enumerate(dispositivos, 2):
            imagen_path = fotos_dict.get((dispositivo.pk, tipo_fila))
            
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border # Borde para las celdas de las imágenes

            if imagen_path and os.path.exists(imagen_path):
                try:
                    img = OpenpyxlImage(imagen_path)
                    img.height = 150
                    img.width = 150
                    cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                    sheet.add_image(img, cell_address)
                except Exception as e:
                    cell.value = "Error img"
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.value = "No hay evidencia"
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Visual_Fijos_{area_general}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_cardex_excel(request, dispositivo_id):
    
    # --- 1. CONFIGURACIÓN INICIAL ---
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    template_path = os.path.join(settings.BASE_DIR, 'cenerisapp', 'templates', 'excel_templates', 'plantilla_cardex.xlsx')
    workbook = load_workbook(template_path)
    
    # ---------------------------------------------------------
    # HOJA 1: SENSORES
    # ---------------------------------------------------------
    sheet_sensores = workbook.active
    sheet_sensores.title = "SENSORES"

    # --- ESTILOS ---
    header_black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid') 
    header_dark_fill = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
    header_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    white_font_bold = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    grey_header_font = Font(name='Arial', size=9, bold=True)
    data_font = Font(name='Arial', size=9)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    empty_border = Border() # Borde vacío para limpiezas

    # =================================================================================
    # --- 2. DATOS FIJOS (ENCABEZADO) ---
    # =================================================================================
    
    def llenar_datos_fijos(hoja):
        hoja['A2'] = dispositivo.id_dispositivo
        hoja['B2'] = dispositivo.nomDisp
        hoja['C2'] = dispositivo.num_serie
        hoja['D2'] = dispositivo.tag

        fecFab = dispositivo.fecFabricacion.replace(tzinfo=None) if isinstance(dispositivo.fecFabricacion, datetime) else dispositivo.fecFabricacion
        fecIng = dispositivo.fecIngreso.replace(tzinfo=None) if isinstance(dispositivo.fecIngreso, datetime) else dispositivo.fecIngreso
        fecGar = dispositivo.fecVencimientoGarantia.replace(tzinfo=None) if isinstance(dispositivo.fecVencimientoGarantia, datetime) else dispositivo.fecVencimientoGarantia
        
        hoja['A4'] = fecFab; hoja['B4'] = fecIng; hoja['C4'] = fecGar

        sensores_str = ", ".join([s.tipGas for s in dispositivo.sensor_set.all() if s.tipGas])
        hoja['A7'] = sensores_str
        hoja['B7'] = dispositivo.area_general if dispositivo.area_general else ''

        ultimo_cert = dispositivo.certificados.order_by('-fechCertificado').first()
        if ultimo_cert:
            f_cert = ultimo_cert.fechCertificado
            hoja['C7'] = f_cert.replace(tzinfo=None) if isinstance(f_cert, datetime) else f_cert
        
        hoja['D7'] = dispositivo.estadoD
        
        # Foto General
        foto_equipo = dispositivo.fotos.filter(tipo_foto='EVIDENCIA', modificacion__isnull=True).order_by('-fecha_carga').first()
        if foto_equipo and foto_equipo.imagen_original:
            try:
                image_data = BytesIO(foto_equipo.imagen_original.read())
                img = OpenpyxlImage(image_data)
                img.height = 120; img.width = 120
                hoja.add_image(img, 'E2')
            except Exception: pass

    llenar_datos_fijos(sheet_sensores)

    # =================================================================================
    # --- 3. CLASIFICACIÓN DE DATOS ---
    # =================================================================================
    
    todas_modificaciones = dispositivo.modificacion_set.select_related(
        'id_trabajador', 'sensor_saliente', 'parte_saliente', 'componente_entrante'
    ).order_by('fecInstalacionMod') # Ordenamos cronológicamente (antiguo a nuevo) para el historial

    mods_sensores = defaultdict(list)
    mods_partes = defaultdict(list) # Clave: Nombre de Parte, Valor: Lista de Mods

    for mod in todas_modificaciones:
        es_sensor = False
        gas_key = None
        parte_key = None

        # Detección: ¿Es Sensor?
        if mod.sensor_saliente:
            gas_key = mod.sensor_saliente.tipGas; es_sensor = True
        elif mod.componente_entrante and hasattr(mod.componente_entrante, 'sensor'):
            gas_key = mod.componente_entrante.sensor.tipGas; es_sensor = True
        
        # Detección: ¿Es Parte?
        if not es_sensor:
            # Determinamos el nombre de la parte para agrupar
            if mod.parte_saliente:
                parte_key = mod.parte_saliente.nomPart
            elif mod.componente_entrante:
                parte_key = mod.componente_entrante.nomComp
            
            if parte_key:
                mods_partes[parte_key].append(mod)
        else:
            if gas_key:
                mods_sensores[gas_key].append(mod)

    # =================================================================================
    # --- 4. DIBUJAR HOJA 1 (SENSORES) ---
    # =================================================================================
    
    keys_sensores = sorted(mods_sensores.keys())
    
    for i, gas in enumerate(keys_sensores):
        col_base = 1 + (i * 2) 
        col_L = get_column_letter(col_base)     
        col_D = get_column_letter(col_base + 1) 
        curr_row = 10
        
        # Invertimos el orden para mostrar el más reciente arriba en sensores (opcional, segun preferencia)
        # Pero normalmente un historial va bajando. Usaremos el orden de la query.
        
        for mod in mods_sensores[gas]:
            # Cabecera Sensor
            sheet_sensores.merge_cells(f'{col_L}{curr_row}:{col_D}{curr_row}')
            c = sheet_sensores[f'{col_L}{curr_row}']; c.value = f"SENSOR {gas}"; c.font = white_font_bold; c.fill = header_dark_fill; c.alignment = center_align; c.border = thin_border
            
            sheet_sensores.merge_cells(f'{col_L}{curr_row+1}:{col_D}{curr_row+1}')
            c = sheet_sensores[f'{col_L}{curr_row+1}']; c.value = "FOTO SENSOR"; c.font = white_font_bold; c.fill = header_dark_fill; c.alignment = center_align; c.border = thin_border
            
            # Foto
            sheet_sensores.merge_cells(f'{col_L}{curr_row+2}:{col_D}{curr_row+3}')
            sheet_sensores.row_dimensions[curr_row+2].height = 60
            sheet_sensores.row_dimensions[curr_row+3].height = 60
            
            foto_cargada = False
            foto_mod = mod.fotos.first()
            if foto_mod and foto_mod.imagen_original:
                try:
                    img = OpenpyxlImage(BytesIO(foto_mod.imagen_original.read()))
                    img.height = 110; img.width = 110
                    sheet_sensores.add_image(img, f'{col_L}{curr_row+2}')
                    foto_cargada = True
                except: pass
            
            if not foto_cargada:
                c = sheet_sensores[f'{col_L}{curr_row+2}']; c.value = "SIN REGISTRO"; c.alignment = center_align
                for r in range(curr_row+2, curr_row+4):
                    sheet_sensores[f'{col_L}{r}'].border = thin_border; sheet_sensores[f'{col_D}{r}'].border = thin_border

            # Detalles
            r_det = curr_row + 4
            sheet_sensores.cell(row=r_det+1, column=col_base, value="N° Serie:").font=grey_header_font; sheet_sensores.cell(row=r_det+1, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+1, column=col_base).border=thin_border
            ns = mod.componente_entrante.sensor.nSerieActual if (mod.componente_entrante and hasattr(mod.componente_entrante, 'sensor')) else ''
            sheet_sensores.cell(row=r_det+1, column=col_base+1, value=ns).border=thin_border
            
            sheet_sensores.cell(row=r_det+2, column=col_base, value="Motivo:").font=grey_header_font; sheet_sensores.cell(row=r_det+2, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+2, column=col_base).border=thin_border
            sheet_sensores.cell(row=r_det+2, column=col_base+1, value=mod.MotivoCambio).border=thin_border
            
            sheet_sensores.cell(row=r_det+3, column=col_base, value="Resp:").font=grey_header_font; sheet_sensores.cell(row=r_det+3, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+3, column=col_base).border=thin_border
            sheet_sensores.cell(row=r_det+3, column=col_base+1, value=mod.id_trabajador.nomEmpleado).border=thin_border
            
            sheet_sensores.cell(row=r_det+4, column=col_base, value="Fecha Inst:").font=grey_header_font; sheet_sensores.cell(row=r_det+4, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+4, column=col_base).border=thin_border
            dt = mod.fecInstalacionMod
            sheet_sensores.cell(row=r_det+4, column=col_base+1, value=dt.replace(tzinfo=None) if isinstance(dt, datetime) else dt).border=thin_border

            curr_row += 10

    # =================================================================================
    # --- 5. GENERAR HOJAS DE MANTENIMIENTO (AGRUPADAS POR PARTE) ---
    # =================================================================================
    
    # Iteramos sobre cada tipo de parte. Ej: 'Carcasa' -> [Mod1, Mod2]
    for index, (nombre_parte, lista_mods) in enumerate(mods_partes.items(), start=1):
        
        # 1. Crear Hoja para esa Parte
        nueva_hoja = workbook.copy_worksheet(sheet_sensores)
        # Nombre de hoja seguro (Excel max 31 chars)
        safe_title = f"MANT {index} {nombre_parte}"[:30]
        nueva_hoja.title = safe_title
        
        # 2. LIMPIEZA
        for merged_range in list(nueva_hoja.merged_cells.ranges):
            if merged_range.min_row >= 10:
                nueva_hoja.unmerge_cells(str(merged_range))

        empty_fill = PatternFill()
        for row in nueva_hoja.iter_rows(min_row=10, max_row=300): # Limpiamos más filas por si acaso
            for cell in row:
                cell.value = None
                cell.fill = empty_fill
                cell.border = empty_border 
        
        llenar_datos_fijos(nueva_hoja) # Restaurar foto general

        # 3. DIBUJAR HISTORIAL VERTICAL (APILADO)
        # Empezamos en la Fila 11
        current_row = 11

        for mod in lista_mods:
            
            # --- TÍTULO (Fila 11 relativa) ---
            # Ocupa A-F
            nueva_hoja.merge_cells(f'A{current_row}:F{current_row}')
            cell = nueva_hoja[f'A{current_row}']
            # Título dinámico: Nombre de la parte + (Fecha) para diferenciar
            cell.value = f"KIT DE MANTENIMIENTO: {nombre_parte} ({mod.fecInstalacionMod.strftime('%d/%m/%Y')})"
            cell.font = white_font_bold; cell.fill = header_black_fill; cell.alignment = center_align; cell.border = thin_border
            
            # --- FOTO (Fila 12 relativa) ---
            row_foto = current_row + 1
            nueva_hoja.merge_cells(f'A{row_foto}:F{row_foto}')
            nueva_hoja.row_dimensions[row_foto].height = 250 
            
            foto_mod = mod.fotos.first()
            foto_insertada = False
            if foto_mod and foto_mod.imagen_original:
                try:
                    img = OpenpyxlImage(BytesIO(foto_mod.imagen_original.read()))
                    img.height = 300; img.width = 450 
                    nueva_hoja.add_image(img, f'A{row_foto}')
                    foto_insertada = True
                except: pass
            
            if not foto_insertada:
                cell = nueva_hoja[f'A{row_foto}']
                cell.value = "SIN REGISTRO FOTOGRAFICO"
                cell.alignment = center_align
                for c_idx in range(1, 7): 
                     col_l = get_column_letter(c_idx)
                     nueva_hoja[f'{col_l}{row_foto}'].border = thin_border

            # --- DETALLES (Filas 13-16 relativas) ---
            def escribir_fila(fila_abs, etiqueta, valor):
                nueva_hoja.merge_cells(f'A{fila_abs}:C{fila_abs}')
                lbl = nueva_hoja[f'A{fila_abs}']
                lbl.value = etiqueta
                lbl.font = grey_header_font; lbl.fill = header_grey_fill; lbl.alignment = center_align; lbl.border = thin_border
                nueva_hoja[f'B{fila_abs}'].border = thin_border; nueva_hoja[f'C{fila_abs}'].border = thin_border

                nueva_hoja.merge_cells(f'D{fila_abs}:F{fila_abs}')
                val_c = nueva_hoja[f'D{fila_abs}']
                val_c.value = valor
                val_c.font = data_font; val_c.alignment = center_align; val_c.border = thin_border
                nueva_hoja[f'E{fila_abs}'].border = thin_border; nueva_hoja[f'F{fila_abs}'].border = thin_border

            # Fila 13: Motivo
            escribir_fila(current_row + 2, "Motivo de Cambio:", mod.MotivoCambio)
            
            # Fila 14: Responsable
            resp = mod.id_trabajador.nomEmpleado if mod.id_trabajador else ""
            escribir_fila(current_row + 3, "Responsable:", resp)
            
            # Fila 15: Fecha Instalación
            dt = mod.fecInstalacionMod
            f_str = dt.strftime('%d/%m/%Y') if dt else ""
            escribir_fila(current_row + 4, "Fecha de instalación:", f_str)
            
            # Fila 16: Fecha Facturación
            escribir_fila(current_row + 5, "Fecha de facturación:", "")

            # --- SALTO PARA EL SIGUIENTE CAMBIO DE LA MISMA PARTE ---
            # La tarjeta ocupó 6 filas (1 titulo + 1 foto + 4 datos).
            # Dejamos 2 filas de espacio libre antes de la siguiente.
            current_row += 8 

    # GUARDAR
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="CARDEX_{dispositivo.tag}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_reportes_excel(request):
    """
    Genera un archivo Excel con el historial de todos los reportes de daños y pérdidas.
    """
    
    ano_seleccionado = request.GET.get('ano')
    if not ano_seleccionado:
        return HttpResponse("Error: Debe seleccionar un año para exportar.", status=400)
    
    
    reportes = Reporte.objects.filter(
        fecReport__year=ano_seleccionado
    ).select_related(
        'id_dispositivo__id_empresa', 
        'id_otro_componente'
    ).order_by('-fecReport')

    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Historial de Reportes"

    
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    header_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid') # Naranja claro
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(wrap_text=True, vertical='center')
    
    
    headers = ['FECHA DE REPORTE', 'EMPRESA', 'EQUIPO', 'DESCRIPCION', 'ESTADO']
    sheet.append(headers)
    
    
    column_widths = {'A': 20, 'B': 25, 'C': 40, 'D': 50, 'E': 15}
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    
    for reporte in reportes:
        
        
        nombre_empresa = 'N/A'
        nombre_equipo = 'N/A'
        estado_final = 'N/A'

        if reporte.id_dispositivo:
            
            dispositivo = reporte.id_dispositivo
            nombre_equipo = f"{dispositivo.nomDisp} ({dispositivo.num_serie})"
            if dispositivo.id_empresa:
                nombre_empresa = dispositivo.id_empresa.nombreE
            estado_final = dispositivo.estadoD
            
        elif reporte.id_otro_componente:
            
            componente = reporte.id_otro_componente
            nombre_equipo = f"{componente.nomComp} ({componente.nSerieActual})"
            
            if componente.inventario and componente.inventario.id_trabajador:
                nombre_empresa = "Empresa del Lote" # Placeholder
            estado_final = componente.estComp

        
        
        row_data = [
            reporte.fecReport,
            nombre_empresa,
            nombre_equipo,
            reporte.razRetiro, # 'razRetiro' parece ser la descripción
            estado_final
        ]
        
        sheet.append(row_data)
        
        last_row = sheet.max_row
        
        
        for cell in sheet[last_row]:
            cell.border = thin_border
            cell.alignment = data_alignment # Aplicar el estilo de alineación de datos
            cell.font = Font(name='Calibri', size=11) # Opcional: para mantener la consistencia

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Historial_Reportes_{ano_seleccionado}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_seguimiento_excel(request):
    # 1. OBTENER FILTROS
    try:
        ano = int(request.GET.get('ano', date.today().year))
        mes = int(request.GET.get('mes', date.today().month))
        area_general = request.GET.get('area_general')
    except (ValueError, TypeError):
        return HttpResponse("Parámetros de año o mes inválidos.", status=400)
    
    if not area_general:
        return HttpResponse("Debe seleccionar un Área General.", status=400)

    # 2. PREPARAR DATOS DE FECHAS
    nombre_mes = calendar.month_name[mes].upper()
    num_dias = calendar.monthrange(ano, mes)[1]
    dias_del_mes = [date(ano, mes, dia) for dia in range(1, num_dias + 1)]

    # 3. OBTENER DATOS DE LA BASE DE DATOS
    dispositivos = Dispositivo.objects.filter(area_general=area_general, tipoDisp='Portatil').order_by('nomDisp')
    seguimientos = SeguimientoDiario.objects.filter(
        dispositivo__in=dispositivos,
        fecha__year=ano,
        fecha__month=mes
    )
    seguimiento_dict = {(s.dispositivo_id, s.fecha): s.estado_texto for s in seguimientos}

    # 4. INICIALIZAR EXCEL Y ESTILOS
    workbook = Workbook()
    sheet = workbook.active
    # Top Header Style (for "SEGUIMIENTO EQUIPOS...")
    header_font = Font(bold=True, size=14, color='000000') # Negro para mejor contraste
    header_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid') # El color solicitado
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fijos_font = Font(bold=True, size=8, color='000000')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 1. Título principal
    # Combina las primeras 3 columnas de la fila 1 para el título
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = sheet.cell(row=1, column=1, value=f"SEGUIMIENTO EQUIPOS {area_general.upper()}")
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border


    # 2. Nombre del mes
    # Combina las siguientes celdas de la fila 1 para el mes
    sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=5 + num_dias)
    month_cell = sheet.cell(row=1, column=6, value=nombre_mes)
    month_cell.font = header_font
    month_cell.fill = header_fill
    month_cell.alignment = center_alignment
    month_cell.border = thin_border

    # 3. Headers detallados (se mantienen en la fila 3)
    headers_fijos = ['N°', 'MODELO', 'SERIE', 'UBICACIÓN', 'ESTADO']
    headers_dias = [d.strftime('%d/%m/%Y') for d in dias_del_mes]
    # Agrega los headers a la tercera fila (índice 3 en Excel)
    sheet.append(headers_fijos + headers_dias)
    
    detailed_header_row_num = sheet.max_row

    # Iterate through the cells in that row and apply the styles
    for cell in sheet[detailed_header_row_num]:
        cell.font = fijos_font
        cell.fill = header_fill    
        cell.alignment = center_alignment
        cell.border = thin_border


    # 6. ESCRIBIR DATOS
    for idx, dispositivo in enumerate(dispositivos, 1):
        row_data = [
            idx,
            dispositivo.nomDisp,
            dispositivo.num_serie,      
            dispositivo.area_general if dispositivo.area_general else '',
            dispositivo.estadoD
        ]
        
        for dia in dias_del_mes:
            estado_del_dia = seguimiento_dict.get((dispositivo.pk, dia), '')
            row_data.append(estado_del_dia)
            
        sheet.append(row_data)
        current_row = sheet.max_row
    
        # Itera sobre todas las celdas de la fila actual y aplica los estilos
        for cell in sheet[current_row]:
            cell.alignment = center_alignment
            cell.border = thin_border

    # 7. FINALIZACIÓN Y RESPUESTA
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Seguimiento_{area_general}_{ano}-{mes:02d}.xlsx"'
    workbook.save(response)
    
    return response
