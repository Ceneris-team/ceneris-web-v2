from datetime import datetime, date, timedelta, time
import os
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.http import JsonResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.forms import formset_factory
from django.db.models import Count, Case, When, Value, IntegerField
from django.core.exceptions import ValidationError
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, HttpResponseForbidden
from django.template.loader import render_to_string
from urllib3 import request
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.template.loader import get_template
from collections import defaultdict
import calendar
from django.db.models import Q
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from django.core.paginator import Paginator
from django.db.models.functions import TruncYear
import locale
from django.views.decorators.http import require_POST
from django.utils.html import format_html 
from weasyprint import HTML, CSS
import json
from django.forms.models import model_to_dict
from django.urls import reverse 
from django.db.models.functions import TruncMonth
from django.http import Http404
from django.db.models import F 
from django.db.models import Prefetch 
from django.contrib.staticfiles import finders
from weasyprint.urls import default_url_fetcher
from django.forms import inlineformset_factory
from django.core.files.storage import default_storage, FileSystemStorage
import uuid
from django.core.files import File
from django.db import transaction

from cenerisapp.models import Dispositivo, Empleado, Registro, Inventario, Componente, OtroComponente, Sensor, Reporte, Calibracion, Alarma, Ventas, Modificacion, Parte, Empresa, AreaTrabajo, Certificado, PatronesCalibracion, Programa, Resultados, FotoDispositivo, Mantenimiento, PuntoExacto, SeguimientoDiario, ObservacionDispositivo, Ocurrencia, ReporteDiario, DatosPDF, InformeCalibracion, AnexoCertificado
from .forms import RegistroSalidaForm, InventarioForm, SensorForm, OtroComponenteForm, DispositivoForm, SensorParaDispositivoForm, BaseSensorParaDispositivoFormSet, AlarmaFijoForm, AlarmaPortatilForm, ReporteForm, CalibracionForm, VentaForm, ModificacionForm, EmpleadoForm, CorreoFormSet, TelefonoFormSet, ParteFormSet, AreaTrabajoForm, CertificadoForm, PatronesFormSet, ResultadosFormSet, ProgramaCreateForm, ProgramaUpdateForm, MantenimientoForm, SensorLoteForm, FotoDispositivoForm, InformeCalibracionForm, SeguimientoDiarioForm, PuntoExactoFormSet, OcurrenciaForm, EmpleadoRapidoForm, ModificacionAntiguaForm, AnexoCertificadoForm

from django.contrib.staticfiles.storage import staticfiles_storage
from django.conf import settings

def static_file_path(path):
    """
    Toma una ruta estática (ej. 'img/logo.png') y devuelve
    una URL 'file://' con la ruta absoluta del sistema de archivos.
    """

    # find() devuelve la ruta absoluta del archivo en el sistema.
    # Esto funciona tanto en DEBUG=True como en DEBUG=False (después de collectstatic)
    absolute_path = staticfiles_storage.path(path)
    
    # WeasyPrint necesita el prefijo 'file://'
    return f'file://{absolute_path}'

@login_required
def inicio(request):
    return render(request, 'inicio_supervisor/inicio.html')

def buscar_puntos_exactos_api(request):
    area_id = request.GET.get('area_id')
    if not area_id:
        return JsonResponse([], safe=False)
    
    puntos = PuntoExacto.objects.filter(area_trabajo_id=area_id).values('id', 'nombre_punto')
    return JsonResponse(list(puntos), safe=False)

@login_required
def get_observaciones_json(request, dispositivo_id):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    observaciones = dispositivo.observaciones.all().values(
        'comentario', 
        'fecha_creacion', 
        'autor__username' # O 'autor__first_name' si lo prefieres
    )
    # Convertimos el QuerySet a una lista de diccionarios para enviarlo como JSON
    return JsonResponse(list(observaciones), safe=False)

# --- VISTA 2: PARA GUARDAR UN NUEVO COMENTARIO ---
@login_required
@require_POST # Esta vista solo debe aceptar peticiones POST
def add_observacion_json(request, dispositivo_id):
    try:
        dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
        data = json.loads(request.body)
        nuevo_comentario = data.get('comentario', '').strip()

        if not nuevo_comentario:
            return JsonResponse({'status': 'error', 'message': 'El comentario no puede estar vacío.'}, status=400)

        # Creamos y guardamos la nueva observación
        obs = ObservacionDispositivo.objects.create(
            dispositivo=dispositivo,
            autor=request.user,
            comentario=nuevo_comentario
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'Observación guardada.',
            'comentario': obs.comentario,
            'autor': obs.autor.username,
            'fecha': obs.fecha_creacion.strftime('%d/%m/%Y %H:%M')
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'¡Cuenta creada para {username}! Ahora puedes iniciar sesión.')
            return redirect('cenerisapp:login') 
    else:
        form = UserCreationForm()
    return render(request, 'registration/registro.html', {'form': form})

@never_cache
@login_required 
def home(request):
    return render(request, 'index.html')
 
 
@login_required
def inventario_dispositivo(request):
    
    opciones_tipos = Dispositivo.objects.values_list('tipoDisp', flat=True).distinct().order_by('tipoDisp')
    opciones_estados = Dispositivo.objects.values_list('estadoD', flat=True).distinct().order_by('estadoD')
    opciones_areas = Dispositivo.objects.exclude(area_general__isnull=True).exclude(area_general__exact='').values_list('area_general', flat=True).distinct().order_by('area_general')
 
    # 2. Capturar los valores de los filtros desde la URL (request.GET)
    modelo_filtro = request.GET.get('modelo', '')
    serie_filtro = request.GET.get('serie', '')
    tag_filtro = request.GET.get('tag', '')
    tipo_filtro = request.GET.get('tipo', '')
    estado_filtro = request.GET.get('estado', '')
    area_filtro = request.GET.get('area', '')
 
    # 3. Construir el queryset base
    dispositivos = Dispositivo.objects.all().prefetch_related('sensor_set').order_by('nomDisp')
 
    # 4. Aplicar los filtros si existen
    if modelo_filtro:
        dispositivos = dispositivos.filter(nomDisp=modelo_filtro)
    if serie_filtro:
        dispositivos = dispositivos.filter(num_serie__icontains=serie_filtro)
    if tag_filtro:
        dispositivos = dispositivos.filter(tag__icontains=tag_filtro)
    if tipo_filtro:
        dispositivos = dispositivos.filter(tipoDisp=tipo_filtro)
    if estado_filtro:
        dispositivos = dispositivos.filter(estadoD=estado_filtro)
    if area_filtro:
        dispositivos = dispositivos.filter(area_general=area_filtro)

    paginator = Paginator(dispositivos, 15) # 15 modificaciones por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
 
    # 5. Preparar el contexto para la plantilla
    context = {
        'page_obj': page_obj,
        'titulo': 'Gestiona los Dispositivos Fijos y Portátiles',
        'opciones_tipos': opciones_tipos,
        'opciones_estados': opciones_estados,
        'opciones_areas': opciones_areas,
        'filtros_aplicados': {
            'modelo': modelo_filtro,
            'serie': serie_filtro,
            'tag': tag_filtro,
            'tipo': tipo_filtro,
            'estado': estado_filtro,
            'area': area_filtro,
        }
    }
 
    return render(request, 'dispositivos/lista_dispositivo.html', context)

def eliminar_dispositivo(request, id_dispositivo):
    dispositivo = Dispositivo.objects.get(id_dispositivo=id_dispositivo)
    dispositivo.delete()
    return redirect('cenerisapp:lista_inventario')

def editar_dispositivo(request, id_dispositivo):
    dispositivo = Dispositivo.objects.get(id_dispositivo=id_dispositivo)
    if request.method == 'POST':
        dispositivo.nomDisp = request.POST.get('nomDisp')
        dispositivo.num_serie = request.POST.get('num_serie')
        dispositivo.tag = request.POST.get('tag')
        dispositivo.tipoDisp = request.POST.get('tipoDisp')
        dispositivo.estadoD = request.POST.get('estadoD')
        dispositivo.fabDisp = request.POST.get('fabDisp')
        dispositivo.fecIngreso = request.POST.get('fecIngreso')
        dispositivo.fecVencimientoGarantia = request.POST.get('fecVencimientoGarantia')
        dispositivo.save()
        return redirect('cenerisapp:lista_inventario')
    context = {
        'dispositivo': dispositivo,
        'titulo': 'Editar Dispositivo'
    }
    return render(request, 'dispositivos/editar_dispositivo.html', context)
 
@login_required
def dashboard_portatiles(request):
    
    # --- PARTE 1: DATOS GENERALES PARA EL DASHBOARD ---
    # (Se calculan siempre, se busque o no un dispositivo)
    
    portatiles_qs = Dispositivo.objects.filter(tipoDisp='Portatil')
    
    # 1.1 Conteo total
    total_portatiles = portatiles_qs.count()
    
    # 1.2 Gráfico de estados
    conteo_por_estado = portatiles_qs.values('estadoD').annotate(total=Count('estadoD')).order_by('-total')
    
    # 1.3 Gráfico de modificaciones por mes (últimos 6 meses)
    modificaciones_por_mes = Modificacion.objects.filter(id_dispositivo__tipoDisp='Portatil') \
        .annotate(mes=TruncMonth('fecInstalacionMod')) \
        .values('mes') \
        .annotate(total=Count('id_modificacion')) \
        .order_by('mes')[:6] # Solo los últimos 6 meses
    
    # 1.4 Dispositivos que NUNCA han sido revisados en Cardex
    pendientes_cardex = portatiles_qs.filter(cardex_revisado=False).count()

    ultimos_mantenimientos = Mantenimiento.objects.filter(
        dispositivo__tipoDisp='Portatil'
    ).select_related('dispositivo', 'tecnico_a_cargo').order_by('-fecha_intervencion')[:4]

    # 1.6 Últimas 7 modificaciones realizadas a portátiles
    ultimas_modificaciones = Modificacion.objects.filter(
        id_dispositivo__tipoDisp='Portatil'
    ).select_related(
        'id_dispositivo', 
        'id_trabajador',
        'parte_saliente',
        'sensor_saliente',
        'componente_entrante'
    ).order_by('-fecInstalacionMod')[:4]

    # --- PARTE 2: BÚSQUEDA Y DATOS ESPECÍFICOS DE UN DISPOSITIVO ---
    
    dispositivo_seleccionado = None
    ultimo_mantenimiento = None
    ultima_modificacion = None
    registros_recientes = None
    certificados_recientes = None
    
    # Obtenemos el ID del dispositivo de la URL (?dispositivo_id=...)
    query_busqueda = request.GET.get('query', '')
    
    if query_busqueda:
        try:
            # prefetch_related es crucial para optimizar y evitar muchas consultas a la BD
            dispositivo_seleccionado = Dispositivo.objects.prefetch_related(
                'mantenimientos__tecnico_a_cargo', 
                'modificacion_set__parte_saliente', 
                'modificacion_set__sensor_saliente', 
                'modificacion_set__componente_entrante',
                'certificados',
                'registro_set__trabajador_receptor'
            ).get(num_serie=query_busqueda, tipoDisp='Portatil')

            # Obtenemos los últimos registros relacionados con este dispositivo
            ultimo_mantenimiento = dispositivo_seleccionado.mantenimientos.first() # Usamos el ordering del modelo
            ultima_modificacion = dispositivo_seleccionado.modificacion_set.first()
            registros_recientes = dispositivo_seleccionado.registro_set.all().order_by('-fecRegistro')[:5]
            certificados_recientes = dispositivo_seleccionado.certificados.all().order_by('-fechCertificado')[:5]
            
        except Dispositivo.DoesNotExist:
            messages.error(request, "El dispositivo portátil seleccionado no existe.")

    # --- PARTE 3: CONTEXTO PARA LA PLANTILLA ---
    
    context = {
        'titulo': 'Dashboard de Dispositivos Portátiles',
        
        # Datos Generales
        'total_portatiles': total_portatiles,
        'conteo_por_estado_json': json.dumps(list(conteo_por_estado)), # Para Chart.js
        'modificaciones_por_mes_json': json.dumps([
            {'mes': m['mes'].strftime('%b %Y'), 'total': m['total']} for m in modificaciones_por_mes
        ]), # Para Chart.js
        'pendientes_cardex': pendientes_cardex,
        'todos_los_portatiles_para_busqueda': portatiles_qs, # Para el <select> de búsqueda

        # Datos Específicos del Dispositivo
        'dispositivo_seleccionado': dispositivo_seleccionado,
        'query_busqueda': query_busqueda, # Para que el <select> recuerde la selección
        'ultimo_mantenimiento': ultimo_mantenimiento,
        'ultima_modificacion': ultima_modificacion,
        'registros_recientes': registros_recientes,
        'certificados_recientes': certificados_recientes,
        'ultimos_mantenimientos': ultimos_mantenimientos,
        'ultimas_modificaciones': ultimas_modificaciones,
    }
    
    return render(request, 'dashboard/dashboard_portatiles.html', context)

@login_required
def dashboard_garantias(request):
    
    # --- 1. Definir los rangos de fechas ---
    today = date.today()
    # "Muy cerca a vencer" es dentro de los próximos 30 días
    un_mes_desde_hoy = today + timedelta(days=30)

    # --- 2. Capturar el filtro de búsqueda (si existe) ---
    query = request.GET.get('q', '')

    # --- 3. Realizar las consultas para Dispositivos ---
    base_dispositivos_qs = Dispositivo.objects.all()
    if query:
        base_dispositivos_qs = base_dispositivos_qs.filter(num_serie__icontains=query)

    # --- 4. Realizar las consultas para Sensores ---
    base_sensores_qs = Sensor.objects.select_related('dispositivo_instalado')
    if query:
        base_sensores_qs = base_sensores_qs.filter(nSerieActual__icontains=query)
    
    # 1. VENCIDAS
    vencidas_dispositivos = base_dispositivos_qs.filter(fecVencimientoGarantia__lt=today).order_by('fecVencimientoGarantia')
    vencidas_sensores = base_sensores_qs.filter(fecVencGarantia__lt=today).exclude(fecVencGarantia__isnull=True).order_by('fecVencGarantia')
    vencidas_items = sorted(
        list(vencidas_dispositivos) + list(vencidas_sensores), 
        key=lambda x: getattr(x, 'fecVencimientoGarantia', None) or getattr(x, 'fecVencGarantia', None)
    )
    vencidas_paginator = Paginator(vencidas_items, 10) # 10 items por página
    vencidas_page_number = request.GET.get('page_vencidas')
    vencidas_page_obj = vencidas_paginator.get_page(vencidas_page_number)
    
    # 2. PRÓXIMAS A VENCER
    proximas_dispositivos = base_dispositivos_qs.filter(fecVencimientoGarantia__range=[today, un_mes_desde_hoy]).order_by('fecVencimientoGarantia')
    proximas_sensores = base_sensores_qs.filter(fecVencGarantia__range=[today, un_mes_desde_hoy]).exclude(fecVencGarantia__isnull=True).order_by('fecVencGarantia')
    proximas_items = sorted(
        list(proximas_dispositivos) + list(proximas_sensores), 
        key=lambda x: getattr(x, 'fecVencimientoGarantia', None) or getattr(x, 'fecVencGarantia', None)
    )
    proximas_paginator = Paginator(proximas_items, 10)
    proximas_page_number = request.GET.get('page_proximas')
    proximas_page_obj = proximas_paginator.get_page(proximas_page_number)

    # 3. VIGENTES
    vigentes_dispositivos = base_dispositivos_qs.filter(fecVencimientoGarantia__gt=un_mes_desde_hoy).order_by('fecVencimientoGarantia')
    vigentes_sensores = base_sensores_qs.filter(fecVencGarantia__gt=un_mes_desde_hoy).exclude(fecVencGarantia__isnull=True).order_by('fecVencGarantia')
    vigentes_items = sorted(
        list(vigentes_dispositivos) + list(vigentes_sensores), 
        key=lambda x: getattr(x, 'fecVencimientoGarantia', None) or getattr(x, 'fecVencGarantia', None)
    )
    vigentes_paginator = Paginator(vigentes_items, 10)
    vigentes_page_number = request.GET.get('page_vigentes')
    vigentes_page_obj = vigentes_paginator.get_page(vigentes_page_number)

    # --- 5. Preparar el contexto para la plantilla ---
    context = {
        'titulo': 'Dashboard de Garantías',
        'query': query,
        
        # Listas de ítems
        'vencidas_page_obj': vencidas_page_obj,
        'proximas_page_obj': proximas_page_obj,
        'vigentes_page_obj': vigentes_page_obj,
        
        # Conteos para las tarjetas principales
        'total_vencidas': vencidas_dispositivos.count() + vencidas_sensores.count(),
        'total_proximas': proximas_dispositivos.count() + proximas_sensores.count(),
        'total_vigentes': vigentes_dispositivos.count() + vigentes_sensores.count(),
        
        # Fecha límite para la categoría "próximas a vencer"
        'fecha_limite_proximas': un_mes_desde_hoy,
    }

    return render(request, 'garantias/dashboard_garantias.html', context)

@require_POST
@login_required
def upload_anexo_temporal(request):
    if 'imagen' not in request.FILES:
        return JsonResponse({'status': 'error', 'message': 'No se encontró ningún archivo.'}, status=400)
    
    imagen = request.FILES['imagen']
    
    try:
        # Guardamos el archivo en una ubicación temporal en S3
        filename = default_storage.save(f"temp_anexos/{uuid.uuid4()}_{imagen.name}", imagen)
        
        # Devolvemos el path que S3 nos dio
        return JsonResponse({'status': 'success', 'path': filename})
    
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@login_required
def agregar_anexos_certificado(request):
    certificado = None
    formset = None
    
    # --- BÚSQUEDA DEL CERTIFICADO (vía GET) ---
    query = request.GET.get('q', '').strip()
    if query:
        try:
            # Buscamos el certificado por su número exacto
            certificado = Certificado.objects.prefetch_related('anexos').get(nro_certificado__iexact=query)
        except Certificado.DoesNotExist:
            messages.error(request, f"No se encontró ningún certificado con el número '{query}'.")
        except Certificado.MultipleObjectsReturned:
            messages.error(request, f"Se encontraron múltiples certificados con el número '{query}'. Verifique los datos.")

    # --- DEFINICIÓN DEL FORMSET ---
    # Lo definimos aquí para usarlo tanto en GET como en POST
    AnexoFormSet = inlineformset_factory(
        Certificado,
        AnexoCertificado,
        form=AnexoCertificadoForm,
        extra=3, # Empezamos con 3 campos de archivo
        can_delete=True # Permitimos borrar anexos existentes
    )
    
    # --- MANEJO DEL FORMULARIO (POST) ---
    if request.method == 'POST':
        # Re-buscamos el certificado usando el ID enviado en el POST para seguridad
        certificado_id = request.POST.get('certificado_id')
        if not certificado_id:
            messages.error(request, "No se especificó un certificado para guardar los anexos.")
            return redirect('cenerisapp:agregar_anexos_certificado')
        
        certificado = get_object_or_404(Certificado, pk=certificado_id)
        
        # Inicializamos el formset con los datos POST, los archivos y la instancia del certificado
        formset = AnexoFormSet(request.POST, request.FILES, instance=certificado, prefix='anexos')

        if formset.is_valid():
            formset.save()
            messages.success(request, f"Anexos para el certificado '{certificado.nro_certificado}' guardados correctamente.")
            # Redirigimos a la misma página con el certificado ya cargado
            return redirect(f"{reverse('cenerisapp:agregar_anexos_certificado')}?q={certificado.nro_certificado}")
        else:
            messages.error(request, "Por favor, corrige los errores en los archivos.")
            # Los errores se mostrarán en la plantilla
            
    # --- VISTA GET (O SI EL POST FALLA) ---
    # Si encontramos un certificado, inicializamos el formset con esa instancia
    if certificado and not formset: # 'not formset' para no sobreescribir si el POST falló
        formset = AnexoFormSet(instance=certificado, prefix='anexos')

    context = {
        'titulo': 'Agregar Anexos a Certificado Existente',
        'query': query,
        'certificado': certificado,
        'formset': formset, # Puede ser None si no se ha buscado, o un formset instanciado
    }
    return render(request, 'certificado/agregar_anexos.html', context)

@login_required
def vista_reporte_fijos(request):
    
    # 1. PREPARAR LOS HEADERS PARA LA PLANTILLA (igual que en el Excel)
    
    # Super-headers para las celdas combinadas
    super_headers = {
        'FECHA DE VENCIMIENTO DEL SENSOR ACTUAL': 2, # Ocupará 2 columnas
        'CALIBRACIÓN ENCONTRADA': 2,                 # Ocupará 2 columnas
        'ALARMAS': 3,                                # Ocupará 3 columnas
        'VALOR SPAM': 3                              # Ocupará 3 columnas
    }

    # Headers principales de cada columna
    headers = [
        'N°', 'NOMBRE DEL DISPOSITIVO', 'MODELO DE SENSOR', 'ÁREA', 'UBICACIÓN EN ÁREA', 'TAG',
        'UBICACIÓN DEL SENSOR EN EL DETECTOR', 'TIPO DE GAS', 
        # Columnas de CALIBRACIÓN ENCONTRADA
        'INFORME', 'ENCONTRADO EN CALIBRACIÓN', 'SENSOR CAMBIADO', 
        # Columnas de FECHA DE VENCIMIENTO
        'MES', 'AÑO',
        # Columnas adicionales de Informe
        'FECHA', 'REALIZADA POR',
        # Columnas de ALARMAS
        '1RA', '2DA', '3RA', 
        # Columnas de VALOR SPAM (ajusta los nombres si son diferentes)
        'EQUIPO', 'CILINDRO', 'UND', 
        # Columnas finales
        'OBSERVACION', 'ESTADO DE CALIBRACIÓN', 'NRO DE CERTIFICADO', 'FECHA DE CALIBRACIÓN POR CENERIS',
    ]

    # --- 1. CAPTURAR VALORES DE FILTRO DE LA URL (request.GET) ---
    busqueda_texto = request.GET.get('q', '')
    area_filtro = request.GET.get('area', '')
    estado_cal_filtro = request.GET.get('estado_cal', '')
    sensor_cambiado_filtro = request.GET.get('sensor_cambiado', '')
    vencimiento_desde = request.GET.get('venc_desde', '')
    vencimiento_hasta = request.GET.get('venc_hasta', '')
    
    # --- 2. CONSTRUIR EL QUERYSET BASE Y APLICAR FILTROS ---
    dispositivos_fijos = Dispositivo.objects.filter(tipoDisp='Fijo').select_related(
        'id_areaTrabajo_fijo'
    ).prefetch_related(
        'sensor_set__certificados_de_componente',
        'sensor_set__informes_calibracion__empresa_realizadora',
        'sensor_set__alarmas',
    ).order_by('nomDisp')

    # Aplicar filtros de texto (Nombre, TAG, Modelo de Sensor)
    if busqueda_texto:
        dispositivos_fijos = dispositivos_fijos.filter(
            Q(nomDisp__icontains=busqueda_texto) |
            Q(tag__icontains=busqueda_texto) |
            Q(sensor__nomComp__icontains=busqueda_texto)
        ).distinct() # .distinct() es importante cuando usamos Q en relaciones

    # Aplicar filtro por Área
    if area_filtro:
        # Asumo que 'area_filtro' contiene el PK de AreaTrabajo
        dispositivos_fijos = dispositivos_fijos.filter(id_areaTrabajo_fijo__pk=area_filtro)
    
    # Aplicar filtro por rango de fecha de vencimiento de garantía del sensor
    if vencimiento_desde:
        dispositivos_fijos = dispositivos_fijos.filter(sensor__fecVencGarantia__gte=vencimiento_desde)
    if vencimiento_hasta:
        dispositivos_fijos = dispositivos_fijos.filter(sensor__fecVencGarantia__lte=vencimiento_hasta)

    paginator = Paginator(dispositivos_fijos, 10) # 15 dispositivos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # --- 3. PROCESAR DATOS PARA LA PLANTILLA (AHORA DESPUÉS DE FILTRAR) ---
    datos_tabla = []
    
    for idx, dispositivo in enumerate(page_obj, start=page_obj.start_index()):
        sensores_del_dispositivo = dispositivo.sensor_set.all()
        
        
        lista_datos_sensores = []
        
        if sensores_del_dispositivo:
            # Si el dispositivo tiene sensores, iteramos sobre ellos
            for sensor in sensores_del_dispositivo:
                ultimo_certificado = sensor.certificados_de_componente.order_by('-fechCertificado').first()
                ultimo_informe = sensor.informes_calibracion.order_by('-fecha_informe').first()
                alarma = sensor.alarmas.first()
                # --- Lógica de Estado de Calibración ---
                estado_calibracion = 'No Calibrado'
                # Asumiendo 6 meses = 183 días para "Calibrado"
                if ultimo_certificado and timezone.now() - ultimo_certificado.fechCertificado <= timedelta(days=183):
                    estado_calibracion = 'Calibrado'
                
                alarma = None
                try:
                    # Hacemos una búsqueda explícita.
                    # Esto es más claro que usar el related_name.
                    alarma = Alarma.objects.get(sensor=sensor)
                    print(f"DEBUG: Alarma encontrada para el Sensor ID: {sensor.pk} ({sensor})")
                except Alarma.DoesNotExist:
                    # Es normal que un sensor no tenga alarma, no hacemos nada.
                    pass 
                except Alarma.MultipleObjectsReturned:
                    # ¡Esto es un problema de datos!
                    # Tienes múltiples alarmas para el mismo sensor.
                    self.stdout.write(self.style.WARNING(f"  - ADVERTENCIA: Se encontraron múltiples alarmas para el sensor {sensor}."))
                    # Tomamos la primera como fallback.
                    alarma = Alarma.objects.filter(sensor=sensor).first()

                # --- PRINT DE DEPURACIÓN ---
                if not alarma:
                    print(f"DEBUG: No se encontró alarma para el Sensor ID: {sensor.pk} ({sensor})")
                # --- FIN DEPURACIÓN ---

                datos_sensor = {
                    'modelo_sensor': sensor.nomComp,
                    'tipo_gas': sensor.tipGas,
                    'informe': ultimo_informe.informe if ultimo_informe else '',
                    'encontrado_calibracion': ultimo_informe.encontrado_calibracion if ultimo_informe else '',
                    'sensor_cambiado': "Sí" if ultimo_informe and ultimo_informe.sensor_cambiado else "No",
                    'mes_vencimiento': sensor.fecVencGarantia.strftime('%B').capitalize() if sensor.fecVencGarantia else '',
                    'anio_vencimiento': sensor.fecVencGarantia.year if sensor.fecVencGarantia else '',
                    'fecha_informe': ultimo_informe.fecha_informe.strftime('%d/%m/%Y') if ultimo_informe and ultimo_informe.fecha_informe else '',
                    'realizada_por': ultimo_informe.empresa_realizadora.nombreE if ultimo_informe and ultimo_informe.empresa_realizadora else '',
                    'observacion': ultimo_informe.observacion if ultimo_informe else '',
                    'estado_calibracion': estado_calibracion,
                    'nro_certificado': ultimo_certificado.nro_certificado if ultimo_certificado else '',
                    'fecha_calibracion_ceneris': ultimo_certificado.fechCertificado.date().strftime('%d/%m/%Y') if ultimo_certificado else '',
                    'alarma_1ra': alarma.primera if alarma else '',
                    'alarma_2da': alarma.segunda if alarma else '',
                    'alarma_3ra': alarma.tercera if alarma else '',
                    'valor_spam_equipo': alarma.equipo if alarma else '',
                    'valor_spam_cilindro': alarma.cilindro if alarma else '',
                    'valor_spam_und': alarma.und if alarma else '',
                }
                lista_datos_sensores.append(datos_sensor)
        else:
            # Si no hay sensores, añadimos una fila de datos vacía para que el dispositivo aparezca
            lista_datos_sensores.append({ 
                'modelo_sensor': 'Sin sensor registrado',
                'estado_calibracion': 'No Calibrado' # Default para filtrado
            })

        # --- FILTROS POST-PROCESAMIENTO (Aplicados a datos calculados en Python) ---
        
        # Filtro 'sensor_cambiado' (a nivel de dispositivo: ¿ALGÚN sensor fue cambiado?)
        #if sensor_cambiado_filtro:
            # Buscamos si ALGÚN informe de CALIBRACIÓN en CUALQUIERA de los sensores tuvo un cambio.
            #hubo_cambio = any(
                #(informe.sensor_cambiado for sensor in dispositivo.sensor_set.all() 
                 #for informe in sensor.informes_calibracion.all())
            #)
            # Convertimos el filtro a booleano (sensor_cambiado_filtro es 'Si' o 'No')
            #filtro_bool = sensor_cambiado_filtro == 'Si'
            #if hubo_cambio != filtro_bool:
                #continue # Saltamos todo el dispositivo si no cumple la condición

        # Filtro 'estado_calibracion' (a nivel de sensor)
        lista_datos_sensores_filtrada = []
        if lista_datos_sensores:
            for sensor_data in lista_datos_sensores:
                if estado_cal_filtro and sensor_data['estado_calibracion'] != estado_cal_filtro:
                    continue # Saltamos este sensor si no cumple con el filtro de estado
                lista_datos_sensores_filtrada.append(sensor_data)
        
        # Si después de filtrar no quedan sensores Y hay un filtro de sensor (para no saltar dispositivos sin filtro)
        # O si el dispositivo no tenía sensores y se está aplicando el filtro de estado de calibración.
        if not lista_datos_sensores_filtrada and (estado_cal_filtro or busqueda_texto):
             continue

        # Si pasamos los filtros, construimos y añadimos el diccionario principal
        num_sensores_final = len(lista_datos_sensores_filtrada) if lista_datos_sensores_filtrada else 1
        
        fila_dispositivo = {
            'numero': idx,
            'objeto_dispositivo': dispositivo,
            'num_sensores': num_sensores_final, # Cuántas filas ocupará
            'datos_comunes': {
                'nombre_dispositivo': dispositivo.nomDisp,
                # Asumo que tienes un campo 'area_general' o lo calculas
                'area': dispositivo.area_general if hasattr(dispositivo, 'area_general') else 'N/A', 
                'ubicacion_area': dispositivo.id_areaTrabajo_fijo.nombreA if dispositivo.id_areaTrabajo_fijo else '',
                'tag': dispositivo.tag,
                'ubicacion_sensor': '', # Este campo debe llenarse con la data de la ubicación del sensor en el detector
            },
            'datos_sensores': lista_datos_sensores_filtrada, # La lista de datos por sensor
        }
        
        # Re-indexamos si el filtro de estado_cal_filtro elimina todos los sensores de un dispositivo
        if fila_dispositivo['datos_sensores']:
            datos_tabla.append(fila_dispositivo)
    
    # --- 4. PREPARAR DATOS PARA LOS DESPLEGABLES DE FILTRO ---
    # Asumo que AreaTrabajo es un modelo disponible
    opciones_areas = AreaTrabajo.objects.all().order_by('nombreA')
    opciones_estado_cal = [('Calibrado', 'Calibrado'), ('No Calibrado', 'No Calibrado')]
    opciones_sensor_cambiado = [('Si', 'Sí'), ('No', 'No')] # Para el filtro booleano
    
    # --- 5. CONSTRUIR EL CONTEXTO FINAL ---
    context = {
        'titulo': 'Reporte de Dispositivos Fijos',
        'headers': headers,
        'super_headers': super_headers,
        'page_obj': page_obj,
        'datos_tabla': datos_tabla,
        'url_exportacion_excel': reverse('cenerisapp:exportar_fijos_excel'),
        
        # Datos para los filtros
        'opciones_areas': opciones_areas,
        'opciones_estado_cal': opciones_estado_cal,
        'opciones_sensor_cambiado': opciones_sensor_cambiado,
        'filtros_aplicados': {
            'q': busqueda_texto,
            'area': int(area_filtro) if area_filtro else None,
            'estado_cal': estado_cal_filtro,
            'sensor_cambiado': sensor_cambiado_filtro,
            'venc_desde': vencimiento_desde,
            'venc_hasta': vencimiento_hasta,
        }
    }
    
    return render(request, 'tabla/vista_reporte_fijos.html', context)

@login_required
def dashboard_fijos(request):
    
    # --- PARTE 1: DATOS GENERALES PARA EL DASHBOARD ---
    
    fijos_qs = Dispositivo.objects.filter(tipoDisp='Fijo')
    
    # 1.1 Conteo de dispositivos fijos
    total_fijos = fijos_qs.count()
    
    # 1.2 Conteo de sensores instalados en dispositivos fijos
    total_sensores_fijos = Sensor.objects.filter(dispositivo_instalado__tipoDisp='Fijo').count()
    
    # 1.3 Conteo de dispositivos "Calibrados"
    # Un dispositivo se considera calibrado si tiene al menos un sensor con un certificado de los últimos 6 meses.
    seis_meses_atras = timezone.now() - timedelta(days=183)
    dispositivos_calibrados_ids = Certificado.objects.filter(
        componente__sensor__dispositivo_instalado__tipoDisp='Fijo',
        fechCertificado__gte=seis_meses_atras
    ).values_list('componente__sensor__dispositivo_instalado_id', flat=True).distinct()
    total_calibrados = len(dispositivos_calibrados_ids)

    # 1.4 Últimos 5 sensores calibrados (basado en la fecha del certificado)
    ultimos_certificados = Certificado.objects.filter(componente__sensor__dispositivo_instalado__tipoDisp='Fijo') \
        .select_related('componente__sensor', 'componente__sensor__dispositivo_instalado') \
        .order_by('-fechCertificado')[:5]

    # 1.5 Gráfico de cantidad de fijos por área de trabajo
    dispositivos_por_area = fijos_qs.values('id_areaTrabajo_fijo__nombreA') \
        .annotate(total=Count('id_dispositivo')).order_by('-total')
    
    # --- PARTE 2: BÚSQUEDA Y DATOS ESPECÍFICOS ---
    dispositivo_seleccionado = None
    lista_certificados_ordenada = []
    query_busqueda = request.GET.get('query', '')
    
    if query_busqueda:
        try:
            dispositivo_seleccionado = Dispositivo.objects.prefetch_related(
                'sensor_set__certificados_de_componente'
            ).get(num_serie__iexact=query_busqueda, tipoDisp='Fijo')

            todos_los_certificados = []
            for sensor in dispositivo_seleccionado.sensor_set.all():
                for cert in sensor.certificados_de_componente.all():
                    # Añadimos el certificado y una referencia a su sensor padre
                    todos_los_certificados.append({'cert': cert, 'sensor': sensor})
            
            # Ordenamos la lista combinada por fecha, de más reciente a más antigua
            todos_los_certificados.sort(key=lambda x: x['cert'].fechCertificado, reverse=True)
            
            # Nos quedamos solo con los 5 más recientes
            lista_certificados_ordenada = todos_los_certificados[:5]

        except Dispositivo.DoesNotExist:
            messages.error(request, f"No se encontró ningún dispositivo fijo con el N° de Serie '{query_busqueda}'.")

    # --- PARTE 3: CONTEXTO PARA LA PLANTILLA ---
    context = {
        'titulo': 'Dashboard de Dispositivos Fijos',
        
        # Datos Generales
        'total_fijos': total_fijos,
        'total_sensores_fijos': total_sensores_fijos,
        'total_calibrados': total_calibrados,
        'ultimos_certificados': ultimos_certificados,
        'dispositivos_por_area_json': json.dumps([
            {'area': item['id_areaTrabajo_fijo__nombreA'] or 'Sin Área', 'total': item['total']}
            for item in dispositivos_por_area
        ]),
        
        # Datos para Búsqueda
        'todos_los_fijos_para_busqueda': fijos_qs,
        'query_busqueda': query_busqueda,
        'dispositivo_seleccionado': dispositivo_seleccionado,
        'lista_certificados_ordenada': lista_certificados_ordenada,
    }
    
    return render(request, 'dashboard/dashboard_fijos.html', context)


@login_required
def muro_ocurrencias(request):
    # Procesamos el envío del formulario si la petición es POST
    if request.method == 'POST':
        form = OcurrenciaForm(request.POST)
        if form.is_valid():
            # Creamos la instancia en memoria
            nueva_ocurrencia = form.save(commit=False)
            # ¡Asignamos el usuario de la sesión actual como autor!
            nueva_ocurrencia.autor = request.user
            # Ahora guardamos en la base de datos
            nueva_ocurrencia.save()
            
            messages.success(request, "Ocurrencia publicada exitosamente.")
            # Redirigimos a la misma página para ver el nuevo comentario y limpiar el formulario
            return redirect('cenerisapp:muro_ocurrencias')
        else:
            messages.error(request, "El mensaje no puede estar vacío.")

    # Para peticiones GET (o si el POST falla), preparamos la página
    
    # Obtenemos todas las ocurrencias para mostrarlas
    # Usamos select_related para optimizar la obtención del nombre del autor
    ocurrencias = Ocurrencia.objects.all().select_related('autor')
    
    # Creamos una instancia vacía del formulario para mostrarla en la página
    form = OcurrenciaForm()
    
    context = {
        'titulo': 'Muro de Ocurrencias y Noticias',
        'form': form,
        'ocurrencias': ocurrencias,
    }
    
    return render(request, 'ocurrencias/muro_ocurrencias.html', context)

@login_required
def borrar_ocurrencia(request, ocurrencia_id):
    # Solo aceptamos peticiones POST para esta acción por seguridad.
    if request.method != 'POST':
        # Si alguien intenta acceder por GET, le negamos el acceso.
        return HttpResponseForbidden("Método no permitido.")

    # 1. Obtenemos la ocurrencia que se quiere borrar.
    ocurrencia = get_object_or_404(Ocurrencia, pk=ocurrencia_id)
    
    # --- LA LÓGICA DE SEGURIDAD MÁS IMPORTANTE ---
    # 2. Verificamos si el usuario de la sesión actual es el autor de la ocurrencia.
    if ocurrencia.autor != request.user:
        # Si no es el autor, le negamos el permiso.
        messages.error(request, "No tienes permiso para borrar esta ocurrencia.")
        return HttpResponseForbidden("No tienes permiso para realizar esta acción.")
    
    # 3. Si las comprobaciones pasan, procedemos a borrar.
    ocurrencia.delete()
    
    messages.success(request, "Ocurrencia borrada exitosamente.")
    
    # 4. Redirigimos de vuelta al muro de ocurrencias.
    return redirect('cenerisapp:muro_ocurrencias')

@login_required
def gestor_reportes(request, tipo_reporte):
    # Validamos que el tipo de reporte sea válido
    tipos_validos = [choice[0] for choice in ReporteDiario.TIPO_CHOICES]
    tipo_reporte_upper = tipo_reporte.upper()
    if tipo_reporte_upper not in tipos_validos:
        raise Http404("Tipo de reporte no válido.")

    # Manejo de la carga de archivos (AJAX)
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        fecha_str = request.POST.get('fecha')
        archivo = request.FILES.get('archivo')

        if not fecha_str or not archivo:
            return JsonResponse({'error': 'Faltan datos.'}, status=400)
        
        try:
            fecha_obj = date.fromisoformat(fecha_str)
            tipo_reporte_upper = tipo_reporte.upper()

            # --- LÓGICA DE GUARDADO CORREGIDA Y ROBUSTA ---
            
            # 1. Usamos get_or_create para ver si ya existe un registro.
            #    Esto es más explícito que update_or_create para FileFields.
            reporte, created = ReporteDiario.objects.get_or_create(
                tipo_reporte=tipo_reporte_upper,
                fecha=fecha_obj,
                defaults={'archivo': archivo} # Asignamos el archivo solo si se está creando
            )
            
            # 2. Si el registro ya existía, actualizamos el archivo explícitamente.
            if not created:
                # Opcional: Borramos el archivo antiguo de S3 antes de subir el nuevo.
                # Esto evita tener archivos huérfanos.
                if reporte.archivo:
                    reporte.archivo.delete(save=False) # save=False evita una consulta extra
                
                # Asignamos el nuevo archivo y guardamos.
                reporte.archivo = archivo
                reporte.save()
            
            # --- FIN DE LA LÓGICA DE GUARDADO ---
            
            # La URL del archivo ahora se obtendrá correctamente desde S3 en producción.
            file_url = reporte.archivo.url

            return JsonResponse({
                'status': 'ok',
                'message': 'Archivo cargado exitosamente.',
                'file_url': file_url
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # Lógica para la vista normal (GET)
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month

    # Construir el calendario
    cal = calendar.Calendar()
    dias_del_mes = cal.monthdatescalendar(year, month)

    # Obtener los archivos ya cargados para este mes y tipo
    archivos_cargados = ReporteDiario.objects.filter(
        tipo_reporte=tipo_reporte.upper(),
        fecha__year=year,
        fecha__month=month
    ) # Obtenemos los objetos completos

    
    # Convertir a un diccionario para búsqueda rápida en la plantilla
    mapa_archivos = {}
    for reporte in archivos_cargados:
        if reporte.archivo:
            # Para cada reporte, llamamos a .url para obtener la URL completa de S3
            mapa_archivos[reporte.fecha] = reporte.archivo.url


    choices_dict = dict(ReporteDiario.TIPO_CHOICES)
    # Obtenemos el texto legible usando la clave
    titulo_legible = choices_dict.get(tipo_reporte_upper, "Reporte Desconocido")

    context = {
        'titulo': f'Gestor de {titulo_legible}',
        'tipo_reporte': tipo_reporte,
        'dias_del_mes': dias_del_mes,
        'current_month': date(year, month, 1),
        'mapa_archivos': mapa_archivos,
        'es_quincenal': tipo_reporte.upper() == 'QUINCENAL',
    }
    
    return render(request, 'archivos/gestor_reportes.html', context)

@login_required
def registro_rapido_in_out(request):
    
    # --- MANEJO DE CONFIGURACIÓN DE SESIÓN (Turno) ---
    if request.method == 'POST' and ('set_turno' in request.POST or 'set_config' in request.POST):
        # Turno
        turno = request.POST.get('turno')
        if turno in ['A', 'B']: request.session['turno_activo'] = turno
        else: request.session.pop('turno_activo', None)
        
        # Área de Trabajo
        area_id = request.POST.get('area_trabajo')
        if area_id: request.session['area_trabajo_id'] = area_id
        else: request.session.pop('area_trabajo_id', None)

        # Punto Exacto
        punto_id = request.POST.get('punto_exacto')
        if punto_id: request.session['punto_exacto_id'] = punto_id
        else: request.session.pop('punto_exacto_id', None)
        
        messages.success(request, "Configuración de sesión actualizada.")
        return redirect('cenerisapp:registro_rapido_in_out')

    # --- FINALIZAR PRÉSTAMO Y LIMPIAR SESIÓN DE TRABAJADOR ---
    if request.method == 'POST' and 'end_prestamo' in request.POST:
        request.session.pop('receptor_prestamo_id', None)
        request.session.pop('receptor_prestamo_nombre', None)
        messages.info(request, "Préstamo finalizado. Escanee el DNI de un nuevo trabajador.")
        return redirect('cenerisapp:registro_rapido_in_out')
        
    # --- PROCESO DE ESCANEO (TRABAJADOR O DISPOSITIVO) ---
    if request.method == 'POST' and 'codigo_escaneado' in request.POST:
        codigo = request.POST.get('codigo_escaneado', '').strip()
        
        if not codigo:
            messages.warning(request, "El campo de escaneo estaba vacío.")
            return redirect('cenerisapp:registro_rapido_in_out')
        
        # Intentamos identificar si es un DNI (8 dígitos numéricos)
        if len(codigo) == 8 and codigo.isdigit():
            # --- LÓGICA PARA FIJAR TRABAJADOR POR DNI ---
            try:
                receptor = Empleado.objects.get(dni=codigo)
                request.session['receptor_prestamo_id'] = receptor.pk
                request.session['receptor_prestamo_nombre'] = receptor.nomEmpleado
                messages.success(request, f"Trabajador receptor fijado: {receptor.nomEmpleado}. Ahora puede escanear los dispositivos.")
            except Empleado.DoesNotExist:
                messages.error(request, f"No se encontró ningún empleado con el DNI '{codigo}'.")
        
        else: # Si no es un DNI, asumimos que es un N° de Serie de dispositivo
            # --- LÓGICA DE PRÉSTAMO/DEVOLUCIÓN DE DISPOSITIVO ---
            try:
                dispositivo = Dispositivo.objects.get(num_serie=codigo)
                registro_abierto = Registro.objects.filter(id_dispositivo=dispositivo, fecDevol__isnull=True).first()
                
                if registro_abierto: # Lógica de DEVOLUCIÓN
                    registro_abierto.fecDevol = timezone.now()
                
                    # 2. Asignamos el operador actual como el que recibe la devolución.
                    registro_abierto.operador_receptor = request.user.empleado
                    
                    # 3. Guardamos los cambios en la base de datos.
                    registro_abierto.save()
    
                    messages.success(request, f"ENTRADA: Dispositivo '{dispositivo.nomDisp}' devuelto.")
                else: # Lógica de PRÉSTAMO
                    receptor_id_sesion = request.session.get('receptor_prestamo_id')
                    turno_sesion = request.session.get('turno_activo')
                    area_id_sesion = request.session.get('area_trabajo_id') # <-- Leemos de la sesión
                    punto_id_sesion = request.session.get('punto_exacto_id')

                    if not receptor_id_sesion:
                        messages.error(request, "Error: Debe escanear primero el DNI de un trabajador antes de prestar un dispositivo.")
                    elif not turno_sesion:
                        messages.error(request, "Error: Debe seleccionar un turno antes de registrar un préstamo.")
                    elif not area_id_sesion: # <-- Nueva validación
                        messages.error(request, "Error: Debe seleccionar un Área de Operación.")
                    else:
                        operador = request.user.empleado
                        receptor = Empleado.objects.get(pk=receptor_id_sesion)
                        Registro.objects.create(
                            id_dispositivo=dispositivo,
                            operador_responsable=operador,
                            trabajador_receptor=receptor,
                            turno=turno_sesion, # ¡Asignamos el turno de la sesión!
                            area_trabajo_operacion_id=area_id_sesion, # <-- Asignamos
                            punto_exacto_operacion_id=punto_id_sesion, # <-- Asignamos
                        )
                        messages.success(request, f"SALIDA: Dispositivo '{dispositivo.nomDisp}' prestado a {receptor.nomEmpleado} en turno {turno_sesion}.")
            
            except Dispositivo.DoesNotExist:
                messages.error(request, f"Dispositivo con N/S '{codigo}' no encontrado.")
            except Empleado.DoesNotExist:
                messages.error(request, "Error de sesión. Por favor, vuelva a escanear al trabajador.")
        
        return redirect('cenerisapp:registro_rapido_in_out')
    
    # --- NUEVA LÓGICA PARA REGISTRAR EMPLEADO RÁPIDO ---
    empleado_form = EmpleadoRapidoForm() # Inicializamos el form
    if request.method == 'POST' and 'registrar_empleado' in request.POST:
        empleado_form = EmpleadoRapidoForm(request.POST)
        if empleado_form.is_valid():
            try:
                empresa_id = empleado_form.cleaned_data.get('empresa_id')
                empresa_nombre = empleado_form.cleaned_data.get('empresa_nombre')
                
                # Buscamos o creamos la empresa
                if empresa_id:
                    empresa = Empresa.objects.get(pk=empresa_id)
                else:
                    # get_or_create para evitar duplicados si se escribe el nombre exacto
                    empresa, created = Empresa.objects.get_or_create(
                        nombreE=empresa_nombre,
                        defaults={
                            # Rellenamos campos obligatorios con valores por defecto si los hay
                            'abreviacion': empresa_nombre[:20], 
                            'direccion': 'N/A',
                            'departamento': 'N/A',
                            'telefono': 'N/A',
                            'ruc': '00000000000',
                        }
                    )
                
                # Creamos el nuevo empleado
                nuevo_empleado = Empleado.objects.create(
                    empresa=empresa,
                    nomEmpleado=empleado_form.cleaned_data.get('nomEmpleado'),
                    dni=empleado_form.cleaned_data.get('dni'),
                    puesto=empleado_form.cleaned_data.get('puesto')
                )
                
                messages.success(request, f"Empleado '{nuevo_empleado.nomEmpleado}' creado y asignado a la empresa '{empresa.nombreE}'.")
                
                # Fijamos al nuevo empleado como el receptor activo en la sesión
                request.session['receptor_prestamo_id'] = nuevo_empleado.pk
                request.session['receptor_prestamo_nombre'] = nuevo_empleado.nomEmpleado

            except Exception as e:
                messages.error(request, f"Error al crear el empleado: {e}")
            
            return redirect('cenerisapp:registro_rapido_in_out')

    # --- VISTA GET (Mostrar la página) ---
    empleados_receptores = Empleado.objects.all().order_by('nomEmpleado')
    receptor_id_activo = request.session.get('receptor_prestamo_id')
    turno_activo = request.session.get('turno_activo')
    receptor_nombre_activo = request.session.get('receptor_prestamo_nombre')
    area_activa_id = request.session.get('area_trabajo_id')
    punto_activo_id = request.session.get('punto_exacto_id')

    query = request.GET.get('q', '') # Capturamos el parámetro de búsqueda 'q'
    
    # Queryset base
    historial_completo = Registro.objects.select_related(
        'id_dispositivo', 'trabajador_receptor'
    )
    
    if query:
        # Filtramos por el número de serie del dispositivo relacionado
        historial_completo = historial_completo.filter(id_dispositivo__num_serie__icontains=query)

    # --- 2. ORDENAMIENTO INTELIGENTE ---
    # Ordena por 'fecDevol' ascendente, poniendo los NULL (no devueltos) primero.
    # Luego, como segundo criterio, ordena por fecha de registro descendente.
    historial_completo = historial_completo.order_by(F('fecDevol').asc(nulls_first=True), '-fecRegistro')
    
    # --- 3. LÓGICA DE PAGINACIÓN ---
    paginator = Paginator(historial_completo, 15) # 15 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    hoy = timezone.now().date()
    
    # 1. Obtenemos todos los registros de SALIDA que ocurrieron hoy.
    salidas_hoy_qs = Registro.objects.filter(fecRegistro__date=hoy)
    
    # 2. Agrupamos estos registros por el nombre del dispositivo para obtener el TOTAL de salidas.
    #    .values() agrupa, .annotate() cuenta.
    salidas_hoy_por_modelo = salidas_hoy_qs.values('id_dispositivo__nomDisp') \
        .annotate(total_salidas=Count('id_dispositivo__nomDisp')) \
        .order_by('id_dispositivo__nomDisp')

    # 3. Obtenemos los registros de SALIDA de hoy que AÚN NO HAN SIDO DEVUELTOS.
    pendientes_hoy_qs = salidas_hoy_qs.filter(fecDevol__isnull=True)
    
    # 4. Agrupamos estos para obtener el TOTAL de pendientes.
    pendientes_hoy_por_modelo = pendientes_hoy_qs.values('id_dispositivo__nomDisp') \
        .annotate(total_pendientes=Count('id_dispositivo__nomDisp')) \
        .order_by('id_dispositivo__nomDisp')
        
    # 5. Combinamos los dos resultados en una sola estructura de datos para la plantilla.
    #    Usamos un diccionario para facilitar la combinación.
    kpis_por_modelo = {
        item['id_dispositivo__nomDisp']: {'salidas': item['total_salidas'], 'pendientes': 0}
        for item in salidas_hoy_por_modelo
    }
    
    for item in pendientes_hoy_por_modelo:
        if item['id_dispositivo__nomDisp'] in kpis_por_modelo:
            kpis_por_modelo[item['id_dispositivo__nomDisp']]['pendientes'] = item['total_pendientes']

    # Mostramos los dispositivos actualmente prestados a esta persona
    prestamos_activos = []
    if receptor_id_activo:
        prestamos_activos = Registro.objects.filter(
            trabajador_receptor_id=receptor_id_activo,
            fecDevol__isnull=True
        ).select_related('id_dispositivo').order_by('-fecRegistro')
    
    todas_las_areas = AreaTrabajo.objects.all().order_by('nombreA')
    
    context = {
        'titulo': 'Registro Rápido IN/OUT (Modo Lote)',
        'turno_choices': Registro._meta.get_field('turno').choices,
        'turno_activo': turno_activo,
        'todas_las_areas': todas_las_areas,
        'area_activa_id': int(area_activa_id) if area_activa_id else None,
        'punto_activo_id': int(punto_activo_id) if punto_activo_id else None,
        'empleados_receptores': empleados_receptores,
        'receptor_id_activo': receptor_id_activo,
        'receptor_nombre_activo': receptor_nombre_activo,
        'prestamos_activos': prestamos_activos,
        'page_obj': page_obj, # Reemplaza a 'ultimos_movimientos'
        'query': query, # Para que el buscador recuerde el término buscado

        'kpis_por_modelo': kpis_por_modelo,
        'fecha_hoy': hoy,
        'empleado_form': empleado_form,
    }
    return render(request, 'registros/registro_rapido.html', context)

@login_required
def cargar_historial_modificaciones(request):
    ModificacionFormSet = formset_factory(ModificacionAntiguaForm, extra=5)

    if request.method == 'POST':
        # CAMBIO IMPORTANTE: Agregar request.FILES para recibir las imágenes
        formset = ModificacionFormSet(request.POST, request.FILES)
        
        if formset.is_valid():
            registros_creados = 0
            registros_omitidos = 0

            for form in formset:
                if not form.cleaned_data:
                    continue

                data = form.cleaned_data
                
                nombre_sensor_saliente = data['sensor_saliente_nombre']
                nombre_sensor_entrante = data['sensor_entrante_nombre']

                # 1. Crear/Obtener Sensores (Igual que antes)
                sensor_saliente, _ = Sensor.objects.get_or_create(
                    nSerieActual=data['sensor_saliente_ns'],
                    defaults={'nomComp': nombre_sensor_saliente, 'tipGas': nombre_sensor_saliente, 'estComp': 'Inoperativo por cambio'}
                )

                sensor_entrante, _ = Sensor.objects.get_or_create(
                    nSerieActual=data['sensor_entrante_ns'],
                    defaults={'nomComp': nombre_sensor_entrante, 'tipGas': nombre_sensor_entrante}
                )
                
                # 2. Verificar duplicados (Igual que antes)
                if Modificacion.objects.filter(
                    id_dispositivo=data['dispositivo'],
                    fecInstalacionMod=data['fecInstalacionMod'],
                    sensor_saliente=sensor_saliente,
                    componente_entrante=sensor_entrante
                ).exists():
                    registros_omitidos += 1
                    continue 
                
                # 3. Crear Modificación
                nueva_modificacion = Modificacion.objects.create(
                    id_dispositivo=data['dispositivo'],
                    fecInstalacionMod=data['fecInstalacionMod'],
                    sensor_saliente=sensor_saliente,
                    componente_entrante=sensor_entrante,
                    id_trabajador=data['id_trabajador'],
                    MotivoCambio=data['MotivoCambio'],
                    tipoServicio='Reparacion'
                )

                # --- 4. NUEVA LÓGICA: GUARDAR LA FOTO Y VINCULARLA ---
                imagen = data.get('evidencia_foto')
                if imagen:
                    # Usamos el nombre del sensor entrante como 'tipo_foto' (ej. 'O2', 'LEL')
                    # Esto asegura que el Excel sepa dónde ponerla.
                    tipo_evidencia = nombre_sensor_entrante if nombre_sensor_entrante else "MANTENIMIENTO"
                    
                    FotoDispositivo.objects.create(
                        dispositivo=data['dispositivo'],
                        modificacion=nueva_modificacion, # ¡VINCULACIÓN CLAVE!
                        imagen_original=imagen,
                        tipo_foto=tipo_evidencia,
                        contexto='CARDEX' # Para que aparezca en el reporte
                    )

                registros_creados += 1
            
            messages.success(request, f"Proceso finalizado. {registros_creados} modificaciones con sus fotos creadas.")
            return redirect('cenerisapp:lista_modificaciones')
        
        else:
            messages.error(request, "Por favor, corrige los errores en los formularios.")
    
    else: 
        formset = ModificacionFormSet()

    context = {
        'titulo': 'Carga Rápida de Historial de Modificaciones',
        'formset': formset,
    }
    return render(request, 'modificaciones/cargar_historial.html', context)
    
@login_required
def buscar_empresas_api(request):
    query = request.GET.get('term', '')
    empresas = Empresa.objects.filter(nombreE__icontains=query)[:10]
    resultados = [
        {'id': empresa.pk, 'label': empresa.nombreE, 'value': empresa.nombreE}
        for empresa in empresas
    ]
    return JsonResponse(resultados, safe=False)

@login_required
def flujo(request): 
    if request.method == 'POST':
        form = RegistroSalidaForm(request.POST)
        
        if form.is_valid():
            # ¡La magia ocurre aquí!
            # El form.cleaned_data ya está validado, y la instancia del modelo ya
            # tiene id_dispositivo_id o id_componente_id asignados.
            registro = form.save(commit=False)
            
            # --- LÓGICA SIMPLIFICADA ---
            if registro.id_dispositivo:
                dispositivo_prestado = registro.id_dispositivo
                if registro.trabajador_receptor and hasattr(registro.trabajador_receptor, 'empresa') and registro.trabajador_receptor.empresa:
                    dispositivo_prestado.id_empresa = registro.trabajador_receptor.empresa
                    dispositivo_prestado.save()
                    registro.save()
                    messages.success(request, f"Salida registrada para el dispositivo '{dispositivo_prestado.nomDisp}'.")
                    return redirect('cenerisapp:lista_registros')
                else:
                    messages.error(request, "El trabajador seleccionado no tiene una empresa asignada.")

            elif registro.id_componente:
                componente_prestado = registro.id_componente
                if hasattr(componente_prestado, 'otrocomponente'):
                    componente_prestado.otrocomponente.estComp = 'Prestado'
                    componente_prestado.otrocomponente.save()
                elif hasattr(componente_prestado, 'sensor'):
                    componente_prestado.sensor.estComp = 'Prestado'
                    componente_prestado.sensor.save()
                
                registro.save()
                messages.success(request, f"Salida registrada y stock actualizado para el componente '{componente_prestado.nomComp}'.")
                return redirect('cenerisapp:lista_registros')
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else: 
        form = RegistroSalidaForm()
 
    context = {
        'form': form,
        'titulo': 'Registro de Prestamo de Dispositivos'
    }
    return render(request, 'flujo/flujo.html', context)
 
@login_required
def get_empleado_info(request, empleado_id):
    try:
        
        empleado = Empleado.objects.select_related('areaTrabajo').get(pk=empleado_id)
        data = {
            'nombre': empleado.nomEmpleado,
            'puesto': empleado.puesto,
            'area_trabajo': empleado.areaTrabajo.nombreA if empleado.areaTrabajo else 'No asignada'
        }
        return JsonResponse(data)
    except Empleado.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)

def get_puntos_por_area_api(request):
    area_id = request.GET.get('area_id')
    puntos = PuntoExacto.objects.filter(area_trabajo_id=area_id).values('id', 'nombre_punto')
    return JsonResponse(list(puntos), safe=False)

@login_required
def lista_registros(request):
    registros_qs = Registro.objects.select_related('trabajador_receptor', 'operador_responsable', 'id_dispositivo').order_by('-fecRegistro')
   
    
    registros_procesados = []
    for registro in registros_qs:
        duracion_en_horas = None 
        if registro.durPrestamo:
            
            
            total_segundos = registro.durPrestamo.total_seconds()
            duracion_en_horas = total_segundos / 3600
       
        
        registro.duracion_horas = duracion_en_horas
        registros_procesados.append(registro)
 
    context = {
        'registros': registros_procesados, 
        'titulo': 'Historial de Prestamo de Dispositivos'
    }
    return render(request, 'flujo/lista_registros.html', context)



@login_required
def vista_inoperativos(request):
    
    # --- 1. CONSULTAS BASE ---
    dispositivos_inoperativos_qs = Dispositivo.objects.filter(estadoD__in=['Inoperativo', 'Extraviado']).order_by('nomDisp')
    sensores_inoperativos_qs = Sensor.objects.filter(estComp__in=['Inoperativo', 'Inoperativo por cambio']).order_by('nomComp')
    otros_componentes_inoperativos_qs = OtroComponente.objects.filter(estComp='Inoperativo').order_by('nomComp')

    # --- 2. PAGINACIÓN INDEPENDIENTE ---
    # Paginador para Dispositivos
    disp_paginator = Paginator(dispositivos_inoperativos_qs, 5) # 5 por página
    disp_page_number = request.GET.get('page_disp')
    disp_page_obj = disp_paginator.get_page(disp_page_number)

    # Paginador para Sensores
    sensor_paginator = Paginator(sensores_inoperativos_qs, 5)
    sensor_page_number = request.GET.get('page_sensor')
    sensor_page_obj = sensor_paginator.get_page(sensor_page_number)

    # Paginador para Otros Componentes
    otro_paginator = Paginator(otros_componentes_inoperativos_qs, 5)
    otro_page_number = request.GET.get('page_otro')
    otro_page_obj = otro_paginator.get_page(otro_page_number)

    # --- 3. CONTEXTO PARA LA PLANTILLA ---
    context = {
        'titulo': 'Dashboard de Ítems Inoperativos',
        
        # Objetos de Paginación para cada sección
        'disp_page_obj': disp_page_obj,
        'sensor_page_obj': sensor_page_obj,
        'otro_page_obj': otro_page_obj,
        
        # Conteos para las tarjetas KPI (usamos los querysets completos, es eficiente)
        'total_dispositivos_inop': dispositivos_inoperativos_qs.count(),
        'total_sensores_inop': sensores_inoperativos_qs.count(),
        'total_otros_inop': otros_componentes_inoperativos_qs.count(),
    }
    return render(request, 'inoperativos/inoperativos.html', context)


@require_POST
def marcar_operativo(request, tipo_item, item_id):
    """
    Cambia el estado de un ítem específico a 'Operativo'.
    Recibe el tipo de ítem y su ID desde la URL.
    """
    try:
        if tipo_item == 'dispositivo':
            # Usamos el nombre de la clave primaria explícita: id_dispositivo
            item = get_object_or_404(Dispositivo, id_dispositivo=item_id)
            item.estadoD = 'Operativo'
            item.save()
            messages.success(request, f'El dispositivo "{item}" ha sido marcado como Operativo.')

        elif tipo_item == 'sensor':
            # Usamos el nombre de la clave primaria explícita: id_componente
            item = get_object_or_404(Sensor, id_componente=item_id)
            item.estComp = 'Operativo'
            item.save()
            messages.success(request, f'El sensor "{item}" ha sido marcado como Operativo.')

        # Cambiamos 'parte' por 'otrocomponente' para que coincida con el modelo
        elif tipo_item == 'otrocomponente':
            # Usamos el nombre de la clave primaria explícita: id_componente
            item = get_object_or_404(OtroComponente, id_componente=item_id)
            item.estComp = 'Operativo'
            item.save()
            messages.success(request, f'El componente "{item}" ha sido marcado como Operativo.')

        else:
            messages.error(request, 'Tipo de ítem no válido.')

    except Exception as e:
        messages.error(request, f'Ocurrió un error al actualizar el ítem: {e}')

    return redirect('cenerisapp:vista_inoperativos')

@login_required
def registrar_devolucion(request, registro_id):
    
    if request.method == 'POST':
        try:

            registro = Registro.objects.get(pk=registro_id, fecDevol__isnull=True)
            
            
            registro.fecDevol = timezone.now()
            
            
            duracion = registro.fecDevol - registro.fecRegistro
            registro.durPrestamo = duracion

            if registro.id_componente:
                componente_devuelto = registro.id_componente
                # Lo devolvemos al stock marcándolo como 'Operativo'
                # (Aquí podrías añadir un formulario para que el usuario confirme
                # si el componente volvió en buen estado o inoperativo)
                componente_devuelto.estComp = 'Operativo'
                componente_devuelto.save()
            
            # (Aquí podrías añadir lógica para el dispositivo, como cambiar su 'id_empresa' a NULL)
            elif registro.id_dispositivo:
                dispositivo_devuelto = registro.id_dispositivo
                dispositivo_devuelto.id_empresa = None # Vuelve a ser propiedad interna
                dispositivo_devuelto.save()

            registro.save()
            
            
            horas_prestamo = duracion.total_seconds() / 3600
            messages.success(request, f'Devolución registrada. Duración: {horas_prestamo:.2f} horas.')

        except Registro.DoesNotExist:
            messages.error(request, 'El registro no es válido o ya ha sido devuelto.')

    
    return redirect('cenerisapp:lista_registros')

@login_required
def crear_inventario_lote(request):
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            
            nuevo_lote = form.save()
        
            messages.success(request, f"Lote de inventario #{nuevo_lote.id_inventario} creado. Ahora, por favor, añade los detalles de los componentes.")
    
            return redirect('cenerisapp:añadir_componentes_a_lote', lote_id=nuevo_lote.id_inventario)
    else:
        form = InventarioForm()

    context = { 'form': form, 'titulo': 'Registrar Nuevo Lote de Inventario' }
    return render(request, 'inventario_general/crear_lote.html', context)

@login_required
def añadir_componentes_a_lote(request, lote_id):
    lote = get_object_or_404(Inventario, pk=lote_id)
    componente_tipo = request.GET.get('tipo', 'sensor')
    SensorFormSet = formset_factory(SensorLoteForm, extra=lote.cantIngreso)
    OtroComponenteFormSet = formset_factory(OtroComponenteForm, extra=lote.cantIngreso)

    
    formset_sensores = SensorFormSet(prefix='sensor')
    formset_otros = OtroComponenteFormSet(prefix='otro')
 
    if request.method == 'POST':
        nombre_comun = lote.descripInv
 
        formset_sensores = SensorFormSet(request.POST, prefix='sensor')
        formset_otros = OtroComponenteFormSet(request.POST, prefix='otro')
 
 
        if 'guardar_sensores' in request.POST:
            formset_sensores = SensorFormSet(request.POST, prefix='sensor')
            if formset_sensores.is_valid():
                try: # --- INICIO DEL BLOQUE TRY ---
                    for form in formset_sensores:
                        if form.has_changed():
                            sensor = form.save(commit=False)
                            sensor.inventario = lote
                            sensor.nomComp = nombre_comun
                            sensor.estComp = 'Operativo'
                            sensor.save() 
                           
                    messages.success(request, "Sensores añadidos exitosamente.")
                    return redirect('cenerisapp:lista_lotes')
 
                except ValidationError as e: # --- AQUÍ ATRAPAMOS EL ERROR ---
                    
                    
                    for error in e.message_dict['__all__']:
                        messages.error(request, error)
                    
           
            else: # Si el formset no es válido desde el principio
                messages.error(request, "Por favor, corrige los errores en el formulario de sensores.")
 
        elif 'guardar_otros' in request.POST:
            formset_otros = OtroComponenteFormSet(request.POST, prefix='otro')
            if formset_otros.is_valid():
                for form in formset_otros:
                    if form.has_changed():
                        
                        otro_comp = form.save(commit=False)
                        otro_comp.inventario = lote
                        otro_comp.nomComp = nombre_comun
                        otro_comp.save()
                messages.success(request, "Componentes añadidos exitosamente.")
                return redirect('cenerisapp:lista_lotes')
            else:
                messages.error(request, "Por favor, corrige los errores en el formulario de otros componentes.")
               
    context = {
        'lote': lote,
        'formset_sensores': formset_sensores,
        'formset_otros': formset_otros,
        'componente_tipo_seleccionado': componente_tipo
    }
 
    return render(request, 'inventario_general/añadir_componentes.html', context)
 
@login_required
def lista_inventario(request):
    
    # --- 1. PREPARAR DATOS PARA LOS FILTROS ---
    opciones_ubicaciones = Inventario.objects.values_list('ubiImv', flat=True).distinct().order_by('ubiImv')
    opciones_trabajadores = Empleado.objects.filter(inventario__isnull=False).distinct().order_by('nomEmpleado')

    # --- 2. CAPTURAR VALORES DE FILTRO ---
    query_desc = request.GET.get('q', '')
    ubicacion_filtro = request.GET.get('ubicacion', '')
    trabajador_filtro = request.GET.get('trabajador', '')
    
    # --- 3. CONSTRUIR QUERYSET BASE Y FILTRAR ---
    lotes_qs = Inventario.objects.annotate(
        num_componentes=Count('componentes')
    )
    
    if query_desc:
        lotes_qs = lotes_qs.filter(descripInv__icontains=query_desc)
    if ubicacion_filtro:
        lotes_qs = lotes_qs.filter(ubiImv=ubicacion_filtro)
    if trabajador_filtro:
        lotes_qs = lotes_qs.filter(id_trabajador__pk=trabajador_filtro)

    # Ordenar al final
    lotes_qs = lotes_qs.order_by('-id_inventario')

    # --- 4. APLICAR PAGINACIÓN ---
    paginator = Paginator(lotes_qs, 9) # 9 lotes por página (se ve bien en un grid de 3 columnas)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # --- 5. CONTEXTO ---
    context = {
        'page_obj': page_obj,
        'titulo': 'Gestión de Lotes de Inventario',
        
        # Para los filtros
        'opciones_ubicaciones': opciones_ubicaciones,
        'opciones_trabajadores': opciones_trabajadores,
        'filtros_aplicados': {
            'q': query_desc,
            'ubicacion': ubicacion_filtro,
            'trabajador': int(trabajador_filtro) if trabajador_filtro else None,
        }
    }
    return render(request, 'inventario_general/lista_lotes.html', context)

@login_required
def vista_stock(request):
    
    stock_sensores = Sensor.objects.filter(dispositivo_instalado__isnull=True)\
                                   .values('nomComp')\
                                   .annotate(cantidad=Count('id_componente'))

    stock_otros = OtroComponente.objects.values('nomComp')\
                                        .annotate(cantidad=Count('id_componente'))
        
    stock_combinado = defaultdict(int)
    for item in stock_sensores:
        stock_combinado[item['nomComp']] += item['cantidad']
    for item in stock_otros:
        stock_combinado[item['nomComp']] += item['cantidad']

    stock_con_estado = []
    
    for nombre, cantidad in sorted(stock_combinado.items()):
        
        umbral_bajo = 5
        umbral_medio = 10

        estado = ''
        if cantidad < umbral_bajo:
            estado = 'rojo'
        elif cantidad < umbral_medio:
            estado = 'amarillo'
        else:
            estado = 'verde'
            
        stock_con_estado.append({
            'nomComp': nombre,
            'cantidad_total': cantidad,
            'estado_stock': estado
        })

    context = {
        'stock_items': stock_con_estado,
        'titulo': 'Stock de Componentes Disponibles'
    }
    return render(request, 'inventario_general/vista_stock.html', context)
 
@login_required
def lista_componentes(request, lote_id):
    lote = get_object_or_404(Inventario, pk=lote_id)
   
    
    componentes_del_lote = lote.componentes.all()
 
    context = {
        'lote': lote,
        'componentes': componentes_del_lote,
        'titulo': f'Componentes del Lote #{lote.id_inventario}'
    }
    return render(request, 'inventario_general/lista_componentes.html', context)
 

 
@login_required
def componentes_indice(request):
    """Muestra la página con las dos tarjetas de selección."""
    titulo = "Índice de Componentes"
    return render(request, 'componentes/indice.html', {'titulo': titulo})
 
 
@login_required
def lista_sensores(request):
    """Muestra una tabla con todos los datos de los sensores."""
    
    
    sensores = Sensor.objects.select_related('inventario').all().order_by('-id_componente')
   
    context = {
        'sensores': sensores,
        'titulo': 'Lista de Todos los Sensores'
    }
    return render(request, 'componentes/lista_sensores.html', context)
 
 
@login_required
def lista_otros_componentes(request):
    """Muestra una tabla con todos los datos de otros componentes."""
    otros_componentes = OtroComponente.objects.select_related('inventario').all().order_by('-id_componente')
   
    context = {
        'otros_componentes': otros_componentes,
        'titulo': 'Lista de Otros Componentes'
    }
    return render(request, 'componentes/lista_otros.html', context)
 
 
@login_required
def crear_dispositivo(request):
    if request.method == 'POST':
        form = DispositivoForm(request.POST)
        if form.is_valid():
            
            dispositivo = form.save(commit=False)
            
            dispositivo.fecIngreso = date.today()  # Fecha del día de hoy
            dispositivo.estadoD = 'Operativo'      # Estado por defecto
            
            dispositivo.save()
            
            cantidad = form.cleaned_data.get('cantidad_sensores') or 0
            messages.success(request, f"Dispositivo '{dispositivo.nomDisp}' creado exitosamente.")
            
            if cantidad > 0:
                return redirect('cenerisapp:asignar_sensores_a_dispositivo', dispositivo_id=dispositivo.id_dispositivo, cantidad=cantidad)
            else:
                return redirect('cenerisapp:lista_dispositivos')
    else:
        form = DispositivoForm()
 
    context = {
        'form': form,
        'titulo': 'Crear Nuevo Dispositivo'
    }
    return render(request, 'dispositivos/crear_dispositivo.html', context)
 
 
@login_required
def asignar_sensores_a_dispositivo(request, dispositivo_id, cantidad):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
   
    SensorFormSet = formset_factory(
        SensorParaDispositivoForm,
        formset=BaseSensorParaDispositivoFormSet, # Mantenemos la validación
        extra=cantidad # Solo usamos 'extra'
    )
 
    if request.method == 'POST':
        formset = SensorFormSet(request.POST)
        if formset.is_valid():
            try: # --- INICIO DEL BLOQUE TRY ---
                for form in formset:
                    if form.has_changed():
                        
                        sensor = form.save(commit=False)
                        sensor.dispositivo_instalado = dispositivo
            
                        sensor.save()
               
                messages.success(request, f"Se asignaron {len(formset)} sensores al dispositivo.")
                return redirect('cenerisapp:lista_dispositivos')

            except ValidationError as e: # --- ATRAPAMOS EL ERROR DEL MODELO ---
                
                if '__all__' in e.message_dict:
                    for error_message in e.message_dict['__all__']:
                        messages.error(request, error_message)
                
               
    else: # Petición GET
        formset = SensorFormSet()
 
    context = {
        'formset': formset,
        'dispositivo': dispositivo,
        'cantidad': cantidad
    }
    return render(request, 'dispositivos/asignar_sensores.html', context)

@login_required
def asignar_partes_a_dispositivo(request, dispositivo_id):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    
    
    formset_prefix = 'partes' 

    if request.method == 'POST':
        
        formset = ParteFormSet(request.POST, instance=dispositivo, prefix=formset_prefix)
        
        if formset.is_valid():
            formset.save()
            messages.success(request, f"Se han actualizado las partes para el dispositivo '{dispositivo.nomDisp}'.")
            return redirect('cenerisapp:lista_dispositivos')
        else:
            # Imprimimos los errores en la consola para depuración
            print("Errores del FormSet:", formset.errors)
            messages.error(request, "No se pudieron guardar los cambios. Por favor, revisa los errores en el formulario.")
    else:
        
        formset = ParteFormSet(instance=dispositivo, prefix=formset_prefix)

    context = {
        'dispositivo': dispositivo,
        'formset': formset,
        'titulo': f'Asignar/Editar Partes de {dispositivo.nomDisp}'
    }
    return render(request, 'dispositivos/asignar_partes.html', context)

@login_required
def configurar_lote_certificacion(request):


    if request.method == 'POST':
        form = CertificadoForm(request.POST)
        patrones_formset = PatronesFormSet(request.POST, prefix='patrones')
        resultados_formset = ResultadosFormSet(request.POST, prefix='resultados')

        form.fields['estadoFinal'].required = False
        form.fields['nro_certificado'].required = False
        if 'dispositivo' in form.fields:
            form.fields['dispositivo'].required = False

        if form.is_valid():
            
            # --- NUEVA LÓGICA DE LIMPIEZA Y SERIALIZACIÓN ---
            
            # 1. Limpiamos el formulario principal
            main_data = form.cleaned_data
            main_data.pop('estadoFinal', None)
            main_data.pop('nro_certificado', None)
            
            # Convertimos las fechas a string
            for key, value in main_data.items():
                if isinstance(value, date):
                    main_data[key] = value.isoformat()
            
            # 2. Limpiamos los formsets de manera explícita
            patrones_data_list = []
            for patron_form_data in patrones_formset.cleaned_data:
                # Nos aseguramos de saltar formularios vacíos o marcados para borrar
                if patron_form_data and not patron_form_data.get('DELETE'):
                    # Creamos un nuevo diccionario solo con los datos que queremos
                    clean_data = {
                        'patronUtil': patron_form_data.get('patronUtil'),
                        'n_p': patron_form_data.get('n_p'),
                        'n_lote': patron_form_data.get('n_lote'),
                        'n_certificado': patron_form_data.get('n_certificado'),
                        # Convertimos la fecha si existe
                        'fechaExpiracion': patron_form_data.get('fechaExpiracion').isoformat() if patron_form_data.get('fechaExpiracion') else None,
                    }
                    patrones_data_list.append(clean_data)

            resultados_data_list = []
            for resultado_form_data in resultados_formset.cleaned_data:
                if resultado_form_data and not resultado_form_data.get('DELETE'):
                    clean_data = {
                        'gas': resultado_form_data.get('gas'),
                        'lecturaPatron': resultado_form_data.get('lecturaPatron'),
                        'lecturaEquipo': resultado_form_data.get('lecturaEquipo'),
                        'prob_error': resultado_form_data.get('prob_error'),
                    }
                    resultados_data_list.append(clean_data)
            
            

            # 3. Construimos el diccionario final para la sesión
            lote_data = {
                'main': main_data,
                'patrones': patrones_data_list,
                'resultados': resultados_data_list,
            }
            
            request.session['lote_certificado_data'] = lote_data
            
            messages.success(request, "Paso 1 completado. Ahora seleccione los dispositivos para certificar.")
            
            # --- ¡CAMBIO CRUCIAL AQUÍ! ---
            # Redirigimos al Paso 2: la selección de dispositivos.
            return redirect('cenerisapp:seleccionar_dispositivos_lote') 

        else:
            messages.error(request, "Por favor, corrige los errores en el formulario para guardar los datos del lote.")

    else: # GET
        form = CertificadoForm()
        patrones_formset = PatronesFormSet(prefix='patrones')
        resultados_formset = ResultadosFormSet(prefix='resultados')

    context = {
        'form': form,
        'patrones_formset': patrones_formset,
        'resultados_formset': resultados_formset,
        'titulo': 'Configurar Datos para Lote de Certificación Diario',
        'modo_configuracion_lote': True,
    }
    return render(request, 'certificado/certificado_form.html', context)

@login_required
def limpiar_lote_certificacion(request):
    """
    Elimina los datos del lote de certificación de la sesión actual.
    """
    if 'lote_certificado_data' in request.session:
        del request.session['lote_certificado_data']
        messages.info(request, "Los datos del lote de certificación han sido limpiados.")
    return redirect(request.META.get('HTTP_REFERER', 'cenerisapp:lista_dispositivos'))

@login_required
def certificado_form(request, dispositivo_id=None, componente_id=None):
    dispositivo = None
    componente = None
    # Determinar si se está certificando un dispositivo completo o un componente específico
    if componente_id:
        componente = get_object_or_404(Componente, pk=componente_id)
        # Si es un componente, buscar su dispositivo asociado
        # Asumiendo que Sensor tiene un ForeignKey a Dispositivo y Componente es padre de Sensor
        if hasattr(componente, 'sensor') and componente.sensor.dispositivo_instalado:
            dispositivo = componente.sensor.dispositivo_instalado
        # Si tienes otros tipos de componentes que puedan estar en un dispositivo, añade lógica aquí
        
        if not dispositivo:
            messages.error(request, "Error: El componente seleccionado no está instalado en ningún dispositivo o el dispositivo no fue encontrado.")
            return redirect('cenerisapp:alguna_pagina_de_error_o_lista_de_componentes') # Ajusta tu URL de redirección
            
    elif dispositivo_id:
        dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
        # Para portátiles, el certificado se asocia al dispositivo completo, no a un componente individual en este contexto
        
    else:
        messages.error(request, "Error: No se proporcionó ID de dispositivo ni de componente.")
        return redirect('cenerisapp:alguna_pagina_de_error_o_lista_de_dispositivos') # Ajusta tu URL de redirección

    # Asegurarse de que tenemos un dispositivo para continuar
    if not dispositivo:
        messages.error(request, "Error: No se pudo determinar el dispositivo asociado para la certificación.")
        return redirect('cenerisapp:alguna_pagina_de_error_o_lista_general')

    empresa_asociada = dispositivo.id_empresa

    # Calcular el estado inicial previo
    estado_inicial_previo = "Primera Calibración"
    if componente: # Si estamos certificando un componente específico
        ultimo_certificado = Certificado.objects.filter(componente=componente).order_by('-fechCertificado', '-pk').first()
    else: # Si estamos certificando un dispositivo completo (portátil)
        ultimo_certificado = Certificado.objects.filter(dispositivo=dispositivo, componente__isnull=True).order_by('-fechCertificado', '-pk').first()

    if ultimo_certificado and ultimo_certificado.estadoFinal:
        estado_inicial_previo = ultimo_certificado.estadoFinal

    lote_data = request.session.get('lote_certificado_data', None)
    lote_activo = lote_data is not None


    if request.method == 'POST':
        form = CertificadoForm(request.POST)
        patrones_formset = PatronesFormSet(request.POST, prefix='patrones')
        resultados_formset = ResultadosFormSet(request.POST, prefix='resultados')

        if form.is_valid():
            lote_data = request.session.get('lote_certificado_data', None)
            
            # La validación condicional de los formsets
            are_formsets_valid = (not lote_data and 
                                  patrones_formset.is_valid() and 
                                  resultados_formset.is_valid())
            if lote_data or are_formsets_valid:
                certificado = form.save(commit=False)
                
                # Fusión de datos del lote
                if lote_data:
                    datos_main = lote_data.get('main', {})
                    certificado.temp = datos_main.get('temp')
                    certificado.presion = datos_main.get('presion')
                    certificado.humedadRelativa = datos_main.get('humedadRelativa')
                    certificado.proxFecha = datos_main.get('proxFecha')
                    certificado.rango_medicion = datos_main.get('rango_medicion')
                
                # Asignación de datos del dispositivo
                certificado.dispositivo = dispositivo
                if componente:
                    certificado.componente = componente 
                else:
                    certificado.componente = None 
                
                certificado.estado_inicial = estado_inicial_previo
                if empresa_asociada:
                    certificado.id_empresa = empresa_asociada
                else:
                    messages.error(request, "Error: El dispositivo no tiene una empresa asignada.")
                    return redirect('cenerisapp:lista_dispositivos')

                # Guardamos el certificado principal
                certificado.save()


                DatosPDF.objects.update_or_create(
                    certificado=certificado,
                    defaults={
                        'num_paginas_pdf': form.cleaned_data.get('num_paginas_pdf', 1),
                        'codigo_pdf': form.cleaned_data.get('codigo_pdf', ''),
                        'version_pdf': form.cleaned_data.get('version_pdf', ''),
                    }
                )
                
                # Procesamiento de formsets
                if lote_data:
                    # Modo Lote: Creamos objetos desde la sesión
                    for patron_data in lote_data.get('patrones', []):
                        if patron_data:
                            # --- CORRECCIÓN AQUÍ ---
                            # El campo en PatronesCalibracion se llama 'certificado'
                            PatronesCalibracion.objects.create(certificado=certificado, **patron_data)
                    
                    for resultado_data in lote_data.get('resultados', []):
                        if resultado_data:
                            # --- CORRECCIÓN AQUÍ ---
                            # El campo en Resultados se llama 'id_certificado'
                            Resultados.objects.create(id_certificado=certificado, **resultado_data)
                else:
                    # Modo Normal: Guardamos desde el POST
                    patrones_formset.instance = certificado
                    patrones_formset.save()
                    
                    resultados_formset.instance = certificado
                    resultados_formset.save()
                
                messages.success(request, f"Certificado N°{certificado.nro_certificado} creado exitosamente.")
                
                # Redirección a la lista de certificados
                if dispositivo:
                    return redirect('cenerisapp:lista_certificados_dispositivo', dispositivo_id=dispositivo.id_dispositivo)
                else:
                    return redirect('cenerisapp:lista_dispositivos')
            else:
                messages.error(request, "Por favor, corrige los errores en el formulario.")
                print("--- ERRORES DEL FORMULARIO PRINCIPAL ---")
                print(form.errors.as_json())
                print("\n--- ERRORES DEL FORMSET DE ANEXOS ---")
                print(formset.errors)
                print("Non-form errors:", formset.non_form_errors())

    else:
        
        if lote_activo:
            # Usamos los datos de la sesión como valores iniciales
            form = CertificadoForm(initial=lote_data.get('main'))
            patrones_formset = PatronesFormSet(prefix='patrones', initial=lote_data.get('patrones'))
            resultados_formset = ResultadosFormSet(prefix='resultados', initial=lote_data.get('resultados'))
        else:
            # Si no hay datos de lote, el formulario se carga vacío
            form = CertificadoForm()
            patrones_formset = PatronesFormSet(prefix='patrones')
            resultados_formset = ResultadosFormSet(prefix='resultados')
    
    context = {
        'form': form,
        'patrones_formset': patrones_formset,
        'resultados_formset': resultados_formset,
        'dispositivo': dispositivo,
        'componente': componente, # Puede ser None para portátiles
        'empresa': empresa_asociada,
        'titulo': f'Registrar Nuevo Certificado para: {dispositivo.nomDisp}' + (f' - {componente.nomComp}' if componente else ''),
        'estado_inicial_para_mostrar': estado_inicial_previo,
        'lote_activo': lote_activo,
    }
    return render(request, 'certificado/certificado_form.html', context)

@login_required
def seleccionar_dispositivos_lote(request):
    lote_data = request.session.get('lote_certificado_data')
    if not lote_data:
        messages.warning(request, "Primero debe configurar los datos del lote.")
        return redirect('configurar_lote_certificacion')

    AnexoFormSet = formset_factory(AnexoCertificadoForm, extra=1)
    if request.method == 'POST':
        seleccionados_str = request.POST.get('todos_los_seleccionados', '')
        dispositivos_ids = seleccionados_str.split(',') if seleccionados_str else []
        # Obtenemos la lista de IDs de los dispositivos SELECCIONADOS
        # dispositivos_ids = request.POST.getlist('dispositivos_seleccionados')
        anexos_paths_str = request.POST.get('anexos_paths', '')
        anexos_paths = anexos_paths_str.split(',') if anexos_paths_str else []
        
        programa_id = request.POST.get('programa')
        
        if not dispositivos_ids:
            messages.error(request, "No ha seleccionado ningún dispositivo.")
            return redirect('cenerisapp:seleccionar_dispositivos_lote')

        programa = get_object_or_404(Programa, pk=programa_id) if programa_id else None
            
        certificados_creados = []
        errores = []
        
        for dispositivo_id in dispositivos_ids:
            try:
                with transaction.atomic():
                        dispositivo = Dispositivo.objects.get(pk=dispositivo_id)
                            
                        # --- ¡NUEVA LÓGICA PARA DATOS INDIVIDUALES! ---
                        # Construimos el 'name' del input para este dispositivo
                        nro_certificado_name = f'nro_certificado_D{dispositivo_id}'
                        estado_final_name = f'estado_final_D{dispositivo_id}'
                        version_name = f'version_D{dispositivo_id}'
                            
                        # Leemos los valores del POST
                        nro_certificado_individual = request.POST.get(nro_certificado_name, '').strip()
                        estado_final_individual = request.POST.get(estado_final_name, 'Operativo') # 'Calibrado' por defecto
                        version_individual = request.POST.get(version_name, '01')

                        # Validación: el número de certificado es obligatorio
                        if not nro_certificado_individual:
                            errores.append(f"Falta el N° de Certificado para el dispositivo '{dispositivo.nomDisp}'.")
                            continue # Saltamos este dispositivo y continuamos con el siguiente

                        # Validación de duplicados
                        if Certificado.objects.filter(nro_certificado=nro_certificado_individual).exists():
                            errores.append(f"El N° de Certificado '{nro_certificado_individual}' ya existe en la base de datos.")
                            continue

                        # --- CÓDIGO COMPLETO PARA CREAR EL CERTIFICADO ---
                        certificado = Certificado(
                            dispositivo=dispositivo,
                            id_empresa=dispositivo.id_empresa,
                            id_programa=programa,
                                
                            # Datos comunes del lote (desde la sesión)
                            temp=lote_data['main'].get('temp'),
                            presion=lote_data['main'].get('presion'),
                            humedadRelativa=lote_data['main'].get('humedadRelativa'),
                            proxFecha=lote_data['main'].get('proxFecha'),
                            rango_medicion=lote_data['main'].get('rango_medicion'),
                                
                            # Datos individuales (leídos del POST)
                            nro_certificado=nro_certificado_individual,
                            estadoFinal=estado_final_individual,
                                
                            # Asumimos que el estado inicial se puede determinar o viene del lote
                            # Para este ejemplo, lo tomaremos del último certificado si existe.
                            estado_inicial=Certificado.objects.filter(dispositivo=dispositivo).order_by('-fechCertificado').first().estadoFinal if Certificado.objects.filter(dispositivo=dispositivo).exists() else "Primera Calibración",
                        )
                            
                        certificado.save()

                        from django.core.files import File
                        for path in anexos_paths:
                            if default_storage.exists(path):
                                with default_storage.open(path) as f:
                                    AnexoCertificado.objects.create(
                                        certificado=certificado,
                                        imagen=File(f, name=os.path.basename(path))
                                    )

                        DatosPDF.objects.create(
                            certificado=certificado,
                            version_pdf=version_individual # <-- Asignamos la versión individual
                        )

                        # Creamos sus Patrones y Resultados desde el lote
                        for patron_data in lote_data.get('patrones', []):
                            if patron_data: PatronesCalibracion.objects.create(certificado=certificado, **patron_data)
                        for resultado_data in lote_data.get('resultados', []):
                            if resultado_data: Resultados.objects.create(id_certificado=certificado, **resultado_data)
                            
                        certificados_creados.append(certificado.nro_certificado)
            
            except Exception as e:
                errores.append(f"Error inesperado con dispositivo ID {dispositivo_id}: {e}")
                print(f"--- TRACEBACK PARA ERROR EN DISPOSITIVO ID {dispositivo_id} ---")
                import traceback
                traceback.print_exc()
                print("---------------------------------------------------------")
        
        for path in anexos_paths:
            if default_storage.exists(path):
                default_storage.delete(path)
        
        if certificados_creados:
            messages.success(request, f"Se crearon exitosamente {len(certificados_creados)} certificados: {', '.join(certificados_creados)}.")
        if errores:
            messages.error(request, f"Ocurrieron errores: {'; '.join(errores)}")
        
        # Limpiamos la sesión del lote después de usarla
        request.session.pop('lote_certificado_data', None)
        return redirect('cenerisapp:lista_dispositivos')

    # Si la petición es GET, mostramos la lista de dispositivos
    else:
        query = request.GET.get('q', '')
        # Mostramos todos los dispositivos portátiles elegibles
        dispositivos_elegibles = Dispositivo.objects.filter(tipoDisp='Portatil')
        
        if query:
            # Filtramos por nombre del dispositivo O por número de serie
            dispositivos_elegibles = dispositivos_elegibles.filter(
                Q(nomDisp__icontains=query) |
                Q(num_serie__icontains=query)
            )
        programas_disponibles = Programa.objects.filter(totalEjecutado__lt=F('totalPrograma'))

        anexos_formset = AnexoFormSet(prefix='anexos')

        context = {
            'titulo': 'Paso 2: Seleccionar Dispositivos para el Lote',
            'dispositivos': dispositivos_elegibles,
            'lote_data': lote_data,
            'query': query, # Pasamos la query para que el buscador la recuerde
            'programas': programas_disponibles,
            'anexos_formset': anexos_formset,
        }
        return render(request, 'certificado/seleccionar_dispositivos_lote.html', context)

def generar_pdf_respuesta(request,certificado):
    """
    Toma un objeto Certificado y los datos limpios de un formulario,
    procesa toda la información y devuelve una respuesta HTTP con el PDF.
    """
    print("\n[PASO 1] Entrando en generar_pdf_respuesta...")
    patrones = certificado.patronescalibracion_set.all()
    resultados = certificado.resultados_set.all()
    dispositivo_asociado = certificado.dispositivo # Obtenemos el dispositivo directamente

    
    nombre_equipo = 'N/A'
    nombre_modelo = 'N/A'
    nombre_fabricante = 'N/A'
    nombre_area = 'N/A'
    tag_dispositivo = 'N/A'
    sensores_texto = 'N/A'
    ns_sensores_texto = 'N/A'
    es_portatil = False
    texto_area_completa = 'N/A'

    if dispositivo_asociado:
        es_portatil = dispositivo_asociado.tipoDisp == 'Portatil'
        nombre_fabricante = dispositivo_asociado.fabDisp
        area_texto = dispositivo_asociado.area_general or ""
        
        # Si tiene un área de trabajo fija asignada, la añadimos
        if dispositivo_asociado.id_areaTrabajo_fijo:
            # Añadimos un separador si ya teníamos un área general
            if area_texto:
                area_texto += " - "
            # Concatenamos el nombre del área específica
            area_texto += dispositivo_asociado.id_areaTrabajo_fijo.nombreA
        
        # Si después de todo, la cadena no está vacía, la usamos.
        if area_texto:
            texto_area_completa = area_texto

        tag_dispositivo = dispositivo_asociado.tag
        
        if es_portatil:
            nombre_equipo = 'Detector multigas'
            nombre_modelo = dispositivo_asociado.nomDisp # Para portátiles, el modelo es el nombre
            
            
            todos_sensores = dispositivo_asociado.sensor_set.all()
            if todos_sensores.exists():
                lista_de_gases = list(filter(None, [s.tipGas for s in todos_sensores]))
                sensores_texto = ", ".join(lista_de_gases) if lista_de_gases else 'No especificado'
                lista_de_series = [s.nSerieActual for s in todos_sensores]
                ns_sensores_texto = ", ".join(lista_de_series)
            else:
                sensores_texto = 'Sin sensores asignados'
                ns_sensores_texto = 'N/A'
        else: # Es Fijo
            nombre_equipo = 'Monitor estacionario'
            nombre_modelo = dispositivo_asociado.nomDisp
            
            # --- ¡LÓGICA CORREGIDA Y MEJORADA PARA FIJOS! ---
            
            # 1. Obtenemos el sensor específico desde el certificado que estamos procesando.
            sensor_calibrado = None
            if certificado.componente and hasattr(certificado.componente, 'sensor'):
                sensor_calibrado = certificado.componente.sensor
                
            if sensor_calibrado:
                # Si encontramos el sensor, llenamos sus datos básicos.
                sensores_texto = sensor_calibrado.tipGas
                
                # 2. Buscamos el último InformeCalibracion para ESE sensor.
                ultimo_informe = InformeCalibracion.objects.filter(sensor=sensor_calibrado).order_by('-fecha_informe').first()
                
                # 3. Asignamos el valor de 'encontrado_calibracion' a la variable de la plantilla.
                if ultimo_informe and ultimo_informe.encontrado_calibracion:
                    ns_sensores_texto = ultimo_informe.encontrado_calibracion
                else:
                    # Si no hay informe o el campo está vacío, usamos el N/S del sensor como un valor de respaldo (fallback).
                    self.stdout.write(self.style.WARNING(f"    - Advertencia: No se encontró 'encontrado_calibracion' para el sensor {sensor_calibrado}. Usando N/S como fallback."))
                    ns_sensores_texto = sensor_calibrado.nSerieActual or 'N/A'
            else:
                # Fallback si el certificado no está vinculado a un sensor.
                sensores_texto = 'Sensor no especificado'
                ns_sensores_texto = 'N/A'   
    
    estado_inicial_texto = certificado.estado_inicial or "Primera Calibración"

    
    try:
        # Usamos el 'related_name' que definimos en el modelo
        datos_pdf = certificado.datos_pdf 
    except DatosPDF.DoesNotExist:
        datos_pdf = None
    
    fecha_generacion = date.today()
    static_root = os.path.join(settings.BASE_DIR, 'cenerisapp', 'static')

    
    context_pdf = {
        'certificado': certificado,
        'patrones': patrones,
        'resultados': resultados,
        'logo_path': os.path.join(static_root, 'img', 'logo_ceneris.jpg'),
        'watermark_path': os.path.join(static_root, 'img', 'marca_de_agua.png'),
        'es_portatil': es_portatil,
        'nombre_equipo': nombre_equipo,
        'nombre_modelo': nombre_modelo,
        'nombre_fabricante': nombre_fabricante,
        'nombre_area': texto_area_completa,
        'tag_dispositivo': tag_dispositivo,
        'ns_dispositivo': dispositivo_asociado.num_serie if dispositivo_asociado else 'N/A',
        'sensores_texto': sensores_texto,
        'ns_sensores_texto': ns_sensores_texto,
        'estado_inicial_texto': estado_inicial_texto,
        'num_paginas_pdf': datos_pdf.num_paginas_pdf if datos_pdf else 'N/A',
        'codigo_pdf': datos_pdf.codigo_pdf if datos_pdf else 'N/A',
        'version_pdf': datos_pdf.version_pdf if datos_pdf else 'N/A',
        'fecha_generacion': fecha_generacion,
    }
    print("[PASO 2] Contexto del PDF preparado con éxito.")

    def s3_and_static_fetcher(url):
        # ---------------------------------------------
        # CASO 1: Archivos ESTÁTICOS (logos, css)
        # ---------------------------------------------
        # Estos SÍ están en el disco local del servidor en la carpeta 'staticfiles'
        if url.startswith(settings.STATIC_URL):
            path = url[len(settings.STATIC_URL):]
            absolute_path = finders.find(path)
            if absolute_path:
                print(f"✅ WeasyPrint (STATIC): Abriendo archivo local: {absolute_path}")
                return default_url_fetcher(f'file://{absolute_path}')
            else:
                print(f"❌ WeasyPrint ERROR (STATIC): No se encontró: {path}")
        
        # ---------------------------------------------
        # CASO 2: Archivos MULTIMEDIA (anexos de S3)
        # ---------------------------------------------
        # La URL completa de S3 o la relativa /media/
        elif settings.MEDIA_URL in url:
            # Extraemos la ruta del archivo dentro del bucket/media
            # ej. de '/media/anexos/img.png' -> 'anexos/img.png'
            path = url.split(settings.MEDIA_URL, 1)[-1]
            try:
                # 'default_storage' en producción es tu S3Boto3Storage.
                # .open() usa la API de boto3 para obtener un stream del archivo,
                # lo cual es mucho más rápido y fiable que una petición HTTP.
                print(f"✅ WeasyPrint (MEDIA): Abriendo desde S3 storage: {path}")
                file = default_storage.open(path)
                
                # Le pasamos el contenido en memoria a WeasyPrint
                return {'file_obj': file}
            
            except Exception as e:
                print(f"❌ WeasyPrint ERROR (MEDIA): No se pudo abrir desde S3: {path}. Error: {e}")

        # Para cualquier otra URL externa, usa la lógica por defecto (descarga HTTP)
        return default_url_fetcher(url)
    
    print("[PASO 3] Renderizando la plantilla HTML...")
    
    try:
        # --- Renderizado del HTML ---
        print("[PASO 4] Renderizando la plantilla HTML a string...")
        html_string = render_to_string('certificado/certificado_detalle.html', context_pdf, request=request)
        print("[PASO 5] Plantilla HTML renderizada con éxito.")
        
        # --- Creación del objeto WeasyPrint ---
        print("[PASO 6] Creando objeto HTML de WeasyPrint...")
        html = HTML(string=html_string, url_fetcher=s3_and_static_fetcher)
        print("[PASO 7] Objeto HTML de WeasyPrint creado.")

        # --- ¡EL PUNTO MÁS PROBABLE DE FALLO! ---
        print("[PASO 8] INICIANDO renderizado del PDF con html.write_pdf()... (Este puede tardar)")
        pdf_file = html.write_pdf()
        print("[PASO 9] ¡PDF renderizado con éxito en memoria!") # Si ves esto, el problema está después

        # --- Creación de la Respuesta HTTP ---
        print("[PASO 10] Creando respuesta HTTP...")
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Certificado-Nro-{certificado.nro_certificado}.pdf"'
        print("[PASO 11] Respuesta HTTP creada. Devolviendo al navegador.")
        return response
    except Exception as e:
        print(f"!!!!!!!!!!!!!!! ERROR INESPERADO DENTRO DE generar_pdf_respuesta !!!!!!!!!!!!!!!")
        print(f"Tipo de error: {type(e).__name__}")
        print(f"Mensaje: {e}")
        import traceback
        traceback.print_exc()
        # Devolvemos un error 500 explícito para que no haya timeout
        return HttpResponse("Ocurrió un error interno al generar el PDF.", status=500)


@login_required
def descargar_certificado(request, certificado_id):
    print("=============================================")
    print(f"INICIANDO DESCARGA DE PDF PARA CERTIFICADO ID: {certificado_id}")
    print("=============================================")
    certificado = get_object_or_404(Certificado, pk=certificado_id)
    
    # Pass the 'request' object as the first argument to the function
    return generar_pdf_respuesta(request, certificado)

@login_required
def lista_certificados_dispositivo(request, dispositivo_id): # <-- AÑADE el argumento aquí
    
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    
    certificados = Certificado.objects.filter(dispositivo=dispositivo).order_by('-fechCertificado')
    sensores_del_dispositivo = Sensor.objects.filter(dispositivo_instalado=dispositivo)
    context = {
        'dispositivo': dispositivo,
        'certificados': certificados,
        'titulo': f'Certificados para {dispositivo.nomDisp}',
        'sensores_del_dispositivo': sensores_del_dispositivo,
    }
    return render(request, 'certificado/lista_certificados_dispositivos.html', context)


def lista_certificados(request):
    certificados = Certificado.objects.all()
    return render(request, 'certificado/lista_certificados.html', {'certificados': certificados})

@login_required
def detalle_componente(request, id_componente):
    componente = get_object_or_404(Componente, pk=id_componente)
    context = {
        'componente': componente,
        'titulo': 'Detalle del Componente'
    }
    return render(request, 'componentes/detalle_componente.html', context)



def modificaciones_componente(request, id_componente):
    componente = Componente.objects.get(pk=id_componente)
    modificaciones = Modificacion.objects.filter(id_componente=componente)
    context = {
        'componente': componente,
        'modificaciones': modificaciones,
    }
    return render(request, 'modificaciones/modificaciones_componente.html', context)
@login_required
def seleccionar_dispositivo_alarma(request):
    """Página 1: El usuario selecciona un dispositivo."""
    dispositivos = Dispositivo.objects.all().order_by('nomDisp')
    context = {
        'dispositivos': dispositivos,
        'titulo': 'Seleccionar Dispositivo para Configurar Alarma'
    }
    return render(request, 'alarmas/seleccionar_dispositivo.html', context)

@login_required
def get_dispositivo_tipo(request, dispositivo_id):
    """Una mini-API para que JS sepa qué tipo de dispositivo es."""
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    return JsonResponse({'tipo': dispositivo.tipoDisp})


@login_required
def configurar_alarma(request, dispositivo_id):
    """Página 2: Muestra el formulario correcto y procesa los datos."""
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    
    if dispositivo.tipoDisp.upper() == 'FIJO':
        
        if request.method == 'POST':
            form = AlarmaFijoForm(request.POST)
            if form.is_valid():
                alarma = form.save(commit=False)
                alarma.id_dispositivo = dispositivo
                alarma.save()
                messages.success(request, f"Alarma configurada para el dispositivo fijo '{dispositivo.nomDisp}'.")
                return redirect('cenerisapp:seleccionar_dispositivo_alarma')
        else:
            form = AlarmaFijoForm()
        
        template_name = 'alarmas/configurar_alarma_fijo.html'

    elif dispositivo.tipoDisp.upper() == 'PORTATIL':
        
        if request.method == 'POST':
            
            form = AlarmaPortatilForm(request.POST, instance=dispositivo)
            if form.is_valid():
                form.save()
                messages.success(request, f"Alarma actualizada para el dispositivo portátil '{dispositivo.nomDisp}'.")
                return redirect('cenerisapp:seleccionar_dispositivo_alarma')
        else:
            form = AlarmaPortatilForm(instance=dispositivo)
        
        template_name = 'alarmas/configurar_alarma_portatil.html'

    else:
        messages.error(request, "Tipo de dispositivo desconocido.")
        return redirect('cenerisapp:seleccionar_dispositivo_alarma')

    context = {
        'form': form,
        'dispositivo': dispositivo
    }
    return render(request, template_name, context)



@login_required
def lista_reportes(request):
    reportes = Reporte.objects.select_related('id_dispositivo', 'id_trabajador').all().order_by('-fecReport')
    return render(request, 'reportes/lista_reportes.html', {'reportes': reportes})

@login_required
def editar_reporte(request, reporte_id):
    """
    Vista para editar un reporte existente.
    """
    
    
    reporte = get_object_or_404(Reporte, pk=reporte_id)
 
    if request.method == 'POST':
        
        
        form = ReporteForm(request.POST, instance=reporte)
       
        if form.is_valid():
            form.save() # Guarda los cambios en el objeto 'reporte'
            messages.success(request, f"Reporte #{reporte.id_reporte} actualizado exitosamente.")
            return redirect('cenerisapp:lista_reportes') # Redirigir a la lista de reportes
        else:
            
            messages.error(request, "Por favor, corrige los errores a continuación.")
 
    else: # Petición GET
        
        
        form = ReporteForm(instance=reporte)
 
    context = {
        'form': form,
        'reporte': reporte, # Pasamos el objeto para usarlo en el título, etc.
        'titulo': f'Editar Reporte #{reporte.id_reporte}'
    }
    return render(request, 'reportes/editar_reporte.html', context)

@login_required
def get_dispositivo_tipo(request, dispositivo_id):
    """Una mini-API para que JS sepa qué tipo de dispositivo es."""
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    return JsonResponse({'tipo': dispositivo.tipoDisp})


@login_required
def configurar_alarma(request, dispositivo_id):
    """Página 2: Muestra el formulario correcto y procesa los datos."""
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    
    if dispositivo.tipoDisp.upper() == 'FIJO':
        
        if request.method == 'POST':
            form = AlarmaFijoForm(request.POST)
            if form.is_valid():
                alarma = form.save(commit=False)
                alarma.id_dispositivo = dispositivo
                alarma.save()
                messages.success(request, f"Alarma configurada para el dispositivo fijo '{dispositivo.nomDisp}'.")
                return redirect('cenerisapp:seleccionar_dispositivo_alarma')
        else:
            form = AlarmaFijoForm()
        
        template_name = 'alarmas/configurar_alarma_fijo.html'

    elif dispositivo.tipoDisp.upper() == 'PORTATIL':
        
        if request.method == 'POST':
            
            form = AlarmaPortatilForm(request.POST, instance=dispositivo)
            if form.is_valid():
                form.save()
                messages.success(request, f"Alarma actualizada para el dispositivo portátil '{dispositivo.nomDisp}'.")
                return redirect('cenerisapp:seleccionar_dispositivo_alarma')
        else:
            form = AlarmaPortatilForm(instance=dispositivo)
        
        template_name = 'alarmas/configurar_alarma_portatil.html'

    else:
        messages.error(request, "Tipo de dispositivo desconocido.")
        return redirect('cenerisapp:seleccionar_dispositivo_alarma')

    context = {
        'form': form,
        'dispositivo': dispositivo
    }
    return render(request, template_name, context)

@login_required
def crear_reporte(request):
    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            reporte = form.save(commit=False)
            
            reporte.fecReport = date.today()
            if hasattr(request.user, 'empleado'):
                reporte.id_trabajador = request.user.empleado
            
            nuevo_estado_seleccionado = form.cleaned_data.get('nuevo_estado')
            
            # La lógica de actualización del estado se mantiene igual
            if reporte.id_dispositivo:
                reporte.id_dispositivo.estadoD = nuevo_estado_seleccionado
                
                # --- ¡CAMBIO CLAVE AQUÍ! ---
                # Le decimos a Django que solo guarde el campo 'estadoD'.
                reporte.id_dispositivo.save(update_fields=['estadoD']) 
                
                item_afectado = reporte.id_dispositivo
                
            elif reporte.id_otro_componente:
                reporte.id_otro_componente.estComp = nuevo_estado_seleccionado
                
                # --- ¡CAMBIO CLAVE AQUÍ! ---
                reporte.id_otro_componente.save(update_fields=['estComp'])
                
                item_afectado = reporte.id_otro_componente
            
            reporte.save()
            
            messages.success(request, f"Reporte guardado y estado de '{item_afectado}' actualizado a '{nuevo_estado_seleccionado}'.")
            return redirect('cenerisapp:lista_reportes')
        # Si el form no es válido, Django automáticamente pasará el form con los errores
        # a la plantilla, y ahora sí se mostrarán debajo del campo correcto.
        else: 
            print("========================================")
            print("El formulario no es válido. Errores:")
            print(form.errors)
            print("========================================")

            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        form = ReporteForm()

    context = {'form': form, 'titulo': 'Registrar Reporte de Daño o Pérdida'}
    return render(request, 'reportes/crear_reporte.html', context)

@login_required
def lista_reportes(request):
    
    # --- 1. CAPTURAR VALORES DE FILTRO ---
    dispositivo_q = request.GET.get('dispositivo', '')
    trabajador_q = request.GET.get('trabajador', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # --- 2. CONSTRUIR QUERYSET BASE Y FILTRAR ---
    reportes_qs = Reporte.objects.select_related(
        'id_dispositivo', 'id_otro_componente', 'id_trabajador'
    ).all()
    
    if dispositivo_q:
        reportes_qs = reportes_qs.filter(
            Q(id_dispositivo__nomDisp__icontains=dispositivo_q) |
            Q(id_dispositivo__num_serie__icontains=dispositivo_q) |
            Q(id_otro_componente__nomComp__icontains=dispositivo_q) |
            Q(id_otro_componente__nSerieActual__icontains=dispositivo_q)
        )

    if trabajador_q:
        reportes_qs = reportes_qs.filter(id_trabajador__nomEmpleado__icontains=trabajador_q)
        
    if fecha_desde:
        reportes_qs = reportes_qs.filter(fecReport__gte=fecha_desde)
    if fecha_hasta:
        reportes_qs = reportes_qs.filter(fecReport__lte=fecha_hasta)
        
    # Ordenar al final
    reportes_qs = reportes_qs.order_by('-fecReport')

    # --- 3. APLICAR PAGINACIÓN ---
    paginator = Paginator(reportes_qs, 10) # 10 reportes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # --- 4. CONTEXTO ---
    context = {
        'page_obj': page_obj,
        'titulo': 'Reportes de Daño o Pérdida',
        'filtros_aplicados': {
            'dispositivo': dispositivo_q,
            'trabajador': trabajador_q,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
    }
    return render(request, 'reportes/lista_reportes.html', context)

@login_required
def editar_reporte(request, reporte_id):
    
    reporte = get_object_or_404(Reporte, pk=reporte_id)

    if request.method == 'POST':
        
        
        form = ReporteForm(request.POST, instance=reporte)
        
        if form.is_valid():
            form.save() # Guarda los cambios en el objeto 'reporte'
            messages.success(request, f"Reporte #{reporte.id_reporte} actualizado exitosamente.")
            return redirect('cenerisapp:lista_reportes') # Redirigir a la lista de reportes
        else:
            
            messages.error(request, "Por favor, corrige los errores a continuación.")

    else: # Petición GET
        
        
        form = ReporteForm(instance=reporte)

    context = {
        'form': form,
        'reporte': reporte, # Pasamos el objeto para usarlo en el título, etc.
        'titulo': f'Editar Reporte #{reporte.id_reporte}'
    }
    return render(request, 'reportes/editar_reporte.html', context)

@login_required
def gestion_calibraciones(request): # Le cambiamos el nombre para que sea más claro
    
    
    dispositivos = Dispositivo.objects.prefetch_related('calibracion_set').all().order_by('nomDisp')
    
    

    context = {
        'dispositivos': dispositivos,
        'titulo': 'Gestión de Calibraciones'
    }
    return render(request, 'calibraciones/gestion_calibraciones.html', context)

@login_required
def registrar_calibracion_ahora(request, dispositivo_id):
    
    if request.method == 'POST':
        dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
        
        
        
        calibracion, created = Calibracion.objects.get_or_create(id_dispositivo=dispositivo)
        
        
        calibracion.estado = 'Calibrado'
        calibracion.fecCalibracionC = timezone.now().date() # Usamos la fecha actual
        
        
        if dispositivo.tipoDisp == 'Portatil':
            calibracion.prox_fecha = calibracion.fecCalibracionC + timedelta(days=1)
        elif dispositivo.tipoDisp == 'Fijo':
            calibracion.prox_fecha = calibracion.fecCalibracionC + timedelta(days=30)
        
        calibracion.save()
        
        messages.success(request, f"El dispositivo '{dispositivo.nomDisp}' ha sido calibrado.")
        
    
    return redirect('cenerisapp:gestion_calibraciones')

@login_required
def crear_venta(request):
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            componente_id = form.cleaned_data.get('id_componente')
            componente_obj = get_object_or_404(Componente, pk=componente_id)
            
            
            venta = form.save(commit=False)
            venta.id_componente = componente_obj
            venta.save()
            
            
            
            nombre_componente_vendido = componente_obj.nomComp # Guardamos el nombre para el mensaje
            componente_obj.delete()
            
            messages.success(request, f"Venta registrada y componente '{nombre_componente_vendido}' descontado del stock.")
            return redirect('cenerisapp:lista_ventas')
    else:
        form = VentaForm()

    context = {
        'form': form,
        'titulo': 'Registrar Nueva Venta'
    }
    return render(request, 'ventas/crear_venta.html', context)

@login_required
def completar_venta(request, venta_id):
    if request.method == 'POST':
        venta = get_object_or_404(Ventas, pk=venta_id)
        
        venta.estado = 'Completado' # O el estado que prefieras
        venta.save()
        messages.success(request, f"La venta #{venta.id_ventas} ha sido marcada como completada.")
    return redirect('cenerisapp:lista_ventas')

@login_required
def search_componentes_disponibles(request):
    query = request.GET.get('q', '')
    results = []
    if query:
        
        condicion_sensor = Q(
            sensor__isnull=False, 
            sensor__dispositivo_instalado__isnull=True,
            sensor__estComp='Operativo' # <-- ¡ESTE FILTRO ES CRUCIAL!
        )
        
        condicion_otro = Q(
            otrocomponente__isnull=False,
            otrocomponente__estComp='Operativo' # <-- ¡ESTE FILTRO ES CRUCIAL!
        )
        
        componentes_disponibles = Componente.objects.filter(
            Q(nomComp__icontains=query) & (condicion_sensor | condicion_otro)
        ).select_related('sensor', 'otrocomponente').distinct()

        results = [{ 'id': comp.pk, 'text': f"{comp.nomComp} (N/S: {comp.nSerieActual})" } 
                   for comp in componentes_disponibles[:10]]

    return JsonResponse(results, safe=False)

@login_required
def lista_ventas(request):
    ventas = Ventas.objects.select_related('id_componente').all().order_by('-fecVenta')
    context = {
        'ventas': ventas,
        'titulo': 'Historial de Ventas'
    }
    return render(request, 'ventas/lista_ventas.html', context)

@login_required
def lista_modificaciones(request):
    
    # --- 1. CAPTURAR VALORES DE FILTRO DE LA URL ---
    dispositivo_query = request.GET.get('dispositivo', '')
    trabajador_query = request.GET.get('trabajador', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo_servicio = request.GET.get('tipo', '')

    # --- 2. CONSTRUIR EL QUERYSET BASE Y APLICAR FILTROS ---
    modificaciones = Modificacion.objects.select_related(
        'id_dispositivo', 'sensor_saliente', 'parte_saliente', 
        'componente_entrante', 'id_trabajador'
    ).all() # Empezamos con .all() y aplicamos filtros

    # Filtro por Dispositivo (busca por nombre o N/S)
    if dispositivo_query:
        modificaciones = modificaciones.filter(
            Q(id_dispositivo__nomDisp__icontains=dispositivo_query) |
            Q(id_dispositivo__num_serie__icontains=dispositivo_query)
        )

    # Filtro por Trabajador (busca por nombre)
    if trabajador_query:
        modificaciones = modificaciones.filter(id_trabajador__nomEmpleado__icontains=trabajador_query)
        
    # Filtro por Tipo de Servicio
    if tipo_servicio:
        modificaciones = modificaciones.filter(tipoServicio=tipo_servicio)

    # Filtro por Rango de Fechas
    if fecha_desde:
        modificaciones = modificaciones.filter(fecInstalacionMod__gte=fecha_desde)
    if fecha_hasta:
        modificaciones = modificaciones.filter(fecInstalacionMod__lte=fecha_hasta)

    # Aplicamos el ordenamiento al final
    modificaciones = modificaciones.order_by('-fecInstalacionMod')

    paginator = Paginator(modificaciones, 15) # 15 modificaciones por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # --- 3. PREPARAR DATOS PARA LOS DESPLEGABLES DE FILTRO ---
    opciones_tipo_servicio = Modificacion.objects.values_list('tipoServicio', flat=True).distinct()
    
    # --- 4. CONSTRUIR EL CONTEXTO FINAL ---
    context = {
        'page_obj': page_obj, 
        'titulo': 'Historial de Modificaciones y Servicios',
        'opciones_tipo_servicio': opciones_tipo_servicio,
        'filtros_aplicados': {
            'dispositivo': dispositivo_query,
            'trabajador': trabajador_query,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'tipo': tipo_servicio,
        }
    }
    return render(request, 'modificaciones/lista_modificaciones.html', context)

@login_required
def crear_modificacion(request):
    opciones_salientes = []

    if request.method == 'POST':
        dispositivo_id = request.POST.get('id_dispositivo')
        if dispositivo_id:
            try:
                # Se necesita Dispositivo.objects.get() para llenar opciones_salientes
                dispositivo = Dispositivo.objects.get(pk=dispositivo_id)
                for parte in dispositivo.partes.all(): 
                    opciones_salientes.append((f'parte_{parte.pk}', f"Parte: {parte.nomPart}"))
                for sensor in dispositivo.sensor_set.all():
                    opciones_salientes.append((f'sensor_{sensor.pk}', f"Sensor: {sensor.nomComp}"))
            except Dispositivo.DoesNotExist:
                pass

        form = ModificacionForm(request.POST, opciones_salientes=opciones_salientes)
        
        if form.is_valid():
            modificacion = form.save(commit=False)
            
            dispositivo = modificacion.id_dispositivo
            item_saliente_str = form.cleaned_data.get('item_saliente')
            tipo_saliente, pk_saliente = item_saliente_str.split('_')
            modificacion.fecInstalacionMod = date.today()
            
            # Recuperamos los nuevos campos del formulario
            pk_entrante = form.cleaned_data.get('reemplazo_id')
            n_serie_nuevo = form.cleaned_data.get('n_serie_reemplazo')

            componente_entrante = None
            if pk_entrante:
                componente_entrante = get_object_or_404(Componente, pk=pk_entrante)
                
            # --- LÓGICA DE ACTUALIZACIÓN DEL N/S ---
            if componente_entrante and n_serie_nuevo:
                # Asignamos el número de serie al componente seleccionado del stock
                componente_entrante.nSerieActual = n_serie_nuevo
                componente_entrante.save() # Guardamos el cambio en el componente

            # --- Lógica para ITEM SALIENTE (Parte o Sensor) ---
            if tipo_saliente == 'parte':
                parte_afectada = get_object_or_404(Parte, pk=pk_saliente)
                modificacion.parte_saliente = parte_afectada
                
                # Guardamos la modificación ANTES de consumir el componente
                modificacion.save()
                
                # Si se usó un componente del stock, se consume (descuenta)
                if componente_entrante:
                    nombre_componente = componente_entrante.nomComp
                    componente_entrante.delete() # Se descuenta del stock
                    messages.success(request, f"Servicio registrado para la parte '{parte_afectada.nomPart}'. El componente '{nombre_componente}' con N/S '{n_serie_nuevo}' fue descontado del stock.")
                else:
                    messages.success(request, f"Servicio registrado para la parte '{parte_afectada.nomPart}'.")
                    
            elif tipo_saliente == 'sensor':
                sensor_saliente = get_object_or_404(Sensor, pk=pk_saliente)
                modificacion.sensor_saliente = sensor_saliente
                
                if componente_entrante and hasattr(componente_entrante, 'sensor'):
                    # El 'componente_entrante' ahora ya tiene su N/S guardado
                    
                    # 1. Marcar el sensor saliente como retirado/inoperativo
                    sensor_saliente.dispositivo_instalado = None
                    sensor_saliente.estComp = 'Inoperativo por cambio'
                    sensor_saliente.save()
                    
                    # 2. Instalar el sensor entrante (Componente -> Sensor)
                    sensor_entrante = componente_entrante.sensor
                    sensor_entrante.dispositivo_instalado = dispositivo
                    sensor_entrante.fecInst = date.today()
                    sensor_entrante.save()
                    
                    # 3. Registrar el sensor entrante en la modificación
                    modificacion.componente_entrante = sensor_entrante

                    messages.success(request, f"Reparación registrada. El sensor '{sensor_saliente.nomComp}' fue reemplazado por N/S '{sensor_entrante.nSerieActual}'.")
                else:
                    messages.success(request, f"Servicio registrado para el sensor '{sensor_saliente.nomComp}'.")
                
                # Guardar la modificación final
                modificacion.save()
                
                # Mensaje final y redirección con enlace de foto
                url_fotos = reverse('cenerisapp:gestionar_fotos_dispositivo', args=[dispositivo.id_dispositivo])
                mensaje = format_html(
                    "Reparación registrada exitosamente. <strong>¡No olvides subir la foto de evidencia!</strong> <a href='{}' class='alert-link'>Cargar foto ahora</a>.",
                    url_fotos
                )
                messages.success(request, mensaje)
                
            return redirect('cenerisapp:lista_modificaciones')
        else:
            # Reutiliza las opciones salientes si el formulario falla, para que no se pierdan
            # (Aunque ModificacionForm debería manejar esto si se pasa en la inicialización)
            messages.error(request, "Por favor, corrige los errores en el formulario.")
            
    else:
        form = ModificacionForm()

    context = {
        'form': form,
        'titulo': 'Registrar Nueva Reparación / Servicio'
    }
    return render(request, 'modificaciones/crear_modificacion.html', context)

@login_required
def editar_modificacion(request, modificacion_id):
    print("\n" + "="*50)
    print(f"INICIO VISTA 'editar_modificacion' - MÉTODO: {request.method}")
    print("="*50)

    modificacion = get_object_or_404(Modificacion, pk=modificacion_id)
    print(f"[PASO 1] Objeto a editar cargado: Modificacion #{modificacion.id_modificacion}")

    opciones = []
    dispositivo = modificacion.id_dispositivo
    
    if dispositivo:
        print(f"[PASO 2] Dispositivo asociado: '{dispositivo}' (ID: {dispositivo.pk})")
        sensores = Sensor.objects.filter(dispositivo_instalado=dispositivo)
        partes = Parte.objects.filter(id_dispositivo=dispositivo)
        
        for s in sensores:
            opciones.append((f'sensor_{s.pk}', f"Sensor: {s.nomComp} ({s.nSerieActual})"))
        for p in partes:
            opciones.append((f'parte_{p.id_parte}', f"Parte: {p.nomPart}"))
        
        print(f"[PASO 3] Lista de 'opciones' generada: {opciones}")
    else:
        print("[PASO 2] ADVERTENCIA: La modificación no tiene un dispositivo asociado.")

    if request.method == 'POST':
        print("\n--- INICIO PROCESO POST ---")
        
        print("[PASO 4] Instanciando el formulario con los datos del POST.")
        form = ModificacionForm(request.POST, instance=modificacion, opciones_involucrados=opciones)
        
        print("[PASO 5] Verificando si el formulario es válido (form.is_valid())...")
        is_valid = form.is_valid()

        if is_valid:
            print("\n  /------------------------------------\\")
            print("  |   ¡EL FORMULARIO ES VÁLIDO!    |")
            print("  \\------------------------------------/")
            form.save()
            messages.success(request, f"Modificación #{modificacion.id_modificacion} actualizada exitosamente.")
            return redirect('cenerisapp:lista_modificaciones')
        else:
            print("\n  /--------------------------------------\\")
            print("  |   ¡EL FORMULARIO NO ES VÁLIDO!     |")
            print("  \\--------------------------------------/")
            print("[PASO 6] Errores del formulario:")
            
            print(form.errors.as_json())
            messages.error(request, "Por favor, corrige los errores en el formulario.")

    else: # Petición GET
        print("\n--- INICIO PROCESO GET ---")
        form = ModificacionForm(instance=modificacion, opciones_involucrados=opciones)
        
        if modificacion.id_sensor:
            initial_value = f'sensor_{modificacion.id_sensor.pk}'
            form.fields['componente_o_parte_involucrada'].initial = initial_value
            print(f"[PASO 4 GET] Pre-seleccionando valor inicial: {initial_value}")
        elif modificacion.id_parte:
            initial_value = f'parte_{modificacion.id_parte.pk}'
            form.fields['componente_o_parte_involucrada'].initial = initial_value
            print(f"[PASO 4 GET] Pre-seleccionando valor inicial: {initial_value}")

    context = {
        'form': form,
        'modificacion': modificacion,
        'titulo': f'Editar Modificación #{modificacion.id_modificacion}'
    }
    print("--- FIN DE LA VISTA. RENDERIZANDO PLANTILLA ---")
    return render(request, 'modificaciones/editar_modificacion.html', context)

@login_required
def get_tipos_componentes(request):
    """
    Devuelve una lista de todos los NOMBRES de componentes únicos
    que existen en el stock y están disponibles.
    """
    # Buscamos sensores disponibles
    nombres_sensores_qs = Sensor.objects.filter(
        dispositivo_instalado__isnull=True,
        estComp='Operativo',
        nSerieActual__isnull=True # ¡Importante! Solo los que no tienen N/S
    ).values_list('nomComp', flat=True)
    
    # Buscamos otros componentes disponibles
    nombres_otros_qs = OtroComponente.objects.filter(
        estComp='Operativo',
        nSerieActual__isnull=True # ¡Importante! Solo los que no tienen N/S
    ).values_list('nomComp', flat=True)
    
    # Unimos los resultados y eliminamos duplicados
    tipos_unicos_set = set(nombres_sensores_qs) | set(nombres_otros_qs)
    
    # Convertimos a lista y ordenamos
    tipos_unicos = sorted(list(tipos_unicos_set))
    
    return JsonResponse(tipos_unicos, safe=False)

@login_required
def get_componentes_sin_ns(request):
    """
    Busca en el stock componentes de un tipo específico que NO tengan un número de serie asignado.
    """
    tipo = request.GET.get('tipo', None)
    if not tipo:
        return JsonResponse([], safe=False)

    componentes = Componente.objects.filter(
        nomComp=tipo,
        nSerieActual__isnull=True
    ).select_related('inventario')

    data = [
        {
            'id': comp.id_componente,
            # --- LÍNEA CORREGIDA ---
            'text': f"ID: {comp.id_componente} (Lote: {comp.inventario.descripInv if comp.inventario else 'N/A'})"
        } 
        for comp in componentes
    ]
    return JsonResponse(data, safe=False)


@login_required
def get_ns_por_tipo_api(request):
    """
    Dado un nombre de componente, devuelve los N/S disponibles y operativos.
    """
    nombre_componente = request.GET.get('tipo', '')
    results = []
    
    if nombre_componente:
        
        

        
        sensores_disponibles = Sensor.objects.filter(
            nomComp=nombre_componente,
            dispositivo_instalado__isnull=True,
            estComp='Operativo'
        )
        
        
        otros_disponibles = OtroComponente.objects.filter(
            nomComp=nombre_componente,
            estComp='Operativo'
        )

        
        
        componentes_combinados = list(sensores_disponibles) + list(otros_disponibles)
        
        
        componentes_combinados.sort(key=lambda x: x.nSerieActual)

        
        results = [{
            'id': c.pk, 
            'text': f"{c.nSerieActual} (ID: {c.pk})" # Añadir el ID puede ayudar a depurar
        } for c in componentes_combinados]
        
    return JsonResponse(results, safe=False)

@login_required # <-- ¡AÑADE ESTO! Es crucial para la seguridad y para evitar errores.
def get_partes_y_sensores_por_dispositivo(request):
    dispositivo_id = request.GET.get('dispositivo_id')
    opciones = []
    
    if not dispositivo_id:
        # Devuelve un JSON de error si no se proporciona el ID
        return JsonResponse({'error': 'No se proporcionó dispositivo_id'}, status=400)

    try:
        # Verificamos que el dispositivo exista
        dispositivo = Dispositivo.objects.get(pk=dispositivo_id)
        
        partes = Parte.objects.filter(id_dispositivo=dispositivo) # Es más limpio pasar el objeto
        for p in partes:
            opciones.append({'id': f'parte_{p.id_parte}', 'nombre': f"Parte: {p.nomPart}"})

        sensores = Sensor.objects.filter(dispositivo_instalado=dispositivo) # Igual aquí
        for s in sensores:
            opciones.append({'id': f'sensor_{s.pk}', 'nombre': f"Sensor: {s.nomComp} ({s.nSerieActual})"})
            
    except Dispositivo.DoesNotExist:
        # Devuelve un JSON de error si el dispositivo no existe
        return JsonResponse({'error': 'Dispositivo no encontrado'}, status=404)

    return JsonResponse(opciones, safe=False)

@login_required
def gestionar_fotos_dispositivo(request, dispositivo_id):
    
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    # Traemos las modificaciones para el desplegable
    modificaciones_del_dispositivo = dispositivo.modificacion_set.order_by('-fecInstalacionMod')

    if request.method == 'POST':
        form = FotoDispositivoForm(request.POST, request.FILES, modificaciones_queryset=modificaciones_del_dispositivo)
        
        if form.is_valid():
            foto = form.save(commit=False)
            foto.dispositivo = dispositivo

            modificacion_asociada = form.cleaned_data.get('modificacion')
            
            if modificacion_asociada:
                # =========================================================
                # LÓGICA AUTOMÁTICA MEJORADA (SENSORES Y PARTES)
                # =========================================================
                foto.contexto = 'CARDEX' # Siempre es Cardex si hay modificación
                
                # CASO 1: Es un cambio de SENSOR (Componente entrante es Sensor)
                if modificacion_asociada.componente_entrante and hasattr(modificacion_asociada.componente_entrante, 'sensor'):
                    foto.tipo_foto = modificacion_asociada.componente_entrante.sensor.tipGas
                
                # CASO 2: Es un cambio de SENSOR (Salió un sensor, aunque no haya entrado nada)
                elif modificacion_asociada.sensor_saliente:
                    foto.tipo_foto = modificacion_asociada.sensor_saliente.tipGas

                # CASO 3: Es un cambio de PARTE / KIT (Entró un componente genérico)
                # Aquí capturamos el nombre: "Carcasa", "Batería", etc.
                elif modificacion_asociada.componente_entrante:
                    foto.tipo_foto = modificacion_asociada.componente_entrante.nomComp
                
                # CASO 4: Es un retiro de PARTE (Sin reemplazo)
                elif modificacion_asociada.parte_saliente:
                    foto.tipo_foto = modificacion_asociada.parte_saliente.nomPart
                
                else:
                    # Fallback por seguridad
                    foto.tipo_foto = 'MANTENIMIENTO'

            # =========================================================
            # LÓGICA PARA FOTOS GENERALES (Sin modificación asociada)
            # =========================================================
            else:
                tipo_foto_nuevo = form.cleaned_data.get('tipo_foto')
                contexto_nuevo = form.cleaned_data.get('contexto')
                
                # Si el usuario sube una nueva foto "General" del mismo tipo (ej. EVIDENCIA),
                # borramos la anterior para no acumular basura.
                foto_existente = FotoDispositivo.objects.filter(
                    dispositivo=dispositivo, 
                    tipo_foto=tipo_foto_nuevo,
                    contexto=contexto_nuevo,
                    modificacion__isnull=True
                ).first()

                if foto_existente:
                    foto_existente.delete()
                    messages.info(request, f"Se ha reemplazado la imagen general anterior para '{tipo_foto_nuevo}' en el contexto '{contexto_nuevo}'.")
            
            # Guardamos la foto finalmente
            foto.save()
            messages.success(request, f"Imagen cargada exitosamente como '{foto.tipo_foto}'.")
            return redirect('cenerisapp:gestionar_fotos_dispositivo', dispositivo_id=dispositivo.id_dispositivo)
        else:
            messages.error(request, "Error al cargar la imagen. Por favor, revisa el formulario.")
            
    else: # GET
        form = FotoDispositivoForm(modificaciones_queryset=modificaciones_del_dispositivo)

    # Listar fotos para la galería
    fotos_list = dispositivo.fotos.all().order_by('-fecha_carga')
    
    paginator = Paginator(fotos_list, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Lista de sugerencias para el frontend
    tipos_validos = [s.tipGas for s in dispositivo.sensor_set.all() if s.tipGas] + ["EVIDENCIA"]
    
    context = {
        'form': form,
        'dispositivo': dispositivo,
        'fotos': page_obj,
        'tipos_validos_para_fotos': tipos_validos,
        'titulo': f"Gestionar Fotos para {dispositivo.nomDisp}"
    }
    return render(request, 'dispositivos/gestionar_fotos.html', context)

@login_required
def tecnicos_indice(request):
    titulo = "Datos Técnicos"
    return render(request, 'tecnicos/indice.html', {'titulo': titulo})

@login_required
def lista_empresas(request):
    empresas = Empresa.objects.all().order_by('nombreE')
    context = {
        'empresas': empresas,
        'titulo': 'Lista de Todas las Empresas'
    }
    return render(request, 'tecnicos/lista_empresas.html', context)

@login_required
def lista_areas(request):
    areas = AreaTrabajo.objects.all().order_by('nombreA')
    context = {
        'areas': areas,
        'titulo': 'Lista de Todas las Áreas de Trabajo' 
    }
    return render(request, 'tecnicos/lista_areas.html', context)

@login_required
def lista_empleados(request):
    empleados = Empleado.objects.select_related('areaTrabajo', 'supervisor').all().order_by('nomEmpleado')
    context = {
        'empleados': empleados,
        'titulo': 'Lista de Todos los Empleados'
    }
    return render(request, 'tecnicos/lista_empleados.html', context)

@login_required
def crear_empleado(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        
        
        correo_formset = CorreoFormSet(request.POST, prefix='correo')
        telefono_formset = TelefonoFormSet(request.POST, prefix='telefono')

        if form.is_valid() and correo_formset.is_valid() and telefono_formset.is_valid():
            
            empleado = form.save()
            
            correos = correo_formset.save(commit=False)
            for correo in correos:
                correo.empleado = empleado
                correo.save()
            
            telefonos = telefono_formset.save(commit=False)
            for telefono in telefonos:
                telefono.empleado = empleado
                telefono.save()

            messages.success(request, f"Empleado '{empleado.nomEmpleado}' creado exitosamente.")
            return redirect('cenerisapp:lista_empleados')
    else:
        form = EmpleadoForm()
        correo_formset = CorreoFormSet(prefix='correo')
        telefono_formset = TelefonoFormSet(prefix='telefono')

    context = {
        'form': form,
        'correo_formset': correo_formset,
        'telefono_formset': telefono_formset,
        'titulo': 'Registrar Nuevo Empleado'
    }
    return render(request, 'tecnicos/crear_empleado.html', context)



@login_required
def gestionar_seguimiento_diario(request):

    anos_disponibles = Programa.objects.values_list('ano', flat=True).distinct().order_by('-ano')
    areas_generales_disponibles = Dispositivo.objects.filter(
        tipoDisp__iexact="Portatil" # Asumiendo que 'tipo_dispositivo' es el campo para "Portatil"
    ).values_list('area_general', flat=True).distinct().order_by('area_general')
    
    meses_disponibles = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]
    
    today = date.today()
    try:
        ano_seleccionado = int(request.GET.get('ano', today.year))
        mes_seleccionado = int(request.GET.get('mes', today.month))
    except (ValueError, TypeError):
        ano_seleccionado = today.year
        mes_seleccionado = today.month
        
    area_general_seleccionada = request.GET.get('area_general')

    # 2. PROCESAR EL GUARDADO (SI ES POST)
    if request.method == 'POST':
        print("\n--- INICIO PROCESO POST (Guardar Seguimiento) ---")
    
        # --- 1. OBTENEMOS EL ESTADO "ANTES" DEL CAMBIO ---
        # Reconstruimos la matriz de datos tal como estaba antes del envío.
        seguimientos_previos = SeguimientoDiario.objects.filter(
            dispositivo__area_general=area_general_seleccionada,
            fecha__year=ano_seleccionado,
            fecha__month=mes_seleccionado
        )
        matriz_previa = {}
        for s in seguimientos_previos:
            if s.dispositivo_id not in matriz_previa:
                matriz_previa[s.dispositivo_id] = {}
            matriz_previa[s.dispositivo_id][s.fecha] = s.estado_texto

        items_guardados = 0
        items_actualizados = 0
        items_borrados = 0

        # Iteramos sobre los datos enviados en el POST
        for key, new_value in request.POST.items():
            if key.startswith('estado_D'):
                try:
                    parts = key.split('_')
                    dispositivo_id = int(parts[1].replace('D', ''))
                    fecha_str = parts[2].replace('F', '')
                    fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()

                    # --- 2. COMPARAMOS EL ESTADO "NUEVO" CON EL "ANTIGUO" ---
                    old_value = matriz_previa.get(dispositivo_id, {}).get(fecha_obj, '')
                    
                    # Solo actuamos si el valor ha cambiado
                    if new_value != old_value:
                        if new_value:
                            # Si el nuevo valor NO está vacío, creamos o actualizamos.
                            obj, created = SeguimientoDiario.objects.update_or_create(
                                dispositivo_id=dispositivo_id,
                                fecha=fecha_obj,
                                defaults={'estado_texto': new_value}
                            )
                            if created: items_guardados += 1
                            else: items_actualizados += 1
                        else:
                            # Si el nuevo valor ESTÁ vacío, significa que el usuario lo borró.
                            SeguimientoDiario.objects.filter(dispositivo_id=dispositivo_id, fecha=fecha_obj).delete()
                            items_borrados += 1
                            
                except (ValueError, IndexError):
                    continue
        
        messages.success(request, f"Seguimiento guardado. Creados: {items_guardados}, Actualizados: {items_actualizados}, Borrados: {items_borrados}.")
        return redirect(f"{request.path}?ano={ano_seleccionado}&mes={mes_seleccionado}&area_general={area_general_seleccionada or ''}")

    historial_qs = SeguimientoDiario.objects.exclude(
        estado_texto__isnull=True
    ).exclude(
        estado_texto__exact=''
    ).select_related('dispositivo').order_by('-fecha', '-id_seguimiento')
    
    # Capturar filtros para el historial
    historial_q = request.GET.get('historial_q', '')
    historial_fecha_desde = request.GET.get('historial_fecha_desde', '')
    historial_fecha_hasta = request.GET.get('historial_fecha_hasta', '')
    
    # Aplicar filtros al historial
    if historial_q:
        historial_qs = historial_qs.filter(dispositivo__nomDisp__icontains=historial_q)
    if historial_fecha_desde:
        historial_qs = historial_qs.filter(fecha__gte=historial_fecha_desde)
    if historial_fecha_hasta:
        historial_qs = historial_qs.filter(fecha__lte=historial_fecha_hasta)

    # Paginación para el historial
    historial_paginator = Paginator(historial_qs, 10)
    page_number_historial = request.GET.get('page_historial') # <-- Usamos un param diferente
    historial_page_obj = historial_paginator.get_page(page_number_historial)


    # 3. PREPARAR DATOS PARA LA PLANTILLA (PETICIÓN GET)
    context = {
        'titulo': 'Gestión de Seguimiento Diario',
        # Pasamos las listas de opciones a la plantilla
        'anos_disponibles': anos_disponibles,
        'meses_disponibles': meses_disponibles,
        'areas_generales_disponibles': [area for area in areas_generales_disponibles if area],
        # Pasamos los valores seleccionados para que los <select> los recuerden
        'ano_seleccionado': ano_seleccionado,
        'mes_seleccionado': mes_seleccionado,
        'area_general_seleccionada': area_general_seleccionada,

        'historial_page_obj': historial_page_obj,
        'filtros_historial_aplicados': {
            'q': historial_q,
            'fecha_desde': historial_fecha_desde,
            'fecha_hasta': historial_fecha_hasta,
        }
    }
    
    # Solo construimos la matriz si el usuario ha seleccionado un área
    if area_general_seleccionada:
        num_dias = calendar.monthrange(ano_seleccionado, mes_seleccionado)[1]
        dias_del_mes = [date(ano_seleccionado, mes_seleccionado, dia) for dia in range(1, num_dias + 1)]
        
        dispositivos_qs = Dispositivo.objects.filter(
            area_general=area_general_seleccionada,
            tipoDisp='Portatil'
        ).order_by('nomDisp') # Es bueno tener un orden consistente

        # 2. Creamos el paginador
        matriz_paginator = Paginator(dispositivos_qs, 15) # Nuevo nombre
        page_number_matriz = request.GET.get('page_matriz') # <-- Usamos un param diferente
        matriz_page_obj = matriz_paginator.get_page(page_number_matriz) # <-- Nuevo nombre
        
        seguimientos = SeguimientoDiario.objects.filter(
            dispositivo__in=matriz_page_obj,
            fecha__year=ano_seleccionado,
            fecha__month=mes_seleccionado
        )
        
        matriz_seguimiento = {}
        for s in seguimientos:
            dispositivo_id = s.dispositivo.id_dispositivo # Obtenemos el ID
            if dispositivo_id not in matriz_seguimiento:
                matriz_seguimiento[dispositivo_id] = {}
            matriz_seguimiento[dispositivo_id][s.fecha] = s.estado_texto
            
        # Añadimos los datos de la matriz al contexto
        context.update({
            'matriz_page_obj': matriz_page_obj,
            'dias_del_mes': dias_del_mes,
            'matriz_seguimiento': matriz_seguimiento,
            'opciones_estado': SeguimientoDiarioForm.ESTADO_CHOICES,
        })

    return render(request, 'seguimiento/gestionar_seguimiento.html', context)

@login_required
def vista_tabla_portatiles(request):
    base_query = Dispositivo.objects.filter(tipoDisp='Portatil')
    
    opciones_modelos = base_query.values_list('nomDisp', flat=True).distinct().order_by('nomDisp')
    opciones_estados = base_query.values_list('estadoD', flat=True).distinct().order_by('estadoD')
    opciones_areas = base_query.exclude(area_general__isnull=True).exclude(area_general__exact='')\
                               .values_list('area_general', flat=True).distinct().order_by('area_general')

    # Opciones para los nuevos filtros (no necesitan consulta a la BD)
    opciones_garantia = [('Vigente', 'Vigente'), ('Caducado', 'Caducado')]
    opciones_observaciones = [('Con', 'Con Observaciones'), ('Sin', 'Sin Observaciones')]

    opciones_tipo_gas = Sensor.objects.filter(dispositivo_instalado__tipoDisp='Portatil')\
                                    .values_list('tipGas', flat=True).distinct().order_by('tipGas')
    opciones_estado_sensor = Sensor.objects.filter(dispositivo_instalado__tipoDisp='Portatil')\
                                       .values_list('estComp', flat=True).distinct().order_by('estComp')
    opciones_garantia_sensor = [('Vigente', 'Vigente'), ('Caducado', 'Caducado')]

    # ========================================================================
    # 2. CAPTURAR TODOS LOS VALORES DE FILTRO DE LA URL (request.GET)
    # ========================================================================
    modelo_filtro = request.GET.get('modelo', '')
    serie_filtro = request.GET.get('serie', '')
    estado_filtro = request.GET.get('estado', '')
    area_filtro = request.GET.get('area', '')
    fecha_desde_filtro = request.GET.get('fecha_desde', '')
    fecha_hasta_filtro = request.GET.get('fecha_hasta', '')
    ingreso_desde_filtro = request.GET.get('ingreso_desde', '')
    ingreso_hasta_filtro = request.GET.get('ingreso_hasta', '')
    
    # --- FECHA DE VENCIMIENTO DE GARANTÍA (RANGO) ---
    vencimiento_desde_filtro = request.GET.get('vencimiento_desde', '')
    vencimiento_hasta_filtro = request.GET.get('vencimiento_hasta', '')

    # Nuevos filtros
    modificacion_desde_filtro = request.GET.get('modificacion_desde', '')
    mantenimiento_desde_filtro = request.GET.get('mantenimiento_desde', '')
    garantia_filtro = request.GET.get('garantia', '')
    irreparable_desde_filtro = request.GET.get('irreparable_desde', '')
    inoperativo_desde_filtro = request.GET.get('inoperativo_desde', '')
    observaciones_filtro = request.GET.get('observaciones', '')

    sensor_gas_filtro = request.GET.get('sensor_gas', '')
    sensor_serie_filtro = request.GET.get('sensor_serie', '')
    sensor_estado_filtro = request.GET.get('sensor_estado', '')
    sensor_instalacion_desde_filtro = request.GET.get('sensor_instalacion_desde', '')
    sensor_instalacion_hasta_filtro = request.GET.get('sensor_instalacion_hasta', '')
    sensor_garantia_filtro = request.GET.get('sensor_garantia', '')

    # ========================================================================
    # 3. CONSTRUIR EL QUERYSET BASE Y APLICAR FILTROS
    # ========================================================================
    dispositivos_portatiles = Dispositivo.objects.filter(tipoDisp='Portatil').select_related(
        'id_areaTrabajo_fijo'
    ).prefetch_related(
        'sensor_set',
        'modificacion_set__id_trabajador',
        'modificacion_set__sensor_saliente',
        'modificacion_set__componente_entrante',
        'partes',
        'mantenimientos__tecnico_a_cargo',
        'observaciones'
    ).order_by('id_dispositivo')
    
    # Bandera para saber si necesitamos usar .distinct() al final
    needs_distinct = False

    # Filtros existentes (con 'modelo' actualizado)
    if modelo_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(nomDisp=modelo_filtro)
    if serie_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(num_serie__icontains=serie_filtro)
    if estado_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(estadoD=estado_filtro)
    if area_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(area_general=area_filtro)
    if fecha_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fecIngreso__gte=fecha_desde_filtro)
    if fecha_hasta_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fecIngreso__lte=fecha_hasta_filtro)

    # --- LÓGICA DE FILTRADO PARA FECHA DE INGRESO ---
    if ingreso_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fecIngreso__gte=ingreso_desde_filtro)
    if ingreso_hasta_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fecIngreso__lte=ingreso_hasta_filtro)
        
    # --- LÓGICA DE FILTRADO PARA VENCIMIENTO DE GARANTÍA ---
    if vencimiento_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fecVencimientoGarantia__gte=vencimiento_desde_filtro)
    if vencimiento_hasta_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fecVencimientoGarantia__lte=vencimiento_hasta_filtro)
        
    # --- APLICACIÓN DE NUEVOS FILTROS ---
    if modificacion_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(modificacion__fecInstalacionMod__gte=modificacion_desde_filtro)
        needs_distinct = True
    if mantenimiento_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(mantenimientos__fecha_intervencion__gte=mantenimiento_desde_filtro)
        needs_distinct = True
    if irreparable_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fec_irreparable__gte=irreparable_desde_filtro)
    if inoperativo_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(fec_inoperativo__gte=inoperativo_desde_filtro)
        
    # Filtro de Estatus de Garantía (lógica especial)
    if garantia_filtro:
        hoy = date.today()
        if garantia_filtro == 'Vigente':
            dispositivos_portatiles = dispositivos_portatiles.filter(fecVencimientoGarantia__gte=hoy)
        elif garantia_filtro == 'Caducado':
            dispositivos_portatiles = dispositivos_portatiles.filter(fecVencimientoGarantia__lt=hoy)
            
    # Filtro por existencia de observaciones
    if observaciones_filtro:
        if observaciones_filtro == 'Con':
            dispositivos_portatiles = dispositivos_portatiles.filter(observaciones__isnull=False)
            needs_distinct = True
        elif observaciones_filtro == 'Sin':
            dispositivos_portatiles = dispositivos_portatiles.filter(observaciones__isnull=True)

    if sensor_gas_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(sensor__tipGas=sensor_gas_filtro)
        needs_distinct = True
    if sensor_serie_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(sensor__nSerieActual__icontains=sensor_serie_filtro)
        needs_distinct = True
    if sensor_estado_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(sensor__estComp=sensor_estado_filtro)
        needs_distinct = True
    if sensor_instalacion_desde_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(sensor__fecInst__gte=sensor_instalacion_desde_filtro)
        needs_distinct = True
    if sensor_instalacion_hasta_filtro:
        dispositivos_portatiles = dispositivos_portatiles.filter(sensor__fecInst__lte=sensor_instalacion_hasta_filtro)
        needs_distinct = True
        
    # Filtro de Estatus de Garantía del Sensor (lógica especial)
    if sensor_garantia_filtro:
        hoy = date.today()
        if sensor_garantia_filtro == 'Vigente':
            dispositivos_portatiles = dispositivos_portatiles.filter(sensor__fecVencGarantia__gte=hoy)
        elif sensor_garantia_filtro == 'Caducado':
            dispositivos_portatiles = dispositivos_portatiles.filter(sensor__fecVencGarantia__lt=hoy)
        needs_distinct = True
            
    # Aplicamos distinct() solo si es necesario para evitar duplicados por los joins
    if needs_distinct:
        dispositivos_portatiles = dispositivos_portatiles.distinct()

    paginator = Paginator(dispositivos_portatiles, 10) # 15 dispositivos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ----------------- El resto de la lógica de la vista permanece igual -----------------

    sensores_unicos_qs = Sensor.objects.filter(dispositivo_instalado__in=dispositivos_portatiles)\
                                      .values('tipGas', 'nomComp')\
                                      .distinct()
    
    ORDEN_SENSORES = ['LEL', 'O2', 'DUAL', 'SO2', 'CO2', 'NH3', 'CL2', 'HCN']
    
    tipos_presentes = [s['tipGas'] for s in sensores_unicos_qs if s['tipGas']]
    tipos_de_sensor_unicos = [tipo for tipo in ORDEN_SENSORES if tipo in tipos_presentes]

    SENSOR_HEADER_CONFIG = {
        sensor_data.get('tipGas'): f"SENSOR {sensor_data.get('tipGas')}"
        for sensor_data in sensores_unicos_qs
    }

    CAMPOS_POR_SENSOR_DETALLES = [
        'Nº DE SERIE ANTERIOR', 'Nº DE SERIE ACTUAL', 'FECHA DE FABRICACIÓN', 'FECHA DE INSTALACIÓN',
        'RESPONSABLE DEL CAMBIO', 'NÚMERO DE GUÍA DE INGRESO', 'ITEM DE GÚIA',
        'VENCIMIENTO DE GARANTIA', 'ESTATUS DE GARANTIA', 'ESTATUS DEL SENSOR'
    ]
    
    hoy = date.today()

    headers_fijos_inicio = [
        'N°', 'MODELO', 'NÚMERO SERIE', 'Fecha de Fabricación', 'Fecha de ingreso',
        'Fecha de vencimiento de garantía', 'NS', 'CÓDIGO DEL EQUIPO', 'Última fecha de Mantto',
        'Responsable de Mantto', 'SENSOR', 'UBICACIÓN DEL EQUIPO', 'OBSERVACION',
        'ESTATUS GARANTÍA DEL EQUIPO'
    ]
    headers_dinamicos_estatus = tipos_de_sensor_unicos
    headers_fijos_finales = [
        'Número de Guía', 'OBSERVACION POR FALTA DE MANTENIMIENTO INDICADO POR MSA',
        'FECHA QUE PASA A IRREPARABLE', 'FECHA QUE PASA A INOPEATIVO', 'ESTADO DEL EQUIPO',
        'PROPIEDAD DEL EQUIPO', 'FECHA DE ULTIMA MODIFICACIÓN'
    ]
    headers_dinamicos_detalles_flat = []
    for tipo in tipos_de_sensor_unicos:
        headers_dinamicos_detalles_flat.extend([f'{campo} {tipo}' for campo in CAMPOS_POR_SENSOR_DETALLES])

    headers_fijos_finales_finales = ['PLACA ELECTRÓNICA - FECHA FABRICACIÓN', 'SENSOR CANIBALIZADO',
        'PCBA', 'CARCASA', 'CLIP', 'CARDEX']
        
    headers_completos = headers_fijos_inicio + headers_dinamicos_estatus + headers_fijos_finales + headers_dinamicos_detalles_flat + headers_fijos_finales_finales

    # ========================================================================
    # OPTIMIZACIÓN: Cachear consultas fuera del loop
    # ========================================================================
    dispositivos_ids = [d.id_dispositivo for d in page_obj]
    
    # Pre-cargar todas las modificaciones relacionadas con sensores de estos dispositivos
    modificaciones_por_sensor = {}
    modificaciones_qs = Modificacion.objects.filter(
        Q(id_dispositivo__in=dispositivos_ids) & 
        (Q(sensor_saliente__isnull=False) | Q(componente_entrante__isnull=False))
    ).select_related('sensor_saliente', 'id_trabajador', 'componente_entrante').order_by('-fecInstalacionMod')
    
    for mod in modificaciones_qs:
        if mod.sensor_saliente:
            sensor_id = mod.sensor_saliente.id_componente
            if sensor_id not in modificaciones_por_sensor:
                modificaciones_por_sensor[sensor_id] = mod
        if mod.componente_entrante:
            comp_id = mod.componente_entrante.id_componente
            if comp_id not in modificaciones_por_sensor:
                modificaciones_por_sensor[comp_id] = mod
    
    # Pre-cargar sensores canibalizados
    sensores_canibalizados = {}
    for sensor in Sensor.objects.filter(
        dispositivo_instalado__in=dispositivos_ids,
        info_canibalizado__isnull=False
    ).exclude(info_canibalizado__exact=''):
        sensores_canibalizados[sensor.dispositivo_instalado_id] = sensor.info_canibalizado

    datos_tabla = []
    for idx, dispositivo in enumerate(page_obj, start=page_obj.start_index()):
        ultima_modificacion = dispositivo.modificacion_set.order_by('-fecInstalacionMod').first()
        sensores_del_dispositivo = {s.tipGas: s for s in dispositivo.sensor_set.all()}
        ultimo_mantenimiento = dispositivo.mantenimientos.order_by('-fecha_intervencion').first()
        estatus_garantia_equipo = 'VIGENTE' if dispositivo.fecVencimientoGarantia and dispositivo.fecVencimientoGarantia >= hoy else 'CADUCADO'
        ns_codigo = ''.join(filter(str.isdigit, dispositivo.tag or ''))
        placa_electronica = next((p for p in dispositivo.partes.all() if 'placa electrónica' in p.nomPart.lower()), None)
        fecha_fab_placa = placa_electronica.fecFab if placa_electronica and placa_electronica.fecFab else ''
        # Usar cache de sensores canibalizados
        sensor_canibalizado_info = sensores_canibalizados.get(dispositivo.id_dispositivo, '')
        ultima_fecha_mant_str = ''
        if ultimo_mantenimiento and ultimo_mantenimiento.fecha_intervencion:
            fecha_local = timezone.localtime(ultimo_mantenimiento.fecha_intervencion)
            ultima_fecha_mant_str = fecha_local.strftime('%d/%m/%Y')
        cambios_partes_clave = { 'PCBA': None, 'CARCASA': None, 'CLIP': None }
        todas_mods_partes = dispositivo.modificacion_set.filter(MotivoCambio__in=cambios_partes_clave.keys()).order_by('MotivoCambio', '-fecInstalacionMod')
        for mod in todas_mods_partes:
            motivo = mod.MotivoCambio.upper()
            if motivo in cambios_partes_clave and cambios_partes_clave[motivo] is None:
                cambios_partes_clave[motivo] = mod
        
        cardex_status = "Revisado" if getattr(dispositivo, 'cardex_revisado', False) else ""
        observacion_html = f"""
            <button type="button" class="btn btn-sm btn-info btn-observacion" 
                    data-bs-toggle="modal" data-bs-target="#observacionesModal"
                    data-device-id="{dispositivo.id_dispositivo}"
                    data-device-name="{dispositivo.nomDisp}"
                    
                    // --- ¡NUEVOS ATRIBUTOS DATA CON LAS URLS! ---
                    data-get-url="{ reverse('cenerisapp:get_observaciones_json', args=[dispositivo.pk]) }"
                    data-add-url="{ reverse('cenerisapp:add_observacion_json', args=[dispositivo.pk]) }">
                Ver/Añadir
            </button>
        """
        row_data = [
            idx, dispositivo.nomDisp, dispositivo.num_serie, dispositivo.fecFabricacion,
            dispositivo.fecIngreso, dispositivo.fecVencimientoGarantia, ns_codigo, dispositivo.tag,
            ultima_fecha_mant_str if ultima_fecha_mant_str else '',
            ultimo_mantenimiento.tecnico_a_cargo.nomEmpleado if ultimo_mantenimiento and ultimo_mantenimiento.tecnico_a_cargo else '',
            ", ".join(sensores_del_dispositivo.keys()),
            dispositivo.area_general if dispositivo.area_general else '',
            observacion_html,
            estatus_garantia_equipo,
        ]
        
        for tipo in tipos_de_sensor_unicos:
            sensor = sensores_del_dispositivo.get(tipo)
            row_data.append('VIGENTE' if sensor and sensor.fecVencGarantia and sensor.fecVencGarantia >= hoy else ('CADUCADO' if sensor else 'N/A'))

        row_data.extend([
            "",
            ultimo_mantenimiento.observacion_msa if ultimo_mantenimiento else '',
            dispositivo.fec_irreparable, dispositivo.fec_inoperativo, dispositivo.estadoD,
            getattr(dispositivo, 'propiedad', 'SMCV'),
            ultima_modificacion.fecInstalacionMod if ultima_modificacion else ''
        ])
        
        for tipo in tipos_de_sensor_unicos:
            sensor = sensores_del_dispositivo.get(tipo)
            if sensor:
                # Usar cache de modificaciones en lugar de consultar
                mod_sensor = modificaciones_por_sensor.get(sensor.id_componente)
                row_data.extend([
                    mod_sensor.sensor_saliente.nSerieActual if mod_sensor and mod_sensor.sensor_saliente else '',
                    sensor.nSerieActual, sensor.fecFabComp, sensor.fecInst,
                    mod_sensor.id_trabajador.nomEmpleado if mod_sensor and mod_sensor.id_trabajador else '',
                    sensor.nro_guia_ingreso, sensor.item_guia, sensor.fecVencGarantia,
                    'VIGENTE' if sensor.fecVencGarantia and sensor.fecVencGarantia >= hoy else 'CADUCADO',
                    sensor.estComp
                ])
            else:
                row_data.extend([''] * len(CAMPOS_POR_SENSOR_DETALLES))
        
        pcba_mod = cambios_partes_clave['PCBA']
        carcasa_mod = cambios_partes_clave['CARCASA']
        clip_mod = cambios_partes_clave['CLIP']
        row_data.extend([
            fecha_fab_placa, sensor_canibalizado_info,
            f"Se cambió por {pcba_mod.id_trabajador.nomEmpleado}" if pcba_mod and pcba_mod.id_trabajador else '',
            f"Se cambió por {carcasa_mod.id_trabajador.nomEmpleado}" if carcasa_mod and carcasa_mod.id_trabajador else '',
            f"Se cambió por {clip_mod.id_trabajador.nomEmpleado}" if clip_mod and clip_mod.id_trabajador else '',
            cardex_status
        ])
        
        datos_tabla.append(row_data)

    # ========================================================================
    # 4. PREPARAR CONTEXTO FINAL PARA LA PLANTILLA
    # ========================================================================
    context = {
        'titulo': 'Base de Datos de Equipos Portátiles',
        'page_obj': page_obj,
        'datos_tabla': datos_tabla,
        'headers_completos': headers_completos,
        # Datos para construir las cabeceras complejas
        'headers_fijos_inicio': headers_fijos_inicio,
        'tipos_de_sensor_unicos': tipos_de_sensor_unicos,
        'headers_fijos_finales': headers_fijos_finales,
        'CAMPOS_POR_SENSOR_DETALLES': CAMPOS_POR_SENSOR_DETALLES,
        'SENSOR_HEADER_CONFIG': SENSOR_HEADER_CONFIG,
        'headers_fijos_finales_finales': headers_fijos_finales_finales,
        
        # --- AÑADIR DATOS PARA EL FORMULARIO DE FILTROS ---
        'opciones_modelos': opciones_modelos,
        'opciones_estados': opciones_estados,
        'opciones_areas': opciones_areas,
        'opciones_garantia': opciones_garantia,
        'opciones_observaciones': opciones_observaciones,
        'opciones_tipo_gas': opciones_tipo_gas,
        'opciones_estado_sensor': opciones_estado_sensor,
        'opciones_garantia_sensor': opciones_garantia_sensor,
        'filtros_aplicados': {
            'modelo': modelo_filtro,
            'serie': serie_filtro,
            'estado': estado_filtro,
            'area': area_filtro,
            'ingreso_desde': ingreso_desde_filtro,       # <-- Añadido/Actualizado
            'ingreso_hasta': ingreso_hasta_filtro,       # <-- Añadido/Actualizado
            'vencimiento_desde': vencimiento_desde_filtro, # <-- Nuevo
            'vencimiento_hasta': vencimiento_hasta_filtro, # <-- Nuevo
            'fecha_desde': fecha_desde_filtro,
            'fecha_hasta': fecha_hasta_filtro,
            'modificacion_desde': modificacion_desde_filtro,
            'mantenimiento_desde': mantenimiento_desde_filtro,
            'garantia': garantia_filtro,
            'irreparable_desde': irreparable_desde_filtro,
            'inoperativo_desde': inoperativo_desde_filtro,
            'observaciones': observaciones_filtro,
            'sensor_gas': sensor_gas_filtro,
            'sensor_serie': sensor_serie_filtro,
            'sensor_estado': sensor_estado_filtro,
            'sensor_instalacion_desde': sensor_instalacion_desde_filtro,
            'sensor_instalacion_hasta': sensor_instalacion_hasta_filtro,
            'sensor_garantia': sensor_garantia_filtro,
        }
    }
    
    return render(request, 'tabla/tabla_portatiles.html', context)

@login_required
def crear_empresa(request):
    if request.method == 'POST':
        abreviacion = request.POST.get('abreviacion')
        nombreE = request.POST.get('nombreE')
        direccion = request.POST.get('direccion')
        departamento = request.POST.get('departamento')
        telefono = request.POST.get('telefono')
        ruc = request.POST.get('ruc')
        Empresa.objects.create(
            abreviacion=abreviacion,
            nombreE=nombreE,
            direccion=direccion,
            departamento=departamento,
            telefono=telefono,
            ruc=ruc
        )
        return redirect('cenerisapp:lista_empresas')
    return render(request, 'tecnicos/crear_empresa.html')


@login_required
def editar_empresa(request, empresa_id):
    empresa = Empresa.objects.get(id=empresa_id)
    if request.method == 'POST':
        empresa.abreviacion = request.POST.get('abreviacion')
        empresa.nombreE = request.POST.get('nombreE')
        empresa.direccion = request.POST.get('direccion')
        empresa.departamento = request.POST.get('departamento')
        empresa.telefono = request.POST.get('telefono')
        empresa.ruc = request.POST.get('ruc')
        empresa.save()
        return redirect('cenerisapp:lista_empresas')
    return render(request, 'tecnicos/editar_empresa.html', {'empresa': empresa})


@login_required
def eliminar_empresa(request, empresa_id):
    empresa = Empresa.objects.get(id=empresa_id)
    empresa.delete()
    return redirect('cenerisapp:lista_empresas')


@login_required
def crear_area(request):
    if request.method == 'POST':
        # Instanciamos el formulario principal y el formset con los datos del POST
        form = AreaTrabajoForm(request.POST)
        formset = PuntoExactoFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            # Primero, guardamos el objeto principal (AreaTrabajo)
            area_trabajo = form.save()
            
            # Asociamos el formset con la instancia del área recién creada
            puntos_exactos = formset.save(commit=False)
            for punto in puntos_exactos:
                punto.area_trabajo = area_trabajo
                punto.save()

            messages.success(request, f"Área de Trabajo '{area_trabajo.nombreA}' y sus puntos exactos han sido creados.")
            return redirect('cenerisapp:lista_areas') # Asume que tienes una URL con este nombre
    else:
        # Petición GET: mostramos los formularios vacíos
        form = AreaTrabajoForm()
        formset = PuntoExactoFormSet()

    context = {
        'form': form,
        'formset': formset,
        'titulo': 'Crear Nueva Área de Trabajo'
    }
    return render(request, 'tecnicos/crear_area.html', context)


# --- VISTA DE EDICIÓN ---
@login_required
def editar_area(request, area_id):
    area_trabajo = get_object_or_404(AreaTrabajo, pk=area_id)
    
    if request.method == 'POST':
        # Pasamos la 'instance' a ambos para indicar que estamos editando
        form = AreaTrabajoForm(request.POST, instance=area_trabajo)
        formset = PuntoExactoFormSet(request.POST, instance=area_trabajo)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save() # Con 'instance', Django maneja la creación, actualización y borrado
            
            messages.success(request, f"Área de Trabajo '{area_trabajo.nombreA}' actualizada.")
            return redirect('cenerisapp:lista_areas')
    else:
        # Petición GET: mostramos los formularios pre-llenados
        form = AreaTrabajoForm(instance=area_trabajo)
        formset = PuntoExactoFormSet(instance=area_trabajo)

    context = {
        'form': form,
        'formset': formset,
        'area': area_trabajo,
        'titulo': f'Editar Área de Trabajo: {area_trabajo.nombreA}'
    }
    return render(request, 'tecnicos/editar_area.html', context)

@login_required
def eliminar_area(request, area_id):
    area = get_object_or_404(AreaTrabajo, pk=area_id)
    area.delete()
    return redirect('cenerisapp:lista_areas')

@login_required
def marcar_cardex_revisado(request, dispositivo_id):
    if request.method == 'POST':
        dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
        dispositivo.cardex_revisado = True
        dispositivo.save()
        messages.success(request, f"CARDEX para {dispositivo.nomDisp} marcado como revisado.")
    return redirect('cenerisapp:lista_dispositivos')

@login_required
def lista_programas(request):
    
    # --- 1. PREPARAR DATOS PARA LA NAVEGACIÓN Y FILTROS ---
    anos_disponibles = Programa.objects.values_list('ano', flat=True).distinct().order_by('-ano')
    meses_disponibles = Programa.MES_CHOICES
    tipos_disponibles = Programa.TIPO_DISPOSITIVO_CHOICES
    
    # --- 2. CAPTURAR EL AÑO Y FILTROS SELECCIONADOS ---
    today = date.today()
    ano_seleccionado = int(request.GET.get('year', today.year))
    
    mes_filtro = request.GET.get('mes')
    tipo_filtro = request.GET.get('tipo')
    
    # --- 3. CONSTRUIR QUERYSET BASE Y APLICAR FILTROS ---
    # La base siempre se filtra por el año seleccionado
    programas_qs = Programa.objects.filter(ano=ano_seleccionado)
    
    if mes_filtro:
        programas_qs = programas_qs.filter(mes=mes_filtro)
    if tipo_filtro:
        programas_qs = programas_qs.filter(tipo_dispositivo=tipo_filtro)
    
    # Anotamos el conteo de certificados (esto es crucial y se mantiene)
    programas_con_conteo = programas_qs.annotate(
        num_certificados=Count('certificado')
    )

    # Actualizamos el totalEjecutado en memoria para la plantilla
    for programa in programas_con_conteo:
        programa.totalEjecutado = programa.num_certificados
    
    # --- 4. PREPARAR CONTEXTO ---
    context = {
        'programas': programas_con_conteo, 
        'titulo': f'Programas de Calibración - Año {ano_seleccionado}',
        
        # Para la navegación por año
        'anos_disponibles': anos_disponibles,
        'ano_seleccionado': ano_seleccionado,
        
        # Para los desplegables de filtros
        'meses_disponibles': meses_disponibles,
        'tipos_disponibles': tipos_disponibles,
        'filtros_aplicados': {
            'mes': int(mes_filtro) if mes_filtro else None,
            'tipo': tipo_filtro,
        }
    }
    return render(request, 'programas/lista_programas.html', context)

@login_required
def crear_programa(request):
    if request.method == 'POST':
        form = ProgramaCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Nuevo programa creado exitosamente.")
            return redirect('cenerisapp:lista_programas')
    else:
        form = ProgramaCreateForm()
    
    context = {'form': form, 'titulo': 'Crear Nuevo Programa'}
    return render(request, 'programas/crear_programa.html', context)

@login_required
def ver_certificados_programa(request, programa_id):
    # 1. Obtenemos el objeto del programa específico. Si no existe, devuelve un error 404.
    programa = get_object_or_404(Programa, pk=programa_id)

    # 2. Hacemos la consulta clave: Filtramos todos los Certificados
    #    cuyo campo 'id_programa' coincida con el programa que obtuvimos.
    #    Usamos 'select_related' para optimizar la consulta y evitar
    #    golpes extra a la BD al acceder al dispositivo en la plantilla.
    certificados_del_programa = Certificado.objects.filter(
        id_programa=programa
    ).select_related('dispositivo').order_by('-fechCertificado')

    # 3. Preparamos el contexto para pasarlo a la plantilla.
    context = {
        'programa': programa,
        'certificados': certificados_del_programa,
        'titulo': f"Certificados para el Programa: {programa.get_mes_display()} {programa.ano}"
    }
    
    # 4. Renderizamos la nueva plantilla que crearemos en el siguiente paso.
    return render(request, 'programas/certificados_por_programa.html', context)

@login_required
def editar_programa(request, programa_id):
    programa = get_object_or_404(Programa, pk=programa_id)
    
    if request.method == 'POST':
        form = ProgramaUpdateForm(request.POST, instance=programa)
        if form.is_valid():
            form.save()
            messages.success(request, f"Progreso del programa #{programa.id_programa} actualizado.")
            return redirect('cenerisapp:lista_programas')
    else:
        form = ProgramaUpdateForm(instance=programa)

    context = {
        'form': form,
        'programa': programa,
        'titulo': f'Actualizar Progreso del Programa #{programa.id_programa}'
    }
    return render(request, 'programas/editar_programa.html', context)


@login_required
def seleccionar_sensor_para_informe(request, dispositivo_id):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    # Obtenemos todos los sensores instalados en este dispositivo
    sensores_del_dispositivo = dispositivo.sensor_set.all()
    
    context = {
        'dispositivo': dispositivo,
        'sensores_del_dispositivo': sensores_del_dispositivo,
        'titulo': f'Seleccionar Sensor para Informe en {dispositivo.nomDisp}'
    }
    return render(request, 'informes/seleccionar_sensor.html', context)

@login_required
def crear_informe_calibracion(request, sensor_id):
    # Buscamos el sensor específico, y a través de él, su dispositivo
    sensor = get_object_or_404(Sensor.objects.select_related('dispositivo_instalado'), pk=sensor_id)
    dispositivo = sensor.dispositivo_instalado

    if request.method == 'POST':
        form = InformeCalibracionForm(request.POST)
        if form.is_valid():
            informe = form.save(commit=False)
            informe.sensor = sensor # <-- Asignamos el sensor
            informe.save()
            messages.success(request, f"Informe guardado para el sensor '{sensor.nomComp}'.")
            
            # Redirigimos a la página de detalles del dispositivo padre
            return redirect('cenerisapp:detalle_dispositivo', pk=dispositivo.pk) 
    else:
        form = InformeCalibracionForm()

    context = {
        'form': form,
        'sensor': sensor,
        'dispositivo': dispositivo, # Pasamos ambos para mostrar info en el título
        'titulo': f'Nuevo Informe para Sensor {sensor.nomComp} (en {dispositivo.nomDisp})'
    }
    return render(request, 'informes/crear_informe.html', context)


@login_required
def dashboard_index(request):
    # Asegúrate de que los campos 'tipoDisp' y 'estadoD' estén en mayúsculas
    # si los estás comparando/agrupando directamente contra cadenas de texto.
    from django.db.models.functions import Upper, Lower
    from django.db.models import Count

    total_dispositivos = Dispositivo.objects.count()
    
    # Aplicar Upper a tipoDisp para agrupar de forma consistente
    dispositivo_por_tipo = Dispositivo.objects.annotate(
        tipo_upper=Upper('tipoDisp')
    ).values('tipo_upper').annotate(
        total=Count('tipo_upper')
    ).order_by('tipo_upper')
    
    entregas_por_empresa = Registro.objects.values(
        'trabajador_receptor__empresa_id__nombreE'
    ).annotate(
        total=Count('id_registro')
    ).order_by('-total')

    # --- CÁLCULO DE OPERATIVOS/INOPERATIVOS ---
    # Usar icontains o una función Upper para la comprobación en el filtro es más robusto si los datos son inconsistentes.
    # Si quieres estrictamente 'Operativo' o 'Inoperativo', el filtro es suficiente si la data está limpia.
    # Asumimos que los datos de estadoD son correctos, pero normalizamos para el conteo:
    dispositivo_por_estado = Dispositivo.objects.annotate(
        estado_upper=Upper('estadoD')
    ).values('estado_upper').annotate(
        total=Count('estado_upper')
    ).order_by('estado_upper')
    
    # Recalculamos los contadores exactos usando Upper() para la consistencia
    operativo = Dispositivo.objects.filter(estadoD__iexact='operativo').count()
    inoperativo = Dispositivo.objects.filter(estadoD__iexact='inoperativo').count()
    
    # === Estadísticas de Préstamos (sin cambios) ===
    prestados = Registro.objects.filter(fecDevol__isnull=True).count()
    
    # Usamos __iexact (case-insensitive) para la disponibilidad
    disponibles_operativos = Dispositivo.objects.filter(estadoD__iexact='operativo').count() - prestados

    
    # --- PROCESAMIENTO POR ÁREA Y ESTADO ---
    datos_crudos_area_estado = Dispositivo.objects.filter(
        tipoDisp__iexact='portatil' # Usamos iexact para el filtro de tipoDisp
    ).exclude(
        area_general__isnull=True
    ).exclude(
        area_general__exact=''
    ).annotate(
        area_upper=Upper('area_general'), # Normalizamos el área para agrupar
        estado_upper=Upper('estadoD')    # Normalizamos el estado
    ).values(
        'area_upper', 'estado_upper'
    ).annotate(
        total=Count('id_dispositivo')
    ).order_by('area_upper')
    
    datos_procesados = {}
    
    for item in datos_crudos_area_estado:
        # Usamos los campos normalizados del ORM
        area = item['area_upper']
        estado = item['estado_upper']
        total = item['total']
        
        # El mapeo de estado debe usar mayúsculas
        if area not in datos_procesados:
            datos_procesados[area] = {'OPERATIVO': 0, 'INOPERATIVO': 0}
        
        # El estado 'Operativo' se convierte en 'OPERATIVO', y 'Inoperativo' en 'INOPERATIVO'
        if estado in datos_procesados[area]:
            datos_procesados[area][estado] = total

    labels_area_estado = list(datos_procesados.keys())
    # NOTA: Debes cambiar 'Operativo'/'Inoperativo' a 'OPERATIVO'/'INOPERATIVO' en esta línea
    data_operativos = [datos['OPERATIVO'] for datos in datos_procesados.values()]
    data_inoperativos = [datos['INOPERATIVO'] for datos in datos_procesados.values()]
    
    # --- PRÉSTAMOS POR TURNO ---
    turnos = ['A', 'B']

    # Normalizamos el turno en el ORM para agrupar de forma consistente
    data_dispositivos = Registro.objects.filter(
        id_dispositivo__tipoDisp__iexact='portatil', # iexact para el filtro
        turno__in=turnos
    ).annotate(
        turno_upper=Upper('turno') # Normalizamos el turno para agrupar
    ).values('turno_upper').annotate(total=Count('id_registro'))
    
    # Contamos los préstamos de COMPONENTES (bombas) por cada turno
    data_bombas = Registro.objects.filter(
        id_componente__nomComp__icontains='bomba', 
        turno__in=turnos
    ).annotate(
        turno_upper=Upper('turno') # Normalizamos el turno para agrupar
    ).values('turno_upper').annotate(total=Count('id_registro'))

    # Convertimos los resultados a un formato fácil de usar para el gráfico
    # Usamos los campos normalizados 'turno_upper'
    dispositivos_por_turno = {item['turno_upper']: item['total'] for item in data_dispositivos}
    bombas_por_turno = {item['turno_upper']: item['total'] for item in data_bombas}
    
    # Las etiquetas de los turnos deben coincidir con la normalización si el input es distinto
    labels_turnos = [t.upper() for t in turnos] # Normalizamos las etiquetas
    values_dispositivos = [dispositivos_por_turno.get(t, 0) for t in labels_turnos]
    values_bombas = [bombas_por_turno.get(t, 0) for t in labels_turnos]


    context = {
        'total_dispositivos': total_dispositivos,
        'por_tipo': list(dispositivo_por_tipo),
        'por_estado': list(dispositivo_por_estado),
        'entregas_por_empresa': list(entregas_por_empresa),
        'operativo': operativo,
        'inoperativo': inoperativo,
        'prestados': prestados,
        'disponibles': disponibles_operativos,
        'labels_area_estado': labels_area_estado,
        'data_operativos': data_operativos,
        'data_inoperativos': data_inoperativos,

        'labels_turnos': labels_turnos,
        'values_dispositivos': values_dispositivos,
        'values_bombas': values_bombas,
    }
    return render(request, 'dashboard/index.html', context)

    
@login_required
def crear_mantenimiento(request, dispositivo_id):
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    
    
    partes_del_dispositivo = list(dispositivo.partes.all().values_list('nomPart', flat=True))
    sensores_del_dispositivo = [f"Sensor {s.tipGas}" for s in dispositivo.sensor_set.all()]
    checklist_items = sorted(partes_del_dispositivo + sensores_del_dispositivo)

    if request.method == 'POST':
        
        form = MantenimientoForm(request.POST)
        
        if form.is_valid():
            mantenimiento = form.save(commit=False)
            mantenimiento.dispositivo = dispositivo
            
            
            checklist_data = {}
            for item in checklist_items:
                estado = request.POST.get(f'checklist_estado_{item}')
                comentario = request.POST.get(f'checklist_comentario_{item}')
                if estado:
                    checklist_data[item] = {'estado': estado, 'comentario': comentario}
            mantenimiento.checklist_partes = checklist_data
            
            mantenimiento.save() # Guardamos el registro de mantenimiento

            
            fec_ino = form.cleaned_data.get('actualizar_fec_inoperativo')
            fec_irr = form.cleaned_data.get('actualizar_fec_irreparable')
            if dispositivo.tipoDisp == 'Portatil' and (fec_ino or fec_irr):
                if fec_ino:
                    dispositivo.fec_inoperativo = fec_ino
                    dispositivo.estadoD = 'Inoperativo'
                if fec_irr:
                    dispositivo.fec_irreparable = fec_irr
                dispositivo.save()
            
            messages.success(request, "Registro de mantenimiento guardado.")
            return redirect('cenerisapp:lista_dispositivos') # O a una lista de mantenimientos
            
    else: # GET
        form = MantenimientoForm()
        
    context = {
        'form': form,
        'dispositivo': dispositivo,
        'checklist_items': checklist_items, # Le pasamos la lista de items
        'titulo': f'Nuevo Mantenimiento para {dispositivo.nomDisp}'
    }
    return render(request, 'mantenimiento/crear_mantenimiento.html', context)

@login_required
def exportar_indice(request):
    
    años_disponibles = Programa.objects.values_list('ano', flat=True).distinct().order_by('-ano')
    tipos_disponibles = Programa.objects.values_list('tipo_dispositivo', flat=True).distinct().order_by('tipo_dispositivo')
    
    fechas_de_reportes = Reporte.objects.values_list('fecReport', flat=True).distinct()
    
    años_unicos = set()
    for fecha in fechas_de_reportes:
        if fecha: # Nos aseguramos de que la fecha no sea Nula
            años_unicos.add(fecha.year)
    
    años_reportes = sorted(list(años_unicos), reverse=True)

    if not años_disponibles:
        años_disponibles = [date.today().year]
    if not tipos_disponibles:
        tipos_disponibles = ['Fijo', 'Portatil']

    tipos_disponibles = Programa.objects.values_list('tipo_dispositivo', flat=True).distinct().order_by('tipo_dispositivo')
    if not tipos_disponibles.exists():
        tipos_disponibles = ['Fijo', 'Portatil']

    areas_generales_disponibles = Dispositivo.objects.values_list('area_general', flat=True)\
                                                                                .distinct()\
                                                                                    .order_by('area_general')

    areas_generales_portatiles = Dispositivo.objects.filter(tipoDisp='Portatil')\
                                                    .values_list('area_general', flat=True)\
                                                    .distinct().order_by('area_general')
    
    turnos_disponibles = Registro.objects.values_list('turno', flat=True).distinct().order_by('turno')

    operadores = Empleado.objects.filter(
        Q(puesto__iexact="Operador de Servicio Tecnico") | 
        Q(puesto__iexact="Supervisor Operativo")
    ).order_by('nomEmpleado')

    areas_generales_fijos = Dispositivo.objects.filter(tipoDisp='Fijo')\
                                               .values_list('area_general', flat=True)\
                                               .distinct()\
                                               .order_by('area_general')
    meses_disponibles = [
        (1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'),
        (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'),
        (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')
    ]

    context = {
        'titulo': "Centro de Exportaciones",
        'años_disponibles': años_disponibles,
        'tipos_disponibles': tipos_disponibles,
        'años_reportes': años_reportes,
        'areas_generales_disponibles': [area for area in areas_generales_disponibles if area],
        'turnos_disponibles': [t for t in turnos_disponibles if t], # Filtramos valores vacíos
        'operadores': operadores,
        'areas_generales_fijos': [area for area in areas_generales_fijos if area],
        'areas_generales_portatiles': [area for area in areas_generales_portatiles if area],
        'meses_disponibles': meses_disponibles,
    }
    return render(request, 'exportar/indice.html', context)


@login_required
def exportar_portatiles_excel(request):
    
    
    dispositivos_portatiles = Dispositivo.objects.filter(tipoDisp='Portatil').prefetch_related(
        'sensor_set', 'modificacion_set'
    ).order_by('id_dispositivo')

    
    sensores_unicos_qs = Sensor.objects.filter(dispositivo_instalado__in=dispositivos_portatiles)\
                                      .values('tipGas', 'nomComp')\
                                      .distinct()
    tipos_de_sensor_unicos = sorted([s['tipGas'] for s in sensores_unicos_qs if s['tipGas']])
    
    SENSOR_HEADER_CONFIG = {
        sensor_data.get('tipGas'): f"SENSOR {sensor_data.get('tipGas')}\n{sensor_data.get('nomComp', '')}"
        for sensor_data in sensores_unicos_qs
    }

    CAMPOS_POR_SENSOR_DETALLES = [
        'Nº DE SERIE ANTERIOR', 'Nº DE SERIE ACTUAL', 'FECHA DE FABRICACIÓN', 'FECHA DE INSTALACIÓN',
        'RESPONSABLE DEL CAMBIO', 'NÚMERO DE GUÍA DE INGRESO', 'ITEM DE GÚIA',
        'VENCIMIENTO DE GARANTIA', 'ESTATUS DE GARANTIA', 'ESTATUS DEL SENSOR'
    ]

    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Base de Datos Portátiles"
    hoy = date.today()

    
    super_header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=8, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    
    
    headers_fijos_inicio = [
        'N°', 'MODELO', 'NÚMERO SERIE', 'Fecha de Fabricación', 'Fecha de ingreso',
        'Fecha de vencimiento de garantía', 'NS', 'CÓDIGO DEL EQUIPO', 'Última fecha de Mantto',
        'Responsable de Mantto', 'SENSOR', 'UBICACIÓN DEL EQUIPO', 'OBSERVACION',
        'ESTATUS GARANTÍA DEL EQUIPO'
    ]
    headers_dinamicos_estatus = tipos_de_sensor_unicos # Ej: ['DUAL', 'LEL', 'O2']
    
    headers_fijos_finales = [
        'Número de Guía', 'OBSERVACION POR FALTA DE MANTENIMIENTO INDICADO POR MSA',
        'FECHA QUE PASA A IRREPARABLE', 'FECHA QUE PASA A INOPEATIVO', 'ESTADO DEL EQUIPO',
        'PROPIEDAD DEL EQUIPO', 'FECHA DE ULTIMA MODIFICACIÓN'
    ]
    
    headers_dinamicos_detalles = []
    for tipo in tipos_de_sensor_unicos:
        headers_dinamicos_detalles.extend([f'{campo} {tipo}' for campo in CAMPOS_POR_SENSOR_DETALLES])

    headers_fijos_finales_finales = [ 'PLACA ELECTRÓNICA - FECHA FABRICACIÓN', 'SENSOR CANIBALIZADO',
        'PCBA', 'CARCASA', 'CLIP', 'CARDEX']
        
    headers_completos = headers_fijos_inicio + headers_dinamicos_estatus + headers_fijos_finales + headers_dinamicos_detalles + headers_fijos_finales_finales
    
    
    
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers_fijos_inicio))
    sheet['B1'].value = 'INFORMACIÓN DEL EQUIPO'
    
    start_col = len(headers_fijos_inicio) + 1

    if tipos_de_sensor_unicos:    

        end_col = start_col + len(headers_dinamicos_estatus) -1
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        sheet['O1'].value = 'ESTATUS GARANTÍA DE SENSORES' # Título para el grupo de estatus
        
        start_col_detalles = len(headers_fijos_inicio) + len(headers_dinamicos_estatus) + len(headers_fijos_finales) + 1
        for tipo in tipos_de_sensor_unicos:
            end_col_detalles = start_col_detalles + len(CAMPOS_POR_SENSOR_DETALLES) - 1
            sheet.merge_cells(start_row=1, start_column=start_col_detalles, end_row=1, end_column=end_col_detalles)
            sheet.cell(row=1, column=start_col_detalles).value = SENSOR_HEADER_CONFIG.get(tipo)
            start_col_detalles = end_col_detalles + 1   
        
    
    sheet.append(headers_completos)
    
    for cell in sheet[2]:  # Fila 2 es la de los encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, dispositivo in enumerate(dispositivos_portatiles, 1):
        
        ultima_modificacion = dispositivo.modificacion_set.order_by('-fecInstalacionMod').first()
        sensores_del_dispositivo = {s.tipGas: s for s in dispositivo.sensor_set.all()}
        ultimo_mantenimiento = dispositivo.mantenimientos.order_by('-fecha_intervencion').first()
        
        ultima_fecha_mant_str = ''
        if ultimo_mantenimiento and ultimo_mantenimiento.fecha_intervencion:
            # 1. Convertir de UTC a tu zona horaria local
            fecha_local = timezone.localtime(ultimo_mantenimiento.fecha_intervencion)
            # 2. Formatear solo la parte de la fecha
            ultima_fecha_mant_str = fecha_local.strftime('%d/%m/%Y')


        estatus_garantia_equipo = 'VIGENTE' if dispositivo.fecVencimientoGarantia and dispositivo.fecVencimientoGarantia >= hoy else 'CADUCADO'
        ns_codigo = ''.join(filter(str.isdigit, dispositivo.tag or ''))

        observaciones = dispositivo.observaciones.all().values_list('comentario', flat=True)
    
        # 2. Unimos todos los comentarios en una sola cadena de texto.
        #    Cada comentario estará en una nueva línea, precedido por un guion.
        comentarios_concatenados = "\n".join([f"- {obs}" for obs in observaciones])
        

        placa_electronica = dispositivo.partes.filter(nomPart__icontains='Placa Electrónica').first()
        fecha_fab_placa = ''
        if placa_electronica:
            fecha_fab_placa = placa_electronica.fecFab if hasattr(placa_electronica, 'fecFab') else ''

        
        sensor_canibalizado_info = ''
        sensor_c = Sensor.objects.filter(dispositivo_instalado=dispositivo, info_canibalizado__isnull=False).exclude(info_canibalizado__exact='').first()
        if sensor_c:
            sensor_canibalizado_info = sensor_c.info_canibalizado

        cambios_partes_clave = { 'PCBA': None, 'CARCASA': None, 'CLIP': None }
    
        todas_mods_partes = dispositivo.modificacion_set.filter(
            MotivoCambio__in=cambios_partes_clave.keys()
        ).order_by('MotivoCambio', '-fecInstalacionMod')

        
        
        cambios_partes_clave = { 'PCBA': None, 'CARCASA': None, 'CLIP': None }
        for mod in todas_mods_partes:
            motivo = mod.MotivoCambio.upper()  # Convertimos el motivo a mayúsculas para la búsqueda en el diccionario
            if motivo in cambios_partes_clave and cambios_partes_clave[motivo] is None:
                cambios_partes_clave[motivo] = mod
            
        
        cardex_status = "Revisado" if getattr(dispositivo, 'cardex_revisado', False) else ""

        
        row_data = [
            idx, dispositivo.nomDisp, dispositivo.num_serie, dispositivo.fecFabricacion,
            dispositivo.fecIngreso, dispositivo.fecVencimientoGarantia, ns_codigo, dispositivo.tag,
            ultima_fecha_mant_str if ultima_fecha_mant_str else '',
            ultimo_mantenimiento.tecnico_a_cargo.nomEmpleado if ultimo_mantenimiento and ultimo_mantenimiento.tecnico_a_cargo else '',
            ", ".join(sensores_del_dispositivo.keys()),
            dispositivo.area_general if dispositivo.area_general else '',
            comentarios_concatenados,
            estatus_garantia_equipo,
        ]
        
        
        for tipo in tipos_de_sensor_unicos:
            sensor = sensores_del_dispositivo.get(tipo)
            if sensor and sensor.fecVencGarantia:
                estatus = 'VIGENTE' if sensor.fecVencGarantia >= hoy else 'CADUCADO'
                row_data.append(estatus)
            else:
                row_data.append('N/A')

        
        row_data.extend([
            next((s.nro_guia_ingreso for s in sensores_del_dispositivo.values() if s.nro_guia_ingreso), ''), # Número de Guía
            ultimo_mantenimiento.observacion_msa if ultimo_mantenimiento else '',
            dispositivo.fec_irreparable,
            dispositivo.fec_inoperativo,
            dispositivo.estadoD,
            getattr(dispositivo, 'propiedad', 'CENERIS'),
            ultima_modificacion.fecInstalacionMod if ultima_modificacion else ''
        ])
        
        
        for tipo in tipos_de_sensor_unicos:
            sensor = sensores_del_dispositivo.get(tipo)
            if sensor:
                mod_sensor = Modificacion.objects.filter(Q(sensor_saliente=sensor) | Q(componente_entrante=sensor.componente_ptr)).order_by('-fecInstalacionMod').first()
                ns_anterior = mod_sensor.sensor_saliente.nSerieActual if mod_sensor and mod_sensor.sensor_saliente else ''
                responsable_cambio = mod_sensor.id_trabajador.nomEmpleado if mod_sensor and mod_sensor.id_trabajador else ''
                estatus_garantia = 'VIGENTE' if sensor.fecVencGarantia and sensor.fecVencGarantia >= hoy else 'CADUCADO'

                row_data.extend([
                    ns_anterior, sensor.nSerieActual, sensor.fecFabComp, sensor.fecInst,
                    responsable_cambio, sensor.nro_guia_ingreso, sensor.item_guia,
                    sensor.fecVencGarantia, estatus_garantia, sensor.estComp
                ])
            else:
                row_data.extend([''] * len(CAMPOS_POR_SENSOR_DETALLES))

        

        pcba_mod = cambios_partes_clave['PCBA']
        carcasa_mod = cambios_partes_clave['CARCASA']
        clip_mod = cambios_partes_clave['CLIP']

        row_data.extend([
            fecha_fab_placa,
            sensor_canibalizado_info,
            f"Se cambió por {pcba_mod.id_trabajador.nomEmpleado}" if pcba_mod and pcba_mod.id_trabajador else '',
            f"Se cambió por {carcasa_mod.id_trabajador.nomEmpleado}" if carcasa_mod and carcasa_mod.id_trabajador else '',
            f"Se cambió por {clip_mod.id_trabajador.nomEmpleado}" if clip_mod and clip_mod.id_trabajador else '',
            cardex_status
        ])
        sheet.append(row_data)

        current_row = sheet.max_row
        
        row_fill = gray_fill if current_row % 2 == 0 else white_fill
        
        
        for col_idx, cell in enumerate(sheet[current_row]):
            
            if col_idx >= 0: # Comienza en la segunda columna (índice 1)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = header_alignment # Puedes usar una alineación diferente si lo necesitas


    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Portatiles_Final_{date.today().strftime("%Y-%m-%d")}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_registros_diarios_excel(request):
    
    fecha_str = request.GET.get('fecha')
    area_general_seleccionada = request.GET.get('area_general')
    if area_general_seleccionada:
        
        area_general_seleccionada = area_general_seleccionada.strip()
    turno_seleccionado = request.GET.get('turno')
    operador_id = request.GET.get('operador_id')

    print(f"[PASO 1] Filtros recibidos desde la URL:")
    print(f"  - fecha: '{fecha_str}'")
    print(f"  - area_general: '{area_general_seleccionada}'")
    print(f"  - turno: '{turno_seleccionado}'")
    print(f"  - operador_id: '{operador_id}'")
    if not all([fecha_str, area_general_seleccionada, turno_seleccionado, operador_id]):
        print(">>> ERROR: Faltan parámetros de filtro. Abortando.")
        return HttpResponse("Faltan parámetros de filtro.", status=400)
    operador_seleccionado = get_object_or_404(Empleado, pk=operador_id)
    try:
        
        fecha_seleccionada = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Formato de fecha inválido.", status=400)

    
    tz = timezone.get_current_timezone()
    start_of_day = timezone.make_aware(datetime.combine(fecha_seleccionada, time.min), tz)
    end_of_day = timezone.make_aware(datetime.combine(fecha_seleccionada, time.max), tz)

    
    
    registros_del_dia = Registro.objects.filter(
        fecRegistro__range=(start_of_day, end_of_day),
        id_dispositivo__area_general__iexact=area_general_seleccionada.strip(),
        id_dispositivo__tipoDisp='Portatil',
        turno=turno_seleccionado
    ).select_related(
        'trabajador_receptor__empresa',
        'operador_responsable',
        'id_dispositivo',
        'area_trabajo_operacion',
        'punto_exacto_operacion'
    ).prefetch_related('trabajador_receptor__telefono_set').order_by('fecRegistro')

    print(f"-> Consulta SQL (aproximada): \n{registros_del_dia.query}\n")
    print(f"\n-> Número de registros encontrados con TODOS los filtros combinados: {registros_del_dia.count()}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Registro {fecha_seleccionada.strftime('%Y-%m-%d')}"
    
    
    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses_ano = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    dia_semana_nombre = dias_semana[fecha_seleccionada.weekday()]
    mes_nombre = meses_ano[fecha_seleccionada.month]
    
    fecha_formateada = f"{dia_semana_nombre}, {fecha_seleccionada.day} de {mes_nombre} de {fecha_seleccionada.year}"
    
    sheet.merge_cells('A1:C1')
    sheet['A1'].value = area_general_seleccionada.upper()
    sheet['A1'].font = Font(bold=True, size=12)

    sheet['A3'].value = "Turno :"
    sheet['A4'].value = "Operador"
    sheet['B3'].value = turno_seleccionado
    operador_encargado = registros_del_dia.first().operador_responsable if registros_del_dia.exists() else None
    if operador_encargado:
        sheet['B4'].value = operador_encargado.nomEmpleado
    
    sheet.merge_cells('F1:I1')
    sheet['F1'].value = fecha_formateada
    sheet['F1'].font = Font(bold=True, size=12)
    sheet['F1'].alignment = Alignment(horizontal='right')

    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid') # Verde
    blue_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')  # Azul
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')# Amarillo
    orange_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')# Naranja
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    
    sheet.merge_cells('A6:B6'); sheet['A6'].value = 'ORDEN'
    sheet.merge_cells('C6:F6'); sheet['C6'].value = 'DATOS PERSONALES'
    sheet.merge_cells('G6:K6'); sheet['G6'].value = 'INFORMACIÓN DEL EQUIPO DETECTOR DE GASES'
    sheet.merge_cells('L6:N6'); sheet['L6'].value = 'INFORMACION DE LA ACTIVIDAD'
    sheet.merge_cells('O6:S6'); sheet['O6'].value = 'CONTROL DE EQUIPO'

    super_headers = ['A6', 'C6', 'G6', 'L6', 'O6']
    fills = [blue_fill, green_fill, blue_fill, yellow_fill, orange_fill]
    for i, cell_coord in enumerate(super_headers):
        cell = sheet[cell_coord]
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = fills[i]
        cell.border = thin_border

    
    headers = [
        'ITEM', 'FECHA', 'IDENTIFICACIÓN', 'NOMBRE',
        'EMPRESA', 'TELÉFONO', 'EQUIPO ENTREGADO',
        'ÁREA ASIGNADA AL EQUIPO', 'MODELO', 'UBICACIÓN DEL EQUIPO', 'EQUIPO DEVUELTO',
        'ÁREA DE TRABAJO', 'PUNTO EXACTO', 'DURACIÓN (HRS)', 'TURNO',
        'ADMIN', 'HORA ENTREGA', 'HORA DEVOLUCIÓN', 'ESTADO'
    ]
    sheet.append(headers)
    
    
    header_row_num = sheet.max_row
    for cell in sheet[header_row_num]:
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    
    for idx, registro in enumerate(registros_del_dia, 1):
        fecha_registro_peru = registro.fecRegistro.strftime('%d/%m/%Y')
        
        trabajador_receptor = registro.trabajador_receptor
        operador_responsable = registro.operador_responsable
        dispositivo = registro.id_dispositivo
        telefonos = ", ".join([t.numero for t in trabajador_receptor.telefono_set.all()])
        duracion_horas = (registro.durPrestamo.total_seconds() / 3600) if registro.durPrestamo else ''
        
        
        row_data = [
            idx, 
            fecha_registro_peru,    
            trabajador_receptor.dni,
            
            trabajador_receptor.nomEmpleado,
            trabajador_receptor.empresa.nombreE if trabajador_receptor.empresa else '',
            
            telefonos,
            dispositivo.tag, 
            registro.id_dispositivo.area_general,
            dispositivo.nomDisp, 
            registro.id_dispositivo.area_general,
            dispositivo.tag if registro.fecDevol else '',
            
            registro.area_trabajo_operacion.nombreA if registro.area_trabajo_operacion else '', 
            registro.punto_exacto_operacion.nombre_punto if registro.punto_exacto_operacion else '', 
            f'{duracion_horas:.2f}' if duracion_horas else '',
            registro.turno,
            operador_responsable.nomEmpleado if operador_responsable else '',
            registro.fecRegistro.time().strftime('%H:%M:%S'),
            registro.fecDevol.time().strftime('%H:%M:%S') if registro.fecDevol else '',
            "DEVUELTO" if registro.fecDevol else "PRESTADO"
        ]
        
        sheet.append(row_data)

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Registro_Diario_{fecha_seleccionada.strftime("%Y-%m-%d")}.xlsx"'
    workbook.save(response)
    
    return response


@login_required
def exportar_fijos_excel_certificado(request):
    
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dispositivos Fijos"
    
    
    super_header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    super_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    super_header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid') # Naranja oscuro
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
    sheet.merge_cells('H2:J2'); sheet['H2'].value = 'CONDICIONES AMBIENTALES'
    sheet.merge_cells('L2:Q2'); sheet['L2'].value = 'GASES DE CALIBRACIÓN UTILIZADO'
    sheet.merge_cells('R2:T2'); sheet['R2'].value = 'RESULTADOS'

    for col in ['H', 'L', 'R']:
        cell = sheet[f'{col}2']
        cell.font = super_header_font
        cell.alignment = super_header_alignment
        cell.border = thin_border
        cell.fill = super_header_fill

    headers = [
        'N°', 'MODELO DEL DETECTOR', 'ÁREA', 'UBICACIÓN EN ÁREA', 'Código SMCV:',
        'TIPO DE SENSOR', 'GASES Y RANGO DE MEDICIÓN', 'TEMPERATURA', 'PRESIÓN', 'HUMEDAD RELATIVA',
        'ESTADO', 'CILINDROS', 'GASES', 'N/P', 'Lote N°', 'N° CERTIFICADO', 'EXPIRACIÓN',
        'LECTURA PATRÓN', 'LECTURA DEL EQUIPO', '% ERROR TRAS CALIBRACIÓN',
        'FECHA DE CALIBRACIÓN POR CENERIS', 'PRÓXIMA CALIBRACIÓN'
    ]
    
    sheet.append(headers)
    
    
    for cell in sheet[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    dispositivos_fijos = Dispositivo.objects.filter(tipoDisp='Fijo').select_related(
        'id_areaTrabajo_fijo'
    ).prefetch_related(
        'certificados__componente__sensor__informes_calibracion__empresa_realizadora' ,'certificados__patronescalibracion_set', 'certificados__resultados_set'
    ).order_by('nomDisp')

    for idx, dispositivo in enumerate(dispositivos_fijos, 1):
        
        ultimo_certificado = dispositivo.certificados.order_by('-fechCertificado').first()

        if ultimo_certificado:
            # Obtenemos el sensor VINCULADO a ESE certificado
            if ultimo_certificado.componente and hasattr(ultimo_certificado.componente, 'sensor'):
                sensor = ultimo_certificado.componente.sensor
        
        patrones = ultimo_certificado.patronescalibracion_set.all() if ultimo_certificado else []
        resultados = ultimo_certificado.resultados_set.all() if ultimo_certificado else []

        estado_calibracion = 'No Calibrado' 
        if ultimo_certificado and ultimo_certificado.estadoFinal:
            estado_calibracion = 'Calibrado'

        row = [
            idx,
            dispositivo.nomDisp,
            dispositivo.area_general,
            dispositivo.id_areaTrabajo_fijo.nombreA if dispositivo.id_areaTrabajo_fijo else '',
            dispositivo.tag,
            sensor.nomComp if sensor else '',
            ultimo_certificado.rango_medicion if ultimo_certificado else '', 
            
            
            ultimo_certificado.temp if ultimo_certificado else '',
            ultimo_certificado.presion if ultimo_certificado else '',
            ultimo_certificado.humedadRelativa if ultimo_certificado else '',
            
            
            estado_calibracion,
            ", ".join([p.numPatron for p in patrones if p.numPatron]),
            ", ".join([p.patronUtil for p in patrones]),
            ", ".join([p.n_p for p in patrones]),
            ", ".join([p.n_lote for p in patrones]),
            ", ".join([p.n_certificado for p in patrones]),
            ", ".join([p.fechaExpiracion.strftime('%d/%m/%Y') for p in patrones if p.fechaExpiracion]),
            
            
            ", ".join([r.lecturaPatron for r in resultados]),
            ", ".join([r.lecturaEquipo for r in resultados]),
            ", ".join([r.prob_error for r in resultados]),
            
            
            ultimo_certificado.fechCertificado.strftime('%d/%m/%Y') if ultimo_certificado and ultimo_certificado.fechCertificado else '',
            ultimo_certificado.proxFecha.strftime('%d/%m/%Y') if ultimo_certificado and ultimo_certificado.proxFecha else '',
        ]
        
        sheet.append(row)


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Complejo_Fijos_{date.today().strftime("%Y-%m-%d")}.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def exportar_fijos_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte Dispositivos Fijos"
    
    # --- ESTILOS (sin cambios) ---
    super_header_font = Font(name='Arial', size=10, bold=True)
    super_header_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
    super_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid') 
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- CABECERAS (ligeramente ajustadas para que coincidan con la vista HTML) ---
    # Fila de Super-Headers
    # NOTA: Los headers se escriben manualmente para controlar las celdas combinadas.
    sheet.merge_cells('A1:H1'); sheet['A1'].value = 'INFORMACIÓN DEL EQUIPO'
    sheet.merge_cells('I1:K1'); sheet['I1'].value = 'CALIBRACIÓN ENCONTRADA'
    sheet.merge_cells('L1:M1'); sheet['L1'].value = 'FECHA VENCIMIENTO SENSOR'
    sheet.merge_cells('N1:O1'); sheet['N1'].value = 'INFO INFORME'
    sheet.merge_cells('P1:R1'); sheet['P1'].value = 'ALARMAS'
    sheet.merge_cells('S1:U1'); sheet['S1'].value = 'VALOR SPAM'
    sheet.merge_cells('V1:Y1'); sheet['V1'].value = 'ESTADO FINAL'

    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = super_header_font
            cell.fill = super_header_fill
            cell.alignment = super_header_alignment
            cell.border = thin_border
    sheet.row_dimensions[1].height = 30

    # Fila de Headers Principales
    headers = [
        'N°', 'NOMBRE DEL DISPOSITIVO','MODELO DE SENSOR', 'ÁREA', 'UBICACIÓN EN ÁREA', 'TAG', 'UBICACIÓN SENSOR',
        'TIPO DE GAS', 'INFORME', 'ENCONTRADO', 'SENSOR CAMBIADO',
        'MES', 'AÑO', 'FECHA', 'REALIZADA POR', '1RA', '2DA', '3RA', 'EQUIPO',
        'CILINDRO', 'UND', 'OBSERVACION', 'ESTADO CALIBRACIÓN', 'NRO CERTIFICADO',
        'FECHA CALIBRACIÓN CENERIS',
    ]
    sheet.append(headers)
    
    header_row_num = 2
    for cell in sheet[header_row_num]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- LÓGICA DE DATOS (NUEVA ESTRUCTURA) ---
    dispositivos_fijos = Dispositivo.objects.filter(tipoDisp='Fijo').select_related(
        'id_areaTrabajo_fijo'
    ).prefetch_related(
        'sensor_set__certificados_de_componente',
        'sensor_set__informes_calibracion__empresa_realizadora',
        'sensor_set__alarmas',
    ).order_by('nomDisp')
    
    current_row = header_row_num + 1
    
    for idx, dispositivo in enumerate(dispositivos_fijos, 1):
        sensores = list(dispositivo.sensor_set.all())
        num_sensores = len(sensores) if sensores else 1
        start_row_for_merge = current_row
        
        # Iteramos sobre los sensores para escribir las filas de datos
        if not sensores:
            sensores = [None] # Creamos una lista con un 'None' para que el bucle se ejecute una vez
        
        for sensor in sensores:

            alarma = None
            # Datos específicos del sensor
            if sensor:
                ultimo_certificado = sensor.certificados_de_componente.order_by('-fechCertificado').first()
                ultimo_informe = sensor.informes_calibracion.order_by('-fecha_informe').first()
                try:
                    alarma = Alarma.objects.get(sensor=sensor)
                except Alarma.DoesNotExist:
                    pass # Es normal, alarma se queda como None
                except Alarma.MultipleObjectsReturned:
                    # Si, a pesar de la limpieza, hay duplicados,
                    # tomamos el más reciente y continuamos.
                    print(f"ADVERTENCIA: Se encontraron múltiples alarmas para el sensor {sensor}. Se usará la más reciente.")
                    alarma = Alarma.objects.filter(sensor=sensor).order_by('-id_alarma').first()

                estado_calibracion = 'No Calibrado'
                if ultimo_certificado and timezone.now() - ultimo_certificado.fechCertificado <= timedelta(days=183):
                    estado_calibracion = 'Calibrado'
            else:
                ultimo_certificado, ultimo_informe, estado_calibracion = None, None, 'N/A'
            
            # Escribimos los datos de esta fila
            row_data = [
                idx, dispositivo.nomDisp,sensor.nomComp if sensor else 'Sin sensor', dispositivo.area_general,
                dispositivo.id_areaTrabajo_fijo.nombreA if dispositivo.id_areaTrabajo_fijo else '',
                dispositivo.tag, '',
                # Datos específicos del sensor
                
                sensor.tipGas if sensor else '',
                ultimo_informe.informe if ultimo_informe else '',
                ultimo_informe.encontrado_calibracion if ultimo_informe else '',
                "Sí" if ultimo_informe and ultimo_informe.sensor_cambiado else "No",
                sensor.fecVencGarantia.strftime('%B').capitalize() if sensor and sensor.fecVencGarantia else '',
                sensor.fecVencGarantia.year if sensor and sensor.fecVencGarantia else '',
                ultimo_informe.fecha_informe.strftime('%d/%m/%Y') if ultimo_informe and ultimo_informe.fecha_informe else '',
                ultimo_informe.empresa_realizadora.nombreE if ultimo_informe and ultimo_informe.empresa_realizadora else '',
                alarma.primera if alarma else '', alarma.segunda if alarma else '', alarma.tercera if alarma else '',
                alarma.equipo if alarma else '', alarma.cilindro if alarma else '', alarma.und if alarma else '',
                ultimo_informe.observacion if ultimo_informe else '',
                estado_calibracion,
                ultimo_certificado.nro_certificado if ultimo_certificado else '',
                ultimo_certificado.fechCertificado.date().strftime('%d/%m/%Y') if ultimo_certificado else '',
            ]
            sheet.append(row_data)
            current_row += 1

        # --- LA MAGIA: COMBINAR CELDAS VERTICALMENTE ---
        # Si hay más de una fila para este dispositivo, combinamos las celdas comunes.
        if num_sensores > 1:
            end_row_for_merge = start_row_for_merge + num_sensores - 1
            # Columnas comunes (índices de columna, 1-based)
            columnas_a_combinar_inicio = [1, 2]
            for col_idx in columnas_a_combinar_inicio:
                sheet.merge_cells(start_row=start_row_for_merge, start_column=col_idx,
                                  end_row=end_row_for_merge, end_column=col_idx)

            # Columnas comunes DESPUÉS de la última columna de sensor
            # Área (4) hasta UND (21), y luego Observación (22)
            # Adaptamos los rangos a la nueva estructura.
            # Los datos específicos de sensor ahora son: Modelo(3), Tipo Gas(8), Informe(9), ..., Fecha Cal Ceneris(25)
            # El resto es común.
            columnas_a_combinar_final = list(range(4, 8)) + list(range(16, 22))
            for col_idx in columnas_a_combinar_final:
                 sheet.merge_cells(start_row=start_row_for_merge, start_column=col_idx,
                                  end_row=end_row_for_merge, end_column=col_idx)

            # Centramos verticalmente todas las celdas combinadas
            for col_idx in (columnas_a_combinar_inicio + columnas_a_combinar_final):
                cell = sheet.cell(row=start_row_for_merge, column=col_idx)
                cell.alignment = Alignment(vertical='center')
    # --- RESPUESTA HTTP (sin cambios) ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Fijos_Completo_{date.today().strftime("%Y%m%d")}.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def exportar_programas_excel(request):
    ano_seleccionado = request.GET.get('ano', date.today().year) 
    tipo_dispositivo_seleccionado = request.GET.get('tipo', 'Fijo')
    
    programas = Programa.objects.filter(
        ano=ano_seleccionado,
        tipo_dispositivo=tipo_dispositivo_seleccionado
    ).order_by('mes')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Programa {ano_seleccionado}"

    
    bold_font_11 = Font(name='Arial', size=11, bold=True)
    bold_font_10 = Font(name='Arial', size=10, bold=True)
    normal_font_10 = Font(name='Arial', size=10)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid') # Naranja claro
    green_fill = PatternFill(start_color='E2F0D5', end_color='E2F0D5', fill_type='solid') # Verde claro

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))



    column_widths = {'A': 20, 'B': 25, 'C': 25, 'D': 10, 'E': 8, 'F': 8, 'G': 18, 'H': 18, 'I': 40}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    
    sheet.merge_cells('A2:H2')
    cell_a2 = sheet['A2']
    cell_a2.value = f'PROGRAMA DE CALIBRACIÓN DE DETECTORES {tipo_dispositivo_seleccionado.upper()} {ano_seleccionado}'
    cell_a2.font = bold_font_11
    cell_a2.alignment = center_align

    
    
    meta_data = {
        'OBJETIVO': 'Calibrar los detectores de fijos del  Proyecto de Gases de Cerro Verde',
        'Presupuesto': 'Proyecto de Gases de Cerro Verde',
        'Mecanismos Legales': 'Ley N° 29783 Ley de SST, D.S. N°005-2012-TR Reglamento de la Ley de SST, 050-2013-TR Registros obligatorios del SGSST',
        'Recursos': 'HHT, Laptop, Internet, Luz, Papel, Impresora , Controlador, Gases, Detectores',
    }
    
    for i, (key, value) in enumerate(meta_data.items(), 5):
        sheet[f'A{i}'].value = key
        sheet[f'A{i}'].font = bold_font_10
        sheet[f'A{i}'].border = thin_border
        sheet.merge_cells(f'B{i}:I{i}')
        sheet[f'B{i}'].value = value
        sheet[f'B{i}'].border = thin_border

    
    sheet.merge_cells('A10:A11'); sheet['A10'].value = 'CERTIFICACIÓN'
    sheet.merge_cells('B10:C11'); sheet['B10'].value = 'RESPONSABLE'
    sheet.merge_cells('D10:D11'); sheet['D10'].value = 'META'
    sheet.merge_cells('E10:F10'); sheet['E10'].value = 'AVANCE'
    sheet.merge_cells('G10:G11'); sheet['G10'].value = 'TOTAL PROGRAMADO'
    sheet.merge_cells('H10:H11'); sheet['H10'].value = 'TOTAL EJECUTADO'
    sheet.merge_cells('I10:I11'); sheet['I10'].value = 'COMENTARIOS'
    
    sheet['E11'].value = 'P'; sheet['F11'].value = 'E'

    
    for row in sheet.iter_rows(min_row=10, max_row=11, min_col=1, max_col=9):
        for cell in row:
            cell.font = bold_font_10
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
    
    
    current_row = 12 # Empezamos a escribir los datos de los meses en la fila 12
    total_programado_anual = 0
    total_ejecutado_anual = 0

    for programa in programas:
        
        sheet.cell(row=current_row, column=1).value = calendar.month_name[programa.mes].upper()
        sheet.cell(row=current_row, column=2).value = "SUPERVISOR" # Placeholder
        sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)

        sheet.cell(row=current_row, column=4).value = "100%" # META
        sheet.cell(row=current_row, column=5).value = "100%" # AVANCE P
        
        porcentaje_ejecutado = programa.porcentaje_progreso
        cell_avance_e = sheet.cell(row=current_row, column=6)
        cell_avance_e.value = f'{porcentaje_ejecutado}%' # AVANCE E
        cell_avance_e.fill = green_fill # Pintamos la celda de verde
        
        sheet.cell(row=current_row, column=7).value = programa.totalPrograma
        sheet.cell(row=current_row, column=8).value = programa.totalEjecutado
        sheet.cell(row=current_row, column=9).value = programa.comentarios

        
        total_programado_anual += programa.totalPrograma
        total_ejecutado_anual += programa.totalEjecutado

        
        for col_num in range(1, 10):
            sheet.cell(row=current_row, column=col_num).border = thin_border
            sheet.cell(row=current_row, column=col_num).alignment = center_align if col_num != 9 else left_align
        
        current_row += 1

    
    current_row += 1
    sheet.merge_cells(f'A{current_row}:F{current_row}')
    sheet[f'A{current_row}'].value = 'CUMPLIMIENTO ACTUAL DEL PROGRAMA:'
    sheet[f'A{current_row}'].font = bold_font_10
    sheet[f'A{current_row}'].alignment = Alignment(horizontal='right')

    sheet[f'G{current_row}'].value = total_programado_anual
    sheet[f'H{current_row}'].value = total_ejecutado_anual
    
    promedio_cumplimiento = 0
    if total_programado_anual > 0:
        promedio_cumplimiento = (total_ejecutado_anual / total_programado_anual) * 100
        
    sheet[f'I{current_row}'].value = f'{promedio_cumplimiento:.0f}%'
    sheet[f'I{current_row}'].font = bold_font_10

    
    for col_num in range(1, 10):
        sheet.cell(row=current_row, column=col_num).border = thin_border
    

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Programa_Calibracion_{ano_seleccionado}.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def exportar_mantenimiento_indice(request):
    
    modelos_disponibles = Dispositivo.objects.values_list('nomDisp', flat=True).distinct().order_by('nomDisp')
    
    context = {
        'modelos_disponibles': modelos_disponibles,
        'titulo': "Exportar Reporte de Mantenimiento"
    }
    return render(request, 'exportar/mantenimiento_indice.html', context)

@login_required
def exportar_mantenimiento_excel(request):
    modelo_seleccionado = request.GET.get('modelo', None)
    
    if not modelo_seleccionado:
        return HttpResponse("Error: Debe seleccionar un modelo de dispositivo.", status=400)
    
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Mantenimiento {modelo_seleccionado}"
    
    
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C55A11', end_color='C55A11', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    super_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
    dispositivo_ejemplo = Dispositivo.objects.filter(nomDisp=modelo_seleccionado).prefetch_related('partes', 'sensor_set').first()
    if not dispositivo_ejemplo:
        return HttpResponse(f"No se encontraron dispositivos del modelo '{modelo_seleccionado}'.", status=404)

    columnas_fijas = [
        'N°', 'MODELO', 'NÚMERO SERIE', 'FECHA DE FABRICACIÓN', 'FECHA DE INGRESO',
        'VENCIMIENTO DE GARANTÍA', 'CÓDIGO', 'SENSOR', 'UBICACIÓN',
        'ESTADO INICIAL DEL EQUIPO', 'FECHA DE INTERVENCIÓN', 'TÉCNICO A CARGO'
    ]
    
    partes_del_modelo = sorted([p.nomPart for p in dispositivo_ejemplo.partes.all()])
    sensores_del_modelo = sorted([f"Sensor {s.tipGas}" for s in dispositivo_ejemplo.sensor_set.all()])
    columnas_checklist = partes_del_modelo + sensores_del_modelo

    
    columnas_fotos_evidencia = ['Fotos de evidencia1', 'Fotos de evidencia2', 'Fotos de evidencia3']

    columnas_finales = [
        'Componentes en mal estado', 'Componentes en estado regular',
        'Cambios Realizados', 'ESTADO DEL EQUIPO'
    ]
    
    headers_completos = columnas_fijas + columnas_checklist + columnas_fotos_evidencia + columnas_finales

    
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_completos))
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"MANTENIMIENTO DE DETECTORES PORTÁTILES ({modelo_seleccionado.upper()})"
    title_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    title_cell.fill = header_fill
    title_cell.alignment = super_header_alignment
    title_cell.border = thin_border
    
    sheet.append(headers_completos)
    header_row_num = sheet.max_row
    for col_idx, cell in enumerate(sheet[header_row_num], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        
        
        if get_column_letter(col_idx) in ['S', 'T', 'U']: # Ajusta estas letras según la posición de las fotos.
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15
        else:
            sheet.column_dimensions[get_column_letter(col_idx)].width = 12

    
    dispositivos_del_modelo = Dispositivo.objects.filter(nomDisp=modelo_seleccionado).prefetch_related(
        'mantenimientos__tecnico_a_cargo',
        'partes',
        'sensor_set'
    ).order_by('num_serie')

    for idx, dispositivo in enumerate(dispositivos_del_modelo, 1):
        ultimo_mantenimiento = dispositivo.mantenimientos.first()
        sensores_str = ", ".join([s.tipGas for s in dispositivo.sensor_set.all() if s.tipGas])
        
        fotos_evidencia = []
        if ultimo_mantenimiento:
            fotos_evidencia = dispositivo.fotos.filter(
                contexto='MANTENIMIENTO'
            ).order_by('fecha_carga')[:3]

        row_data = [
            idx,
            dispositivo.nomDisp,
            dispositivo.num_serie,
            dispositivo.fecFabricacion,
            dispositivo.fecIngreso,
            dispositivo.fecVencimientoGarantia,
            dispositivo.tag,
            sensores_str,
            dispositivo.area_general if dispositivo.area_general else '',
            ultimo_mantenimiento.estado_inicial_equipo if ultimo_mantenimiento else 'Operativo',
            ultimo_mantenimiento.fecha_intervencion.strftime('%d/%m/%Y') if ultimo_mantenimiento else '',
            ultimo_mantenimiento.tecnico_a_cargo.nomEmpleado if ultimo_mantenimiento and ultimo_mantenimiento.tecnico_a_cargo else ''
        ]
        
        checklist_data = ultimo_mantenimiento.checklist_partes if ultimo_mantenimiento else {}
        for item_columna in columnas_checklist:
            datos_parte = checklist_data.get(item_columna, {})
            estado_parte = datos_parte.get('estado', 'N/A')
            comentario_parte = datos_parte.get('comentario', '')
            celda_texto = estado_parte
            if comentario_parte:
                celda_texto += f":\n{comentario_parte}"
            row_data.append(celda_texto)

        
        row_data.extend([''] * len(columnas_fotos_evidencia))
        
        
        if ultimo_mantenimiento:
            row_data.extend([
                ultimo_mantenimiento.componentes_mal_estado,
                ultimo_mantenimiento.componentes_estado_regular,
                ultimo_mantenimiento.cambios_realizados,
                ultimo_mantenimiento.estado_final_equipo
            ])
        else:
            row_data.extend([''] * len(columnas_finales))

        sheet.append(row_data) 
        current_row = sheet.max_row
        
        
        
        start_col_photos_idx = len(columnas_fijas) + len(columnas_checklist) + 1
        
        
        sheet.row_dimensions[current_row].height = 85 

        
        for i, foto in enumerate(fotos_evidencia):
            
            col_idx_photos = start_col_photos_idx + i
            col_letter_photos = get_column_letter(col_idx_photos)

            if foto.imagen_original and os.path.exists(foto.imagen_original.path):
                try:
                    img = OpenpyxlImage(foto.imagen_original.path)
                    
                    
                    img.height = 60 # Altura de la imagen en píxeles
                    img.width = 90 # Ancho de la imagen en píxeles
                    
                    sheet.add_image(img, f"{col_letter_photos}{current_row}")
                except Exception as e:
                    print(f"Error al insertar imagen: {e}")
                    sheet.cell(row=current_row, column=col_idx_photos).value = "Error img"
            else:
                 sheet.cell(row=current_row, column=col_idx_photos).value = "No img"

    
    for row in sheet.iter_rows(min_row=3, min_col=len(columnas_fijas) + 1, max_col=len(columnas_fijas) + len(columnas_checklist)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border
    
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Mantenimiento_{modelo_seleccionado}.xlsx"'
    workbook.save(response)
    return response



@login_required
def exportar_fijos_por_area(request, area_general):
    
    
    dispositivos = Dispositivo.objects.filter(
        tipoDisp='Fijo', 
        #aqui va area general
    ).order_by('tag').prefetch_related('sensor_set', 'fotos')

    if not dispositivos.exists():
        return HttpResponse(f"No se encontraron dispositivos fijos en el área '{area_general}'.")

    tipos_sensor_unicos = set()
    for dispositivo in dispositivos:
        for sensor in dispositivo.sensor_set.all():
            if sensor.tipGas:
                tipos_sensor_unicos.add(sensor.tipGas)
    
    filas_reporte = sorted(list(tipos_sensor_unicos)) + ['EVIDENCIA']
    
    print(f"Filas a generar en el reporte: {filas_reporte}")

    fotos_dict = {}
    fotos_relevantes = FotoDispositivo.objects.filter(
        dispositivo__in=dispositivos,
        tipo_foto__in=filas_reporte
    )
    for foto in fotos_relevantes:
        fotos_dict[(foto.dispositivo_id, foto.tipo_foto)] = foto.imagen_original.path

    workbook = Workbook()
    sheet = workbook.active
    
    
    title_font = Font(name='Calibri', size=16, bold=True)
    title_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    title_alignment = Alignment(horizontal='center', vertical='center')

    header_font = Font(name='Calibri', size=11, bold=True)
    detector_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    sensor_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    ubicacion_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    header_col_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    content_header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dispositivos) + 2)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"PLANTA CONCENTRADORA {area_general.upper()}"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = thin_border
    
    
    sheet.cell(row=2, column=1).value = "MODELO DETECTOR"
    sheet.cell(row=3, column=1).value = "MODELO SENSOR"
    sheet.cell(row=4, column=1).value = "UBICACIÓN"

    for row_idx in range(2, 5):
        cell = sheet.cell(row=row_idx, column=1)
        cell.font = header_font
        cell.fill = detector_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    
    for col_idx, dispositivo in enumerate(dispositivos, 2):
        
        cell_r2 = sheet.cell(row=2, column=col_idx)
        cell_r2.value = dispositivo.nomDisp
        cell_r2.font = header_font
        cell_r2.fill = header_col_fill
        cell_r2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_r2.border = thin_border

        
        modelos_sensor = ", ".join([s.nomComp for s in dispositivo.sensor_set.all()])
        cell_r3 = sheet.cell(row=3, column=col_idx)
        cell_r3.value = modelos_sensor
        cell_r3.font = header_font
        cell_r3.fill = header_col_fill
        cell_r3.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_r3.border = thin_border

        
        cell_r4 = sheet.cell(row=4, column=col_idx)
        cell_r4.value = dispositivo.tag
        cell_r4.font = header_font
        cell_r4.fill = header_col_fill
        cell_r4.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_r4.border = thin_border
        
        sheet.column_dimensions[get_column_letter(col_idx)].width = 25

    
    
    
    
    start_row = 5
    for row_idx, tipo_fila in enumerate(filas_reporte, start_row):
        
        cell_header_dinamica = sheet.cell(row=row_idx, column=1)
        cell_header_dinamica.value = tipo_fila
        cell_header_dinamica.font = header_font
        cell_header_dinamica.fill = content_header_fill
        cell_header_dinamica.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_header_dinamica.border = thin_border

        sheet.row_dimensions[row_idx].height = 120
        
        for col_idx, dispositivo in enumerate(dispositivos, 2):
            imagen_path = fotos_dict.get((dispositivo.pk, tipo_fila))
            
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border # Borde para las celdas de las imágenes

            if imagen_path and os.path.exists(imagen_path):
                try:
                    img = OpenpyxlImage(imagen_path)
                    img.height = 150
                    img.width = 150
                    cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                    sheet.add_image(img, cell_address)
                except Exception as e:
                    cell.value = "Error img"
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.value = "No hay evidencia"
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Visual_Fijos_{area_general}.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def exportar_cardex_excel(request, dispositivo_id):
    
    # --- 1. CONFIGURACIÓN INICIAL ---
    dispositivo = get_object_or_404(Dispositivo, pk=dispositivo_id)
    template_path = os.path.join(settings.BASE_DIR, 'cenerisapp', 'templates', 'excel_templates', 'plantilla_cardex.xlsx')
    workbook = load_workbook(template_path)
    
    # ---------------------------------------------------------
    # HOJA 1: SENSORES
    # ---------------------------------------------------------
    sheet_sensores = workbook.active
    sheet_sensores.title = "SENSORES"

    # --- ESTILOS ---
    header_black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid') 
    header_dark_fill = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
    header_grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    
    white_font_bold = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    grey_header_font = Font(name='Arial', size=9, bold=True)
    data_font = Font(name='Arial', size=9)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    empty_border = Border() # Borde vacío para limpiezas

    # =================================================================================
    # --- 2. DATOS FIJOS (ENCABEZADO) ---
    # =================================================================================
    
    def llenar_datos_fijos(hoja):
        hoja['A2'] = dispositivo.id_dispositivo
        hoja['B2'] = dispositivo.nomDisp
        hoja['C2'] = dispositivo.num_serie
        hoja['D2'] = dispositivo.tag

        fecFab = dispositivo.fecFabricacion.replace(tzinfo=None) if isinstance(dispositivo.fecFabricacion, datetime) else dispositivo.fecFabricacion
        fecIng = dispositivo.fecIngreso.replace(tzinfo=None) if isinstance(dispositivo.fecIngreso, datetime) else dispositivo.fecIngreso
        fecGar = dispositivo.fecVencimientoGarantia.replace(tzinfo=None) if isinstance(dispositivo.fecVencimientoGarantia, datetime) else dispositivo.fecVencimientoGarantia
        
        hoja['A4'] = fecFab; hoja['B4'] = fecIng; hoja['C4'] = fecGar

        sensores_str = ", ".join([s.tipGas for s in dispositivo.sensor_set.all() if s.tipGas])
        hoja['A7'] = sensores_str
        hoja['B7'] = dispositivo.area_general if dispositivo.area_general else ''

        ultimo_cert = dispositivo.certificados.order_by('-fechCertificado').first()
        if ultimo_cert:
            f_cert = ultimo_cert.fechCertificado
            hoja['C7'] = f_cert.replace(tzinfo=None) if isinstance(f_cert, datetime) else f_cert
        
        hoja['D7'] = dispositivo.estadoD
        
        # Foto General
        foto_equipo = dispositivo.fotos.filter(tipo_foto='EVIDENCIA', modificacion__isnull=True).order_by('-fecha_carga').first()
        if foto_equipo and foto_equipo.imagen_original:
            try:
                image_data = BytesIO(foto_equipo.imagen_original.read())
                img = OpenpyxlImage(image_data)
                img.height = 120; img.width = 120
                hoja.add_image(img, 'E2')
            except Exception: pass

    llenar_datos_fijos(sheet_sensores)

    # =================================================================================
    # --- 3. CLASIFICACIÓN DE DATOS ---
    # =================================================================================
    
    todas_modificaciones = dispositivo.modificacion_set.select_related(
        'id_trabajador', 'sensor_saliente', 'parte_saliente', 'componente_entrante'
    ).order_by('fecInstalacionMod') # Ordenamos cronológicamente (antiguo a nuevo) para el historial

    mods_sensores = defaultdict(list)
    mods_partes = defaultdict(list) # Clave: Nombre de Parte, Valor: Lista de Mods

    for mod in todas_modificaciones:
        es_sensor = False
        gas_key = None
        parte_key = None

        # Detección: ¿Es Sensor?
        if mod.sensor_saliente:
            gas_key = mod.sensor_saliente.tipGas; es_sensor = True
        elif mod.componente_entrante and hasattr(mod.componente_entrante, 'sensor'):
            gas_key = mod.componente_entrante.sensor.tipGas; es_sensor = True
        
        # Detección: ¿Es Parte?
        if not es_sensor:
            # Determinamos el nombre de la parte para agrupar
            if mod.parte_saliente:
                parte_key = mod.parte_saliente.nomPart
            elif mod.componente_entrante:
                parte_key = mod.componente_entrante.nomComp
            
            if parte_key:
                mods_partes[parte_key].append(mod)
        else:
            if gas_key:
                mods_sensores[gas_key].append(mod)

    # =================================================================================
    # --- 4. DIBUJAR HOJA 1 (SENSORES) ---
    # =================================================================================
    
    keys_sensores = sorted(mods_sensores.keys())
    
    for i, gas in enumerate(keys_sensores):
        col_base = 1 + (i * 2) 
        col_L = get_column_letter(col_base)     
        col_D = get_column_letter(col_base + 1) 
        curr_row = 10
        
        # Invertimos el orden para mostrar el más reciente arriba en sensores (opcional, segun preferencia)
        # Pero normalmente un historial va bajando. Usaremos el orden de la query.
        
        for mod in mods_sensores[gas]:
            # Cabecera Sensor
            sheet_sensores.merge_cells(f'{col_L}{curr_row}:{col_D}{curr_row}')
            c = sheet_sensores[f'{col_L}{curr_row}']; c.value = f"SENSOR {gas}"; c.font = white_font_bold; c.fill = header_dark_fill; c.alignment = center_align; c.border = thin_border
            
            sheet_sensores.merge_cells(f'{col_L}{curr_row+1}:{col_D}{curr_row+1}')
            c = sheet_sensores[f'{col_L}{curr_row+1}']; c.value = "FOTO SENSOR"; c.font = white_font_bold; c.fill = header_dark_fill; c.alignment = center_align; c.border = thin_border
            
            # Foto
            sheet_sensores.merge_cells(f'{col_L}{curr_row+2}:{col_D}{curr_row+3}')
            sheet_sensores.row_dimensions[curr_row+2].height = 60
            sheet_sensores.row_dimensions[curr_row+3].height = 60
            
            foto_cargada = False
            foto_mod = mod.fotos.first()
            if foto_mod and foto_mod.imagen_original:
                try:
                    img = OpenpyxlImage(BytesIO(foto_mod.imagen_original.read()))
                    img.height = 110; img.width = 110
                    sheet_sensores.add_image(img, f'{col_L}{curr_row+2}')
                    foto_cargada = True
                except: pass
            
            if not foto_cargada:
                c = sheet_sensores[f'{col_L}{curr_row+2}']; c.value = "SIN REGISTRO"; c.alignment = center_align
                for r in range(curr_row+2, curr_row+4):
                    sheet_sensores[f'{col_L}{r}'].border = thin_border; sheet_sensores[f'{col_D}{r}'].border = thin_border

            # Detalles
            r_det = curr_row + 4
            sheet_sensores.cell(row=r_det+1, column=col_base, value="N° Serie:").font=grey_header_font; sheet_sensores.cell(row=r_det+1, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+1, column=col_base).border=thin_border
            ns = mod.componente_entrante.sensor.nSerieActual if (mod.componente_entrante and hasattr(mod.componente_entrante, 'sensor')) else ''
            sheet_sensores.cell(row=r_det+1, column=col_base+1, value=ns).border=thin_border
            
            sheet_sensores.cell(row=r_det+2, column=col_base, value="Motivo:").font=grey_header_font; sheet_sensores.cell(row=r_det+2, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+2, column=col_base).border=thin_border
            sheet_sensores.cell(row=r_det+2, column=col_base+1, value=mod.MotivoCambio).border=thin_border
            
            sheet_sensores.cell(row=r_det+3, column=col_base, value="Resp:").font=grey_header_font; sheet_sensores.cell(row=r_det+3, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+3, column=col_base).border=thin_border
            sheet_sensores.cell(row=r_det+3, column=col_base+1, value=mod.id_trabajador.nomEmpleado).border=thin_border
            
            sheet_sensores.cell(row=r_det+4, column=col_base, value="Fecha Inst:").font=grey_header_font; sheet_sensores.cell(row=r_det+4, column=col_base).fill=header_grey_fill; sheet_sensores.cell(row=r_det+4, column=col_base).border=thin_border
            dt = mod.fecInstalacionMod
            sheet_sensores.cell(row=r_det+4, column=col_base+1, value=dt.replace(tzinfo=None) if isinstance(dt, datetime) else dt).border=thin_border

            curr_row += 10

    # =================================================================================
    # --- 5. GENERAR HOJAS DE MANTENIMIENTO (AGRUPADAS POR PARTE) ---
    # =================================================================================
    
    # Iteramos sobre cada tipo de parte. Ej: 'Carcasa' -> [Mod1, Mod2]
    for index, (nombre_parte, lista_mods) in enumerate(mods_partes.items(), start=1):
        
        # 1. Crear Hoja para esa Parte
        nueva_hoja = workbook.copy_worksheet(sheet_sensores)
        # Nombre de hoja seguro (Excel max 31 chars)
        safe_title = f"MANT {index} {nombre_parte}"[:30]
        nueva_hoja.title = safe_title
        
        # 2. LIMPIEZA
        for merged_range in list(nueva_hoja.merged_cells.ranges):
            if merged_range.min_row >= 10:
                nueva_hoja.unmerge_cells(str(merged_range))

        empty_fill = PatternFill()
        for row in nueva_hoja.iter_rows(min_row=10, max_row=300): # Limpiamos más filas por si acaso
            for cell in row:
                cell.value = None
                cell.fill = empty_fill
                cell.border = empty_border 
        
        llenar_datos_fijos(nueva_hoja) # Restaurar foto general

        # 3. DIBUJAR HISTORIAL VERTICAL (APILADO)
        # Empezamos en la Fila 11
        current_row = 11

        for mod in lista_mods:
            
            # --- TÍTULO (Fila 11 relativa) ---
            # Ocupa A-F
            nueva_hoja.merge_cells(f'A{current_row}:F{current_row}')
            cell = nueva_hoja[f'A{current_row}']
            # Título dinámico: Nombre de la parte + (Fecha) para diferenciar
            cell.value = f"KIT DE MANTENIMIENTO: {nombre_parte} ({mod.fecInstalacionMod.strftime('%d/%m/%Y')})"
            cell.font = white_font_bold; cell.fill = header_black_fill; cell.alignment = center_align; cell.border = thin_border
            
            # --- FOTO (Fila 12 relativa) ---
            row_foto = current_row + 1
            nueva_hoja.merge_cells(f'A{row_foto}:F{row_foto}')
            nueva_hoja.row_dimensions[row_foto].height = 250 
            
            foto_mod = mod.fotos.first()
            foto_insertada = False
            if foto_mod and foto_mod.imagen_original:
                try:
                    img = OpenpyxlImage(BytesIO(foto_mod.imagen_original.read()))
                    img.height = 300; img.width = 450 
                    nueva_hoja.add_image(img, f'A{row_foto}')
                    foto_insertada = True
                except: pass
            
            if not foto_insertada:
                cell = nueva_hoja[f'A{row_foto}']
                cell.value = "SIN REGISTRO FOTOGRAFICO"
                cell.alignment = center_align
                for c_idx in range(1, 7): 
                     col_l = get_column_letter(c_idx)
                     nueva_hoja[f'{col_l}{row_foto}'].border = thin_border

            # --- DETALLES (Filas 13-16 relativas) ---
            def escribir_fila(fila_abs, etiqueta, valor):
                nueva_hoja.merge_cells(f'A{fila_abs}:C{fila_abs}')
                lbl = nueva_hoja[f'A{fila_abs}']
                lbl.value = etiqueta
                lbl.font = grey_header_font; lbl.fill = header_grey_fill; lbl.alignment = center_align; lbl.border = thin_border
                nueva_hoja[f'B{fila_abs}'].border = thin_border; nueva_hoja[f'C{fila_abs}'].border = thin_border

                nueva_hoja.merge_cells(f'D{fila_abs}:F{fila_abs}')
                val_c = nueva_hoja[f'D{fila_abs}']
                val_c.value = valor
                val_c.font = data_font; val_c.alignment = center_align; val_c.border = thin_border
                nueva_hoja[f'E{fila_abs}'].border = thin_border; nueva_hoja[f'F{fila_abs}'].border = thin_border

            # Fila 13: Motivo
            escribir_fila(current_row + 2, "Motivo de Cambio:", mod.MotivoCambio)
            
            # Fila 14: Responsable
            resp = mod.id_trabajador.nomEmpleado if mod.id_trabajador else ""
            escribir_fila(current_row + 3, "Responsable:", resp)
            
            # Fila 15: Fecha Instalación
            dt = mod.fecInstalacionMod
            f_str = dt.strftime('%d/%m/%Y') if dt else ""
            escribir_fila(current_row + 4, "Fecha de instalación:", f_str)
            
            # Fila 16: Fecha Facturación
            escribir_fila(current_row + 5, "Fecha de facturación:", "")

            # --- SALTO PARA EL SIGUIENTE CAMBIO DE LA MISMA PARTE ---
            # La tarjeta ocupó 6 filas (1 titulo + 1 foto + 4 datos).
            # Dejamos 2 filas de espacio libre antes de la siguiente.
            current_row += 8 

    # GUARDAR
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="CARDEX_{dispositivo.tag}.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def exportar_reportes_excel(request):
    """
    Genera un archivo Excel con el historial de todos los reportes de daños y pérdidas.
    """
    
    ano_seleccionado = request.GET.get('ano')
    if not ano_seleccionado:
        return HttpResponse("Error: Debe seleccionar un año para exportar.", status=400)
    
    
    reportes = Reporte.objects.filter(
        fecReport__year=ano_seleccionado
    ).select_related(
        'id_dispositivo__id_empresa', 
        'id_otro_componente'
    ).order_by('-fecReport')

    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Historial de Reportes"

    
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    header_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid') # Naranja claro
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    data_alignment = Alignment(wrap_text=True, vertical='center')
    
    
    headers = ['FECHA DE REPORTE', 'EMPRESA', 'EQUIPO', 'DESCRIPCION', 'ESTADO']
    sheet.append(headers)
    
    
    column_widths = {'A': 20, 'B': 25, 'C': 40, 'D': 50, 'E': 15}
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    
    for reporte in reportes:
        
        
        nombre_empresa = 'N/A'
        nombre_equipo = 'N/A'
        estado_final = 'N/A'

        if reporte.id_dispositivo:
            
            dispositivo = reporte.id_dispositivo
            nombre_equipo = f"{dispositivo.nomDisp} ({dispositivo.num_serie})"
            if dispositivo.id_empresa:
                nombre_empresa = dispositivo.id_empresa.nombreE
            estado_final = dispositivo.estadoD
            
        elif reporte.id_otro_componente:
            
            componente = reporte.id_otro_componente
            nombre_equipo = f"{componente.nomComp} ({componente.nSerieActual})"
            
            if componente.inventario and componente.inventario.id_trabajador:
                nombre_empresa = "Empresa del Lote" # Placeholder
            estado_final = componente.estComp

        
        
        row_data = [
            reporte.fecReport,
            nombre_empresa,
            nombre_equipo,
            reporte.razRetiro, # 'razRetiro' parece ser la descripción
            estado_final
        ]
        
        sheet.append(row_data)
        
        last_row = sheet.max_row
        
        
        for cell in sheet[last_row]:
            cell.border = thin_border
            cell.alignment = data_alignment # Aplicar el estilo de alineación de datos
            cell.font = Font(name='Calibri', size=11) # Opcional: para mantener la consistencia

    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Historial_Reportes_{ano_seleccionado}.xlsx"'
    workbook.save(response)
    
    return response

@login_required
def exportar_seguimiento_excel(request):
    # 1. OBTENER FILTROS
    try:
        ano = int(request.GET.get('ano', date.today().year))
        mes = int(request.GET.get('mes', date.today().month))
        area_general = request.GET.get('area_general')
    except (ValueError, TypeError):
        return HttpResponse("Parámetros de año o mes inválidos.", status=400)
    
    if not area_general:
        return HttpResponse("Debe seleccionar un Área General.", status=400)

    # 2. PREPARAR DATOS DE FECHAS
    nombre_mes = calendar.month_name[mes].upper()
    num_dias = calendar.monthrange(ano, mes)[1]
    dias_del_mes = [date(ano, mes, dia) for dia in range(1, num_dias + 1)]

    # 3. OBTENER DATOS DE LA BASE DE DATOS
    dispositivos = Dispositivo.objects.filter(area_general=area_general, tipoDisp='Portatil').order_by('nomDisp')
    seguimientos = SeguimientoDiario.objects.filter(
        dispositivo__in=dispositivos,
        fecha__year=ano,
        fecha__month=mes
    )
    seguimiento_dict = {(s.dispositivo_id, s.fecha): s.estado_texto for s in seguimientos}

    # 4. INICIALIZAR EXCEL Y ESTILOS
    workbook = Workbook()
    sheet = workbook.active
    # Top Header Style (for "SEGUIMIENTO EQUIPOS...")
    header_font = Font(bold=True, size=14, color='000000') # Negro para mejor contraste
    header_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid') # El color solicitado
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fijos_font = Font(bold=True, size=8, color='000000')
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 1. Título principal
    # Combina las primeras 3 columnas de la fila 1 para el título
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = sheet.cell(row=1, column=1, value=f"SEGUIMIENTO EQUIPOS {area_general.upper()}")
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border


    # 2. Nombre del mes
    # Combina las siguientes celdas de la fila 1 para el mes
    sheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=5 + num_dias)
    month_cell = sheet.cell(row=1, column=6, value=nombre_mes)
    month_cell.font = header_font
    month_cell.fill = header_fill
    month_cell.alignment = center_alignment
    month_cell.border = thin_border

    # 3. Headers detallados (se mantienen en la fila 3)
    headers_fijos = ['N°', 'MODELO', 'SERIE', 'UBICACIÓN', 'ESTADO']
    headers_dias = [d.strftime('%d/%m/%Y') for d in dias_del_mes]
    # Agrega los headers a la tercera fila (índice 3 en Excel)
    sheet.append(headers_fijos + headers_dias)
    
    detailed_header_row_num = sheet.max_row

    # Iterate through the cells in that row and apply the styles
    for cell in sheet[detailed_header_row_num]:
        cell.font = fijos_font
        cell.fill = header_fill    
        cell.alignment = center_alignment
        cell.border = thin_border


    # 6. ESCRIBIR DATOS
    for idx, dispositivo in enumerate(dispositivos, 1):
        row_data = [
            idx,
            dispositivo.nomDisp,
            dispositivo.num_serie,      
            dispositivo.area_general if dispositivo.area_general else '',
            dispositivo.estadoD
        ]
        
        for dia in dias_del_mes:
            estado_del_dia = seguimiento_dict.get((dispositivo.pk, dia), '')
            row_data.append(estado_del_dia)
            
        sheet.append(row_data)
        current_row = sheet.max_row
    
        # Itera sobre todas las celdas de la fila actual y aplica los estilos
        for cell in sheet[current_row]:
            cell.alignment = center_alignment
            cell.border = thin_border

    # 7. FINALIZACIÓN Y RESPUESTA
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Seguimiento_{area_general}_{ano}-{mes:02d}.xlsx"'
    workbook.save(response)
    
    return response

