# calidad/views.py
# --- 1. Importaciones de Python Estándar ---
import json
import datetime
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import threading
import os
from collections import defaultdict
import logging
import traceback
import sys

# --- 2. Importaciones de Django ---
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Q, Max, Subquery, OuterRef, Value
from django.db.models.functions import TruncMonth, Concat
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.urls import reverse
from recursoshumanos.models import Trabajador, Empresa 
from django.conf import settings
from django.forms import HiddenInput
from django.core.validators import RegexValidator
from openpyxl.drawing.image import Image as OpenpyxlImage
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string

# --- 3. Importaciones de Librerías de Terceros (openpyxl) ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- 4. Importaciones de tu Propio Proyecto ---
# Modelos
from .models import Clinica, Control, EMO, ArchivoAdjuntoEMO, ArchivoClinica
from recursoshumanos.models import Empresa, Proyecto, Cargo, Trabajador, Area, Sede
from recursoshumanos.forms import ProyectoForm
# Formularios (¡Lista Única y Limpia!)
from .forms import (
    ProgramarEmoForm, 
    RegistrarResultadoEmoForm,
    ControlForm, 
    EmpresaForm, 
    ClinicaForm, 
    CargoForm,
    SubsanacionEmoForm,
    ConfirmacionLecturaForm,
    RegistrarControlForm,
    ArchivoClinicaForm
)
# Otros
from .decorators import group_required
from django.utils import timezone

logger = logging.getLogger(__name__)


def _trabajador_es_notificable(trabajador, fecha_referencia=None):
    """
    Un trabajador es notificable si está activo y no tiene fecha de cese vencida.
    """
    if not trabajador or not trabajador.activo:
        return False

    ref = fecha_referencia or timezone.now().date()
    if trabajador.fecha_cese and trabajador.fecha_cese <= ref:
        return False

    return True




@login_required
@group_required('Calidad')
def gestion_calidad(request):
    """
    Dashboard de tarjetas para la app Calidad. Filtra el contenido
    y establece títulos dinámicos según el parámetro 'view'.
    """
    current_view = request.GET.get('view', 'dashboard_calidad')

    # Diccionario para almacenar los títulos y subtítulos
    page_info = {
        'title': 'Panel de Gestión de Calidad',
        'subtitle': 'Bienvenido. Selecciona una opción para continuar.'
    }

    # Cambiamos los textos según la vista seleccionada
    if current_view == 'gestion_trabajadores':
        page_info['title'] = 'Gestión de Trabajadores'
        page_info['subtitle'] = 'Herramientas para registrar y consultar la información del personal.'
    elif current_view == 'gestion_certificados':
        page_info['title'] = 'Gestionar EMOs'
        page_info['subtitle'] = 'Herramienta para registrar resultados y programar EMOs'
    elif current_view == 'filtro_emos':
        page_info['title'] = 'Listas de EMOs'
        page_info['subtitle'] = 'Consulta el estado de EMOs de los empleados'

    context = {
        'current_view': current_view,
        'page_info': page_info, # Pasamos el diccionario completo a la plantilla
    }
    return render(request, 'calidad/dashboards/dashboard_calidad.html', context)



@group_required('Calidad')
def lista_emos(request):
    """
    Muestra una lista con el ESTADO ACTUAL del último EMO registrado
    para cada trabajador, con filtros interactivos por estado.
    """

    # --- FILTRADO DE DATOS ---
    
    # Obtiene el parámetro de filtro de la URL. Si no hay, no se filtra nada.
    filtro_estado = request.GET.get('filtro_estado', None)
    
    # Queryset base: Obtiene el último EMO de cada trabajador.
    ultimos_emos = EMO.objects.order_by(
        'trabajador__dni', 
        '-fecha_realizacion', 
        '-fecha_programada'
    ).distinct('trabajador__dni')

    # Aplicamos el filtro de estado si se ha seleccionado uno.
    hoy = timezone.now().date()
    limite_30_dias = hoy + timezone.timedelta(days=30)
    
    if filtro_estado == 'vigente':
        ultimos_emos = ultimos_emos.filter(estado='Realizado', fecha_vencimiento__gte=hoy)
    elif filtro_estado == 'por_vencer':
        ultimos_emos = ultimos_emos.filter(estado='Realizado', fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite_30_dias)
    elif filtro_estado == 'vencido':
        ultimos_emos = ultimos_emos.filter(estado='Realizado', fecha_vencimiento__lt=hoy)
    elif filtro_estado == 'programado':
        ultimos_emos = ultimos_emos.filter(estado='Programado')
    
    context = {
        'current_view': 'filtro_emos',
        'titulo': 'Estado actual de Vigencia de EMOs',
        'emos': ultimos_emos,
        'filtro_activo': filtro_estado, # Pasamos el filtro actual a la plantilla
    }
    return render(request, 'calidad/reportes/lista_emos.html', context)



# --- NUEVA VISTA PARA EL DASHBOARD DE ESTADÍSTICAS ---

@login_required
@group_required('Calidad')
def registrar_resultado_emo(request, emo_id):
    # Query normal
    query_condition = Q(estado='Programado') | Q(estado='Realizado')
    emo_programado = get_object_or_404(EMO, query_condition, id=emo_id)

    if request.method == 'POST':
        # --- LOGGING FORZADO ---
        print("\n" + "x"*50, flush=True)
        print(f"[DEBUG] INICIO POST - Editar EMO ID: {emo_id}", flush=True)
        
        # 1. Ver qué llega exactamente en el HTML (Raw Data)
        # Los checkboxes envían 'on' si están marcados, o nada si no lo están.
        checkbox_raw = request.POST.get('eliminar_archivo_pdf')
        archivo_raw = request.FILES.get('archivo_pdf')
        
        print(f"[DEBUG] Checkbox 'eliminar_archivo_pdf' RAW: '{checkbox_raw}'", flush=True)
        print(f"[DEBUG] Archivo 'archivo_pdf' RAW: {archivo_raw}", flush=True)

        form = RegistrarResultadoEmoForm(request.POST, request.FILES, instance=emo_programado)

        if form.is_valid():
            print("[DEBUG] Formulario VÁLIDO. Procesando...", flush=True)
            
            # Guardamos commit=False
            emo_realizado = form.save(commit=False)
            
            # 2. Determinar si debemos borrar
            # A veces cleaned_data falla si el campo no está en Meta, así que usamos doble chequeo
            eliminar_pdf_form = form.cleaned_data.get('eliminar_archivo_pdf')
            eliminar_pdf_manual = (checkbox_raw == 'on')
            
            # La condición final de borrado
            debe_borrar = (eliminar_pdf_form or eliminar_pdf_manual) and not archivo_raw
            
            print(f"[DEBUG] Decisión de borrado: {debe_borrar}", flush=True)

            # --- LÓGICA DE BORRADO ---
            if debe_borrar:
                print("[DEBUG] -> EJECUTANDO BORRADO...", flush=True)
                
                # Si hay un archivo actualmente en la base de datos
                if emo_programado.archivo_pdf:
                    nombre_archivo = emo_programado.archivo_pdf.name
                    print(f"[DEBUG] Archivo actual en DB: {nombre_archivo}", flush=True)
                    
                    try:
                        # 1. Borrar físico (si usas S3 esto es importante)
                        emo_programado.archivo_pdf.delete(save=False)
                        print(f"[DEBUG] delete(save=False) ejecutado.", flush=True)
                    except Exception as e:
                        print(f"[DEBUG] Error al borrar archivo físico (no crítico): {e}", flush=True)
                
                # 2. Forzar valor None explícitamente
                emo_realizado.archivo_pdf = None
                print("[DEBUG] Asignado None a emo_realizado.archivo_pdf", flush=True)
            
            # --- GUARDADO FINAL ---
            emo_realizado.estado = 'Realizado'
            emo_realizado.save()
            print("[DEBUG] emo_realizado.save() ejecutado.", flush=True)
            
            # Verificación inmediata post-guardado
            emo_chequeo = EMO.objects.get(id=emo_id)
            print(f"[DEBUG] Verificación en DB tras guardar: {emo_chequeo.archivo_pdf}", flush=True)
            print("x"*50 + "\n", flush=True)

            # --- Lógica de Controles ---
            if emo_realizado.tipo_emo != 'Retiro':
                emo_realizado.controles.filter(realizado=False).delete()
                cantidad = form.cleaned_data.get('cantidad_controles', 0)
                # Validación extra por si llega None
                if cantidad is None: cantidad = 0
                
                if cantidad > 0:
                    for i in range(cantidad):
                        fecha_str = request.POST.get(f'control_fecha_{i+1}')
                        descripcion = request.POST.get(f'control_descripcion_{i+1}')
                        if fecha_str and descripcion:
                            Control.objects.create(emo=emo_realizado, fecha_programada=fecha_str, descripcion=descripcion)

            # --- Actualización Automática ---
            trabajador = emo_realizado.trabajador
            trabajador.aptitud_actual = emo_realizado.aptitud
            trabajador.save(update_fields=['aptitud_actual'])
            
            messages.success(request, f"Resultados guardados correctamente.")
            return redirect('calidad:seleccionar_edicion_emo', emo_id=emo_realizado.id)
        else:
            print("[DEBUG] El formulario NO es válido", flush=True)
            print(f"[DEBUG] Errores: {form.errors}", flush=True)
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        form = RegistrarResultadoEmoForm(instance=emo_programado)
        controles_existentes = emo_programado.controles.all().order_by('id')
        controles_data = []
        if controles_existentes.exists():
            form.initial['cantidad_controles'] = controles_existentes.count()
            for control in controles_existentes:
                controles_data.append({
                    'fecha': control.fecha_programada.strftime('%Y-%m-%d'),
                    'descripcion': control.descripcion
                })
        
        controles_data_json = json.dumps(controles_data)
        is_edit_mode = True if controles_existentes.exists() else False

    context = {
        'form': form,
        'trabajador': emo_programado.trabajador,
        'emo': emo_programado,
        'current_view': 'gestion_emos',
        'form_title': 'Registrar/Editar Resultados de EMO',
        'form_subtitle': f'Actualizando el examen para {emo_programado.trabajador}',
        'controles_data_json': controles_data_json if 'controles_data_json' in locals() else '[]',
        'is_edit_mode': is_edit_mode if 'is_edit_mode' in locals() else False,
    }
    
    return render(request, 'calidad/formularios/crear_emo_multistep.html', context)

@login_required
@group_required('Calidad')
def crear_emo(request):
    if request.method == 'POST':
        form = ProgramarEmoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, f"EMO registrado con éxito para {form.cleaned_data['trabajador']}.")
            return redirect('lista_emos') # Necesitaremos esta URL
    else:
        form = ProgramarEmoForm()

    context = {
        'current_view': 'gestion_emos', # Para el menú
        'form': form,
    }
    return render(request, 'calidad/crear_emo_multistep.html', context)

def filtro_trabajador(request):
    return render(request, 'calidad/filtro_trabajador.html')


@login_required
@group_required('Calidad')
def programar_emo(request):
    
    # Variables de control
    trabajador_preseleccionado = None
    emo_instance = None
    modo_edicion = False
    titulo_vista = "Programar Nuevo EMO"
    empresas_trabajador = []

    # 1. CAPTURAR PARÁMETROS DE URL
    emo_id_editar = request.GET.get('editar') # ID para editar
    dni_trabajador = request.GET.get('dni')   # DNI para crear nuevo
    
    initial_data = {}

    # 2. DETERMINAR MODO (EDICIÓN O CREACIÓN)
    if emo_id_editar:
        # --- MODO EDICIÓN ---
        emo_instance = get_object_or_404(EMO, pk=emo_id_editar)
        trabajador_preseleccionado = emo_instance.trabajador
        modo_edicion = True
        titulo_vista = f"Editar Datos EMO: {trabajador_preseleccionado}"
        # Pre-cargar la empresa del EMO
        initial_data['empresa'] = emo_instance.empresa_id if emo_instance.empresa else None
        
    elif dni_trabajador:
        # --- MODO CREACIÓN ---
        try:
            trabajador_preseleccionado = Trabajador.objects.get(dni=dni_trabajador)
            initial_data['trabajador'] = trabajador_preseleccionado
            # Pre-cargar la empresa por defecto del trabajador
            if trabajador_preseleccionado.empresa:
                initial_data['empresa'] = trabajador_preseleccionado.empresa_id
        except Trabajador.DoesNotExist:
            messages.error(request, "El trabajador a programar no fue encontrado.")
            return redirect('calidad:reporte_maestro_emos')

    # Obtener las empresas del trabajador (si existe)
    if trabajador_preseleccionado:
        # Opción 1: Usar la empresa por defecto del trabajador
        if trabajador_preseleccionado.empresa:
            empresas_trabajador = [trabajador_preseleccionado.empresa]
        
        # Opción 2: (FUTURO) Si el trabajador tuviera múltiples empresas en una M2M,
        # podrías hacer: empresas_trabajador = trabajador_preseleccionado.empresas.all()

    # PREPARAR DATOS PARA EL CONTEXTO (necesarios antes del POST)
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent')
    
    for sub in all_subproyectos:
        pid = sub.parent_id
        if pid not in subproyectos_map:
            subproyectos_map[pid] = []
        
        subproyectos_map[pid].append({
            'id': sub.id, 
            'nombre': sub.nombre, 
            'codigo': sub.codigo if sub.codigo else ''
        })

    clinicas_fachadas_map = {}
    for c in Clinica.objects.all().only('id', 'imagen_fachada'):
        if c.imagen_fachada:
            clinicas_fachadas_map[str(c.id)] = c.imagen_fachada.url

    # Crea un formulario "default" para inicializar context
    if request.method == 'POST':
        form = ProgramarEmoForm(request.POST, instance=emo_instance)
    else:
        form = ProgramarEmoForm(instance=emo_instance, initial=initial_data)

    # Ocultar el campo trabajador si ya está definido (en ambos modos)
    if trabajador_preseleccionado:
        form.fields['trabajador'].widget = HiddenInput()

    # Construir contexto ANTES de procesar el POST (para poder usarlo en preview)
    context = {
        'form': form, 
        'trabajador_preseleccionado': trabajador_preseleccionado,
        'empresas_trabajador': empresas_trabajador,
        'subproyectos_map': subproyectos_map, 
        'clinicas_fachadas_map': clinicas_fachadas_map,
        'current_view': 'gestion_emos',
        'modo_edicion': modo_edicion,
        'titulo': titulo_vista,
        'emo': emo_instance
    }

    # 3. PROCESAMIENTO DEL FORMULARIO
    if request.method == 'POST':
        # Pasamos 'instance' para que Django sepa si actualizar o crear
        form = ProgramarEmoForm(request.POST, instance=emo_instance)
        context['form'] = form  # Actualizar el form en el contexto
        
        if form.is_valid():
            emo = form.save(commit=False)
            comentario_alerta = (form.cleaned_data.get('comentario_alerta') or '').strip()
            confirmar_envio = (request.POST.get('confirm_send') == '1')
            
            # Solo forzamos el estado inicial si es un registro NUEVO
            if not modo_edicion:
                emo.estado = 'Programado'
                emo.aptitud = 'Pendiente'

            # MODO EDICIÓN: guardar directo (sin enviar correo)
            if modo_edicion:
                # Guardar EMO sin enviar notificaciones
                emo.save()
                
                # Mensaje confirmando actualización sin envío de correo
                messages.success(request, "Datos del EMO actualizados correctamente. (No se ha notificado al trabajador)")
                
                return redirect('calidad:seleccionar_edicion_emo', emo_id=emo.id)

            # MODO CREACIÓN: preview -> confirmar_envio guarda + envía
            clinica_fachada_url = (
                request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                if (emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None))
                else ''
            )
            subject = f'EMO Programado - {emo.get_tipo_emo_display()}'
            from_email = settings.DEFAULT_FROM_EMAIL
            trabajador_notificable = _trabajador_es_notificable(emo.trabajador)
            to_email = [emo.trabajador.email] if trabajador_notificable and emo.trabajador.email else []

            # Renderizamos siempre el HTML del correo para la vista previa
            html_content = render_to_string('calidad/emails/notificacion_emo_programado.html', {
                'trabajador': emo.trabajador,
                'emo': emo,
                'comentario_alerta': comentario_alerta,
                'clinica_fachada_url': clinica_fachada_url,
            })

            if not confirmar_envio:
                # NO guardamos todavía: solo guardamos datos en sesión
                # Luego redirigimos a la vista de preview
                # El EMO se guardará solo cuando se confirme en preview_emo_email()
                
                comentario_alerta = (form.cleaned_data.get('comentario_alerta') or '').strip()
                clinica_fachada_url = (
                    request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                    if (emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None))
                    else ''
                )
                
                subject = f'EMO Programado - {emo.get_tipo_emo_display()}'
                trabajador_notificable = _trabajador_es_notificable(emo.trabajador)
                to_email = [emo.trabajador.email] if trabajador_notificable and emo.trabajador.email else []
                
                html_content = render_to_string('calidad/emails/notificacion_emo_programado.html', {
                    'trabajador': emo.trabajador,
                    'emo': emo,
                    'comentario_alerta': comentario_alerta,
                    'clinica_fachada_url': clinica_fachada_url,
                })
                
                # Guardar datos del EMO NO guardado en sesión
                fecha_str = None
                if emo.fecha_programada:
                    try:
                        if hasattr(emo.fecha_programada, 'strftime'):
                            fecha_str = emo.fecha_programada.strftime('%Y-%m-%d')
                        else:
                            fecha_str = str(emo.fecha_programada)
                    except Exception as e:
                        fecha_str = None
                
                hora_str = None
                if emo.hora_examen:
                    try:
                        if hasattr(emo.hora_examen, 'strftime'):
                            hora_str = emo.hora_examen.strftime('%H:%M')
                        else:
                            hora_str = str(emo.hora_examen)
                    except Exception as e:
                        hora_str = None
                
                request.session['emo_preview'] = {
                    'emo_data': {
                        'trabajador_id': emo.trabajador.id,
                        'tipo_emo': emo.tipo_emo,
                        'fecha_programada': fecha_str,
                        'hora_examen': hora_str,
                        'lugar_examen_id': emo.lugar_examen_id,
                        'empresa_id': emo.empresa_id,
                        'proyecto_id': emo.proyecto_id,
                        'subproyecto_id': emo.subproyecto_id,
                        'cargo_id': emo.cargo_id,
                        'comentario_alerta': comentario_alerta,
                    },
                    'tipo_emo': emo.get_tipo_emo_display(),
                    'email_subject': subject,
                    'email_to': to_email[0] if to_email else '',
                    'email_html': html_content,
                    'comentario_alerta': comentario_alerta,
                    'can_send': bool(to_email),
                }
                return redirect('calidad:preview_emo_email')
            else:
                # Guardar + enviar
                emo.save()

                if to_email:
                    try:
                        # CORREO DE COPIA (Igual que en edición)
                        correo_para_copia = 'habilitaciones@ceneris.com'
                        lista_cc = [correo_para_copia] if correo_para_copia else []
                        
                        # Preparar BCC (Copia Oculta) si tienes settings
                        bcc_list = []
                        if hasattr(settings, 'EMAIL_COPIA_EMOS') and settings.EMAIL_COPIA_EMOS:
                            bcc_list = [settings.EMAIL_COPIA_EMOS]

                        email = EmailMultiAlternatives(
                            subject=subject,
                            body=(
                                'Tienes un EMO programado. Por favor revisa los detalles en este correo.'
                                + (f"\n\nComentario: {comentario_alerta}" if comentario_alerta else "")
                            ),
                            from_email=from_email,
                            to=to_email,
                            cc=lista_cc,
                            bcc=bcc_list
                        )
                        email.attach_alternative(html_content, "text/html")
                        email.send()

                        messages.success(request, f"EMO programado y notificación enviada a {emo.trabajador.email}")
                    except Exception as e:
                        import traceback
                        traceback.print_exc()
                        messages.warning(request, f"EMO programado correctamente, pero no se pudo enviar el correo: {str(e)}")
                else:
                    messages.info(request, "EMO programado correctamente. El trabajador no tiene correo electrónico registrado.")

                messages.success(request, f"EMO ({emo.get_tipo_emo_display()}) programado con éxito para {emo.trabajador}.")
                return redirect('calidad:reporte_maestro_emos')
    else:
        # Carga inicial del formulario (con instancia si es edición)
        form = ProgramarEmoForm(instance=emo_instance, initial=initial_data)

    # Ocultar el campo trabajador si ya está definido (en ambos modos)
    if trabajador_preseleccionado:
        form.fields['trabajador'].widget = HiddenInput()

    context = {
        'form': form, 
        'trabajador_preseleccionado': trabajador_preseleccionado,
        'empresas_trabajador': empresas_trabajador,
        'subproyectos_map': subproyectos_map, 
        'clinicas_fachadas_map': clinicas_fachadas_map,
        'current_view': 'gestion_emos',
        'modo_edicion': modo_edicion,
        'titulo': titulo_vista,
        'emo': emo_instance
    }

    return render(request, 'calidad/emos/programar_emo.html', context)

@login_required
@group_required('Calidad')
def preview_emo_email(request):
    """
    Vista de previsualización estética del correo a enviar.
    Permite editar el comentario y confirmar el envío.
    """
    # Obtener datos de la sesión
    preview_data = request.session.get('emo_preview')
    if not preview_data:
        messages.error(request, "No hay datos para previsualizar.")
        return redirect('calidad:reporte_maestro_emos')

    if request.method == 'POST':
        accion = request.POST.get('accion')
        emo_id = preview_data.get('emo_id')
        emo_data = preview_data.get('emo_data')
        
        if accion == 'editar_datos':
            # Usuario está editando cualquier dato: actualizar EMO y re-renderizar HTML
            try:
                from datetime import datetime, time
                
                # Si es nuevo EMO (sin emo_id), crear temporal para validar cambios
                if not emo_id and emo_data:
                    trabajador_id = emo_data.get('trabajador_id')
                    trabajador = Trabajador.objects.get(id=trabajador_id)
                    
                    # Convertir strings de fecha y hora
                    fecha_programada = None
                    if emo_data.get('fecha_programada'):
                        try:
                            fecha_programada = datetime.strptime(emo_data.get('fecha_programada'), '%Y-%m-%d').date()
                        except:
                            pass
                    
                    hora_examen = None
                    if emo_data.get('hora_examen'):
                        try:
                            hora_examen = datetime.strptime(emo_data.get('hora_examen'), '%H:%M').time()
                        except:
                            pass
                    
                    emo = EMO(
                        trabajador=trabajador,
                        tipo_emo=emo_data.get('tipo_emo'),
                        fecha_programada=fecha_programada,
                        hora_examen=hora_examen,
                        lugar_examen_id=emo_data.get('lugar_examen_id'),
                        empresa_id=emo_data.get('empresa_id'),
                        proyecto_id=emo_data.get('proyecto_id'),
                        subproyecto_id=emo_data.get('subproyecto_id'),
                        cargo_id=emo_data.get('cargo_id'),
                        estado='Programado',
                        aptitud='Pendiente',
                    )
                    es_nuevo = True
                else:
                    emo = EMO.objects.get(id=emo_id)
                    es_nuevo = False
                
                # Actualizar campos editables
                if request.POST.get('tipo_emo'):
                    emo.tipo_emo = request.POST.get('tipo_emo')
                if request.POST.get('fecha_programada'):
                    emo.fecha_programada = request.POST.get('fecha_programada')
                if request.POST.get('hora_examen'):
                    emo.hora_examen = request.POST.get('hora_examen')
                if request.POST.get('lugar_examen_id'):
                    try:
                        emo.lugar_examen_id = request.POST.get('lugar_examen_id') or None
                    except:
                        pass
                if request.POST.get('empresa_id'):
                    try:
                        emo.empresa_id = request.POST.get('empresa_id') or None
                    except:
                        pass
                if request.POST.get('proyecto_id'):
                    try:
                        emo.proyecto_id = request.POST.get('proyecto_id') or None
                    except:
                        pass
                if request.POST.get('subproyecto_id'):
                    try:
                        emo.subproyecto_id = request.POST.get('subproyecto_id') or None
                    except:
                        pass
                if request.POST.get('cargo_id'):
                    try:
                        emo.cargo_id = request.POST.get('cargo_id') or None
                    except:
                        pass
                
                # Guardar cambios solo si es EMO existente (no nuevo)
                if not es_nuevo:
                    emo.save()
                
                # NO actualizar emo_data en sesión - mantener el estado original
                # Solo regenerar el HTML del correo con los cambios
                # El emo_data se guarda completo solo cuando se confirma
                
                # Actualizar datos en sesión
                nuevo_comentario = request.POST.get('comentario_alerta', '').strip()
                preview_data['comentario_alerta'] = nuevo_comentario
                preview_data['tipo_emo'] = emo.get_tipo_emo_display()
                
                # Actualizar emo_data en sesión con todos los datos actualizados
                from datetime import date, time as time_obj
                
                fecha_str = None
                if emo.fecha_programada:
                    try:
                        if hasattr(emo.fecha_programada, 'strftime'):
                            fecha_str = emo.fecha_programada.strftime('%Y-%m-%d')
                        else:
                            fecha_str = str(emo.fecha_programada)
                    except Exception as e:
                        fecha_str = None
                
                hora_str = None
                if emo.hora_examen:
                    try:
                        if hasattr(emo.hora_examen, 'strftime'):
                            hora_str = emo.hora_examen.strftime('%H:%M')
                        else:
                            hora_str = str(emo.hora_examen)
                    except Exception as e:
                        hora_str = None
                
                preview_data['emo_data'] = {
                    'trabajador_id': emo.trabajador.id,
                    'tipo_emo': emo.tipo_emo,
                    'fecha_programada': fecha_str,
                    'hora_examen': hora_str,
                    'lugar_examen_id': emo.lugar_examen_id,
                    'empresa_id': emo.empresa_id,
                    'proyecto_id': emo.proyecto_id,
                    'subproyecto_id': emo.subproyecto_id,
                    'cargo_id': emo.cargo_id,
                    'comentario_alerta': preview_data.get('comentario_alerta'),
                }
                
                # Re-renderizar el HTML del correo
                clinica_fachada_url = ''
                if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
                    clinica_fachada_url = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                
                html_content = render_to_string('calidad/emails/notificacion_emo_programado.html', {
                    'trabajador': emo.trabajador,
                    'emo': emo,
                    'comentario_alerta': nuevo_comentario,
                    'clinica_fachada_url': clinica_fachada_url,
                })
                
                preview_data['email_html'] = html_content
                request.session['emo_preview'] = preview_data
                messages.info(request, "Datos actualizados. Revisa el correo abajo para confirmar los cambios.")
                
            except EMO.DoesNotExist:
                messages.error(request, "El EMO no se encontró.")
                return redirect('calidad:reporte_maestro_emos')
            except Exception as e:
                messages.error(request, f"Error al actualizar: {str(e)}")
                return redirect('calidad:reporte_maestro_emos')
            
        elif accion == 'confirmar':
            # Usuario confirma: guardar EMO si es nuevo, luego enviar correo
            try:
                from datetime import datetime
                
                emo_id = preview_data.get('emo_id')
                emo_data = preview_data.get('emo_data')
                
                # Si es nuevo EMO (sin emo_id), guardarlo ahora
                if not emo_id and emo_data:
                    trabajador_id = emo_data.get('trabajador_id')
                    trabajador = Trabajador.objects.get(id=trabajador_id)
                    
                    # Convertir strings de fecha y hora
                    fecha_programada = None
                    if emo_data.get('fecha_programada'):
                        try:
                            fecha_programada = datetime.strptime(emo_data.get('fecha_programada'), '%Y-%m-%d').date()
                        except Exception as e:
                            pass
                    
                    hora_examen = None
                    if emo_data.get('hora_examen'):
                        try:
                            hora_examen = datetime.strptime(emo_data.get('hora_examen'), '%H:%M').time()
                        except:
                            pass
                    
                    emo = EMO(
                        trabajador=trabajador,
                        tipo_emo=emo_data.get('tipo_emo'),
                        fecha_programada=fecha_programada,
                        hora_examen=hora_examen,
                        lugar_examen_id=emo_data.get('lugar_examen_id'),
                        empresa_id=emo_data.get('empresa_id'),
                        proyecto_id=emo_data.get('proyecto_id'),
                        subproyecto_id=emo_data.get('subproyecto_id'),
                        cargo_id=emo_data.get('cargo_id'),
                        estado='Programado',
                        aptitud='Pendiente',
                    )
                    emo.save()
                    preview_data['emo_id'] = emo.id
                    request.session['emo_preview'] = preview_data

                    try:
                        from .services import notificar_doctores
                        notificar_doctores(nuevo_emo=emo)
                    except Exception:
                        logger.exception(f"Error al notificar a doctores sobre el nuevo EMO {emo.id}")
                else:
                    emo = EMO.objects.get(id=emo_id)
                
                comentario_alerta = preview_data.get('comentario_alerta', '')
                email_html = preview_data.get('email_html', '')
                email_subject = preview_data.get('email_subject', '')
                email_to = preview_data.get('email_to', '')
                
                # Verificar si se debe enviar el correo (checkbox)
                enviar_notification = request.POST.get('enviar_correo') == 'on'
                
                trabajador_notificable = _trabajador_es_notificable(emo.trabajador)

                if email_to and enviar_notification and trabajador_notificable:
                    try:
                        # CORREO DE COPIA (Igual que en edición)
                        correo_para_copia = 'habilitaciones@ceneris.com'
                        lista_cc = [correo_para_copia] if correo_para_copia else []

                        # Preparar BCC (copia oculta para supervisión)
                        bcc_list = []
                        if hasattr(settings, 'EMAIL_COPIA_EMOS') and settings.EMAIL_COPIA_EMOS:
                            bcc_list = [settings.EMAIL_COPIA_EMOS]
                        
                        email = EmailMultiAlternatives(
                            subject=email_subject,
                            body=(
                                'Tienes un EMO programado. Por favor revisa los detalles en este correo.'
                                + (f"\n\nComentario: {comentario_alerta}" if comentario_alerta else "")
                            ),
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            to=[email_to],
                            cc=lista_cc,
                            bcc=bcc_list
                        )
                        email.attach_alternative(email_html, "text/html")
                        
                        # Adjuntar PDFs de la clínica
                        if emo.lugar_examen:
                            archivos_clinica = ArchivoClinica.objects.filter(
                                clinica=emo.lugar_examen,
                                activo=True
                            )
                            for archivo in archivos_clinica:
                                try:
                                    with archivo.archivo_pdf.open('rb') as f:
                                        email.attach(
                                            f"{archivo.descripcion}.pdf",
                                            f.read(),
                                            'application/pdf'
                                        )
                                except Exception as e:
                                    pass
                        
                        email.send()
                        
                        messages.success(request, f"EMO programado y correo enviado a {email_to}.")
                    except Exception as e:
                        messages.warning(request, f"EMO programado, pero el correo no se pudo enviar: {str(e)}")
                else:
                    # Caso: No se envió correo (ya sea porque no se marcó checkbox o no tiene email)
                    if email_to and trabajador_notificable:
                        # Tiene email pero se desmarcó la casilla
                        messages.success(request, "EMO programado y guardado correctamente (Sin enviar correo).")
                    elif email_to and not trabajador_notificable:
                        messages.warning(request, "EMO programado, pero no se envió correo porque el trabajador está cesado/inactivo.")
                    else:
                        # No tiene email
                        messages.info(request, "EMO programado. El trabajador no tiene email registrado.")
                
                # Limpiar sesión y redirigir
                del request.session['emo_preview']
                return redirect('calidad:reporte_maestro_emos')
                
            except EMO.DoesNotExist:
                messages.error(request, "El EMO no se encontró.")
                return redirect('calidad:reporte_maestro_emos')
            except Exception as e:
                messages.error(request, f"Error al procesar el EMO: {str(e)}")
                return redirect('calidad:reporte_maestro_emos')
        
        elif accion == 'cancelar':
            # Usuario cancela: eliminar el EMO y limpiar sesión
            try:
                if emo_id:
                    emo = EMO.objects.get(id=emo_id)
                    emo.delete()
            except:
                pass
            
            if 'emo_preview' in request.session:
                del request.session['emo_preview']
            messages.info(request, "Se canceló la programación del EMO.")
            return redirect('calidad:reporte_maestro_emos')

    context = {
        'preview_data': preview_data,
        'current_view': 'gestion_emos',
        'tipo_emo_choices': EMO.TIPO_EMO_CHOICES,
        'clinicas': Clinica.objects.all().order_by('nombre'),
        'empresas': Empresa.objects.all().order_by('nombre'),
        'proyectos': Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
        'cargos': Cargo.objects.all().order_by('nombre'),
    }
    
    # Obtener el EMO para mostrar valores actuales en los campos
    emo_id = preview_data.get('emo_id')
    emo_data = preview_data.get('emo_data')
    
    if emo_id:
        # EMO ya existe en BD (edición)
        try:
            emo = EMO.objects.get(id=emo_id)
            context['emo'] = emo
            context['trabajador'] = emo.trabajador
            
            # Renderizar HTML del email si no existe en preview_data
            if not preview_data.get('email_html'):
                clinica_fachada_url = ''
                if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
                    try:
                        clinica_fachada_url = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                    except:
                        clinica_fachada_url = ''
                
                html_content = render_to_string('calidad/emails/notificacion_emo_programado.html', {
                    'trabajador': emo.trabajador,
                    'emo': emo,
                    'comentario_alerta': preview_data.get('comentario_alerta', ''),
                    'clinica_fachada_url': clinica_fachada_url,
                })
                
                preview_data['email_html'] = html_content
            
            # Pasar imagen de clínica a la sesión si existe
            if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
                preview_data['clinica_fachada_url'] = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                request.session['emo_preview'] = preview_data
            # Pasar subproyectos si el proyecto está seleccionado
            if emo.proyecto:
                context['subproyectos'] = emo.proyecto.subproyectos.filter(activo=True).order_by('nombre')
            else:
                context['subproyectos'] = []
        except EMO.DoesNotExist:
            context['subproyectos'] = []
    elif emo_data:
        # EMO es nuevo, aún no guardado - construir objeto temporal para renderizar
        try:
            from datetime import datetime
            
            trabajador_id = emo_data.get('trabajador_id')
            trabajador = Trabajador.objects.get(id=trabajador_id)
            
            # Convertir strings de fecha y hora
            fecha_programada = None
            if emo_data.get('fecha_programada'):
                try:
                    fecha_str = emo_data.get('fecha_programada')
                    fecha_programada = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except Exception as e:
                    pass
            
            hora_examen = None
            if emo_data.get('hora_examen'):
                try:
                    hora_str = emo_data.get('hora_examen')
                    hora_examen = datetime.strptime(hora_str, '%H:%M').time()
                except Exception as e:
                    pass
            
            # Crear EMO temporal (sin guardar en BD)
            emo = EMO(
                trabajador=trabajador,
                tipo_emo=emo_data.get('tipo_emo'),
                fecha_programada=fecha_programada,
                hora_examen=hora_examen,
                lugar_examen_id=emo_data.get('lugar_examen_id'),
                empresa_id=emo_data.get('empresa_id'),
                proyecto_id=emo_data.get('proyecto_id'),
                subproyecto_id=emo_data.get('subproyecto_id'),
                cargo_id=emo_data.get('cargo_id'),
                estado='Programado',
                aptitud='Pendiente',
            )
            context['emo'] = emo
            context['trabajador'] = emo.trabajador
            
            # Pasar imagen de clínica si existe
            if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
                preview_data['clinica_fachada_url'] = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                request.session['emo_preview'] = preview_data
            
            # **IMPORTANTE**: Renderizar el HTML del email para el preview
            clinica_fachada_url = ''
            if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
                try:
                    clinica_fachada_url = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                except:
                    clinica_fachada_url = ''
            
            html_content = render_to_string('calidad/emails/notificacion_emo_programado.html', {
                'trabajador': emo.trabajador,
                'emo': emo,
                'comentario_alerta': preview_data.get('comentario_alerta', ''),
                'clinica_fachada_url': clinica_fachada_url,
            })
            
            # Guardar HTML renderizado en preview_data
            preview_data['email_html'] = html_content
            context['preview_data'] = preview_data
            
            # Pasar subproyectos si el proyecto está seleccionado
            if emo.proyecto:
                context['subproyectos'] = emo.proyecto.subproyectos.filter(activo=True).order_by('nombre')
            else:
                context['subproyectos'] = []
        except Exception as e:
            context['subproyectos'] = []
    else:
        context['subproyectos'] = []
    
    return render(request, 'calidad/emos/preview_emo_email.html', context)


@login_required
@group_required('Calidad')
def preview_correo_edicion_emo(request, emo_id):
    """
    Vista para previsualizar el correo que se envía cuando se edita 
    un EMO en estado 'Realizado' o 'Programado'
    """
    try:
        emo = EMO.objects.get(id=emo_id)
    except EMO.DoesNotExist:
        messages.error(request, "El EMO no fue encontrado.")
        return redirect('calidad:reporte_maestro_emos')
    
    # Recuperar datos de sesión si existen (datos editados no guardados)
    datos_sesion = request.session.get(f'emo_{emo_id}_datos', None)
    
    # Si hay datos en sesión, crear un objeto EMO temporal con esos datos para la preview
    if datos_sesion:
        from datetime import datetime
        from copy import copy
        
        # Clonar el EMO y actualizar con datos de la sesión
        emo_preview = copy(emo)
        
        if datos_sesion.get('tipo_emo'):
            emo_preview.tipo_emo = datos_sesion['tipo_emo']
        
        if datos_sesion.get('fecha_programada'):
            try:
                emo_preview.fecha_programada = datetime.strptime(datos_sesion['fecha_programada'], '%Y-%m-%d').date()
            except:
                pass
        
        if datos_sesion.get('hora_examen'):
            try:
                emo_preview.hora_examen = datetime.strptime(datos_sesion['hora_examen'], '%H:%M').time()
            except:
                pass
        
        if datos_sesion.get('lugar_examen_id'):
            try:
                clinica = Clinica.objects.get(id=int(datos_sesion['lugar_examen_id']))
                emo_preview.lugar_examen = clinica
            except:
                pass
        
        if datos_sesion.get('empresa_id'):
            try:
                empresa = Empresa.objects.get(id=int(datos_sesion['empresa_id']))
                emo_preview.empresa = empresa
            except:
                pass
        
        if datos_sesion.get('proyecto_id'):
            try:
                proyecto = Proyecto.objects.get(id=int(datos_sesion['proyecto_id']))
                emo_preview.proyecto = proyecto
            except:
                pass
        
        if datos_sesion.get('subproyecto_id'):
            try:
                subproyecto = Proyecto.objects.get(id=int(datos_sesion['subproyecto_id']))
                emo_preview.subproyecto = subproyecto
            except:
                pass
        
        if datos_sesion.get('cargo_id'):
            try:
                cargo = Cargo.objects.get(id=int(datos_sesion['cargo_id']))
                emo_preview.cargo = cargo
            except:
                pass
        
        # Guardar comentario_alerta para uso posterior
        comentario_alerta = datos_sesion.get('comentario_alerta', '')
        
        emo = emo_preview
    else:
        comentario_alerta = ''
    
    # Renderizar HTML del correo
    try:
        clinica_fachada_url = ''
        if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
            clinica_fachada_url = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
        
        html_content = render_to_string('calidad/emails/notificacion_emo_programado.html', {
            'trabajador': emo.trabajador,
            'emo': emo,
            'comentario_alerta': comentario_alerta if comentario_alerta else 'Se han actualizado los datos de tu examen médico. Por favor, revisa la información.',
            'clinica_fachada_url': clinica_fachada_url,
        })
    except Exception as e:
        html_content = f"<p>Error al renderizar el correo: {str(e)}</p>"
    
    context = {
        'emo': emo,
        'trabajador': emo.trabajador,
        'email_html': html_content,
        'from_email': 'calidad@ceneris.com',
        'to_email': emo.trabajador.email,
        'comentario_alerta': comentario_alerta,
    }
    
    return render(request, 'calidad/emos/preview_correo_edicion_emo.html', context)

    
@login_required
@group_required('Calidad')
def buscar_trabajador_info(request):
    trabajador = None
    ultimo_emo_realizado = None
    historial_emos = None
    dni_buscado = request.GET.get('dni', '')
    
    if dni_buscado:
        try:
            trabajador = Trabajador.objects.select_related('empresa', 'area', 'sede').get(dni=dni_buscado)

            emos_base_qs = trabajador.historial_emo.select_related(
                'empresa',
                'proyecto',
                'subproyecto',
                'lugar_examen',
            )
            
            # --- ¡LÓGICA CORREGIDA! ---
            # Ordenamos primero por fecha, y luego por ID para desempatar.
            ultimo_emo_realizado = emos_base_qs.filter(
                estado='Realizado'
            ).order_by('-fecha_realizacion', '-pk').first() # <-- Se añade '-pk'
            
            # Historial completo ordenado por fecha de realización descendente
            historial_emos = emos_base_qs.all().order_by('-fecha_realizacion', '-pk')
            
        except Trabajador.DoesNotExist:
            messages.error(request, f"No se encontró ningún trabajador con el DNI {dni_buscado}.")
    
    context = {
        'trabajador': trabajador,
        'ultimo_emo_realizado': ultimo_emo_realizado,
        'historial_emos': historial_emos,
        'dni_buscado': dni_buscado,
        'current_view': 'info_trabajador'
    }
    return render(request, 'calidad/reportes/info_trabajador.html', context)

@login_required
@group_required('Calidad')
def eliminar_emo(request, emo_id):
    # Obtenemos el EMO o devolvemos error 404 si no existe
    emo = get_object_or_404(EMO, pk=emo_id)
    
    # Guardamos el DNI para redirigir al usuario a la misma pantalla
    dni_trabajador = emo.trabajador.dni
    
    if request.method == 'POST':
        try:
            # Al eliminar el EMO, Django eliminará automáticamente 
            # los 'Controles' asociados si usaste on_delete=models.CASCADE en tu modelo
            emo.delete()
            messages.success(request, "El EMO y sus controles asociados han sido eliminados.")
        except Exception as e:
            messages.error(request, f"Error al eliminar: {e}")
    
    # Redirigimos a la vista de búsqueda con el DNI pre-cargado
    url_destino = reverse('calidad:buscar_trabajador_info')
    return redirect(f"{url_destino}?dni={dni_trabajador}")

@login_required
@group_required('Calidad')
def lista_emos_pendientes(request):
    """
    Muestra una lista de todos los Exámenes Médicos Ocupacionales (EMOs)
    que tienen el estado 'Programado', con filtros y coloración por urgencia.
    """
    # Obtener parámetros de filtro
    filtro_buscador = request.GET.get('apellidos', '').strip()
    filtro_empresa = request.GET.get('empresa', '')
    filtro_tipo_emo = request.GET.get('tipo_emo', '')
    filtro_estado_urgencia = request.GET.get('estado_urgencia', '')
    filtro_lugar = request.GET.get('lugar', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    filtro_proyecto = request.GET.get('proyecto', '')
    filtro_subproyecto = request.GET.get('subproyecto', '')
    filtro_cargo = request.GET.get('cargo', '')
    
    # Queryset base: EMOs 'Programado' O 'Realizado' sin archivo_pdf (CAMO) cargado
    emos_queryset = EMO.objects.filter(
        Q(estado='Programado') | 
        (Q(estado='Realizado') & (Q(archivo_pdf__isnull=True) | Q(archivo_pdf='')))
    ).select_related(
        'trabajador', 
        'trabajador__empresa',
        'lugar_examen',
        'proyecto',
        'subproyecto',
        'cargo'
    )
    
    # Aplicar filtros
    if filtro_buscador:
        terminos = filtro_buscador.split()
        query_trabajador = Q()
        for termino in terminos:
            query_trabajador &= (
                Q(trabajador__dni__icontains=termino) |
                Q(trabajador__nombres__icontains=termino) |
                Q(trabajador__apellido_paterno__icontains=termino) |
                Q(trabajador__apellido_materno__icontains=termino)
            )
        emos_queryset = emos_queryset.filter(query_trabajador)
    
    if filtro_empresa:
        emos_queryset = emos_queryset.filter(empresa__pk=filtro_empresa)
    
    if filtro_tipo_emo:
        emos_queryset = emos_queryset.filter(tipo_emo=filtro_tipo_emo)
    
    if filtro_lugar:
        emos_queryset = emos_queryset.filter(lugar_examen__pk=filtro_lugar)
    
    if filtro_proyecto:
        emos_queryset = emos_queryset.filter(proyecto__pk=filtro_proyecto)
    
    if filtro_subproyecto:
        emos_queryset = emos_queryset.filter(subproyecto__pk=filtro_subproyecto)
    
    if filtro_cargo:
        emos_queryset = emos_queryset.filter(cargo__pk=filtro_cargo)
    
    if filtro_fecha_desde:
        from datetime import datetime
        try:
            fecha_desde_obj = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            emos_queryset = emos_queryset.filter(fecha_programada__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        from datetime import datetime
        try:
            fecha_hasta_obj = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            emos_queryset = emos_queryset.filter(fecha_programada__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # Calcular urgencia y convertir a lista
    from datetime import date, timedelta
    hoy = date.today()
    emos_list = list(emos_queryset.order_by('fecha_programada'))
    
    for emo in emos_list:
        if emo.fecha_programada:
            dias_hasta_vencimiento = (emo.fecha_programada - hoy).days
            emo.dias_hasta_vencimiento = dias_hasta_vencimiento

            # Asignar clase según urgencia:
            # Ya venció (fecha pasada): rojo
            # Cualquier otro caso (programado, sin importar cuántos días faltan): amarillo
            if dias_hasta_vencimiento < 0:
                emo.clase_urgencia = 'emo-vencido'
            else:
                emo.clase_urgencia = 'emo-por-vencer'
        else:
            emo.dias_hasta_vencimiento = None
            emo.clase_urgencia = 'emo-por-vencer'
    
    # Filtro por estado de urgencia (después de calcular clases)
    if filtro_estado_urgencia:
        emos_list = [emo for emo in emos_list if emo.clase_urgencia == filtro_estado_urgencia]
    
    # Obtener subproyectos si hay un proyecto seleccionado
    opciones_subproyectos = []
    if filtro_proyecto:
        opciones_subproyectos = Proyecto.objects.filter(
            parent_id=filtro_proyecto, activo=True
        ).order_by('nombre')
    
    context = {
        'emos': emos_list,
        'current_view': 'gestion_emos',
        'opciones_empresas': Empresa.objects.all().order_by('nombre'),
        'opciones_tipo_emo': EMO.TIPO_EMO_CHOICES,
        'opciones_lugares': Clinica.objects.all().order_by('nombre'),
        'opciones_proyectos': Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
        'opciones_subproyectos': opciones_subproyectos,
        'opciones_cargos': Cargo.objects.all().order_by('nombre'),
        'opciones_estado_urgencia': [
            ('', 'Todos'),
            ('emo-por-vencer', 'Programados'),
            ('emo-vencido', 'Vencidos'),
        ],
        'filtro_activo_apellidos': filtro_buscador,
        'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else None,
        'filtro_activo_tipo_emo': filtro_tipo_emo,
        'filtro_activo_estado_urgencia': filtro_estado_urgencia,
        'filtro_activo_lugar': int(filtro_lugar) if filtro_lugar else None,
        'filtro_activo_proyecto': int(filtro_proyecto) if filtro_proyecto else None,
        'filtro_activo_subproyecto': int(filtro_subproyecto) if filtro_subproyecto else None,
        'filtro_activo_cargo': int(filtro_cargo) if filtro_cargo else None,
        'filtro_activo_fecha_desde': filtro_fecha_desde,
        'filtro_activo_fecha_hasta': filtro_fecha_hasta,
    }
    
    # Renderiza la plantilla HTML con el contexto
    return render(request, 'calidad/reportes/lista_emos_pendientes.html', context)

@login_required
@group_required('Calidad')
def get_subproyectos_ajax(request):
    """Vista AJAX para obtener subproyectos de un proyecto específico"""
    proyecto_id = request.GET.get('proyecto_id')
    
    if proyecto_id:
        subproyectos = Proyecto.objects.filter(
            parent_id=proyecto_id, 
            activo=True
        ).values('id', 'nombre').order_by('nombre')
        
        return JsonResponse({
            'subproyectos': list(subproyectos)
        })
    
    return JsonResponse({'subproyectos': []})

@login_required
@group_required('Calidad')
def eliminar_emo_pendiente(request, emo_id):
    # Solo permite peticiones POST para mayor seguridad
    if request.method == 'POST':
        # Permitimos eliminar EMOs Programados O EMOs Realizados vacíos (sin PDF)
        # Esto coincide con lo que se muestra en 'lista_emos_pendientes'
        query_condition = Q(estado='Programado') | (Q(estado='Realizado') & (Q(archivo_pdf__isnull=True) | Q(archivo_pdf='')))
        
        emo_a_eliminar = get_object_or_404(EMO, query_condition, id=emo_id)
        
        trabajador_nombre = emo_a_eliminar.trabajador
        emo_a_eliminar.delete()
        messages.success(request, f"El EMO pendiente de {trabajador_nombre} ha sido eliminado.")
    else:
        # Si alguien intenta acceder por GET, simplemente redirige
        messages.warning(request, "Acción no permitida.")
        
    return redirect('calidad:lista_emos_pendientes')

@login_required
@group_required('Calidad')
def lista_emos_vigentes(request):
    hoy = timezone.now().date()
    emos = EMO.objects.filter(estado='Realizado', fecha_vencimiento__gte=hoy).select_related('trabajador')
    context = {'emos': emos, 'titulo': 'Trabajadores con EMO Vigente'}
    return render(request, 'calidad/lista_emos_generica.html', context)

@login_required
@group_required('Calidad')
def lista_aptos_con_restriccion(request):
    trabajadores = Trabajador.objects.filter(
        Q(historial_emo__aptitud='Apto') | Q(historial_emo__aptitud='Apto con Restricción')
    ).distinct()
    context = {'trabajadores': trabajadores, 'titulo': 'Trabajadores Aptos o con Restricción'}
    return render(request, 'calidad/lista_trabajadores_simple.html', context)


@login_required
@group_required('Calidad')
def lista_emos_por_vencer(request):
    hoy = timezone.now().date()
    limite_30_dias = hoy + timezone.timedelta(days=30)
    emos = EMO.objects.filter(estado='Realizado', fecha_vencimiento__gte=hoy, fecha_vencimiento__lte=limite_30_dias)
    return render(request, 'calidad/lista_emos_vencimiento.html', {'emos': emos, 'titulo': 'EMOs por Vencer (Próximos 30 días)'})

@login_required
@group_required('Calidad')
def lista_emos_vencidos(request):
    hoy = timezone.now().date()
    emos = EMO.objects.filter(estado='Realizado', fecha_vencimiento__lt=hoy)
    return render(request, 'calidad/lista_emos_vencimiento.html', {'emos': emos, 'titulo': 'EMOs Vencidos'})

@login_required
@group_required('Calidad')
def reporte_aptitud_emos(request):
    """
    Muestra una lista con el último EMO de cada trabajador, filtrada
    por el resultado de la aptitud.
    """
    filtro_aptitud = request.GET.get('filtro_aptitud', None)

    # Queryset base: Obtiene el último EMO de cada trabajador (incluyendo los 'Programado')
    ultimos_emos = EMO.objects.order_by(
        'trabajador__dni', 
        '-fecha_realizacion',
        '-fecha_programada' # Añadido para desempatar si hay varios programados
    ).distinct('trabajador__dni')

    if filtro_aptitud:
        ultimos_emos = ultimos_emos.filter(aptitud=filtro_aptitud)

    opciones_aptitud = [choice[0] for choice in EMO.APTITUD_CHOICES]

    context = {
        'current_view': 'filtro_emos',
        'titulo': 'Reporte de Aptitud por Trabajador',
        'emos': ultimos_emos,
        'opciones_aptitud': opciones_aptitud,
        'filtro_activo': filtro_aptitud,
    }
    return render(request, 'calidad/reporte_aptitud_emos.html', context)

@login_required
@group_required('Calidad')
def lista_empresas(request):
    empresas = Empresa.objects.all().order_by('nombre')
    context = {
        'empresas': empresas,
        'current_view': 'gestion_empresas',
    }
    return render(request, 'calidad/reportes/lista_empresas.html', context)

@login_required
@group_required('Calidad')
def crear_empresa(request):
    if request.method == 'POST':
        form = EmpresaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa creada con éxito.')
            return redirect('lista_empresas')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EmpresaForm()
    context = {
        'form': form,
        'form_title': 'Añadir Nueva Empresa',
        'current_view': 'gestion_empresas',
    }
    # Apuntamos a la nueva plantilla multi-step
    return render(request, 'calidad/formularios/empresa_form_multistep.html', context)

@login_required
@group_required('Calidad')
def editar_empresa(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa actualizada con éxito.')
            return redirect('lista_empresas')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EmpresaForm(instance=empresa)
    context = {
        'form': form,
        'form_title': f'Editar Empresa: {empresa.nombre}',
        'current_view': 'gestion_empresas',
    }
    # La vista de edición también reutiliza la plantilla multi-step
    return render(request, 'calidad/formularios/empresa_form_multistep.html', context)

@login_required
@group_required('Calidad')
def eliminar_empresa(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)
    if request.method == 'POST':
        # Verificamos si hay trabajadores asociados antes de borrar
        if empresa.trabajador_set.exists():
            messages.error(request, f'No se puede eliminar la empresa "{empresa.nombre}" porque tiene trabajadores asociados.')
        else:
            empresa.delete()
            messages.success(request, f'Empresa "{empresa.nombre}" eliminada con éxito.')
        return redirect('lista_empresas')
    # No se debería llegar aquí por GET, pero por seguridad redirigimos
    return redirect('lista_empresas')


@login_required
@group_required('Calidad')
def lista_programacion_emos(request):
    """
    Muestra una lista priorizada de trabajadores que necesitan un nuevo EMO.
    Incluye:
    1. Trabajadores sin ningún EMO.
    2. Trabajadores cuyo último EMO está Vencido o Por Vencer.
    """
    hoy = timezone.now().date()
    limite_30_dias = hoy + timezone.timedelta(days=30)
    
    # --- LÓGICA SIMPLIFICADA Y CORREGIDA ---

    # 1. Obtenemos todos los trabajadores para analizar
    todos_trabajadores = Trabajador.objects.all().prefetch_related('historial_emo')
    
    lista_final_emos = []

    # 2. Iteramos sobre cada trabajador
    for trabajador in todos_trabajadores:
        # Buscamos el último EMO 'Realizado' de este trabajador
        ultimo_emo_realizado = trabajador.historial_emo.filter(estado='Realizado').order_by('-fecha_realizacion', '-pk').first()

        # CASO 1: Trabajador nuevo (nunca ha tenido un EMO 'Realizado')
        if not ultimo_emo_realizado:
            # Comprobamos si ya tiene uno 'Programado' para no duplicar
            if not trabajador.historial_emo.filter(estado='Programado').exists():
                emo_fantasma = EMO(trabajador=trabajador, aptitud="Sin EMO")
                lista_final_emos.append(emo_fantasma)
            continue # Pasamos al siguiente trabajador

        # CASO 2: Trabajador con historial. Comprobamos si el último EMO está por renovar.
        necesita_renovacion = False
        if ultimo_emo_realizado.fecha_vencimiento and ultimo_emo_realizado.fecha_vencimiento <= limite_30_dias:
            necesita_renovacion = True

        # Si necesita renovación, comprobamos que no tenga ya uno nuevo programado
        if necesita_renovacion:
            if not trabajador.historial_emo.filter(estado='Programado').exists():
                lista_final_emos.append(ultimo_emo_realizado)

    # 3. Ordenamos la lista final para mostrar los más urgentes primero
    lista_final_emos_ordenada = sorted(
        lista_final_emos,
        key=lambda x: x.fecha_vencimiento if x.fecha_vencimiento else hoy - timezone.timedelta(days=999)
    )

    context = {
        'current_view': 'gestion_certificados',
        'emos': lista_final_emos_ordenada,
    }
    return render(request, 'calidad/reportes/lista_programacion_emos.html', context)

@login_required
@group_required('Calidad')  # <-- decorador para roles
@login_required
@group_required('Calidad')
def lista_emos_por_confirmar(request):
    """
    Muestra al Habilitador la lista de EMOs que ya han sido 'Realizados'
    por el doctor, pero que aún no tienen confirmación.
    Incluye filtros por: empresa, proyecto, aptitud, tipo de EMO y búsqueda por trabajador.
    """
    # 1. Obtener todos los parámetros de filtro del formulario
    filtro_buscador = request.GET.get('apellidos', '').strip()
    filtro_empresa = request.GET.get('empresa', '')
    filtro_aptitud = request.GET.get('aptitud', '')
    filtro_tipo_emo = request.GET.get('tipo_emo', '')
    filtro_estado_urgencia = request.GET.get('estado_urgencia', '')
    filtro_clinica = request.GET.get('clinica', '')
    filtro_cargo = request.GET.get('cargo', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Filtros Jerárquicos
    filtro_proyecto_padre_id = request.GET.get('proyecto_padre', '')
    filtro_subproyecto_id = request.GET.get('subproyecto', '')
    
    # --- LÓGICA PRINCIPAL DE CONSULTAS ---
    
    # Queryset base: EMOs pendientes de confirmación (sin archivo cargado)
    emos_queryset = EMO.objects.filter(
        estado='Realizado'
    ).filter(
        Q(archivo_confirmacion__isnull=True) | Q(archivo_confirmacion='')
    ).select_related(
        'trabajador',
        'trabajador__empresa',
        'proyecto',
        'proyecto__parent',
        'lugar_examen'
    ).prefetch_related(
        'trabajador__asignaciones',
        'trabajador__asignaciones__proyecto',
        'trabajador__asignaciones__cargo'
    ).distinct()

    # --- APLICACIÓN DE FILTROS ---

    # 1. Filtro de Búsqueda por Trabajador
    if filtro_buscador:
        terminos = filtro_buscador.split()
        query_trabajador = Q()
        for termino in terminos:
            query_trabajador &= (
                Q(trabajador__dni__icontains=termino) |
                Q(trabajador__nombres__icontains=termino) |
                Q(trabajador__apellido_paterno__icontains=termino) |
                Q(trabajador__apellido_materno__icontains=termino)
            )
        
        emos_queryset = emos_queryset.filter(query_trabajador)

    # 2. Filtro de Empresa
    if filtro_empresa:
        emos_queryset = emos_queryset.filter(empresa__pk=filtro_empresa)

    # 3. Filtro de Aptitud
    if filtro_aptitud:
        emos_queryset = emos_queryset.filter(aptitud=filtro_aptitud)
    
    # 4. Filtro de Tipo de EMO
    if filtro_tipo_emo:
        emos_queryset = emos_queryset.filter(tipo_emo=filtro_tipo_emo)
    
    # 5. Filtro por Clínica
    if filtro_clinica:
        emos_queryset = emos_queryset.filter(lugar_examen__pk=filtro_clinica)
    
    # 5.1 Filtro por Cargo
    if filtro_cargo:
        emos_queryset = emos_queryset.filter(cargo__pk=filtro_cargo)
    
    # 6. Filtro por Rango de Fechas de Realización
    if filtro_fecha_desde:
        from datetime import datetime
        try:
            fecha_desde_obj = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            emos_queryset = emos_queryset.filter(fecha_realizacion__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    if filtro_fecha_hasta:
        from datetime import datetime
        try:
            fecha_hasta_obj = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            emos_queryset = emos_queryset.filter(fecha_realizacion__lte=fecha_hasta_obj)
        except ValueError:
            pass
    
    # 7. Filtro Jerárquico de Proyecto (M2M)
    if filtro_subproyecto_id:
        # Filtramos EMOs ligados al proyecto O trabajadores asignados a él
        pid = filtro_subproyecto_id
        emos_queryset = emos_queryset.filter(
            Q(proyecto__pk=pid) | 
            Q(trabajador__asignaciones__proyecto__pk=pid)
        ).distinct()
        
    elif filtro_proyecto_padre_id:
        # Filtro por Padre (Padre o Hijos)
        pid = filtro_proyecto_padre_id
        emos_queryset = emos_queryset.filter(
            Q(proyecto__pk=pid) | Q(proyecto__parent__pk=pid) |
            Q(trabajador__asignaciones__proyecto__pk=pid) | 
            Q(trabajador__asignaciones__proyecto__parent__pk=pid)
        ).distinct()

    # --- PREPARAR DATOS PARA EL TEMPLATE ---
    proyectos_padre = Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre')
    
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent')
    for sub in all_subproyectos:
        if sub.parent_id not in subproyectos_map:
            subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({
            'id': sub.id, 
            'nombre': sub.nombre, 
            'codigo': sub.codigo or ''
        })
    
    # Calcular días desde realización y asignar clase de urgencia para cada EMO
    from datetime import date, timedelta
    hoy = date.today()
    
    # IMPORTANTE: Convertir a lista ANTES de agregar atributos dinámicos
    emos_list = list(emos_queryset.order_by('-fecha_realizacion'))
    
    for emo in emos_list:
        if emo.fecha_realizacion:
            dias_transcurridos = (hoy - emo.fecha_realizacion).days
            emo.dias_transcurridos = dias_transcurridos
            emo.fecha_limite_lectura = emo.fecha_realizacion + timedelta(days=15)
            
            # Asignar clase según urgencia:
            # < 7 días: sin clase (normal)
            # ≥ 7 días y < 15 días: amarillo (advertencia)
            # ≥ 15 días: rojo (urgente/vencido)
            if dias_transcurridos >= 15:
                emo.clase_urgencia = 'emo-urgente'
            elif dias_transcurridos >= 7:
                emo.clase_urgencia = 'emo-advertencia'
            else:
                emo.clase_urgencia = ''
        else:
            emo.dias_transcurridos = None
            emo.fecha_limite_lectura = None
            emo.clase_urgencia = 'emo-pendiente'
    
    # 8. Filtro por Estado de Urgencia (después de calcular clases)
    if filtro_estado_urgencia:
        emos_list = [emo for emo in emos_list if emo.clase_urgencia == filtro_estado_urgencia]

    context = {
        'current_view': 'confirmar_emos',
        'titulo': 'EMOs Pendientes de Confirmación de Lectura',
        'emos': emos_list,
        
        'opciones_aptitud': EMO.APTITUD_CHOICES,
        'opciones_tipo_emo': EMO.TIPO_EMO_CHOICES,
        'opciones_empresas': Empresa.objects.all().order_by('nombre'),
        'opciones_proyectos_padre': proyectos_padre,
        'opciones_clinicas': Clinica.objects.all().order_by('nombre'),
        'opciones_cargos': Cargo.objects.all().order_by('nombre'),
        'opciones_estado_urgencia': [
            ('', 'Todos'),
            ('emo-advertencia', 'Advertencia (Amarillo)'),
            ('emo-urgente', 'Urgente (Rojo)'),
        ],
        
        'subproyectos_map_json': json.dumps(subproyectos_map),
        
        # Mantener filtros seleccionados en el HTML
        'filtro_activo_apellidos': filtro_buscador,
        'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else None,
        'filtro_activo_aptitud': filtro_aptitud,
        'filtro_activo_tipo_emo': filtro_tipo_emo,
        'filtro_activo_estado_urgencia': filtro_estado_urgencia,
        'filtro_activo_clinica': int(filtro_clinica) if filtro_clinica else None,
        'filtro_activo_cargo': int(filtro_cargo) if filtro_cargo else None,
        'filtro_activo_fecha_desde': filtro_fecha_desde,
        'filtro_activo_fecha_hasta': filtro_fecha_hasta,
        'filtro_activo_proyecto_padre': int(filtro_proyecto_padre_id) if filtro_proyecto_padre_id else None,
        'filtro_activo_subproyecto': int(filtro_subproyecto_id) if filtro_subproyecto_id else None,
    }
    return render(request, 'calidad/reportes/lista_emos_por_confirmar.html', context)

@login_required
@group_required('Calidad')
def confirmar_lectura_emo(request, emo_id):
    emo_a_confirmar = get_object_or_404(EMO, id=emo_id, estado='Realizado')
    
    # --- LÓGICA DINÁMICA ---
    # Detectamos si es una edición (ya estaba confirmado) o un registro nuevo.
    es_edicion = emo_a_confirmar.confirmado_por_habilitador
    
    if request.method == 'POST':
        form = ConfirmacionLecturaForm(request.POST, request.FILES, instance=emo_a_confirmar)
        
        if form.is_valid():
            emo_confirmado = form.save(commit=False)
            
            # Lógica para eliminar el archivo si se marcó el checkbox
            if form.cleaned_data.get('eliminar_archivo'):
                if emo_confirmado.archivo_confirmacion:
                    emo_confirmado.archivo_confirmacion.delete(save=False)
                emo_confirmado.archivo_confirmacion = None

            # Aseguramos los datos de confirmación
            emo_confirmado.confirmado_por_habilitador = True
            emo_confirmado.confirmado_por = request.user
            emo_confirmado.save() 
            
            # --- REDIRECCIÓN INTELIGENTE ---
            if es_edicion:
                messages.success(request, f"La lectura del EMO se ha actualizado correctamente.")
                # Si estamos editando, volvemos al menú de tarjetas del EMO
                return redirect('calidad:seleccionar_edicion_emo', emo_id=emo_a_confirmar.id)
            else:
                messages.success(request, f"Lectura confirmada exitosamente.")
                # Si es nuevo, volvemos a la lista de pendientes
                return redirect('calidad:lista_emos_por_confirmar')
        else:
            messages.error(request, "Por favor, corrige los errores en el formulario.")
    else:
        form = ConfirmacionLecturaForm(instance=emo_a_confirmar)

    context = {
        'form': form,
        'emo': emo_a_confirmar,
        # Variables para el template
        'es_edicion': es_edicion,
        'form_title': 'Editar Lectura de EMO' if es_edicion else 'Confirmación de Lectura',
    }
    return render(request, 'calidad/formularios/confirmar_lectura_form.html', context)

@login_required
@group_required('Calidad')
def gestion_empresas(request):
    """
    Muestra el dashboard de tarjetas para la gestión de empresas.
    """
    context = {
        'current_view': 'gestion_empresas',
    }
    return render(request, 'calidad/dashboards/gestion_empresas.html', context)

from django.db.models import OuterRef, Subquery # Asegúrate de tener estas importaciones

@login_required
@group_required('Calidad')
def reporte_maestro_emos(request):
    """
    Vista del Reporte Maestro con filtros jerárquicos para proyectos y subproyectos.
    Adaptada para relación Muchos-a-Muchos (Asignaciones).
    """
    # 1. Obtener todos los parámetros de filtro del formulario
    filtro_estado = request.GET.get('estado', '')
    filtro_aptitud = request.GET.get('aptitud', '')
    filtro_empresa = request.GET.get('empresa', '')
    filtro_lectura = request.GET.get('lectura', '')
    filtro_hospital = request.GET.get('hospital', '')
    filtro_buscador = request.GET.get('apellidos', '').strip()
    filtro_estado_laboral = request.GET.get('estado_laboral', '')
    filtro_cargo = request.GET.get('cargo', '')
    
    # Filtros de fecha de realización
    filtro_fecha_realizacion_desde = request.GET.get('fecha_realizacion_desde', '')
    filtro_fecha_realizacion_hasta = request.GET.get('fecha_realizacion_hasta', '')
    
    # Filtros Jerárquicos
    filtro_proyecto_padre_id = request.GET.get('proyecto_padre', '')
    filtro_subproyecto_id = request.GET.get('subproyecto', '')
    
    # --- LÓGICA PRINCIPAL DE CONSULTAS ---
    
    # A. Queryset base para EMOs (sin caché para mostrar datos actualizados)
    emos_queryset = EMO.objects.select_related(
        'trabajador', 'empresa', 'proyecto', 'subproyecto', 'cargo', 'lugar_examen'
    ).all()

    # B. Queryset base para trabajadores sin EMO
    trabajadores_sin_emo_queryset = Trabajador.objects.filter(
        historial_emo__isnull=True, 
        activo=True
    ).select_related(
        'empresa', 
        'centro_costo'
    ).prefetch_related(
        'asignaciones',
        'asignaciones__proyecto',
        'asignaciones__cargo'
    ).distinct()

    # --- APLICACIÓN DE FILTROS COMUNES (Aplican a ambos grupos) ---

    # 1. Filtro de Búsqueda por Trabajador
    if filtro_buscador:
        terminos = filtro_buscador.split()
        query_trabajador = Q()
        for termino in terminos:
            query_trabajador &= (
                Q(dni__icontains=termino) |
                Q(nombres__icontains=termino) |
                Q(apellido_paterno__icontains=termino) |
                Q(apellido_materno__icontains=termino)
            )
        
        emos_queryset = emos_queryset.filter(trabajador__in=Trabajador.objects.filter(query_trabajador))
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(query_trabajador)

    # 2. Filtro de Empresa
    if filtro_empresa:
        emos_queryset = emos_queryset.filter(empresa__pk=filtro_empresa)
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(empresa__pk=filtro_empresa)

    # 2.1 Filtro de Estado Laboral
    if filtro_estado_laboral == 'activo':
        emos_queryset = emos_queryset.filter(trabajador__activo=True)
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(activo=True)
    elif filtro_estado_laboral in ('inactivo', 'cesado'):
        emos_queryset = emos_queryset.filter(trabajador__activo=False)
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(activo=False)

    # 2.2 Filtro de Cargo
    if filtro_cargo:
        emos_queryset = emos_queryset.filter(cargo__pk=filtro_cargo)
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(asignaciones__cargo__pk=filtro_cargo).distinct()

    # 3. Filtro Jerárquico de Proyecto (M2M)
    if filtro_subproyecto_id:
        # En reporte de EMOs: cada fila es un EMO, así que filtramos por el proyecto/subproyecto del EMO.
        # (Usar asignaciones aquí mete EMOs “de rebote” y luego quedan sin proyecto en la tabla.)
        pid = filtro_subproyecto_id
        emos_queryset = emos_queryset.filter(
            Q(subproyecto__pk=pid) |
            Q(proyecto__pk=pid)
        ).distinct()
        
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(
            asignaciones__proyecto__pk=pid
        ).distinct()
        
    elif filtro_proyecto_padre_id:
        # Filtro por Padre (Padre o Hijos)
        pid = filtro_proyecto_padre_id
        emos_queryset = emos_queryset.filter(
            Q(proyecto__pk=pid) |
            Q(proyecto__parent__pk=pid) |
            Q(subproyecto__parent__pk=pid)
        ).distinct()
        
        trabajadores_sin_emo_queryset = trabajadores_sin_emo_queryset.filter(
            Q(asignaciones__proyecto__pk=pid) | 
            Q(asignaciones__proyecto__parent__pk=pid)
        ).distinct()

    # --- FILTROS ESPECÍFICOS DE EMO ---
    # Variable bandera para saber si ocultamos a los "Sin EMO"
    filtro_emo_activo = False

    # 4. Estado de Vigencia
    if filtro_estado and filtro_estado != 'sin_emo':
        filtro_emo_activo = True # Activamos bandera
        hoy = timezone.now().date()
        limite_30_dias = hoy + timezone.timedelta(days=30)
        
        if filtro_estado == 'vigente': 
            emos_queryset = emos_queryset.filter(estado='Realizado', fecha_vencimiento__gte=hoy)
        elif filtro_estado == 'por_vencer': 
            emos_queryset = emos_queryset.filter(estado='Realizado', fecha_vencimiento__range=(hoy, limite_30_dias))
        elif filtro_estado == 'vencido': 
            emos_queryset = emos_queryset.filter(estado='Realizado', fecha_vencimiento__lt=hoy)
        elif filtro_estado == 'programado': 
            emos_queryset = emos_queryset.filter(estado='Programado')
        
    if filtro_aptitud: 
        filtro_emo_activo = True # Activamos bandera
        emos_queryset = emos_queryset.filter(aptitud=filtro_aptitud)
    
    if filtro_hospital: 
        filtro_emo_activo = True # Activamos bandera
        emos_queryset = emos_queryset.filter(lugar_examen__pk=filtro_hospital)
        
    if filtro_lectura: 
        filtro_emo_activo = True 
        
        if filtro_lectura == 'confirmado': 
            emos_queryset = emos_queryset.filter(confirmado_por_habilitador=True)
            
        elif filtro_lectura == 'pendiente': 
            # --- CORRECCIÓN ---
            # 1. Filtramos que sean False O que sean Nulos (None)
            # 2. Mantenemos estado='Realizado', pero asegúrate que tus EMOs tengan ese estado.
            emos_queryset = emos_queryset.filter(
                estado='Realizado'
            ).filter(
                Q(confirmado_por_habilitador=False) | Q(confirmado_por_habilitador__isnull=True)
            )
    
    # Filtro de fecha de realización
    if filtro_fecha_realizacion_desde:
        filtro_emo_activo = True
        emos_queryset = emos_queryset.filter(fecha_realizacion__gte=filtro_fecha_realizacion_desde)
    
    if filtro_fecha_realizacion_hasta:
        filtro_emo_activo = True
        emos_queryset = emos_queryset.filter(fecha_realizacion__lte=filtro_fecha_realizacion_hasta)
    
    # --- LIMPIEZA DE LISTAS FINAL Y AGRUPACIÓN POR TRABAJADOR-PROYECTO ---
    
    # Si se seleccionó explícitamente "Sin EMO", ocultamos los EMOs
    if filtro_estado == 'sin_emo': 
        emos_queryset = EMO.objects.none()
    
    # Si se usó CUALQUIER filtro de EMO (Clínica, Aptitud, Vigencia), ocultamos los "Sin EMO"
    if filtro_emo_activo: 
        trabajadores_sin_emo_queryset = Trabajador.objects.none()

    # --- MOSTRAR SOLO EL ÚLTIMO EMO POR COMBINACIÓN TRABAJADOR-PROYECTO-EMPRESA ---
    emos_unicos = []
    emos_vistos = {}  # Dict para rastrear: {(trabajador_id, proyecto_id, empresa_id): emo}
    
    # Ordenamos por trabajador e id descendente (para que los más recientes aparezcan primero)
    # Usamos -id en lugar de fecha_realizacion para que EMOs Programados (sin fecha_realizacion) también se ordenen correctamente
    for emo in emos_queryset.order_by('trabajador__id', '-id'):
        trabajador_id = emo.trabajador.id
        # Clave del proyecto real del EMO (prioriza subproyecto si existe)
        proyecto_id = emo.subproyecto_id or emo.proyecto_id
        # Usar empresa del EMO, si no existe usar la del trabajador (fallback)
        empresa_id = emo.empresa_id or emo.trabajador.empresa_id
        
        # Creamos una clave única para trabajador-proyecto-empresa
        clave = (trabajador_id, proyecto_id, empresa_id)
        
        # Solo agregamos si es el primer EMO que vemos para esta combinación
        if clave not in emos_vistos:
            emos_vistos[clave] = emo
            emos_unicos.append(emo)
    
    # Nota: no “inventamos” proyectos vía asignaciones cuando el EMO no tiene proyecto/subproyecto.
    # Esos registros deben mostrarse como sin proyecto, o corregirse en el dato.

    # --- PREPARAR DATOS PARA EL TEMPLATE ---
    proyectos_padre = Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre')
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent')
    for sub in all_subproyectos:
        if sub.parent_id not in subproyectos_map:
            subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({
            'id': sub.id, 
            'nombre': sub.nombre, 
            'codigo': sub.codigo or ''
        })

    context = {
        'current_view': 'filtro_emos',
        'titulo': 'Reporte de EMOs',
        'emos': emos_unicos,  # Usar la lista filtrada de EMOs únicos
        'trabajadores_sin_emo': trabajadores_sin_emo_queryset.order_by('apellido_paterno'),
        'opciones_aptitud': EMO.APTITUD_CHOICES,
        'opciones_empresas': Empresa.objects.all().order_by('nombre'),
        'opciones_clinicas': Clinica.objects.all().order_by('nombre'),
        'opciones_cargos': Cargo.objects.all().order_by('nombre'),
        'opciones_proyectos_padre': proyectos_padre,
        'subproyectos_map_json': json.dumps(subproyectos_map),
        
        # Mantener filtros seleccionados en el HTML
        'filtro_activo_apellidos': filtro_buscador,
        'filtro_activo_estado': filtro_estado,
        'filtro_activo_aptitud': filtro_aptitud,
        'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else None,
        'filtro_activo_lectura': filtro_lectura,
        'filtro_activo_hospital': int(filtro_hospital) if filtro_hospital else None,
        'filtro_activo_cargo': int(filtro_cargo) if filtro_cargo else None,
        'filtro_activo_proyecto_padre': int(filtro_proyecto_padre_id) if filtro_proyecto_padre_id else None,
        'filtro_activo_subproyecto': int(filtro_subproyecto_id) if filtro_subproyecto_id else None,
        'filtro_activo_fecha_realizacion_desde': filtro_fecha_realizacion_desde,
        'filtro_activo_fecha_realizacion_hasta': filtro_fecha_realizacion_hasta,
        'filtro_activo_estado_laboral': filtro_estado_laboral,
    }
    
    return render(request, 'calidad/emos/reporte_maestro_emos.html', context)

@login_required
@group_required('Calidad')
def lista_proyectos(request):
    proyectos = Proyecto.objects.all().order_by('-activo', 'nombre')
    context = {
        'proyectos': proyectos, 
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'calidad/reportes/lista_proyectos.html', context)

@login_required
@group_required('Recursos Humanos', 'Calidad')
def crear_proyecto(request):
    """
    Formulario para crear un nuevo Proyecto Principal o un Subproyecto.
    """
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proyecto creado con éxito.')
            return redirect('calidad:lista_proyectos') # Asumiendo que tienes esta URL
    else:
        form = ProyectoForm()

    context = {
        'form': form,
        'form_title': 'Crear Nuevo Proyecto',
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'calidad/proyectos/proyecto_form_tarjetas.html', context)

@login_required
@group_required('Calidad')
def editar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proyecto actualizado con éxito.')
            return redirect('lista_proyectos')
        else:
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = ProyectoForm(instance=proyecto)
    context = {
        'form': form,
        'form_title': f'Editar Proyecto: {proyecto.nombre}',
        'current_view': 'gestion_proyectos',
    }
    # También reutiliza la plantilla multi-step
    return render(request, 'calidad/formularios/proyecto_form_multistep.html', context)

@login_required
@group_required('Calidad')
def eliminar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    if request.method == 'POST':
        if proyecto.trabajadores.exists():
            messages.error(request, f'No se puede eliminar "{proyecto.nombre}" porque tiene trabajadores asignados.')
        else:
            proyecto.delete()
            messages.success(request, f'Proyecto "{proyecto.nombre}" eliminado con éxito.')
        return redirect('lista_proyectos')
    return redirect('lista_proyectos')

@login_required
@group_required('Calidad')
def gestion_proyectos(request):
    """
    Muestra el dashboard de tarjetas para la gestión de proyectos.
    """
    context = {
        'current_view': 'gestion_proyectos',
    }
    return render(request, 'calidad/dashboards/gestion_proyectos.html', context)

@login_required
@group_required('Calidad')
def subir_subsana_emo(request, emo_id):
    """
    Procesa la subida del archivo de subsanación y cambia el estado de aptitud
    inmediatamente según las reglas de negocio.
    """
    # Obtenemos el EMO que necesita la subsanación.
    emo = get_object_or_404(EMO, id=emo_id)
    
    if request.method == 'POST':
        form = SubsanacionEmoForm(request.POST, request.FILES, instance=emo)
        if form.is_valid():
            # Guardamos el formulario. Esto ya asocia el archivo al objeto 'emo'.
            subsanacion = form.save(commit=False)
            
            # --- ¡LA LÓGICA QUE PEDISTE ESTÁ AQUÍ! ---
            aptitud_original = subsanacion.aptitud
            mensaje_cambio = ""

            # Solo aplicamos la lógica si el estado original es 'Observado'.
            if aptitud_original == 'Observado':
                # Comprobamos si el campo 'restriccion' tiene contenido.
                # .strip() elimina espacios en blanco al inicio y final.
                if subsanacion.restriccion and subsanacion.restriccion.strip() != '':
                    subsanacion.aptitud = 'Apto con Restricción'
                    mensaje_cambio = "La aptitud del EMO ha sido actualizada a 'Apto con Restricción'."
                else:
                    subsanacion.aptitud = 'Apto'
                    mensaje_cambio = "La aptitud del EMO ha sido actualizada a 'Apto'."
            else:
                mensaje_cambio = "La aptitud original no era 'Observado', por lo que no se realizó ningún cambio de estado."

            # También marcamos los campos de validación para mantener un registro
            subsanacion.subsana_validado = True
            subsanacion.fecha_validacion_subsana = timezone.now()
            
            # Guardamos todos los cambios en la base de datos
            subsanacion.save()
            
            messages.success(request, f"Documento de subsanación para {emo.trabajador} subido con éxito. {mensaje_cambio}")
            
            # Redirigimos a la lista principal de observaciones, donde este EMO ya no aparecerá.
            return redirect('calidad:lista_observaciones')
    else:
        form = SubsanacionEmoForm(instance=emo)
        
    context = {
        'form': form, 
        'emo': emo,
        'current_view': 'observaciones'
    }
    return render(request, 'calidad/formularios/subir_subsana_emo.html', context)


@login_required
@group_required('Calidad')
def validar_subsana_emo(request, emo_id):
    """
    Acción para marcar la subsanación como validada y cambiar la aptitud.
    """
    if request.method == 'POST':
        emo = get_object_or_404(EMO, id=emo_id)
        
        # Cambiamos la aptitud del EMO a 'Apto'
        emo.aptitud = 'Apto'
        emo.subsana_validado = True
        emo.fecha_validacion_subsana = timezone.now()
        emo.save(update_fields=['aptitud', 'subsana_validado', 'fecha_validacion_subsana'])
        
        # ¡IMPORTANTE! Actualizamos la aptitud general del trabajador
        trabajador = emo.trabajador
        trabajador.aptitud_actual = 'Apto'
        trabajador.save(update_fields=['aptitud_actual'])
        
        messages.success(request, f"Subsanación para {trabajador} validada. Su aptitud ahora es 'Apto'.")
    
    return redirect('lista_emos_por_validar')


@login_required
@group_required('Calidad')
def lista_para_subsanar(request):
    """
    [MODO DEPURACIÓN] Muestra TODOS los EMOs 'Realizados' que tienen un resultado problemático.
    """
    print("\n--- INICIANDO VISTA 'lista_para_subsanar' (MODO DEPURACIÓN) ---")

    # --- CONSULTA BASE: SOLO POR APTITUD ---
    # Empezamos con el filtro más simple. Esto debería mostrar TODO, incluso los ya subsanados.
    emos_base = EMO.objects.filter(
        estado='Realizado',
        aptitud__in=['Apto con Restricción', 'Observado']
    ).select_related('trabajador', 'trabajador__empresa')

    print(f"1. EMOs encontrados SOLO por aptitud: {emos_base.count()}")
    for emo in emos_base:
        # Imprimimos el estado del campo 'archivo_subsana' para cada uno.
        # .name nos da el nombre del archivo, si existe.
        print(f"  - EMO ID {emo.id} para {emo.trabajador}: archivo_subsana='{emo.archivo_subsana.name if emo.archivo_subsana else 'None'}'")

    # --- CONSULTA CON FILTRO ISNULL ---
    # Ahora aplicamos el filtro que sospechamos que está fallando.
    emos_filtrados_isnull = emos_base.filter(
        archivo_subsana__isnull=True
    )
    print(f"2. EMOs restantes DESPUÉS de filtrar por archivo_subsana__isnull=True: {emos_filtrados_isnull.count()}")
    
    # --- CONSULTA CON FILTRO DE CADENA VACÍA (Alternativa) ---
    # Probamos una forma diferente de buscar campos de archivo vacíos.
    emos_filtrados_blank = emos_base.filter(
        archivo_subsana=''
    )
    print(f"3. EMOs restantes DESPUÉS de filtrar por archivo_subsana='': {emos_filtrados_blank.count()}")

    # Para la prueba, enviaremos la lista SIN el filtro de 'archivo_subsana'
    # para que puedas ver todos los registros problemáticos en la tabla.
    emos_para_subsanar = emos_base.order_by('trabajador__apellido', '-fecha_realizacion')
    print("--- FIN DEPURACIÓN. Enviando a plantilla. ---")

    context = {
        'current_view': 'gestion_certificados',
        'emos': emos_para_subsanar,
    }
    return render(request, 'calidad/reportes/lista_para_subsanar.html', context)


@login_required
@group_required('Calidad')
def gestion_lecturas(request):
    context = {'current_view': 'gestion_lecturas'}
    return render(request, 'calidad/dashboards/gestion_lecturas.html', context)

@login_required
@group_required('Calidad')
def rechazar_subsana_emo(request, emo_id):
    """
    Rechaza una subsanación, eliminando el archivo de evidencia
    para que pueda ser subido de nuevo.
    """
    if request.method == 'POST':
        emo = get_object_or_404(EMO, id=emo_id)
        
        # Opcional: Elimina el archivo físico del almacenamiento.
        if emo.archivo_subsana:
            emo.archivo_subsana.delete(save=False) # save=False para no guardar el objeto aún

        # Limpiamos el campo en la base de datos
        emo.archivo_subsana = None
        emo.save(update_fields=['archivo_subsana'])
        
        messages.warning(request, f"La evidencia de subsanación para {emo.trabajador} ha sido rechazada. El usuario puede volver a subirla.")
    
    return redirect('lista_emos_por_validar')

@login_required
@group_required('Calidad')
def lista_observaciones(request):
    """
    Dashboard de seguimiento para EMOs observados y controles pendientes,
    con filtros elaborados por empresa, proyecto, clínica, aptitud, etc.
    """
    # --- OBTENER PARÁMETROS DE FILTRO ---
    filtro_buscador = request.GET.get('apellidos', '').strip()
    filtro_empresa = request.GET.get('empresa', '')
    filtro_clinica = request.GET.get('clinica', '')
    filtro_aptitud = request.GET.get('aptitud', '')
    filtro_cargo = request.GET.get('cargo', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Filtros Jerárquicos
    filtro_proyecto_padre_id = request.GET.get('proyecto_padre', '')
    filtro_subproyecto_id = request.GET.get('subproyecto', '')

    # --- CONSULTA 1: EMOs 'Observados' que requieren subsanación ---
    emos_para_subsanar = EMO.objects.filter(
        Q(archivo_subsana__isnull=True) | Q(archivo_subsana=''),
        aptitud='Observado'
    ).select_related(
        'trabajador', 
        'trabajador__empresa',
        'empresa',
        'proyecto',
        'proyecto__parent',
        'subproyecto',
        'subproyecto__parent',
        'cargo',
        'lugar_examen'
    ).prefetch_related(
        'trabajador__asignaciones',
        'trabajador__asignaciones__proyecto',
        'trabajador__asignaciones__cargo'
    ).distinct()

    # --- CONSULTA 2: Controles pendientes ---
    controles_pendientes = Control.objects.filter(
        realizado=False
    ).select_related(
        'emo',
        'emo__trabajador',
        'emo__trabajador__empresa',
        'emo__empresa',
        'emo__proyecto',
        'emo__proyecto__parent',
        'emo__subproyecto',
        'emo__subproyecto__parent',
        'emo__cargo',
        'emo__lugar_examen'
    ).prefetch_related(
        'emo__trabajador__asignaciones',
        'emo__trabajador__asignaciones__proyecto',
        'emo__trabajador__asignaciones__cargo'
    ).distinct()

    # --- APLICAR FILTROS COMUNES A AMBAS LISTAS ---

    # 1. Filtro de Búsqueda por Trabajador
    if filtro_buscador:
        terminos = filtro_buscador.split()
        query_trabajador = Q()
        for termino in terminos:
            query_trabajador &= (
                Q(dni__icontains=termino) |
                Q(nombres__icontains=termino) |
                Q(apellido_paterno__icontains=termino) |
                Q(apellido_materno__icontains=termino)
            )
        
        emos_para_subsanar = emos_para_subsanar.filter(trabajador__in=Trabajador.objects.filter(query_trabajador))
        controles_pendientes = controles_pendientes.filter(emo__trabajador__in=Trabajador.objects.filter(query_trabajador))

    # 2. Filtro de Empresa
    if filtro_empresa:
        emos_para_subsanar = emos_para_subsanar.filter(empresa__pk=filtro_empresa)
        controles_pendientes = controles_pendientes.filter(emo__empresa__pk=filtro_empresa)

    # 3. Filtro de Clínica
    if filtro_clinica:
        emos_para_subsanar = emos_para_subsanar.filter(lugar_examen__pk=filtro_clinica)
        controles_pendientes = controles_pendientes.filter(emo__lugar_examen__pk=filtro_clinica)

    # 4. Filtro de Aptitud (solo para EMOs)
    if filtro_aptitud:
        emos_para_subsanar = emos_para_subsanar.filter(aptitud=filtro_aptitud)
    
    # 4.1 Filtro de Cargo
    if filtro_cargo:
        emos_para_subsanar = emos_para_subsanar.filter(cargo__pk=filtro_cargo)
        controles_pendientes = controles_pendientes.filter(emo__cargo__pk=filtro_cargo)
    
    # 4.2 Filtro de Fecha de Realización (Desde)
    if filtro_fecha_desde:
        from datetime import datetime
        try:
            fecha_desde_obj = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d').date()
            emos_para_subsanar = emos_para_subsanar.filter(fecha_realizacion__gte=fecha_desde_obj)
            controles_pendientes = controles_pendientes.filter(emo__fecha_realizacion__gte=fecha_desde_obj)
        except ValueError:
            pass
    
    # 4.3 Filtro de Fecha de Realización (Hasta)
    if filtro_fecha_hasta:
        from datetime import datetime
        try:
            fecha_hasta_obj = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d').date()
            emos_para_subsanar = emos_para_subsanar.filter(fecha_realizacion__lte=fecha_hasta_obj)
            controles_pendientes = controles_pendientes.filter(emo__fecha_realizacion__lte=fecha_hasta_obj)
        except ValueError:
            pass

    # 5. Filtro Jerárquico de Proyecto (M2M)
    if filtro_subproyecto_id:
        pid = filtro_subproyecto_id
        emos_para_subsanar = emos_para_subsanar.filter(
            Q(proyecto__pk=pid) | 
            Q(trabajador__asignaciones__proyecto__pk=pid)
        ).distinct()
        
        controles_pendientes = controles_pendientes.filter(
            Q(emo__proyecto__pk=pid) | 
            Q(emo__trabajador__asignaciones__proyecto__pk=pid)
        ).distinct()
        
    elif filtro_proyecto_padre_id:
        pid = filtro_proyecto_padre_id
        emos_para_subsanar = emos_para_subsanar.filter(
            Q(proyecto__pk=pid) | Q(proyecto__parent__pk=pid) |
            Q(trabajador__asignaciones__proyecto__pk=pid) | 
            Q(trabajador__asignaciones__proyecto__parent__pk=pid)
        ).distinct()
        
        controles_pendientes = controles_pendientes.filter(
            Q(emo__proyecto__pk=pid) | Q(emo__proyecto__parent__pk=pid) |
            Q(emo__trabajador__asignaciones__proyecto__pk=pid) | 
            Q(emo__trabajador__asignaciones__proyecto__parent__pk=pid)
        ).distinct()

    # --- PREPARAR DATOS PARA EL TEMPLATE ---
    proyectos_padre = Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre')
    
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent')
    for sub in all_subproyectos:
        if sub.parent_id not in subproyectos_map:
            subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({
            'id': sub.id, 
            'nombre': sub.nombre, 
            'codigo': sub.codigo or ''
        })

    context = {
        'current_view': 'observaciones',
        'titulo': 'Seguimiento de Observaciones y Controles',
        'emos_para_subsanar': emos_para_subsanar.order_by('-fecha_realizacion'),
        'controles_pendientes': controles_pendientes.order_by('fecha_programada'),
        
        'opciones_empresas': Empresa.objects.all().order_by('nombre'),
        'opciones_clinicas': Clinica.objects.all().order_by('nombre'),
        'opciones_cargos': Cargo.objects.all().order_by('nombre'),
        'opciones_aptitud': EMO.APTITUD_CHOICES,
        'opciones_proyectos_padre': proyectos_padre,
        
        'subproyectos_map_json': json.dumps(subproyectos_map),
        
        # Mantener filtros seleccionados en el HTML
        'filtro_activo_apellidos': filtro_buscador,
        'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else None,
        'filtro_activo_clinica': int(filtro_clinica) if filtro_clinica else None,
        'filtro_activo_aptitud': filtro_aptitud,
        'filtro_activo_cargo': int(filtro_cargo) if filtro_cargo else None,
        'filtro_activo_fecha_desde': filtro_fecha_desde,
        'filtro_activo_fecha_hasta': filtro_fecha_hasta,
        'filtro_activo_proyecto_padre': int(filtro_proyecto_padre_id) if filtro_proyecto_padre_id else None,
        'filtro_activo_subproyecto': int(filtro_subproyecto_id) if filtro_subproyecto_id else None,
    }
    return render(request, 'calidad/observaciones/lista_observaciones.html', context)

@login_required
@group_required('Calidad')
def lista_controles(request):
    """
    Vista para listar y gestionar los Controles Médicos pendientes.
    Filtra controles donde realizado=False.
    """
    # 1. Obtener parámetros de filtro del GET
    filtro_buscador = request.GET.get('apellidos', '').strip()
    filtro_empresa = request.GET.get('empresa', '')
    filtro_clinica = request.GET.get('clinica', '')
    filtro_proyecto_padre = request.GET.get('proyecto_padre', '')
    filtro_subproyecto = request.GET.get('subproyecto', '')
    
    # 2. QuerySet Base: Solo controles NO realizados (pendientes)
    controles_queryset = Control.objects.filter(realizado=False).select_related(
        'emo', 
        'emo__trabajador', 
        'emo__empresa', 
        'emo__lugar_examen',
        'emo__proyecto',
        'emo__trabajador__empresa'
    ).order_by('fecha_programada')

    # 3. Aplicación de Filtros
    
    # A. Buscador
    if filtro_buscador:
        terminos = filtro_buscador.split()
        query_trabajador = Q()
        for termino in terminos:
            query_trabajador &= (
                Q(emo__trabajador__dni__icontains=termino) |
                Q(emo__trabajador__nombres__icontains=termino) |
                Q(emo__trabajador__apellido_paterno__icontains=termino) |
                Q(emo__trabajador__apellido_materno__icontains=termino)
            )
        controles_queryset = controles_queryset.filter(query_trabajador)

    # B. Filtro Empresa
    if filtro_empresa:
        controles_queryset = controles_queryset.filter(
            Q(emo__empresa__pk=filtro_empresa) | 
            Q(emo__trabajador__empresa__pk=filtro_empresa)
        )

    # C. Filtro Clínica
    if filtro_clinica:
        controles_queryset = controles_queryset.filter(emo__lugar_examen__pk=filtro_clinica)

    # D. Filtros de Proyecto
    if filtro_subproyecto:
        controles_queryset = controles_queryset.filter(
            Q(emo__proyecto__pk=filtro_subproyecto) |
            Q(emo__subproyecto__pk=filtro_subproyecto) |
            Q(emo__trabajador__asignaciones__proyecto__pk=filtro_subproyecto)
        ).distinct()
        
    elif filtro_proyecto_padre:
        controles_queryset = controles_queryset.filter(
            Q(emo__proyecto__pk=filtro_proyecto_padre) |
            Q(emo__proyecto__parent__pk=filtro_proyecto_padre) |
            Q(emo__trabajador__asignaciones__proyecto__pk=filtro_proyecto_padre) |
            Q(emo__trabajador__asignaciones__proyecto__parent__pk=filtro_proyecto_padre)
        ).distinct()
    
    # E. Filtro por Rango de Fechas
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    if fecha_desde:
        controles_queryset = controles_queryset.filter(fecha_programada__gte=fecha_desde)
    if fecha_hasta:
        controles_queryset = controles_queryset.filter(fecha_programada__lte=fecha_hasta)

    # F. Filtro por Estado
    filtro_estado = request.GET.get('estado', '').strip()
    hoy = timezone.now().date()
    
    if filtro_estado == 'vencido':
        controles_queryset = controles_queryset.filter(fecha_programada__lt=hoy)
    elif filtro_estado == 'hoy':
        controles_queryset = controles_queryset.filter(fecha_programada=hoy)
    elif filtro_estado == 'pendiente':
        controles_queryset = controles_queryset.filter(fecha_programada__gt=hoy)

    # 4. Datos para selectores
    proyectos_padre = Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre')
    
    import json
    subproyectos_map = {}
    for sub in Proyecto.objects.filter(parent__isnull=False, activo=True).select_related('parent'):
        if sub.parent_id not in subproyectos_map: subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({'id': sub.id, 'nombre': sub.nombre, 'codigo': sub.codigo or ''})

    # --- NUEVA LÓGICA: Verificar Permisos de Doctor ---
    es_doctor = request.user.groups.filter(name='Doctor').exists() or request.user.is_superuser

    # 5. Contexto final
    context = {
        'current_view': 'lista_controles',
        'titulo': 'Controles Médicos Pendientes',
        'controles_pendientes': controles_queryset,
        'today': timezone.now().date(),
        'es_doctor': es_doctor, # <--- Enviamos la variable al template
        
        'opciones_empresas': Empresa.objects.all().order_by('nombre'),
        'opciones_clinicas': Clinica.objects.all().order_by('nombre'),
        'opciones_proyectos_padre': proyectos_padre,
        'subproyectos_map_json': json.dumps(subproyectos_map),

        'filtro_activo_apellidos': filtro_buscador,
        'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else None,
        'filtro_activo_clinica': int(filtro_clinica) if filtro_clinica else None,
        'filtro_activo_proyecto_padre': int(filtro_proyecto_padre) if filtro_proyecto_padre else None,
        'filtro_activo_subproyecto': int(filtro_subproyecto) if filtro_subproyecto else None,
    }

    return render(request, 'calidad/observaciones/lista_controles.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def hub_observaciones(request):
    conteo_observaciones = EMO.objects.filter(
        estado='Realizado',
        aptitud__in=['Observado', 'Apto con Restricción', 'No Apto']
    ).count()

    try:
        conteo_controles = Control.objects.filter(realizado=False).count()
    except:
        conteo_controles = 0

    context = {
        'current_view': 'observaciones',
        'conteo_observaciones': conteo_observaciones,
        'conteo_controles': conteo_controles,
    }
    return render(request, 'calidad/observaciones/hub_observaciones.html', context)
    
@login_required
@group_required('Calidad')
def lista_clinicas(request):
    """Muestra una lista de todas las clínicas registradas."""
    clinicas = Clinica.objects.all().order_by('nombre')
    context = {
        'clinicas': clinicas,
        'current_view': 'gestion_clinicas', # Para el menú del sidebar
    }
    return render(request, 'calidad/clinicas/lista_clinicas.html', context)

@login_required
@group_required('Calidad')
def crear_clinica(request):
    """Muestra y procesa el formulario para crear una nueva clínica."""
    if request.method == 'POST':
        form = ClinicaForm(request.POST, request.FILES)
        archivo_form = ArchivoClinicaForm(request.POST, request.FILES) if request.POST.get('agregar_archivo') else None
        
        if form.is_valid():
            clinica = form.save()
            
            # Procesar archivos de la clínica si se cargaron
            if archivo_form and archivo_form.is_valid():
                archivo = archivo_form.save(commit=False)
                archivo.clinica = clinica
                archivo.save()
                messages.success(request, 'Clínica y archivo registrado con éxito.')
            else:
                messages.success(request, 'Clínica registrada con éxito.')
            
            return redirect('calidad:lista_clinicas')
    else:
        form = ClinicaForm()
        archivo_form = ArchivoClinicaForm()
    
    # Obtener archivos existentes si es edición
    archivos = []
    
    context = {
        'form': form,
        'archivo_form': archivo_form,
        'archivos': archivos,
        'form_title': 'Añadir Nueva Clínica',
        'current_view': 'gestion_clinicas',
    }
    return render(request, 'calidad/clinicas/clinica_form_multistep.html', context)

@login_required
@group_required('Calidad')
def editar_clinica(request, pk):
    """Muestra y procesa el formulario para editar una clínica existente."""
    clinica = get_object_or_404(Clinica, pk=pk)
    
    if request.method == 'POST':
        form = ClinicaForm(request.POST, request.FILES, instance=clinica)
        archivo_form = ArchivoClinicaForm(request.POST, request.FILES) if request.POST.get('agregar_archivo') else None
        
        if form.is_valid():
            form.save()
            
            # Procesar archivos de la clínica si se cargaron
            if archivo_form and archivo_form.is_valid():
                archivo = archivo_form.save(commit=False)
                archivo.clinica = clinica
                archivo.save()
                messages.success(request, 'Clínica y archivo actualizado con éxito.')
            else:
                messages.success(request, 'Clínica actualizada con éxito.')
            
            return redirect('calidad:lista_clinicas')
    else:
        form = ClinicaForm(instance=clinica)
        archivo_form = ArchivoClinicaForm()
    
    # Obtener archivos existentes
    archivos = ArchivoClinica.objects.filter(clinica=clinica).order_by('-fecha_carga')
    
    context = {
        'form': form,
        'archivo_form': archivo_form,
        'archivos': archivos,
        'form_title': 'Editar Clínica',
        'current_view': 'gestion_clinicas',
        'clinica': clinica,
    }
    return render(request, 'calidad/clinicas/clinica_form_multistep.html', context)

@login_required
@group_required('Calidad')
def eliminar_clinica(request, pk):
    """Elimina una clínica."""
    clinica = get_object_or_404(Clinica, pk=pk)
    if request.method == 'POST':
        # Comprobamos si la clínica está siendo usada en algún EMO
        if clinica.emo_set.exists():
            messages.error(request, f'No se puede eliminar "{clinica.nombre}" porque está asociada a uno o más EMOs.')
        else:
            clinica.delete()
            messages.success(request, f'Clínica "{clinica.nombre}" eliminada con éxito.')
        return redirect('calidad:lista_clinicas')
    return redirect('calidad:lista_clinicas')

@login_required
@group_required('Calidad')
def gestion_clinicas(request):
    """
    Muestra el dashboard de tarjetas para la gestión de clínicas.
    """
    context = {
        'current_view': 'gestion_clinicas',
    }
    return render(request, 'calidad/clinicas/gestion_clinicas.html', context)

@login_required
@group_required('Calidad')
def editar_emo_programado(request, emo_id):
    """
    Permite REPROGRAMAR un EMO en estado 'Programado'.
    Detecta cambios y envía correo de reprogramación al trabajador.
    (Este es el lápiz verde - para reprogramaciones oficiales)
    """
    from datetime import datetime
    
    emo = get_object_or_404(EMO, id=emo_id, estado='Programado')
    trabajador = emo.trabajador
    
    if request.method == 'POST':
        # Verificar si es solicitud de preview
        ir_a_preview = request.POST.get('ir_a_preview') == 'true'

        enviar_correo = request.POST.get('enviar_correo') == 'on'
        
        if ir_a_preview:
            # Guardar datos editados en sesión para mostrar en preview
            request.session[f'emo_{emo_id}_datos'] = {
                'tipo_emo': request.POST.get('tipo_emo', emo.tipo_emo),
                'fecha_programada': request.POST.get('fecha_programada', emo.fecha_programada.strftime('%Y-%m-%d') if emo.fecha_programada else ''),
                'hora_examen': request.POST.get('hora_examen', emo.hora_examen.strftime('%H:%M') if emo.hora_examen else ''),
                'lugar_examen_id': request.POST.get('lugar_examen_id', emo.lugar_examen_id or ''),
                'empresa_id': request.POST.get('empresa_id', emo.empresa_id or ''),
                'proyecto_id': request.POST.get('proyecto_id', emo.proyecto_id or ''),
                'subproyecto_id': request.POST.get('subproyecto_id', emo.subproyecto_id or ''),
                'cargo_id': request.POST.get('cargo_id', emo.cargo_id or ''),
                'comentario_alerta': request.POST.get('comentario_alerta', ''),
            }
            request.session.modified = True
            return redirect('calidad:preview_correo_edicion_emo', emo_id=emo_id)
        
        # Detectar cambios ANTES de actualizar
        cambios = {}
        emo_original = EMO.objects.get(pk=emo.id)
        
        # Actualizar datos del EMO
        nuevo_tipo_emo = request.POST.get('tipo_emo')
        if nuevo_tipo_emo:
            emo.tipo_emo = nuevo_tipo_emo
        
        nueva_fecha = request.POST.get('fecha_programada')
        if nueva_fecha:
            try:
                emo.fecha_programada = datetime.strptime(nueva_fecha, '%Y-%m-%d').date()
            except:
                pass
        
        nueva_hora = request.POST.get('hora_examen')
        if nueva_hora:
            try:
                emo.hora_examen = datetime.strptime(nueva_hora, '%H:%M').time()
            except:
                pass
        
        nuevo_lugar_id = request.POST.get('lugar_examen_id')
        if nuevo_lugar_id:
            try:
                emo.lugar_examen_id = int(nuevo_lugar_id) if nuevo_lugar_id else None
            except:
                pass
        
        nueva_empresa_id = request.POST.get('empresa_id')
        if nueva_empresa_id:
            try:
                emo.empresa_id = int(nueva_empresa_id) if nueva_empresa_id else None
            except:
                pass
        
        nuevo_proyecto_id = request.POST.get('proyecto_id')
        if nuevo_proyecto_id:
            try:
                emo.proyecto_id = int(nuevo_proyecto_id) if nuevo_proyecto_id else None
            except:
                pass
        
        nuevo_subproyecto_id = request.POST.get('subproyecto_id')
        if nuevo_subproyecto_id:
            try:
                emo.subproyecto_id = int(nuevo_subproyecto_id) if nuevo_subproyecto_id else None
            except:
                pass
        
        nuevo_cargo_id = request.POST.get('cargo_id')
        if nuevo_cargo_id:
            try:
                emo.cargo_id = int(nuevo_cargo_id) if nuevo_cargo_id else None
            except:
                pass
        
        # Detectar qué cambió
        campos_monitoreados = [
            ('fecha_programada', 'Fecha Programada'),
            ('hora_examen', 'Hora del Examen'),
            ('lugar_examen', 'Clínica'),
            ('empresa', 'Empresa'),
            ('proyecto', 'Proyecto'),
            ('subproyecto', 'Subproyecto'),
            ('tipo_emo', 'Tipo de EMO'),
            ('cargo', 'Cargo'),
        ]
        
        for campo, label in campos_monitoreados:
            valor_original = getattr(emo_original, campo)
            valor_nuevo = getattr(emo, campo)
            
            if valor_original != valor_nuevo:
                if campo in ['lugar_examen', 'empresa', 'proyecto', 'subproyecto', 'cargo']:
                    valor_original_str = str(valor_original) if valor_original else "Sin asignar"
                    valor_nuevo_str = str(valor_nuevo) if valor_nuevo else "Sin asignar"
                elif campo == 'tipo_emo':
                    valor_original_str = dict(EMO.TIPO_EMO_CHOICES).get(valor_original, valor_original)
                    valor_nuevo_str = dict(EMO.TIPO_EMO_CHOICES).get(valor_nuevo, valor_nuevo)
                elif campo == 'hora_examen':
                    valor_original_str = str(valor_original) if valor_original else "No asignada"
                    valor_nuevo_str = str(valor_nuevo) if valor_nuevo else "No asignada"
                else:
                    valor_original_str = str(valor_original) if valor_original else "Sin valor"
                    valor_nuevo_str = str(valor_nuevo) if valor_nuevo else "Sin valor"
                
                cambios[campo] = {
                    'label': label,
                    'antes': valor_original_str,
                    'ahora': valor_nuevo_str
                }
        
        # Guardar EMO actualizado
        emo.save()
        
        # Si hay cambios y tiene email, enviar correo de reprogramación
        trabajador_notificable = _trabajador_es_notificable(emo.trabajador)

        if cambios and emo.trabajador and emo.trabajador.email and enviar_correo and trabajador_notificable:
            try:
                # Obtener URL de la imagen de la clínica
                clinica_fachada_url = ''
                if emo.lugar_examen and getattr(emo.lugar_examen, 'imagen_fachada', None):
                    clinica_fachada_url = request.build_absolute_uri(emo.lugar_examen.imagen_fachada.url)
                
                html_reprogramacion = render_to_string('calidad/emails/notificacion_reprogramacion_emo.html', {
                    'trabajador': emo.trabajador,
                    'emo': emo,
                    'cambios': cambios,
                    'es_reprogramacion': True,
                    'clinica_fachada_url': clinica_fachada_url,
                })
                
                # 2. DEFINIR EL CORREO DE COPIA AQUÍ
                # Puedes ponerlo fijo o traerlo de settings
                correo_para_copia = 'habilitaciones@ceneris.com'  # <--- PON AQUÍ EL CORREO
                
                # Preparar lista de CC (Copia Visible)
                lista_cc = [correo_para_copia] if correo_para_copia else []

                # Preparar BCC (Copia Oculta - Opcional, si quieres mantener lo que tenías)
                bcc_list = []
                if hasattr(settings, 'EMAIL_COPIA_EMOS') and settings.EMAIL_COPIA_EMOS:
                    bcc_list = [settings.EMAIL_COPIA_EMOS]
                
                email_reprog = EmailMultiAlternatives(
                    subject=f'🔄 REPROGRAMACIÓN - EMO {emo.get_tipo_emo_display()}',
                    body='Tu examen médico ha sido reprogramado. Revisa los detalles en este correo.',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[emo.trabajador.email],
                    cc=lista_cc,   # <--- AGREGADO: Envía copia a este correo
                    bcc=bcc_list   # (Opcional: mantiene la copia oculta si la usas)
                )
                email_reprog.attach_alternative(html_reprogramacion, "text/html")
                
                # Adjuntar PDFs de recomendaciones de la clínica
                if emo.lugar_examen:
                    archivos_clinica = ArchivoClinica.objects.filter(
                        clinica=emo.lugar_examen,
                        activo=True
                    )
                    for archivo in archivos_clinica:
                        try:
                            with archivo.archivo_pdf.open('rb') as f:
                                email_reprog.attach(
                                    f"{archivo.descripcion}.pdf",
                                    f.read(),
                                    'application/pdf'
                                )
                        except Exception as e:
                            pass
                
                email_reprog.send()
                
                messages.success(request, f"EMO reprogramado y correo enviado a {emo.trabajador.email} (Copia a: {correo_para_copia})")
            except Exception as e:
                messages.warning(request, f"EMO actualizado, pero el correo no se pudo enviar: {str(e)}")
        else:
            if cambios:
                if not enviar_correo:
                    messages.success(request, "EMO actualizado correctamente. (No se envió correo por elección del usuario)")
                elif emo.trabajador and emo.trabajador.email and not trabajador_notificable:
                    messages.warning(request, "EMO actualizado correctamente. No se envió correo porque el trabajador está cesado/inactivo.")
                else:
                    messages.success(request, "EMO actualizado. (El trabajador no tiene email registrado)")
            else:
                messages.info(request, "No se realizaron cambios.")
        
        return redirect('calidad:reporte_maestro_emos')
    
    # GET: Mostrar formulario con datos actuales
    # Recuperar datos de la sesión si existen (cuando vuelves desde preview)
    datos_sesion = request.session.pop(f'emo_{emo_id}_datos', None)
    
    # Determinar qué proyecto usar para cargar subproyectos
    if datos_sesion and datos_sesion.get('proyecto_id'):
        try:
            proyecto_seleccionado = Proyecto.objects.get(id=datos_sesion['proyecto_id'])
            subproyectos_list = proyecto_seleccionado.subproyectos.filter(activo=True).order_by('nombre')
        except Proyecto.DoesNotExist:
            subproyectos_list = emo.proyecto.subproyectos.filter(activo=True).order_by('nombre') if emo.proyecto else []
    else:
        subproyectos_list = emo.proyecto.subproyectos.filter(activo=True).order_by('nombre') if emo.proyecto else []
    
    context = {
        'emo': emo,
        'trabajador': trabajador,
        'clinicas': Clinica.objects.all().order_by('nombre'),
        'empresas': Empresa.objects.all().order_by('nombre'),
        'proyectos': Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
        'subproyectos': subproyectos_list,
        'cargos': Cargo.objects.all().order_by('nombre'),
        'tipo_emo_choices': EMO._meta.get_field('tipo_emo').choices,
        'datos_sesion': datos_sesion,
    }
    
    return render(request, 'calidad/emos/editar_emo_programado.html', context)

@login_required
@group_required('Calidad')
def editar_emo(request, emo_id):
    """
    Permite editar los detalles de un EMO que ya ha sido 'Realizado'.
    Incluye lógica para eliminar o reemplazar el archivo PDF.
    """
    emo_a_editar = get_object_or_404(EMO, id=emo_id, estado='Realizado')
    trabajador = emo_a_editar.trabajador
    
    # Cargar controles para el modo edición (JS)
    controles_existentes = emo_a_editar.controles.all().order_by('fecha_programada')
    controles_data = [
        {'fecha': c.fecha_programada.strftime('%Y-%m-%d'), 'descripcion': c.descripcion}
        for c in controles_existentes
    ]
    
    if request.method == 'POST':
        print("\n" + "="*50, flush=True)
        print(f"[DEBUG] EDITAR EMO ID: {emo_id}", flush=True)
        
        form = RegistrarResultadoEmoForm(request.POST, request.FILES, instance=emo_a_editar)
        
        if form.is_valid():
            # 1. Guardar temporalmente (sin ir a BD aún)
            emo_editado = form.save(commit=False)
            
            # 2. Lógica de Archivos (Eliminar / Reemplazar)
            eliminar_pdf = form.cleaned_data.get('eliminar_archivo_pdf')
            nuevo_archivo = request.FILES.get('archivo_pdf')
            
            print(f"[DEBUG] Checkbox Eliminar: {eliminar_pdf}", flush=True)
            print(f"[DEBUG] Nuevo Archivo: {nuevo_archivo}", flush=True)

            # Si marcaron eliminar y NO subieron uno nuevo -> Borramos el actual
            if eliminar_pdf and not nuevo_archivo:
                print("[DEBUG] -> Ejecutando borrado de archivo...", flush=True)
                if emo_a_editar.archivo_pdf:
                    try:
                        # Borrado físico (importante para S3/Local)
                        emo_a_editar.archivo_pdf.delete(save=False)
                    except Exception as e:
                        print(f"[DEBUG] Error borrando archivo físico: {e}", flush=True)
                
                # Seteamos el campo en None
                emo_editado.archivo_pdf = None
            
            # Si subieron uno nuevo, Django maneja el reemplazo automáticamente en el save()
            
            # 3. Guardar cambios definitivos del EMO
            emo_editado.save()
            print("[DEBUG] EMO Guardado.", flush=True)

            # Actualizar la aptitud del trabajador
            trabajador.aptitud_actual = emo_editado.aptitud
            trabajador.save(update_fields=['aptitud_actual'])

            # --- Lógica de Controles (Reemplazo completo) ---
            # Borramos los anteriores
            controles_existentes.delete()
            
            cantidad_controles = int(request.POST.get('cantidad_controles', 0))
            if cantidad_controles > 0:
                for i in range(cantidad_controles):
                    fecha = request.POST.get(f'control_fecha_{i+1}')
                    descripcion = request.POST.get(f'control_descripcion_{i+1}')
                    if fecha and descripcion:
                        Control.objects.create(emo=emo_editado, fecha_programada=fecha, descripcion=descripcion)

            # --- Guardar cambios SIN enviar correo ---
            messages.success(request, f"El EMO para {trabajador} ha sido actualizado con éxito.")
            
            print("="*50 + "\n", flush=True)
            return redirect(f"{reverse('calidad:buscar_trabajador_info')}?dni={trabajador.dni}")
        else:
            print(f"[DEBUG] Errores formulario: {form.errors}", flush=True)
            messages.error(request, "Por favor, corrige los errores en el formulario.")
            
    else: # Método GET
        form = RegistrarResultadoEmoForm(instance=emo_a_editar)
        form.initial['cantidad_controles'] = controles_existentes.count()

    context = {
        'form': form,
        'emo': emo_a_editar,
        'trabajador': trabajador, # Aseguramos pasar trabajador
        'is_edit_mode': True,
        'controles_data_json': json.dumps(controles_data),
        'current_view': 'filtro_emos',
        'form_title': 'Editar Resultado de EMO',
    }
    
    return render(request, 'calidad/formularios/crear_emo_multistep.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def gestion_cargos(request):
    """Muestra el dashboard de tarjetas para la gestión de cargos."""
    context = {
        'current_view': 'gestion_cargos',
    }
    return render(request, 'calidad/cargos/gestion_cargos.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def lista_cargos(request):
    """Muestra una lista de todos los cargos registrados."""
    cargos = Cargo.objects.all().order_by('nombre')
    context = {
        'cargos': cargos,
        'current_view': 'gestion_cargos',
    }
    return render(request, 'calidad/cargos/lista_cargos.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def crear_cargo(request):
    """Muestra y procesa el formulario para crear un nuevo cargo."""
    if request.method == 'POST':
        form = CargoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo creado con éxito.')
            return redirect('lista_cargos')
    else:
        form = CargoForm()
    context = {
        'form': form,
        'form_title': 'Añadir Nuevo Cargo',
        'current_view': 'gestion_cargos',
    }
    return render(request, 'calidad/cargos/cargo_form_multistep.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def editar_cargo(request, pk):
    """Muestra y procesa el formulario para editar un cargo existente."""
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cargo actualizado con éxito.')
            return redirect('lista_cargos')
    else:
        form = CargoForm(instance=cargo)
    context = {
        'form': form,
        'form_title': 'Editar Cargo',
        'current_view': 'gestion_cargos',
    }
    return render(request, 'calidad/cargos/cargo_form.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def eliminar_cargo(request, pk):
    """Elimina un cargo, con protección si está en uso."""
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        # Comprobamos si el cargo está siendo usado por algún trabajador
        if cargo.trabajador_set.exists():
            messages.error(request, f'No se puede eliminar "{cargo.nombre}" porque está asignado a uno o más trabajadores.')
        else:
            cargo.delete()
            messages.success(request, f'Cargo "{cargo.nombre}" eliminado con éxito.')
        return redirect('lista_cargos')
    return redirect('lista_cargos')



@login_required
@group_required('Calidad')
def registrar_control(request, control_id):
    """
    Vista para marcar un control como realizado y subir su archivo de evidencia.
    """
    control = get_object_or_404(Control, id=control_id, realizado=False)
    
    if request.method == 'POST':
        # 1. Usa el nuevo formulario: RegistrarControlForm
        form = RegistrarControlForm(request.POST, request.FILES, instance=control)
        if form.is_valid():
            # El formulario ya guarda el archivo en la instancia de 'control'
            control_realizado = form.save(commit=False)
            
            # 2. Actualizamos los campos que no están en el formulario
            control_realizado.realizado = True
            control_realizado.fecha_realizacion = timezone.now()
            control_realizado.save()
            
            messages.success(request, f"Control '{control.descripcion}' para {control.emo.trabajador} registrado con éxito.")
            return redirect('calidad:lista_controles')
    else:
        # 3. Al cargar la página, también usamos el nuevo formulario
        form = RegistrarControlForm(instance=control)
        
    context = {
        'form': form,
        'control': control,
        'current_view': 'lista_controles' # Para el menú
    }
    return render(request, 'calidad/formularios/registrar_control.html', context)

# ==============================================================================
# 1. VISTA DE ESTADÍSTICAS DE CALIDAD
# ==============================================================================
       
@login_required
@group_required('Calidad', 'Recursos Humanos')
def estadisticas_calidad(request):
    try:
        # --- 1. OBTENCIÓN DE FILTROS ---
        filtro_empresa = request.GET.get('empresa', '')
        fecha_ini_str = request.GET.get('fecha_inicio', '')
        fecha_fin_str = request.GET.get('fecha_fin', '')

        hoy = timezone.now().date()

        # Configuración del rango de fechas
        if fecha_ini_str and fecha_fin_str:
            fecha_inicio = datetime.datetime.strptime(fecha_ini_str, '%Y-%m-%d').date()
            fecha_fin = datetime.datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        else:
            fecha_fin = hoy
            fecha_inicio = hoy - relativedelta(months=6)

        # --- 2. GENERAR LISTA MAESTRA DE MESES (EJE X) ---
        fechas_maestras = []
        labels_cronologicos = []
        
        fecha_cursor = fecha_inicio.replace(day=1)
        fecha_limite = fecha_fin.replace(day=1)

        while fecha_cursor <= fecha_limite:
            fechas_maestras.append(fecha_cursor)
            labels_cronologicos.append(fecha_cursor.strftime('%b %Y'))
            fecha_cursor += relativedelta(months=1)

        # --- 3. PREPARAR FILTROS BASE ---
        q_trabajador = Q(activo=True)
        if filtro_empresa:
            q_trabajador &= Q(empresa_id=filtro_empresa)

        q_base_relacionada = Q()
        if filtro_empresa:
            q_base_relacionada &= Q(trabajador__empresa_id=filtro_empresa)
        
        q_control_base = Q()
        if filtro_empresa:
            q_control_base &= Q(emo__trabajador__empresa_id=filtro_empresa)

        # ==============================================================================
        # 4. KPIs GLOBALES
        # ==============================================================================
        trabajadores_activos = Trabajador.objects.filter(q_trabajador)
        total_trabajadores = trabajadores_activos.count()
        
        trabajadores_con_emo_vencido = 0
        trabajadores_con_emo_vigente = 0
        estado_trabajadores = {} 

        for trabajador in trabajadores_activos.prefetch_related('historial_emo'):
            ultimo_emo = trabajador.historial_emo.filter(estado='Realizado').order_by('-fecha_realizacion').first()
            
            if ultimo_emo:
                if ultimo_emo.fecha_vencimiento and ultimo_emo.fecha_vencimiento < hoy:
                    trabajadores_con_emo_vencido += 1
                    estado_trabajadores[trabajador.id] = 'Vencido'
                else:
                    trabajadores_con_emo_vigente += 1
                    estado_trabajadores[trabajador.id] = 'Vigente'
            else:
                if trabajador.historial_emo.filter(estado='Programado').exists():
                     estado_trabajadores[trabajador.id] = 'Programado'
                else:
                     estado_trabajadores[trabajador.id] = 'Sin EMO'

        trabajadores_sin_emo = trabajadores_activos.filter(historial_emo__isnull=True).count()
        emos_pendientes_total = EMO.objects.filter(q_base_relacionada, estado='Programado').count()

        # ==============================================================================
        # 5. GRÁFICO 1: LECTURAS (CORREGIDO)
        # ==============================================================================
        
        # A. LECTURAS REALIZADAS (Confirmadas y con archivo)
        lecturas_realizadas_qs = EMO.objects.filter(
            q_base_relacionada,
            estado='Realizado', 
            confirmado_por_habilitador=True,
            fecha_realizacion__range=[fecha_inicio, fecha_fin]
        ).exclude(
            Q(archivo_confirmacion='') | Q(archivo_confirmacion__isnull=True)
        ).annotate(mes=TruncMonth('fecha_realizacion')).values('mes').annotate(total=Count('id'))

        # B. LECTURAS PENDIENTES TIPO 1: Programados (Aún no se hacen el examen)
        lecturas_programadas_qs = EMO.objects.filter(
            q_base_relacionada,
            estado='Programado', 
            fecha_programada__range=[fecha_inicio, fecha_fin]
        ).annotate(mes=TruncMonth('fecha_programada')).values('mes').annotate(total=Count('id'))

        # C. LECTURAS PENDIENTES TIPO 2: Realizados pero NO COMPLETADOS (Falta confirmar o falta archivo)
        lecturas_sin_leer_qs = EMO.objects.filter(
            q_base_relacionada,
            estado='Realizado',
            fecha_realizacion__range=[fecha_inicio, fecha_fin]
        ).filter(
            Q(confirmado_por_habilitador=False) | 
            Q(archivo_confirmacion__isnull=True) | 
            Q(archivo_confirmacion='')
        ).annotate(mes=TruncMonth('fecha_realizacion')).values('mes').annotate(total=Count('id'))

        data_lecturas_realizadas = []
        data_lecturas_pendientes = []

        # Helper para manejar fechas (evita errores de comparación datetime vs date)
        get_date = lambda d: d.date() if isinstance(d, datetime.datetime) else d

        for fecha_ref in fechas_maestras:
            # 1. Obtener Realizadas
            val_real = next((x['total'] for x in lecturas_realizadas_qs if x['mes'] and get_date(x['mes']).replace(day=1) == fecha_ref), 0)
            data_lecturas_realizadas.append(val_real)
            
            # 2. Obtener Pendientes (Programados + Sin Leer)
            val_prog = next((x['total'] for x in lecturas_programadas_qs if x['mes'] and get_date(x['mes']).replace(day=1) == fecha_ref), 0)
            val_sin_leer = next((x['total'] for x in lecturas_sin_leer_qs if x['mes'] and get_date(x['mes']).replace(day=1) == fecha_ref), 0)
            
            # Sumamos ambos como "Pendientes" para el gráfico
            data_lecturas_pendientes.append(val_prog + val_sin_leer)

        # ==============================================================================
        # 6. GRÁFICO 2: CONTROLES
        # ==============================================================================
        controles_realizados_qs = Control.objects.filter(
            q_control_base,
            realizado=True, 
            fecha_realizacion__range=[fecha_inicio, fecha_fin]
        ).annotate(mes=TruncMonth('fecha_realizacion')).values('mes').annotate(total=Count('id'))

        controles_pendientes_qs = Control.objects.filter(
            q_control_base,
            realizado=False, 
            fecha_programada__range=[fecha_inicio, fecha_fin]
        ).annotate(mes=TruncMonth('fecha_programada')).values('mes').annotate(total=Count('id'))

        data_controles_realizados = []
        data_controles_pendientes = []

        for fecha_ref in fechas_maestras:
            val_real = next((x['total'] for x in controles_realizados_qs if x['mes'] and get_date(x['mes']).replace(day=1) == fecha_ref), 0)
            data_controles_realizados.append(val_real)
            
            val_pend = next((x['total'] for x in controles_pendientes_qs if x['mes'] and get_date(x['mes']).replace(day=1) == fecha_ref), 0)
            data_controles_pendientes.append(val_pend)

        # ==============================================================================
        # 7. DATOS PARA GRÁFICO TORTA EMPRESAS
        # ==============================================================================
        empresas = Empresa.objects.all()
        if filtro_empresa:
            empresas = empresas.filter(id=filtro_empresa)

        labels_empresas = []
        data_empresas_total = []
        data_empresas_detalle = []

        for emp in empresas:
            trabs_empresa = trabajadores_activos.filter(empresa=emp)
            count = trabs_empresa.count()
            
            if count > 0:
                labels_empresas.append(emp.nombre)
                data_empresas_total.append(count)
                
                vencidos = 0
                vigentes = 0
                sin_emo = 0
                for t in trabs_empresa:
                    est = estado_trabajadores.get(t.id, 'Sin EMO')
                    if est == 'Vencido': vencidos += 1
                    elif est == 'Vigente': vigentes += 1
                    else: sin_emo += 1
                
                data_empresas_detalle.append({
                    'vigentes': vigentes,
                    'vencidos': vencidos,
                    'sin_emo': sin_emo
                })

        # ==============================================================================
        # 8. EVOLUCIÓN MENSUAL
        # ==============================================================================
        evolucion_mensual = EMO.objects.filter(
            q_base_relacionada,
            estado='Realizado', 
            fecha_realizacion__range=[fecha_inicio, fecha_fin]
        ).annotate(mes=TruncMonth('fecha_realizacion')).values('mes').annotate(
            nuevos_aptos=Count('id', filter=Q(aptitud='Apto')),
            nuevos_con_restriccion=Count('id', filter=Q(aptitud='Apto con Restricción')),
            nuevos_problemas=Count('id', filter=Q(aptitud__in=['Observado', 'No Apto']))
        )

        data_aptos_acum = []
        data_restriccion_acum = []
        data_problemas_acum = []
        
        for fecha_ref in fechas_maestras:
            dato = next((x for x in evolucion_mensual if x['mes'] and get_date(x['mes']).replace(day=1) == fecha_ref), None)
            if dato:
                data_aptos_acum.append(dato['nuevos_aptos'])
                data_restriccion_acum.append(dato['nuevos_con_restriccion'])
                data_problemas_acum.append(dato['nuevos_problemas'])
            else:
                data_aptos_acum.append(0)
                data_restriccion_acum.append(0)
                data_problemas_acum.append(0)

        # ==============================================================================
        # 9. CONTEXTO FINAL
        # ==============================================================================
        context = {
            'current_view': 'dashboard_calidad',
            'opciones_empresas': Empresa.objects.all().order_by('nombre'),
            'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else '',
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
            'total_trabajadores': total_trabajadores,
            'cobertura_emo_vigente': trabajadores_con_emo_vigente,
            'trabajadores_con_emo_vencido': trabajadores_con_emo_vencido,
            'trabajadores_sin_emo': trabajadores_sin_emo,
            'emos_pendientes_programados': emos_pendientes_total,
            'labels_lecturas': json.dumps(labels_cronologicos),
            'data_lecturas_realizadas': json.dumps(data_lecturas_realizadas),
            'data_lecturas_pendientes': json.dumps(data_lecturas_pendientes),
            'labels_obs': json.dumps(labels_cronologicos),
            'data_obs_pendientes': json.dumps(data_controles_pendientes),
            'data_obs_realizadas': json.dumps(data_controles_realizados),
            'labels_empresas': json.dumps(labels_empresas),
            'data_empresas_total': json.dumps(data_empresas_total),
            'data_empresas_detalle': json.dumps(data_empresas_detalle),
            'labels_evolucion': json.dumps(labels_cronologicos),
            'data_aptos': json.dumps(data_aptos_acum),
            'data_con_restriccion': json.dumps(data_restriccion_acum),
            'data_problemas': json.dumps(data_problemas_acum),
        }
        return render(request, 'calidad/estadisticas_calidaddos.html', context)
    
    except Exception as e:
        print(f"!!! ERROR EN ESTADISTICAS !!!: {str(e)}")
        raise e


# ==========================================
# APIs JSON PARA MODALES DEL DASHBOARD
# ==========================================

# Funcion auxiliar para construir data por proyectos
def construir_data_por_proyectos(trabajadores_queryset, filtro_contexto=None):
    """
    Recibe un QuerySet de trabajadores y construye la estructura JSON
    agrupada por Proyectos para las tarjetas del modal.
    Muestra TARJETAS tanto para Proyectos Asignados como para Proyectos con Historial de EMOs.
    param filtro_contexto: 'vencidos' | None. Si es 'vencidos', solo muestra tarjetas con estado 'vencido'.
    """
    hoy = timezone.now().date()
    data = []

    for t in trabajadores_queryset:
        proyectos_data = []
        
        # 1. Obtener Proyectos Asignados (Activos)
        # Convertimos a diccionario para acceso rápido a objetos
        asignaciones = t.asignaciones.select_related('proyecto', 'proyecto__parent').all()
        assigned_projects_map = {a.proyecto.id: a.proyecto for a in asignaciones}
        assigned_pids = set(assigned_projects_map.keys())

        # 2. Obtener Historial de EMOs Realizados (Para mostrar EMOs de proyectos antiguos o genéricos)
        # Traemos todos los EMOs realizados, ordenados por fecha (el primero será el más reciente)
        all_emos = t.historial_emo.filter(estado='Realizado').select_related('proyecto', 'proyecto__parent', 'lugar_examen').order_by('-fecha_realizacion', '-id')
        
        # Mapa: ProjectID -> Último EMO Realizado
        # Nota: ProjectID puede ser None (EMO Genérico)
        last_emo_map = {}
        projects_from_emos_map = {} # Mapa para tener los objetos proyecto de los EMOs históricos

        for emo in all_emos:
            pid = emo.proyecto_id
            if pid not in last_emo_map:
                last_emo_map[pid] = emo
                if pid is not None and emo.proyecto:
                    projects_from_emos_map[pid] = emo.proyecto

        # 3. Determinar lista unificada de proyectos a mostrar (Unión de Asignados + Históricos)
        all_pids = assigned_pids.union(set(last_emo_map.keys()))
        
        # Ordenamos: Primero proyectos con nombre, al final None (Genérico)
        # Helper para ordenar
        def sort_key(pid):
            if pid is None: return "ZZZZZ" # Al final
            p = assigned_projects_map.get(pid) or projects_from_emos_map.get(pid)
            return p.nombre if p else "ZZZZZ"

        sorted_pids = sorted(list(all_pids), key=sort_key)

        # Si no hay nada (ni asignación ni EMOs), mostrar tarjeta vacía
        if not sorted_pids:
             proyectos_data.append({
                'nombre': 'Sin Proyecto Asignado',
                'codigo': 'N/A',
                'estado': 'sin_registro',
                'emo': None
            })
        
        emo_ids_mostrados = set() # Para evitar duplicidad visual excesiva si se desea (opcional)

        for pid in sorted_pids:
            # Recuperar objeto Proyecto
            proyecto = assigned_projects_map.get(pid) or projects_from_emos_map.get(pid)
            
            # Datos básicos tarjeta
            nombre_proy = "Sin Proyecto (General)"
            codigo_proy = "N/A"
            if proyecto:
                nombre_proy = proyecto.nombre
                codigo_proy = proyecto.codigo
                if proyecto.parent:
                    codigo_proy = f"{proyecto.codigo} ({proyecto.parent.codigo})"
            
            # Buscar el EMO correspondiente
            # Prioridad 1: EMO exacto del proyecto (o None si es genérico)
            emo = last_emo_map.get(pid)
            
            # Prioridad 2: Si es un proyecto asignado pero NO tiene EMO propio, buscamos fallback
            # (Solo para mostrar cumplimiento en proyectos actuales)
            is_assigned = pid in assigned_pids
            if not emo and is_assigned and proyecto:
                # Intentamos buscar EMO del Padre
                if proyecto.parent_id and proyecto.parent_id in last_emo_map:
                    emo = last_emo_map[proyecto.parent_id]
                # Intentamos buscar EMO Genérico
                elif None in last_emo_map:
                    emo = last_emo_map[None]
            
            # Lógica para ocultar tarjeta genérica si ya se usó para cubrir un proyecto asignado?
            # El usuario pidió ver "2 EMOs". Si mostramos el EMO genérico en la tarjeta del proyecto,
            # y luego otra tarjeta "General" con el mismo EMO, es redundante.
            # Vamos a filtrar: Si pid es None (Genérico) y ese EMO ya se mostró en otra tarjeta, lo saltamos.
            if pid is None and emo and emo.id in emo_ids_mostrados:
                continue

            # Construir info del EMO
            estado_key = 'sin_registro'
            emo_info = None

            if emo:
                # Registrar que mostramos este EMO
                emo_ids_mostrados.add(emo.id)

                fecha_venc = emo.fecha_vencimiento
                if fecha_venc:
                    if fecha_venc < hoy:
                        estado_key = 'vencido'
                    elif fecha_venc <= (hoy + timezone.timedelta(days=30)):
                        estado_key = 'por_vencer'
                    else:
                        estado_key = 'vigente'
                
                emo_info = {
                    'tipo': emo.get_tipo_emo_display(),
                    'fecha': emo.fecha_realizacion.strftime('%d/%m/%Y') if emo.fecha_realizacion else '-',
                    'vencimiento': fecha_venc.strftime('%d/%m/%Y') if fecha_venc else '-',
                    'clinica': emo.lugar_examen.nombre if emo.lugar_examen else '-',
                    'aptitud': emo.aptitud
                }
            
            # --- FILTRO DE CONTEXTO ---
            # Si estamos en la vista de 'vencidos', solo mostrar tarjetas vencidas.
            if filtro_contexto == 'vencidos':
                if estado_key != 'vencido':
                    continue
            # --------------------------

            proyectos_data.append({
                'nombre': nombre_proy,
                'codigo': codigo_proy,
                'estado': estado_key,
                'emo': emo_info
            })

        # Agregamos la fila del trabajador con sus tarjetas
        # Solo agregamos si tiene proyectos para mostrar tras el filtro
        if proyectos_data or not filtro_contexto: 
        # Si hay filtro activo y no quedo ningun proyecto, no agregamos al trabajador (consistencia visual)
        # Ojo: Si el trabajador entró al QuerySet es por tener algo vencido. Si aquí filtramos y queda vacio,
        # significaría que la lógica de "estado_key" difiere de la Query.
            if proyectos_data:
                data.append({
                    'dni': t.dni,
                    'nombre_completo': f"{t.apellido_paterno} {t.apellido_materno}, {t.nombres}",
                    'empresa': t.empresa.nombre if t.empresa else "-",
                    'proyectos': proyectos_data 
                })
        
    return data


@login_required
def api_trabajadores_vigentes(request):
    """
    Retorna trabajadores que tienen AL MENOS UN EMO vigente.
    El detalle mostrará el estado de TODOS sus proyectos.
    """
    hoy = timezone.now().date()
    filtro_empresa = request.GET.get('empresa', '').strip()
    
    # Filtramos trabajadores activos que tengan algún EMO vigente hoy
    trabajadores = Trabajador.objects.filter(
        activo=True,
        historial_emo__estado='Realizado',
        historial_emo__fecha_vencimiento__gte=hoy
    ).distinct().select_related('empresa').prefetch_related('asignaciones__proyecto')

    if filtro_empresa:
        trabajadores = trabajadores.filter(empresa_id=filtro_empresa)

    data_procesada = construir_data_por_proyectos(trabajadores)
    return JsonResponse({'data': data_procesada, 'total': len(data_procesada)})

@login_required
def api_trabajadores_vencidos(request):
    """
    Retorna trabajadores que tienen AL MENOS UN EMO vencido.
    """
    hoy = timezone.now().date()
    filtro_empresa = request.GET.get('empresa', '').strip()
    
    # Filtramos trabajadores activos con algún EMO vencido
    trabajadores = Trabajador.objects.filter(
        activo=True,
        historial_emo__estado='Realizado',
        historial_emo__fecha_vencimiento__lt=hoy
    ).distinct().select_related('empresa').prefetch_related('asignaciones__proyecto')

    if filtro_empresa:
        trabajadores = trabajadores.filter(empresa_id=filtro_empresa)
    
    data_procesada = construir_data_por_proyectos(trabajadores, filtro_contexto='vencidos')
    return JsonResponse({'data': data_procesada, 'total': len(data_procesada)})

@login_required
def api_emos_programados(request):
    """
    Muestra los trabajadores con EMOs programados.
    Ahora muestra tarjetas unificadas: Proyectos Asignados + Proyectos Programados
    """
    hoy = timezone.now().date()
    filtro_empresa = request.GET.get('empresa', '').strip()
    
    # Trabajadores con EMOs programados pendientes
    trabajadores = Trabajador.objects.filter(
        activo=True,
        historial_emo__estado='Programado'
    ).distinct().select_related('empresa').prefetch_related('asignaciones__proyecto')

    if filtro_empresa:
        trabajadores = trabajadores.filter(empresa_id=filtro_empresa)

    data = []
    for t in trabajadores:
        proyectos_data = []
        
        # 1. Obtener Proyectos Asignados
        asignaciones = t.asignaciones.select_related('proyecto', 'proyecto__parent').all()
        assigned_projects_map = {a.proyecto.id: a.proyecto for a in asignaciones}
        assigned_pids = set(assigned_projects_map.keys())

        # 2. Obtener Programaciones Pendientes
        all_programados = t.historial_emo.filter(estado='Programado').select_related('proyecto', 'proyecto__parent', 'lugar_examen').order_by('fecha_programada')
        
        # Mapa: ProjectID -> Primer EMO Programado
        first_prog_map = {}
        projects_from_emos_map = {} 

        for emo in all_programados:
            pid = emo.proyecto_id
            if pid not in first_prog_map:
                first_prog_map[pid] = emo
                if pid is not None and emo.proyecto:
                    projects_from_emos_map[pid] = emo.proyecto

        # 3. Determinar lista unificada
        all_pids = assigned_pids.union(set(first_prog_map.keys()))
        
        def sort_key(pid):
            if pid is None: return "ZZZZZ"
            p = assigned_projects_map.get(pid) or projects_from_emos_map.get(pid)
            return p.nombre if p else "ZZZZZ"

        sorted_pids = sorted(list(all_pids), key=sort_key)
        
        if not sorted_pids:
             proyectos_data.append({
                 'nombre': 'Sin Proyecto', 
                 'codigo': '-', 
                 'estado': 'sin_registro', 
                 'emo': None
             })

        emo_ids_mostrados = set()

        for pid in sorted_pids:
            # Recuperar Proyecto
            proyecto = assigned_projects_map.get(pid) or projects_from_emos_map.get(pid)
            
            nombre_proy = "Sin Proyecto (Programado)"
            codigo_proy = "-"
            if proyecto:
                nombre_proy = proyecto.nombre
                codigo_proy = proyecto.codigo

            # Buscar Programación correspondiente
            prog = first_prog_map.get(pid)
            
            # Fallback para asignados
            is_assigned = pid in assigned_pids
            if not prog and is_assigned and proyecto:
                if proyecto.parent_id and proyecto.parent_id in first_prog_map:
                    prog = first_prog_map[proyecto.parent_id]
                elif None in first_prog_map:
                    prog = first_prog_map[None]
            
            # Evitar duplicados de generic
            if pid is None and prog and prog.id in emo_ids_mostrados:
                continue

            estado_key = 'sin_registro'
            emo_info = None
            
            if prog:
                emo_ids_mostrados.add(prog.id)
                estado_key = 'por_vencer'
                
                dias_restantes = None
                if prog.fecha_programada:
                    dias_restantes = (prog.fecha_programada - hoy).days
                
                msg_vencimiento = f"Faltan {dias_restantes} días" if dias_restantes and dias_restantes >= 0 else "Programación Atrasada"
                if dias_restantes == 0: msg_vencimiento = "ES HOY"

                emo_info = {
                    'tipo': prog.get_tipo_emo_display(),
                    'fecha': 'Programado',
                    'vencimiento': prog.fecha_programada.strftime('%d/%m/%Y') if prog.fecha_programada else '-',
                    'clinica': prog.lugar_examen.nombre if prog.lugar_examen else '-',
                    'aptitud': msg_vencimiento
                }
            
            proyectos_data.append({
                'nombre': nombre_proy,
                'codigo': codigo_proy,
                'estado': estado_key,
                'emo': emo_info
            })
            
        data.append({
            'dni': t.dni,
            'nombre_completo': f"{t.apellido_paterno} {t.apellido_materno}, {t.nombres}",
            'empresa': t.empresa.nombre if t.empresa else "-",
            'proyectos': proyectos_data
        })

    return JsonResponse({'data': data, 'total': len(data)})

@login_required
def api_trabajadores_sin_emo(request):
    """API que retorna trabajadores sin EMO en formato JSON"""
    from django.http import JsonResponse
    filtro_empresa = request.GET.get('empresa', '').strip()

    trabajadores = Trabajador.objects.filter(
        activo=True,
        historial_emo__isnull=True
    ).select_related('empresa', 'area').distinct()

    if filtro_empresa:
        trabajadores = trabajadores.filter(empresa_id=filtro_empresa)
    
    data = []
    for trabajador in trabajadores:
        data.append({
            'dni': trabajador.dni,
            'nombre_completo': f"{trabajador.apellido_paterno} {trabajador.apellido_materno}, {trabajador.nombres}",
            'empresa': trabajador.empresa.nombre if trabajador.empresa else "-",
            'area': trabajador.area.nombre if trabajador.area else "-",
            'email': trabajador.email or "-",
            'telefono': trabajador.telefono or "-"
        })
    
    return JsonResponse({'data': data, 'total': len(data)})

def trigger_daily_report(request):
    """
    Esta vista es llamada por el scheduler externo.
    Verifica la clave secreta y lanza la tarea de envío de reportes en un hilo separado.
    """
    secret_key = request.GET.get('secret_key')
    
    if secret_key != settings.CRON_SECRET_KEY:
        return HttpResponseForbidden('Acceso denegado: Clave secreta inválida.')

    # ¡CAMBIA LA FUNCIÓN QUE SE EJECUTA!
    email_thread = threading.Thread(target=enviar_reportes_programados)
    email_thread.daemon = True
    email_thread.start()
    
    return HttpResponse('OK: Tarea de reportes programados iniciada en segundo plano.')


#=======================================================
# EXPORTACIÓN DE REPORTES EMOS
#=======================================================

class DummyEMO:
    """ Clase auxiliar para representar trabajadores sin EMO en al exportar a Excel """
    def __init__(self, trabajador):
        self.trabajador = trabajador
        self.trabajador_id = trabajador.id
        self.id = None
        asig = trabajador.asignaciones.filter(activo=True).first()
        self.proyecto = asig.proyecto if asig else None
        self.cargo = asig.cargo if asig else None
        
        self.lugar_examen = None
        self.fecha_realizacion = None
        self.fecha_vencimiento = None
        self.fecha_programada = None
        self.fecha_confirmacion = None
        self.fecha_validacion_subsana = None
        self.restriccion = ""
        self.comentario = ""
        
        self.esta_vencido = False
        self.esta_por_vencer = False
        self.esta_vigente = False

    def get_tipo_emo_display(self): 
        return ""
    
    def get_aptitud_display(self): 
        return ""

    @property
    def controles(self):
        class _Controles:
            def all(self): return []
        return _Controles()


def _obtener_emos_unicos_por_trabajador_proyecto_empresa(emos_queryset):
    """
    Devuelve solo el EMO más reciente por combinación trabajador-proyecto-empresa,
    replicando la lógica usada en Reporte de EMOs para evitar duplicados en exportación.
    """
    emos_unicos = []
    emos_vistos = set()

    for emo in emos_queryset.order_by('trabajador__id', '-id'):
        trabajador_id = emo.trabajador_id
        proyecto_id = emo.subproyecto_id or emo.proyecto_id
        empresa_id = emo.empresa_id or emo.trabajador.empresa_id
        clave = (trabajador_id, proyecto_id, empresa_id)

        if clave not in emos_vistos:
            emos_vistos.add(clave)
            emos_unicos.append(emo)

    return emos_unicos

@login_required
@group_required('Calidad')
def pagina_exportar_reportes(request):
    """
    Muestra un formulario con filtros para generar y descargar un reporte de EMOs en Excel.
    """
    filtro_proyecto_padre_id = request.GET.get('proyecto_padre', '')
    filtro_subproyecto_id = request.GET.get('subproyecto', '')  # NUEVO FILTRO
    filtro_empresa = request.GET.get('empresa', '')            # NUEVO FILTRO
    filtro_mes = request.GET.get('mes', '')
    # Año exacto y rango de años (todo opcional)
    filtro_ano = request.GET.get('ano', '')
    filtro_ano_desde = request.GET.get('ano_desde', '')
    filtro_ano_hasta = request.GET.get('ano_hasta', '')
    
    if 'exportar' in request.GET:
        # --- CONSULTA OPTIMIZADA PARA M2M ---
        emos_queryset = EMO.objects.select_related(
            'trabajador', 'trabajador__empresa', 
            'empresa',
            'proyecto', 'proyecto__parent', 
            'subproyecto', 'subproyecto__parent',
            'lugar_examen'
        ).prefetch_related(
            'controles',
        ).all()

        # --- APLICACIÓN DE FILTROS ---
        if filtro_empresa:
            # Igual que en Reporte de EMOs: empresa se toma del propio EMO.
            emos_queryset = emos_queryset.filter(empresa__pk=filtro_empresa)

        if filtro_subproyecto_id:
            # Igual que en Reporte de EMOs: filtrar por proyecto/subproyecto del EMO.
            emos_queryset = emos_queryset.filter(
                Q(subproyecto__pk=filtro_subproyecto_id) |
                Q(proyecto__pk=filtro_subproyecto_id)
            ).distinct()
        elif filtro_proyecto_padre_id:
            # Igual que en Reporte de EMOs: padre/hijos en campos del EMO.
            emos_queryset = emos_queryset.filter(
                Q(proyecto__pk=filtro_proyecto_padre_id) | 
                Q(proyecto__parent__pk=filtro_proyecto_padre_id) |
                Q(subproyecto__parent__pk=filtro_proyecto_padre_id)
            ).distinct()

        # Filtro por rango de años tiene prioridad sobre año exacto
        ano_desde_int = int(filtro_ano_desde) if str(filtro_ano_desde).isdigit() else None
        ano_hasta_int = int(filtro_ano_hasta) if str(filtro_ano_hasta).isdigit() else None

        if ano_desde_int and ano_hasta_int and ano_desde_int > ano_hasta_int:
            ano_desde_int, ano_hasta_int = ano_hasta_int, ano_desde_int

        if ano_desde_int or ano_hasta_int:
            # IMPORTANTE: el rango de años filtra SOLO por fecha_programada
            cond_programada = Q()
            if ano_desde_int:
                cond_programada &= Q(fecha_programada__year__gte=ano_desde_int)
            if ano_hasta_int:
                cond_programada &= Q(fecha_programada__year__lte=ano_hasta_int)
            emos_queryset = emos_queryset.filter(cond_programada)
        elif filtro_ano and str(filtro_ano).isdigit():
            # IMPORTANTE: el año exacto filtra SOLO por fecha_programada
            ano_int = int(filtro_ano)
            emos_queryset = emos_queryset.filter(fecha_programada__year=ano_int)

        if filtro_mes and str(filtro_mes).isdigit():
            # IMPORTANTE: el mes filtra SOLO por fecha_programada
            mes_int = int(filtro_mes)
            emos_queryset = emos_queryset.filter(fecha_programada__month=mes_int)
        
        # Empaquetamos los filtros para usarlos en la lógica del Excel
        filtros_aplicados = {
            'proyecto_id': filtro_subproyecto_id,
            'proyecto_padre_id': filtro_proyecto_padre_id,
            'empresa_id': filtro_empresa
        }

        # Evita que un mismo trabajador salga varias veces por historial de EMOs.
        emos_list = _obtener_emos_unicos_por_trabajador_proyecto_empresa(emos_queryset)

        # --- REQUERIMIENTO: incluir trabajadores sin EMO en blanco ---
        # Se consideran solo trabajadores realmente sin historial de EMO,
        # para mantener congruencia con Reporte de EMOs.
        trabajadores_con_emo = [emo.trabajador_id for emo in emos_list if emo.trabajador_id]

        trabajadores_faltantes = Trabajador.objects.filter(
            historial_emo__isnull=True,
            activo=True
        )

        if filtro_empresa:
            trabajadores_faltantes = trabajadores_faltantes.filter(empresa__pk=filtro_empresa)

        if filtro_subproyecto_id:
            trabajadores_faltantes = trabajadores_faltantes.filter(
                asignaciones__proyecto__pk=filtro_subproyecto_id
            ).distinct()
        elif filtro_proyecto_padre_id:
            trabajadores_faltantes = trabajadores_faltantes.filter(
                Q(asignaciones__proyecto__pk=filtro_proyecto_padre_id)
                | Q(asignaciones__proyecto__parent__pk=filtro_proyecto_padre_id)
            ).distinct()

        trabajadores_faltantes = trabajadores_faltantes.exclude(id__in=trabajadores_con_emo).prefetch_related(
            'asignaciones__proyecto',
            'asignaciones__cargo'
        ).distinct()

        for t in trabajadores_faltantes:
            emos_list.append(DummyEMO(t))
        
        return exportar_emos_a_excel(emos_list, filtros_aplicados)

    # --- PREPARACIÓN DE DATOS PARA EL FORMULARIO ---
    subproyectos_map = {}
    all_subproyectos = Proyecto.objects.filter(parent__isnull=False, activo=True)
    for sub in all_subproyectos:
        if sub.parent_id not in subproyectos_map: subproyectos_map[sub.parent_id] = []
        subproyectos_map[sub.parent_id].append({'id': sub.id, 'nombre': sub.nombre, 'codigo': sub.codigo or ''})
    
    import json
    
    context = {
        'current_view': 'exportar_reportes',
        'titulo': 'Exportar Reporte de EMOs',
        
        'opciones_empresas': Empresa.objects.all(),
        'opciones_proyectos_padre': Proyecto.objects.filter(parent__isnull=True, activo=True).order_by('nombre'),
        'opciones_anos': range(timezone.now().year, 2022, -1),
        'opciones_meses': [(1, 'Enero'), (2, 'Febrero'), (3, 'Marzo'), (4, 'Abril'), (5, 'Mayo'), (6, 'Junio'), (7, 'Julio'), (8, 'Agosto'), (9, 'Septiembre'), (10, 'Octubre'), (11, 'Noviembre'), (12, 'Diciembre')],
        
        'subproyectos_map_json': json.dumps(subproyectos_map),
        
        'filtro_activo_empresa': int(filtro_empresa) if filtro_empresa else None,
        'filtro_activo_proyecto_padre': int(filtro_proyecto_padre_id) if filtro_proyecto_padre_id else None,
        'filtro_activo_subproyecto': int(filtro_subproyecto_id) if filtro_subproyecto_id else None,
        'filtro_activo_ano': int(filtro_ano) if str(filtro_ano).isdigit() else None,
        'filtro_activo_ano_desde': int(filtro_ano_desde) if str(filtro_ano_desde).isdigit() else None,
        'filtro_activo_ano_hasta': int(filtro_ano_hasta) if str(filtro_ano_hasta).isdigit() else None,
        'filtro_activo_mes': int(filtro_mes) if filtro_mes.isdigit() else None,
    }
    
    return render(request, 'calidad/reportes/pagina_exportar_reportes.html', context)


def exportar_emos_a_excel(emos, filtros):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_EMOs_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte EMOs"

    # ==============================================================================
    # 1. ESTILOS DEFINIDOS
    # ==============================================================================
    font_bold = Font(bold=True)
    font_bold_white = Font(bold=True, color="FFFFFF")
    font_main_title = Font(name='Arial Narrow', size=24, bold=True) 
    
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_center_middle = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    fill_dark_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # --- ESTILOS DE ALERTA ---
    fill_red = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid") 
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    font_white_data = Font(color="FFFFFF", name='Arial Narrow', size=11)
    font_black_data = Font(color="000000", name='Arial Narrow', size=11)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    font_arial_narrow = Font(name='Arial Narrow', size=11)

    # ==============================================================================
    # 2. DIMENSIONES DE COLUMNAS Y FILAS
    # ==============================================================================
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[4].height = 31
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[7].height = 45

    # Ajustamos anchos incluyendo las nuevas columnas D y E
    # Ahora la tabla llega hasta la letra W
    column_widths = {
        'A': 35, 'B': 11, 'C': 25, 
        'D': 18, 'E': 12, # NUEVAS: TIPO EMO y ESTADO
        'F': 11, 'G': 11, # Antes D y E (Vigencia)
        'H': 24, 'I': 25, 'J': 18, 'K': 18, 'L': 22, 'M': 18, 'N': 25, 
        'O': 15, 'P': 25, 'Q': 25, 'R': 30, 'S': 15, 'T': 15, 'U': 15, 'V': 15, 'W': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ==============================================================================
    # 3. LOGO Y TÍTULO PRINCIPAL
    # ==============================================================================
    try:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_excel.png')
        if os.path.exists(logo_path):
            ws.merge_cells('A1:B3')
            img = OpenpyxlImage(logo_path)
            img.width = 200
            img.height = 80
            ws.add_image(img, 'A1')
    except Exception as e:
        print(f"No se pudo cargar el logo: {e}")
    
    # Expandimos el título principal hasta la W
    ws.merge_cells('C1:W3') 
    cell = ws['C1']
    cell.value = "CRONOGRAMA DE EXAMENES MEDICOS"
    cell.font = font_main_title
    cell.alignment = alignment_center
    
    # ==============================================================================
    # 4. LÓGICA DE ENCABEZADO DINÁMICO
    # ==============================================================================
    proyecto_titulo = "S/C" 
    empresa_titulo = "CENERIS E.I.R.L." 
    empresa_obj = None 
    
    target_project_id = None
    if filtros.get('proyecto_id'):
        target_project_id = int(filtros['proyecto_id'])
    elif filtros.get('proyecto_padre_id'):
        target_project_id = int(filtros['proyecto_padre_id'])

    if target_project_id:
        p = Proyecto.objects.filter(pk=target_project_id).first()
        if p:
            if p.parent:
                proyecto_titulo = p.parent.codigo if p.parent.codigo else p.parent.nombre
            else:
                proyecto_titulo = p.codigo if p.codigo else p.nombre
            if p.empresa:
                empresa_obj = p.empresa

    if filtros.get('empresa_id'):
        try:
            empresa_filtro = Empresa.objects.filter(pk=int(filtros['empresa_id'])).first()
            if empresa_filtro:
                empresa_obj = empresa_filtro
        except (ValueError, TypeError):
            pass

    if empresa_obj:
        empresa_titulo = empresa_obj.nombre
    else:
        empresa_default = Empresa.objects.filter(nombre__icontains="CENERIS").first()
        if empresa_default:
            empresa_obj = empresa_default

    # Dibujar Encabezados (Ajustamos los merges para cubrir hasta la W)
    ws.merge_cells('A4:E4'); ws['A4'].value = "1. RAZÓN SOCIAL O DENOMINACIÓN SOCIAL"
    ws.merge_cells('F4:M4'); ws['F4'].value = "2. RUC"
    ws.merge_cells('N4:P4'); ws['N4'].value = "3. DOMICILIO"
    ws.merge_cells('Q4:R4'); ws['Q4'].value = "4.- NOMBRE Y N° DE CONTRATO"
    ws.merge_cells('S4:T4'); ws['S4'].value = "5. N° DE TRABAJADORES"
    ws.merge_cells('U4:U4'); ws['U4'].value = "6. FECHA"
    ws.merge_cells('V4:W4'); ws['V4'].value = "7. RESPONSABLE"
    
    for cell in ws[4]:
        cell.fill = fill_dark_gray
        cell.font = font_bold_white
        cell.alignment = alignment_center_wrap
    
    ws.merge_cells('A5:E5'); cell = ws['A5']; cell.value = empresa_titulo
    ws.merge_cells('F5:M5'); cell = ws['F5']; cell.value = empresa_obj.ruc if empresa_obj else ""
    ws.merge_cells('N5:P5'); cell = ws['N5']; cell.value = empresa_obj.direccion if empresa_obj else ""
    ws.merge_cells('Q5:R5'); cell = ws['Q5']; cell.value = proyecto_titulo
    
    unique_workers = set(e.trabajador_id for e in emos)
    ws.merge_cells('S5:T5'); cell = ws['S5']; cell.value = len(unique_workers)
    ws.merge_cells('U5:U5'); cell = ws['U5']; cell.value = timezone.now().date(); cell.number_format = 'DD/MM/YYYY'
    ws.merge_cells('V5:W5'); cell = ws['V5']; cell.value = ""

    for cell in ws[5]:
        cell.alignment = alignment_center

    # ==============================================================================
    # 5. CABECERAS DE LA TABLA PRINCIPAL (Ajustado con nuevas columnas)
    # ==============================================================================
    headers = {
        'A6': ("APELLIDOS, Nombres", 'A6:A7'), 
        'B6': ("DNI", 'B6:B7'), 
        'C6': ("Cargo(s)", 'C6:C7'),
        
        # --- NUEVAS COLUMNAS ---
        'D6': ("TIPO EMO", 'D6:D7'), 
        'E6': ("ESTADO", 'E6:E7'), 
        
        # --- DESPLAZAMIENTO DE LAS DEMÁS (Todo +2 columnas) ---
        'F6': ("VIGENCIA", 'F6:G6'), 'F7': ("Inicio", None), 'G7': ("Final", None),
        'H6': ("CLINICA", 'H6:H7'), 
        'I6': ("PROYECTO", 'I6:I7'), 
        'J6': ("Codigo", 'J6:J7'),
        'K6': ("Telefono", 'K6:K7'),
        'L6': ("LECTURA DE EMO", 'L6:M6'), 'L7': ("Programada", None), 'M7': ("Realizada", None),
        'N6': ("PRE INICIO", 'N6:O6'), 'N7': ("Detalle", None), 'O7': ("Levantamiento", None),
        'P6': ("RESTRICCIONES", 'P6:P7'), 
        'Q6': ("OBSERVACIONES", 'Q6:Q7'),
        'R6': ("CONTROLES", 'R6:S6'), 'R7': ("Detalle", None), 'S7': ("Fecha", None),
        'T6': ("EMOR", 'T6:U6'), 'T7': ("Programada", None), 'U7': ("Realizada", None),
        'V6': ("APTO", 'V6:V7'), 
        'W6': ("Programación", 'W6:W7'), # Antes era U
    }
    for cell_ref, (value, merge_range) in headers.items():
        if merge_range: ws.merge_cells(merge_range)
        cell = ws[cell_ref]
        cell.value = value
        cell.fill = fill_light_gray
        cell.font = font_bold
        cell.alignment = alignment_center_wrap
    
    # ==============================================================================
    # 6. LLENADO DE DATOS (BUCLE PRINCIPAL)
    # ==============================================================================
    
    def make_naive(value):
        if isinstance(value, datetime.datetime) and timezone.is_aware(value):
            return timezone.localtime(value).replace(tzinfo=None)
        return value
    
    current_row = 7 
    for emo in emos:
        current_row += 1
        
        # --- Datos Proyecto ---
        val_proyecto_padre = ""
        val_subproyecto = ""
        if emo.proyecto:
            if emo.proyecto.parent:
                obj_padre = emo.proyecto.parent
                obj_hijo = emo.proyecto
            else:
                obj_padre = emo.proyecto
                obj_hijo = None 
            val_proyecto_padre = obj_padre.codigo if (obj_padre.codigo and obj_padre.codigo.strip()) else obj_padre.nombre 
            if obj_hijo:
                val_subproyecto = obj_hijo.codigo if (obj_hijo.codigo and obj_hijo.codigo.strip()) else obj_hijo.nombre
            else:
                val_subproyecto = ""
        else:
            val_proyecto_padre = "Sin Asignación"

        # --- Otros Datos ---
        nombre_cargo = emo.cargo.nombre if emo.cargo else "Sin Cargo"
        telefono = getattr(emo.trabajador, 'telefono', '')
        
        # Estado Programación
        estado_programacion = ""
        if emo.esta_vencido: estado_programacion = "Vencido"
        elif emo.esta_por_vencer: estado_programacion = "Por Vencer"
        elif emo.esta_vigente: estado_programacion = "Vigente"

        # --- Estado Trabajador ---
        es_activo = getattr(emo.trabajador, 'activo', True) 
        estado_trabajador_str = "ACTIVO" if es_activo else "CESADO"
        
        nombre_completo = f"{emo.trabajador.apellido_paterno} {emo.trabajador.apellido_materno}, {emo.trabajador.nombres}"
        if not es_activo:
            nombre_completo += " (CESADO)"

        row_data = [
            nombre_completo,    # A
            emo.trabajador.dni, # B
            nombre_cargo,       # C
            
            # --- NUEVAS COLUMNAS ---
            emo.get_tipo_emo_display(), # D: Tipo de EMO
            estado_trabajador_str,      # E: Estado del Trabajador
            
            # --- DATOS DESPLAZADOS ---
            make_naive(emo.fecha_realizacion), # F (Inicio)
            make_naive(emo.fecha_vencimiento), # G (Final)
            emo.lugar_examen.nombre if emo.lugar_examen else "N/A", # H
            val_proyecto_padre,  # I
            val_subproyecto,     # J 
            telefono, # K
            make_naive(emo.fecha_programada), # L
            make_naive(emo.fecha_confirmacion), # M
            emo.get_aptitud_display(), # N
            make_naive(emo.fecha_validacion_subsana), # O
            emo.restriccion, # P
            emo.comentario,  # Q
            "\n".join([c.descripcion for c in emo.controles.all()]),  # R
            "\n".join([c.fecha_programada.strftime("%d/%m/%Y") if c.fecha_programada else "" for c in emo.controles.all()]), # S
            "", "", # T, U (EMOR)
            emo.get_aptitud_display(), # V (Apto)
            estado_programacion, # W (Programación)
        ]
        ws.append(row_data)
        ws.row_dimensions[current_row].height = 36

        # ==============================================================================
        # 7. COLOREADO CONDICIONAL DE CELDAS
        # ==============================================================================
        
        # A) NOMBRE EN ROJO SI ESTÁ CESADO (Columna A)
        celda_nombre = ws[f'A{current_row}']
        if not es_activo:
            celda_nombre.fill = fill_red
            celda_nombre.font = font_white_data 

        # B) ESTADO PROGRAMACIÓN EN ROJO/AMARILLO (AHORA COLUMNA W)
        celda_programacion = ws[f'W{current_row}']
        
        if estado_programacion == "Vencido":
            celda_programacion.fill = fill_red
            celda_programacion.font = font_white_data
        elif estado_programacion == "Por Vencer":
            celda_programacion.fill = fill_yellow
            celda_programacion.font = font_black_data

    # ==============================================================================
    # 8. ESTILOS FINALES (BORDES Y FUENTES GENERALES)
    # ==============================================================================
    full_range = f'A4:W{ws.max_row}' # Rango extendido hasta W
    for row in ws[full_range]:
        for cell in row:
            cell.border = thin_border

    if ws.max_row >= 8:
        for row in ws[f'A8:W{ws.max_row}']:
            for cell in row:
                
                es_celda_cesado = (cell.column_letter == 'A' and not getattr(cell.value, 'es_activo', True) and "CESADO" in str(cell.value))
                # La columna W ahora es la de Vencido
                es_celda_vencido = (cell.column_letter == 'W' and cell.value == "Vencido")
                
                # Solo aplicamos estilo base si NO es una celda coloreada por alerta
                # (Mejoramos la lógica comprobando el fill actual)
                if cell.fill.start_color.index == '00000000': # Si no tiene color de fondo (transparente/blanco)
                     cell.font = font_arial_narrow
                
                cell.alignment = alignment_center_middle
                
                if isinstance(cell.value, datetime.datetime):
                    cell.number_format = 'DD/MM/YYYY'
                elif isinstance(cell.value, datetime.date):
                    cell.number_format = 'DD/MM/YYYY'
    
    wb.save(response)
    return response

@login_required
@group_required('Calidad')
def seleccionar_edicion_emo(request, emo_id):
    emo = get_object_or_404(EMO, id=emo_id)
    
    context = {
        'emo': emo,
        'form_title': f'Gestión de EMO: {emo.trabajador}',
    }
    return render(request, 'calidad/emos/editar_emo_opciones.html', context)

@login_required
@group_required('Calidad')
def ver_documentos_emo(request, emo_id):
    emo = get_object_or_404(EMO, id=emo_id)
    
    # --- LÓGICA DE PERMISOS ---
    # Verificamos si el usuario pertenece al grupo 'Doctor' (o es superusuario)
    # Cambia 'Doctor' por el nombre exacto de tu grupo en el Admin
    es_doctor = request.user.groups.filter(name='Doctor').exists() or request.user.is_superuser

    context = {
        'emo': emo,
        'titulo': f'Documentación EMO: {emo.trabajador}',
        'es_doctor': es_doctor, # <--- Pasamos esta variable al HTML
    }
    return render(request, 'calidad/documentos/ver_documentos_emo.html', context)

@login_required
@group_required('Calidad')
def eliminar_emo(request, emo_id):
    emo = get_object_or_404(EMO, id=emo_id)
    
    if request.method == 'POST':
        trabajador_nombre = str(emo.trabajador)
        emo.delete()
        messages.success(request, f"El EMO de {trabajador_nombre} ha sido eliminado correctamente.")
        return redirect('calidad:reporte_maestro_emos')
        
    context = {
        'emo': emo,
        'titulo': 'Confirmar Eliminación'
    }
    return render(request, 'calidad/emos/confirmar_eliminar_emo.html', context)

# Vista para gestionar trabajadores
@login_required
@group_required('Recursos Humanos', 'Calidad')
def gestion_empleados(request):
    """Muestra el dashboard de tarjetas para la gestión de empleados."""
    context = {'current_view': 'gestion_empleados'}
    return render(request, 'calidad/trabajadores/gestion_empleados.html', context)


@login_required
def gestionar_jefaturas(request):
    """Permite asignar niveles jerárquicos y áreas a trabajadores."""
    search_query = request.GET.get('search', '').strip()
    area_id = request.GET.get('area')
    ver_inactivos = request.GET.get('ver_inactivos') == 'on'
    solo_jefes = request.GET.get('solo_jefes') == 'on'

    if request.method == 'POST':
        trabajador_id = request.POST.get('trabajador_id')
        accion = request.POST.get('accion')
        querystring = request.POST.get('querystring', '')

        trabajador = get_object_or_404(Trabajador, pk=trabajador_id)

        if accion == 'guardar_config_jefatura':
            rol_tipo = (request.POST.get('rol_tipo') or 'trabajador').strip().lower()
            alias_roles = {
                'colaborador': 'trabajador',
                'jefe': 'supervisor',
                'jefe_gerente': 'responsable',
            }
            rol_tipo = alias_roles.get(rol_tipo, rol_tipo)

            if rol_tipo not in ['trabajador', 'supervisor', 'responsable', 'gerente']:
                rol_tipo = 'trabajador'

            areas_validas = []

            # Gestion de areas: Supervisor y Responsable usan asignacion manual de areas.
            if rol_tipo in ['supervisor', 'responsable']:
                areas_ids = request.POST.getlist('areas_supervisadas')
                
                # BLOQUEO: Excluimos explícitamente Gerencia General
                areas_validas = list(Area.objects.filter(id__in=areas_ids).exclude(nombre__istartswith='Gerencia General').order_by('nombre'))

                # Fallback: Si no elige áreas, intentamos asignarle su área base
                if not areas_validas and trabajador.area_id:
                    area_principal = Area.objects.filter(id=trabajador.area_id).exclude(nombre__istartswith='Gerencia General').first()
                    if area_principal:
                        areas_validas = [area_principal]

                # Regla obligatoria: un Responsable debe tener al menos un area valida.
                if rol_tipo == 'responsable' and not areas_validas:
                    messages.warning(
                        request,
                        f"{trabajador}: para asignar el rol Responsable debes seleccionar al menos un área.",
                    )
                    redirect_url = reverse('calidad:gestionar_jefaturas')
                    if querystring:
                        redirect_url = f"{redirect_url}?{querystring}"
                    return redirect(redirect_url)

            trabajador.set_cargo_jerarquico(rol_tipo)
            trabajador.save(update_fields=['es_jefe', 'es_gerente'])

            if rol_tipo in ['supervisor', 'responsable']:

                trabajador.areas_supervisadas.set(areas_validas)
                if rol_tipo == 'responsable':
                    messages.success(request, f"{trabajador} configurado como Responsable con áreas asignadas.")
                else:
                    messages.success(request, f"{trabajador} configurado como Supervisor exitosamente.")
            else:
                # Gerente/Trabajador no usan asignacion manual de areas.
                trabajador.areas_supervisadas.clear()
                if rol_tipo == 'gerente':
                    messages.success(request, f"{trabajador} configurado como Gerente.")
                else:
                    messages.success(request, f"{trabajador} configurado como Trabajador.")

        redirect_url = reverse('calidad:gestionar_jefaturas')
        if querystring:
            redirect_url = f"{redirect_url}?{querystring}"
        return redirect(redirect_url)

    # --- CONSULTA GET ---
    trabajadores_queryset = Trabajador.objects.select_related(
        'empresa', 'area', 'sede'
    ).prefetch_related(
        'areas_supervisadas'
    ).all().order_by('apellido_paterno', 'apellido_materno', 'nombres')

    if search_query:
        trabajadores_queryset = trabajadores_queryset.filter(
            Q(nombres__icontains=search_query)
            | Q(apellido_paterno__icontains=search_query)
            | Q(apellido_materno__icontains=search_query)
            | Q(dni__icontains=search_query)
        )

    if area_id:
        trabajadores_queryset = trabajadores_queryset.filter(area_id=area_id)

    if not ver_inactivos:
        trabajadores_queryset = trabajadores_queryset.filter(activo=True)

    if solo_jefes:
        trabajadores_queryset = trabajadores_queryset.filter(Q(es_jefe=True) | Q(es_gerente=True))

    trabajadores_lista = list(trabajadores_queryset)

    for trabajador in trabajadores_lista:
        cargo = trabajador.cargo_jerarquico

        areas_gestion = []
        if cargo in [Trabajador.CargoJerarquico.SUPERVISOR, Trabajador.CargoJerarquico.RESPONSABLE]:
            areas_gestion = list(trabajador.areas_supervisadas.all().order_by('nombre'))
            if trabajador.area and all(a.id != trabajador.area.id for a in areas_gestion):
                if not trabajador.area.nombre.lower().startswith('gerencia general'):
                    areas_gestion.insert(0, trabajador.area)

        trabajador.areas_gestion = areas_gestion

        if cargo == Trabajador.CargoJerarquico.GERENTE:
            trabajador.rol_display = 'gerente'
        elif cargo == Trabajador.CargoJerarquico.RESPONSABLE:
            trabajador.rol_display = 'responsable'
        elif cargo == Trabajador.CargoJerarquico.SUPERVISOR:
            trabajador.rol_display = 'supervisor'
        else:
            trabajador.rol_display = 'colaborador'

    # BLOQUEO EN EL SELECTOR: No mostramos Gerencia General en la lista de áreas para asignar
    opciones_areas = list(Area.objects.exclude(nombre__istartswith='Gerencia General').order_by('nombre'))

    context = {
        'trabajadores': trabajadores_lista,
        'opciones_areas': opciones_areas,
        'opciones_areas_json': [{'id': area.id, 'nombre': area.nombre} for area in opciones_areas],
        'filtro_busqueda': search_query,
        'filtro_activo_area': int(area_id) if area_id else None,
        'ver_inactivos': ver_inactivos,
        'solo_jefes': solo_jefes,
        'querystring_actual': request.GET.urlencode(),
        'current_view': 'gestion_empleados',
    }
    return render(request, 'calidad/trabajadores/gestionar_jefaturas.html', context)


@login_required
@group_required('Recursos Humanos', 'Calidad')
def gestionar_sedes_trabajadores(request):
    """Permite crear sedes y asignarlas a trabajadores desde un solo listado."""
    search_query = request.GET.get('search', '').strip()
    sede_id = request.GET.get('sede')
    solo_activos = request.GET.get('solo_activos') == 'on'

    if request.method == 'POST':
        accion = (request.POST.get('accion') or '').strip().lower()
        querystring = request.POST.get('querystring', '')

        if accion == 'crear_sede':
            nombre_sede = (request.POST.get('nombre_sede') or '').strip()
            direccion_sede = (request.POST.get('direccion_sede') or '').strip()

            if not nombre_sede:
                messages.warning(request, 'Debes ingresar un nombre para la sede.')
            else:
                sede_existente = Sede.objects.filter(nombre__iexact=nombre_sede).first()
                if sede_existente:
                    actualizado = False
                    if direccion_sede and sede_existente.direccion != direccion_sede:
                        sede_existente.direccion = direccion_sede
                        actualizado = True
                    if not sede_existente.activo:
                        sede_existente.activo = True
                        actualizado = True
                    if actualizado:
                        sede_existente.save(update_fields=['direccion', 'activo'])
                    messages.info(request, f'La sede "{sede_existente.nombre}" ya existia y fue actualizada.')
                else:
                    Sede.objects.create(
                        nombre=nombre_sede,
                        direccion=direccion_sede or None,
                        activo=True,
                    )
                    messages.success(request, f'Sede "{nombre_sede}" creada correctamente.')

        elif accion == 'asignar_sede':
            trabajador_id = request.POST.get('trabajador_id')
            sede_asignada_id = request.POST.get('sede_asignada')

            trabajador = get_object_or_404(Trabajador, pk=trabajador_id)
            if sede_asignada_id:
                sede = get_object_or_404(Sede, pk=sede_asignada_id)
                trabajador.sede = sede
                trabajador.save(update_fields=['sede'])
                messages.success(request, f'Sede de {trabajador.nombre_completo} actualizada a {sede.nombre}.')
            else:
                trabajador.sede = None
                trabajador.save(update_fields=['sede'])
                messages.success(request, f'Se retiró la sede asignada a {trabajador.nombre_completo}.')

        redirect_url = reverse('calidad:gestionar_sedes_trabajadores')
        if querystring:
            redirect_url = f"{redirect_url}?{querystring}"
        return redirect(redirect_url)

    trabajadores_queryset = Trabajador.objects.select_related(
        'empresa', 'area', 'sede'
    ).all().order_by('apellido_paterno', 'apellido_materno', 'nombres')

    if search_query:
        trabajadores_queryset = trabajadores_queryset.filter(
            Q(nombres__icontains=search_query)
            | Q(apellido_paterno__icontains=search_query)
            | Q(apellido_materno__icontains=search_query)
            | Q(dni__icontains=search_query)
        )

    if sede_id:
        if sede_id == 'sin_sede':
            trabajadores_queryset = trabajadores_queryset.filter(sede__isnull=True)
        else:
            trabajadores_queryset = trabajadores_queryset.filter(sede_id=sede_id)

    if solo_activos:
        trabajadores_queryset = trabajadores_queryset.filter(activo=True)

    opciones_sedes = Sede.objects.order_by('nombre')

    context = {
        'trabajadores': trabajadores_queryset,
        'opciones_sedes': opciones_sedes,
        'filtro_busqueda': search_query,
        'filtro_sede': sede_id or '',
        'solo_activos': solo_activos,
        'querystring_actual': request.GET.urlencode(),
        'current_view': 'gestion_empleados',
    }
    return render(request, 'calidad/trabajadores/gestionar_sedes_trabajadores.html', context)


def _exportar_trabajadores_a_excel(trabajadores_queryset, filtros):
    """Genera un Excel de trabajadores respetando los filtros aplicados en la lista."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Lista_Trabajadores_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Trabajadores'

    font_bold = Font(bold=True)
    fill_header = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:K1')
    ws['A1'] = 'LISTA DE TRABAJADORES'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    resumen_filtros = (
        f"Búsqueda: {filtros.get('busqueda', 'Todos')} | "
        f"Empresa: {filtros.get('empresa', 'Todas')} | "
        f"Área: {filtros.get('area', 'Todas')} | "
        f"Estado: {filtros.get('estado', 'Activos')}"
    )
    ws.merge_cells('A2:K2')
    ws['A2'] = resumen_filtros

    headers = [
        'N°',
        'DNI',
        'Nombre Completo',
        'Empresa',
        'Área de trabajo',
        'Sede',
        'Fecha de nacimiento',
        'Fecha de ingreso',
        'Estado',
        'Teléfono',
        'Correo electrónico',
    ]

    ws.append([])
    ws.append(headers)

    header_row = 4
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = center_alignment
        cell.border = thin_border

    for idx, trabajador in enumerate(trabajadores_queryset, start=1):
        ws.append([
            idx,
            trabajador.dni,
            str(trabajador),
            trabajador.empresa.nombre if trabajador.empresa else '--',
            trabajador.area.nombre if trabajador.area else '--',
            trabajador.sede.nombre if trabajador.sede else '--',
            trabajador.fecha_nacimiento,
            trabajador.fecha_ingreso,
            'Activo' if trabajador.activo else 'Inactivo',
            trabajador.telefono or '--',
            trabajador.email or '--',
        ])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = thin_border

    for col_idx in range(1, 12):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

    # La columna N° no necesita autoajuste amplio.
    ws.column_dimensions['A'].width = 6

    wb.save(response)
    return response

# Vista para listar trabajadores
@login_required
@group_required('Calidad')
def lista_trabajadores(request):
    # 1. Filtros
    search_query = request.GET.get('search', '')
    empresa_id = request.GET.get('empresa')
    area_id = request.GET.get('area')
    # estado = request.GET.get('estado') # Reemplazado por checkbox
    ver_inactivos = request.GET.get('ver_inactivos') # 'on' si está marcado, None si no

    # 2. Consulta (Sin 'correo_set' para evitar el error por ahora)
    trabajadores_queryset = Trabajador.objects.select_related(
        'empresa', 
        'centro_costo',
        'area',
        'sede'
    ).prefetch_related(
        'asignaciones',             
        'asignaciones__proyecto',   
        'asignaciones__cargo'
    ).all().order_by('apellido_paterno')

    # 3. Aplicar Filtros
    if search_query:
        trabajadores_queryset = trabajadores_queryset.filter(
            Q(nombres__icontains=search_query) |
            Q(apellido_paterno__icontains=search_query) |
            Q(apellido_materno__icontains=search_query) |
            Q(dni__icontains=search_query)
        )

    if empresa_id:
        trabajadores_queryset = trabajadores_queryset.filter(empresa_id=empresa_id)

    if area_id:
        trabajadores_queryset = trabajadores_queryset.filter(area_id=area_id)

    # Filtro de Activo/Inactivo
    if ver_inactivos == 'on':
        # Si el check está marcado, mostramos SOLO los inactivos
        trabajadores_queryset = trabajadores_queryset.filter(activo=False)
    else:
        # Si el check NO está marcado (por defecto), mostramos SOLO los activos
        trabajadores_queryset = trabajadores_queryset.filter(activo=True)

    # 4. Opciones para el select (Traídas de la app RRHH)
    opciones_empresas = Empresa.objects.all().order_by('nombre')
    opciones_areas = Area.objects.all().order_by('nombre')

    empresa_nombre = 'Todas'
    if empresa_id:
        empresa_obj = Empresa.objects.filter(pk=empresa_id).first()
        if empresa_obj:
            empresa_nombre = empresa_obj.nombre

    area_nombre = 'Todas'
    if area_id:
        area_obj = Area.objects.filter(pk=area_id).first()
        if area_obj:
            area_nombre = area_obj.nombre

    if request.GET.get('exportar') == 'excel':
        filtros_excel = {
            'busqueda': search_query or 'Todos',
            'empresa': empresa_nombre,
            'area': area_nombre,
            'estado': 'Inactivos' if ver_inactivos == 'on' else 'Activos',
        }
        return _exportar_trabajadores_a_excel(trabajadores_queryset, filtros_excel)

    context = {
        'trabajadores': trabajadores_queryset,
        'opciones_empresas': opciones_empresas,
        'opciones_areas': opciones_areas,
        'filtro_busqueda': search_query,
        'filtro_activo_empresa': int(empresa_id) if empresa_id else None,
        'filtro_activo_area': int(area_id) if area_id else None,
        # 'filtro_activo_estado': estado,
        'ver_inactivos': ver_inactivos == 'on',
    }

    # Asegúrate de que la ruta del template sea la correcta según tu estructura
    return render(request, 'calidad/trabajadores/lista_trabajadores.html', context)

@login_required
@group_required('Calidad', 'Recursos Humanos')
def gestion_observaciones(request):
    """
    Panel central para la gestión de observaciones y controles médicos.
    """
    # 1. Contar EMOs con Aptitud 'Observado' o 'Apto con Restricción'
    # Asumimos que estos requieren seguimiento
    conteo_observaciones = EMO.objects.filter(
        estado='Realizado',
        aptitud__in=['Observado', 'Apto con Restricción', 'No Apto']
    ).count()

    # 2. Contar Controles Pendientes (Modelo Control que vimos antes)
    # Si no tienes el modelo Control activo, puedes omitir esta parte o usar lógica de EMOs programados
    try:
        hoy = timezone.now().date()
        # Controles no realizados y cuya fecha ya pasó o es hoy
        conteo_controles = Control.objects.filter(
            realizado=False,
            fecha_programada__lte=hoy
        ).count()
    except:
        conteo_controles = 0

    context = {
        'current_view': 'observaciones', # Para resaltar en el sidebar
        'conteo_observaciones': conteo_observaciones,
        'conteo_controles': conteo_controles,
    }
    return render(request, 'calidad/observaciones/gestion_observaciones.html', context)

